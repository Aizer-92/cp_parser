#!/usr/bin/env python3
"""
Исправленный парсинг Excel файла с правильной обработкой объединенных ячеек
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl
from pathlib import Path
import re

def parse_excel_data(excel_path):
    """Парсинг Excel файла с правильной обработкой объединенных ячеек"""
    
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        print(f"❌ Файл не найден: {excel_path}")
        return []
    
    print(f"🔄 Парсинг Excel файла: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        print(f"📊 Размер файла: {ws.max_row} строк x {ws.max_column} колонок")
        
        # Определяем колонки динамически из заголовков
        product_name_col = 2  # B
        characteristics_col = 3  # C
        custom_design_col = 4  # D
        quantity_col = 5  # E
        
        print(f"🔍 Структура колонок:")
        print(f"  - Название товара: {product_name_col} ({chr(64 + product_name_col)})")
        print(f"  - Характеристики: {characteristics_col} ({chr(64 + characteristics_col)})")
        print(f"  - Кастом: {custom_design_col} ({chr(64 + custom_design_col)})")
        print(f"  - Тираж: {quantity_col} ({chr(64 + quantity_col)})")
        
        # Ищем варианты доставки по заголовкам (строка 2) и подзаголовкам (строка 3)
        delivery_variants = {}
        header_row = 2
        subheader_row = 3

        for col in range(6, ws.max_column + 1):  # Начинаем с колонки F
            # Проверяем заголовок (строка 2)
            header_cell = ws.cell(row=header_row, column=col)
            header_value = header_cell.value
            if not header_value:
                continue

            header_text = str(header_value).strip().lower()

            # Определяем тип маршрута по ключевым словам
            route_type = None
            if 'жд' in header_text or 'железнодорожн' in header_text:
                route_type = 'ЖД'
            elif 'авиа' in header_text or 'самолет' in header_text:
                route_type = 'АВИА'
            elif 'контейнер' in header_text:
                route_type = 'Контейнер'
            elif 'образец' in header_text or 'sample' in header_text:
                route_type = 'Образец'

            if route_type and route_type not in delivery_variants:
                delivery_variants[route_type] = {
                    'price_usd_col': None,
                    'price_rub_col': None,
                    'delivery_time_col': None,
                    'sample_price_col': None,
                    'sample_time_col': None
                }

            # Проверяем подзаголовки (строка 3) для определения типа колонки
            subheader_cell = ws.cell(row=subheader_row, column=col)
            subheader_value = subheader_cell.value
            if subheader_value:
                subheader_text = str(subheader_value).strip().lower()

                if route_type:
                    if '$' in subheader_text or 'usd' in subheader_text or 'доллар' in subheader_text:
                        delivery_variants[route_type]['price_usd_col'] = col
                    elif 'руб' in subheader_text or 'rub' in subheader_text:
                        delivery_variants[route_type]['price_rub_col'] = col
                    elif 'срок' in subheader_text and 'тиража' in subheader_text:
                        delivery_variants[route_type]['delivery_time_col'] = col
                    elif 'срок' in subheader_text and ('фото' in subheader_text or 'видео' in subheader_text):
                        delivery_variants[route_type]['sample_time_col'] = col
                    elif 'срок' in subheader_text and 'доставкой' in subheader_text:
                        delivery_variants[route_type]['sample_time_col'] = col

        # Дополнительный поиск по соседним колонкам для рублевых цен и сроков
        for variant, cols in delivery_variants.items():
            if cols['price_usd_col'] and not cols['price_rub_col']:
                # Ищем рублевую цену в следующей колонке
                next_col = cols['price_usd_col'] + 1
                if next_col <= ws.max_column:
                    next_cell = ws.cell(row=subheader_row, column=next_col)
                    if next_cell.value and 'руб' in str(next_cell.value).lower():
                        delivery_variants[variant]['price_rub_col'] = next_col

            if cols['price_rub_col'] and not cols['delivery_time_col']:
                # Ищем срок в следующей колонке
                next_col = cols['price_rub_col'] + 1
                if next_col <= ws.max_column:
                    next_cell = ws.cell(row=subheader_row, column=next_col)
                    if next_cell.value and 'срок' in str(next_cell.value).lower():
                        delivery_variants[variant]['delivery_time_col'] = next_col

        print(f"🚚 Найденные варианты доставки:")
        for variant, cols in delivery_variants.items():
            print(f"  - {variant}:")
            print(f"    USD: {cols['price_usd_col']}, RUB: {cols['price_rub_col']}")
            print(f"    Срок: {cols['delivery_time_col']}, Образец: {cols['sample_price_col']}")
            print(f"    Срок образца: {cols['sample_time_col']}")
        
        # Собираем все строки с товарами
        product_rows = []
        current_product_name = None
        current_characteristics = None
        current_custom_design = None
        
        for row_num in range(4, ws.max_row + 1):
            # Получаем название товара
            product_name_cell = ws.cell(row=row_num, column=product_name_col)
            product_name = product_name_cell.value
            
            # Если название есть, обновляем текущее название и характеристики
            if product_name and str(product_name).strip() != '' and str(product_name).strip() != ' ':
                current_product_name = str(product_name).strip()
                
                # Получаем характеристики
                characteristics_cell = ws.cell(row=row_num, column=characteristics_col)
                current_characteristics = str(characteristics_cell.value).strip() if characteristics_cell.value else None
                
                # Получаем кастом
                custom_design_cell = ws.cell(row=row_num, column=custom_design_col)
                current_custom_design = str(custom_design_cell.value).strip() if custom_design_cell.value else None
            
            # Если нет текущего названия, пропускаем строку
            if not current_product_name:
                continue
            
            # Получаем тираж
            quantity_cell = ws.cell(row=row_num, column=quantity_col)
            quantity = quantity_cell.value
            if quantity:
                try:
                    # Убираем пробелы и конвертируем в число
                    quantity_str = str(quantity).replace(' ', '').replace(',', '')
                    quantity = int(float(quantity_str))
                except (ValueError, TypeError):
                    quantity = None
            
            # Собираем ценовые предложения для этой строки
            price_offers = []
            
            for variant, cols in delivery_variants.items():
                # Получаем цены и сроки
                price_usd = None
                price_rub = None
                delivery_time = None
                sample_price = None
                sample_time = None
                
                if cols['price_usd_col']:
                    usd_cell = ws.cell(row=row_num, column=cols['price_usd_col'])
                    if usd_cell.value:
                        try:
                            price_usd = float(usd_cell.value)
                        except (ValueError, TypeError):
                            pass
                
                if cols['price_rub_col']:
                    rub_cell = ws.cell(row=row_num, column=cols['price_rub_col'])
                    if rub_cell.value:
                        try:
                            price_rub = float(rub_cell.value)
                        except (ValueError, TypeError):
                            pass
                
                if cols['delivery_time_col']:
                    time_cell = ws.cell(row=row_num, column=cols['delivery_time_col'])
                    if time_cell.value:
                        delivery_time = str(time_cell.value).strip()
                
                if cols['sample_price_col']:
                    sample_cell = ws.cell(row=row_num, column=cols['sample_price_col'])
                    if sample_cell.value:
                        try:
                            sample_price = float(sample_cell.value)
                        except (ValueError, TypeError):
                            pass
                
                if cols['sample_time_col']:
                    sample_time_cell = ws.cell(row=row_num, column=cols['sample_time_col'])
                    if sample_time_cell.value:
                        sample_time = str(sample_time_cell.value).strip()
                
                # Определяем, является ли это предложением по образцу
                is_sample = variant == 'Образец'
                
                # Для образцов используем рублевую цену как основную
                if is_sample and price_rub:
                    sample_price = price_rub
                    price_rub = None
                
                # Создаем ценовое предложение только если есть хотя бы одна цена
                if price_usd or price_rub or sample_price:
                    offer = {
                        'route_name': variant,
                        'quantity': quantity,
                        'price_usd': price_usd,
                        'price_rub': price_rub,
                        'delivery_time': delivery_time,
                        'sample_price': sample_price,
                        'sample_time': sample_time,
                        'is_sample': is_sample
                    }
                    price_offers.append(offer)
            
            # Добавляем строку товара только если есть ценовые предложения
            if price_offers:
                product_rows.append({
                    'name': current_product_name,
                    'description': current_characteristics,
                    'characteristics': current_characteristics,
                    'custom_design': current_custom_design,
                    'quantity': quantity,
                    'price_offers': price_offers,
                    'images': []
                })
        
        # Группируем товары по названию и тиражу
        grouped_products = {}
        
        for product_row in product_rows:
            key = f"{product_row['name']}_{product_row['quantity']}"
            
            if key not in grouped_products:
                grouped_products[key] = {
                    'name': product_row['name'],
                    'description': product_row['description'],
                    'characteristics': product_row['characteristics'],
                    'custom_design': product_row['custom_design'],
                    'quantity': product_row['quantity'],
                    'price_offers': [],
                    'images': product_row['images']
                }
            
            # Добавляем ценовые предложения
            grouped_products[key]['price_offers'].extend(product_row['price_offers'])
        
        # Преобразуем в список
        products = list(grouped_products.values())
        
        print(f"✅ Парсинг завершен!")
        print(f"📊 Результаты:")
        print(f"  - Найдено товаров: {len(products)}")
        
        total_offers = sum(len(p['price_offers']) for p in products)
        print(f"  - Найдено ценовых предложений: {total_offers}")
        
        # Показываем детали по товарам
        for product in products:
            print(f"\n📦 {product['name']} (тираж: {product['quantity']})")
            print(f"  Ценовых предложений: {len(product['price_offers'])}")
            for offer in product['price_offers']:
                if offer['is_sample']:
                    print(f"    🎯 {offer['route_name']}: {offer['sample_price']} руб ({offer['sample_time']} дней)")
                else:
                    print(f"    📦 {offer['route_name']}: ${offer['price_usd']} / {offer['price_rub']} руб ({offer['delivery_time']} дней)")
        
        return products
        
    except Exception as e:
        print(f"❌ Ошибка при парсинге: {e}")
        import traceback
        traceback.print_exc()
        return []

if __name__ == "__main__":
    # Тестируем парсинг
    excel_path = "storage/excel_files/Вторая таблица_1757933504.xlsx"
    products = parse_excel_corrected(excel_path)
    print(f"\n📊 Итого найдено товаров: {len(products)}")
