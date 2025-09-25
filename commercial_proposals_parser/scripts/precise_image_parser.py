#!/usr/bin/env python3
"""
Точный парсер изображений с детальным анализом позиционирования в Excel
"""

import os
import sys
import zipfile
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import json

# Добавляем путь к модулям
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL

def clear_existing_images():
    """Очищает существующие изображения в базе данных"""
    print("🧹 Очищаем существующие изображения...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import ProductImage
        
        # Удаляем все записи об изображениях
        deleted_count = session.query(ProductImage).count()
        session.query(ProductImage).delete()
        session.commit()
        print(f"✅ Удалено записей об изображениях из БД: {deleted_count}")
    
    # Очищаем папку с изображениями
    images_dir = "storage/images/products_parsed"
    if os.path.exists(images_dir):
        shutil.rmtree(images_dir)
        print(f"✅ Очищена папка {images_dir}")
    
    os.makedirs(images_dir, exist_ok=True)
    print(f"✅ Создана новая папка {images_dir}")

def analyze_cell_content(ws, row, col):
    """Анализирует содержимое ячейки и соседних ячеек"""
    try:
        cell_value = ws.cell(row=row, column=col).value
        if not cell_value:
            return None
            
        cell_text = str(cell_value).strip()
        if not cell_text or cell_text == 'None':
            return None
            
        # Проверяем, является ли это названием товара
        product_keywords = [
            'ежедневник', 'худи', 'шапка', 'бини', 'шоппер', 'брелок', 'таблетница', 
            'кардхолдер', 'зонт', 'письменный', 'набор', 'термокружка', 'сумка',
            'футболка', 'кружка', 'ручка', 'карандаш', 'футляр', 'обложка', 'косметичка',
            'чехол', 'ноутбук', 'дутый', 'обложка для паспорта', 'набор карандашей',
            'футляр для очков', 'дутый чехол для ноутбука'
        ]
        
        for keyword in product_keywords:
            if keyword.lower() in cell_text.lower():
                return {
                    'name': cell_text,
                    'row': row,
                    'col': col,
                    'keyword': keyword,
                    'confidence': len(keyword) / len(cell_text) if cell_text else 0
                }
        
        return None
        
    except Exception as e:
        return None

def find_product_in_row_range(ws, start_row, end_row, search_columns=range(1, 25)):
    """Ищет товар в диапазоне строк"""
    best_match = None
    best_confidence = 0
    
    for row in range(start_row, end_row + 1):
        for col in search_columns:
            match = analyze_cell_content(ws, row, col)
            if match and match['confidence'] > best_confidence:
                best_confidence = match['confidence']
                best_match = match
    
    return best_match

def extract_images_from_excel_zip(excel_path, output_dir):
    """Извлекает изображения из Excel файла через zipfile"""
    print(f"📤 Извлекаем изображения из {excel_path}")
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and f.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
            
            print(f"  🖼️  Найдено изображений в архиве: {len(image_files)}")
            
            extracted_images = []
            
            for i, image_file in enumerate(image_files):
                try:
                    image_data = zip_ref.read(image_file)
                    original_name = os.path.basename(image_file)
                    new_name = f"image_{i+1}_{original_name}"
                    output_path = os.path.join(output_dir, new_name)
                    
                    with open(output_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_images.append({
                        'original_path': image_file,
                        'saved_path': output_path,
                        'filename': new_name,
                        'index': i + 1
                    })
                    
                    print(f"    ✅ Извлечено: {original_name} -> {new_name}")
                    
                except Exception as e:
                    print(f"    ❌ Ошибка при извлечении {image_file}: {e}")
            
            return extracted_images
            
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений из {excel_path}: {e}")
        return []

def analyze_excel_precise_positioning(excel_path):
    """Точный анализ позиционирования изображений в Excel"""
    print(f"🎯 Точный анализ позиционирования: {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        analysis = {
            'file_path': excel_path,
            'sheets': {},
            'images_with_products': []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📊 Лист: {sheet_name}")
            
            sheet_info = {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'images': []
            }
            
            # Собираем все изображения с их точными координатами
            images_data = []
            for i, image in enumerate(ws._images):
                try:
                    if hasattr(image, 'anchor') and image.anchor:
                        anchor = image.anchor
                        
                        # Получаем точные координаты
                        if hasattr(anchor, '_from'):
                            from_cell = anchor._from
                            row = from_cell.row + 1
                            col = from_cell.col + 1
                        else:
                            row = getattr(anchor, 'row', 1)
                            col = getattr(anchor, 'col', 1)
                        
                        # Получаем размеры изображения
                        width = getattr(image, 'width', 0)
                        height = getattr(image, 'height', 0)
                        
                        images_data.append({
                            'index': i + 1,
                            'row': row,
                            'col': col,
                            'width': width,
                            'height': height
                        })
                        
                        print(f"    🖼️  Изображение {i+1}: строка {row}, колонка {col}, размер {width}x{height}")
                        
                except Exception as e:
                    print(f"    ❌ Ошибка при анализе изображения {i+1}: {e}")
            
            # Сортируем изображения по строкам
            images_data.sort(key=lambda x: (x['row'], x['col']))
            
            # Для каждого изображения ищем товар в той же строке
            for img_data in images_data:
                row = img_data['row']
                col = img_data['col']
                
                print(f"    🔍 Анализируем изображение в строке {row}, колонке {col}")
                
                # Ищем товар в той же строке
                product_match = find_product_in_row_range(ws, row, row, range(1, 25))
                
                if product_match:
                    print(f"      ✅ Найден товар: '{product_match['name']}' (колонка {product_match['col']}, уверенность: {product_match['confidence']:.2f})")
                    
                    img_data['product_name'] = product_match['name']
                    img_data['product_col'] = product_match['col']
                    img_data['product_confidence'] = product_match['confidence']
                else:
                    print(f"      ⚠️  Товар не найден в строке {row}")
                    img_data['product_name'] = f'Unknown_row_{row}'
                    img_data['product_col'] = 0
                    img_data['product_confidence'] = 0
                
                sheet_info['images'].append(img_data)
                analysis['images_with_products'].append({
                    'sheet': sheet_name,
                    **img_data
                })
            
            analysis['sheets'][sheet_name] = sheet_info
        
        return analysis
        
    except Exception as e:
        print(f"❌ Ошибка при анализе {excel_path}: {e}")
        return None

def create_precise_image_mapping(excel_files):
    """Создает точный маппинг изображений к товарам"""
    print("🎯 Создаем точный маппинг изображений к товарам...")
    
    all_images = []
    
    for excel_file in excel_files:
        if os.path.exists(excel_file):
            print(f"\n📁 Обрабатываем: {excel_file}")
            
            # Анализируем структуру с точным позиционированием
            analysis = analyze_excel_precise_positioning(excel_file)
            if analysis:
                # Извлекаем изображения
                output_dir = f"storage/images/products_parsed/{os.path.basename(excel_file).replace('.xlsx', '')}"
                os.makedirs(output_dir, exist_ok=True)
                
                extracted_images = extract_images_from_excel_zip(excel_file, output_dir)
                
                # Связываем извлеченные изображения с анализом
                for i, extracted in enumerate(extracted_images):
                    if i < len(analysis['images_with_products']):
                        image_info = analysis['images_with_products'][i]
                        all_images.append({
                            'excel_file': excel_file,
                            'sheet_name': image_info['sheet'],
                            'row': image_info['row'],
                            'col': image_info['col'],
                            'product_name': image_info['product_name'],
                            'product_col': image_info['product_col'],
                            'product_confidence': image_info['product_confidence'],
                            'image_path': extracted['saved_path'],
                            'image_filename': extracted['filename']
                        })
    
    return all_images

def map_images_to_products_precise(image_mapping):
    """Точно привязывает изображения к товарам"""
    print("🎯 Точно привязываем изображения к товарам...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import Product, ProductImage
        
        # Получаем все товары
        products = session.query(Product).all()
        print(f"📦 Найдено товаров в БД: {len(products)}")
        
        # Создаем словарь для поиска товаров
        products_by_name = {}
        for product in products:
            normalized_name = product.name.strip().lower()
            if normalized_name not in products_by_name:
                products_by_name[normalized_name] = []
            products_by_name[normalized_name].append(product)
        
        mapped_count = 0
        unmatched_images = []
        
        for image_info in image_mapping:
            product_name = image_info['product_name'].strip()
            confidence = image_info['product_confidence']
            
            if product_name.startswith('Unknown_') or confidence < 0.3:
                unmatched_images.append(image_info)
                print(f"  ⚠️  Пропускаем: {product_name} (уверенность: {confidence:.2f})")
                continue
            
            # Ищем товар по названию
            normalized_name = product_name.lower()
            matching_products = products_by_name.get(normalized_name, [])
            
            if not matching_products:
                # Пробуем найти по частичному совпадению
                for norm_name, prods in products_by_name.items():
                    if normalized_name in norm_name or norm_name in normalized_name:
                        matching_products = prods
                        break
            
            if matching_products:
                product = matching_products[0]
                
                # Создаем новое имя файла
                new_image_name = f"product_{product.id}_precise_{mapped_count + 1}.jpg"
                new_image_path = f"storage/images/products_parsed/{new_image_name}"
                
                try:
                    # Копируем изображение
                    shutil.copy2(image_info['image_path'], new_image_path)
                    
                    # Создаем запись в БД
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=new_image_path,
                        image_type='main' if mapped_count == 0 else 'additional'
                    )
                    session.add(product_image)
                    mapped_count += 1
                    
                    print(f"  ✅ {product_name} -> {product.name} #{product.id} -> {new_image_name} (уверенность: {confidence:.2f})")
                    
                except Exception as e:
                    print(f"  ❌ Ошибка при копировании изображения: {e}")
            else:
                unmatched_images.append(image_info)
                print(f"  ⚠️  Товар не найден: '{product_name}'")
        
        # Пытаемся привязать несоответствующие изображения к товарам без изображений
        print(f"\n🔄 Пытаемся привязать {len(unmatched_images)} несоответствующих изображений...")
        
        products_without_images = session.query(Product).outerjoin(ProductImage).filter(ProductImage.id.is_(None)).all()
        
        for i, image_info in enumerate(unmatched_images):
            if i < len(products_without_images):
                product = products_without_images[i]
                
                new_image_name = f"product_{product.id}_unmatched_{i+1}.jpg"
                new_image_path = f"storage/images/products_parsed/{new_image_name}"
                
                try:
                    shutil.copy2(image_info['image_path'], new_image_path)
                    
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=new_image_path,
                        image_type='main'
                    )
                    session.add(product_image)
                    mapped_count += 1
                    
                    print(f"  🔄 {image_info['product_name']} -> {product.name} #{product.id} -> {new_image_name} (unmatched)")
                    
                except Exception as e:
                    print(f"  ❌ Ошибка при привязке несоответствующего изображения: {e}")
        
        session.commit()
        print(f"\n✅ Привязано изображений: {mapped_count}")

def main():
    """Основная функция"""
    print("🎯 Точный парсинг изображений с детальным позиционированием")
    
    # Очищаем существующие данные
    clear_existing_images()
    
    # Список Excel файлов для обработки
    excel_files = [
        "storage/excel_files/original_sheet.xlsx",
        "storage/excel_files/Вторая таблица_1757933430.xlsx", 
        "storage/excel_files/Мерч для Sense_1757934153.xlsx"
    ]
    
    # Создаем точный маппинг изображений
    image_mapping = create_precise_image_mapping(excel_files)
    
    print(f"\n📊 Всего найдено изображений: {len(image_mapping)}")
    
    # Точно привязываем изображения к товарам
    map_images_to_products_precise(image_mapping)
    
    print("\n✅ Точный парсинг изображений завершен!")

if __name__ == "__main__":
    main()
