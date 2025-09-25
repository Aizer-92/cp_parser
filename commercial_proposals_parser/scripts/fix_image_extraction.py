#!/usr/bin/env python3
"""
Правильное извлечение изображений из Excel с сохранением порядка
"""

import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import shutil
from PIL import Image
import io

sys.path.append('.')
from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL_V4
from sqlalchemy import text

def extract_images_with_order(excel_file_path, output_dir):
    """Извлечение изображений с сохранением порядка по позиции в Excel"""
    
    print(f"🖼️  Извлекаем изображения из Excel файла с сохранением порядка...")
    
    # Загружаем Excel файл
    workbook = load_workbook(excel_file_path, data_only=True)
    sheet = workbook.active
    
    # Создаем словарь товаров по строкам
    products_by_row = {}
    for row_num in range(4, sheet.max_row + 1):
        cell_c = sheet.cell(row=row_num, column=3)  # Колонка C - название товара
        if cell_c.value and isinstance(cell_c.value, str) and cell_c.value.strip():
            product_name = cell_c.value.strip()
            if product_name != 'Наименование':
                products_by_row[row_num] = product_name
    
    print(f"📦 Найдено товаров в Excel: {len(products_by_row)}")
    
    # Анализируем позиции изображений и извлекаем их
    image_positions = []
    extracted_images = []
    
    try:
        with zipfile.ZipFile(excel_file_path, 'r') as z:
            # Сначала получаем все изображения
            image_files = [name for name in z.namelist() if name.startswith('xl/media/image')]
            print(f"📷 Найдено изображений в Excel: {len(image_files)}")
            
            # Извлекаем изображения
            for i, name in enumerate(image_files):
                image_data = z.read(name)
                try:
                    img = Image.open(io.BytesIO(image_data))
                    img.verify()
                    
                    # Сохраняем изображение
                    temp_name = f"image_{i+1:02d}_{os.path.basename(name)}"
                    temp_path = os.path.join(output_dir, temp_name)
                    with open(temp_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_images.append({
                        'original_name': name,
                        'temp_path': temp_path,
                        'format': img.format,
                        'index': i
                    })
                    print(f"  ✅ Извлечено: {temp_name}")
                except Exception as e:
                    print(f"  ❌ Пропускаем поврежденное изображение: {name}")
            
            # Теперь анализируем позиции изображений
            drawing_files = [name for name in z.namelist() if 'xl/drawings/drawing' in name]
            print(f"📐 Анализируем позиции изображений...")
            
            for drawing_file in drawing_files:
                try:
                    drawing_data = z.read(drawing_file)
                    root = ET.fromstring(drawing_data)
                    
                    for anchor in root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}oneCellAnchor'):
                        from_elem = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
                        if from_elem is not None:
                            row_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
                            col_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col')
                            if row_elem is not None and col_elem is not None:
                                row = int(row_elem.text) + 1
                                col = int(col_elem.text) + 1
                                
                                # Находим товар в этой строке
                                product_name = products_by_row.get(row, f"Unknown_row_{row}")
                                
                                # Определяем тип изображения
                                if col == 1:  # Колонка A - основное изображение
                                    image_type = 'main'
                                elif col == 16:  # Колонка P - дополнительное изображение
                                    image_type = 'additional'
                                else:
                                    image_type = 'other'
                                
                                image_positions.append({
                                    'row': row,
                                    'col': col,
                                    'product_name': product_name,
                                    'image_type': image_type,
                                    'drawing_file': drawing_file
                                })
                except Exception as e:
                    print(f"  ❌ Ошибка при анализе {drawing_file}: {e}")
                    
    except Exception as e:
        print(f"❌ Ошибка при работе с Excel: {e}")
        return [], []
    
    print(f"✅ Извлечено {len(extracted_images)} изображений")
    print(f"✅ Найдено {len(image_positions)} позиций изображений")
    
    return extracted_images, image_positions

def create_correct_mapping_with_order(extracted_images, image_positions, output_dir):
    """Создание правильной привязки изображений с учетом порядка"""
    
    print(f"\n🔗 Создаем правильную привязку изображений...")
    
    # Группируем изображения по товарам
    products_images = {}
    for pos in image_positions:
        product_name = pos['product_name']
        if product_name not in products_images:
            products_images[product_name] = {'main': [], 'additional': []}
        products_images[product_name][pos['image_type']].append(pos)
    
    print(f"📦 Товары с изображениями:")
    for product_name, images in products_images.items():
        main_count = len(images['main'])
        additional_count = len(images['additional'])
        print(f"  {product_name}: {main_count} основных, {additional_count} дополнительных")
    
    # Создаем маппинг
    mappings = []
    image_index = 0
    
    for product_name, images in products_images.items():
        # Основные изображения
        for i, pos in enumerate(images['main']):
            if image_index < len(extracted_images):
                img_info = extracted_images[image_index]
                
                # Создаем новое имя файла
                new_image_name = f"product_{product_name.replace(' ', '_').replace('/', '_')}_main_{i+1}.jpg"
                new_image_path = os.path.join(output_dir, new_image_name)
                
                # Копируем изображение
                shutil.copy2(img_info['temp_path'], new_image_path)
                
                mappings.append({
                    'product_name': product_name,
                    'image_path': new_image_path,
                    'relative_path': f"storage/images/products_parsed/{new_image_name}",
                    'image_type': 'main',
                    'row': pos['row'],
                    'original_index': img_info['index']
                })
                
                print(f"  ✅ {product_name} (строка {pos['row']}) -> {new_image_name} (основное, индекс {img_info['index']})")
                image_index += 1
        
        # Дополнительные изображения
        for i, pos in enumerate(images['additional']):
            if image_index < len(extracted_images):
                img_info = extracted_images[image_index]
                
                # Создаем новое имя файла
                new_image_name = f"product_{product_name.replace(' ', '_').replace('/', '_')}_additional_{i+1}.jpg"
                new_image_path = os.path.join(output_dir, new_image_name)
                
                # Копируем изображение
                shutil.copy2(img_info['temp_path'], new_image_path)
                
                mappings.append({
                    'product_name': product_name,
                    'image_path': new_image_path,
                    'relative_path': f"storage/images/products_parsed/{new_image_name}",
                    'image_type': 'additional',
                    'row': pos['row'],
                    'original_index': img_info['index']
                })
                
                print(f"  ✅ {product_name} (строка {pos['row']}) -> {new_image_name} (дополнительное, индекс {img_info['index']})")
                image_index += 1
    
    return mappings

def fix_image_extraction():
    """Основная функция для правильного извлечения изображений"""
    
    db = CommercialProposalsDB(DATABASE_URL_V4)
    
    # 1. Очищаем существующие изображения для товаров "Мерч для Sense"
    print("🧹 Очищаем существующие изображения для товаров 'Мерч для Sense'...")
    session = db.get_session()
    try:
        session.execute(text('DELETE FROM product_images WHERE product_id >= 29'))
        session.commit()
        print("✅ Очищены изображения для товаров 'Мерч для Sense'")
    finally:
        session.close()
    
    # 2. Очищаем папку для изображений
    images_output_dir = "storage/images/products_parsed"
    if os.path.exists(images_output_dir):
        # Удаляем только изображения товаров "Мерч для Sense"
        for file in os.listdir(images_output_dir):
            if file.startswith('product_') and any(name in file for name in ['Худи', 'ШАПКА', 'ШОППЕР', 'Брелоки', 'ДУТЫЙ', 'Зонт', 'Письменный', 'Термокружка', 'Сумка', 'Шоппер']):
                os.remove(os.path.join(images_output_dir, file))
        print(f"✅ Очищены изображения товаров 'Мерч для Sense' из папки")
    
    # 3. Извлекаем изображения с сохранением порядка
    excel_file = "storage/excel_files/Мерч для Sense_1758096973.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл не найден: {excel_file}")
        return
    
    extracted_images, image_positions = extract_images_with_order(excel_file, images_output_dir)
    
    if not extracted_images or not image_positions:
        print("❌ Не удалось извлечь изображения или позиции")
        return
    
    # 4. Создаем правильную привязку
    mappings = create_correct_mapping_with_order(extracted_images, image_positions, images_output_dir)
    
    # 5. Получаем товары из БД
    products_with_details = db.get_all_products_with_details(limit=50)
    merch_sense_products = [(p, po, im) for p, po, im in products_with_details if p.id >= 29]
    
    # Создаем словарь для сопоставления имен товаров
    product_name_map = {}
    for product, price_offers, images in merch_sense_products:
        product_name_map[product.name.strip().lower()] = product
    
    # 6. Сохраняем привязки в базу данных
    print(f"\n💾 Сохраняем привязки в базу данных...")
    
    mapped_count = 0
    for mapping in mappings:
        product_name = mapping['product_name'].strip().lower()
        
        if product_name in product_name_map:
            product = product_name_map[product_name]
            
            try:
                db.create_product_image(
                    product_id=product.id,
                    image_path=mapping['relative_path'],
                    image_type=mapping['image_type']
                )
                mapped_count += 1
                print(f"  ✅ {product.name} (ID: {product.id}) -> {os.path.basename(mapping['image_path'])} ({mapping['image_type']})")
            except Exception as e:
                print(f"  ❌ Ошибка при сохранении изображения для товара {product.name}: {e}")
        else:
            print(f"  ⚠️ Товар '{mapping['product_name']}' не найден в БД")
    
    print(f"\n✅ Правильное извлечение завершено! Привязано {mapped_count} изображений.")
    
    # 7. Проверяем результат
    print(f"\n📊 Проверяем результат:")
    for product, price_offers, images in merch_sense_products:
        main_images = [img for img in images if img.image_type == 'main']
        additional_images = [img for img in images if img.image_type == 'additional']
        print(f"  {product.name} (ID: {product.id}) - {len(main_images)} основных, {len(additional_images)} дополнительных")

if __name__ == "__main__":
    fix_image_extraction()
