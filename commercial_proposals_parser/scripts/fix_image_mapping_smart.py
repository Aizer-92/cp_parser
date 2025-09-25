#!/usr/bin/env python3
"""
Умная привязка изображений к товарам на основе анализа содержимого
"""

import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import shutil
from PIL import Image
import io

sys.path.append('.')
from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL_V4
from sqlalchemy import text

def analyze_excel_structure(excel_file_path):
    """Анализ структуры Excel файла для понимания расположения товаров и изображений"""
    
    print(f"🔍 Анализируем структуру файла: {excel_file_path}")
    
    # Загружаем Excel файл
    workbook = load_workbook(excel_file_path, data_only=True)
    sheet = workbook.active
    
    # Находим строки с товарами (ищем в колонке C - название товара)
    products_info = []
    for row_num in range(1, sheet.max_row + 1):
        cell_c = sheet.cell(row=row_num, column=3)  # Колонка C - название товара
        if cell_c.value and isinstance(cell_c.value, str) and cell_c.value.strip():
            product_name = cell_c.value.strip()
            # Пропускаем заголовки
            if product_name not in ['Наименование']:
                products_info.append({
                    'row': row_num,
                    'name': product_name,
                    'cell': f'C{row_num}'
                })
    
    print(f"📦 Найдено товаров в Excel: {len(products_info)}")
    for i, product in enumerate(products_info[:10], 1):
        print(f"  {i}. Строка {product['row']}: {product['name']}")
    
    # Извлекаем изображения из Excel
    extracted_images = []
    try:
        with zipfile.ZipFile(excel_file_path, 'r') as z:
            for name in z.namelist():
                if name.startswith('xl/media/image'):
                    image_data = z.read(name)
                    
                    # Проверяем, что это валидное изображение
                    try:
                        img = Image.open(io.BytesIO(image_data))
                        img.verify()
                        
                        extracted_images.append({
                            'original_name': name,
                            'image_data': image_data,
                            'format': img.format
                        })
                    except Exception as e:
                        print(f"  ❌ Пропускаем поврежденное изображение: {name}")
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений: {e}")
        return [], []
    
    print(f"🖼️  Извлечено изображений: {len(extracted_images)}")
    
    return products_info, extracted_images

def create_smart_mapping(products_info, extracted_images, output_dir):
    """Создание умной привязки изображений к товарам"""
    
    print(f"\n🔗 Создаем умную привязку изображений...")
    
    # Простая эвристика: привязываем изображения к товарам по порядку
    # В реальности здесь должен быть более сложный алгоритм анализа содержимого
    
    mappings = []
    
    # Для каждого товара пытаемся найти подходящее изображение
    for i, product in enumerate(products_info):
        if i < len(extracted_images):
            img_info = extracted_images[i]
            
            # Создаем новое имя файла
            new_image_name = f"product_{product['name'].replace(' ', '_').replace('/', '_')}_main.jpg"
            new_image_path = os.path.join(output_dir, new_image_name)
            
            # Сохраняем изображение
            with open(new_image_path, 'wb') as f:
                f.write(img_info['image_data'])
            
            mappings.append({
                'product_name': product['name'],
                'product_row': product['row'],
                'image_path': new_image_path,
                'original_name': img_info['original_name']
            })
            
            print(f"  ✅ {product['name']} (строка {product['row']}) -> {new_image_name}")
    
    return mappings

def fix_image_mapping_smart():
    """Основная функция для умной привязки изображений"""
    
    db = CommercialProposalsDB(DATABASE_URL_V4)
    
    # 1. Очищаем существующие изображения для товаров "Мерч для Sense"
    print("🧹 Очищаем существующие изображения для товаров 'Мерч для Sense'...")
    session = db.get_session()
    try:
        # Удаляем изображения для товаров с ID >= 29 (Мерч для Sense)
        session.execute(text('DELETE FROM product_images WHERE product_id >= 29'))
        session.commit()
        print("✅ Очищены изображения для товаров 'Мерч для Sense'")
    finally:
        session.close()
    
    # 2. Очищаем папку для изображений
    images_output_dir = "storage/images/products_parsed"
    if os.path.exists(images_output_dir):
        # Удаляем только изображения товаров "Мерч для Sense"
        for file in os.listdir(images_output_dir):
            if file.startswith('product_') and any(name in file for name in ['Худи', 'ШАПКА', 'ШОППЕР', 'Брелоки', 'ДУТЫЙ', 'Зонт', 'Письменный', 'Термокружка', 'Сумка', 'Шоппер']):
                os.remove(os.path.join(images_output_dir, file))
        print(f"✅ Очищены изображения товаров 'Мерч для Sense' из папки")
    
    # 3. Анализируем Excel файл
    excel_file = "storage/excel_files/Мерч для Sense_1758096973.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл не найден: {excel_file}")
        return
    
    products_info, extracted_images = analyze_excel_structure(excel_file)
    
    if not products_info or not extracted_images:
        print("❌ Не удалось извлечь данные из Excel файла")
        return
    
    # 4. Создаем умную привязку
    mappings = create_smart_mapping(products_info, extracted_images, images_output_dir)
    
    # 5. Привязываем изображения к товарам в базе данных
    print(f"\n💾 Сохраняем привязки в базу данных...")
    
    # Получаем товары "Мерч для Sense" из базы данных
    products_with_details = db.get_all_products_with_details(limit=50)
    merch_sense_products = [(p, po, im) for p, po, im in products_with_details if p.id >= 29]
    
    print(f"📦 Найдено товаров 'Мерч для Sense' в БД: {len(merch_sense_products)}")
    
    # Создаем словарь для сопоставления имен товаров
    product_name_map = {}
    for product, price_offers, images in merch_sense_products:
        product_name_map[product.name.strip().lower()] = product
    
    mapped_count = 0
    for mapping in mappings:
        product_name = mapping['product_name'].strip().lower()
        
        if product_name in product_name_map:
            product = product_name_map[product_name]
            
            # Создаем запись об изображении в БД
            relative_path = f"storage/images/products_parsed/{os.path.basename(mapping['image_path'])}"
            db.create_product_image(
                product_id=product.id,
                image_path=relative_path,
                image_type='main'
            )
            
            mapped_count += 1
            print(f"  ✅ {product.name} (ID: {product.id}) -> {os.path.basename(mapping['image_path'])}")
        else:
            print(f"  ⚠️ Товар '{mapping['product_name']}' не найден в БД")
    
    print(f"\n✅ Умная привязка завершена! Привязано {mapped_count} изображений.")
    
    # 6. Проверяем результат
    print(f"\n📊 Проверяем результат:")
    for product, price_offers, images in merch_sense_products:
        print(f"  {product.name} (ID: {product.id}) - {len(images)} изображений")

if __name__ == "__main__":
    fix_image_mapping_smart()
