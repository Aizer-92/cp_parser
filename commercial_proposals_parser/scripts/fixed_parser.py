#!/usr/bin/env python3
"""
Исправленный парсер Excel файлов
"""

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import json
from typing import List, Dict, Any, Optional, Tuple
from PIL import Image
import io

class FixedExcelParser:
    """Исправленный парсер Excel файлов"""
    
    def __init__(self):
        self.image_columns = []
        self.price_columns = {}
        self.route_columns = {}
        self.quantity_column = None
        self.product_name_column = None
        self.characteristics_column = None
        self.custom_design_column = None
        
    def analyze_structure(self, file_path: str) -> Dict[str, Any]:
        """Анализирует структуру Excel файла"""
        print(f"🔍 Анализ структуры: {os.path.basename(file_path)}")
        
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Анализируем заголовки
        headers = {}
        for row in range(1, min(4, ws.max_row + 1)):
            row_headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    row_headers.append((col, str(cell.value).strip()))
            headers[f'row_{row}'] = row_headers
        
        # Определяем колонки
        structure = {
            'file_name': os.path.basename(file_path),
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'headers': headers,
            'image_columns': [],
            'price_columns': {},
            'route_columns': {},
            'quantity_column': None,
            'product_name_column': None,
            'characteristics_column': None,
            'custom_design_column': None
        }
        
        # Ищем колонки с изображениями
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(10, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['фото', 'photo', 'изображение']):
                    structure['image_columns'].append(col)
                    break
        
        # Ищем колонку с названием товара
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(4, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['наименование', 'название', 'товар', 'product', 'name']):
                    structure['product_name_column'] = col
                    break
        
        # Ищем колонку с характеристиками
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(4, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['характеристики', 'описание', 'материал', 'characteristics', 'description']):
                    structure['characteristics_column'] = col
                    break
        
        # Ищем колонку с кастомом
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(4, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['кастом', 'дизайн', 'custom', 'design']):
                    structure['custom_design_column'] = col
                    break
        
        # Ищем колонку с тиражом
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(4, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['тираж', 'количество', 'quantity', 'шт', 'штук']):
                    structure['quantity_column'] = col
                    break
        
        # Ищем колонки с маршрутами доставки
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(4, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_value = str(cell.value).lower()
                    if 'жд' in cell_value or 'железнодорожн' in cell_value:
                        structure['route_columns']['ЖД'] = col
                    elif 'авиа' in cell_value or 'самолет' in cell_value:
                        structure['route_columns']['АВИА'] = col
                    elif 'контейнер' in cell_value:
                        structure['route_columns']['Контейнер'] = col
                    elif 'образец' in cell_value:
                        structure['route_columns']['Образец'] = col
                    elif 'доставка' in cell_value and 'жд' in cell_value:
                        structure['route_columns']['ЖД'] = col
                    elif 'доставка' in cell_value and 'авиа' in cell_value:
                        structure['route_columns']['АВИА'] = col
        
        # Ищем ценовые колонки для каждого маршрута
        for route, route_col in structure['route_columns'].items():
            structure['price_columns'][route] = {}
            
            # Ищем USD цены (обычно в той же колонке или следующей)
            for col in range(route_col, min(route_col + 3, ws.max_column + 1)):
                for row in range(1, min(4, ws.max_row + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and ('$' in str(cell.value) or 'usd' in str(cell.value).lower()):
                        structure['price_columns'][route]['usd'] = col
                        break
            
            # Ищем RUB цены
            for col in range(route_col, min(route_col + 3, ws.max_column + 1)):
                for row in range(1, min(4, ws.max_row + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and ('₽' in str(cell.value) or 'rub' in str(cell.value).lower() or 'руб' in str(cell.value).lower()):
                        structure['price_columns'][route]['rub'] = col
                        break
            
            # Ищем колонки со сроками
            for col in range(route_col, min(route_col + 3, ws.max_column + 1)):
                for row in range(1, min(4, ws.max_row + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and any(keyword in str(cell.value).lower() for keyword in ['срок', 'дней', 'к.д.', 'дн']):
                        structure['price_columns'][route]['delivery_time'] = col
                        break
        
        print(f"  ✅ Структура определена: {len(structure['route_columns'])} маршрутов, {len(structure['image_columns'])} колонок с изображениями")
        return structure
    
    def parse_products(self, file_path: str, structure: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Парсит товары из Excel файла"""
        print(f"📦 Парсинг товаров из: {os.path.basename(file_path)}")
        
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        products = []
        current_product = None
        
        # Начинаем с 4-й строки (после заголовков)
        for row in range(4, ws.max_row + 1):
            # Получаем название товара
            name_cell = ws.cell(row=row, column=structure['product_name_column'])
            product_name = name_cell.value
            
            if not product_name or str(product_name).strip() == '':
                # Если название пустое, это продолжение предыдущего товара
                if current_product:
                    # Добавляем ценовое предложение к текущему товару
                    self._add_price_offer_to_product(ws, row, structure, current_product)
                continue
            
            # Новый товар
            if current_product:
                products.append(current_product)
            
            current_product = {
                'name': str(product_name).strip(),
                'characteristics': '',
                'custom_design': '',
                'price_offers': [],
                'images': []
            }
            
            # Получаем характеристики
            if structure['characteristics_column']:
                char_cell = ws.cell(row=row, column=structure['characteristics_column'])
                if char_cell.value:
                    current_product['characteristics'] = str(char_cell.value).strip()
            
            # Получаем кастом
            if structure['custom_design_column']:
                custom_cell = ws.cell(row=row, column=structure['custom_design_column'])
                if custom_cell.value:
                    current_product['custom_design'] = str(custom_cell.value).strip()
            
            # Добавляем ценовое предложение
            self._add_price_offer_to_product(ws, row, structure, current_product)
        
        # Добавляем последний товар
        if current_product:
            products.append(current_product)
        
        print(f"  ✅ Найдено товаров: {len(products)}")
        return products
    
    def _add_price_offer_to_product(self, ws, row: int, structure: Dict[str, Any], product: Dict[str, Any]):
        """Добавляет ценовое предложение к товару"""
        # Получаем тираж из ПРАВИЛЬНОЙ колонки (не из образца!)
        quantity = None
        if structure['quantity_column']:
            qty_cell = ws.cell(row=row, column=structure['quantity_column'])
            if qty_cell.value:
                try:
                    quantity_str = str(qty_cell.value).replace(' ', '').replace(',', '')
                    quantity = int(float(quantity_str))
                except (ValueError, TypeError):
                    quantity = None
        
        # Создаем ценовые предложения для каждого маршрута
        for route_name, route_col in structure['route_columns'].items():
            price_offer = {
                'route_name': route_name,
                'quantity': quantity,
                'price_usd': None,
                'price_rub': None,
                'delivery_time': None,
                'is_sample': route_name == 'Образец',
                'sample_price': None
            }
            
            # Получаем цены
            if route_name in structure['price_columns']:
                route_prices = structure['price_columns'][route_name]
                
                # USD цена
                if 'usd' in route_prices:
                    usd_cell = ws.cell(row=row, column=route_prices['usd'])
                    if usd_cell.value:
                        try:
                            price_offer['price_usd'] = float(usd_cell.value)
                        except (ValueError, TypeError):
                            pass
                
                # RUB цена
                if 'rub' in route_prices:
                    rub_cell = ws.cell(row=row, column=route_prices['rub'])
                    if rub_cell.value:
                        try:
                            price_offer['price_rub'] = float(rub_cell.value)
                        except (ValueError, TypeError):
                            pass
                
                # Срок доставки
                if 'delivery_time' in route_prices:
                    time_cell = ws.cell(row=row, column=route_prices['delivery_time'])
                    if time_cell.value:
                        price_offer['delivery_time'] = str(time_cell.value).strip()
            
            # Для образцов используем RUB цену как sample_price
            if price_offer['is_sample'] and price_offer['price_rub']:
                price_offer['sample_price'] = price_offer['price_rub']
                price_offer['price_rub'] = None
            
            # Добавляем только если есть хотя бы одна цена
            if price_offer['price_usd'] or price_offer['price_rub'] or price_offer['sample_price']:
                product['price_offers'].append(price_offer)
    
    def extract_images(self, file_path: str, structure: Dict[str, Any], output_dir: str) -> List[str]:
        """Извлекает изображения из Excel файла с правильным форматом"""
        print(f"🖼️  Извлечение изображений из: {os.path.basename(file_path)}")
        
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        os.makedirs(output_dir, exist_ok=True)
        image_files = []
        
        for image in ws._images:
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                row = image.anchor._from.row + 1  # openpyxl использует 0-based индексы
                col = image.anchor._from.col + 1
                
                # Проверяем, что изображение в нужной колонке
                if col in structure['image_columns']:
                    filename = f"product_row_{row}_col_{col}_{len(image_files)}.jpg"
                    filepath = os.path.join(output_dir, filename)
                    
                    try:
                        # Получаем данные изображения
                        image_data = image.ref.read()
                        
                        # Создаем изображение с помощью PIL
                        img = Image.open(io.BytesIO(image_data))
                        
                        # Конвертируем в RGB если нужно
                        if img.mode in ('RGBA', 'LA', 'P'):
                            img = img.convert('RGB')
                        
                        # Сохраняем как JPEG
                        img.save(filepath, 'JPEG', quality=95)
                        
                        image_files.append(filepath)
                        print(f"  ✅ Сохранено: {filename} (строка {row}, колонка {col})")
                    except Exception as e:
                        print(f"  ❌ Ошибка сохранения {filename}: {e}")
        
        print(f"  ✅ Извлечено изображений: {len(image_files)}")
        return image_files

def main():
    """Основная функция"""
    parser = FixedExcelParser()
    
    # Анализируем все Excel файлы
    excel_dir = "storage/excel_files"
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    all_products = []
    
    for excel_file in excel_files:
        file_path = os.path.join(excel_dir, excel_file)
        
        try:
            # Анализируем структуру
            structure = parser.analyze_structure(file_path)
            
            # Парсим товары
            products = parser.parse_products(file_path, structure)
            
            # Извлекаем изображения
            image_dir = f"storage/images/products_{os.path.splitext(excel_file)[0]}"
            images = parser.extract_images(file_path, structure, image_dir)
            
            # Добавляем информацию о файле
            for product in products:
                product['source_file'] = excel_file
                product['structure'] = structure
            
            all_products.extend(products)
            
        except Exception as e:
            print(f"❌ Ошибка при обработке {excel_file}: {e}")
            continue
    
    # Сохраняем результаты
    with open('fixed_parsed_products.json', 'w', encoding='utf-8') as f:
        json.dump(all_products, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Парсинг завершен. Всего товаров: {len(all_products)}")
    print("📄 Результаты сохранены в fixed_parsed_products.json")

if __name__ == "__main__":
    main()
