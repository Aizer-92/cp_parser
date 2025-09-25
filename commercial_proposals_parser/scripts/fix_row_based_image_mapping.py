#!/usr/bin/env python3
"""
Исправление привязки изображений по строкам - изображения в одной строке привязываются к товару из этой строки
"""

import os
import sys
import zipfile
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import json

# Добавляем путь к модулям
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL

def clear_existing_images():
    """Очищает существующие изображения в базе данных"""
    print("🧹 Очищаем существующие изображения...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import ProductImage
        
        # Удаляем все записи об изображениях
        deleted_count = session.query(ProductImage).count()
        session.query(ProductImage).delete()
        session.commit()
        print(f"✅ Удалено записей об изображениях из БД: {deleted_count}")
    
    # Очищаем папку с изображениями
    images_dir = "storage/images/products_parsed"
    if os.path.exists(images_dir):
        shutil.rmtree(images_dir)
        print(f"✅ Очищена папка {images_dir}")
    
    os.makedirs(images_dir, exist_ok=True)
    print(f"✅ Создана новая папка {images_dir}")

def find_product_name_in_row(ws, row, search_columns=range(1, 20)):
    """Ищет название товара в указанной строке"""
    
    # Список возможных названий товаров
    product_keywords = [
        'ежедневник', 'худи', 'шапка', 'бини', 'шоппер', 'брелок', 'таблетница', 
        'кардхолдер', 'зонт', 'письменный', 'набор', 'термокружка', 'сумка',
        'футболка', 'кружка', 'ручка', 'карандаш', 'футляр', 'обложка', 'косметичка',
        'чехол', 'ноутбук', 'дутый', 'обложка для паспорта', 'набор карандашей'
    ]
    
    best_match = None
    best_score = 0
    
    # Ищем в указанных колонках строки
    for col in search_columns:
        try:
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip()
            if not cell_text or cell_text == 'None':
                continue
                
            # Проверяем, содержит ли ячейка название товара
            for keyword in product_keywords:
                if keyword.lower() in cell_text.lower():
                    # Вычисляем "качество" совпадения
                    score = len(keyword) / len(cell_text) if cell_text else 0
                    
                    if score > best_score:
                        best_score = score
                        best_match = {
                            'name': cell_text,
                            'row': row,
                            'col': col,
                            'score': score
                        }
        except:
            continue
    
    return best_match

def extract_images_from_excel_zip(excel_path, output_dir):
    """Извлекает изображения из Excel файла через zipfile"""
    print(f"📤 Извлекаем изображения из {excel_path}")
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and f.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
            
            print(f"  🖼️  Найдено изображений в архиве: {len(image_files)}")
            
            extracted_images = []
            
            for i, image_file in enumerate(image_files):
                try:
                    image_data = zip_ref.read(image_file)
                    original_name = os.path.basename(image_file)
                    new_name = f"image_{i+1}_{original_name}"
                    output_path = os.path.join(output_dir, new_name)
                    
                    with open(output_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_images.append({
                        'original_path': image_file,
                        'saved_path': output_path,
                        'filename': new_name,
                        'index': i + 1
                    })
                    
                    print(f"    ✅ Извлечено: {original_name} -> {new_name}")
                    
                except Exception as e:
                    print(f"    ❌ Ошибка при извлечении {image_file}: {e}")
            
            return extracted_images
            
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений из {excel_path}: {e}")
        return []

def analyze_excel_by_rows(excel_path):
    """Анализирует Excel файл по строкам - группирует изображения по строкам"""
    print(f"🔍 Анализ по строкам: {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        analysis = {
            'file_path': excel_path,
            'sheets': {},
            'row_images': {}  # Словарь: строка -> список изображений
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📊 Лист: {sheet_name}")
            
            sheet_info = {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'images_by_row': {}
            }
            
            # Группируем изображения по строкам
            for i, image in enumerate(ws._images):
                try:
                    if hasattr(image, 'anchor') and image.anchor:
                        anchor = image.anchor
                        
                        if hasattr(anchor, '_from'):
                            from_cell = anchor._from
                            row = from_cell.row + 1
                            col = from_cell.col + 1
                        else:
                            row = getattr(anchor, 'row', 1)
                            col = getattr(anchor, 'col', 1)
                        
                        # Добавляем изображение к строке
                        if row not in sheet_info['images_by_row']:
                            sheet_info['images_by_row'][row] = []
                        
                        sheet_info['images_by_row'][row].append({
                            'index': i + 1,
                            'row': row,
                            'col': col,
                            'width': getattr(image, 'width', 0),
                            'height': getattr(image, 'height', 0)
                        })
                        
                        print(f"    🖼️  Строка {row}, колонка {col}: изображение {i+1}")
                        
                except Exception as e:
                    print(f"    ❌ Ошибка при анализе изображения {i+1}: {e}")
            
            # Для каждой строки с изображениями ищем название товара
            for row, images in sheet_info['images_by_row'].items():
                product_match = find_product_name_in_row(ws, row)
                
                if product_match:
                    print(f"    🎯 Строка {row}: '{product_match['name']}' (колонка {product_match['col']})")
                    
                    # Добавляем информацию о товаре к изображениям
                    for img in images:
                        img['product_name'] = product_match['name']
                        img['product_col'] = product_match['col']
                        img['product_score'] = product_match['score']
                else:
                    print(f"    ⚠️  Строка {row}: товар не найден")
                    
                    # Добавляем пустое название
                    for img in images:
                        img['product_name'] = f'Unknown_row_{row}'
                        img['product_col'] = 0
                        img['product_score'] = 0
            
            analysis['sheets'][sheet_name] = sheet_info
            
            # Собираем все изображения по строкам
            for row, images in sheet_info['images_by_row'].items():
                if row not in analysis['row_images']:
                    analysis['row_images'][row] = []
                analysis['row_images'][row].extend(images)
        
        return analysis
        
    except Exception as e:
        print(f"❌ Ошибка при анализе {excel_path}: {e}")
        return None

def create_row_based_image_mapping(excel_files):
    """Создает маппинг изображений по строкам"""
    print("🗺️  Создаем маппинг изображений по строкам...")
    
    all_images = []
    
    for excel_file in excel_files:
        if os.path.exists(excel_file):
            print(f"\n📁 Обрабатываем: {excel_file}")
            
            # Анализируем структуру по строкам
            analysis = analyze_excel_by_rows(excel_file)
            if analysis:
                # Извлекаем изображения
                output_dir = f"storage/images/products_parsed/{os.path.basename(excel_file).replace('.xlsx', '')}"
                os.makedirs(output_dir, exist_ok=True)
                
                extracted_images = extract_images_from_excel_zip(excel_file, output_dir)
                
                # Связываем извлеченные изображения с анализом по строкам
                image_index = 0
                for row, images in analysis['row_images'].items():
                    for img_info in images:
                        if image_index < len(extracted_images):
                            extracted = extracted_images[image_index]
                            
                            all_images.append({
                                'excel_file': excel_file,
                                'sheet_name': analysis['sheets']['Copy of Просчет']['name'],
                                'row': row,
                                'col': img_info['col'],
                                'product_name': img_info['product_name'],
                                'product_col': img_info['product_col'],
                                'product_score': img_info['product_score'],
                                'image_path': extracted['saved_path'],
                                'image_filename': extracted['filename']
                            })
                            
                            image_index += 1
    
    return all_images

def map_images_to_products_by_rows(image_mapping):
    """Привязывает изображения к товарам по строкам"""
    print("🔗 Привязываем изображения к товарам по строкам...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import Product, ProductImage
        
        # Получаем все товары
        products = session.query(Product).all()
        print(f"📦 Найдено товаров в БД: {len(products)}")
        
        # Создаем словарь для поиска товаров
        products_by_name = {}
        for product in products:
            normalized_name = product.name.strip().lower()
            if normalized_name not in products_by_name:
                products_by_name[normalized_name] = []
            products_by_name[normalized_name].append(product)
        
        # Группируем изображения по строкам
        row_groups = {}
        for image_info in image_mapping:
            row = image_info['row']
            if row not in row_groups:
                row_groups[row] = []
            row_groups[row].append(image_info)
        
        print(f"📊 Групп строк с изображениями: {len(row_groups)}")
        
        mapped_count = 0
        
        for row, images in row_groups.items():
            print(f"\n📋 Строка {row}: {len(images)} изображений")
            
            # Берем первое изображение для определения товара
            first_image = images[0]
            product_name = first_image['product_name']
            
            if product_name.startswith('Unknown_'):
                print(f"  ⚠️  Товар не найден: {product_name}")
                continue
            
            # Ищем товар по названию
            normalized_name = product_name.lower()
            matching_products = products_by_name.get(normalized_name, [])
            
            if not matching_products:
                # Пробуем найти по частичному совпадению
                for norm_name, prods in products_by_name.items():
                    if normalized_name in norm_name or norm_name in normalized_name:
                        matching_products = prods
                        break
            
            if matching_products:
                product = matching_products[0]
                print(f"  ✅ Товар найден: {product_name} -> {product.name} #{product.id}")
                
                # Привязываем все изображения из этой строки к товару
                for i, image_info in enumerate(images):
                    new_image_name = f"product_{product.id}_row_{row}_{i+1}.jpg"
                    new_image_path = f"storage/images/products_parsed/{new_image_name}"
                    
                    try:
                        # Копируем изображение
                        shutil.copy2(image_info['image_path'], new_image_path)
                        
                        # Создаем запись в БД
                        image_type = 'main' if i == 0 else 'additional'
                        product_image = ProductImage(
                            product_id=product.id,
                            local_path=new_image_path,
                            image_type=image_type
                        )
                        session.add(product_image)
                        mapped_count += 1
                        
                        print(f"    🖼️  {image_info['image_filename']} -> {new_image_name} ({image_type})")
                        
                    except Exception as e:
                        print(f"    ❌ Ошибка при копировании изображения: {e}")
            else:
                print(f"  ⚠️  Товар не найден: '{product_name}'")
        
        session.commit()
        print(f"\n✅ Привязано изображений: {mapped_count}")

def main():
    """Основная функция"""
    print("🚀 Исправление привязки изображений по строкам")
    
    # Очищаем существующие данные
    clear_existing_images()
    
    # Список Excel файлов для обработки
    excel_files = [
        "storage/excel_files/original_sheet.xlsx",
        "storage/excel_files/Вторая таблица_1757933430.xlsx", 
        "storage/excel_files/Мерч для Sense_1757934153.xlsx"
    ]
    
    # Создаем маппинг изображений по строкам
    image_mapping = create_row_based_image_mapping(excel_files)
    
    print(f"\n📊 Всего найдено изображений: {len(image_mapping)}")
    
    # Привязываем изображения к товарам по строкам
    map_images_to_products_by_rows(image_mapping)
    
    print("\n✅ Исправление привязки изображений по строкам завершено!")

if __name__ == "__main__":
    main()
