#!/usr/bin/env python3
"""
Финальный правильный парсер для v4 структуры
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl
from pathlib import Path
from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL
import re

def parse_excel_final(excel_path):
    """Финальный правильный парсинг Excel файла"""
    
    # Подключение к базе данных
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    # Путь к Excel файлу
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        print(f"❌ Файл не найден: {excel_path}")
        return
    
    print(f"🔄 Парсинг Excel файла: {excel_path}")
    
    # Читаем Excel файл с правильными параметрами
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        print(f"📊 Размер файла: {ws.max_row} строк x {ws.max_column} колонок")
        
        # Определяем колонки по структуре
        # Строка 2: Фото, Наименование, Характеристики, Кастом, Тираж, Доставка ЖД, ..., Доставка АВИА
        # Строка 3: ..., Цена за шт., $, Цена за шт., руб, Срок тиража, к.д., Цена за шт., $
        
        product_name_col = 2  # B - Наименование
        characteristics_col = 3  # C - Характеристики
        custom_design_col = 4  # D - Кастом
        quantity_col = 5  # E - Тираж
        
        # Варианты доставки
        delivery_variants = {
            'ЖД': {
                'price_usd_col': 6,  # F - Цена за шт., $
                'price_rub_col': 7,  # G - Цена за шт., руб
                'delivery_time_col': 8  # H - Срок тиража, к.д.
            },
            'АВИА': {
                'price_usd_col': 9,  # I - Цена за шт., $
                'price_rub_col': 10,  # J - Цена за шт., руб
                'delivery_time_col': 11  # K - Срок тиража, к.д.
            },
            'Контейнер': {
                'price_usd_col': 12,  # L - Цена за шт., $
                'price_rub_col': 13,  # M - Цена за шт., руб
                'delivery_time_col': 14  # N - Срок тиража, к.д.
            }
        }
        
        print(f"🔍 Структура колонок:")
        print(f"  - Название товара: {product_name_col} (B)")
        print(f"  - Характеристики: {characteristics_col} (C)")
        print(f"  - Кастом: {custom_design_col} (D)")
        print(f"  - Тираж: {quantity_col} (E)")
        print(f"🚚 Варианты доставки:")
        for variant, cols in delivery_variants.items():
            print(f"  - {variant}: USD={cols['price_usd_col']}, RUB={cols['price_rub_col']}, Срок={cols['delivery_time_col']}")
        
        # Обрабатываем данные начиная с 4-й строки (индекс 4)
        products_created = 0
        offers_created = 0
        
        for row_num in range(4, ws.max_row + 1):
            # Получаем название товара
            product_name_cell = ws.cell(row=row_num, column=product_name_col)
            product_name = product_name_cell.value
            
            if not product_name or str(product_name).strip() == '' or str(product_name).strip() == ' ':
                continue
            
            product_name = str(product_name).strip()
            
            # Получаем характеристики
            characteristics_cell = ws.cell(row=row_num, column=characteristics_col)
            characteristics = str(characteristics_cell.value).strip() if characteristics_cell.value else None
            
            # Получаем кастом
            custom_design_cell = ws.cell(row=row_num, column=custom_design_col)
            custom_design = str(custom_design_cell.value).strip() if custom_design_cell.value else None
            
            # Получаем тираж
            quantity_cell = ws.cell(row=row_num, column=quantity_col)
            quantity = None
            if quantity_cell.value:
                try:
                    # Убираем пробелы и преобразуем в число
                    quantity_str = str(quantity_cell.value).replace(' ', '').replace(',', '')
                    quantity = int(float(quantity_str))
                except:
                    pass
            
            print(f"📦 Обрабатываем товар: {product_name}")
            if quantity:
                print(f"  📊 Тираж: {quantity:,} шт")
            
            # Создаем товар
            product = db.create_product(
                name=product_name,
                description=characteristics,
                custom_design=custom_design,
                sheet_id=1
            )
            products_created += 1
            
            # Создаем ценовые предложения для каждого варианта доставки
            for variant, cols in delivery_variants.items():
                # Получаем цену в USD
                price_usd_cell = ws.cell(row=row_num, column=cols['price_usd_col'])
                price_usd = None
                if price_usd_cell.value:
                    try:
                        price_usd = float(price_usd_cell.value)
                    except:
                        pass
                
                # Получаем цену в RUB
                price_rub_cell = ws.cell(row=row_num, column=cols['price_rub_col'])
                price_rub = None
                if price_rub_cell.value:
                    try:
                        price_rub = float(price_rub_cell.value)
                    except:
                        pass
                
                # Получаем срок доставки
                delivery_time_cell = ws.cell(row=row_num, column=cols['delivery_time_col'])
                delivery_time = str(delivery_time_cell.value).strip() if delivery_time_cell.value else None
                
                # Создаем ценовое предложение только если есть хотя бы одна цена
                if price_usd is not None or price_rub is not None:
                    offer = db.create_price_offer(
                        product_id=product.id,
                        route_name=variant,
                        quantity=quantity,
                        price_usd=price_usd,
                        price_rub=price_rub,
                        delivery_time=delivery_time
                    )
                    offers_created += 1
                    
                    print(f"  ✅ {variant}: ${price_usd} / {price_rub} руб / {delivery_time} дней")
        
        print(f"✅ Парсинг завершен!")
        print(f"📊 Результаты:")
        print(f"  - Создано товаров: {products_created}")
        print(f"  - Создано ценовых предложений: {offers_created}")
        
        # Статистика
        stats = db.get_statistics()
        print(f"\n📈 Статистика базы данных:")
        print(f"  - Товаров: {stats['total_products']}")
        print(f"  - Ценовых предложений: {stats['total_price_offers']}")
        print(f"  - Изображений: {stats['total_images']}")
        
    except Exception as e:
        print(f"❌ Ошибка при парсинге: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    parse_excel_final()
