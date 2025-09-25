#!/usr/bin/env python3
"""
Исправление парсинга изображений - правильная привязка к товарам
"""

import os
import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import shutil

# Добавляем путь к модулям
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL

def analyze_excel_images(excel_path):
    """Анализирует изображения в Excel файле"""
    print(f"🔍 Анализируем изображения в {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        images_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📊 Лист: {sheet_name}")
            
            # Получаем все изображения на листе
            for image in ws._images:
                if hasattr(image, 'anchor') and image.anchor:
                    # Получаем ячейку, к которой привязано изображение
                    cell_ref = image.anchor._from
                    row = cell_ref.row + 1  # openpyxl использует 0-based индексы
                    col = cell_ref.col + 1
                    
                    # Получаем значение ячейки (название товара)
                    cell_value = ws.cell(row=row, column=col).value
                    
                    # Получаем информацию об изображении
                    image_info = {
                        'sheet_name': sheet_name,
                        'row': row,
                        'col': col,
                        'cell_value': str(cell_value) if cell_value else '',
                        'image_name': getattr(image, 'filename', f'image_{row}_{col}'),
                        'width': getattr(image, 'width', 0),
                        'height': getattr(image, 'height', 0)
                    }
                    
                    images_data.append(image_info)
                    print(f"    🖼️  Строка {row}, колонка {col}: '{cell_value}' -> {image_info['image_name']}")
        
        return images_data
        
    except Exception as e:
        print(f"❌ Ошибка при анализе {excel_path}: {e}")
        return []

def extract_images_from_excel(excel_path, output_dir):
    """Извлекает изображения из Excel файла"""
    print(f"📤 Извлекаем изображения из {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        extracted_images = []
        
        # Создаем папку для изображений
        os.makedirs(output_dir, exist_ok=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📊 Лист: {sheet_name}")
            
            for i, image in enumerate(ws._images):
                try:
                    # Создаем уникальное имя файла
                    image_name = f"{sheet_name}_{i+1}.jpg"
                    image_path = os.path.join(output_dir, image_name)
                    
                    # Сохраняем изображение
                    if hasattr(image, 'ref') and image.ref:
                        # Копируем изображение
                        shutil.copy2(image.ref.path, image_path)
                        
                        # Получаем информацию о ячейке
                        if hasattr(image, 'anchor') and image.anchor:
                            cell_ref = image.anchor._from
                            row = cell_ref.row + 1
                            col = cell_ref.col + 1
                            cell_value = ws.cell(row=row, column=col).value
                            
                            image_info = {
                                'file_path': image_path,
                                'sheet_name': sheet_name,
                                'row': row,
                                'col': col,
                                'cell_value': str(cell_value) if cell_value else '',
                                'product_name': str(cell_value) if cell_value else f'Unknown_{i+1}'
                            }
                            
                            extracted_images.append(image_info)
                            print(f"    ✅ Извлечено: {image_name} -> '{cell_value}'")
                        else:
                            print(f"    ⚠️  Изображение без привязки к ячейке: {image_name}")
                            
                except Exception as e:
                    print(f"    ❌ Ошибка при извлечении изображения {i+1}: {e}")
        
        return extracted_images
        
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений из {excel_path}: {e}")
        return []

def clear_existing_images():
    """Очищает существующие изображения в базе данных"""
    print("🧹 Очищаем существующие изображения...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import ProductImage
        
        # Удаляем все записи об изображениях
        session.query(ProductImage).delete()
        session.commit()
        print("✅ Очищены записи об изображениях в БД")
    
    # Очищаем папку с изображениями
    images_dir = "storage/images/products_correct"
    if os.path.exists(images_dir):
        shutil.rmtree(images_dir)
        print(f"✅ Очищена папка {images_dir}")
    
    os.makedirs(images_dir, exist_ok=True)
    print(f"✅ Создана новая папка {images_dir}")

def map_images_to_products(extracted_images):
    """Привязывает изображения к товарам в базе данных"""
    print("🔗 Привязываем изображения к товарам...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import Product, SheetMetadata, ProductImage
        
        # Получаем все товары
        products = session.query(Product).all()
        print(f"📦 Найдено товаров в БД: {len(products)}")
        
        # Создаем словарь для быстрого поиска товаров по названию
        products_by_name = {}
        for product in products:
            # Нормализуем название для поиска
            normalized_name = product.name.strip().lower()
            if normalized_name not in products_by_name:
                products_by_name[normalized_name] = []
            products_by_name[normalized_name].append(product)
        
        mapped_count = 0
        
        for image_info in extracted_images:
            product_name = image_info['product_name'].strip()
            if not product_name or product_name == 'None':
                continue
                
            # Ищем товар по названию
            normalized_name = product_name.lower()
            matching_products = products_by_name.get(normalized_name, [])
            
            if not matching_products:
                # Пробуем найти по частичному совпадению
                for norm_name, prods in products_by_name.items():
                    if normalized_name in norm_name or norm_name in normalized_name:
                        matching_products = prods
                        break
            
            if matching_products:
                # Берем первый подходящий товар
                product = matching_products[0]
                
                # Копируем изображение в правильную папку
                new_image_name = f"product_{product.id}_{mapped_count}.jpg"
                new_image_path = f"storage/images/products_correct/{new_image_name}"
                
                try:
                    shutil.copy2(image_info['file_path'], new_image_path)
                    
                    # Создаем запись в БД
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=new_image_path,
                        image_type='main' if mapped_count == 0 else 'additional'
                    )
                    session.add(product_image)
                    mapped_count += 1
                    
                    print(f"  ✅ {product_name} -> {product.name} #{product.id} -> {new_image_name}")
                    
                except Exception as e:
                    print(f"  ❌ Ошибка при копировании изображения: {e}")
            else:
                print(f"  ⚠️  Товар не найден: '{product_name}'")
        
        session.commit()
        print(f"✅ Привязано изображений: {mapped_count}")

def main():
    """Основная функция"""
    print("🚀 Исправление парсинга изображений")
    
    # Очищаем существующие данные
    clear_existing_images()
    
    # Анализируем все Excel файлы
    excel_files = [
        "storage/excel_files/original_sheet.xlsx",
        "storage/excel_files/Вторая таблица_1757933430.xlsx", 
        "storage/excel_files/Мерч для Sense_1757934153.xlsx"
    ]
    
    all_extracted_images = []
    
    for excel_file in excel_files:
        if os.path.exists(excel_file):
            print(f"\n📁 Обрабатываем: {excel_file}")
            
            # Анализируем структуру
            images_data = analyze_excel_images(excel_file)
            
            # Извлекаем изображения
            extracted_images = extract_images_from_excel(excel_file, "storage/images/temp_extraction")
            all_extracted_images.extend(extracted_images)
        else:
            print(f"⚠️  Файл не найден: {excel_file}")
    
    print(f"\n📊 Всего извлечено изображений: {len(all_extracted_images)}")
    
    # Привязываем изображения к товарам
    map_images_to_products(all_extracted_images)
    
    # Очищаем временную папку
    if os.path.exists("storage/images/temp_extraction"):
        shutil.rmtree("storage/images/temp_extraction")
        print("🧹 Очищена временная папка")
    
    print("\n✅ Исправление парсинга изображений завершено!")

if __name__ == "__main__":
    main()
