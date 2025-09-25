#!/usr/bin/env python3
"""
Специальный парсер для таблицы "Мерч для Sense"
"""

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
import json

def parse_merch_sense():
    """Парсит таблицу Мерч для Sense"""
    file_path = "storage/excel_files/Мерч для Sense_1757934153.xlsx"
    
    print(f"🔍 Парсинг таблицы: {os.path.basename(file_path)}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    products = []
    current_product = None
    
    # Начинаем с 4-й строки (после заголовков)
    for row in range(4, ws.max_row + 1):
        # Получаем название товара (колонка 3)
        name_cell = ws.cell(row=row, column=3)
        product_name = name_cell.value
        
        if not product_name or str(product_name).strip() == '':
            # Если название пустое, это продолжение предыдущего товара
            if current_product:
                # Добавляем ценовое предложение к текущему товару
                add_price_offer_to_product(ws, row, current_product)
            continue
        
        # Новый товар
        if current_product:
            products.append(current_product)
        
        current_product = {
            'name': str(product_name).strip(),
            'characteristics': '',
            'custom_design': '',
            'price_offers': [],
            'images': []
        }
        
        # Получаем характеристики (колонка 4)
        char_cell = ws.cell(row=row, column=4)
        if char_cell.value:
            current_product['characteristics'] = str(char_cell.value).strip()
        
        # Получаем кастом (колонка 5)
        custom_cell = ws.cell(row=row, column=5)
        if custom_cell.value:
            current_product['custom_design'] = str(custom_cell.value).strip()
        
        # Добавляем ценовое предложение
        print(f"  🔍 Обработка строки {row} для товара {current_product['name']}")
        add_price_offer_to_product(ws, row, current_product)
        print(f"    ✅ Добавлено ценовых предложений: {len(current_product['price_offers'])}")
    
    # Добавляем последний товар
    if current_product:
        products.append(current_product)
    
    print(f"✅ Найдено товаров: {len(products)}")
    
    # Показываем примеры
    for product in products[:3]:
        print(f"  - {product['name']}: {len(product['price_offers'])} ценовых предложений")
        for offer in product['price_offers'][:2]:
            print(f"    * {offer['route_name']}: USD={offer['price_usd']}, RUB={offer['price_rub']}")
    
    return products

def add_price_offer_to_product(ws, row: int, product: dict):
    """Добавляет ценовое предложение к товару"""
    # Получаем тираж (колонка 6)
    quantity = None
    qty_cell = ws.cell(row=row, column=6)
    if qty_cell.value:
        try:
            quantity_str = str(qty_cell.value).replace(' ', '').replace(',', '')
            quantity = int(float(quantity_str))
        except (ValueError, TypeError):
            quantity = None
    
    # Создаем ценовые предложения для каждого маршрута
    
    # ЖД маршрут (колонки 7, 8, 9)
    col7_value = ws.cell(row=row, column=7).value
    if col7_value is not None and str(col7_value).strip() != '':
        price_offer = {
            'route_name': 'ЖД',
            'quantity': quantity,
            'price_usd': None,
            'price_rub': None,
            'delivery_time': None,
            'is_sample': False,
            'sample_price': None
        }
        
        # USD цена (колонка 7)
        usd_cell = ws.cell(row=row, column=7)
        if usd_cell.value:
            try:
                # Заменяем запятую на точку для корректного парсинга
                usd_value = str(usd_cell.value).replace(',', '.')
                price_offer['price_usd'] = float(usd_value)
            except (ValueError, TypeError):
                pass
        
        # RUB цена (колонка 8)
        rub_cell = ws.cell(row=row, column=8)
        if rub_cell.value:
            try:
                # Заменяем запятую на точку для корректного парсинга
                rub_value = str(rub_cell.value).replace(',', '.')
                price_offer['price_rub'] = float(rub_value)
            except (ValueError, TypeError):
                pass
        
        # Срок доставки (колонка 9)
        time_cell = ws.cell(row=row, column=9)
        if time_cell.value:
            price_offer['delivery_time'] = str(time_cell.value).strip()
        
        # Добавляем только если есть хотя бы одна цена
        if price_offer['price_usd'] or price_offer['price_rub']:
            product['price_offers'].append(price_offer)
    
    # АВИА маршрут (колонки 10, 11, 12)
    if ws.cell(row=row, column=10).value is not None and str(ws.cell(row=row, column=10).value).strip() != '':
        price_offer = {
            'route_name': 'АВИА',
            'quantity': quantity,
            'price_usd': None,
            'price_rub': None,
            'delivery_time': None,
            'is_sample': False,
            'sample_price': None
        }
        
        # USD цена (колонка 10)
        usd_cell = ws.cell(row=row, column=10)
        if usd_cell.value:
            try:
                # Заменяем запятую на точку для корректного парсинга
                usd_value = str(usd_cell.value).replace(',', '.')
                price_offer['price_usd'] = float(usd_value)
            except (ValueError, TypeError):
                pass
        
        # RUB цена (колонка 11)
        rub_cell = ws.cell(row=row, column=11)
        if rub_cell.value:
            try:
                # Заменяем запятую на точку для корректного парсинга
                rub_value = str(rub_cell.value).replace(',', '.')
                price_offer['price_rub'] = float(rub_value)
            except (ValueError, TypeError):
                pass
        
        # Срок доставки (колонка 12)
        time_cell = ws.cell(row=row, column=12)
        if time_cell.value:
            price_offer['delivery_time'] = str(time_cell.value).strip()
        
        # Добавляем только если есть хотя бы одна цена
        if price_offer['price_usd'] or price_offer['price_rub']:
            product['price_offers'].append(price_offer)
    
    # Образец (колонка 13)
    if ws.cell(row=row, column=13).value is not None and str(ws.cell(row=row, column=13).value).strip() != '':
        price_offer = {
            'route_name': 'Образец',
            'quantity': quantity,
            'price_usd': None,
            'price_rub': None,
            'delivery_time': None,
            'is_sample': True,
            'sample_price': None
        }
        
        # RUB цена (колонка 13)
        rub_cell = ws.cell(row=row, column=13)
        if rub_cell.value:
            try:
                # Заменяем запятую на точку для корректного парсинга
                rub_value = str(rub_cell.value).replace(',', '.')
                price_offer['sample_price'] = float(rub_value)
            except (ValueError, TypeError):
                pass
        
        # Срок доставки (колонка 14)
        time_cell = ws.cell(row=row, column=14)
        if time_cell.value:
            price_offer['delivery_time'] = str(time_cell.value).strip()
        
        # Добавляем только если есть цена
        if price_offer['sample_price']:
            product['price_offers'].append(price_offer)

def main():
    """Основная функция"""
    products = parse_merch_sense()
    
    # Сохраняем результаты
    with open('merch_sense_products.json', 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Парсинг завершен. Товаров: {len(products)}")
    print("📄 Результаты сохранены в merch_sense_products.json")

if __name__ == "__main__":
    main()
