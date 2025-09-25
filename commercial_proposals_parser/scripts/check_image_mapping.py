#!/usr/bin/env python3
"""
Проверка соответствия изображений товарам
"""

import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

sys.path.append('.')
from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL_V4

def check_image_mapping():
    """Проверка соответствия изображений товарам"""
    
    print("🔍 Проверяем соответствие изображений товарам...")
    
    # 1. Получаем товары из БД
    db = CommercialProposalsDB(DATABASE_URL_V4)
    products_with_details = db.get_all_products_with_details(limit=50)
    merch_sense_products = [(p, po, im) for p, po, im in products_with_details if p.id >= 29]
    
    print(f"📦 Товары 'Мерч для Sense' в БД:")
    for product, price_offers, images in merch_sense_products:
        main_images = [img for img in images if img.image_type == 'main']
        additional_images = [img for img in images if img.image_type == 'additional']
        print(f"  {product.id}. {product.name} - {len(main_images)} основных, {len(additional_images)} дополнительных")
        for img in main_images:
            print(f"    Основное: {img.local_path}")
    
    # 2. Анализируем Excel файл
    excel_file = "storage/excel_files/Мерч для Sense_1758096973.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл не найден: {excel_file}")
        return
    
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active
    
    print(f"\n📊 Товары в Excel файле:")
    for row_num in range(4, sheet.max_row + 1):
        cell_c = sheet.cell(row=row_num, column=3)  # Колонка C - название товара
        if cell_c.value and isinstance(cell_c.value, str) and cell_c.value.strip():
            product_name = cell_c.value.strip()
            if product_name != 'Наименование':
                print(f"  Строка {row_num}: {product_name}")
    
    # 3. Анализируем позиции изображений
    print(f"\n🖼️  Позиции изображений в Excel:")
    try:
        with zipfile.ZipFile(excel_file, 'r') as z:
            drawing_files = [name for name in z.namelist() if 'xl/drawings/drawing' in name]
            
            for drawing_file in drawing_files:
                try:
                    drawing_data = z.read(drawing_file)
                    root = ET.fromstring(drawing_data)
                    
                    for anchor in root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}oneCellAnchor'):
                        from_elem = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
                        if from_elem is not None:
                            row_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
                            col_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col')
                            if row_elem is not None and col_elem is not None:
                                row = int(row_elem.text) + 1
                                col = int(col_elem.text) + 1
                                
                                # Находим товар в этой строке
                                cell_c = sheet.cell(row=row, column=3)
                                product_name = cell_c.value if cell_c.value else f"Unknown_row_{row}"
                                
                                image_type = 'main' if col == 1 else 'additional' if col == 16 else 'other'
                                print(f"    Строка {row}, колонка {col} ({image_type}) -> {product_name}")
                except Exception as e:
                    print(f"    Ошибка при анализе {drawing_file}: {e}")
    except Exception as e:
        print(f"❌ Ошибка при работе с Excel: {e}")
    
    # 4. Проверяем соответствие
    print(f"\n🔗 Проверяем соответствие изображений товарам:")
    
    # Создаем словарь товаров по именам
    products_by_name = {}
    for product, price_offers, images in merch_sense_products:
        products_by_name[product.name.strip().lower()] = product
    
    # Проверяем каждое изображение
    for product, price_offers, images in merch_sense_products:
        main_images = [img for img in images if img.image_type == 'main']
        if main_images:
            img = main_images[0]
            print(f"  {product.name} (ID: {product.id})")
            print(f"    Изображение: {img.local_path}")
            
            # Проверяем, что файл существует
            if os.path.exists(img.local_path):
                file_size = os.path.getsize(img.local_path)
                print(f"    ✅ Файл существует, размер: {file_size} байт")
            else:
                print(f"    ❌ Файл не найден: {img.local_path}")

if __name__ == "__main__":
    check_image_mapping()
