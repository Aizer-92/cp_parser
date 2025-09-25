#!/usr/bin/env python3
"""
Продвинутый парсер изображений с точным позиционированием
Использует комбинацию openpyxl и zipfile для точного извлечения
"""

import os
import sys
import zipfile
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import json
import xml.etree.ElementTree as ET

# Добавляем путь к модулям
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL

def clear_existing_images():
    """Очищает существующие изображения в базе данных"""
    print("🧹 Очищаем существующие изображения...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import ProductImage
        
        # Удаляем все записи об изображениях
        deleted_count = session.query(ProductImage).count()
        session.query(ProductImage).delete()
        session.commit()
        print(f"✅ Удалено записей об изображениях из БД: {deleted_count}")
    
    # Очищаем папку с изображениями
    images_dir = "storage/images/products_parsed"
    if os.path.exists(images_dir):
        shutil.rmtree(images_dir)
        print(f"✅ Очищена папка {images_dir}")
    
    os.makedirs(images_dir, exist_ok=True)
    print(f"✅ Создана новая папка {images_dir}")

def extract_images_from_zip(excel_path, output_dir):
    """Извлекает изображения из Excel zip архива"""
    print(f"📤 Извлекаем изображения из zip: {excel_path}")
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            # Ищем изображения в папке xl/media/
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and f.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
            
            print(f"  🖼️  Найдено изображений в архиве: {len(image_files)}")
            
            extracted_images = []
            
            for i, image_file in enumerate(image_files):
                try:
                    image_data = zip_ref.read(image_file)
                    original_name = os.path.basename(image_file)
                    new_name = f"image_{i+1}_{original_name}"
                    output_path = os.path.join(output_dir, new_name)
                    
                    with open(output_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_images.append({
                        'original_path': image_file,
                        'saved_path': output_path,
                        'filename': new_name,
                        'index': i + 1
                    })
                    
                    print(f"    ✅ Извлечено: {original_name} -> {new_name}")
                    
                except Exception as e:
                    print(f"    ❌ Ошибка при извлечении {image_file}: {e}")
            
            return extracted_images
            
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений из {excel_path}: {e}")
        return []

def analyze_drawing_relationships(excel_path):
    """Анализирует связи между изображениями и ячейками через drawing relationships"""
    print(f"🔍 Анализ связей изображений: {excel_path}")
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            # Ищем файлы drawing relationships
            drawing_files = [f for f in zip_ref.namelist() if 'xl/drawings/drawing' in f and f.endswith('.xml')]
            
            print(f"  📊 Найдено drawing файлов: {len(drawing_files)}")
            
            image_relationships = []
            
            for drawing_file in drawing_files:
                try:
                    # Читаем XML файл
                    xml_content = zip_ref.read(drawing_file)
                    root = ET.fromstring(xml_content)
                    
                    # Ищем элементы с изображениями
                    for elem in root.iter():
                        if 'pic' in elem.tag or 'picture' in elem.tag.lower():
                            # Извлекаем информацию о позиции
                            for child in elem.iter():
                                if 'anchor' in child.tag.lower():
                                    # Получаем координаты
                                    for coord in child.iter():
                                        if 'col' in coord.tag.lower():
                                            col = int(coord.text) if coord.text else 0
                                        elif 'row' in coord.tag.lower():
                                            row = int(coord.text) if coord.text else 0
                                    
                                    image_relationships.append({
                                        'file': drawing_file,
                                        'row': row,
                                        'col': col
                                    })
                                    print(f"    🎯 Найдена связь: строка {row}, колонка {col}")
                    
                except Exception as e:
                    print(f"    ❌ Ошибка при анализе {drawing_file}: {e}")
                    continue
            
            return image_relationships
            
    except Exception as e:
        print(f"❌ Ошибка при анализе связей: {e}")
        return []

def analyze_excel_structure_advanced(excel_path):
    """Продвинутый анализ структуры Excel файла"""
    print(f"🔍 Продвинутый анализ структуры: {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        
        table_name = os.path.basename(excel_path).replace('.xlsx', '')
        
        analysis = {
            'file_path': excel_path,
            'table_name': table_name,
            'sheets': {},
            'products_found': [],
            'images_found': [],
            'image_relationships': []
        }
        
        # Анализируем связи изображений
        image_relationships = analyze_drawing_relationships(excel_path)
        analysis['image_relationships'] = image_relationships
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📊 Лист: {sheet_name}")
            
            sheet_info = {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'products': [],
                'images': []
            }
            
            # Ищем товары в первых 25 строках
            for row in range(1, min(26, ws.max_row + 1)):
                for col in range(1, min(26, ws.max_column + 1)):
                    try:
                        cell_value = ws.cell(row=row, column=col).value
                        if not cell_value:
                            continue
                            
                        cell_text = str(cell_value).strip()
                        if not cell_text or cell_text == 'None':
                            continue
                        
                        # Проверяем, является ли это названием товара
                        product_keywords = [
                            'ежедневник', 'худи', 'шапка', 'бини', 'шоппер', 'брелок', 'таблетница', 
                            'кардхолдер', 'зонт', 'письменный', 'набор', 'термокружка', 'сумка',
                            'футболка', 'кружка', 'ручка', 'карандаш', 'футляр', 'обложка', 'косметичка',
                            'чехол', 'ноутбук', 'дутый', 'обложка для паспорта', 'набор карандашей',
                            'футляр для очков', 'дутый чехол для ноутбука'
                        ]
                        
                        for keyword in product_keywords:
                            if keyword.lower() in cell_text.lower():
                                product_info = {
                                    'name': cell_text,
                                    'row': row,
                                    'col': col,
                                    'keyword': keyword,
                                    'sheet': sheet_name
                                }
                                sheet_info['products'].append(product_info)
                                analysis['products_found'].append(product_info)
                                print(f"    🎯 Товар: '{cell_text}' (строка {row}, колонка {col})")
                                break
                    except Exception as e:
                        continue
            
            # Анализируем изображения на листе
            for i, image in enumerate(ws._images):
                try:
                    if hasattr(image, 'anchor') and image.anchor:
                        anchor = image.anchor
                        
                        if hasattr(anchor, '_from'):
                            from_cell = anchor._from
                            row = from_cell.row + 1
                            col = from_cell.col + 1
                        else:
                            row = getattr(anchor, 'row', 1)
                            col = getattr(anchor, 'col', 1)
                        
                        image_info = {
                            'index': i + 1,
                            'row': row,
                            'col': col,
                            'width': getattr(image, 'width', 0),
                            'height': getattr(image, 'height', 0),
                            'sheet': sheet_name
                        }
                        
                        sheet_info['images'].append(image_info)
                        analysis['images_found'].append(image_info)
                        print(f"    🖼️  Изображение {i+1}: строка {row}, колонка {col}")
                        
                except Exception as e:
                    print(f"    ❌ Ошибка при анализе изображения {i+1}: {e}")
            
            analysis['sheets'][sheet_name] = sheet_info
        
        return analysis
        
    except Exception as e:
        print(f"❌ Ошибка при анализе {excel_path}: {e}")
        return None

def find_closest_product_to_image(products, image_row, image_col, max_distance=5):
    """Находит ближайший товар к изображению в пределах максимального расстояния"""
    if not products:
        return None
    
    best_match = None
    best_distance = float('inf')
    
    for product in products:
        # Вычисляем расстояние от изображения до товара
        distance = abs(product['row'] - image_row) + abs(product['col'] - image_col)
        
        if distance <= max_distance and distance < best_distance:
            best_distance = distance
            best_match = product
    
    return best_match

def process_single_excel_advanced(excel_path):
    """Обрабатывает один Excel файл продвинутым способом"""
    print(f"\n📁 Продвинутая обработка файла: {excel_path}")
    
    # Анализируем структуру
    analysis = analyze_excel_structure_advanced(excel_path)
    if not analysis:
        return []
    
    # Извлекаем изображения
    output_dir = f"storage/images/products_parsed/{analysis['table_name']}"
    os.makedirs(output_dir, exist_ok=True)
    
    extracted_images = extract_images_from_zip(excel_path, output_dir)
    
    # Связываем изображения с товарами
    linked_images = []
    
    for i, extracted in enumerate(extracted_images):
        if i < len(analysis['images_found']):
            image_info = analysis['images_found'][i]
            
            # Ищем ближайший товар к изображению
            closest_product = find_closest_product_to_image(
                analysis['products_found'], 
                image_info['row'], 
                image_info['col'],
                max_distance=3  # Уменьшаем максимальное расстояние
            )
            
            if closest_product:
                linked_images.append({
                    'excel_file': excel_path,
                    'table_name': analysis['table_name'],
                    'sheet_name': image_info['sheet'],
                    'row': image_info['row'],
                    'col': image_info['col'],
                    'product_name': closest_product['name'],
                    'product_row': closest_product['row'],
                    'product_col': closest_product['col'],
                    'distance': abs(closest_product['row'] - image_info['row']) + abs(closest_product['col'] - image_info['col']),
                    'image_path': extracted['saved_path'],
                    'image_filename': extracted['filename']
                })
                print(f"    ✅ {closest_product['name']} -> {extracted['filename']} (расстояние: {abs(closest_product['row'] - image_info['row']) + abs(closest_product['col'] - image_info['col'])})")
            else:
                print(f"    ⚠️  Товар не найден для изображения {extracted['filename']} (строка {image_info['row']}, колонка {image_info['col']})")
    
    return linked_images

def map_images_to_products_by_table(image_mapping):
    """Привязывает изображения к товарам с учетом таблицы"""
    print("🔗 Привязываем изображения к товарам с учетом таблицы...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import Product, ProductImage
        
        # Получаем все товары
        products = session.query(Product).all()
        print(f"📦 Найдено товаров в БД: {len(products)}")
        
        # Создаем словарь для поиска товаров по названию
        products_by_name = {}
        for product in products:
            normalized_name = product.name.strip().lower()
            if normalized_name not in products_by_name:
                products_by_name[normalized_name] = []
            products_by_name[normalized_name].append(product)
        
        # Группируем изображения по таблицам
        table_groups = {}
        for image_info in image_mapping:
            table_name = image_info['table_name']
            if table_name not in table_groups:
                table_groups[table_name] = []
            table_groups[table_name].append(image_info)
        
        print(f"📊 Групп таблиц: {len(table_groups)}")
        
        mapped_count = 0
        
        for table_name, images in table_groups.items():
            print(f"\n📋 Таблица: {table_name} ({len(images)} изображений)")
            
            for image_info in images:
                product_name = image_info['product_name'].strip()
                
                # Ищем товар по названию
                normalized_name = product_name.lower()
                matching_products = products_by_name.get(normalized_name, [])
                
                if not matching_products:
                    # Пробуем найти по частичному совпадению
                    for norm_name, prods in products_by_name.items():
                        if normalized_name in norm_name or norm_name in normalized_name:
                            matching_products = prods
                            break
                
                if matching_products:
                    product = matching_products[0]
                    
                    # Создаем новое имя файла с указанием таблицы
                    new_image_name = f"product_{product.id}_{table_name}_{mapped_count + 1}.jpg"
                    new_image_path = f"storage/images/products_parsed/{new_image_name}"
                    
                    try:
                        # Копируем изображение
                        shutil.copy2(image_info['image_path'], new_image_path)
                        
                        # Создаем запись в БД
                        product_image = ProductImage(
                            product_id=product.id,
                            local_path=new_image_path,
                            image_type='main' if mapped_count == 0 else 'additional'
                        )
                        session.add(product_image)
                        mapped_count += 1
                        
                        print(f"  ✅ {product_name} -> {product.name} #{product.id} -> {new_image_name}")
                        
                    except Exception as e:
                        print(f"  ❌ Ошибка при копировании изображения: {e}")
                else:
                    print(f"  ⚠️  Товар не найден: '{product_name}'")
        
        session.commit()
        print(f"\n✅ Привязано изображений: {mapped_count}")

def main():
    """Основная функция"""
    print("🚀 Продвинутый парсер изображений с точным позиционированием")
    
    # Очищаем существующие данные
    clear_existing_images()
    
    # Список Excel файлов для обработки
    excel_files = [
        "storage/excel_files/original_sheet.xlsx",
        "storage/excel_files/Вторая таблица_1757933430.xlsx", 
        "storage/excel_files/Мерч для Sense_1757934153.xlsx"
    ]
    
    all_linked_images = []
    
    # Обрабатываем каждый файл отдельно
    for excel_file in excel_files:
        if os.path.exists(excel_file):
            linked_images = process_single_excel_advanced(excel_file)
            all_linked_images.extend(linked_images)
        else:
            print(f"⚠️  Файл не найден: {excel_file}")
    
    print(f"\n📊 Всего найдено изображений: {len(all_linked_images)}")
    
    # Привязываем изображения к товарам с учетом таблицы
    map_images_to_products_by_table(all_linked_images)
    
    print("\n✅ Продвинутый парсинг изображений завершен!")

if __name__ == "__main__":
    main()
