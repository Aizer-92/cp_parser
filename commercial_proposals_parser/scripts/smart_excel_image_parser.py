#!/usr/bin/env python3
"""
Умный парсер изображений из Excel файлов с поиском названий товаров в соседних ячейках
"""

import os
import sys
import zipfile
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import json

# Добавляем путь к модулям
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL

def find_product_name_near_image(ws, image_row, image_col, search_radius=5):
    """Ищет название товара в соседних ячейках от изображения"""
    
    # Список возможных названий товаров
    product_keywords = [
        'ежедневник', 'худи', 'шапка', 'бини', 'шоппер', 'брелок', 'таблетница', 
        'кардхолдер', 'зонт', 'письменный', 'набор', 'термокружка', 'сумка',
        'футболка', 'кружка', 'ручка', 'карандаш', 'футляр', 'обложка', 'косметичка',
        'чехол', 'ноутбук', 'дутый'
    ]
    
    best_match = None
    best_score = 0
    
    # Ищем в радиусе вокруг изображения
    for row_offset in range(-search_radius, search_radius + 1):
        for col_offset in range(-search_radius, search_radius + 1):
            check_row = image_row + row_offset
            check_col = image_col + col_offset
            
            if check_row < 1 or check_col < 1:
                continue
                
            try:
                cell_value = ws.cell(row=check_row, column=check_col).value
                if not cell_value:
                    continue
                    
                cell_text = str(cell_value).strip().lower()
                
                # Проверяем, содержит ли ячейка название товара
                for keyword in product_keywords:
                    if keyword in cell_text:
                        # Вычисляем "близость" к изображению
                        distance = abs(row_offset) + abs(col_offset)
                        score = len(keyword) / (distance + 1)
                        
                        if score > best_score:
                            best_score = score
                            best_match = {
                                'name': str(cell_value).strip(),
                                'row': check_row,
                                'col': check_col,
                                'distance': distance
                            }
            except:
                continue
    
    return best_match

def extract_images_from_excel_zip(excel_path, output_dir):
    """Извлекает изображения из Excel файла через zipfile"""
    print(f"📤 Извлекаем изображения из {excel_path}")
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and f.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
            
            print(f"  🖼️  Найдено изображений в архиве: {len(image_files)}")
            
            extracted_images = []
            
            for i, image_file in enumerate(image_files):
                try:
                    image_data = zip_ref.read(image_file)
                    original_name = os.path.basename(image_file)
                    new_name = f"image_{i+1}_{original_name}"
                    output_path = os.path.join(output_dir, new_name)
                    
                    with open(output_path, 'wb') as f:
                        f.write(image_data)
                    
                    extracted_images.append({
                        'original_path': image_file,
                        'saved_path': output_path,
                        'filename': new_name,
                        'index': i + 1
                    })
                    
                    print(f"    ✅ Извлечено: {original_name} -> {new_name}")
                    
                except Exception as e:
                    print(f"    ❌ Ошибка при извлечении {image_file}: {e}")
            
            return extracted_images
            
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений из {excel_path}: {e}")
        return []

def analyze_excel_with_smart_mapping(excel_path):
    """Анализирует Excel файл с умным поиском названий товаров"""
    print(f"🔍 Умный анализ структуры {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        analysis = {
            'file_path': excel_path,
            'sheets': {},
            'images_info': []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📊 Лист: {sheet_name}")
            
            sheet_info = {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'images': []
            }
            
            # Анализируем изображения на листе
            for i, image in enumerate(ws._images):
                try:
                    if hasattr(image, 'anchor') and image.anchor:
                        anchor = image.anchor
                        
                        if hasattr(anchor, '_from'):
                            from_cell = anchor._from
                            row = from_cell.row + 1
                            col = from_cell.col + 1
                        else:
                            row = getattr(anchor, 'row', 1)
                            col = getattr(anchor, 'col', 1)
                        
                        # Ищем название товара в соседних ячейках
                        product_match = find_product_name_near_image(ws, row, col)
                        
                        if product_match:
                            product_name = product_match['name']
                            print(f"    🎯 Строка {row}, колонка {col}: '{product_name}' (найдено в строке {product_match['row']})")
                        else:
                            # Пробуем получить значение самой ячейки
                            cell_value = ws.cell(row=row, column=col).value
                            product_name = str(cell_value).strip() if cell_value else f'Unknown_{i+1}'
                            print(f"    ⚠️  Строка {row}, колонка {col}: '{product_name}' (не найдено)")
                        
                        image_info = {
                            'index': i + 1,
                            'row': row,
                            'col': col,
                            'cell_value': str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value else '',
                            'product_name': product_name,
                            'width': getattr(image, 'width', 0),
                            'height': getattr(image, 'height', 0)
                        }
                        
                        sheet_info['images'].append(image_info)
                        analysis['images_info'].append({
                            'sheet': sheet_name,
                            **image_info
                        })
                        
                except Exception as e:
                    print(f"    ❌ Ошибка при анализе изображения {i+1}: {e}")
            
            analysis['sheets'][sheet_name] = sheet_info
        
        return analysis
        
    except Exception as e:
        print(f"❌ Ошибка при анализе {excel_path}: {e}")
        return None

def create_smart_image_mapping(excel_files):
    """Создает умный маппинг изображений к товарам"""
    print("🧠 Создаем умный маппинг изображений к товарам...")
    
    all_images = []
    
    for excel_file in excel_files:
        if os.path.exists(excel_file):
            print(f"\n📁 Обрабатываем: {excel_file}")
            
            # Анализируем структуру
            analysis = analyze_excel_with_smart_mapping(excel_file)
            if analysis:
                # Извлекаем изображения
                output_dir = f"storage/images/products_parsed/{os.path.basename(excel_file).replace('.xlsx', '')}"
                os.makedirs(output_dir, exist_ok=True)
                
                extracted_images = extract_images_from_excel_zip(excel_file, output_dir)
                
                # Связываем извлеченные изображения с анализом
                for i, extracted in enumerate(extracted_images):
                    if i < len(analysis['images_info']):
                        image_info = analysis['images_info'][i]
                        all_images.append({
                            'excel_file': excel_file,
                            'sheet_name': image_info['sheet'],
                            'row': image_info['row'],
                            'col': image_info['col'],
                            'product_name': image_info['product_name'],
                            'image_path': extracted['saved_path'],
                            'image_filename': extracted['filename']
                        })
    
    return all_images

def map_images_to_products_smart(image_mapping):
    """Умно привязывает изображения к товарам в базе данных"""
    print("🔗 Умно привязываем изображения к товарам...")
    
    db = CommercialProposalsDB(DATABASE_URL.replace('.db', '_v4.db'))
    
    with db.get_session() as session:
        from database.models_v4 import Product, ProductImage
        
        # Получаем все товары
        products = session.query(Product).all()
        print(f"📦 Найдено товаров в БД: {len(products)}")
        
        # Создаем словарь для поиска товаров
        products_by_name = {}
        for product in products:
            normalized_name = product.name.strip().lower()
            if normalized_name not in products_by_name:
                products_by_name[normalized_name] = []
            products_by_name[normalized_name].append(product)
        
        mapped_count = 0
        unmatched_images = []
        
        for image_info in image_mapping:
            product_name = image_info['product_name'].strip()
            if not product_name or product_name == 'None' or product_name == '':
                unmatched_images.append(image_info)
                continue
            
            # Ищем товар по названию
            normalized_name = product_name.lower()
            matching_products = products_by_name.get(normalized_name, [])
            
            if not matching_products:
                # Пробуем найти по частичному совпадению
                for norm_name, prods in products_by_name.items():
                    if normalized_name in norm_name or norm_name in normalized_name:
                        matching_products = prods
                        break
            
            if matching_products:
                product = matching_products[0]
                
                # Создаем новое имя файла
                new_image_name = f"product_{product.id}_{mapped_count + 1}.jpg"
                new_image_path = f"storage/images/products_parsed/{new_image_name}"
                
                try:
                    # Копируем изображение
                    shutil.copy2(image_info['image_path'], new_image_path)
                    
                    # Создаем запись в БД
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=new_image_path,
                        image_type='main' if mapped_count == 0 else 'additional'
                    )
                    session.add(product_image)
                    mapped_count += 1
                    
                    print(f"  ✅ {product_name} -> {product.name} #{product.id} -> {new_image_name}")
                    
                except Exception as e:
                    print(f"  ❌ Ошибка при копировании изображения: {e}")
            else:
                unmatched_images.append(image_info)
                print(f"  ⚠️  Товар не найден: '{product_name}'")
        
        # Пытаемся привязать несоответствующие изображения к товарам без изображений
        print(f"\n🔄 Пытаемся привязать {len(unmatched_images)} несоответствующих изображений...")
        
        products_without_images = session.query(Product).outerjoin(ProductImage).filter(ProductImage.id.is_(None)).all()
        
        for i, image_info in enumerate(unmatched_images):
            if i < len(products_without_images):
                product = products_without_images[i]
                
                new_image_name = f"product_{product.id}_unmatched_{i+1}.jpg"
                new_image_path = f"storage/images/products_parsed/{new_image_name}"
                
                try:
                    shutil.copy2(image_info['image_path'], new_image_path)
                    
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=new_image_path,
                        image_type='main'
                    )
                    session.add(product_image)
                    mapped_count += 1
                    
                    print(f"  🔄 {image_info['product_name']} -> {product.name} #{product.id} -> {new_image_name} (unmatched)")
                    
                except Exception as e:
                    print(f"  ❌ Ошибка при привязке несоответствующего изображения: {e}")
        
        session.commit()
        print(f"\n✅ Привязано изображений: {mapped_count}")

def main():
    """Основная функция"""
    print("🧠 Умный парсинг изображений из Excel файлов")
    
    # Список Excel файлов для обработки
    excel_files = [
        "storage/excel_files/original_sheet.xlsx",
        "storage/excel_files/Вторая таблица_1757933430.xlsx", 
        "storage/excel_files/Мерч для Sense_1757934153.xlsx"
    ]
    
    # Создаем умный маппинг изображений
    image_mapping = create_smart_image_mapping(excel_files)
    
    print(f"\n📊 Всего найдено изображений: {len(image_mapping)}")
    
    # Умно привязываем изображения к товарам
    map_images_to_products_smart(image_mapping)
    
    print("\n✅ Умный парсинг изображений завершен!")

if __name__ == "__main__":
    main()
