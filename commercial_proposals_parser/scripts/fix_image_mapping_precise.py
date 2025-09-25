#!/usr/bin/env python3
"""
Точная привязка изображений к товарам на основе порядка в Excel
"""

import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import shutil
from PIL import Image
import io

sys.path.append('.')
from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL_V4
from sqlalchemy import text

def get_excel_products_in_order(excel_file_path):
    """Получение товаров из Excel в правильном порядке"""
    
    print(f"🔍 Анализируем товары в Excel файле: {excel_file_path}")
    
    workbook = load_workbook(excel_file_path, data_only=True)
    sheet = workbook.active
    
    # Находим товары в колонке C, пропуская заголовки
    products = []
    for row_num in range(4, sheet.max_row + 1):  # Начинаем с строки 4 (первый товар)
        cell_c = sheet.cell(row=row_num, column=3)  # Колонка C - название товара
        if cell_c.value and isinstance(cell_c.value, str) and cell_c.value.strip():
            product_name = cell_c.value.strip()
            # Пропускаем пустые строки и заголовки
            if product_name and product_name != 'Наименование':
                products.append({
                    'row': row_num,
                    'name': product_name,
                    'excel_order': len(products) + 1
                })
    
    print(f"📦 Найдено товаров в Excel: {len(products)}")
    for i, product in enumerate(products, 1):
        print(f"  {i}. Строка {product['row']}: {product['name']}")
    
    return products

def get_db_products_merch_sense():
    """Получение товаров 'Мерч для Sense' из базы данных в правильном порядке"""
    
    db = CommercialProposalsDB(DATABASE_URL_V4)
    products_with_details = db.get_all_products_with_details(limit=50)
    
    # Фильтруем только товары "Мерч для Sense" (ID >= 29)
    merch_sense_products = []
    for product, price_offers, images in products_with_details:
        if product.id >= 29:
            merch_sense_products.append({
                'id': product.id,
                'name': product.name,
                'db_order': len(merch_sense_products) + 1
            })
    
    print(f"📦 Найдено товаров 'Мерч для Sense' в БД: {len(merch_sense_products)}")
    for i, product in enumerate(merch_sense_products, 1):
        print(f"  {i}. ID {product['id']}: {product['name']}")
    
    return merch_sense_products

def extract_images_from_excel(excel_file_path, output_dir):
    """Извлечение изображений из Excel файла"""
    
    print(f"🖼️  Извлекаем изображения из Excel файла...")
    
    extracted_images = []
    try:
        with zipfile.ZipFile(excel_file_path, 'r') as z:
            for name in z.namelist():
                if name.startswith('xl/media/image'):
                    image_data = z.read(name)
                    
                    # Проверяем, что это валидное изображение
                    try:
                        img = Image.open(io.BytesIO(image_data))
                        img.verify()
                        
                        # Сохраняем изображение
                        temp_name = os.path.basename(name)
                        temp_path = os.path.join(output_dir, temp_name)
                        with open(temp_path, 'wb') as f:
                            f.write(image_data)
                        
                        extracted_images.append({
                            'original_name': name,
                            'temp_path': temp_path,
                            'format': img.format
                        })
                    except Exception as e:
                        print(f"  ❌ Пропускаем поврежденное изображение: {name}")
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений: {e}")
        return []
    
    print(f"✅ Извлечено {len(extracted_images)} изображений")
    return extracted_images

def create_precise_mapping(excel_products, db_products, extracted_images, output_dir):
    """Создание точной привязки изображений к товарам"""
    
    print(f"\n🔗 Создаем точную привязку изображений...")
    
    # Создаем маппинг по порядку
    mappings = []
    
    # Привязываем изображения к товарам по порядку
    for i, excel_product in enumerate(excel_products):
        if i < len(extracted_images) and i < len(db_products):
            img_info = extracted_images[i]
            db_product = db_products[i]
            
            # Создаем новое имя файла
            new_image_name = f"product_{db_product['id']}_main.jpg"
            new_image_path = os.path.join(output_dir, new_image_name)
            
            # Копируем изображение
            shutil.copy2(img_info['temp_path'], new_image_path)
            
            mappings.append({
                'excel_product': excel_product['name'],
                'db_product_id': db_product['id'],
                'db_product_name': db_product['name'],
                'image_path': new_image_path,
                'relative_path': f"storage/images/products_parsed/{new_image_name}"
            })
            
            print(f"  ✅ {excel_product['name']} -> {db_product['name']} (ID: {db_product['id']}) -> {new_image_name}")
    
    return mappings

def fix_image_mapping_precise():
    """Основная функция для точной привязки изображений"""
    
    db = CommercialProposalsDB(DATABASE_URL_V4)
    
    # 1. Очищаем существующие изображения для товаров "Мерч для Sense"
    print("🧹 Очищаем существующие изображения для товаров 'Мерч для Sense'...")
    session = db.get_session()
    try:
        session.execute(text('DELETE FROM product_images WHERE product_id >= 29'))
        session.commit()
        print("✅ Очищены изображения для товаров 'Мерч для Sense'")
    finally:
        session.close()
    
    # 2. Очищаем папку для изображений
    images_output_dir = "storage/images/products_parsed"
    if os.path.exists(images_output_dir):
        # Удаляем только изображения товаров "Мерч для Sense"
        for file in os.listdir(images_output_dir):
            if file.startswith('product_') and any(name in file for name in ['Худи', 'ШАПКА', 'ШОППЕР', 'Брелоки', 'ДУТЫЙ', 'Зонт', 'Письменный', 'Термокружка', 'Сумка', 'Шоппер']):
                os.remove(os.path.join(images_output_dir, file))
        print(f"✅ Очищены изображения товаров 'Мерч для Sense' из папки")
    
    # 3. Получаем товары из Excel и БД
    excel_file = "storage/excel_files/Мерч для Sense_1758096973.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл не найден: {excel_file}")
        return
    
    excel_products = get_excel_products_in_order(excel_file)
    db_products = get_db_products_merch_sense()
    
    if not excel_products or not db_products:
        print("❌ Не удалось получить данные о товарах")
        return
    
    # 4. Извлекаем изображения
    extracted_images = extract_images_from_excel(excel_file, images_output_dir)
    
    if not extracted_images:
        print("❌ Не удалось извлечь изображения")
        return
    
    # 5. Создаем точную привязку
    mappings = create_precise_mapping(excel_products, db_products, extracted_images, images_output_dir)
    
    # 6. Сохраняем привязки в базу данных
    print(f"\n💾 Сохраняем привязки в базу данных...")
    
    mapped_count = 0
    for mapping in mappings:
        try:
            db.create_product_image(
                product_id=mapping['db_product_id'],
                image_path=mapping['relative_path'],
                image_type='main'
            )
            mapped_count += 1
        except Exception as e:
            print(f"  ❌ Ошибка при сохранении изображения для товара {mapping['db_product_name']}: {e}")
    
    print(f"\n✅ Точная привязка завершена! Привязано {mapped_count} изображений.")
    
    # 7. Проверяем результат
    print(f"\n📊 Проверяем результат:")
    products_with_details = db.get_all_products_with_details(limit=50)
    merch_sense_products = [(p, po, im) for p, po, im in products_with_details if p.id >= 29]
    
    for product, price_offers, images in merch_sense_products:
        print(f"  {product.name} (ID: {product.id}) - {len(images)} изображений")

if __name__ == "__main__":
    fix_image_mapping_precise()
