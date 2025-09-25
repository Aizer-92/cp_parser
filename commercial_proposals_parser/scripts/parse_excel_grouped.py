#!/usr/bin/env python3
"""
Парсинг Excel файла с правильной группировкой товаров по названию и тиражу
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl
from pathlib import Path
import re

def parse_excel_grouped(excel_path):
    """Парсинг Excel файла с группировкой товаров"""
    
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        print(f"❌ Файл не найден: {excel_path}")
        return []
    
    print(f"🔄 Парсинг Excel файла: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        print(f"📊 Размер файла: {ws.max_row} строк x {ws.max_column} колонок")
        
        # Определяем колонки динамически из заголовков
        product_name_col = 2  # B - Наименование
        characteristics_col = 3  # C - Характеристики
        custom_design_col = 4  # D - Кастом
        quantity_col = 5  # E - Тираж
        
        # Ищем варианты доставки по заголовкам (строка 2) и подзаголовкам (строка 3)
        delivery_variants = {}
        header_row = 2
        subheader_row = 3
        
        for col in range(6, ws.max_column + 1):
            # Проверяем заголовок (строка 2)
            header_cell = ws.cell(row=header_row, column=col)
            header_value = header_cell.value
            if not header_value:
                continue
                
            header_text = str(header_value).strip().lower()
            
            # Определяем тип маршрута по ключевым словам
            route_type = None
            if 'жд' in header_text or 'железнодорожн' in header_text:
                route_type = 'ЖД'
            elif 'авиа' in header_text or 'самолет' in header_text:
                route_type = 'АВИА'
            elif 'контейнер' in header_text:
                route_type = 'Контейнер'
            elif 'образец' in header_text or 'sample' in header_text:
                route_type = 'Образец'
            
            if route_type and route_type not in delivery_variants:
                delivery_variants[route_type] = {
                    'price_usd_col': None,
                    'price_rub_col': None,
                    'delivery_time_col': None,
                    'sample_price_col': None,
                    'sample_time_col': None
                }
            
            # Проверяем подзаголовки (строка 3) для определения типа колонки
            subheader_cell = ws.cell(row=subheader_row, column=col)
            subheader_value = subheader_cell.value
            if subheader_value:
                subheader_text = str(subheader_value).strip().lower()
                
                if route_type:
                    if '$' in subheader_text or 'usd' in subheader_text or 'доллар' in subheader_text:
                        delivery_variants[route_type]['price_usd_col'] = col
                    elif 'руб' in subheader_text or 'rub' in subheader_text:
                        delivery_variants[route_type]['price_rub_col'] = col
                    elif 'срок' in subheader_text and 'тиража' in subheader_text:
                        delivery_variants[route_type]['delivery_time_col'] = col
                    elif 'срок' in subheader_text and ('фото' in subheader_text or 'видео' in subheader_text):
                        delivery_variants[route_type]['sample_time_col'] = col
                    elif 'срок' in subheader_text and 'доставкой' in subheader_text:
                        delivery_variants[route_type]['sample_time_col'] = col
        
        # Дополнительный поиск по соседним колонкам для рублевых цен и сроков
        for variant, cols in delivery_variants.items():
            if cols['price_usd_col'] and not cols['price_rub_col']:
                next_col = cols['price_usd_col'] + 1
                if next_col <= ws.max_column:
                    next_cell = ws.cell(row=subheader_row, column=next_col)
                    if next_cell.value and 'руб' in str(next_cell.value).lower():
                        delivery_variants[variant]['price_rub_col'] = next_col
            
            if cols['price_rub_col'] and not cols['delivery_time_col']:
                next_col = cols['price_rub_col'] + 1
                if next_col <= ws.max_column:
                    next_cell = ws.cell(row=subheader_row, column=next_col)
                    if next_cell.value and 'срок' in str(next_cell.value).lower():
                        delivery_variants[variant]['delivery_time_col'] = next_col
        
        print(f"🔍 Структура колонок:")
        print(f"  - Название товара: {product_name_col} (B)")
        print(f"  - Характеристики: {characteristics_col} (C)")
        print(f"  - Кастом: {custom_design_col} (D)")
        print(f"  - Тираж: {quantity_col} (E)")
        print(f"🚚 Найденные варианты доставки:")
        for variant, cols in delivery_variants.items():
            print(f"  - {variant}:")
            print(f"    USD: {cols['price_usd_col']}, RUB: {cols['price_rub_col']}")
            print(f"    Срок: {cols['delivery_time_col']}, Образец: {cols['sample_price_col']}")
            print(f"    Срок образца: {cols['sample_time_col']}")
        
        # Собираем все строки с товарами
        product_rows = []
        current_product_name = None
        
        for row_num in range(4, ws.max_row + 1):
            # Получаем название товара
            product_name_cell = ws.cell(row=row_num, column=product_name_col)
            product_name = product_name_cell.value
            
            # Если название есть, обновляем текущее название
            if product_name and str(product_name).strip() != '' and str(product_name).strip() != ' ':
                current_product_name = str(product_name).strip()
            
            # Если нет текущего названия, пропускаем строку
            if not current_product_name:
                continue
            
            # Получаем характеристики
            characteristics_cell = ws.cell(row=row_num, column=characteristics_col)
            characteristics = str(characteristics_cell.value).strip() if characteristics_cell.value else None
            
            # Получаем кастом
            custom_design_cell = ws.cell(row=row_num, column=custom_design_col)
            custom_design = str(custom_design_cell.value).strip() if custom_design_cell.value else None
            
            # Получаем тираж
            quantity_cell = ws.cell(row=row_num, column=quantity_col)
            quantity = None
            if quantity_cell.value:
                try:
                    quantity_str = str(quantity_cell.value).replace(' ', '').replace(',', '')
                    quantity = int(float(quantity_str))
                except:
                    pass
            
            # Собираем ценовые предложения для этой строки
            price_offers = []
            for variant, cols in delivery_variants.items():
                # Получаем цену в USD
                price_usd = None
                if cols['price_usd_col']:
                    price_usd_cell = ws.cell(row=row_num, column=cols['price_usd_col'])
                    if price_usd_cell.value:
                        try:
                            price_usd = float(price_usd_cell.value)
                        except:
                            pass
                
                # Получаем цену в RUB
                price_rub = None
                if cols['price_rub_col']:
                    price_rub_cell = ws.cell(row=row_num, column=cols['price_rub_col'])
                    if price_rub_cell.value:
                        try:
                            price_rub = float(price_rub_cell.value)
                        except:
                            pass
                
                # Получаем срок доставки
                delivery_time = None
                if cols['delivery_time_col']:
                    delivery_time_cell = ws.cell(row=row_num, column=cols['delivery_time_col'])
                    if delivery_time_cell.value:
                        delivery_time = str(delivery_time_cell.value).strip()
                
                # Получаем цену образца
                sample_price = None
                if cols['sample_price_col']:
                    sample_price_cell = ws.cell(row=row_num, column=cols['sample_price_col'])
                    if sample_price_cell.value:
                        try:
                            sample_price = float(sample_price_cell.value)
                        except:
                            pass
                
                # Получаем срок образца
                sample_time = None
                if cols['sample_time_col']:
                    sample_time_cell = ws.cell(row=row_num, column=cols['sample_time_col'])
                    if sample_time_cell.value:
                        sample_time = str(sample_time_cell.value).strip()
                
                # Создаем ценовое предложение только если есть хотя бы одна цена
                if price_usd is not None or price_rub is not None or sample_price is not None:
                    is_sample = variant == 'Образец'
                    
                    # Для образцов используем рублевую цену как основную
                    if is_sample and price_rub is not None:
                        sample_price = price_rub
                        price_rub = None
                    
                    offer = {
                        'route_name': variant,
                        'quantity': quantity,
                        'price_usd': price_usd,
                        'price_rub': price_rub,
                        'delivery_time': delivery_time,
                        'sample_price': sample_price,
                        'sample_time': sample_time,
                        'is_sample': is_sample
                    }
                    price_offers.append(offer)
            
            # Добавляем строку товара
            product_rows.append({
                'name': current_product_name,
                'description': characteristics,
                'characteristics': characteristics,
                'custom_design': custom_design,
                'quantity': quantity,
                'price_offers': price_offers,
                'images': []
            })
        
        # Группируем товары по названию и тиражу
        grouped_products = {}
        
        for product_row in product_rows:
            # Создаем ключ для группировки
            key = f"{product_row['name']}_{product_row['quantity']}"
            
            if key not in grouped_products:
                # Создаем новый товар
                grouped_products[key] = {
                    'name': product_row['name'],
                    'description': product_row['description'],
                    'characteristics': product_row['characteristics'],
                    'custom_design': product_row['custom_design'],
                    'quantity': product_row['quantity'],
                    'price_offers': [],
                    'images': []
                }
            
            # Добавляем ценовые предложения
            grouped_products[key]['price_offers'].extend(product_row['price_offers'])
        
        # Преобразуем в список
        products = list(grouped_products.values())
        
        print(f"✅ Парсинг завершен!")
        print(f"📊 Результаты:")
        print(f"  - Найдено товаров: {len(products)}")
        
        total_offers = sum(len(p['price_offers']) for p in products)
        print(f"  - Найдено ценовых предложений: {total_offers}")
        
        # Показываем детали по каждому товару
        for product in products:
            print(f"\n📦 {product['name']} (тираж: {product['quantity']})")
            print(f"  Ценовых предложений: {len(product['price_offers'])}")
            for offer in product['price_offers']:
                if offer['is_sample']:
                    print(f"    🎯 {offer['route_name']}: {offer['sample_price']} руб ({offer['sample_time']} дней)")
                else:
                    price_usd = f"${offer['price_usd']}" if offer['price_usd'] else "N/A"
                    price_rub = f"{offer['price_rub']} руб" if offer['price_rub'] else "N/A"
                    delivery = f"({offer['delivery_time']} дней)" if offer['delivery_time'] else ""
                    print(f"    📦 {offer['route_name']}: {price_usd} / {price_rub} {delivery}")
        
        return products
        
    except Exception as e:
        print(f"❌ Ошибка при парсинге: {e}")
        import traceback
        traceback.print_exc()
        return []

if __name__ == "__main__":
    # Тестирование
    test_file = "storage/excel_files/Вторая таблица_1757933504.xlsx"
    products = parse_excel_grouped(test_file)
    print(f"\n📊 Итого найдено товаров: {len(products)}")
