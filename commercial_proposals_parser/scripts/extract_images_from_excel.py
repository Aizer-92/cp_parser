#!/usr/bin/env python3
"""
Извлечение изображений из Excel файла
"""

import openpyxl
from openpyxl.drawing.image import Image
import os
from pathlib import Path
import sys

def extract_images_from_excel(excel_path, output_dir):
    """Извлекает все изображения из Excel файла"""
    
    # Создаем папку для изображений
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    # Открываем Excel файл
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"📊 Извлечение изображений из: {excel_path}")
    print(f"📁 Папка для сохранения: {output_dir}")
    
    # Счетчик изображений
    image_count = 0
    
    # Проходим по всем строкам и колонкам
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            # Проверяем, есть ли изображение в ячейке
            if hasattr(cell, 'value') and cell.value is not None:
                # Ищем изображения в ячейке
                for image in ws._images:
                    if (image.anchor._from.row == row - 1 and 
                        image.anchor._from.col == col - 1):
                        
                        # Создаем имя файла
                        filename = f"product_row_{row}_col_{col}_{image_count}.jpg"
                        filepath = output_path / filename
                        
                        try:
                            # Сохраняем изображение
                            with open(filepath, 'wb') as f:
                                f.write(image.ref.read())
                            
                            print(f"✅ Сохранено: {filename} (строка {row}, колонка {col})")
                            image_count += 1
                            
                        except Exception as e:
                            print(f"❌ Ошибка сохранения {filename}: {e}")
    
    print(f"📸 Всего извлечено изображений: {image_count}")
    return image_count

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Использование: python extract_images_from_excel.py <excel_file> <output_dir>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_dir = sys.argv[2]
    
    if not os.path.exists(excel_path):
        print(f"❌ Файл не найден: {excel_path}")
        sys.exit(1)
    
    extract_images_from_excel(excel_path, output_dir)
