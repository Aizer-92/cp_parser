#!/usr/bin/env python3
"""
Анализ структуры Excel файлов для создания универсального парсера
"""

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl import load_workbook
import json

def analyze_excel_file(file_path):
    """Анализирует структуру Excel файла"""
    print(f"\n📊 Анализ файла: {os.path.basename(file_path)}")
    print("=" * 60)
    
    try:
        # Загружаем Excel файл
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"📋 Лист: {ws.title}")
        print(f"📏 Размер: {ws.max_row} строк x {ws.max_column} колонок")
        
        # Анализируем первые 10 строк
        print("\n🔍 Первые 10 строк:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(21, ws.max_column + 1)):  # Первые 20 колонок
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    row_data.append(f"{col}: {str(value)[:30]}")
            if row_data:
                print(f"  Строка {row}: {' | '.join(row_data)}")
        
        # Ищем заголовки и подзаголовки
        print("\n📝 Анализ заголовков:")
        
        # Строка 1 - основные заголовки
        headers_row1 = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                headers_row1.append((col, str(cell.value)))
        
        if headers_row1:
            print(f"  Строка 1 (заголовки): {headers_row1}")
        
        # Строка 2 - подзаголовки
        headers_row2 = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value:
                headers_row2.append((col, str(cell.value)))
        
        if headers_row2:
            print(f"  Строка 2 (подзаголовки): {headers_row2}")
        
        # Строка 3 - дополнительные подзаголовки
        headers_row3 = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=3, column=col)
            if cell.value:
                headers_row3.append((col, str(cell.value)))
        
        if headers_row3:
            print(f"  Строка 3 (доп. подзаголовки): {headers_row3}")
        
        # Ищем колонки с изображениями
        print("\n🖼️  Анализ изображений:")
        image_columns = []
        for col in range(1, ws.max_column + 1):
            has_images = False
            for row in range(1, min(20, ws.max_row + 1)):
                if ws.cell(row=row, column=col).value is None:
                    # Проверяем, есть ли изображение в этой ячейке
                    for image in ws._images:
                        if (hasattr(image, 'anchor') and 
                            hasattr(image.anchor, '_from') and
                            image.anchor._from.row == row - 1 and 
                            image.anchor._from.col == col - 1):
                            has_images = True
                            break
            if has_images:
                image_columns.append(col)
        
        if image_columns:
            print(f"  Колонки с изображениями: {image_columns}")
        else:
            print("  Изображения не найдены в первых 20 строках")
        
        # Ищем колонки с ценами
        print("\n💰 Анализ ценовых колонок:")
        price_columns = []
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['цена', 'price', 'usd', 'rub', '₽', '$']):
                    price_columns.append((col, str(cell.value)))
                    break
        
        if price_columns:
            print(f"  Колонки с ценами: {price_columns}")
        
        # Ищем колонки с тиражами
        print("\n📦 Анализ колонок с тиражами:")
        quantity_columns = []
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['тираж', 'количество', 'quantity', 'шт', 'штук']):
                    quantity_columns.append((col, str(cell.value)))
                    break
        
        if quantity_columns:
            print(f"  Колонки с тиражами: {quantity_columns}")
        
        # Ищем колонки с маршрутами доставки
        print("\n🚚 Анализ колонок с маршрутами:")
        route_columns = []
        for col in range(1, ws.max_column + 1):
            for row in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and any(keyword in str(cell.value).lower() for keyword in ['жд', 'авиа', 'контейнер', 'образец', 'доставка', 'маршрут']):
                    route_columns.append((col, str(cell.value)))
                    break
        
        if route_columns:
            print(f"  Колонки с маршрутами: {route_columns}")
        
        return {
            'file_name': os.path.basename(file_path),
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'headers_row1': headers_row1,
            'headers_row2': headers_row2,
            'headers_row3': headers_row3,
            'image_columns': image_columns,
            'price_columns': price_columns,
            'quantity_columns': quantity_columns,
            'route_columns': route_columns
        }
        
    except Exception as e:
        print(f"❌ Ошибка при анализе файла: {e}")
        return None

def main():
    """Основная функция анализа"""
    excel_dir = "storage/excel_files"
    
    if not os.path.exists(excel_dir):
        print(f"❌ Директория {excel_dir} не найдена")
        return
    
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not excel_files:
        print(f"❌ Excel файлы не найдены в {excel_dir}")
        return
    
    print(f"📁 Найдено Excel файлов: {len(excel_files)}")
    
    analysis_results = []
    
    for excel_file in excel_files:
        file_path = os.path.join(excel_dir, excel_file)
        result = analyze_excel_file(file_path)
        if result:
            analysis_results.append(result)
    
    # Сохраняем результаты анализа
    with open('excel_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(analysis_results, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Анализ завершен. Результаты сохранены в excel_analysis.json")
    
    # Сводка по всем файлам
    print("\n📊 СВОДКА ПО ВСЕМ ФАЙЛАМ:")
    print("=" * 60)
    
    for result in analysis_results:
        print(f"\n📄 {result['file_name']}")
        print(f"  Размер: {result['max_row']} x {result['max_column']}")
        print(f"  Изображения: {result['image_columns']}")
        print(f"  Цены: {len(result['price_columns'])} колонок")
        print(f"  Тиражи: {len(result['quantity_columns'])} колонок")
        print(f"  Маршруты: {len(result['route_columns'])} колонок")

if __name__ == "__main__":
    main()
