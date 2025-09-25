#!/usr/bin/env python3
"""
Финальный правильный парсер Excel файлов с точным извлечением данных
"""

import os
import sys
import json
import pandas as pd
import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional

# Добавляем корневую директорию проекта в путь
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

from database.manager import CommercialProposalsDB
import config

def parse_excel_final(excel_path: str) -> List[Dict[str, Any]]:
    """
    Финальный правильный парсинг Excel файла с точным извлечением данных
    """
    print(f"🔄 Финальный парсинг Excel файла: {excel_path}")
    
    # Читаем Excel файл с openpyxl для точного извлечения данных
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"📊 Размер файла: {ws.max_row} строк x {ws.max_column} колонок")
    
    products = []
    
    # Проходим по всем строкам, начиная с 4-й (где начинаются данные товаров)
    for row in range(4, ws.max_row + 1):
        # Получаем название товара (колонка B)
        product_name = ws.cell(row=row, column=2).value
        if not product_name or not str(product_name).strip():
            continue
        
        product_name = str(product_name).strip()
        
        # Пропускаем служебные строки
        if product_name in ['Наименование', 'Характеристики', 'Кастом']:
            continue
        
        print(f"📦 Обрабатываем товар: {product_name}")
        
        # Извлекаем основные данные
        characteristics = ws.cell(row=row, column=3).value or ""
        custom_design = ws.cell(row=row, column=4).value or ""
        quantity = ws.cell(row=row, column=5).value
        
        # Обрабатываем тираж
        if quantity:
            try:
                if isinstance(quantity, str):
                    quantity = int(quantity.replace(' ', '').replace(',', ''))
                else:
                    quantity = int(quantity)
            except:
                quantity = None
        
        # Извлекаем данные по видам доставки
        
        # ЖД доставка (колонки 6-8)
        jd_price_usd = ws.cell(row=row, column=6).value
        jd_price_rub = ws.cell(row=row, column=7).value
        jd_delivery_time = ws.cell(row=row, column=8).value
        
        # АВИА доставка (колонки 9-11)
        avia_price_usd = ws.cell(row=row, column=9).value
        avia_price_rub = ws.cell(row=row, column=10).value
        avia_delivery_time = ws.cell(row=row, column=11).value
        
        # Контейнер доставка (колонки 12-14)
        container_price_usd = ws.cell(row=row, column=12).value
        container_price_rub = ws.cell(row=row, column=13).value
        container_delivery_time = ws.cell(row=row, column=14).value
        
        # Образец (колонки 15-16)
        sample_price = ws.cell(row=row, column=15).value
        sample_time = ws.cell(row=row, column=16).value
        
        # Обрабатываем числовые значения
        def safe_float(value):
            if value is None:
                return None
            try:
                return float(value)
            except:
                return None
        
        def safe_str(value):
            if value is None:
                return None
            return str(value).strip()
        
        # Создаем предложения для каждого вида доставки, если есть цены
        
        # ЖД доставка
        if jd_price_usd is not None:
            product_data = {
                'name': product_name,
                'description': str(characteristics).strip(),
                'custom_design': str(custom_design).strip(),
                'quantity': quantity,
                'delivery_type': 'ЖД',
                'price_usd': safe_float(jd_price_usd),
                'price_rub': safe_float(jd_price_rub),
                'delivery_time': safe_str(jd_delivery_time),
                'sample_price': safe_float(sample_price),
                'sample_time': safe_str(sample_time),
                'row_number': row,
                'variant': 'ЖД'
            }
            products.append(product_data)
            print(f"  ✅ ЖД: ${product_data['price_usd']} / {product_data['price_rub']} руб / {product_data['delivery_time']} дней")
        
        # АВИА доставка
        if avia_price_usd is not None:
            product_data = {
                'name': product_name,
                'description': str(characteristics).strip(),
                'custom_design': str(custom_design).strip(),
                'quantity': quantity,
                'delivery_type': 'АВИА',
                'price_usd': safe_float(avia_price_usd),
                'price_rub': safe_float(avia_price_rub),
                'delivery_time': safe_str(avia_delivery_time),
                'sample_price': safe_float(sample_price),
                'sample_time': safe_str(sample_time),
                'row_number': row,
                'variant': 'АВИА'
            }
            products.append(product_data)
            print(f"  ✅ АВИА: ${product_data['price_usd']} / {product_data['price_rub']} руб / {product_data['delivery_time']} дней")
        
        # Контейнер доставка
        if container_price_usd is not None:
            product_data = {
                'name': product_name,
                'description': str(characteristics).strip(),
                'custom_design': str(custom_design).strip(),
                'quantity': quantity,
                'delivery_type': 'Контейнер',
                'price_usd': safe_float(container_price_usd),
                'price_rub': safe_float(container_price_rub),
                'delivery_time': safe_str(container_delivery_time),
                'sample_price': safe_float(sample_price),
                'sample_time': safe_str(sample_time),
                'row_number': row,
                'variant': 'Контейнер'
            }
            products.append(product_data)
            print(f"  ✅ Контейнер: ${product_data['price_usd']} / {product_data['price_rub']} руб / {product_data['delivery_time']} дней")
        
        # Если нет ни одного варианта доставки, создаем базовое предложение
        if not any([jd_price_usd, avia_price_usd, container_price_usd]):
            product_data = {
                'name': product_name,
                'description': str(characteristics).strip(),
                'custom_design': str(custom_design).strip(),
                'quantity': quantity,
                'delivery_type': 'Стандартная',
                'price_usd': None,
                'price_rub': None,
                'delivery_time': None,
                'sample_price': safe_float(sample_price),
                'sample_time': safe_str(sample_time),
                'row_number': row,
                'variant': 'Стандартная'
            }
            products.append(product_data)
            print(f"  ✅ Стандартная: без цен")
    
    print(f"📊 Найдено предложений: {len(products)}")
    return products

def save_final_products_to_db(products: List[Dict[str, Any]], sheet_url: str, sheet_title: str):
    """
    Сохраняет финальные данные в базу данных
    """
    print(f"💾 Сохранение {len(products)} предложений в базу данных...")
    
    db = CommercialProposalsDB(config.DATABASE_URL)
    
    try:
        # Создаем запись о таблице
        sheet_id = 'final_sheet'
        sheet = db.add_sheet_metadata(sheet_url, sheet_title, sheet_id)
        print(f"  ✅ Создана запись таблицы: {sheet.id}")
        
        # Сохраняем товары
        for product_data in products:
            db_product_data = {
                'name': f"{product_data['name']} ({product_data['delivery_type']})",
                'description': product_data['description'],
                'characteristics': {
                    'custom_design': product_data['custom_design'],
                    'delivery_type': product_data['delivery_type'],
                    'delivery_time': product_data['delivery_time'],
                    'sample_price': product_data['sample_price'],
                    'sample_time': product_data['sample_time'],
                    'row_number': product_data['row_number'],
                    'variant': product_data['variant']
                },
                'price_usd': product_data['price_usd'],
                'price_rub': product_data['price_rub'],
                'quantity': product_data['quantity'],
                'delivery_method': product_data['delivery_type'],
                'delivery_time': product_data['delivery_time'],
                'custom_design': product_data['custom_design'],
                'sample_price': product_data['sample_price'],
                'sample_time': product_data['sample_time'],
                'row_number': product_data['row_number']
            }
            
            product = db.add_product(sheet.id, db_product_data)
            print(f"  ✅ Создано предложение: {product.name}")
        
        print(f"✅ Все предложения сохранены в базу данных")
        
    except Exception as e:
        print(f"❌ Ошибка сохранения в базу данных: {e}")
        raise

def main():
    """Основная функция"""
    
    # Путь к Excel файлу
    excel_path = project_root / 'storage' / 'excel_files' / 'downloaded_sheet_1757491580.xlsx'
    
    if not excel_path.exists():
        print(f"❌ Excel файл не найден: {excel_path}")
        return 1
    
    try:
        print(f"🔄 Финальный парсинг Excel файла: {excel_path}")
        
        # Парсим товары
        products = parse_excel_final(str(excel_path))
        
        # Сохраняем в базу данных
        sheet_url = 'https://docs.google.com/spreadsheets/d/1mxn4kARVecduFhscX1UMOxoG-d0tlfP676F5jhw_Qw/edit'
        sheet_title = 'VTB Test Sheet - Final'
        save_final_products_to_db(products, sheet_url, sheet_title)
        
        print(f"✅ Финальный парсинг завершен успешно!")
        print(f"📊 Результаты:")
        print(f"  - Предложений: {len(products)}")
        
        # Статистика по типам доставки
        delivery_stats = {}
        for product in products:
            delivery_type = product['delivery_type']
            delivery_stats[delivery_type] = delivery_stats.get(delivery_type, 0) + 1
        
        print(f"  - По типам доставки:")
        for delivery_type, count in delivery_stats.items():
            print(f"    • {delivery_type}: {count}")
        
        # Показываем примеры данных
        print(f"\n📋 Примеры данных:")
        for i, product in enumerate(products[:5]):
            print(f"  {i+1}. {product['name']} ({product['delivery_type']})")
            print(f"     Цена USD: {product['price_usd']}")
            print(f"     Цена RUB: {product['price_rub']}")
            print(f"     Срок: {product['delivery_time']}")
            print(f"     Образец: {product['sample_price']} руб, {product['sample_time']} дней")
            print(f"     Тираж: {product['quantity']}")
            print()
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())



