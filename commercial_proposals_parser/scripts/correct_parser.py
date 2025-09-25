#!/usr/bin/env python3
"""
Правильный парсер для исправления проблем с тиражом
"""

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
import json

def parse_original_sheet():
    """Парсит оригинальную таблицу с правильным тиражом"""
    file_path = "storage/excel_files/original_sheet.xlsx"
    
    print(f"🔍 Парсинг оригинальной таблицы: {os.path.basename(file_path)}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    products = []
    
    # Начинаем с 4-й строки (после заголовков)
    for row in range(4, ws.max_row + 1):
        # Получаем название товара (колонка 2)
        name_cell = ws.cell(row=row, column=2)
        product_name = name_cell.value
        
        if not product_name or str(product_name).strip() == '':
            continue
        
        product = {
            'name': str(product_name).strip(),
            'characteristics': '',
            'custom_design': '',
            'price_offers': [],
            'images': []
        }
        
        # Получаем характеристики (колонка 3)
        char_cell = ws.cell(row=row, column=3)
        if char_cell.value:
            product['characteristics'] = str(char_cell.value).strip()
        
        # Получаем кастом (колонка 4)
        custom_cell = ws.cell(row=row, column=4)
        if custom_cell.value:
            product['custom_design'] = str(custom_cell.value).strip()
        
        # Получаем ПРАВИЛЬНЫЙ тираж (колонка 5)
        qty_cell = ws.cell(row=row, column=5)
        quantity = None
        if qty_cell.value:
            try:
                quantity_str = str(qty_cell.value).replace(' ', '').replace(',', '')
                quantity = int(float(quantity_str))
            except (ValueError, TypeError):
                quantity = None
        
        print(f"  📦 {product['name']}: тираж={quantity}")
        
        # Создаем ценовые предложения
        # ЖД маршрут (колонки 6, 7, 8)
        if ws.cell(row=row, column=6).value is not None:
            price_offer = {
                'route_name': 'ЖД',
                'quantity': quantity,
                'price_usd': None,
                'price_rub': None,
                'delivery_time': None,
                'is_sample': False,
                'sample_price': None
            }
            
            # USD цена (колонка 6)
            usd_cell = ws.cell(row=row, column=6)
            if usd_cell.value:
                try:
                    price_offer['price_usd'] = float(usd_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # RUB цена (колонка 7)
            rub_cell = ws.cell(row=row, column=7)
            if rub_cell.value:
                try:
                    price_offer['price_rub'] = float(rub_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # Срок доставки (колонка 8)
            time_cell = ws.cell(row=row, column=8)
            if time_cell.value:
                price_offer['delivery_time'] = str(time_cell.value).strip()
            
            if price_offer['price_usd'] or price_offer['price_rub']:
                product['price_offers'].append(price_offer)
        
        # АВИА маршрут (колонки 9, 10, 11)
        if ws.cell(row=row, column=9).value is not None:
            price_offer = {
                'route_name': 'АВИА',
                'quantity': quantity,
                'price_usd': None,
                'price_rub': None,
                'delivery_time': None,
                'is_sample': False,
                'sample_price': None
            }
            
            # USD цена (колонка 9)
            usd_cell = ws.cell(row=row, column=9)
            if usd_cell.value:
                try:
                    price_offer['price_usd'] = float(usd_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # RUB цена (колонка 10)
            rub_cell = ws.cell(row=row, column=10)
            if rub_cell.value:
                try:
                    price_offer['price_rub'] = float(rub_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # Срок доставки (колонка 11)
            time_cell = ws.cell(row=row, column=11)
            if time_cell.value:
                price_offer['delivery_time'] = str(time_cell.value).strip()
            
            if price_offer['price_usd'] or price_offer['price_rub']:
                product['price_offers'].append(price_offer)
        
        # Контейнер маршрут (колонки 12, 13, 14)
        if ws.cell(row=row, column=12).value is not None:
            price_offer = {
                'route_name': 'Контейнер',
                'quantity': quantity,
                'price_usd': None,
                'price_rub': None,
                'delivery_time': None,
                'is_sample': False,
                'sample_price': None
            }
            
            # USD цена (колонка 12)
            usd_cell = ws.cell(row=row, column=12)
            if usd_cell.value:
                try:
                    price_offer['price_usd'] = float(usd_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # RUB цена (колонка 13)
            rub_cell = ws.cell(row=row, column=13)
            if rub_cell.value:
                try:
                    price_offer['price_rub'] = float(rub_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # Срок доставки (колонка 14)
            time_cell = ws.cell(row=row, column=14)
            if time_cell.value:
                price_offer['delivery_time'] = str(time_cell.value).strip()
            
            if price_offer['price_usd'] or price_offer['price_rub']:
                product['price_offers'].append(price_offer)
        
        # Образец (колонка 15) - НЕ МЕНЯЕМ ТИРАЖ!
        if ws.cell(row=row, column=15).value is not None:
            price_offer = {
                'route_name': 'Образец',
                'quantity': None,  # У образца нет тиража
                'price_usd': None,
                'price_rub': None,
                'delivery_time': None,
                'is_sample': True,
                'sample_price': None
            }
            
            # RUB цена (колонка 15)
            rub_cell = ws.cell(row=row, column=15)
            if rub_cell.value:
                try:
                    price_offer['sample_price'] = float(rub_cell.value)
                except (ValueError, TypeError):
                    pass
            
            # Срок доставки (колонка 16)
            time_cell = ws.cell(row=row, column=16)
            if time_cell.value:
                price_offer['delivery_time'] = str(time_cell.value).strip()
            
            if price_offer['sample_price']:
                product['price_offers'].append(price_offer)
        
        products.append(product)
    
    print(f"✅ Найдено товаров: {len(products)}")
    return products

def main():
    """Основная функция"""
    print("🚀 Правильный парсинг оригинальной таблицы")
    print("=" * 50)
    
    # Парсим оригинальную таблицу
    products = parse_original_sheet()
    
    # Добавляем source_file
    for product in products:
        product['source_file'] = 'original_sheet.xlsx'
    
    # Сохраняем результаты
    with open('correct_parsed_products.json', 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Парсинг завершен. Товаров: {len(products)}")
    print("📄 Результаты сохранены в correct_parsed_products.json")
    
    # Показываем примеры
    print("\n🎯 Примеры товаров:")
    for product in products[:3]:
        print(f"  - {product['name']}: {len(product['price_offers'])} ценовых предложений")
        for offer in product['price_offers'][:2]:
            print(f"    * {offer['route_name']}: тираж={offer['quantity']}, USD={offer['price_usd']}, RUB={offer['price_rub']}")

if __name__ == "__main__":
    main()
