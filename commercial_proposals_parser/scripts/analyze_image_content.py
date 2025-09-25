#!/usr/bin/env python3
"""
Анализ содержимого изображений для правильной привязки к товарам
"""

import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import shutil
from PIL import Image
import io

sys.path.append('.')
from database.manager_v4 import CommercialProposalsDB
from config import DATABASE_URL_V4
from sqlalchemy import text

def analyze_image_content():
    """Анализ содержимого изображений"""
    
    print("🔍 Анализируем содержимое изображений...")
    
    # 1. Получаем товары из БД
    db = CommercialProposalsDB(DATABASE_URL_V4)
    products_with_details = db.get_all_products_with_details(limit=50)
    merch_sense_products = [(p, po, im) for p, po, im in products_with_details if p.id >= 29]
    
    print(f"📦 Товары 'Мерч для Sense' в БД:")
    for product, price_offers, images in merch_sense_products:
        print(f"  {product.id}. {product.name}")
    
    # 2. Анализируем Excel файл
    excel_file = "storage/excel_files/Мерч для Sense_1758096973.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл не найден: {excel_file}")
        return
    
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active
    
    print(f"\n📊 Товары в Excel файле:")
    products_by_row = {}
    for row_num in range(4, sheet.max_row + 1):
        cell_c = sheet.cell(row=row_num, column=3)  # Колонка C - название товара
        if cell_c.value and isinstance(cell_c.value, str) and cell_c.value.strip():
            product_name = cell_c.value.strip()
            if product_name != 'Наименование':
                products_by_row[row_num] = product_name
                print(f"  Строка {row_num}: {product_name}")
    
    # 3. Анализируем позиции изображений
    print(f"\n🖼️  Позиции изображений в Excel:")
    image_positions = []
    
    try:
        with zipfile.ZipFile(excel_file, 'r') as z:
            drawing_files = [name for name in z.namelist() if 'xl/drawings/drawing' in name]
            
            for drawing_file in drawing_files:
                try:
                    drawing_data = z.read(drawing_file)
                    root = ET.fromstring(drawing_data)
                    
                    for anchor in root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}oneCellAnchor'):
                        from_elem = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
                        if from_elem is not None:
                            row_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
                            col_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col')
                            if row_elem is not None and col_elem is not None:
                                row = int(row_elem.text) + 1
                                col = int(col_elem.text) + 1
                                
                                # Находим товар в этой строке
                                product_name = products_by_row.get(row, f"Unknown_row_{row}")
                                
                                # Определяем тип изображения
                                if col == 1:  # Колонка A - основное изображение
                                    image_type = 'main'
                                elif col == 16:  # Колонка P - дополнительное изображение
                                    image_type = 'additional'
                                else:
                                    image_type = 'other'
                                
                                image_positions.append({
                                    'row': row,
                                    'col': col,
                                    'product_name': product_name,
                                    'image_type': image_type,
                                    'drawing_file': drawing_file
                                })
                                
                                print(f"    Строка {row}, колонка {col} ({image_type}) -> {product_name}")
                except Exception as e:
                    print(f"    Ошибка при анализе {drawing_file}: {e}")
                    
    except Exception as e:
        print(f"❌ Ошибка при работе с Excel: {e}")
        return
    
    # 4. Извлекаем изображения с правильной привязкой
    print(f"\n📷 Извлекаем изображения с правильной привязкой...")
    
    try:
        with zipfile.ZipFile(excel_file, 'r') as z:
            image_files = [name for name in z.namelist() if name.startswith('xl/media/image')]
            print(f"Найдено изображений в Excel: {len(image_files)}")
            
            # Создаем маппинг изображений по позициям
            image_mapping = {}
            
            for i, name in enumerate(image_files):
                image_data = z.read(name)
                try:
                    img = Image.open(io.BytesIO(image_data))
                    img.verify()
                    
                    # Сохраняем изображение
                    temp_name = f"image_{i+1:02d}_{os.path.basename(name)}"
                    temp_path = os.path.join("storage/images/products_parsed", temp_name)
                    with open(temp_path, 'wb') as f:
                        f.write(image_data)
                    
                    # Анализируем размер изображения
                    img_size = img.size
                    file_size = len(image_data)
                    
                    print(f"  {temp_name}: {img_size[0]}x{img_size[1]}px, {file_size} байт")
                    
                    image_mapping[i] = {
                        'original_name': name,
                        'temp_path': temp_path,
                        'size': img_size,
                        'file_size': file_size,
                        'index': i
                    }
                except Exception as e:
                    print(f"  ❌ Пропускаем поврежденное изображение: {name}")
            
            # 5. Создаем правильную привязку
            print(f"\n🔗 Создаем правильную привязку...")
            
            # Группируем изображения по товарам
            products_images = {}
            for pos in image_positions:
                product_name = pos['product_name']
                if product_name not in products_images:
                    products_images[product_name] = {'main': [], 'additional': []}
                products_images[product_name][pos['image_type']].append(pos)
            
            # Привязываем изображения к товарам
            image_index = 0
            for product_name, images in products_images.items():
                print(f"\n📦 {product_name}:")
                
                # Основные изображения
                for i, pos in enumerate(images['main']):
                    if image_index < len(image_mapping):
                        img_info = image_mapping[image_index]
                        print(f"  Основное {i+1}: {img_info['temp_path']} ({img_info['size'][0]}x{img_info['size'][1]}px, {img_info['file_size']} байт)")
                        image_index += 1
                
                # Дополнительные изображения
                for i, pos in enumerate(images['additional']):
                    if image_index < len(image_mapping):
                        img_info = image_mapping[image_index]
                        print(f"  Дополнительное {i+1}: {img_info['temp_path']} ({img_info['size'][0]}x{img_info['size'][1]}px, {img_info['file_size']} байт)")
                        image_index += 1
                        
    except Exception as e:
        print(f"❌ Ошибка при извлечении изображений: {e}")

if __name__ == "__main__":
    analyze_image_content()
