#!/usr/bin/env python3
"""
Умный адаптивный парсер - парсит таблицы согласно их структуре
"""

from adaptive_table_analyzer import TableStructureAnalyzer, TableStructure
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, PriceOffer, ProductImage
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import logging
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class SmartAdaptiveParser:
    """Умный адаптивный парсер таблиц"""
    
    def __init__(self):
        self.analyzer = TableStructureAnalyzer()
        self.session = DatabaseManager.get_session()
        
        # Стек неопознанных таблиц
        self.unknown_tables = []
        
        # Минимальные требования для парсинга
        self.min_confidence = 0.6
        self.required_columns = ['name']  # Минимум название товара
    
    def parse_excel_smart(self, sheet_id: int) -> bool:
        """Умный парсинг Excel файла с анализом структуры"""
        
        # Получаем информацию о таблице
        sheet_meta = self.session.query(SheetMetadata).get(sheet_id)
        if not sheet_meta or not sheet_meta.local_file_path:
            logger.error(f"Файл для sheet_id {sheet_id} не найден")
            return False
        
        file_path = sheet_meta.local_file_path
        logger.info(f"🤖 УМНЫЙ ПАРСИНГ: {Path(file_path).name}")
        
        try:
            # 1. АНАЛИЗИРУЕМ СТРУКТУРУ
            structure = self.analyzer.analyze_file_structure(file_path)
            
            # 2. ПРОВЕРЯЕМ ВОЗМОЖНОСТЬ ПАРСИНГА
            if not self._can_parse_table(structure):
                logger.warning(f"❌ Таблица не может быть спарсена: {structure.issues}")
                self._add_to_unknown_stack(structure)
                return False
            
            # 3. ПАРСИМ ПО НАЙДЕННОЙ СТРУКТУРЕ
            success = self._parse_with_structure(sheet_id, structure)
            
            if success:
                logger.info(f"✅ Таблица успешно спарсена")
                return True
            else:
                logger.warning(f"❌ Ошибка парсинга таблицы")
                self._add_to_unknown_stack(structure)
                return False
                
        except Exception as e:
            logger.error(f"Критическая ошибка парсинга {file_path}: {e}")
            return False
    
    def _can_parse_table(self, structure: TableStructure) -> bool:
        """Проверяет можно ли парсить таблицу"""
        
        # Проверяем минимальную уверенность
        if structure.confidence < self.min_confidence:
            return False
        
        # Проверяем наличие обязательных колонок
        for required_col in self.required_columns:
            if required_col not in structure.columns:
                return False
        
        # Проверяем что нет критических проблем
        critical_issues = [
            "Не найдена колонка с названиями товаров",
            "Слишком мало строк данных"
        ]
        
        for issue in structure.issues:
            if any(critical in issue for critical in critical_issues):
                return False
        
        return True
    
    def _parse_with_structure(self, sheet_id: int, structure: TableStructure) -> bool:
        """Парсит таблицу согласно найденной структуре"""
        
        wb = openpyxl.load_workbook(structure.file_path, data_only=True)
        ws = wb.active
        
        try:
            # Получаем расширенную информацию о колонках (включая множественные цены)
            extended_columns = self._find_all_price_columns(ws, structure)
            
            logger.info(f"Найдены расширенные колонки: {list(extended_columns.keys())}")
            
            # Ищем товары
            products_data = self._extract_products_smart(ws, structure, extended_columns)
            
            logger.info(f"Извлечено товаров: {len(products_data)}")
            
            # Сохраняем в БД
            saved_count = 0
            for product_data in products_data:
                if self._save_product_smart(product_data, sheet_id):
                    saved_count += 1
            
            wb.close()
            
            logger.info(f"Сохранено товаров: {saved_count}/{len(products_data)}")
            
            # Обновляем статус таблицы
            sheet_meta = self.session.query(SheetMetadata).get(sheet_id)
            sheet_meta.status = 'completed'
            self.session.commit()
            
            return saved_count > 0
            
        except Exception as e:
            logger.error(f"Ошибка парсинга с структурой: {e}")
            wb.close()
            return False
    
    def _find_all_price_columns(self, ws, structure: TableStructure) -> Dict[str, Any]:
        """Находит все колонки с ценами (для разных маршрутов доставки)"""
        
        columns = dict(structure.columns)
        
        # Ищем дополнительные ценовые колонки
        price_groups = []  # Группы связанных ценовых колонок
        
        # Анализируем заголовки для поиска маршрутов доставки
        delivery_routes = []
        for row in structure.header_rows:
            for col in range(1, min(21, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    header_text = str(cell_value).lower()
                    
                    # Ищем упоминания маршрутов
                    if any(route in header_text for route in ['жд', 'авиа', 'контейнер', 'доставка']):
                        delivery_routes.append({
                            'name': str(cell_value).strip(),
                            'start_col': col,
                            'type': self._detect_delivery_type(header_text)
                        })
        
        logger.info(f"Найдены маршруты доставки: {[r['name'] for r in delivery_routes]}")
        
        # Для каждого маршрута ищем колонки с ценами
        for route in delivery_routes:
            # Проверяем следующие 5 колонок после маршрута
            for offset in range(0, 6):
                col = route['start_col'] + offset
                if col > ws.max_column:
                    break
                
                # Анализируем заголовок колонки
                header_texts = []
                for row in structure.header_rows:
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        header_texts.append(str(cell_value).lower())
                
                combined_header = ' '.join(header_texts)
                
                # Определяем тип ценовой колонки
                if any(word in combined_header for word in ['$', 'usd', 'долл']) and 'цена' in combined_header:
                    key = f"price_usd_{route['type']}"
                    from adaptive_table_analyzer import ColumnDefinition
                    columns[key] = ColumnDefinition(
                        index=col,
                        letter=openpyxl.utils.get_column_letter(col), 
                        name=combined_header,
                        type='price_usd',
                        confidence=0.9,
                        sample_values=[]
                    )
                elif any(word in combined_header for word in ['руб', '₽', 'rub']) and 'цена' in combined_header:
                    key = f"price_rub_{route['type']}"
                    from adaptive_table_analyzer import ColumnDefinition
                    columns[key] = ColumnDefinition(
                        index=col,
                        letter=openpyxl.utils.get_column_letter(col),
                        name=combined_header, 
                        type='price_rub',
                        confidence=0.9,
                        sample_values=[]
                    )
        
        return columns
    
    def _detect_delivery_type(self, header_text: str) -> str:
        """Определяет тип доставки по заголовку"""
        if 'жд' in header_text or 'железн' in header_text:
            return 'railway'
        elif 'авиа' in header_text:
            return 'air'
        elif 'контейнер' in header_text:
            return 'container'
        elif 'образец' in header_text or 'sample' in header_text:
            return 'sample'
        else:
            return 'standard'
    
    def _extract_products_smart(self, ws, structure: TableStructure, columns: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Извлекает товары с учетом структуры"""
        
        products = []
        name_col = columns['name'].index if 'name' in columns else None
        
        if not name_col:
            logger.error("Не найдена колонка с названиями")
            return products
        
        # Ищем строки с товарами
        for row in range(structure.data_start_row, structure.data_end_row + 1):
            name_value = ws.cell(row=row, column=name_col).value
            
            if not name_value or not self._is_valid_product_name(str(name_value)):
                continue
            
            # Извлекаем данные товара
            product_data = {
                'name': str(name_value).strip(),
                'row': row,
                'prices': []
            }
            
            # Извлекаем основные данные
            if 'quantity' in columns:
                qty_value = ws.cell(row=row, column=columns['quantity'].index).value
                if qty_value:
                    product_data['quantity'] = self._validate_quantity(qty_value)
            
            # Извлекаем все ценовые предложения
            self._extract_all_prices(ws, row, columns, product_data)
            
            # Добавляем товар только если есть валидные данные
            if product_data['prices'] or product_data.get('quantity'):
                products.append(product_data)
        
        return products
    
    def _extract_all_prices(self, ws, row: int, columns: Dict[str, Any], product_data: Dict[str, Any]):
        """Извлекает все ценовые предложения для товара"""
        
        # Группируем ценовые колонки по маршрутам
        price_groups = {}
        base_quantity = product_data.get('quantity')
        
        for col_key, col_info in columns.items():
            if col_key.startswith('price_'):
                # Правильный доступ к атрибутам ColumnDefinition
                col_type = getattr(col_info, 'type', '') if hasattr(col_info, 'type') else col_info.get('type', '')
                route = getattr(col_info, 'route', 'standard') if hasattr(col_info, 'route') else 'standard'
                
                if route not in price_groups:
                    price_groups[route] = {}
                
                # Извлекаем значение
                col_index = getattr(col_info, 'index', None) if hasattr(col_info, 'index') else col_info.get('index')
                cell_value = ws.cell(row=row, column=col_index).value
                if cell_value:
                    if col_type == 'price_usd':
                        price = self._validate_price(cell_value)
                        if price:
                            price_groups[route]['price_usd'] = price
                    elif col_type == 'price_rub':
                        price = self._validate_price(cell_value)
                        if price:
                            price_groups[route]['price_rub'] = price
        
        # Создаем ценовые предложения
        for route, prices in price_groups.items():
            if prices:  # Есть хотя бы одна цена
                price_offer = {
                    'route_name': route,
                    'quantity': base_quantity or 1,
                    'price_usd': prices.get('price_usd'),
                    'price_rub': prices.get('price_rub'),
                    'delivery_time': None  # TODO: извлекать если есть колонка
                }
                
                product_data['prices'].append(price_offer)
    
    def _is_valid_product_name(self, name: str) -> bool:
        """Проверяет валидность названия товара"""
        name = name.strip()
        
        if len(name) < 3:
            return False
        
        # Исключаемые паттерны
        exclude_patterns = [
            r'^\d+$',  # Только числа
            r'^фото$',
            r'^наименование$',
            r'^название$',
            r'^товар$',
            r'^продукт$',
            r'^unnamed:\s*\d+$'
        ]
        
        for pattern in exclude_patterns:
            if re.match(pattern, name.lower()):
                return False
        
        return True
    
    def _validate_quantity(self, value: Any) -> Optional[int]:
        """Валидация тиража"""
        if isinstance(value, (int, float)):
            qty = int(value)
            if 1 <= qty <= 100000:
                return qty
        return None
    
    def _validate_price(self, value: Any) -> Optional[float]:
        """Валидация цены"""
        if isinstance(value, (int, float)):
            price = float(value)
            if 0.01 <= price <= 10000:
                return price
        return None
    
    def _save_product_smart(self, product_data: Dict[str, Any], sheet_id: int) -> bool:
        """Сохраняет товар в БД с валидацией"""
        
        try:
            # Создаем товар
            product = Product(
                name=product_data['name'],
                sheet_id=sheet_id,
                start_row=product_data['row'],
                end_row=product_data['row']
            )
            
            self.session.add(product)
            self.session.flush()  # Получаем ID
            
            # Сохраняем ценовые предложения
            saved_prices = 0
            for price_data in product_data['prices']:
                if price_data.get('price_usd') or price_data.get('price_rub'):
                    price_offer = PriceOffer(
                        product_id=product.id,
                        route_name=price_data['route_name'],
                        quantity=price_data['quantity'],
                        price_usd=price_data.get('price_usd'),
                        price_rub=price_data.get('price_rub'),
                        delivery_time=price_data.get('delivery_time')
                    )
                    self.session.add(price_offer)
                    saved_prices += 1
            
            self.session.commit()
            
            logger.info(f"✅ Сохранен товар: {product_data['name']} с {saved_prices} ценовыми предложениями")
            return True
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"❌ Ошибка сохранения товара {product_data['name']}: {e}")
            return False
    
    def _add_to_unknown_stack(self, structure: TableStructure):
        """Добавляет неопознанную таблицу в стек"""
        self.unknown_tables.append(structure)
        logger.warning(f"📚 Таблица добавлена в стек неопознанных: {Path(structure.file_path).name}")
    
    def get_unknown_tables_report(self) -> str:
        """Создает отчет по неопознанным таблицам"""
        if not self.unknown_tables:
            return "✅ Все таблицы успешно опознаны и спарсены"
        
        report = f"📚 НЕОПОЗНАННЫЕ ТАБЛИЦЫ ({len(self.unknown_tables)}):\n"
        report += "=" * 60 + "\n"
        
        for i, structure in enumerate(self.unknown_tables, 1):
            report += f"\n{i}. {Path(structure.file_path).name}\n"
            report += f"   Тип: {structure.table_type} (уверенность: {structure.confidence:.2f})\n"
            report += f"   Колонки: {list(structure.columns.keys())}\n"
            report += f"   Проблемы: {structure.issues}\n"
        
        return report

if __name__ == "__main__":
    # Тестируем умный парсер
    parser = SmartAdaptiveParser()
    
    # Тестируем на одной таблице
    test_sheet_id = 1  # Замените на реальный ID
    success = parser.parse_excel_smart(test_sheet_id)
    
    print(f"Результат парсинга: {'✅ Успешно' if success else '❌ Ошибка'}")
    print(parser.get_unknown_tables_report())
