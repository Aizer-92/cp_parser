#!/usr/bin/env python3
"""
Скрипт для повторного скачивания изображений с правильным определением позиции
"""

import sys
import os
import re
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage
from scripts.google_sheets_downloader import GoogleSheetsDownloader

def extract_image_position_from_filename(filename):
    """Извлекает позицию изображения из имени файла"""
    # Ищем паттерн _image{number} в имени файла
    match = re.search(r'_image(\d+)', filename)
    if match:
        return int(match.group(1))
    return None

def find_product_for_image_position(excel_file_path, image_position):
    """Находит товар для изображения по его позиции в Excel файле"""
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Столбец с названиями товаров (обычно 2)
        name_column = 2
        
        # Ищем товар, который начинается в строке image_position или раньше
        # и заканчивается в строке image_position или позже
        current_product = None
        current_start = None
        current_end = None
        
        for row in range(1, min(100, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=name_column).value
            
            if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower() and 'менеджер' not in str(cell_value).lower():
                # Найден новый товар
                if current_product and current_start and current_end:
                    # Проверяем, попадает ли image_position в диапазон предыдущего товара
                    if current_start <= image_position <= current_end:
                        wb.close()
                        return current_product, current_start, current_end
                
                # Начинаем новый товар
                current_product = str(cell_value).strip()
                current_start = row
                current_end = row
                
                # Ищем конец диапазона товара (следующую непустую ячейку в столбце названий)
                for next_row in range(row + 1, min(row + 20, ws.max_row + 1)):
                    next_cell_value = ws.cell(row=next_row, column=name_column).value
                    if next_cell_value and str(next_cell_value).strip() and 'наименование' not in str(next_cell_value).lower() and 'менеджер' not in str(next_cell_value).lower():
                        # Найдена следующая непустая ячейка
                        current_end = next_row - 1
                        break
                    else:
                        # Пустая ячейка - продолжаем диапазон
                        current_end = next_row
            elif current_product and current_start and current_end:
                # Проверяем, попадает ли image_position в диапазон текущего товара
                if current_start <= image_position <= current_end:
                    wb.close()
                    return current_product, current_start, current_end
        
        # Проверяем последний товар
        if current_product and current_start and current_end and current_start <= image_position <= current_end:
            wb.close()
            return current_product, current_start, current_end
            
        wb.close()
        return None, None, None
        
    except Exception as e:
        print(f"Ошибка при анализе файла {excel_file_path}: {e}")
        return None, None, None

def redownload_images_with_correct_positioning():
    """Повторно скачивает изображения с правильным определением позиции"""
    session = DatabaseManager.get_session()
    
    print("=== ПОВТОРНОЕ СКАЧИВАНИЕ ИЗОБРАЖЕНИЙ С ПРАВИЛЬНОЙ ПОЗИЦИЕЙ ===\n")
    
    # Получаем все таблицы
    sheets = session.query(SheetMetadata).all()
    
    total_downloaded = 0
    total_errors = 0
    
    for sheet in sheets:
        print(f"📊 Обрабатываем таблицу: {sheet.sheet_title}")
        
        if not os.path.exists(sheet.local_file_path):
            print(f"   ❌ Файл не найден: {sheet.local_file_path}")
            continue
        
        try:
            # Скачиваем изображения для этой таблицы
            downloader = GoogleSheetsDownloader()
            images_info = downloader.download_images_from_sheet(sheet.sheet_id, sheet.sheet_title)
            
            print(f"   📥 Скачано изображений: {len(images_info)}")
            
            # Обрабатываем каждое изображение
            for image_info in images_info:
                try:
                    # Извлекаем позицию изображения из имени файла
                    image_position = extract_image_position_from_filename(image_info['filename'])
                    
                    if image_position is None:
                        print(f"   ⚠️  Не удалось определить позицию для: {image_info['filename']}")
                        total_errors += 1
                        continue
                    
                    # Находим товар для этой позиции
                    product_name, start_row, end_row = find_product_for_image_position(
                        sheet.local_file_path, image_position
                    )
                    
                    if product_name is None:
                        print(f"   ⚠️  Не найден товар для позиции {image_position}: {image_info['filename']}")
                        total_errors += 1
                        continue
                    
                    # Находим товар в базе данных
                    product = session.query(Product).filter(
                        Product.name == product_name,
                        Product.sheet_id == sheet.id
                    ).first()
                    
                    if not product:
                        print(f"   ⚠️  Товар '{product_name}' не найден в базе данных")
                        total_errors += 1
                        continue
                    
                    # Определяем тип изображения (main если из столбца A, additional иначе)
                    # Для простоты считаем, что все изображения из столбца A - main
                    image_type = 'main'  # В реальности нужно определять по столбцу
                    
                    # Создаем запись об изображении
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=image_info['local_path'],
                        image_type=image_type,
                        file_size=image_info.get('file_size', 0),
                        width=image_info.get('width', 0),
                        height=image_info.get('height', 0),
                        format=image_info.get('format', 'unknown')
                    )
                    
                    session.add(product_image)
                    total_downloaded += 1
                    
                    print(f"   ✅ {image_info['filename']} → {product_name} (строки {start_row}-{end_row})")
                    
                except Exception as e:
                    print(f"   ❌ Ошибка при обработке {image_info['filename']}: {e}")
                    total_errors += 1
            
            session.commit()
            print(f"   ✅ Таблица {sheet.sheet_title} обработана")
            
        except Exception as e:
            print(f"   ❌ Ошибка при обработке таблицы {sheet.sheet_title}: {e}")
            total_errors += 1
    
    session.close()
    
    print(f"\n=== СТАТИСТИКА ===")
    print(f"Скачано изображений: {total_downloaded}")
    print(f"Ошибок: {total_errors}")
    print(f"\n✅ Обработка завершена!")

if __name__ == "__main__":
    try:
        redownload_images_with_correct_positioning()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

