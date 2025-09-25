#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Комплексный скрипт для полного парсинга Google Sheets файлов в 4 этапа:
1. Скачивание Google Sheets с изображениями
2. Извлечение изображений с позициями и ID таблицы
3. Парсинг товаров с диапазонами строк
4. Привязка изображений к товарам и парсинг матрицы цен
"""

import os
import sys
import time
import re
from pathlib import Path
from datetime import datetime

# Добавляем корневую директорию в путь
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import DatabaseManager
from database.models_v4 import Product, ProductImage, PriceOffer, SheetMetadata, Base, ProjectMetadata
from sqlalchemy import text
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image, ImageDraw
import requests
import json

class CompleteParsing:
    def __init__(self, google_sheets_urls_file="google_sheets_urls.txt"):
        """
        Инициализация парсера
        
        Args:
            google_sheets_urls_file: файл со списком URL Google Sheets
        """
        self.base_dir = Path(__file__).parent.parent
        self.storage_dir = self.base_dir / "storage"
        self.images_dir = self.storage_dir / "images"
        self.excel_dir = self.storage_dir / "excel_files"
        
        # Создаем директории
        self.images_dir.mkdir(parents=True, exist_ok=True)
        self.excel_dir.mkdir(parents=True, exist_ok=True)
        
        self.session = DatabaseManager.get_session()
        self.urls_file = self.base_dir / google_sheets_urls_file
        
        print(f"📁 Рабочие директории:")
        print(f"   📊 Excel файлы: {self.excel_dir}")
        print(f"   🖼️ Изображения: {self.images_dir}")
        
    def __del__(self):
        """Закрытие сессии при завершении"""
        if hasattr(self, 'session'):
            self.session.close()
    
    def load_google_sheets_urls(self):
        """Загружает список URL Google Sheets из файла"""
        if not self.urls_file.exists():
            print(f"❌ Файл {self.urls_file} не найден!")
            print("Создайте файл google_sheets_urls.txt со списком URL, по одному на строку")
            return []
        
        urls = []
        with open(self.urls_file, 'r', encoding='utf-8') as f:
            for line in f:
                url = line.strip()
                if url and not url.startswith('#'):  # Игнорируем комментарии
                    urls.append(url)
        
        print(f"📋 Загружено {len(urls)} URL для обработки")
        return urls
    
    def get_processed_sheets(self):
        """Получает список уже обработанных таблиц"""
        processed = set()
        sheets = self.session.query(SheetMetadata).all()
        for sheet in sheets:
            if sheet.sheet_url:
                processed.add(sheet.sheet_url)
        
        print(f"✅ Уже обработано таблиц: {len(processed)}")
        return processed
    
    def convert_google_sheets_url_to_excel(self, url):
        """Конвертирует URL Google Sheets в URL для скачивания Excel"""
        # Извлекаем ID таблицы из URL
        if '/spreadsheets/d/' in url:
            sheet_id = url.split('/spreadsheets/d/')[1].split('/')[0]
            # Формируем URL для скачивания в формате Excel
            download_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
            return download_url
        return None
    
    def download_excel_file(self, url, filename):
        """Скачивает Excel файл с Google Sheets"""
        download_url = self.convert_google_sheets_url_to_excel(url)
        if not download_url:
            print(f"❌ Не удалось получить URL для скачивания: {url}")
            return None
        
        file_path = self.excel_dir / filename
        
        try:
            print(f"📥 Скачиваем: {filename}")
            response = requests.get(download_url, timeout=60)
            response.raise_for_status()
            
            with open(file_path, 'wb') as f:
                f.write(response.content)
            
            print(f"✅ Скачан: {file_path} ({len(response.content)} bytes)")
            return file_path
            
        except Exception as e:
            print(f"❌ Ошибка скачивания {filename}: {e}")
            return None
    
    def extract_images_from_excel(self, excel_path, sheet_metadata_id):
        """
        ЭТАП 1: Извлекает изображения из Excel файла с позициями
        """
        print(f"\n📸 ЭТАП 1: Извлечение изображений из {excel_path.name}")
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            extracted_count = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(f"   📋 Обрабатываем лист: {sheet_name}")
                
                # Извлекаем изображения из листа
                if hasattr(worksheet, '_images') and worksheet._images:
                    for img in worksheet._images:
                        try:
                            # Определяем позицию изображения
                            anchor = img.anchor
                            if hasattr(anchor, '_from'):
                                col_idx = anchor._from.col
                                row_idx = anchor._from.row + 1  # openpyxl использует 0-based индексы
                                
                                # Конвертируем номер столбца в букву
                                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                                
                                # Пропускаем изображения из A1 (логотипы)
                                if col_letter == 'A' and row_idx == 1:
                                    print(f"   🚫 Пропускаем логотип из A1")
                                    continue
                                
                                # Определяем тип изображения
                                image_type = 'main' if col_letter == 'A' else 'additional'
                                
                                # Генерируем имя файла
                                timestamp = int(time.time())
                                safe_sheet_name = re.sub(r'[^\w\-_]', '_', sheet_name)
                                filename = f"{safe_sheet_name}_{col_letter}{row_idx}_{image_type}_{timestamp}.png"
                                
                                # Сохраняем изображение
                                image_path = self.images_dir / filename
                                
                                # Извлекаем данные изображения
                                if hasattr(img, '_data'):
                                    image_data = img._data()
                                    with open(image_path, 'wb') as f:
                                        f.write(image_data)
                                elif hasattr(img, 'ref'):
                                    # Альтернативный способ извлечения
                                    pil_img = Image.open(img.ref)
                                    pil_img.save(image_path, 'PNG')
                                
                                # Добавляем в базу данных
                                product_image = ProductImage(
                                    sheet_id=sheet_metadata_id,
                                    local_path=str(image_path.relative_to(self.base_dir)),
                                    image_type=image_type,
                                    row=row_idx,
                                    column=col_letter,
                                    position={
                                        'row': row_idx,
                                        'col': col_idx,
                                        'cell': f"{col_letter}{row_idx}",
                                        'worksheet': sheet_name,
                                        'found': True
                                    }
                                )
                                
                                self.session.add(product_image)
                                extracted_count += 1
                                
                                print(f"   ✅ Извлечено: {filename} ({col_letter}{row_idx}, {image_type})")
                                
                        except Exception as e:
                            print(f"   ❌ Ошибка обработки изображения: {e}")
                            continue
            
            self.session.commit()
            print(f"✅ ЭТАП 1 завершен: извлечено {extracted_count} изображений")
            return extracted_count
            
        except Exception as e:
            print(f"❌ Ошибка извлечения изображений: {e}")
            self.session.rollback()
            return 0
    
    def parse_products_from_excel(self, excel_path, sheet_metadata_id):
        """
        ЭТАП 2: Парсит товары из Excel с диапазонами строк
        """
        print(f"\n📦 ЭТАП 2: Парсинг товаров из {excel_path.name}")
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            products_count = 0
            
            # Обрабатываем только основные листы с товарами (исключаем служебные)
            main_sheets = []
            for sheet_name in workbook.sheetnames:
                # Исключаем служебные листы
                skip_keywords = ['цена', 'price', 'доставка', 'delivery', 'образец', 'sample', 'условия', 'terms']
                if not any(keyword in sheet_name.lower() for keyword in skip_keywords):
                    main_sheets.append(sheet_name)
            
            if not main_sheets:
                main_sheets = [workbook.sheetnames[0]]  # Если все исключены, берем первый
            
            print(f"   📋 Найдено основных листов: {len(main_sheets)} из {len(workbook.sheetnames)}")
            
            for sheet_name in main_sheets:
                worksheet = workbook[sheet_name]
                print(f"   📋 Обрабатываем лист: {sheet_name}")
                
                # Определяем структуру столбцов - ищем более широко
                header_row = None
                for row_idx in range(1, 10):  # Ищем заголовки в первых 10 строках
                    row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                    # Расширенный список ключевых слов для русских и английских таблиц
                    header_keywords = [
                        'название', 'товар', 'product', 'name', 'наименование', 'артикул', 'код',
                        'item', 'title', 'description', 'goods', 'merchandise'
                    ]
                    if any(cell and any(keyword in str(cell).lower() for keyword in header_keywords) for cell in row):
                        header_row = row_idx
                        print(f"   ✅ Найдена строка заголовков: {row_idx}")
                        break
                
                if not header_row:
                    print(f"   ⚠️ Заголовки не найдены в листе {sheet_name}")
                    continue
                
                # Получаем заголовки
                headers = list(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
                
                # Ищем колонку с названием товара
                name_col = None
                for col_idx, header in enumerate(headers):
                    if header and any(keyword in str(header).lower() for keyword in header_keywords):
                        name_col = col_idx
                        print(f"   ✅ Найден столбец с названием: {col_idx} ('{header}')")
                        break
                
                if name_col is None:
                    print(f"   ⚠️ Колонка с названием товара не найдена в листе {sheet_name}")
                    continue
                
                # Парсим товары
                current_product = None
                product_start_row = None
                
                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                    
                    if len(row) > name_col and row[name_col]:
                        product_name = str(row[name_col]).strip()
                        
                        # Пропускаем служебные строки
                        if any(skip_word in product_name.lower() for skip_word in ['менеджер', 'итого', 'всего', 'сумма']):
                            continue
                        
                        # Завершаем предыдущий товар
                        if current_product and product_start_row:
                            current_product.end_row = row_idx - 1
                            self.session.add(current_product)
                            products_count += 1
                            print(f"   ✅ Товар: {current_product.name} (строки {current_product.start_row}-{current_product.end_row})")
                        
                        # Создаем новый товар
                        current_product = Product(
                            sheet_id=sheet_metadata_id,
                            name=product_name,
                            start_row=row_idx,
                            end_row=row_idx  # Будет обновлено позже
                        )
                        product_start_row = row_idx
                
                # Завершаем последний товар
                if current_product and product_start_row:
                    current_product.end_row = worksheet.max_row
                    self.session.add(current_product)
                    products_count += 1
                    print(f"   ✅ Товар: {current_product.name} (строки {current_product.start_row}-{current_product.end_row})")
            
            self.session.commit()
            print(f"✅ ЭТАП 2 завершен: добавлено {products_count} товаров")
            return products_count
            
        except Exception as e:
            print(f"❌ Ошибка парсинга товаров: {e}")
            self.session.rollback()
            return 0
    
    def assign_images_to_products(self, sheet_metadata_id):
        """
        ЭТАП 3: Привязывает изображения к товарам по позициям
        ЖЕСТКАЯ ПРОВЕРКА: только изображения и товары из одной таблицы
        """
        print(f"\n🔗 ЭТАП 3: Привязка изображений к товарам (СТРОГАЯ ПРОВЕРКА)")
        
        try:
            # Получаем все изображения для данной таблицы
            images = self.session.query(ProductImage).filter(
                ProductImage.sheet_id == sheet_metadata_id,
                ProductImage.row.isnot(None)
            ).all()
            
            # Получаем все товары для данной таблицы
            products = self.session.query(Product).filter(
                Product.sheet_id == sheet_metadata_id
            ).all()
            
            print(f"   📊 Изображений в таблице {sheet_metadata_id}: {len(images)}")
            print(f"   📦 Товаров в таблице {sheet_metadata_id}: {len(products)}")
            
            # КРИТИЧНО: Сначала отвязываем ВСЕ изображения от товаров из других таблиц
            wrong_images = self.session.query(ProductImage).filter(
                ProductImage.product_id.isnot(None)
            ).join(Product, ProductImage.product_id == Product.id).filter(
                Product.sheet_id != sheet_metadata_id,
                ProductImage.sheet_id == sheet_metadata_id
            ).all()
            
            if wrong_images:
                print(f"   ❌ Найдено {len(wrong_images)} изображений с неправильной привязкой - отвязываем")
                for img in wrong_images:
                    img.product_id = None
            
            assigned_count = 0
            
            for image in images:
                # Пропускаем изображения из A1 (логотипы)
                if image.column == 'A' and image.row == 1:
                    continue
                
                # СТРОГАЯ ПРОВЕРКА: ищем товар ТОЛЬКО в той же таблице
                target_product = None
                for product in products:
                    # Дополнительная проверка что товар из той же таблицы
                    if (product.sheet_id == sheet_metadata_id and 
                        product.start_row <= image.row <= product.end_row):
                        target_product = product
                        break
                
                if target_product:
                    # Дополнительная проверка перед назначением
                    if target_product.sheet_id == sheet_metadata_id and image.sheet_id == sheet_metadata_id:
                        image.product_id = target_product.id
                        assigned_count += 1
                        print(f"   ✅ {image.column}{image.row} → {target_product.name} (таблица {sheet_metadata_id})")
                    else:
                        print(f"   ❌ БЛОКИРОВАНО: попытка привязать изображение таблицы {image.sheet_id} к товару таблицы {target_product.sheet_id}")
                else:
                    print(f"   ⚠️ Не найден товар для изображения {image.column}{image.row} в таблице {sheet_metadata_id}")
            
            self.session.commit()
            print(f"✅ ЭТАП 3 завершен: привязано {assigned_count} изображений (СТРОГО в рамках таблицы {sheet_metadata_id})")
            return assigned_count
            
        except Exception as e:
            print(f"❌ Ошибка привязки изображений: {e}")
            self.session.rollback()
            return 0
    
    def parse_price_matrix(self, excel_path, sheet_metadata_id):
        """
        ЭТАП 4: Парсит матрицу цен (тираж + маршрут)
        """
        print(f"\n💰 ЭТАП 4: Парсинг матрицы цен из {excel_path.name}")
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            variants_count = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(f"   📋 Обрабатываем лист: {sheet_name}")
                
                # Получаем все товары для этой таблицы
                products = self.session.query(Product).filter(
                    Product.sheet_id == sheet_metadata_id
                ).all()
                
                # Ищем матрицу цен для каждого товара
                for product in products:
                    product_variants = self.extract_price_variants_for_product(
                        worksheet, product, sheet_metadata_id
                    )
                    variants_count += len(product_variants)
            
            print(f"✅ ЭТАП 4 завершен: добавлено {variants_count} вариантов цен")
            return variants_count
            
        except Exception as e:
            print(f"❌ Ошибка парсинга матрицы цен: {e}")
            self.session.rollback()
            return 0
    
    def extract_price_variants_for_product(self, worksheet, product, sheet_metadata_id):
        """Извлекает варианты цен для конкретного товара"""
        variants = []
        
        try:
            # Сначала найдем заголовки столбцов для определения структуры
            header_row = None
            for row_idx in range(1, 10):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                # Ищем строку с ценами (русские и английские варианты)
                price_keywords = ['цена', 'price', 'тираж', 'quantity', '$', 'usd', 'руб', 'rub']
                if any(cell and any(keyword in str(cell).lower() for keyword in price_keywords) for cell in row):
                    header_row = row_idx
                    break
            
            # Определяем индексы столбцов
            quantity_cols = []
            usd_cols = []
            rub_cols = []
            delivery_cols = []
            
            if header_row:
                headers = list(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
                for col_idx, header in enumerate(headers):
                    if header:
                        header_str = str(header).lower()
                        # Поиск столбцов с тиражами (русский и английский)
                        if any(word in header_str for word in ['тираж', 'шт', 'quantity', 'pcs']):
                            quantity_cols.append(col_idx)
                        # Поиск столбцов с ценами в долларах
                        elif any(word in header_str for word in ['$', 'usd', 'dollar']) or ('price' in header_str and '$' in header_str):
                            usd_cols.append(col_idx)
                        # Поиск столбцов с ценами в рублях/дирхамах
                        elif any(word in header_str for word in ['руб', 'rub', 'ruble', 'aed', 'дирхам']):
                            rub_cols.append(col_idx)
                        # Поиск столбцов со сроками доставки
                        elif any(word in header_str for word in ['доставка', 'срок', 'delivery', 'period', 'time']):
                            delivery_cols.append(col_idx)
            
            # Ищем строки с ценами в диапазоне товара
            for row_idx in range(product.start_row, product.end_row + 1):
                row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                
                quantity = None
                price_usd = None
                price_rub = None
                delivery_time = None
                
                # Используем найденные индексы столбцов
                for col_idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    
                    cell_str = str(cell_value).strip()
                    
                    # Тираж из определенных столбцов
                    if col_idx in quantity_cols and cell_str.replace('.', '').isdigit():
                        potential_quantity = int(float(cell_str))
                        if 10 <= potential_quantity <= 100000:
                            quantity = potential_quantity
                    
                    # USD из определенных столбцов
                    elif col_idx in usd_cols and self.is_decimal_number(cell_str):
                        try:
                            price_usd = float(cell_str.replace(',', '.'))
                        except:
                            pass
                    
                    # RUB из определенных столбцов
                    elif col_idx in rub_cols and self.is_decimal_number(cell_str):
                        try:
                            price_rub = float(cell_str.replace(',', '.'))
                        except:
                            pass
                    
                    # Доставка из определенных столбцов
                    elif col_idx in delivery_cols:
                        delivery_time = cell_str
                    
                    # Fallback: поиск по содержимому (старая логика)
                    elif not quantity and cell_str.isdigit():
                        potential_quantity = int(cell_str)
                        if 10 <= potential_quantity <= 100000:
                            quantity = potential_quantity
                    
                    elif not price_usd and ('$' in cell_str or self.is_decimal_number(cell_str)):
                        try:
                            price_usd = float(re.sub(r'[^\d.,]', '', cell_str).replace(',', '.'))
                        except:
                            pass
                    
                    elif not price_rub and ('₽' in cell_str or 'руб' in cell_str):
                        try:
                            price_rub = float(re.sub(r'[^\d.,]', '', cell_str).replace(',', '.'))
                        except:
                            pass
                    
                    elif not delivery_time and ('дн' in cell_str or 'день' in cell_str or 'раб' in cell_str):
                        delivery_time = cell_str
                
                # Если нашли хотя бы тираж и одну цену, создаем вариант
                if quantity and (price_usd or price_rub):
                    # Определяем маршрут доставки по цене или времени
                    route_name = "Standard"
                    if delivery_time:
                        if any(word in delivery_time.lower() for word in ['авиа', 'air', 'fast']):
                            route_name = "АВИА"
                        elif any(word in delivery_time.lower() for word in ['жд', 'rail', 'train']):
                            route_name = "ЖД"
                        elif any(word in delivery_time.lower() for word in ['контейнер', 'container', 'sea']):
                            route_name = "Контейнер"
                    
                    variant = PriceOffer(
                        product_id=product.id,
                        route_name=route_name,
                        quantity=quantity,
                        price_usd=price_usd,
                        price_rub=price_rub,
                        delivery_time=delivery_time
                    )
                    
                    self.session.add(variant)
                    variants.append(variant)
                    
                    print(f"     ✅ Вариант: {quantity} шт, ${price_usd or '—'}, {price_rub or '—'} ₽, {delivery_time or '—'}")
        
        except Exception as e:
            print(f"   ❌ Ошибка извлечения вариантов для {product.name}: {e}")
        
        try:
            self.session.commit()
        except Exception as e:
            print(f"   ❌ Ошибка сохранения вариантов: {e}")
            self.session.rollback()
        
        return variants
    
    def is_decimal_number(self, s):
        """Проверяет, является ли строка десятичным числом"""
        try:
            float(s.replace(',', '.'))
            return True
        except:
            return False
    
    def process_single_sheet(self, url):
        """Обрабатывает одну Google Sheets таблицу полностью"""
        print(f"\n{'='*60}")
        print(f"🚀 НАЧИНАЕМ ОБРАБОТКУ ТАБЛИЦЫ")
        print(f"📄 URL: {url}")
        print(f"{'='*60}")
        
        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"google_sheet_{timestamp}.xlsx"
        
        # Скачиваем файл
        excel_path = self.download_excel_file(url, filename)
        if not excel_path:
            return False
        
        # Проверяем, есть ли уже запись о таблице в БД
        google_sheet_id = url.split('/')[-2] if '/d/' in url else filename.replace('.xlsx', '')
        existing_sheet = self.session.query(SheetMetadata).filter(
            SheetMetadata.sheet_id == google_sheet_id
        ).first()
        
        if existing_sheet:
            print(f"   📋 Таблица уже есть в БД (ID: {existing_sheet.id})")
            sheet_metadata = existing_sheet
        else:
            # Создаем запись о таблице в БД
            sheet_metadata = SheetMetadata(
                sheet_title=filename.replace('.xlsx', ''),
                sheet_url=url,
                sheet_id=google_sheet_id,
                local_file_path=str(excel_path.relative_to(self.base_dir)),
                products_count=0
            )
            
            self.session.add(sheet_metadata)
            self.session.commit()
            print(f"   📋 Создана новая запись в БД (ID: {sheet_metadata.id})")
        
        sheet_id = sheet_metadata.id
        
        try:
            # ЭТАП 1: Парсим товары СНАЧАЛА
            products_count = self.parse_products_from_excel(excel_path, sheet_id)
            
            # ЭТАП 2: Извлекаем изображения
            images_count = self.extract_images_from_excel(excel_path, sheet_id)
            
            # ЭТАП 3: Привязываем изображения к товарам
            assigned_count = self.assign_images_to_products(sheet_id)
            
            # ЭТАП 4: Парсим матрицу цен
            variants_count = self.parse_price_matrix(excel_path, sheet_id)
            
            # Обновляем статистику
            sheet_metadata.products_count = products_count
            self.session.commit()
            
            print(f"\n🎉 ТАБЛИЦА ОБРАБОТАНА УСПЕШНО!")
            print(f"   📊 Товаров: {products_count}")
            print(f"   🖼️ Изображений: {images_count}")
            print(f"   🔗 Привязано: {assigned_count}")
            print(f"   💰 Вариантов цен: {variants_count}")
            
            return True
            
        except Exception as e:
            print(f"❌ Ошибка обработки таблицы: {e}")
            # Удаляем запись о таблице при ошибке
            self.session.delete(sheet_metadata)
            self.session.commit()
            return False
    
    def run_complete_parsing(self):
        """Запускает полный цикл парсинга всех новых таблиц"""
        print("🚀 ЗАПУСК ПОЛНОГО ПАРСИНГА GOOGLE SHEETS")
        print("=" * 60)
        
        # Загружаем список URL
        urls = self.load_google_sheets_urls()
        if not urls:
            return
        
        # Получаем уже обработанные таблицы
        processed_urls = self.get_processed_sheets()
        
        # Фильтруем новые URL
        new_urls = [url for url in urls if url not in processed_urls]
        
        if not new_urls:
            print("✅ Все таблицы уже обработаны!")
            return
        
        print(f"📋 К обработке: {len(new_urls)} новых таблиц")
        
        # Обрабатываем каждую таблицу
        success_count = 0
        for i, url in enumerate(new_urls, 1):
            print(f"\n📊 Таблица {i}/{len(new_urls)}")
            
            if self.process_single_sheet(url):
                success_count += 1
                print(f"✅ Успешно обработано: {success_count}/{i}")
            else:
                print(f"❌ Ошибка обработки таблицы {i}")
            
            # Пауза между обработкой
            if i < len(new_urls):
                print("⏳ Пауза 2 секунды...")
                time.sleep(2)
        
        print(f"\n🎉 ПАРСИНГ ЗАВЕРШЕН!")
        print(f"✅ Успешно обработано: {success_count}/{len(new_urls)} таблиц")

def main():
    """Главная функция"""
    try:
        # Создаем экземпляр парсера
        parser = CompleteParsing()
        
        # Запускаем полный парсинг
        parser.run_complete_parsing()
        
    except KeyboardInterrupt:
        print("\n⚠️ Парсинг прерван пользователем")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()

def parse_single_sheet(google_sheets_url):
    """
    Парсит одну Google Sheets таблицу
    
    Args:
        google_sheets_url: URL Google Sheets таблицы
        
    Returns:
        dict: Результат парсинга с информацией о успехе/ошибке
    """
    try:
        parser = CompleteParsing()
        
        # Создаем временный файл с одним URL
        temp_urls_file = parser.base_dir / "temp_single_url.txt"
        with open(temp_urls_file, 'w', encoding='utf-8') as f:
            f.write(google_sheets_url + '\n')
        
        # Меняем файл URL в парсере
        parser.urls_file = temp_urls_file
        
        print(f"🎯 Парсинг одной таблицы: {google_sheets_url}")
        
        # Запускаем парсинг
        result = parser.run_complete_parsing()
        
        # Удаляем временный файл
        if temp_urls_file.exists():
            temp_urls_file.unlink()
        
        # Получаем статистику
        session = DatabaseManager.get_session()
        try:
            # Находим последнюю созданную таблицу
            latest_sheet = session.query(SheetMetadata).filter(
                SheetMetadata.sheet_url == google_sheets_url
            ).first()
            
            if latest_sheet:
                products_count = session.query(Product).filter(
                    Product.sheet_id == latest_sheet.id
                ).count()
                
                images_count = session.query(ProductImage).filter(
                    ProductImage.sheet_id == latest_sheet.id
                ).count()
                
                return {
                    'success': True,
                    'sheet_id': latest_sheet.id,
                    'products_count': products_count,
                    'images_count': images_count,
                    'message': f'Успешно обработано: {products_count} товаров, {images_count} изображений'
                }
            else:
                return {
                    'success': False,
                    'error': 'Таблица не была создана в базе данных'
                }
                
        finally:
            session.close()
            
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

if __name__ == "__main__":
    main()
