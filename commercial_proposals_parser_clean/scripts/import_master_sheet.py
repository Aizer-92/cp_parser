#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для импорта данных из мастер-таблицы с проектами
"""

import os
import sys
import re
import requests
from datetime import datetime

# Добавляем корневую директорию в путь
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import DatabaseManager
from database.models_v4 import ProjectMetadata
import openpyxl

class MasterSheetImporter:
    def __init__(self):
        self.session = DatabaseManager.get_session()
        self.master_sheet_url = "https://docs.google.com/spreadsheets/d/1UJXbLdFEwsCHhzPPkU6wfiq2bMGZMHWEzCd4bJ5TtpY/export?format=xlsx&gid=1417349258"
        
    def __del__(self):
        if hasattr(self, 'session'):
            self.session.close()
    
    def download_master_sheet(self):
        """Скачивает мастер-таблицу в формате Excel"""
        print("📥 Скачивание мастер-таблицы...")
        
        try:
            response = requests.get(self.master_sheet_url, timeout=60)
            response.raise_for_status()
            
            # Сохраняем во временный файл
            temp_file = "temp_master_sheet.xlsx"
            with open(temp_file, 'wb') as f:
                f.write(response.content)
            
            print(f"✅ Мастер-таблица скачана: {len(response.content)} bytes")
            return temp_file
            
        except Exception as e:
            print(f"❌ Ошибка скачивания мастер-таблицы: {e}")
            return None
    
    def parse_excel_file(self, file_path):
        """Парсит Excel файл мастер-таблицы"""
        print("📊 Парсинг мастер-таблицы...")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            print(f"   📋 Листов в файле: {len(workbook.sheetnames)}")
            print(f"   📄 Имена листов: {workbook.sheetnames}")
            
            worksheet = workbook.active
            print(f"   📊 Активный лист: {worksheet.title}")
            print(f"   📏 Размер листа: {worksheet.max_row} строк, {worksheet.max_column} столбцов")
            
            projects = []
            
            # Ищем строку с заголовками
            header_row = None
            for check_row in range(1, 20):  # Проверяем первые 20 строк
                row_cells = list(worksheet.iter_rows(min_row=check_row, max_row=check_row, values_only=True))
                if row_cells and row_cells[0]:
                    row = row_cells[0]
                    # Ищем строку, содержащую ключевые слова заголовков
                    if any(cell and ('тираж' in str(cell).lower() or 'товары' in str(cell).lower() or 'стоимость' in str(cell).lower()) for cell in row):
                        header_row = check_row
                        print(f"   📋 Найдена строка заголовков: {header_row}")
                        print(f"   📄 Заголовки: {[str(cell)[:20] for cell in row[:10] if cell]}")
                        break
            
            if not header_row:
                print("   ❌ Не найдена строка заголовков")
                return []
            
            # Поиск первой строки с данными - ищем строку с текстом, содержащим "[Просчет"
            data_start_row = None
            for check_row in range(header_row + 1, min(worksheet.max_row + 1, header_row + 100)):
                row_cells = list(worksheet.iter_rows(min_row=check_row, max_row=check_row, values_only=True))
                if row_cells and row_cells[0]:
                    row = row_cells[0]
                    # Ищем строку, содержащую "[Просчет" - характерный признак данных проекта
                    if any(cell and '[Просчет' in str(cell) for cell in row):
                        data_start_row = check_row
                        print(f"   📋 Найдена первая строка с данными: {data_start_row}")
                        # Находим непустые ячейки
                        non_empty_cells = [(i, str(cell)[:50]) for i, cell in enumerate(row) if cell and str(cell).strip()]
                        print(f"   📄 Непустые ячейки: {non_empty_cells[:5]}")
                        break
            
            if not data_start_row:
                print("   ❌ Не найдены строки с данными проектов")
                return []
            
            # Парсим данные, начиная с найденной строки
            for row_idx in range(data_start_row, worksheet.max_row + 1):
                try:
                    row_cells = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
                    
                    if not row_cells:
                        continue
                        
                    row = row_cells[0]
                    
                    # Пропускаем пустые строки
                    if not row or not any(str(cell).strip() for cell in row if cell):
                        continue
                        
                    print(f"   🔍 Обрабатываем строку {row_idx}")
                except Exception as e:
                    print(f"   ❌ Ошибка обработки строки {row_idx}: {e}")
                    continue
                
                # Безопасно извлекаем данные из строки
                def safe_get(index):
                    return row[index] if len(row) > index else None
                
                # Судя по найденным данным, структура такая:
                # 0: Тираж (название проекта) = столбец 9 (J)
                # 1: Тираж MIN = столбец 10 (K) 
                # 2: Тираж MAX = столбец 11 (L)
                # 3: Товары = столбец 12 (M)
                # 4-7: Цены
                # 8: Описание
                # 9: Название = столбец A (видимо проект)
                # 10: Постановщик = столбец B
                # 11: Исполнители = столбец C  
                # 12: URL = столбец D
                # 13: Дата = столбец E
                # 14: Контрагент = столбец F
                # 15: Регион = столбец G
                
                # Судя по выводу, данные находятся в позициях 9-15
                # 9: project_name (основное название проекта)
                # 10: manager (постановщик)
                # 11: executors (исполнители) 
                # 12: google_sheets_url (URL)
                # 13: creation_date (дата)
                # 14: counterparty (контрагент)
                # 15: region (регион)
                
                project_data = {
                    'project_title': self.clean_text(safe_get(9)),   # J: Основное название проекта
                    'min_quantity': self.clean_text(safe_get(10)),   # K: Постановщик (используем как тираж MIN)
                    'max_quantity': self.clean_text(safe_get(11)),   # L: Исполнители (используем как тираж MAX)
                    'products_info': self.clean_text(safe_get(3)),   # D: Товары (если есть)
                    'min_price_usd': self.parse_float(safe_get(4)),  # E: Мин цена USD (если есть)
                    'max_price_usd': self.parse_float(safe_get(5)),  # F: Макс цена USD (если есть)
                    'min_price_rub': self.parse_float(safe_get(6)),  # G: Мин цена RUB (если есть)
                    'max_price_rub': self.parse_float(safe_get(7)),  # H: Макс цена RUB (если есть)
                    'description': self.clean_text(safe_get(8)),     # I: Описание (если есть)
                    'project_name': self.clean_text(safe_get(0)),    # A: Доп название (тираж)
                    'manager': self.clean_text(safe_get(10)),        # K: Постановщик
                    'executors': self.clean_text(safe_get(11)),      # L: Исполнители
                    'google_sheets_url': self.clean_url(safe_get(12)), # M: URL
                    'creation_date': self.clean_text(safe_get(13)),  # N: Дата
                    'counterparty': self.clean_text(safe_get(14)),   # O: Контрагент
                    'region': self.clean_text(safe_get(15))          # P: Регион
                }
                
                # Добавляем только если есть название проекта или URL
                if project_data['project_title'] or project_data['google_sheets_url']:
                    projects.append(project_data)
                    title_preview = str(project_data['project_title'])[:50] if project_data['project_title'] else "Без названия"
                    print(f"   ✅ Проект: {title_preview}...")
            
            print(f"📋 Найдено проектов: {len(projects)}")
            return projects
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Excel файла: {e}")
            return []
    
    def clean_text(self, value):
        """Очищает текстовое значение"""
        if value is None:
            return None
        
        text = str(value).strip()
        return text if text else None
    
    def clean_url(self, value):
        """Очищает и валидирует URL"""
        if value is None:
            return None
        
        url = str(value).strip()
        if url and ('docs.google.com/spreadsheets' in url or 'drive.google.com' in url):
            return url
        return None
    
    def parse_float(self, value):
        """Парсит числовое значение"""
        if value is None:
            return None
        
        try:
            # Удаляем лишние символы и конвертируем
            cleaned = re.sub(r'[^\d.,]', '', str(value))
            if cleaned:
                return float(cleaned.replace(',', '.'))
        except:
            pass
        
        return None
    
    def save_projects_to_db(self, projects):
        """Сохраняет проекты в базу данных"""
        print("💾 Сохранение проектов в базу данных...")
        
        try:
            saved_count = 0
            updated_count = 0
            
            for project_data in projects:
                # Проверяем, существует ли уже такой проект
                existing_project = None
                
                if project_data['google_sheets_url']:
                    existing_project = self.session.query(ProjectMetadata).filter(
                        ProjectMetadata.google_sheets_url == project_data['google_sheets_url']
                    ).first()
                
                if existing_project:
                    # Обновляем существующий проект
                    for key, value in project_data.items():
                        setattr(existing_project, key, value)
                    existing_project.updated_at = datetime.utcnow()
                    updated_count += 1
                    title_preview = str(project_data['project_title'])[:50] if project_data['project_title'] else "Без названия"
                    print(f"   🔄 Обновлен: {title_preview}...")
                else:
                    # Создаем новый проект
                    project = ProjectMetadata(**project_data)
                    self.session.add(project)
                    saved_count += 1
                    title_preview = str(project_data['project_title'])[:50] if project_data['project_title'] else "Без названия"
                    print(f"   ➕ Добавлен: {title_preview}...")
            
            self.session.commit()
            print(f"✅ Сохранено: {saved_count} новых, {updated_count} обновлено")
            
            return saved_count + updated_count
            
        except Exception as e:
            print(f"❌ Ошибка сохранения в БД: {e}")
            self.session.rollback()
            return 0
    
    def get_urls_for_parsing(self):
        """Получает список URL для парсинга"""
        print("🔍 Получение URL для парсинга...")
        
        try:
            # Получаем проекты с URL, которые еще не были обработаны
            projects = self.session.query(ProjectMetadata).filter(
                ProjectMetadata.google_sheets_url.isnot(None),
                ProjectMetadata.processing_status == 'pending'
            ).all()
            
            urls = []
            for project in projects:
                urls.append({
                    'url': project.google_sheets_url,
                    'project_id': project.id,
                    'project_title': project.project_title,
                    'counterparty': project.counterparty
                })
            
            print(f"📋 Найдено URL для парсинга: {len(urls)}")
            return urls
            
        except Exception as e:
            print(f"❌ Ошибка получения URL: {e}")
            return []
    
    def create_urls_file(self, urls):
        """Создает файл google_sheets_urls.txt с URL для парсинга"""
        print("📝 Создание файла google_sheets_urls.txt...")
        
        try:
            with open('google_sheets_urls.txt', 'w', encoding='utf-8') as f:
                f.write("# Список URL Google Sheets для парсинга\n")
                f.write("# Импортировано из мастер-таблицы\n\n")
                
                for url_data in urls:
                    f.write(f"# Проект: {url_data['project_title']}\n")
                    f.write(f"# Контрагент: {url_data['counterparty']}\n")
                    f.write(f"# ID: {url_data['project_id']}\n")
                    f.write(f"{url_data['url']}\n\n")
            
            print(f"✅ Файл создан: {len(urls)} URL добавлено")
            
        except Exception as e:
            print(f"❌ Ошибка создания файла: {e}")
    
    def run_import(self):
        """Запускает полный импорт мастер-таблицы"""
        print("🚀 ИМПОРТ МАСТЕР-ТАБЛИЦЫ")
        print("=" * 50)
        
        try:
            # 1. Скачиваем мастер-таблицу
            excel_file = self.download_master_sheet()
            if not excel_file:
                return False
            
            # 2. Парсим данные
            projects = self.parse_excel_file(excel_file)
            if not projects:
                print("⚠️ Не найдено проектов для импорта")
                return False
            
            # 3. Сохраняем в БД
            saved_count = self.save_projects_to_db(projects)
            
            # 4. Получаем URL для парсинга
            urls = self.get_urls_for_parsing()
            
            # 5. Создаем файл с URL
            if urls:
                self.create_urls_file(urls)
            
            # 6. Удаляем временный файл
            if os.path.exists(excel_file):
                os.remove(excel_file)
            
            print(f"\n🎉 ИМПОРТ ЗАВЕРШЕН!")
            print(f"📊 Всего проектов: {len(projects)}")
            print(f"💾 Сохранено/обновлено: {saved_count}")
            print(f"🔗 URL для парсинга: {len(urls)}")
            
            return True
            
        except Exception as e:
            print(f"❌ Критическая ошибка импорта: {e}")
            return False

def main():
    """Главная функция"""
    try:
        importer = MasterSheetImporter()
        importer.run_import()
        
    except KeyboardInterrupt:
        print("\n⚠️ Импорт прерван пользователем")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
