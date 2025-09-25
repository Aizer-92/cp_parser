#!/usr/bin/env python3
"""
Улучшенный парсер товаров с автоматическим определением диапазонов строк
"""

import sys
import logging
from pathlib import Path
import openpyxl
import json
from decimal import Decimal, InvalidOperation

# Добавляем путь к проекту
sys.path.append(str(Path(__file__).parent.parent))

from database.manager_v4 import DatabaseManager
from database.models_v4 import Product, SheetMetadata, ProductImage, PriceOffer

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s:%(message)s')
logger = logging.getLogger(__name__)

def parse_price_value(cell_value):
    """Парсит числовое значение цены из ячейки."""
    if isinstance(cell_value, (int, float, Decimal)):
        return float(cell_value)
    
    s_value = str(cell_value).strip().replace(' ', '').replace(',', '.')
    try:
        import re
        match = re.match(r'(\d+(\.\d+)?)', s_value)
        if match:
            return float(match.group(1))
    except ValueError:
        pass
    return None

def parse_time_value(cell_value):
    """Парсит значение срока из ячейки, возвращая строку."""
    if cell_value is None:
        return None
    return str(cell_value).strip()

def find_product_row_range(ws, product_name, name_column, start_search_row=1, max_search_rows=50):
    """Находит диапазон строк для товара по его названию.
    Диапазон = строка с названием + все пустые ячейки в столбце названия до следующего непустого значения."""
    
    for row in range(start_search_row, min(start_search_row + max_search_rows, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=name_column).value
        if cell_value and str(cell_value).strip().lower() == product_name.lower():
            # Найдена строка с названием товара
            start_row = row
            
            # Ищем конец диапазона - следующую непустую ячейку в столбце названия
            end_row = start_row
            for next_row in range(start_row + 1, min(start_row + 20, ws.max_row + 1)):  # Проверяем до 20 строк вперед
                next_cell_value = ws.cell(row=next_row, column=name_column).value
                if next_cell_value and str(next_cell_value).strip():
                    # Найдена следующая непустая ячейка в столбце названия
                    end_row = next_row - 1
                    break
                else:
                    # Пустая ячейка - продолжаем диапазон
                    end_row = next_row
            
            return start_row, end_row
    
    return None, None

class ImprovedProductParser:
    def __init__(self):
        self.db_manager = DatabaseManager

    def parse_all_products(self):
        """Улучшенный парсинг товаров с определением диапазонов строк."""
        logger.info("=== УЛУЧШЕННЫЙ ПАРСИНГ ТОВАРОВ ===")
        
        session = self.db_manager.get_session()
        
        try:
            # Очищаем старые данные
            session.query(Product).delete()
            session.query(PriceOffer).delete()
            session.commit()
            logger.info("Очистка старых данных завершена")

            sheets = session.query(SheetMetadata).all()
            logger.info(f"Найдено таблиц: {len(sheets)}")
            
            total_products = 0
            total_samples = 0

            for sheet in sheets:
                if not sheet.local_file_path or not Path(sheet.local_file_path).exists():
                    logger.warning(f"Файл не найден для таблицы {sheet.sheet_title}: {sheet.local_file_path}. Пропускаем.")
                    continue
                
                logger.info(f"\n=== ОБРАБОТКА ТАБЛИЦЫ: {sheet.sheet_title} ===")
                
                try:
                    wb = openpyxl.load_workbook(sheet.local_file_path, data_only=True)
                    ws = wb.active
                    
                    # Ищем столбцы
                    name_column = None
                    characteristics_column = None
                    custom_column = None
                    sample_column = None
                    sample_time_column = None
                    
                    # Ищем в строках 1-3
                    for row_idx in range(1, 4):
                        for col in range(1, min(ws.max_column + 1, 20)):
                            cell_value = ws.cell(row=row_idx, column=col).value
                            if cell_value:
                                lower_value = str(cell_value).lower()
                                if 'наименование' in lower_value or 'название' in lower_value:
                                    name_column = col
                                elif 'характеристики' in lower_value:
                                    characteristics_column = col
                                elif 'кастом' in lower_value:
                                    custom_column = col
                                elif 'образец' in lower_value:
                                    sample_column = col
                                elif 'срок' in lower_value and ('фото' in lower_value or 'видео' in lower_value):
                                    sample_time_column = col
                    
                    logger.info(f"  Найденные столбцы: наименование={name_column}, характеристики={characteristics_column}, кастом={custom_column}, образец={sample_column}, срок образца={sample_time_column}")
                    
                    if not name_column:
                        logger.warning("  Столбец с наименованием не найден, пропускаем таблицу")
                        continue
                    
                    # Ищем товары
                    sheet_products = 0
                    sheet_samples = 0
                    
                    for row in range(2, ws.max_row + 1):  # Начинаем с строки 2
                        name_cell = ws.cell(row=row, column=name_column).value
                        if not name_cell or not str(name_cell).strip():
                            continue
                        
                        product_name = str(name_cell).strip()
                        
                        # Определяем диапазон строк для товара
                        start_row, end_row = find_product_row_range(ws, product_name, name_column, row)
                        if not start_row:
                            start_row = row
                            end_row = row
                        
                        # Получаем характеристики
                        characteristics = None
                        if characteristics_column:
                            char_cell = ws.cell(row=start_row, column=characteristics_column).value
                            if char_cell:
                                characteristics = str(char_cell).strip()
                        
                        # Получаем кастом
                        custom_design = None
                        if custom_column:
                            custom_cell = ws.cell(row=start_row, column=custom_column).value
                            if custom_cell:
                                custom_design = str(custom_cell).strip()
                        
                        # Создаем товар
                        product = Product(
                            name=product_name,
                            characteristics=characteristics,
                            custom_design=custom_design,
                            sheet_id=sheet.id,
                            start_row=start_row,
                            end_row=end_row
                        )
                        session.add(product)
                        session.flush()  # Получаем ID
                        
                        sheet_products += 1
                        total_products += 1
                        
                        # Парсим образец
                        if sample_column:
                            sample_cell = ws.cell(row=start_row, column=sample_column).value
                            if sample_cell and str(sample_cell).strip():
                                sample_price = parse_price_value(str(sample_cell))
                                
                                sample_time = None
                                if sample_time_column:
                                    sample_time_cell = ws.cell(row=start_row, column=sample_time_column).value
                                    sample_time = parse_time_value(sample_time_cell)
                                
                                if sample_price:
                                    offer = PriceOffer(
                                        product_id=product.id,
                                        route_name="Образец",
                                        is_sample=True,
                                        sample_price=sample_price,
                                        sample_time=sample_time,
                                        is_available=True
                                    )
                                    session.add(offer)
                                    sheet_samples += 1
                                    total_samples += 1
                        
                        logger.info(f"    {product_name} (строки {start_row}-{end_row})")
                    
                    session.commit()
                    logger.info(f"  Обработано товаров: {sheet_products}, образцов: {sheet_samples}")
                    
                except Exception as e:
                    session.rollback()
                    logger.error(f"  Ошибка парсинга таблицы {sheet.sheet_title}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            logger.info(f"\n=== ИТОГО ОБРАБОТАНО ===")
            logger.info(f"Товаров: {total_products}")
            logger.info(f"Образцов: {total_samples}")
            
            # Финальная проверка
            final_products = session.query(Product).count()
            final_samples = session.query(PriceOffer).filter(PriceOffer.is_sample == True).count()
            logger.info(f"В базе данных: товаров={final_products}, образцов={final_samples}")

        except Exception as e:
            session.rollback()
            logger.error(f"Общая ошибка парсинга товаров: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            session.close()

if __name__ == "__main__":
    parser = ImprovedProductParser()
    parser.parse_all_products()