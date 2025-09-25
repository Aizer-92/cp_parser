#!/usr/bin/env python3
"""
Улучшенный парсер v5 с правильным определением колонок по заголовкам и валидацией
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage, PriceOffer
import openpyxl
from pathlib import Path
import re
import logging
from typing import Dict, List, Tuple, Optional, Any
from collections import defaultdict

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ImprovedParser:
    """Улучшенный парсер с определением колонок по заголовкам"""
    
    def __init__(self):
        self.session = DatabaseManager.get_session()
        
        # Ключевые слова для определения колонок
        self.column_keywords = {
            'name': ['название', 'товар', 'продукт', 'наименование', 'item', 'product', 'name', 'title', 'goods'],
            'material': ['материал', 'material', 'состав', 'composition'],
            'size': ['размер', 'size', 'габариты', 'dimensions', 'см', 'cm'],
            'color': ['цвет', 'color', 'colours', 'фабричные цвета'],
            'packaging': ['упаковка', 'package', 'packaging', 'пакет'],
            'quantity': ['тираж', 'количество', 'qty', 'quantity', 'шт', 'pcs', 'pieces'],
            'price_usd': ['цена', 'price', 'usd', '$', 'dollar', 'стоимость'],
            'price_rub': ['руб', 'rub', 'рубль', 'ruble', 'р.'],
            'price_aed': ['aed', 'дирхам', 'dirham'],
            'delivery': ['доставка', 'delivery', 'срок', 'дн', 'days', 'день'],
            'route': ['маршрут', 'route', 'путь', 'направление']
        }
    
    def detect_columns(self, worksheet, max_header_rows=5) -> Dict[str, int]:
        """Определяет колонки по заголовкам"""
        columns = {}
        
        # Анализируем первые несколько строк для поиска заголовков
        for row in range(1, max_header_rows + 1):
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if not cell_value:
                    continue
                
                cell_text = str(cell_value).lower().strip()
                
                # Проверяем каждый тип колонки
                for col_type, keywords in self.column_keywords.items():
                    if col_type in columns:  # Уже найдена
                        continue
                    
                    for keyword in keywords:
                        if keyword in cell_text:
                            columns[col_type] = col
                            logger.info(f"Найдена колонка {col_type} в {openpyxl.utils.get_column_letter(col)}{row}: '{cell_value}'")
                            break
        
        return columns
    
    def validate_product_name(self, name: str) -> bool:
        """Валидация названия товара"""
        if not name or len(name.strip()) < 2:
            return False
        
        name_clean = name.strip().lower()
        
        # Проверяем, что это не цена или тираж
        if re.match(r'^\d+[.,]?\d*\s*[$₽руб]?$', name_clean):
            return False
        
        # Проверяем, что это не служебная информация
        service_patterns = [
            r'^(материал|размер|цвет|упаковка)',
            r'^\d+\s*[x×]\s*\d+',  # размеры
            r'^\d+\s*(см|mm|мм|cm)$',    # размеры
            r'менеджер|email|телефон|примечание',
            r'цена указана|срок тиража|производство',
            r'^(пластик|металл|дерево|картон|нейлон|полиэстер)$',
            # Добавляем заголовки таблиц
            r'^(name|наименование|товар|продукт|item|product|title|goods)$',
            r'^(quantity|тираж|количество|шт|pcs|pieces|qty)$',
            r'^(price|цена|стоимость|cost|\$|usd|руб|rub|aed)$',
            r'^(delivery|доставка|срок|дн|days|день)$',
            r'^(material|материал|состав)$',
            r'^(size|размер|габариты|dimensions)$',
            r'^(color|цвет|colours)$',
            r'^(packaging|упаковка|package|пакет)$',
            r'^(custom|кастом|дизайн|design)$',
            r'^(photo|фото|image|изображение)$'
        ]
        
        for pattern in service_patterns:
            if re.search(pattern, name_clean):
                return False
        
        return True
    
    def validate_quantity(self, value: Any) -> Optional[int]:
        """Валидация тиража"""
        if value is None:
            return None
        
        # Извлекаем числовое значение
        if isinstance(value, (int, float)):
            qty = int(value)
        else:
            # Извлекаем число из строки
            match = re.search(r'(\\d+)', str(value).replace(',', '').replace(' ', ''))
            if not match:
                return None
            qty = int(match.group(1))
        
        # Валидация диапазона
        if qty < 1 or qty > 100000:  # Разумные пределы
            logger.warning(f"Подозрительный тираж: {qty}")
            return None
        
        return qty
    
    def validate_price(self, value: Any) -> Optional[float]:
        """Валидация цены"""
        if value is None:
            return None
        
        # Извлекаем числовое значение
        if isinstance(value, (int, float)):
            price = float(value)
        else:
            # Извлекаем число из строки, убираем валютные символы
            text = str(value).replace('$', '').replace('₽', '').replace(',', '').strip()
            match = re.search(r'(\\d+[.,]?\\d*)', text)
            if not match:
                return None
            price = float(match.group(1).replace(',', '.'))
        
        # Валидация диапазона
        if price < 0.01 or price > 1000:  # Разумные пределы для цены за единицу
            logger.warning(f"Подозрительная цена: ${price}")
            return None
        
        return price
    
    def find_product_rows(self, worksheet, columns: Dict[str, int]) -> List[Tuple[int, int]]:
        """Находит строки с товарами"""
        if 'name' not in columns:
            logger.error("Не найдена колонка с названиями товаров")
            return []
        
        name_col = columns['name']
        product_rows = []
        current_start = None
        
        logger.info(f"🔍 ПОИСК ТОВАРОВ в колонке {name_col} (первые 20 строк):")
        
        for row in range(1, min(21, worksheet.max_row + 1)):  # Ограничиваем для диагностики
            cell_value = worksheet.cell(row=row, column=name_col).value
            cell_str = str(cell_value) if cell_value else "NULL"
            if len(cell_str) > 50:
                cell_str = cell_str[:47] + "..."
            
            is_valid = cell_value and self.validate_product_name(str(cell_value))
            logger.info(f"   Row {row:2d}: {cell_str} -> {'✅' if is_valid else '❌'}")
            
            if cell_value and self.validate_product_name(str(cell_value)):
                if current_start is None:
                    current_start = row
                    logger.info(f"   🟢 Начало товара: строка {row}")
            else:
                if current_start is not None:
                    # Завершаем текущий товар
                    product_rows.append((current_start, row - 1))
                    logger.info(f"   🔴 Конец товара: строки {current_start}-{row - 1}")
                    current_start = None
        
        # Обработаем остальные строки без подробного логирования
        for row in range(21, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=name_col).value
            
            if cell_value and self.validate_product_name(str(cell_value)):
                if current_start is None:
                    current_start = row
            else:
                if current_start is not None:
                    # Завершаем текущий товар
                    product_rows.append((current_start, row - 1))
                    current_start = None
        
        # Добавляем последний товар, если он есть
        if current_start is not None:
            product_rows.append((current_start, worksheet.max_row))
            logger.info(f"   🔴 Последний товар: строки {current_start}-{worksheet.max_row}")
        
        # Фильтруем слишком большие диапазоны (вероятно, ошибки)
        filtered_rows = []
        for start, end in product_rows:
            if end - start + 1 <= 50:  # Максимум 50 строк на товар
                filtered_rows.append((start, end))
            else:
                logger.warning(f"Пропускаем слишком большой диапазон строк: {start}-{end}")
        
        return filtered_rows
    
    def extract_product_data(self, worksheet, columns: Dict[str, int], start_row: int, end_row: int) -> Dict[str, Any]:
        """Извлекает данные товара"""
        data = {
            'name': None,
            'characteristics': [],
            'prices': [],
            'images': []
        }
        
        # Извлекаем название (первая непустая ячейка в колонке названий)
        if 'name' in columns:
            for row in range(start_row, end_row + 1):
                name_value = worksheet.cell(row=row, column=columns['name']).value
                if name_value and self.validate_product_name(str(name_value)):
                    data['name'] = str(name_value).strip()
                    break
        
        # Собираем характеристики из разных колонок
        char_parts = []
        for col_type in ['material', 'size', 'color', 'packaging']:
            if col_type in columns:
                col = columns[col_type]
                for row in range(start_row, end_row + 1):
                    value = worksheet.cell(row=row, column=col).value
                    if value and str(value).strip():
                        char_parts.append(f"{col_type.title()}: {str(value).strip()}")
        
        data['characteristics'] = '\\n'.join(char_parts)
        
        # Извлекаем варианты цен
        self.extract_price_variants(worksheet, columns, start_row, end_row, data)
        
        return data
    
    def extract_price_variants(self, worksheet, columns: Dict[str, int], start_row: int, end_row: int, data: Dict[str, Any]):
        """Извлекает варианты цен"""
        # Ищем строки с ценами
        for row in range(start_row, end_row + 1):
            price_data = {}
            
            # Тираж
            if 'quantity' in columns:
                qty_value = worksheet.cell(row=row, column=columns['quantity']).value
                qty = self.validate_quantity(qty_value)
                if qty:
                    price_data['quantity'] = qty
            
            # Цена в USD
            if 'price_usd' in columns:
                price_value = worksheet.cell(row=row, column=columns['price_usd']).value
                price = self.validate_price(price_value)
                if price:
                    price_data['price_usd'] = price
            
            # Маршрут доставки
            if 'route' in columns:
                route_value = worksheet.cell(row=row, column=columns['route']).value
                if route_value:
                    price_data['route'] = str(route_value).strip()
            
            # Срок доставки
            if 'delivery' in columns:
                delivery_value = worksheet.cell(row=row, column=columns['delivery']).value
                if delivery_value:
                    # Извлекаем число дней
                    match = re.search(r'(\\d+)', str(delivery_value))
                    if match:
                        days = int(match.group(1))
                        if 1 <= days <= 365:  # Разумные пределы
                            price_data['delivery_time'] = f"{days} дн"
            
            # Добавляем вариант цены, если есть основные данные
            if 'quantity' in price_data and 'price_usd' in price_data:
                data['prices'].append(price_data)
    
    def parse_sheet_improved(self, sheet_id: int) -> bool:
        """Парсит таблицу с улучшенным алгоритмом"""
        try:
            sheet = self.session.query(SheetMetadata).filter(SheetMetadata.id == sheet_id).first()
            if not sheet:
                logger.error(f"Таблица {sheet_id} не найдена")
                return False
            
            if not sheet.local_file_path:
                logger.error(f"Нет локального файла для таблицы {sheet_id}")
                return False
            
            # Исправляем путь, если он содержит дублирование
            if 'storage/excel_files' in sheet.local_file_path:
                excel_path = Path(sheet.local_file_path)
            else:
                excel_path = Path('storage/excel_files') / sheet.local_file_path
            if not excel_path.exists():
                logger.error(f"Excel файл не найден: {excel_path}")
                return False
            
            logger.info(f"Парсим таблицу {sheet_id}: {sheet.sheet_title}")
            
            # Загружаем Excel
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            # Определяем колонки по заголовкам
            columns = self.detect_columns(worksheet)
            logger.info(f"Найдены колонки: {columns}")
            
            if not columns:
                logger.warning(f"Не найдено колонок в таблице {sheet_id}")
                return False
            
            # Находим строки с товарами
            product_rows = self.find_product_rows(worksheet, columns)
            logger.info(f"Найдено товаров: {len(product_rows)}")
            
            if not product_rows:
                logger.warning(f"Не найдено товаров в таблице {sheet_id}")
                return False
            
            # Удаляем старые данные
            old_products = self.session.query(Product).filter(Product.sheet_id == sheet_id).all()
            for product in old_products:
                # Удаляем связанные данные
                self.session.query(ProductImage).filter(ProductImage.product_id == product.id).delete()
                self.session.query(PriceOffer).filter(PriceOffer.product_id == product.id).delete()
                self.session.delete(product)
            
            # Парсим каждый товар
            parsed_count = 0
            for start_row, end_row in product_rows:
                try:
                    product_data = self.extract_product_data(worksheet, columns, start_row, end_row)
                    
                    if not product_data['name']:
                        logger.warning(f"Пропускаем товар без названия в строках {start_row}-{end_row}")
                        continue
                    
                    # Создаем товар
                    product = Product(
                        sheet_id=sheet_id,
                        name=product_data['name'],
                        characteristics=product_data['characteristics'],
                        start_row=start_row,
                        end_row=end_row
                    )
                    
                    self.session.add(product)
                    self.session.flush()  # Получаем ID
                    
                    # Добавляем варианты цен
                    for price_data in product_data['prices']:
                        price_offer = PriceOffer(
                            product_id=product.id,
                            quantity=price_data['quantity'],
                            price_usd=price_data['price_usd'],
                            route_name=price_data.get('route', 'Стандарт'),
                            delivery_time=price_data.get('delivery_time')
                        )
                        self.session.add(price_offer)
                    
                    parsed_count += 1
                    logger.info(f"Создан товар: {product_data['name']} ({len(product_data['prices'])} цен)")
                
                except Exception as e:
                    logger.error(f"Ошибка парсинга товара в строках {start_row}-{end_row}: {e}")
                    continue
            
            # Сохраняем изменения
            self.session.commit()
            
            # Обновляем статистику таблицы
            sheet.products_count = parsed_count
            sheet.status = 'completed'
            self.session.commit()
            
            logger.info(f"Успешно обработана таблица {sheet_id}: {parsed_count} товаров")
            return True
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Ошибка парсинга таблицы {sheet_id}: {e}")
            return False
    
    def close(self):
        """Закрывает сессию"""
        self.session.close()

def test_improved_parser():
    """Тестирует улучшенный парсер на проблемных таблицах"""
    parser = ImprovedParser()
    
    try:
        # Тестируем на существующей таблице
        problem_sheet_id = 1  # original_sheet
        
        print(f"🧪 ТЕСТИРОВАНИЕ УЛУЧШЕННОГО ПАРСЕРА")
        print(f"=" * 60)
        print(f"Тестируем таблицу ID {problem_sheet_id}")
        
        success = parser.parse_sheet_improved(problem_sheet_id)
        
        if success:
            print(f"✅ Парсинг успешен!")
            
            # Проверяем результат
            session = DatabaseManager.get_session()
            products = session.query(Product).filter(Product.sheet_id == problem_sheet_id).all()
            print(f"📦 Создано товаров: {len(products)}")
            
            for product in products[:5]:  # Показываем первые 5
                prices_count = session.query(PriceOffer).filter(PriceOffer.product_id == product.id).count()
                print(f"  - {product.name} (строки {product.start_row}-{product.end_row}, {prices_count} цен)")
            
            session.close()
        else:
            print(f"❌ Парсинг не удался")
    
    finally:
        parser.close()

if __name__ == "__main__":
    test_improved_parser()
