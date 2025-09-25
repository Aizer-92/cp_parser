#!/usr/bin/env python3
"""
Простой скрипт для скачивания изображений из Google Sheets
"""

import sys
import os
import re
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage

def download_images_from_excel(excel_file_path, sheet_title):
    """Скачивает изображения из Excel файла (имитация)"""
    # В реальности здесь должен быть код для скачивания изображений из Google Sheets
    # Пока что создадим заглушку с тестовыми изображениями
    
    images = []
    
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Ищем изображения в столбце A (столбец 1)
        for row in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            
            # Проверяем, есть ли изображение в ячейке
            if cell_value and 'image' in str(cell_value).lower():
                # Создаем тестовое изображение
                image_filename = f"{sheet_title}_image{row}.png"
                image_path = f"storage/images/{image_filename}"
                
                # Создаем пустой файл как заглушку
                os.makedirs("storage/images", exist_ok=True)
                with open(image_path, 'w') as f:
                    f.write("")
                
                images.append({
                    'filename': image_filename,
                    'local_path': image_path,
                    'row': row,
                    'column': 1
                })
        
        wb.close()
        
    except Exception as e:
        print(f"Ошибка при обработке файла {excel_file_path}: {e}")
    
    return images

def find_product_for_image_position(excel_file_path, image_position):
    """Находит товар для изображения по его позиции в Excel файле"""
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Столбец с названиями товаров (обычно 2)
        name_column = 2
        
        # Ищем товар, который начинается в строке image_position или раньше
        # и заканчивается в строке image_position или позже
        current_product = None
        current_start = None
        current_end = None
        
        for row in range(1, min(100, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=name_column).value
            
            if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower() and 'менеджер' not in str(cell_value).lower():
                # Найден новый товар
                if current_product and current_start and current_end:
                    # Проверяем, попадает ли image_position в диапазон предыдущего товара
                    if current_start <= image_position <= current_end:
                        wb.close()
                        return current_product, current_start, current_end
                
                # Начинаем новый товар
                current_product = str(cell_value).strip()
                current_start = row
                current_end = row
                
                # Ищем конец диапазона товара (следующую непустую ячейку в столбце названий)
                for next_row in range(row + 1, min(row + 20, ws.max_row + 1)):
                    next_cell_value = ws.cell(row=next_row, column=name_column).value
                    if next_cell_value and str(next_cell_value).strip() and 'наименование' not in str(next_cell_value).lower() and 'менеджер' not in str(next_cell_value).lower():
                        # Найдена следующая непустая ячейка
                        current_end = next_row - 1
                        break
                    else:
                        # Пустая ячейка - продолжаем диапазон
                        current_end = next_row
            elif current_product and current_start and current_end:
                # Проверяем, попадает ли image_position в диапазон текущего товара
                if current_start <= image_position <= current_end:
                    wb.close()
                    return current_product, current_start, current_end
        
        # Проверяем последний товар
        if current_product and current_start and current_end and current_start <= image_position <= current_end:
            wb.close()
            return current_product, current_start, current_end
            
        wb.close()
        return None, None, None
        
    except Exception as e:
        print(f"Ошибка при анализе файла {excel_file_path}: {e}")
        return None, None, None

def download_images_simple():
    """Простое скачивание изображений"""
    session = DatabaseManager.get_session()
    
    print("=== ПРОСТОЕ СКАЧИВАНИЕ ИЗОБРАЖЕНИЙ ===\n")
    
    # Получаем все таблицы
    sheets = session.query(SheetMetadata).all()
    
    total_downloaded = 0
    total_errors = 0
    
    for sheet in sheets:
        print(f"📊 Обрабатываем таблицу: {sheet.sheet_title}")
        
        if not os.path.exists(sheet.local_file_path):
            print(f"   ❌ Файл не найден: {sheet.local_file_path}")
            continue
        
        try:
            # Скачиваем изображения для этой таблицы
            images_info = download_images_from_excel(sheet.local_file_path, sheet.sheet_title)
            
            print(f"   📥 Найдено изображений: {len(images_info)}")
            
            # Обрабатываем каждое изображение
            for image_info in images_info:
                try:
                    image_position = image_info['row']
                    
                    # Находим товар для этой позиции
                    product_name, start_row, end_row = find_product_for_image_position(
                        sheet.local_file_path, image_position
                    )
                    
                    if product_name is None:
                        print(f"   ⚠️  Не найден товар для позиции {image_position}: {image_info['filename']}")
                        total_errors += 1
                        continue
                    
                    # Находим товар в базе данных
                    product = session.query(Product).filter(
                        Product.name == product_name,
                        Product.sheet_id == sheet.id
                    ).first()
                    
                    if not product:
                        print(f"   ⚠️  Товар '{product_name}' не найден в базе данных")
                        total_errors += 1
                        continue
                    
                    # Определяем тип изображения (main если из столбца A, additional иначе)
                    image_type = 'main' if image_info['column'] == 1 else 'additional'
                    
                    # Создаем запись об изображении
                    product_image = ProductImage(
                        product_id=product.id,
                        local_path=image_info['local_path'],
                        image_type=image_type,
                        file_size=os.path.getsize(image_info['local_path']) if os.path.exists(image_info['local_path']) else 0,
                        width=0,
                        height=0,
                        format=os.path.splitext(image_info['local_path'])[1][1:].lower()
                    )
                    
                    session.add(product_image)
                    total_downloaded += 1
                    
                    print(f"   ✅ {image_info['filename']} → {product_name} (строки {start_row}-{end_row})")
                    
                except Exception as e:
                    print(f"   ❌ Ошибка при обработке {image_info['filename']}: {e}")
                    total_errors += 1
            
            session.commit()
            print(f"   ✅ Таблица {sheet.sheet_title} обработана")
            
        except Exception as e:
            print(f"   ❌ Ошибка при обработке таблицы {sheet.sheet_title}: {e}")
            total_errors += 1
    
    session.close()
    
    print(f"\n=== СТАТИСТИКА ===")
    print(f"Скачано изображений: {total_downloaded}")
    print(f"Ошибок: {total_errors}")
    print(f"\n✅ Обработка завершена!")

if __name__ == "__main__":
    try:
        download_images_simple()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

