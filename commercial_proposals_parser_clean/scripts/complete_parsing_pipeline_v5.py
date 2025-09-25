#!/usr/bin/env python3
"""
Полный пайплайн парсинга v5 с улучшенным определением колонок и валидацией
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.manager_v4 import DatabaseManager
from database.models_v4 import ProjectMetadata, SheetMetadata, Product, ProductImage, PriceOffer
import openpyxl
from pathlib import Path
import re
import logging
import time
from typing import Dict, List, Tuple, Optional, Any
from collections import defaultdict
from PIL import Image
import io

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class EnhancedParser:
    """Улучшенный парсер с правильным определением колонок и валидацией"""
    
    def __init__(self):
        self.session = DatabaseManager.get_session()
        
        # Ключевые слова для определения колонок (расширенные)
        self.column_keywords = {
            'name': ['название', 'товар', 'продукт', 'наименование', 'item', 'product', 'name', 'title', 'goods', 'merchandise'],
            'material': ['материал', 'material', 'состав', 'composition'],
            'size': ['размер', 'size', 'габариты', 'dimensions', 'см', 'cm', 'мм', 'mm'],
            'color': ['цвет', 'color', 'colours', 'фабричные цвета'],
            'packaging': ['упаковка', 'package', 'packaging', 'пакет'],
            'quantity': ['тираж, шт', 'тираж шт', 'quantity, pcs', 'quantity pcs', 'кол-во шт'],  # ИСПРАВЛЕНО: только специфичные комбинации
            'price_usd': ['цена', 'price', 'usd', '$', 'dollar', 'стоимость', 'долл'],
            'price_rub': ['руб', 'rub', 'рубль', 'ruble', 'р.', 'рублей'],
            'delivery': ['доставка', 'delivery', 'срок', 'дн', 'days', 'день', 'время'],
            'route': ['маршрут', 'route', 'путь', 'направление', 'способ доставки']
        }
        
        # Исключаемые листы
        self.excluded_sheets = ['цена', 'доставка', 'образец', 'price', 'delivery', 'sample', 'copy of']
    
    def should_skip_sheet(self, sheet_name: str) -> bool:
        """Проверяет, нужно ли пропустить лист"""
        sheet_name_lower = sheet_name.lower()
        return any(excluded in sheet_name_lower for excluded in self.excluded_sheets)
    
    def detect_columns_smart(self, worksheet, max_header_rows=8) -> Dict[str, int]:
        """Умное определение колонок по заголовкам"""
        columns = {}
        
        # Анализируем заголовки построчно
        for row in range(1, max_header_rows + 1):
            for col in range(1, min(worksheet.max_column + 1, 20)):  # Ограничиваем до 20 колонок
                cell_value = worksheet.cell(row=row, column=col).value
                if not cell_value:
                    continue
                
                cell_text = str(cell_value).lower().strip()
                
                # Проверяем каждый тип колонки (ПРИОРИТЕТНЫЙ ПОРЯДОК)
                # Сначала цены - они более специфичны
                priority_order = ['price_usd', 'price_rub', 'quantity', 'name', 'material', 'size', 'color', 'packaging', 'delivery', 'route']
                
                for col_type in priority_order:
                    if col_type in columns:  # Уже найдена
                        continue
                    
                    keywords = self.column_keywords[col_type]
                    # Проверяем точное совпадение или вхождение
                    for keyword in keywords:
                        if keyword == cell_text or keyword in cell_text:
                            # ДОПОЛНИТЕЛЬНАЯ ПРОВЕРКА: если это quantity, убедимся что это не цена
                            if col_type == 'quantity' and any(price_word in cell_text for price_word in ['цена', 'price', '$', 'руб', 'usd']):
                                continue  # Пропускаем - это цена, а не тираж
                            
                            columns[col_type] = col
                            logger.info(f"Найдена колонка {col_type} в {openpyxl.utils.get_column_letter(col)}{row}: '{cell_value}'")
                            break
        
        # Дополнительная логика для определения колонок по содержимому
        if 'name' not in columns:
            # Ищем колонку с наибольшим количеством текстовых значений
            text_counts = defaultdict(int)
            for row in range(1, min(worksheet.max_row + 1, 20)):
                for col in range(1, min(worksheet.max_column + 1, 10)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 3:
                        if not re.match(r'^\\d+[.,]?\\d*', cell_value.strip()):  # Не число
                            text_counts[col] += 1
            
            if text_counts:
                name_col = max(text_counts.items(), key=lambda x: x[1])[0]
                columns['name'] = name_col
                logger.info(f"Автоопределена колонка name в {openpyxl.utils.get_column_letter(name_col)} по содержимому")
        
        return columns
    
    def validate_product_name_enhanced(self, name: str) -> bool:
        """Расширенная валидация названия товара"""
        if not name or len(name.strip()) < 2:
            return False
        
        name_clean = name.strip().lower()
        
        # Проверяем, что это не цена или тираж
        if re.match(r'^\d+[.,]?\d*\s*[$₽руб]?$', name_clean):
            return False
        
        # Проверяем, что это не служебная информация
        service_patterns = [
            r'^(материал|размер|цвет|упаковка)',
            r'^\d+\s*[x×]\s*\d+',  # размеры
            r'^\d+\s*(см|mm|мм|cm)$',    # размеры
            r'менеджер|email|телефон|примечание|фабрика',
            r'цена указана|срок тиража|производство|доставка',
            r'^(пластик|металл|дерево|картон|нейлон|полиэстер)$',
            r'^наименование$',  # заголовок
            r'copy of|цена|доставка|образец',
            # Добавляем заголовки таблиц
            r'^(name|товар|продукт|item|product|title|goods)$',
            r'^(quantity|тираж|количество|шт|pcs|pieces|qty)$',
            r'^(price|цена|стоимость|cost|\$|usd|руб|rub|aed)$',
            r'^(delivery|доставка|срок|дн|days|день)$',
            r'^(material|материал|состав)$',
            r'^(size|размер|габариты|dimensions)$',
            r'^(color|цвет|colours)$',
            r'^(packaging|упаковка|package|пакет)$',
            r'^(custom|кастом|дизайн|design)$',
            r'^(photo|фото|image|изображение)$'
        ]
        
        for pattern in service_patterns:
            if re.search(pattern, name_clean):
                return False
        
        return True
    
    def validate_quantity_enhanced(self, value: Any) -> Optional[int]:
        """Расширенная валидация тиража"""
        if value is None:
            return None
        
        # Извлекаем числовое значение
        if isinstance(value, (int, float)):
            if value <= 0:
                return None
            qty = int(value)
        else:
            # Извлекаем первое число из строки
            text = str(value).replace(',', '').replace(' ', '')
            match = re.search(r'(\d+)', text)
            if not match:
                return None
            qty = int(match.group(1))
        
        # Валидация диапазона (расширенная)
        if qty < 1 or qty > 100000:  # Минимум 1 штука, максимум 100k
            logger.warning(f"Подозрительный тираж: {qty}")
            return None
        
        return qty
    
    def validate_price_enhanced(self, value: Any) -> Optional[float]:
        """Расширенная валидация цены"""
        if value is None:
            return None
        
        # Извлекаем числовое значение
        if isinstance(value, (int, float)):
            price = float(value)
        else:
            # Извлекаем число из строки, убираем валютные символы
            text = str(value).replace('$', '').replace('₽', '').replace(',', '').replace(' ', '').strip()
            match = re.search(r'(\d+[.,]?\d*)', text)
            if not match:
                return None
            price = float(match.group(1).replace(',', '.'))
        
        # Валидация диапазона (расширенная)
        if price < 0.01 or price > 2000:  # Увеличенные разумные пределы 
            logger.warning(f"Подозрительная цена: ${price}")
            return None
        
        return price
    
    def find_product_rows_enhanced(self, worksheet, columns: Dict[str, int]) -> List[Tuple[int, int]]:
        """Улучшенный поиск строк с товарами"""
        if 'name' not in columns:
            logger.error("Не найдена колонка с названиями товаров")
            return []
        
        name_col = columns['name']
        product_rows = []
        current_start = None
        
        logger.info(f"🔍 ПОИСК ТОВАРОВ в колонке {name_col} (первые 15 строк):")
        
        # Ищем с первой строки, но с валидацией заголовков
        for row in range(1, min(16, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row=row, column=name_col).value
            cell_str = str(cell_value) if cell_value else "NULL"
            if len(cell_str) > 30:
                cell_str = cell_str[:27] + "..."
            
            is_valid = cell_value and self.validate_product_name_enhanced(str(cell_value))
            logger.info(f"   Row {row:2d}: {cell_str} -> {'✅' if is_valid else '❌'}")
            
            if cell_value and self.validate_product_name_enhanced(str(cell_value)):
                if current_start is None:
                    current_start = row
                    logger.info(f"   🟢 Начало товара: строка {row}")
            else:
                if current_start is not None:
                    # Завершаем текущий товар
                    product_rows.append((current_start, row - 1))
                    logger.info(f"   🔴 Конец товара: строки {current_start}-{row - 1}")
                    current_start = None
        
        # Обрабатываем остальные строки без подробного лога
        for row in range(16, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=name_col).value
            
            if cell_value and self.validate_product_name_enhanced(str(cell_value)):
                if current_start is None:
                    current_start = row
            else:
                if current_start is not None:
                    # Завершаем текущий товар
                    product_rows.append((current_start, row - 1))
                    current_start = None
        
        # Добавляем последний товар, если он есть
        if current_start is not None:
            product_rows.append((current_start, worksheet.max_row))
            logger.info(f"   🔴 Последний товар: строки {current_start}-{worksheet.max_row}")
        
        # Фильтруем слишком большие диапазоны (вероятно, ошибки)
        filtered_rows = []
        for start, end in product_rows:
            row_count = end - start + 1
            if row_count <= 20:  # Максимум 20 строк на товар
                filtered_rows.append((start, end))
            else:
                logger.warning(f"Пропускаем слишком большой диапазон строк: {start}-{end} ({row_count} строк)")
        
        return filtered_rows
    
    def extract_product_data_enhanced(self, worksheet, columns: Dict[str, int], start_row: int, end_row: int) -> Dict[str, Any]:
        """Улучшенное извлечение данных товара"""
        data = {
            'name': None,
            'characteristics': [],
            'prices': []
        }
        
        # Извлекаем название (первая валидная ячейка в колонке названий)
        if 'name' in columns:
            for row in range(start_row, end_row + 1):
                name_value = worksheet.cell(row=row, column=columns['name']).value
                if name_value and self.validate_product_name_enhanced(str(name_value)):
                    data['name'] = str(name_value).strip()
                    break
        
        # Собираем характеристики из разных колонок
        char_parts = []
        for col_type in ['material', 'size', 'color', 'packaging']:
            if col_type in columns:
                col = columns[col_type]
                values = []
                for row in range(start_row, end_row + 1):
                    value = worksheet.cell(row=row, column=col).value
                    if value and str(value).strip() and str(value).strip() not in values:
                        values.append(str(value).strip())
                
                if values:
                    char_parts.append(f"{col_type.title()}: {', '.join(values)}")
        
        data['characteristics'] = '\\n'.join(char_parts)
        
        # Извлекаем варианты цен
        self.extract_price_variants_enhanced(worksheet, columns, start_row, end_row, data)
        
        return data
    
    def extract_price_variants_enhanced(self, worksheet, columns: Dict[str, int], start_row: int, end_row: int, data: Dict[str, Any]):
        """Улучшенное извлечение вариантов цен"""
        # Ищем строки с ценами
        for row in range(start_row, end_row + 1):
            price_data = {}
            
            # Тираж
            if 'quantity' in columns:
                qty_value = worksheet.cell(row=row, column=columns['quantity']).value
                qty = self.validate_quantity_enhanced(qty_value)
                if qty:
                    price_data['quantity'] = qty
            
            # Цена в USD
            if 'price_usd' in columns:
                price_value = worksheet.cell(row=row, column=columns['price_usd']).value
                price = self.validate_price_enhanced(price_value)
                if price:
                    price_data['price_usd'] = price
            
            # Цена в RUB (ИСПРАВЛЕНО - не перезаписываем USD!)
            if 'price_rub' in columns:
                price_value = worksheet.cell(row=row, column=columns['price_rub']).value
                price_rub = self.validate_price_enhanced(price_value)
                if price_rub:
                    # Сохраняем рублевую цену отдельно
                    price_data['price_rub'] = price_rub
                    # Если нет USD цены, конвертируем для совместимости с БД
                    if 'price_usd' not in price_data:
                        price_data['price_usd'] = round(price_rub / 95, 2)
                    logger.info(f"Найдена RUB цена: {price_rub} ₽ (USD: ${price_data['price_usd']})")
            
            # Цена в AED (трактуется как USD - в таблицах знак $ используется для AED)
            if 'price_aed' in columns:
                price_value = worksheet.cell(row=row, column=columns['price_aed']).value  
                price = self.validate_price_enhanced(price_value)
                if price:
                    # Сохраняем AED как USD (поскольку в таблицах используется знак $)
                    price_data['price_usd'] = price
                    logger.info(f"Найдена цена в AED: ${price} (сохранена как USD)")
            
            # Маршрут доставки
            route_name = 'Стандарт'
            if 'route' in columns:
                route_value = worksheet.cell(row=row, column=columns['route']).value
                if route_value and str(route_value).strip():
                    route_name = str(route_value).strip()
            price_data['route'] = route_name
            
            # Срок доставки
            if 'delivery' in columns:
                delivery_value = worksheet.cell(row=row, column=columns['delivery']).value
                if delivery_value:
                    # Извлекаем число дней
                    match = re.search(r'(\\d+)', str(delivery_value))
                    if match:
                        days = int(match.group(1))
                        if 1 <= days <= 365:  # Разумные пределы
                            price_data['delivery_time'] = f"{days} дн"
            
            # Добавляем вариант цены, если есть основные данные
            if 'quantity' in price_data and 'price_usd' in price_data:
                data['prices'].append(price_data)
    
    def extract_images_from_excel(self, excel_path: Path, sheet_id: int) -> List[Dict]:
        """Извлекает изображения из Excel файла"""
        images_data = []
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            if not hasattr(worksheet, '_images') or not worksheet._images:
                return images_data
            
            storage_path = Path('storage/images')
            storage_path.mkdir(exist_ok=True)
            
            # Группируем изображения по позициям
            images_by_position = defaultdict(list)
            for img in worksheet._images:
                row = img.anchor._from.row + 1
                col = img.anchor._from.col + 1
                col_letter = openpyxl.utils.get_column_letter(col)
                position = f'{col_letter}{row}'
                images_by_position[position].append(img)
            
            # Сохраняем изображения
            for position, imgs in images_by_position.items():
                row = int(''.join(filter(str.isdigit, position)))
                col_letter = position.replace(str(row), '')
                
                for i, img in enumerate(imgs):
                    try:
                        # Извлекаем данные изображения
                        img_data = img._data()
                        
                        # Создаем уникальное имя файла
                        timestamp = int(time.time())
                        filename = f"{excel_path.stem}_{position}"
                        if i > 0:
                            filename += f"_{i}"
                        filename += ".png"
                        
                        file_path = storage_path / filename
                        
                        # Сохраняем изображение
                        with open(file_path, 'wb') as f:
                            f.write(img_data)
                        
                        # Определяем тип изображения
                        image_type = 'main' if col_letter == 'A' else 'additional'
                        
                        images_data.append({
                            'local_path': str(file_path),
                            'row': row,
                            'column': col_letter,
                            'image_type': image_type,
                            'sheet_id': sheet_id
                        })
                        
                    except Exception as e:
                        logger.error(f"Ошибка сохранения изображения {position}: {e}")
            
            logger.info(f"Извлечено изображений: {len(images_data)}")
            
        except Exception as e:
            logger.error(f"Ошибка извлечения изображений: {e}")
        
        return images_data
    
    def parse_sheet_complete(self, sheet_id: int) -> bool:
        """Полный парсинг таблицы с изображениями"""
        try:
            sheet = self.session.query(SheetMetadata).filter(SheetMetadata.id == sheet_id).first()
            if not sheet:
                logger.error(f"Таблица {sheet_id} не найдена")
                return False
            
            if not sheet.local_file_path:
                logger.error(f"Нет локального файла для таблицы {sheet_id}")
                return False
            
            # Исправляем путь, если он содержит дублирование
            if 'storage/excel_files' in sheet.local_file_path:
                excel_path = Path(sheet.local_file_path)
            else:
                excel_path = Path('storage/excel_files') / sheet.local_file_path
                
            if not excel_path.exists():
                logger.error(f"Excel файл не найден: {excel_path}")
                return False
            
            # Проверяем, нужно ли пропустить лист
            if self.should_skip_sheet(sheet.sheet_title):
                logger.info(f"Пропускаем служебный лист: {sheet.sheet_title}")
                return True
            
            logger.info(f"Парсим таблицу {sheet_id}: {sheet.sheet_title}")
            
            # Загружаем Excel
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            # Определяем колонки по заголовкам
            columns = self.detect_columns_smart(worksheet)
            logger.info(f"Найдены колонки: {list(columns.keys())}")
            
            if not columns:
                logger.warning(f"Не найдено колонок в таблице {sheet_id}")
                return False
            
            # Находим строки с товарами
            product_rows = self.find_product_rows_enhanced(worksheet, columns)
            logger.info(f"Найдено товаров: {len(product_rows)}")
            
            if not product_rows:
                logger.warning(f"Не найдено товаров в таблице {sheet_id}")
                return False
            
            # Извлекаем изображения
            images_data = self.extract_images_from_excel(excel_path, sheet_id)
            
            # Удаляем старые данные
            old_products = self.session.query(Product).filter(Product.sheet_id == sheet_id).all()
            for product in old_products:
                # Удаляем связанные данные
                self.session.query(ProductImage).filter(ProductImage.product_id == product.id).delete()
                self.session.query(PriceOffer).filter(PriceOffer.product_id == product.id).delete()
                self.session.delete(product)
            
            # Парсим каждый товар
            parsed_count = 0
            for start_row, end_row in product_rows:
                try:
                    product_data = self.extract_product_data_enhanced(worksheet, columns, start_row, end_row)
                    
                    if not product_data['name']:
                        logger.warning(f"Пропускаем товар без названия в строках {start_row}-{end_row}")
                        continue
                    
                    # Создаем товар
                    product = Product(
                        sheet_id=sheet_id,
                        name=product_data['name'],
                        characteristics=product_data['characteristics'],
                        start_row=start_row,
                        end_row=end_row
                    )
                    
                    self.session.add(product)
                    self.session.flush()  # Получаем ID
                    
                    # Добавляем варианты цен
                    for price_data in product_data['prices']:
                        price_offer = PriceOffer(
                            product_id=product.id,
                            quantity=price_data['quantity'],
                            price_usd=price_data.get('price_usd'),
                            price_rub=price_data.get('price_rub'),
                            route_name=price_data['route'],
                            delivery_time=price_data.get('delivery_time')
                        )
                        self.session.add(price_offer)
                    
                    # Привязываем изображения к товару
                    product_images = [img for img in images_data if start_row <= img['row'] <= end_row]
                    for img_data in product_images:
                        product_image = ProductImage(
                            product_id=product.id,
                            sheet_id=sheet_id,
                            local_path=img_data['local_path'],
                            row=img_data['row'],
                            column=img_data['column'],
                            image_type=img_data['image_type']
                        )
                        self.session.add(product_image)
                    
                    parsed_count += 1
                    logger.info(f"Создан товар: {product_data['name']} ({len(product_data['prices'])} цен, {len(product_images)} изображений)")
                
                except Exception as e:
                    logger.error(f"Ошибка парсинга товара в строках {start_row}-{end_row}: {e}")
                    continue
            
            # Сохраняем изменения
            self.session.commit()
            
            # Обновляем статистику таблицы
            sheet.products_count = parsed_count
            sheet.status = 'completed'
            self.session.commit()
            
            logger.info(f"Успешно обработана таблица {sheet_id}: {parsed_count} товаров")
            return True
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Ошибка парсинга таблицы {sheet_id}: {e}")
            return False
    
    def close(self):
        """Закрывает сессию"""
        self.session.close()

def parse_next_sheets_batch(limit: int = 10):
    """Парсит следующую партию таблиц"""
    parser = EnhancedParser()
    
    try:
        # Получаем необработанные таблицы
        session = DatabaseManager.get_session()
        unprocessed_sheets = session.query(SheetMetadata).filter(
            SheetMetadata.status.in_(['pending', 'error'])
        ).limit(limit).all()
        
        print(f"🚀 ПАРСИНГ СЛЕДУЮЩИХ {len(unprocessed_sheets)} ТАБЛИЦ")
        print("=" * 60)
        
        success_count = 0
        error_count = 0
        
        for i, sheet in enumerate(unprocessed_sheets, 1):
            print(f"\\n📊 [{i}/{len(unprocessed_sheets)}] Обрабатываем: {sheet.sheet_title}")
            
            try:
                success = parser.parse_sheet_complete(sheet.id)
                if success:
                    success_count += 1
                    print(f"✅ Успешно обработана")
                else:
                    error_count += 1
                    print(f"❌ Ошибка обработки")
                    
            except Exception as e:
                error_count += 1
                print(f"❌ Исключение: {e}")
        
        print("\\n" + "=" * 60)
        print("📊 ИТОГИ ПАРСИНГА")
        print("=" * 60)
        print(f"✅ Успешно: {success_count}")
        print(f"❌ Ошибок: {error_count}")
        
        session.close()
        
    finally:
        parser.close()

if __name__ == "__main__":
    parse_next_sheets_batch(5)  # Парсим 5 таблиц для тестирования
