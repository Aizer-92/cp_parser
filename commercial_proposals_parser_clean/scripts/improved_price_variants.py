#!/usr/bin/env python3
"""
Улучшенный парсер вариантов цен с лучшей обработкой тиражей
"""

import sys
import logging
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
import re

# Добавляем путь к проекту
sys.path.append(str(Path(__file__).parent.parent))

from database.manager_v4 import DatabaseManager
from database.models_v4 import Product, SheetMetadata, PriceOffer

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s:%(message)s')
logger = logging.getLogger(__name__)

def parse_numeric_value(cell_value):
    """Улучшенный парсинг числового значения из ячейки."""
    if isinstance(cell_value, (int, float, Decimal)):
        return float(cell_value)
    
    if cell_value is None:
        return None
    
    # Очищаем строку от пробелов и заменяем запятые на точки
    s_value = str(cell_value).strip().replace(' ', '').replace(',', '.')
    
    # Удаляем все символы кроме цифр, точек и дефисов
    s_value = re.sub(r'[^\d\.\-]', '', s_value)
    
    try:
        # Пробуем разные паттерны
        patterns = [
            r'(\d+(?:\.\d+)?)',  # Простое число
            r'(\d+(?:\.\d+)?)\s*шт',  # Число с "шт"
            r'(\d+(?:\.\d+)?)\s*штук',  # Число с "штук"
            r'(\d+(?:\.\d+)?)\s*шт\.',  # Число с "шт."
        ]
        
        for pattern in patterns:
            match = re.search(pattern, s_value, re.IGNORECASE)
            if match:
                return float(match.group(1))
        
        # Если ничего не найдено, пробуем простое преобразование
        return float(s_value)
    except (ValueError, TypeError):
        return None

def parse_string_value(cell_value):
    """Парсит строковое значение из ячейки."""
    if cell_value is None:
        return None
    return str(cell_value).strip()

class ImprovedPriceVariantParser:
    def __init__(self):
        self.db_manager = DatabaseManager

    def parse_all_price_variants(self):
        """Улучшенный парсинг вариантов цен из всех Excel файлов."""
        logger.info("=== УЛУЧШЕННЫЙ ПАРСИНГ ВАРИАНТОВ ЦЕН ===")
        
        session = self.db_manager.get_session()
        
        try:
            # Очищаем старые данные по вариантам цен (не образцам)
            deleted_count = session.query(PriceOffer).filter(PriceOffer.is_sample == False).delete()
            session.commit()
            logger.info(f"Очистка старых данных по вариантам цен... Удалено записей: {deleted_count}")

            sheets = session.query(SheetMetadata).all()
            logger.info(f"Найдено таблиц: {len(sheets)}")
            
            total_price_variants = 0

            for sheet in sheets:
                if not sheet.local_file_path or not Path(sheet.local_file_path).exists():
                    logger.warning(f"Файл не найден для таблицы {sheet.sheet_title}: {sheet.local_file_path}. Пропускаем.")
                    continue
                
                logger.info(f"\n=== ОБРАБОТКА ТАБЛИЦЫ: {sheet.sheet_title} ===")
                
                # Получаем товары с определенными диапазонами строк
                products = session.query(Product).filter(
                    Product.sheet_id == sheet.id,
                    Product.start_row.isnot(None),
                    Product.end_row.isnot(None)
                ).order_by(Product.start_row).all()
                logger.info(f"  Товаров в таблице: {len(products)}")

                if not products:
                    logger.info(f"  В таблице '{sheet.sheet_title}' нет товаров с определенными диапазонами строк. Пропускаем.")
                    continue

                try:
                    wb = openpyxl.load_workbook(sheet.local_file_path, data_only=True)
                    ws = wb.active
                    
                    # Определяем заголовки маршрутов и их начальные столбцы
                    route_headers = []
                    for row_idx in range(1, 4):
                        for col in range(1, min(ws.max_column + 1, 25)):
                            cell_value = ws.cell(row=row_idx, column=col).value
                            if cell_value:
                                lower_value = str(cell_value).lower()
                                if 'доставка жд' in lower_value or 'жд' == lower_value:
                                    route_headers.append({'column': col, 'name': 'ЖД', 'header_row': row_idx})
                                elif 'доставка авиа' in lower_value or 'авиа' == lower_value:
                                    route_headers.append({'column': col, 'name': 'АВИА', 'header_row': row_idx})
                                elif 'доставка контейнером' in lower_value or 'контейнер' == lower_value:
                                    route_headers.append({'column': col, 'name': 'Контейнер', 'header_row': row_idx})
                    
                    if not route_headers:
                        logger.info("  Маршруты не найдены, пропускаем таблицу")
                        continue
                    
                    logger.info(f"  Найдены маршруты: {route_headers}")

                    sheet_price_variants = 0
                    
                    for product in products:
                        logger.info(f"    Товар: {product.name} (строки {product.start_row}-{product.end_row})")
                        
                        # Ищем тиражи в диапазоне строк товара
                        quantities = []
                        for row_idx in range(product.start_row, product.end_row + 1):
                            # Проверяем несколько столбцов для тиража
                            for col in [5, 6, 7, 8]:  # Столбцы 5-8
                                quantity_cell_value = ws.cell(row=row_idx, column=col).value
                                quantity = parse_numeric_value(quantity_cell_value)
                                if quantity and quantity > 0:
                                    quantities.append({'row': row_idx, 'value': quantity, 'column': col})
                                    break  # Найден тираж в этой строке
                        
                        if not quantities:
                            logger.warning(f"      Тираж не найден для товара {product.name} в строках {product.start_row}-{product.end_row}")
                            # Показываем содержимое ячеек для отладки
                            for row_idx in range(product.start_row, product.end_row + 1):
                                for col in [5, 6, 7, 8]:
                                    cell_value = ws.cell(row=row_idx, column=col).value
                                    if cell_value is not None:
                                        logger.info(f"        Строка {row_idx}, столбец {col}: '{cell_value}' (тип: {type(cell_value)})")
                            continue

                        for q_info in quantities:
                            current_row = q_info['row']
                            quantity = q_info['value']

                            for route in route_headers:
                                # Определяем столбцы для цены в $, цены в рублях и срока
                                price_usd_col = route['column']
                                price_rub_col = route['column'] + 1
                                delivery_time_col = route['column'] + 2
                                
                                # Проверяем, что эти столбцы существуют в Excel
                                if price_usd_col > ws.max_column or price_rub_col > ws.max_column or delivery_time_col > ws.max_column:
                                    logger.warning(f"      Пропускаем маршрут {route['name']} для тиража {quantity} в строке {current_row}: не все столбцы найдены.")
                                    continue

                                price_usd = parse_numeric_value(ws.cell(row=current_row, column=price_usd_col).value)
                                price_rub = parse_numeric_value(ws.cell(row=current_row, column=price_rub_col).value)
                                delivery_time = parse_string_value(ws.cell(row=current_row, column=delivery_time_col).value)
                                
                                if price_usd or price_rub:
                                    offer = PriceOffer(
                                        product_id=product.id,
                                        route_name=route['name'],
                                        quantity=quantity,
                                        price_usd=price_usd,
                                        price_rub=price_rub,
                                        delivery_time=delivery_time,
                                        is_available=True,
                                        is_sample=False
                                    )
                                    session.add(offer)
                                    sheet_price_variants += 1
                                    total_price_variants += 1
                                    logger.info(f"      {route['name']}: {quantity} шт, ${price_usd}, {price_rub} руб, {delivery_time}")
                    
                    session.commit()
                    logger.info(f"  Обработано вариантов цен: {sheet_price_variants}")
                    
                except Exception as e:
                    session.rollback()
                    logger.error(f"  Ошибка парсинга вариантов цен для таблицы {sheet.sheet_title}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            logger.info(f"\n=== ИТОГО ОБРАБОТАНО ВАРИАНТОВ ЦЕН: {total_price_variants} ===")

            # Финальная проверка
            logger.info(f"\n=== ПРОВЕРКА РЕЗУЛЬТАТА ===")
            total_offers_in_db = session.query(PriceOffer).filter(PriceOffer.is_sample == False).count()
            logger.info(f"Всего вариантов цен в базе: {total_offers_in_db}")

            logger.info(f"\n=== ПРИМЕРЫ ВАРИАНТОВ ЦЕН ===")
            price_offers_in_db = session.query(PriceOffer).filter(PriceOffer.is_sample == False).limit(10).all()
            for i, offer in enumerate(price_offers_in_db):
                product = session.query(Product).filter(Product.id == offer.product_id).first()
                sheet = session.query(SheetMetadata).filter(SheetMetadata.id == product.sheet_id).first()
                sheet_name = sheet.sheet_title if sheet else f"ID:{product.sheet_id}"
                logger.info(f"{i+1}. {product.name} ({sheet_name}): {offer.route_name}, {offer.quantity} шт, ${offer.price_usd}, {offer.price_rub} руб")

        except Exception as e:
            session.rollback()
            logger.error(f"Общая ошибка парсинга вариантов цен: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            session.close()

if __name__ == "__main__":
    parser = ImprovedPriceVariantParser()
    parser.parse_all_price_variants()
