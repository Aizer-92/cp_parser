#!/usr/bin/env python3
"""
Скрипт для повторной обработки изображений с правильным определением позиции
"""

import sys
import os
import re
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage

def extract_image_position_from_filename(filename):
    """Извлекает позицию изображения из имени файла"""
    # Ищем паттерн _image{number} в имени файла
    match = re.search(r'_image(\d+)', filename)
    if match:
        return int(match.group(1))
    return None

def find_product_for_image_position(excel_file_path, image_position):
    """Находит товар для изображения по его позиции в Excel файле"""
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Столбец с названиями товаров (обычно 2)
        name_column = 2
        
        # Ищем товар, который начинается в строке image_position или раньше
        # и заканчивается в строке image_position или позже
        current_product = None
        current_start = None
        current_end = None
        
        for row in range(1, min(100, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=name_column).value
            
            if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower() and 'менеджер' not in str(cell_value).lower():
                # Найден новый товар
                if current_product and current_start and current_end:
                    # Проверяем, попадает ли image_position в диапазон предыдущего товара
                    if current_start <= image_position <= current_end:
                        wb.close()
                        return current_product, current_start, current_end
                
                # Начинаем новый товар
                current_product = str(cell_value).strip()
                current_start = row
                current_end = row
                
                # Ищем конец диапазона товара (следующую непустую ячейку в столбце названий)
                for next_row in range(row + 1, min(row + 20, ws.max_row + 1)):
                    next_cell_value = ws.cell(row=next_row, column=name_column).value
                    if next_cell_value and str(next_cell_value).strip() and 'наименование' not in str(next_cell_value).lower() and 'менеджер' not in str(next_cell_value).lower():
                        # Найдена следующая непустая ячейка
                        current_end = next_row - 1
                        break
                    else:
                        # Пустая ячейка - продолжаем диапазон
                        current_end = next_row
            elif current_product and current_start and current_end:
                # Проверяем, попадает ли image_position в диапазон текущего товара
                if current_start <= image_position <= current_end:
                    wb.close()
                    return current_product, current_start, current_end
        
        # Проверяем последний товар
        if current_product and current_start and current_end and current_start <= image_position <= current_end:
            wb.close()
            return current_product, current_start, current_end
            
        wb.close()
        return None, None, None
        
    except Exception as e:
        print(f"Ошибка при анализе файла {excel_file_path}: {e}")
        return None, None, None

def find_existing_images():
    """Находит существующие изображения в папке storage/images"""
    images_dir = Path("storage/images")
    if not images_dir.exists():
        return []
    
    images = []
    for file_path in images_dir.rglob("*.png"):
        images.append(str(file_path))
    for file_path in images_dir.rglob("*.jpg"):
        images.append(str(file_path))
    for file_path in images_dir.rglob("*.jpeg"):
        images.append(str(file_path))
    
    return images

def reprocess_images_with_correct_positioning():
    """Повторно обрабатывает изображения с правильным определением позиции"""
    session = DatabaseManager.get_session()
    
    print("=== ПОВТОРНАЯ ОБРАБОТКА ИЗОБРАЖЕНИЙ С ПРАВИЛЬНОЙ ПОЗИЦИЕЙ ===\n")
    
    # Находим существующие изображения
    existing_images = find_existing_images()
    print(f"Найдено изображений: {len(existing_images)}")
    
    # Получаем все таблицы
    sheets = session.query(SheetMetadata).all()
    
    total_processed = 0
    total_errors = 0
    
    for image_path in existing_images:
        try:
            # Извлекаем позицию изображения из имени файла
            image_position = extract_image_position_from_filename(image_path)
            
            if image_position is None:
                print(f"⚠️  Не удалось определить позицию для: {image_path}")
                total_errors += 1
                continue
            
            # Определяем, к какой таблице относится изображение по имени файла
            image_filename = os.path.basename(image_path)
            sheet = None
            
            for s in sheets:
                if s.sheet_title in image_filename:
                    sheet = s
                    break
            
            if not sheet:
                print(f"⚠️  Не найдена таблица для изображения: {image_path}")
                total_errors += 1
                continue
            
            # Находим товар для этой позиции
            product_name, start_row, end_row = find_product_for_image_position(
                sheet.local_file_path, image_position
            )
            
            if product_name is None:
                print(f"⚠️  Не найден товар для позиции {image_position}: {image_path}")
                total_errors += 1
                continue
            
            # Находим товар в базе данных
            product = session.query(Product).filter(
                Product.name == product_name,
                Product.sheet_id == sheet.id
            ).first()
            
            if not product:
                print(f"⚠️  Товар '{product_name}' не найден в базе данных")
                total_errors += 1
                continue
            
            # Определяем тип изображения (main если из столбца A, additional иначе)
            # Для простоты считаем, что все изображения из столбца A - main
            image_type = 'main'  # В реальности нужно определять по столбцу
            
            # Создаем запись об изображении
            product_image = ProductImage(
                product_id=product.id,
                local_path=image_path,
                image_type=image_type,
                file_size=os.path.getsize(image_path) if os.path.exists(image_path) else 0,
                width=0,  # Можно добавить определение размеров
                height=0,
                format=os.path.splitext(image_path)[1][1:].lower()
            )
            
            session.add(product_image)
            total_processed += 1
            
            print(f"✅ {os.path.basename(image_path)} → {product_name} (строки {start_row}-{end_row})")
            
        except Exception as e:
            print(f"❌ Ошибка при обработке {image_path}: {e}")
            total_errors += 1
    
    session.commit()
    session.close()
    
    print(f"\n=== СТАТИСТИКА ===")
    print(f"Обработано изображений: {total_processed}")
    print(f"Ошибок: {total_errors}")
    print(f"\n✅ Обработка завершена!")

if __name__ == "__main__":
    try:
        reprocess_images_with_correct_positioning()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

