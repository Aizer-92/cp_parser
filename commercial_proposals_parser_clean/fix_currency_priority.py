#!/usr/bin/env python3
"""
Исправление приоритета выбора ценовых колонок - основная цена VS цена образца
"""

from database.manager_v4 import DatabaseManager
from database.models_v4 import Product, PriceOffer, SheetMetadata
from adaptive_table_analyzer import TableStructureAnalyzer
import openpyxl
import re

class FixedCurrencyParser:
    """Исправленный парсер с правильным приоритетом валютных колонок"""
    
    def __init__(self):
        self.analyzer = TableStructureAnalyzer()
    
    def fix_currency_priorities(self):
        """Исправляет приоритеты валютных колонок"""
        
        session = DatabaseManager.get_session()
        
        try:
            print("🔧 ИСПРАВЛЕНИЕ ПРИОРИТЕТА ВАЛЮТНЫХ КОЛОНОК")
            print("=" * 60)
            
            # Получаем все товары с подозрительными ценами (курс > 200)
            suspicious_products = []
            products = session.query(Product).all()
            
            for product in products:
                prices = session.query(PriceOffer).filter(PriceOffer.product_id == product.id).all()
                for price in prices:
                    if price.price_usd and price.price_rub:
                        rate = price.price_rub / price.price_usd
                        if rate > 200:  # Подозрительно высокий курс
                            suspicious_products.append((product, price, rate))
            
            print(f"Найдено товаров с подозрительными курсами: {len(suspicious_products)}")
            
            fixed_count = 0
            
            for product, price_offer, rate in suspicious_products:
                print(f"\n🔍 Товар: {product.name}")
                print(f"   Текущий курс: {rate:.2f} ₽/$")
                
                # Получаем информацию о файле
                sheet = session.query(SheetMetadata).get(product.sheet_id)
                if not sheet or not sheet.local_file_path:
                    continue
                
                # Ищем правильную цену в Excel
                correct_rub = self._find_correct_rub_price(sheet.local_file_path, product.start_row)
                
                if correct_rub and correct_rub != price_offer.price_rub:
                    print(f"   Исправляем: {price_offer.price_rub} ₽ → {correct_rub} ₽")
                    
                    # Обновляем цену
                    price_offer.price_rub = correct_rub
                    session.add(price_offer)
                    
                    new_rate = correct_rub / price_offer.price_usd
                    print(f"   Новый курс: {new_rate:.2f} ₽/$")
                    
                    fixed_count += 1
            
            # Сохраняем изменения
            session.commit()
            
            print(f"\n📊 РЕЗУЛЬТАТЫ ИСПРАВЛЕНИЯ:")
            print(f"✅ Исправлено цен: {fixed_count}")
            
            return fixed_count
            
        except Exception as e:
            session.rollback()
            print(f"❌ Ошибка: {e}")
            return 0
            
        finally:
            session.close()
    
    def _find_correct_rub_price(self, file_path, product_row):
        """Находит правильную рублевую цену (НЕ образец)"""
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Ищем все рублевые колонки
            rub_columns = []
            
            for col in range(1, min(16, ws.max_column + 1)):
                # Анализируем заголовки
                headers = []
                for row in range(1, 4):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        headers.append(str(cell_value).lower())
                
                combined_header = ' '.join(headers)
                
                # Проверяем является ли колонка рублевой
                if any(word in combined_header for word in ['руб', '₽', 'rub']) and 'цена' in combined_header:
                    
                    # Определяем приоритет
                    priority = self._get_rub_column_priority(combined_header)
                    
                    # Получаем значение
                    cell_value = ws.cell(row=product_row, column=col).value
                    if cell_value and isinstance(cell_value, (int, float)):
                        rub_columns.append({
                            'column': col,
                            'value': float(cell_value),
                            'header': combined_header,
                            'priority': priority
                        })
            
            wb.close()
            
            if not rub_columns:
                return None
            
            # Сортируем по приоритету (больший приоритет = лучше)
            rub_columns.sort(key=lambda x: x['priority'], reverse=True)
            
            best_column = rub_columns[0]
            print(f"   Найдено {len(rub_columns)} рублевых колонок, выбираем: {best_column['header']} = {best_column['value']}")
            
            return best_column['value']
            
        except Exception as e:
            print(f"   ❌ Ошибка чтения файла: {e}")
            return None
    
    def _get_rub_column_priority(self, header_text):
        """Определяет приоритет рублевой колонки"""
        
        # ВЫСОКИЙ приоритет - основная цена
        if any(phrase in header_text for phrase in [
            'цена за шт., руб',
            'цена за шт руб', 
            'стоимость руб',
            'price rub'
        ]) and 'образец' not in header_text:
            return 100
        
        # СРЕДНИЙ приоритет - другие основные цены
        if any(phrase in header_text for phrase in [
            'цена руб',
            'стоимость',
            'price'
        ]) and 'образец' not in header_text:
            return 50
        
        # НИЗКИЙ приоритет - образцы, опционы
        if any(phrase in header_text for phrase in [
            'образец',
            'sample',
            'опция',
            'доп'
        ]):
            return 10
        
        # По умолчанию
        return 30

def test_currency_fix():
    """Тестируем исправление валют на конкретных товарах"""
    
    parser = FixedCurrencyParser()
    
    # Сначала проверяем текущее состояние товаров 104, 114
    session = DatabaseManager.get_session()
    
    print("🔍 СОСТОЯНИЕ ТОВАРОВ ДО ИСПРАВЛЕНИЯ:")
    for product_id in [104, 114]:
        product = session.query(Product).get(product_id)
        if product:
            prices = session.query(PriceOffer).filter(PriceOffer.product_id == product_id).all()
            for price in prices:
                rate = price.price_rub / price.price_usd if price.price_usd else 0
                print(f"ID {product_id}: ${price.price_usd} / {price.price_rub} ₽ = {rate:.2f}")
    
    session.close()
    
    # Исправляем
    fixed = parser.fix_currency_priorities()
    
    # Проверяем после исправления
    session = DatabaseManager.get_session()
    
    print(f"\n✅ СОСТОЯНИЕ ТОВАРОВ ПОСЛЕ ИСПРАВЛЕНИЯ:")
    for product_id in [104, 114]:
        product = session.query(Product).get(product_id)
        if product:
            prices = session.query(PriceOffer).filter(PriceOffer.product_id == product_id).all()
            for price in prices:
                rate = price.price_rub / price.price_usd if price.price_usd else 0
                print(f"ID {product_id}: ${price.price_usd} / {price.price_rub} ₽ = {rate:.2f}")
    
    session.close()

if __name__ == "__main__":
    test_currency_fix()

