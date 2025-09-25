#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Массовая нормализация ВСЕХ таблиц Excel в проекте
"""

import os
import sys
import logging
import glob
from pathlib import Path

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from table_normalizer import TableNormalizer

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mass_normalize.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

def normalize_all_tables():
    """Нормализует ВСЕ Excel таблицы в проекте"""
    
    # Находим все Excel файлы 
    excel_patterns = [
        'storage/excel_files/*.xlsx',
        'storage/excel_files/*.xls'
    ]
    
    all_files = []
    for pattern in excel_patterns:
        files = glob.glob(pattern)
        all_files.extend(files)
    
    # Исключаем уже нормализованные файлы
    original_files = [f for f in all_files if '_normalized' not in f]
    
    print("🏭 МАССОВАЯ НОРМАЛИЗАЦИЯ ВСЕХ ТАБЛИЦ")
    print("=" * 90)
    print(f"📊 Найдено файлов для нормализации: {len(original_files)}")
    print()
    
    normalizer = TableNormalizer()
    
    results = {
        'success': [],
        'failed': [],
        'skipped': []
    }
    
    # Проверяем каждый файл
    for i, file_path in enumerate(original_files, 1):
        
        print(f"📄 [{i:3}/{len(original_files)}] {os.path.basename(file_path)}")
        print("-" * 60)
        
        # Проверяем есть ли уже нормализованная версия
        base_name = os.path.splitext(file_path)[0]
        normalized_path = f"{base_name}_normalized.xlsx"
        
        if os.path.exists(normalized_path):
            print("⚠️  Уже нормализован - пропускаем")
            results['skipped'].append(file_path)
            print()
            continue
        
        try:
            # Нормализуем файл
            result_path = normalizer.normalize_table(file_path)
            
            if result_path:
                print(f"✅ Успешно → {os.path.basename(result_path)}")
                results['success'].append(file_path)
                
                # Быстрая проверка результата
                check_result = quick_check_normalized(result_path)
                print(f"   📊 Колонок: {check_result['columns']}, Строк: {check_result['rows']}")
                if check_result['routes']:
                    print(f"   🚚 Маршруты: {', '.join(check_result['routes'])}")
                
            else:
                print("❌ Ошибка нормализации")
                results['failed'].append(file_path)
                
        except Exception as e:
            logger.error(f"Критическая ошибка при нормализации {file_path}: {e}")
            print(f"❌ Критическая ошибка: {e}")
            results['failed'].append(file_path)
        
        print()
    
    # Итоговый отчет
    print("=" * 90)
    print("📋 ИТОГОВЫЙ ОТЧЕТ МАССОВОЙ НОРМАЛИЗАЦИИ")
    print("=" * 90)
    
    total = len(original_files)
    success = len(results['success'])
    failed = len(results['failed'])
    skipped = len(results['skipped'])
    
    print(f"📊 ОБЩАЯ СТАТИСТИКА:")
    print(f"   • Всего файлов: {total}")
    print(f"   • ✅ Успешно: {success}")
    print(f"   • ⚠️  Пропущено: {skipped}")  
    print(f"   • ❌ Ошибки: {failed}")
    print(f"   • 📈 Успешность: {success/total*100:.1f}%")
    
    if results['failed']:
        print(f"\n❌ ФАЙЛЫ С ОШИБКАМИ:")
        for file_path in results['failed']:
            print(f"   • {os.path.basename(file_path)}")
    
    if success > 0:
        print(f"\n✅ УСПЕШНО НОРМАЛИЗОВАНЫ:")
        for file_path in results['success']:
            base_name = os.path.basename(file_path)
            print(f"   • {base_name} → {base_name.replace('.xlsx', '_normalized.xlsx')}")
    
    print(f"\n🎯 ИТОГ: {success} таблиц готовы к идеальному парсингу!")
    
    return results

def quick_check_normalized(normalized_path):
    """Быстрая проверка результата нормализации"""
    
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(normalized_path, data_only=True)
        ws = wb.active
        
        # Считаем колонки
        columns = ws.max_column
        
        # Считаем строки с данными
        rows = 0
        for row in range(3, ws.max_row + 1):
            has_data = False
            for col in range(1, min(6, columns + 1)):
                if ws.cell(row=row, column=col).value:
                    has_data = True
                    break
            if has_data:
                rows += 1
        
        # Определяем маршруты из заголовков
        routes = set()
        for col in range(1, columns + 1):
            header = ws.cell(row=2, column=col).value
            if header:
                header_str = str(header)
                if 'АВИА' in header_str:
                    routes.add('АВИА')
                elif 'ЖД' in header_str:
                    routes.add('ЖД')
                elif 'ОБРАЗЕЦ' in header_str:
                    routes.add('ОБРАЗЕЦ')
        
        wb.close()
        
        return {
            'columns': columns,
            'rows': rows,
            'routes': list(routes)
        }
        
    except Exception as e:
        logger.error(f"Ошибка проверки {normalized_path}: {e}")
        return {'columns': 0, 'rows': 0, 'routes': []}

if __name__ == "__main__":
    normalize_all_tables()
