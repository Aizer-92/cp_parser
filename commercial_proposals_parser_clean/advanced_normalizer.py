#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Продвинутый нормализатор таблиц с поддержкой множественных тиражей
"""

import os
import sys
import logging
import openpyxl
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from dataclasses import dataclass

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from table_normalizer import TableNormalizer

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ProductVariant:
    """Вариант товара с конкретным тиражом и ценами"""
    name: str
    quantity: Optional[int]
    prices_usd: Dict[str, float]  # {'АВИА': 1.5, 'ЖД': 1.3}
    prices_rub: Dict[str, float]  # {'АВИА': 120, 'ЖД': 110}
    delivery_times: Dict[str, str]  # {'АВИА': '20-25', 'ЖД': '35-40'}
    base_data: Dict[str, str]  # description, material, size, etc.
    original_row: int

class AdvancedNormalizer(TableNormalizer):
    """Продвинутый нормализатор с поддержкой множественных тиражей"""
    
    def __init__(self):
        super().__init__()
        self.product_variants = []
        
    def _detect_product_variants(self, source_ws, structure: Dict) -> List[ProductVariant]:
        """Обнаруживает варианты товаров с разными тиражами, включая дополнительные строки"""
        
        variants = []
        data_start = structure['data_start_row']
        
        logger.info("🔍 Анализируем варианты товаров с множественными тиражами (включая пустые строки)...")
        logger.info(f"   Начинаем с строки: {data_start}")
        
        current_product_name = None
        current_base_data = {}
        current_product_rows = []  # Все строки текущего товара
        
        # ИСПРАВЛЕНИЕ: проверяем строки начиная с data_start-1 на случай если первые данные пропущены
        # Иногда анализатор структуры определяет начало данных не точно
        actual_start = max(data_start - 1, 3)  # Но не раньше 3-й строки (после заголовков)
        logger.info(f"   Корректируем начало на строку: {actual_start}")
        
        for row in range(actual_start, source_ws.max_row + 1):
            
            # Проверяем есть ли данные в строке (тиражи, цены)
            has_quantity_or_price = self._has_quantity_or_price_data(source_ws, row, structure)
            
            if not has_quantity_or_price:
                continue
            
            # Извлекаем данные товара
            variant_data = self._extract_variant_data(source_ws, row, structure)
            
            if variant_data['name']:
                # НОВЫЙ ТОВАР - завершаем предыдущий если был
                if current_product_name and current_product_rows:
                    self._finalize_product_variants(variants, current_product_name, current_base_data, current_product_rows)
                
                # Начинаем новый товар
                current_product_name = variant_data['name']
                current_base_data = variant_data['base_data']
                current_product_rows = [{'row': row, 'data': variant_data}]
                
                logger.debug(f"🆕 Новый товар: {current_product_name} (строка {row})")
                
            else:
                # ДОПОЛНИТЕЛЬНЫЙ ВАРИАНТ ТИРАЖА для текущего товара
                if current_product_name and (variant_data['quantity'] or variant_data['prices_usd'] or variant_data['prices_rub']):
                    current_product_rows.append({'row': row, 'data': variant_data})
                    logger.debug(f"📋 Дополнительный тираж для {current_product_name}: {variant_data['quantity']} шт (строка {row})")
        
        # Завершаем последний товар
        if current_product_name and current_product_rows:
            self._finalize_product_variants(variants, current_product_name, current_base_data, current_product_rows)
        
        logger.info(f"📊 Найдено вариантов товаров: {len(variants)}")
        
        # Статистика по товарам с множественными тиражами
        from collections import defaultdict
        products_with_variants = defaultdict(int)
        products_with_rows = defaultdict(list)
        
        for variant in variants:
            products_with_variants[variant.name] += 1
            products_with_rows[variant.name].append(variant.original_row)
        
        multiple_variants = [(name, count) for name, count in products_with_variants.items() if count > 1]
        if multiple_variants:
            logger.info(f"🎯 Товары с несколькими тиражами:")
            for name, count in multiple_variants[:7]:
                rows = products_with_rows[name]
                logger.info(f"   • {name}: {count} вариантов (строки: {rows})")
        
        return variants
    
    def _has_quantity_or_price_data(self, ws, row: int, structure: Dict) -> bool:
        """Проверяет есть ли в строке данные о тиражах или ценах"""
        
        # Проверяем базовые колонки количества
        if 'quantity' in structure['base_columns']:
            qty_col = structure['base_columns']['quantity']
            qty_value = ws.cell(row=row, column=qty_col).value
            if qty_value and str(qty_value).strip():
                return True
        
        # Проверяем маршрут-специфичные колонки
        for route_data in structure['routes'].values():
            for field_type, col in route_data.items():
                if field_type in ['quantity', 'price_usd', 'price_rub']:
                    value = ws.cell(row=row, column=col).value
                    if value and str(value).strip():
                        return True
        
        return False
    
    def _finalize_product_variants(self, variants: List[ProductVariant], product_name: str, 
                                 base_data: Dict, product_rows: List[Dict]):
        """Завершает обработку товара - создает варианты для всех его строк"""
        
        logger.debug(f"🔚 Завершаем товар {product_name}: {len(product_rows)} строк")
        
        for row_info in product_rows:
            row = row_info['row']
            variant_data = row_info['data']
            
            # Создаем вариант для каждой строки товара
            variants.append(ProductVariant(
                name=product_name,
                quantity=variant_data['quantity'],
                prices_usd=variant_data['prices_usd'],
                prices_rub=variant_data['prices_rub'],
                delivery_times=variant_data['delivery_times'],
                base_data=base_data,
                original_row=row
            ))
    
    def _extract_variant_data(self, ws, row: int, structure: Dict) -> Dict:
        """Извлекает данные варианта товара из строки"""
        
        result = {
            'name': '',
            'quantity': None,
            'prices_usd': {},
            'prices_rub': {},
            'delivery_times': {},
            'base_data': {}
        }
        
        logger.debug(f"Извлекаем данные из строки {row}")
        
        # Базовые поля
        for field_type, col in structure['base_columns'].items():
            value = ws.cell(row=row, column=col).value
            if value and str(value).strip():
                logger.debug(f"  Базовое поле {field_type} (колонка {col}): {value}")
                if field_type == 'name':
                    result['name'] = str(value).strip()
                elif field_type == 'quantity':
                    try:
                        # Парсим количество
                        qty_str = str(value).replace(' ', '').replace(',', '')
                        if qty_str.replace('.', '').isdigit():
                            result['quantity'] = int(float(qty_str))
                            logger.debug(f"    Найден тираж: {result['quantity']}")
                    except Exception as e:
                        logger.debug(f"    Ошибка парсинга тиража: {e}")
                else:
                    result['base_data'][field_type] = str(value).strip()
        
        # Маршрут-специфичные поля
        for route, route_data in structure['routes'].items():
            logger.debug(f"  Обрабатываем маршрут {route}: {route_data}")
            for field_type, col in route_data.items():
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    logger.debug(f"    {route} {field_type} (колонка {col}): {value}")
                    if field_type == 'quantity' and not result['quantity']:
                        try:
                            qty_str = str(value).replace(' ', '').replace(',', '')
                            if qty_str.replace('.', '').isdigit():
                                result['quantity'] = int(float(qty_str))
                                logger.debug(f"      Найден тираж маршрута: {result['quantity']}")
                        except Exception as e:
                            logger.debug(f"      Ошибка парсинга тиража маршрута: {e}")
                    elif field_type == 'price_usd':
                        try:
                            result['prices_usd'][route] = float(value)
                            logger.debug(f"      Найдена USD цена: {result['prices_usd'][route]}")
                        except Exception as e:
                            logger.debug(f"      Ошибка парсинга USD: {e}")
                    elif field_type == 'price_rub':
                        try:
                            result['prices_rub'][route] = float(value)
                            logger.debug(f"      Найдена RUB цена: {result['prices_rub'][route]}")
                        except Exception as e:
                            logger.debug(f"      Ошибка парсинга RUB: {e}")
                    elif field_type == 'delivery_time':
                        result['delivery_times'][route] = str(value).strip()
        
        logger.debug(f"  Результат извлечения: name='{result['name']}', quantity={result['quantity']}, prices_usd={result['prices_usd']}, prices_rub={result['prices_rub']}")
        return result
    
    def _copy_variants_to_normalized(self, variants: List[ProductVariant], target_ws, 
                                   headers: List[str], structure: Dict):
        """Копирует варианты товаров в нормализованную таблицу"""
        
        header_to_col = {header: col for col, header in enumerate(headers, 1)}
        target_row = 3  # Начинаем с 3-й строки (после заголовков)
        
        # Сохраняем маппинг для изображений
        structure['row_mappings'] = {}
        structure['variant_mappings'] = {}  # оригинал_строка → [нормализованные_строки]
        structure['product_row_groups'] = {}  # товар → [все_его_оригинальные_строки]
        
        logger.info(f"📋 Копируем {len(variants)} вариантов товаров...")
        
        # Группируем строки по товарам для правильной привязки изображений
        from collections import defaultdict
        product_original_rows = defaultdict(list)
        
        for variant in variants:
            product_original_rows[variant.name].append(variant.original_row)
        
        # Сохраняем группы строк товаров
        structure['product_row_groups'] = dict(product_original_rows)
        
        for variant in variants:
            
            logger.debug(f"Копируем: {variant.name} (тираж: {variant.quantity}, строка: {variant.original_row})")
            
            # Маппинг строк
            structure['row_mappings'][variant.original_row] = target_row
            
            if variant.original_row not in structure['variant_mappings']:
                structure['variant_mappings'][variant.original_row] = []
            structure['variant_mappings'][variant.original_row].append(target_row)
            
            # Базовые поля
            target_ws.cell(row=target_row, column=header_to_col['Наименование товара'], value=variant.name)
            
            for field, value in variant.base_data.items():
                if field == 'description':
                    target_col = header_to_col.get('Описание')
                elif field == 'material':
                    target_col = header_to_col.get('Материал')
                elif field == 'size':
                    target_col = header_to_col.get('Размер')
                elif field == 'color':
                    target_col = header_to_col.get('Цвет')
                elif field == 'packaging':
                    target_col = header_to_col.get('Упаковка')
                else:
                    target_col = None
                
                if target_col and value:
                    target_ws.cell(row=target_row, column=target_col, value=value)
            
            # Данные по маршрутам
            for route in structure['routes'].keys():
                
                # Тираж
                qty_col = header_to_col.get(f'Тираж ({route}), шт')
                if qty_col and variant.quantity:
                    target_ws.cell(row=target_row, column=qty_col, value=variant.quantity)
                
                # Цена USD
                usd_col = header_to_col.get(f'Цена ({route}), $')
                if usd_col and route in variant.prices_usd:
                    target_ws.cell(row=target_row, column=usd_col, value=variant.prices_usd[route])
                
                # Цена RUB
                rub_col = header_to_col.get(f'Цена ({route}), ₽')
                if rub_col and route in variant.prices_rub:
                    target_ws.cell(row=target_row, column=rub_col, value=variant.prices_rub[route])
                
                # Срок доставки
                time_col = header_to_col.get(f'Срок ({route}), дн.')
                if time_col and route in variant.delivery_times:
                    target_ws.cell(row=target_row, column=time_col, value=variant.delivery_times[route])
            
            target_row += 1
        
        # Сохраняем информацию о диапазоне
        structure['normalized_data_range'] = {
            'start_row': 3,
            'end_row': target_row - 1,
            'total_products': target_row - 3
        }
        
        logger.info(f"✅ Скопировано {target_row - 3} строк товаров")
    
    def normalize_table(self, file_path: str) -> str:
        """Нормализует таблицу с поддержкой множественных тиражей"""
        
        logger.info(f"🔧 Продвинутая нормализация: {file_path}")
        
        # Открываем исходную таблицу
        source_wb = openpyxl.load_workbook(file_path, data_only=True)
        source_ws = source_wb.active
        
        # Анализируем структуру (вызываем напрямую родительский метод)
        structure = TableNormalizer._analyze_current_structure(self, source_ws)
        structure['original_file'] = os.path.basename(file_path)
        structure['original_end_row'] = source_ws.max_row
        
        if not structure['routes'] and not structure['base_columns']:
            logger.warning(f"❌ Не удалось определить структуру таблицы: {file_path}")
            source_wb.close()
            return None
        
        # Обнаруживаем варианты товаров
        variants = self._detect_product_variants(source_ws, structure)
        
        if not variants:
            logger.warning(f"❌ Не найдено товаров для нормализации: {file_path}")
            source_wb.close()
            return None
        
        # Создаем нормализованную таблицу
        target_wb = openpyxl.Workbook()
        target_ws = target_wb.active
        
        # Генерируем заголовки
        headers = TableNormalizer._create_normalized_headers(self, structure)
        
        # Записываем заголовки
        TableNormalizer._write_headers(self, target_ws, headers)
        
        # Копируем варианты товаров
        self._copy_variants_to_normalized(variants, target_ws, headers, structure)
        
        # Применяем форматирование
        TableNormalizer._apply_formatting(self, target_ws, headers)
        
        # Сохраняем
        target_path = file_path.replace('.xlsx', '_normalized.xlsx')
        target_wb.save(target_path)
        target_wb.close()
        source_wb.close()
        
        # Сохраняем расширенные метаданные с группами строк товаров
        self._save_advanced_metadata(target_path, structure)
        
        logger.info(f"✅ Продвинутая нормализация завершена: {target_path}")
        return target_path
    
    def _save_advanced_metadata(self, normalized_path: str, structure: Dict):
        """Сохраняет расширенные метаданные с информацией о группах строк товаров"""
        
        import json
        
        metadata = {
            'normalized_file': os.path.basename(normalized_path),
            'original_file': structure.get('original_file', ''),
            'data_ranges': {
                'original_start_row': structure['data_start_row'],
                'original_end_row': structure.get('original_end_row', 0),
                'normalized_start_row': structure['normalized_data_range']['start_row'],
                'normalized_end_row': structure['normalized_data_range']['end_row'],
                'total_products': structure['normalized_data_range']['total_products']
            },
            'row_mappings': structure['row_mappings'],  # оригинальная_строка → нормализованная_строка
            'variant_mappings': structure.get('variant_mappings', {}),  # оригинальная_строка → [нормализованные_строки]
            'product_row_groups': structure.get('product_row_groups', {}),  # товар → [все_его_оригинальные_строки]
            'found_routes': list(structure['routes'].keys()),
            'base_columns': list(structure['base_columns'].keys()),
            'advanced_parsing': True,  # Флаг продвинутого парсинга
            'multiple_quantity_support': True,  # Поддержка множественных тиражей
            'created_at': str(datetime.now())
        }
        
        # Сохраняем в файл рядом с нормализованной таблицей
        metadata_path = normalized_path.replace('.xlsx', '_metadata.json')
        
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
            
        logger.info(f"📋 Расширенные метаданные сохранены: {metadata_path}")
        
        return metadata_path

if __name__ == "__main__":
    
    # Тестируем на известной таблице с множественными тиражами
    test_file = 'storage/excel_files/original_sheet.xlsx'
    
    if os.path.exists(test_file):
        normalizer = AdvancedNormalizer()
        result = normalizer.normalize_table(test_file)
        print(f"✅ Тест завершен: {result}")
    else:
        print(f"❌ Тестовый файл не найден: {test_file}")
