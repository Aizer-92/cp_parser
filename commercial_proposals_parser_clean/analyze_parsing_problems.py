#!/usr/bin/env python3
"""
Анализ проблем парсинга - почему из 145 файлов получилось только 19 товаров
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def analyze_parsing_problems():
    """Анализируем проблемы парсинга"""
    
    print("🔍 АНАЛИЗ ПРОБЛЕМ ПАРСИНГА")
    print("=" * 80)
    
    # Проверяем файлы
    storage_path = Path("storage/excel_files")
    all_files = list(storage_path.glob("*.xlsx"))
    non_empty_files = [f for f in all_files if f.stat().st_size > 0]
    
    print(f"📁 Всего файлов: {len(all_files)}")
    print(f"📁 Непустых файлов: {len(non_empty_files)}")
    print(f"📁 Пустых файлов: {len(all_files) - len(non_empty_files)}")
    
    # Проверяем БД
    session = DatabaseManager.get_session()
    try:
        total_sheets = session.query(SheetMetadata).count()
        completed_sheets = session.query(SheetMetadata).filter(
            SheetMetadata.status == 'completed'
        ).count()
        downloaded_sheets = session.query(SheetMetadata).filter(
            SheetMetadata.local_file_path.isnot(None)
        ).count()
        
        print(f"\n📊 СТАТИСТИКА БД:")
        print(f"   Всего записей SheetMetadata: {total_sheets}")
        print(f"   Завершенных: {completed_sheets}")
        print(f"   С локальными файлами: {downloaded_sheets}")
        
        # Проверяем товары
        total_products = session.query(Product).count()
        print(f"   Всего товаров: {total_products}")
        
        # Анализируем последние таблицы
        print(f"\n🔍 АНАЛИЗ ПОСЛЕДНИХ ТАБЛИЦ:")
        recent_sheets = session.query(SheetMetadata).filter(
            SheetMetadata.local_file_path.isnot(None)
        ).order_by(SheetMetadata.id.desc()).limit(20).all()
        
        for i, sheet in enumerate(recent_sheets, 1):
            file_path = Path(sheet.local_file_path)
            exists = file_path.exists() if file_path else False
            size = file_path.stat().st_size if exists else 0
            
            print(f"   {i:2d}. {sheet.sheet_title[:50]:<50} "
                  f"| {sheet.status:<10} | {size:>8} bytes")
            
            # Пробуем открыть файл
            if exists and size > 0:
                try:
                    wb = load_workbook(file_path, read_only=True)
                    sheets_count = len(wb.sheetnames)
                    first_sheet = wb[wb.sheetnames[0]]
                    rows_count = first_sheet.max_row
                    cols_count = first_sheet.max_column
                    
                    print(f"       └─ Листов: {sheets_count}, Строк: {rows_count}, Колонок: {cols_count}")
                    
                    # Проверяем первые несколько строк
                    sample_data = []
                    for row in range(1, min(6, rows_count + 1)):
                        row_data = []
                        for col in range(1, min(6, cols_count + 1)):
                            cell_value = first_sheet.cell(row=row, column=col).value
                            if cell_value:
                                row_data.append(str(cell_value)[:20])
                        if row_data:
                            sample_data.append(" | ".join(row_data))
                    
                    if sample_data:
                        print(f"       └─ Образец данных:")
                        for j, row_data in enumerate(sample_data[:3], 1):
                            print(f"          {j}: {row_data}")
                    
                    wb.close()
                    
                except Exception as e:
                    print(f"       └─ ❌ Ошибка чтения: {e}")
    
    finally:
        session.close()
    
    print(f"\n🎯 РЕКОМЕНДАЦИИ:")
    print(f"   1. Проверить почему многие файлы пустые (0 bytes)")
    print(f"   2. Изучить структуру CSV файлов от AdvancedDownloader") 
    print(f"   3. Улучшить логику парсинга для нестандартных таблиц")
    print(f"   4. Добавить детальное логирование ошибок парсинга")

if __name__ == "__main__":
    analyze_parsing_problems()


