#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Тестирование продвинутого нормализатора на таблице с множественными тиражами
"""

import openpyxl
from advanced_normalizer import AdvancedNormalizer

# Создаем тестовую таблицу с множественными тиражами
def create_test_table():
    """Создает тестовую таблицу с товарами имеющими несколько вариантов тиражей"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Заголовки
    ws.cell(row=1, column=1, value="Фото")
    ws.cell(row=1, column=2, value="Менеджер: Тест\nНаименование")
    ws.cell(row=1, column=3, value="Характеристики")
    ws.cell(row=1, column=4, value="Тираж, шт")
    ws.cell(row=1, column=5, value="Цена ЖД, $")
    ws.cell(row=1, column=6, value="Цена ЖД, ₽")
    ws.cell(row=1, column=7, value="Цена АВИА, $") 
    ws.cell(row=1, column=8, value="Цена АВИА, ₽")
    ws.cell(row=1, column=9, value="Образец, ₽")
    
    # Заголовки 2-я строка
    ws.cell(row=2, column=2, value="")
    ws.cell(row=2, column=3, value="")
    ws.cell(row=2, column=4, value="")
    ws.cell(row=2, column=5, value="")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="")
    ws.cell(row=2, column=8, value="")
    ws.cell(row=2, column=9, value="")
    
    # Данные - Товар 1 с несколькими тиражами
    ws.cell(row=3, column=2, value="Кружка керамическая")  # Основная строка товара
    ws.cell(row=3, column=3, value="Материал: керамика\nРазмер: 330мл")
    ws.cell(row=3, column=4, value=1000)  # 1000 шт
    ws.cell(row=3, column=5, value=2.50)  # ЖД $2.50
    ws.cell(row=3, column=6, value=225)   # ЖД 225₽
    ws.cell(row=3, column=7, value=2.80)  # АВИА $2.80
    ws.cell(row=3, column=8, value=252)   # АВИА 252₽
    ws.cell(row=3, column=9, value=1500)  # Образец 1500₽
    
    # Дополнительные тиражи для той же кружки (без названия!)
    ws.cell(row=4, column=2, value="")    # Пустое название - это дополнительный вариант
    ws.cell(row=4, column=3, value="")    # Пустые характеристики
    ws.cell(row=4, column=4, value=5000)  # 5000 шт - другая цена
    ws.cell(row=4, column=5, value=2.20)  # ЖД $2.20
    ws.cell(row=4, column=6, value=198)   # ЖД 198₽
    ws.cell(row=4, column=7, value=2.45)  # АВИА $2.45
    ws.cell(row=4, column=8, value=220)   # АВИА 220₽
    ws.cell(row=4, column=9, value=1500)  # Образец тот же
    
    ws.cell(row=5, column=2, value="")    # Еще один тираж той же кружки
    ws.cell(row=5, column=3, value="")
    ws.cell(row=5, column=4, value=10000) # 10000 шт - еще дешевле
    ws.cell(row=5, column=5, value=1.95)  # ЖД $1.95
    ws.cell(row=5, column=6, value=176)   # ЖД 176₽
    ws.cell(row=5, column=7, value=2.15)  # АВИА $2.15
    ws.cell(row=5, column=8, value=194)   # АВИА 194₽
    ws.cell(row=5, column=9, value=1500)  # Образец тот же
    
    # Товар 2 с несколькими тиражами
    ws.cell(row=6, column=2, value="Ручка шариковая")  # Новый товар
    ws.cell(row=6, column=3, value="Материал: пластик\nЦвет: синий")
    ws.cell(row=6, column=4, value=500)   # 500 шт
    ws.cell(row=6, column=5, value=0.65)  # ЖД $0.65
    ws.cell(row=6, column=6, value=58)    # ЖД 58₽
    ws.cell(row=6, column=7, value=0.75)  # АВИА $0.75
    ws.cell(row=6, column=8, value=68)    # АВИА 68₽
    ws.cell(row=6, column=9, value=150)   # Образец 150₽
    
    # Дополнительный тираж ручки
    ws.cell(row=7, column=2, value="")    # Пустое название
    ws.cell(row=7, column=3, value="")
    ws.cell(row=7, column=4, value=2000)  # 2000 шт
    ws.cell(row=7, column=5, value=0.55)  # ЖД $0.55
    ws.cell(row=7, column=6, value=50)    # ЖД 50₽
    ws.cell(row=7, column=7, value=0.62)  # АВИА $0.62
    ws.cell(row=7, column=8, value=56)    # АВИА 56₽
    ws.cell(row=7, column=9, value=150)   # Образец тот же
    
    # Товар 3 с одним тиражом (для сравнения)
    ws.cell(row=8, column=2, value="Блокнот А5")
    ws.cell(row=8, column=3, value="Материал: бумага\n50 листов")
    ws.cell(row=8, column=4, value=1000)
    ws.cell(row=8, column=5, value=1.25)
    ws.cell(row=8, column=6, value=112)
    ws.cell(row=8, column=7, value=1.40)
    ws.cell(row=8, column=8, value=126)
    ws.cell(row=8, column=9, value=800)
    
    # Сохраняем
    test_file = 'storage/excel_files/test_multiple_quantities.xlsx'
    wb.save(test_file)
    wb.close()
    
    return test_file

def test_advanced_normalizer():
    """Тестирует продвинутый нормализатор"""
    
    print("🧪 ТЕСТИРОВАНИЕ ПРОДВИНУТОГО НОРМАЛИЗАТОРА")
    print("=" * 80)
    
    # Создаем тестовую таблицу
    test_file = create_test_table()
    print(f"✅ Создана тестовая таблица: {test_file}")
    
    print("\n📋 СТРУКТУРА ТЕСТОВОЙ ТАБЛИЦЫ:")
    print("   • Кружка керамическая: 3 варианта тиражей (1000, 5000, 10000 шт)")
    print("   • Ручка шариковая: 2 варианта тиражей (500, 2000 шт)")
    print("   • Блокнот А5: 1 вариант (1000 шт)")
    print("   • Ожидаем: 6 строк в нормализованной таблице")
    
    # Включаем отладку
    import logging
    logging.getLogger('advanced_normalizer').setLevel(logging.DEBUG)
    logging.getLogger('table_normalizer').setLevel(logging.DEBUG)
    
    # Тестируем нормализатор
    normalizer = AdvancedNormalizer()
    result_file = normalizer.normalize_table(test_file)
    
    if result_file:
        print(f"\n✅ Нормализация успешна: {result_file}")
        
        # Проверяем результат
        check_normalized_result(result_file)
    else:
        print("\n❌ Ошибка нормализации")

def check_normalized_result(normalized_file):
    """Проверяет результат нормализации"""
    
    import json
    
    print("\n🔍 АНАЛИЗ РЕЗУЛЬТАТА:")
    
    # Читаем метаданные
    metadata_file = normalized_file.replace('.xlsx', '_metadata.json')
    with open(metadata_file, 'r', encoding='utf-8') as f:
        metadata = json.load(f)
    
    print(f"📊 Товаров в результате: {metadata['data_ranges']['total_products']}")
    print(f"🗺️ Групп строк товаров: {len(metadata['product_row_groups'])}")
    
    print("\n🛍️ ГРУППЫ СТРОК ТОВАРОВ:")
    for product_name, original_rows in metadata['product_row_groups'].items():
        print(f"   • {product_name}: строки {original_rows}")
    
    print("\n🔗 МАППИНГ СТРОК (оригинал → нормализованная):")
    for orig_row, norm_row in sorted(metadata['row_mappings'].items(), key=lambda x: int(x[0])):
        print(f"   • Строка {orig_row} → Строка {norm_row}")
    
    # Читаем нормализованную таблицу
    wb = openpyxl.load_workbook(normalized_file, data_only=True)
    ws = wb.active
    
    print("\n📋 НОРМАЛИЗОВАННЫЕ ДАННЫЕ:")
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value or ""
        qty_avia = ws.cell(row=row, column=8).value or ""
        price_avia_usd = ws.cell(row=row, column=9).value or ""
        price_avia_rub = ws.cell(row=row, column=10).value or ""
        print(f"   {row-2}: {name[:25]:<25} | {qty_avia} шт | ${price_avia_usd} | ₽{price_avia_rub}")
    
    wb.close()

if __name__ == "__main__":
    test_advanced_normalizer()
