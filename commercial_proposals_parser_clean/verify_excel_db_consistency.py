#!/usr/bin/env python3
"""
Проверка соответствия Excel файлов и базы данных
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product
import logging
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def analyze_excel_file(file_path):
    """Анализ Excel файла для подсчета товаров"""
    try:
        # Пробуем открыть как Excel
        wb = load_workbook(file_path, read_only=True)
        first_sheet = wb[wb.sheetnames[0]]
        
        if first_sheet.max_row is None:
            return {"products": 0, "error": "Empty sheet"}
        
        # Ищем товары - строки с данными исключая заголовки
        product_count = 0
        for row in range(1, min(first_sheet.max_row + 1, 100)):  # Ограничиваем поиск первыми 100 строками
            # Проверяем есть ли данные в первых колонках
            row_data = []
            for col in range(1, min(6, first_sheet.max_column + 1)):
                cell_value = first_sheet.cell(row=row, column=col).value
                if cell_value:
                    row_data.append(str(cell_value).strip())
            
            if row_data:
                # Простая эвристика: если в строке есть данные и это не заголовок
                row_text = " ".join(row_data).lower()
                if (not any(header in row_text for header in ['manager', 'менеджер', 'name', 'наименование', 'photo', 'фото', 'email', 'phone']) 
                    and len(row_text) > 5):
                    product_count += 1
        
        wb.close()
        return {"products": max(0, product_count - 2), "error": None}  # -2 для заголовков
        
    except Exception as e:
        # Пробуем как CSV если Excel не работает
        try:
            df = pd.read_csv(file_path, encoding='utf-8', sep=None, engine='python')
            # Фильтруем строки с данными
            non_empty_rows = len([row for _, row in df.iterrows() 
                                if row.notna().sum() > 1])  # Строки с более чем одним непустым значением
            return {"products": max(0, non_empty_rows - 1), "error": None}  # -1 для заголовка
        except:
            return {"products": 0, "error": str(e)}

def verify_consistency():
    """Проверка соответствия Excel файлов и базы данных"""
    
    print("🔍 ПРОВЕРКА СООТВЕТСТВИЯ EXCEL ФАЙЛОВ И БАЗЫ ДАННЫХ")
    print("=" * 80)
    
    session = DatabaseManager.get_session()
    
    try:
        # Получаем все файлы из папки
        storage_path = Path("storage/excel_files")
        excel_files = list(storage_path.glob("*.xlsx"))
        excel_files.extend(list(storage_path.glob("*.csv")))
        
        print(f"📁 Найдено файлов: {len(excel_files)}")
        
        # Анализируем файлы
        file_analysis = {}
        total_products_in_files = 0
        
        for file_path in excel_files:
            if file_path.stat().st_size == 0:
                continue  # Пропускаем пустые файлы
            
            analysis = analyze_excel_file(file_path)
            file_analysis[file_path.name] = analysis
            total_products_in_files += analysis["products"]
            
            if analysis["error"]:
                logger.warning(f"❌ {file_path.name}: {analysis['error']}")
            else:
                logger.info(f"✅ {file_path.name}: {analysis['products']} товаров")
        
        # Получаем данные из БД
        all_sheets = session.query(SheetMetadata).all()
        completed_sheets = session.query(SheetMetadata).filter(SheetMetadata.status == 'completed').all()
        all_products = session.query(Product).all()
        
        # Группируем товары по таблицам
        products_by_sheet = {}
        for product in all_products:
            sheet_id = product.sheet_id
            if sheet_id not in products_by_sheet:
                products_by_sheet[sheet_id] = []
            products_by_sheet[sheet_id].append(product)
        
        print(f"\n📊 СВОДНАЯ СТАТИСТИКА:")
        print(f"   📁 Всего Excel файлов: {len([f for f in excel_files if f.stat().st_size > 0])}")
        print(f"   🧮 Товаров в файлах (подсчет): {total_products_in_files}")
        print(f"   📑 Записей SheetMetadata: {len(all_sheets)}")
        print(f"   ✅ Завершенных таблиц: {len(completed_sheets)}")
        print(f"   🛍️  Товаров в БД: {len(all_products)}")
        
        # Детальный анализ по файлам
        print(f"\n🔍 ДЕТАЛЬНЫЙ АНАЛИЗ ПО ФАЙЛАМ:")
        print("-" * 80)
        
        matched_files = 0
        unmatched_files = 0
        
        for file_name, analysis in file_analysis.items():
            if analysis["products"] == 0 and analysis["error"]:
                continue
                
            # Ищем соответствующую запись в БД по названию файла
            matching_sheet = None
            for sheet in all_sheets:
                if sheet.local_file_path and file_name in sheet.local_file_path:
                    matching_sheet = sheet
                    break
            
            if matching_sheet:
                products_in_db = len(products_by_sheet.get(matching_sheet.id, []))
                status_icon = "✅" if matching_sheet.status == 'completed' else "⏳"
                match_icon = "✅" if products_in_db > 0 else "❌" if analysis["products"] > 0 else "⭕"
                
                print(f"   {match_icon} {status_icon} {file_name[:50]:<50} | Файл: {analysis['products']:2d} | БД: {products_in_db:2d}")
                
                if products_in_db > 0 or analysis["products"] == 0:
                    matched_files += 1
                else:
                    unmatched_files += 1
            else:
                print(f"   ❓ ⭕ {file_name[:50]:<50} | Файл: {analysis['products']:2d} | БД: нет записи")
                if analysis["products"] > 0:
                    unmatched_files += 1
        
        # Итоговая оценка
        print(f"\n🎯 ИТОГОВАЯ ОЦЕНКА:")
        coverage = (matched_files / max(1, matched_files + unmatched_files)) * 100 if (matched_files + unmatched_files) > 0 else 0
        print(f"   ✅ Обработанных файлов: {matched_files}")
        print(f"   ❌ Необработанных файлов: {unmatched_files}")
        print(f"   📊 Покрытие: {coverage:.1f}%")
        
        if coverage > 90:
            print(f"   🎉 Отлично! Почти все файлы обработаны")
        elif coverage > 70:
            print(f"   👍 Хорошо! Большинство файлов обработано")
        else:
            print(f"   🔧 Нужна работа: много файлов не обработано")
        
        # Рекомендации
        if unmatched_files > 0:
            print(f"\n💡 РЕКОМЕНДАЦИИ:")
            print(f"   📥 Запустить: python3 parse_existing_tables.py")
            print(f"   🔄 Или исправить проблемные файлы и перепарсить")
        
        return {
            "total_files": len(excel_files),
            "products_in_files": total_products_in_files,
            "products_in_db": len(all_products),
            "coverage": coverage,
            "matched_files": matched_files,
            "unmatched_files": unmatched_files
        }
        
    finally:
        session.close()

if __name__ == "__main__":
    verify_consistency()


