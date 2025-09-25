#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Парсер для нормализованных таблиц с идеальной структурой
"""

import os
import sys
import json
import logging
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from database.manager_v4 import CommercialProposalsDB
from database.models_v4 import Product, PriceOffer, ProductImage, SheetMetadata
from config import DATABASE_URL_V4

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class NormalizedTableParser:
    """Парсер для нормализованных таблиц с четкой структурой"""
    
    def __init__(self):
        self.db = CommercialProposalsDB(DATABASE_URL_V4)
        
        # Стандартная структура нормализованных таблиц
        self.standard_columns = {
            1: 'name',                    # Наименование товара
            2: 'article',                 # Артикул  
            3: 'description',             # Описание
            4: 'material',                # Материал
            5: 'size',                    # Размер
            6: 'color',                   # Цвет
            7: 'packaging',               # Упаковка
            8: 'quantity_avia',           # Тираж (АВИА), шт
            9: 'price_avia_usd',          # Цена (АВИА), $
            10: 'price_avia_rub',         # Цена (АВИА), ₽
            11: 'delivery_avia',          # Срок (АВИА), дн.
            12: 'quantity_jd',            # Тираж (ЖД), шт
            13: 'price_jd_usd',           # Цена (ЖД), $
            14: 'price_jd_rub',           # Цена (ЖД), ₽
            15: 'delivery_jd',            # Срок (ЖД), дн.
            16: 'quantity_sample',        # Тираж (ОБРАЗЕЦ), шт
            17: 'price_sample_usd',       # Цена (ОБРАЗЕЦ), $
            18: 'price_sample_rub',       # Цена (ОБРАЗЕЦ), ₽
            19: 'delivery_sample',        # Срок (ОБРАЗЕЦ), дн.
        }
    
    def parse_normalized_table(self, file_path: str, metadata_path: str) -> Tuple[int, int]:
        """Парсит одну нормализованную таблицу"""
        
        logger.info(f"📋 Парсим нормализованную таблицу: {Path(file_path).name}")
        
        # Загружаем метаданные
        with open(metadata_path, 'r', encoding='utf-8') as f:
            metadata = json.load(f)
        
        session = self.db.get_session()
        
        try:
            # Определяем или создаем SheetMetadata (передаем сессию)
            sheet_metadata = self._get_or_create_sheet_metadata(file_path, metadata, session)
        
        
            # Парсим таблицу
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            products_count = 0
            price_offers_count = 0
            # Читаем все строки товаров (начиная с 3-й строки)
            for row in range(3, ws.max_row + 1):
                
                # Извлекаем базовые данные товара
                product_data = self._extract_product_data(ws, row)
                
                if not product_data['name']:
                    continue  # Пропускаем пустые строки
                
                # Создаем товар (убираем row_number - его нет в модели)
                product = Product(
                    name=product_data['name'],
                    description=product_data['description'],
                    characteristics=product_data['characteristics'],
                    custom_design=product_data['custom_design'],
                    sheet_id=sheet_metadata.id
                )
                
                session.add(product)
                session.flush()  # Получаем ID товара
                
                products_count += 1
                
                # Извлекаем варианты цен для всех маршрутов
                price_offers = self._extract_price_offers(ws, row, product.id)
                
                for offer in price_offers:
                    session.add(offer)
                    price_offers_count += 1
            
            # Сохраняем изменения
            session.commit()
            
            wb.close()
            
            logger.info(f"✅ Обработано: {products_count} товаров, {price_offers_count} цен")
            return products_count, price_offers_count
            
        except Exception as e:
            session.rollback()
            wb.close()
            logger.error(f"❌ Ошибка парсинга {file_path}: {e}")
            return 0, 0
            
        finally:
            session.close()
    
    def _extract_product_data(self, ws, row: int) -> Dict:
        """Извлекает базовые данные товара из строки"""
        
        def get_cell_value(col: int) -> Optional[str]:
            value = ws.cell(row=row, column=col).value
            return str(value).strip() if value else None
        
        # Соединяем все дополнительные характеристики
        characteristics_parts = []
        for field in ['material', 'size', 'color', 'packaging']:
            col = None
            for c, field_name in self.standard_columns.items():
                if field_name == field:
                    col = c
                    break
            if col:
                value = get_cell_value(col)
                if value:
                    characteristics_parts.append(f"{field.title()}: {value}")
        
        characteristics = "; ".join(characteristics_parts) if characteristics_parts else None
        
        return {
            'name': get_cell_value(1),
            'description': get_cell_value(3),
            'characteristics': characteristics,
            'custom_design': None,  # Пока не используем
        }
    
    def _extract_price_offers(self, ws, row: int, product_id: int) -> List[PriceOffer]:
        """Извлекает все варианты цен для товара"""
        
        offers = []
        
        # Проверяем каждый маршрут
        routes = [
            ('АВИА', 'quantity_avia', 'price_avia_usd', 'price_avia_rub', 'delivery_avia'),
            ('ЖД', 'quantity_jd', 'price_jd_usd', 'price_jd_rub', 'delivery_jd'), 
            ('ОБРАЗЕЦ', 'quantity_sample', 'price_sample_usd', 'price_sample_rub', 'delivery_sample')
        ]
        
        for route_name, qty_field, usd_field, rub_field, delivery_field in routes:
            
            # Находим колонки для этого маршрута
            qty_col = self._find_column_by_field(qty_field)
            usd_col = self._find_column_by_field(usd_field)
            rub_col = self._find_column_by_field(rub_field)
            delivery_col = self._find_column_by_field(delivery_field)
            
            # Извлекаем данные
            quantity = self._parse_number(ws.cell(row=row, column=qty_col).value) if qty_col else None
            price_usd = self._parse_number(ws.cell(row=row, column=usd_col).value) if usd_col else None
            price_rub = self._parse_number(ws.cell(row=row, column=rub_col).value) if rub_col else None
            delivery_time = str(ws.cell(row=row, column=delivery_col).value).strip() if delivery_col and ws.cell(row=row, column=delivery_col).value else None
            
            # Создаем предложение если есть хоть какие-то данные
            if quantity or price_usd or price_rub:
                
                offer = PriceOffer(
                    product_id=product_id,
                    route_name=route_name,
                    quantity=int(quantity) if quantity else None,
                    price_usd=float(price_usd) if price_usd else None,
                    price_rub=float(price_rub) if price_rub else None,
                    delivery_time=delivery_time,
                    is_sample=(route_name == 'ОБРАЗЕЦ'),
                    is_available=True
                )
                
                offers.append(offer)
        
        return offers
    
    def _find_column_by_field(self, field_name: str) -> Optional[int]:
        """Находит номер колонки по имени поля"""
        
        for col, name in self.standard_columns.items():
            if name == field_name:
                return col
        return None
    
    def _parse_number(self, value) -> Optional[float]:
        """Парсит число из ячейки"""
        
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        # Убираем все лишнее и пытаемся парсить
        clean_str = str(value).replace(' ', '').replace(',', '.').strip()
        
        # Убираем символы валют и единицы
        for symbol in ['$', '₽', 'руб', 'шт', 'pcs']:
            clean_str = clean_str.replace(symbol, '')
        
        try:
            return float(clean_str)
        except:
            return None
    
    def _get_or_create_sheet_metadata(self, file_path: str, metadata: Dict, session) -> SheetMetadata:
        """Получает или создает SheetMetadata - используем переданную сессию!"""
        
        # Генерируем уникальный sheet_id из пути
        sheet_id_str = Path(file_path).stem.replace('_normalized', '')
        
        # Ищем существующий
        sheet_meta = session.query(SheetMetadata).filter_by(sheet_id=sheet_id_str).first()
        
        if not sheet_meta:
            # Создаем новый
            sheet_meta = SheetMetadata(
                sheet_url=f"file://{file_path}",
                sheet_title=Path(file_path).name,
                sheet_id=sheet_id_str,
                status='normalized_parsed',
                products_count=0,  # Пока 0, обновим позже
                local_file_path=file_path
            )
            session.add(sheet_meta)
            session.flush()  # Получаем ID без коммита
        
        return sheet_meta

def clear_old_products_and_prices():
    """Очищает старые данные товаров и цен (изображения НЕ трогаем!)"""
    
    print("🗑️ ОЧИСТКА СТАРЫХ ДАННЫХ ТОВАРОВ И ЦЕН")
    print("🖼️ ИЗОБРАЖЕНИЯ ОСТАВЛЯЕМ - ОНИ УЖЕ ХОРОШО ПРИВЯЗАНЫ!")
    print("=" * 60)
    
    db = CommercialProposalsDB(DATABASE_URL_V4)
    session = db.get_session()
    
    try:
        # Считаем что удаляем
        products_count = session.query(Product).count()
        prices_count = session.query(PriceOffer).count()
        images_count = session.query(ProductImage).count()
        
        print(f"📊 К удалению:")
        print(f"   • Товаров: {products_count}")
        print(f"   • Вариантов цен: {prices_count}")
        print(f"🖼️ ИЗОБРАЖЕНИЯ ОСТАЮТСЯ: {images_count} (не трогаем)")
        
        # Удаляем только цены и товары - изображения НЕ ТРОГАЕМ!
        session.query(PriceOffer).delete()
        session.query(Product).delete()
        
        session.commit()
        
        print("✅ Старые товары и цены удалены!")
        print("🖼️ Изображения сохранены с их привязками!")
        
    except Exception as e:
        session.rollback()
        print(f"❌ Ошибка очистки: {e}")
        raise
        
    finally:
        session.close()

def parse_all_normalized_tables():
    """Парсит все нормализованные таблицы"""
    
    print("📋 ПАРСИНГ ВСЕХ НОРМАЛИЗОВАННЫХ ТАБЛИЦ")
    print("=" * 80)
    
    # Находим все нормализованные файлы
    excel_dir = Path("storage/excel_files")
    normalized_files = list(excel_dir.glob("*_normalized.xlsx"))
    
    print(f"📊 Найдено нормализованных таблиц: {len(normalized_files)}")
    
    if not normalized_files:
        print("❌ Нормализованные таблицы не найдены!")
        return
    
    parser = NormalizedTableParser()
    
    total_products = 0
    total_prices = 0
    success_count = 0
    error_count = 0
    
    for i, file_path in enumerate(normalized_files, 1):
        
        # Ищем файл метаданных
        metadata_path = str(file_path).replace('_normalized.xlsx', '_normalized_metadata.json')
        
        if not Path(metadata_path).exists():
            logger.error(f"❌ Нет метаданных для {file_path.name}")
            error_count += 1
            continue
        
        print(f"\n📋 [{i}/{len(normalized_files)}] {file_path.name}")
        
        try:
            products, prices = parser.parse_normalized_table(str(file_path), metadata_path)
            
            if products > 0:
                total_products += products
                total_prices += prices
                success_count += 1
                print(f"   ✅ {products} товаров, {prices} цен")
            else:
                error_count += 1
                print(f"   ⚠️ Пустая таблица")
                
        except Exception as e:
            error_count += 1
            logger.error(f"   ❌ Ошибка: {e}")
        
        # Прогресс каждые 25 файлов
        if i % 25 == 0:
            print(f"   📊 Прогресс: {i}/{len(normalized_files)} | Товаров: {total_products} | Цен: {total_prices}")
    
    # Итоговая статистика
    print(f"\n🎯 ИТОГИ ПАРСИНГА:")
    print("=" * 60)
    print(f"✅ Успешно: {success_count}")
    print(f"❌ Ошибок: {error_count}")
    print(f"📊 Товаров создано: {total_products}")
    print(f"💰 Вариантов цен: {total_prices}")
    print(f"📋 Среднее цен на товар: {total_prices/total_products:.1f}" if total_products > 0 else "")

if __name__ == "__main__":
    
    print("🎯 ОБНОВЛЕНИЕ БД ИЗ НОРМАЛИЗОВАННЫХ ТАБЛИЦ")
    print("=" * 80)
    
    # Подтверждение
    response = input("⚠️ Это удалит все текущие товары и цены! Продолжить? (y/N): ")
    
    if response.lower() in ['y', 'yes', 'да']:
        
        # 1. Очищаем старые данные
        clear_old_products_and_prices()
        
        # 2. Парсим нормализованные таблицы
        parse_all_normalized_tables()
        
        print(f"\n🎉 ОБНОВЛЕНИЕ ЗАВЕРШЕНО!")
        print("🖼️ Изображения остались с оригинальными привязками!")
        
    else:
        print("❌ Отменено пользователем")
