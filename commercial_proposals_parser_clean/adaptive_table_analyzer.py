#!/usr/bin/env python3
"""
Адаптивный анализатор структуры таблиц
Определяет тип и структуру каждой таблицы для правильного парсинга
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import json
import re
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ColumnDefinition:
    """Определение колонки"""
    index: int
    letter: str
    name: str
    type: str  # 'name', 'quantity', 'price_usd', 'price_rub', 'delivery', 'route', 'other'
    confidence: float  # 0.0-1.0
    sample_values: List[Any]

@dataclass
class TableStructure:
    """Структура таблицы"""
    file_path: str
    sheet_name: str
    header_rows: List[int]  # Строки с заголовками
    data_start_row: int     # Первая строка данных
    data_end_row: int       # Последняя строка данных  
    columns: Dict[str, ColumnDefinition]
    table_type: str         # 'standard', 'multi_price', 'vertical', 'unknown'
    confidence: float       # Общая уверенность в определении
    issues: List[str]       # Найденные проблемы

class TableStructureAnalyzer:
    """Анализатор структуры таблиц"""
    
    def __init__(self):
        # Шаблоны заголовков для разных типов колонок
        self.column_patterns = {
            'name': [
                r'наименование',
                r'название',
                r'товар',
                r'продукт',
                r'name',
                r'product',
                r'item'
            ],
            'quantity': [
                r'тираж.*шт',
                r'количество.*шт',
                r'qty',
                r'quantity',
                r'^тираж$',
                r'^количество$',
                r'шт\s*$'  # Только если в конце строки
            ],
            'price_usd': [
                r'цена.*\$',
                r'price.*\$',
                r'цена.*usd',
                r'стоимость.*\$',
                r'доллар',
                r'dollar'
            ],
            'price_rub': [
                r'цена.*руб',
                r'price.*руб',
                r'цена.*₽',
                r'стоимость.*руб',
                r'рубл'
            ],
            'delivery': [
                r'срок.*доставк',
                r'доставк.*срок',
                r'delivery.*time',
                r'срок.*к\.д',
                r'дн\.',
                r'days'
            ],
            'route': [
                r'маршрут',
                r'route',
                r'доставка.*жд',
                r'доставка.*авиа',
                r'железн.*дорог',
                r'авиа'
            ]
        }
        
        # Известные структуры таблиц
        self.known_structures = []
        self.load_known_structures()
    
    def analyze_file_structure(self, file_path: str) -> TableStructure:
        """Анализирует структуру Excel файла"""
        
        logger.info(f"📊 Анализ структуры: {Path(file_path).name}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Анализируем заголовки
            header_info = self._analyze_headers(ws)
            
            # Определяем колонки
            columns = self._detect_columns(ws, header_info['header_rows'])
            
            # Определяем границы данных
            data_bounds = self._find_data_bounds(ws, header_info['data_start_row'])
            
            # Определяем тип таблицы
            table_type, confidence = self._classify_table_type(columns, ws)
            
            # Выявляем проблемы
            issues = self._detect_issues(columns, ws, data_bounds)
            
            structure = TableStructure(
                file_path=file_path,
                sheet_name=ws.title,
                header_rows=header_info['header_rows'],
                data_start_row=data_bounds['start'],
                data_end_row=data_bounds['end'],
                columns=columns,
                table_type=table_type,
                confidence=confidence,
                issues=issues
            )
            
            wb.close()
            
            logger.info(f"   Тип: {table_type} (уверенность: {confidence:.2f})")
            logger.info(f"   Колонки: {list(columns.keys())}")
            if issues:
                logger.warning(f"   Проблемы: {issues}")
            
            return structure
            
        except Exception as e:
            logger.error(f"Ошибка анализа {file_path}: {e}")
            return TableStructure(
                file_path=file_path,
                sheet_name="unknown",
                header_rows=[],
                data_start_row=1,
                data_end_row=1,
                columns={},
                table_type="error",
                confidence=0.0,
                issues=[f"Ошибка анализа: {str(e)}"]
            )
    
    def _analyze_headers(self, ws) -> Dict[str, Any]:
        """Анализирует заголовки таблицы"""
        
        header_rows = []
        data_start_row = 1
        
        # Ищем заголовки в первых 10 строках
        for row in range(1, min(11, ws.max_row + 1)):
            text_count = 0
            number_count = 0
            
            for col in range(1, min(21, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    if isinstance(cell_value, str) and len(cell_value.strip()) > 2:
                        text_count += 1
                    elif isinstance(cell_value, (int, float)):
                        number_count += 1
            
            # Строка с заголовками: много текста, мало чисел
            if text_count >= 3 and text_count > number_count:
                header_rows.append(row)
            # Строка с данными: числа преобладают
            elif number_count > text_count and text_count > 0:
                if not data_start_row or row < data_start_row:
                    data_start_row = row
                break
        
        if not header_rows:
            header_rows = [1, 2, 3]  # По умолчанию первые 3 строки
        
        if data_start_row <= max(header_rows, default=0):
            data_start_row = max(header_rows, default=3) + 1
        
        return {
            'header_rows': header_rows,
            'data_start_row': data_start_row
        }
    
    def _detect_columns(self, ws, header_rows: List[int]) -> Dict[str, ColumnDefinition]:
        """Определяет колонки по заголовкам"""
        
        columns = {}
        
        # Анализируем каждую колонку
        for col in range(1, min(21, ws.max_column + 1)):
            col_letter = get_column_letter(col)
            
            # Собираем заголовки из всех строк заголовков
            headers = []
            for row in header_rows:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            combined_header = ' '.join(headers).lower()
            
            # Определяем тип колонки и уверенность
            col_type, confidence = self._classify_column(combined_header)
            
            if col_type != 'other' and confidence > 0.5:
                # Собираем примеры значений
                sample_values = []
                for row in range(min(header_rows) + 1, min(ws.max_row + 1, min(header_rows) + 21)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        sample_values.append(cell_value)
                        if len(sample_values) >= 10:
                            break
                
                columns[col_type] = ColumnDefinition(
                    index=col,
                    letter=col_letter,
                    name=combined_header,
                    type=col_type,
                    confidence=confidence,
                    sample_values=sample_values[:5]  # Первые 5 примеров
                )
        
        return columns
    
    def _classify_column(self, header_text: str) -> Tuple[str, float]:
        """Классифицирует тип колонки по заголовку"""
        
        # Специальные исключения для избежания конфликтов
        if any(word in header_text for word in ['цена', 'price']) and 'шт' in header_text:
            if any(word in header_text for word in ['$', 'usd', 'долл']):
                return 'price_usd', 0.9
            elif any(word in header_text for word in ['руб', '₽', 'rub']):
                return 'price_rub', 0.9
        
        # Обычная классификация
        for col_type, patterns in self.column_patterns.items():
            for pattern in patterns:
                if re.search(pattern, header_text, re.IGNORECASE):
                    confidence = 0.8
                    
                    # Повышаем уверенность для точных совпадений
                    if pattern.strip('$^') == header_text.strip():
                        confidence = 0.95
                    
                    return col_type, confidence
        
        return 'other', 0.0
    
    def _find_data_bounds(self, ws, start_row: int) -> Dict[str, int]:
        """Находит границы данных в таблице"""
        
        end_row = start_row
        
        # Ищем последнюю строку с данными
        for row in range(start_row, ws.max_row + 1):
            non_empty_cells = 0
            for col in range(1, min(11, ws.max_column + 1)):
                if ws.cell(row=row, column=col).value is not None:
                    non_empty_cells += 1
            
            if non_empty_cells >= 2:  # Строка с данными
                end_row = row
            elif row - end_row > 5:  # 5+ пустых строк подряд
                break
        
        return {'start': start_row, 'end': end_row}
    
    def _classify_table_type(self, columns: Dict[str, ColumnDefinition], ws) -> Tuple[str, float]:
        """Определяет тип таблицы"""
        
        has_name = 'name' in columns
        has_quantity = 'quantity' in columns  
        has_price_usd = 'price_usd' in columns
        has_price_rub = 'price_rub' in columns
        
        confidence = 0.0
        
        if has_name and has_quantity and (has_price_usd or has_price_rub):
            if has_price_usd and has_price_rub:
                table_type = "multi_currency"
                confidence = 0.9
            else:
                table_type = "standard"
                confidence = 0.8
        elif has_name and (has_price_usd or has_price_rub):
            table_type = "simple"
            confidence = 0.6
        elif has_name:
            table_type = "basic"
            confidence = 0.4
        else:
            table_type = "unknown"
            confidence = 0.1
        
        return table_type, confidence
    
    def _detect_issues(self, columns: Dict[str, ColumnDefinition], ws, data_bounds: Dict[str, int]) -> List[str]:
        """Выявляет потенциальные проблемы"""
        
        issues = []
        
        # Проверяем наличие ключевых колонок
        if 'name' not in columns:
            issues.append("Не найдена колонка с названиями товаров")
        
        if 'quantity' not in columns and 'price_usd' not in columns and 'price_rub' not in columns:
            issues.append("Не найдены колонки с ценами или тиражами")
        
        # Проверяем качество данных
        for col_type, col_def in columns.items():
            if col_type == 'quantity':
                # Проверяем что в колонке quantity действительно числа-тиражи
                numeric_count = 0
                for value in col_def.sample_values:
                    if isinstance(value, (int, float)) and 10 <= value <= 100000:
                        numeric_count += 1
                
                if numeric_count / len(col_def.sample_values) < 0.5:
                    issues.append(f"Колонка quantity содержит подозрительные значения: {col_def.sample_values}")
        
        # Проверяем размер данных
        data_rows = data_bounds['end'] - data_bounds['start'] + 1
        if data_rows < 3:
            issues.append(f"Слишком мало строк данных: {data_rows}")
        
        return issues
    
    def save_structure(self, structure: TableStructure, output_dir: str = "analysis_results"):
        """Сохраняет структуру таблицы в JSON"""
        
        Path(output_dir).mkdir(exist_ok=True)
        
        filename = f"{Path(structure.file_path).stem}_structure.json"
        output_path = Path(output_dir) / filename
        
        # Конвертируем в словарь для JSON
        data = asdict(structure)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Структура сохранена: {output_path}")
    
    def load_known_structures(self):
        """Загружает известные структуры таблиц"""
        # В будущем здесь будет загрузка из файлов
        pass

if __name__ == "__main__":
    analyzer = TableStructureAnalyzer()
    
    # Тестируем на проблемной таблице
    test_file = "/Users/bakirovresad/Downloads/Reshad 1/projects/commercial_proposals_parser_clean/storage/excel_files/sheet_1nav9w2d_public.xlsx"
    
    if Path(test_file).exists():
        structure = analyzer.analyze_file_structure(test_file)
        analyzer.save_structure(structure)
        
        print(f"\n📊 СТРУКТУРА ТАБЛИЦЫ:")
        print(f"Тип: {structure.table_type}")
        print(f"Уверенность: {structure.confidence:.2f}")
        print(f"Колонки:")
        for col_type, col_def in structure.columns.items():
            print(f"  {col_type}: {col_def.letter} '{col_def.name}' (уверенность: {col_def.confidence:.2f})")
        
        if structure.issues:
            print(f"Проблемы: {structure.issues}")
    else:
        print("❌ Тестовый файл не найден")


