#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Тестирование улучшенной нормализации на нескольких разноплановых таблицах
"""

import os
import sys
import logging
from pathlib import Path

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from table_normalizer import TableNormalizer

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('test_normalization.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

def test_multiple_tables():
    """Тестируем нормализацию на разных таблицах"""
    
    # Список таблиц для тестирования
    test_files = [
        'storage/excel_files/original_sheet.xlsx',
        'storage/excel_files/google_sheet_20250923_100121.xlsx', 
        'storage/excel_files/google_sheet_20250923_095052.xlsx',
        'storage/excel_files/google_sheet_20250923_154510.xlsx',
        'storage/excel_files/google_sheet_20250923_152539.xlsx'
    ]
    
    normalizer = TableNormalizer()
    
    print("🧪 ТЕСТИРОВАНИЕ НОРМАЛИЗАЦИИ НА МНОЖЕСТВЕННЫХ ТАБЛИЦАХ")
    print("=" * 90)
    
    results = []
    
    for i, file_path in enumerate(test_files, 1):
        if not os.path.exists(file_path):
            logger.warning(f"Файл не найден: {file_path}")
            continue
            
        print(f"\n📊 ТЕСТ {i}: {os.path.basename(file_path)}")
        print("-" * 60)
        
        try:
            # Нормализуем таблицу
            normalized_path = normalizer.normalize_table(file_path)
            
            if normalized_path:
                # Анализируем результат
                result = analyze_normalized_table(normalized_path, file_path)
                results.append({
                    'original': file_path,
                    'normalized': normalized_path,
                    'result': result
                })
                
                print(f"✅ Успешно нормализовано → {normalized_path}")
                print_result_summary(result)
                
            else:
                print("❌ Ошибка нормализации")
                results.append({
                    'original': file_path,
                    'normalized': None,
                    'result': None
                })
                
        except Exception as e:
            logger.error(f"Ошибка при нормализации {file_path}: {e}")
            print(f"❌ Ошибка: {e}")
    
    # Общий отчет
    print("\n" + "=" * 90)
    print("📋 ИТОГОВЫЙ ОТЧЕТ НОРМАЛИЗАЦИИ")
    print("=" * 90)
    
    successful = sum(1 for r in results if r['result'] is not None)
    total = len(results)
    
    print(f"✅ Успешно нормализовано: {successful}/{total} таблиц")
    
    if successful > 0:
        print(f"\n📊 СТАТИСТИКА ПО УСПЕШНЫМ ТАБЛИЦАМ:")
        
        total_rows = sum(r['result']['data_rows'] for r in results if r['result'])
        total_columns = sum(r['result']['normalized_columns'] for r in results if r['result'])
        
        print(f"   • Общих строк данных: {total_rows}")
        print(f"   • Средних колонок на таблицу: {total_columns/successful:.1f}")
        
        # Популярные маршруты
        all_routes = []
        for r in results:
            if r['result'] and r['result']['routes']:
                all_routes.extend(r['result']['routes'])
        
        if all_routes:
            from collections import Counter
            route_counts = Counter(all_routes)
            print(f"\n🚚 НАЙДЕННЫЕ МАРШРУТЫ:")
            for route, count in route_counts.most_common():
                print(f"   • {route}: {count} таблиц")
    
    return results

def analyze_normalized_table(normalized_path, original_path):
    """Анализирует результат нормализации"""
    
    try:
        import openpyxl
        
        # Загружаем нормализованную таблицу
        wb_norm = openpyxl.load_workbook(normalized_path, data_only=True)
        ws_norm = wb_norm.active
        
        # Анализируем структуру
        headers = []
        for col in range(1, ws_norm.max_column + 1):
            header = ws_norm.cell(row=2, column=col).value
            if header:
                headers.append(str(header).strip())
        
        # Считаем данные
        data_rows = 0
        for row in range(3, ws_norm.max_row + 1):
            has_data = False
            for col in range(1, min(6, ws_norm.max_column + 1)):  # Проверяем первые 5 колонок
                if ws_norm.cell(row=row, column=col).value:
                    has_data = True
                    break
            if has_data:
                data_rows += 1
        
        # Определяем найденные маршруты
        routes = []
        for header in headers:
            if 'АВИА' in header:
                routes.append('АВИА')
            elif 'ЖД' in header:
                routes.append('ЖД')
            elif 'ОБРАЗЕЦ' in header:
                routes.append('ОБРАЗЕЦ')
        
        routes = list(set(routes))  # Уникальные
        
        # Проверяем качество нормализации
        has_names = any('Наименование' in h for h in headers)
        has_prices = any('Цена' in h for h in headers)
        has_quantities = any('Тираж' in h for h in headers)
        
        wb_norm.close()
        
        return {
            'normalized_columns': len(headers),
            'data_rows': data_rows,
            'routes': routes,
            'has_names': has_names,
            'has_prices': has_prices,
            'has_quantities': has_quantities,
            'headers': headers[:10]  # Первые 10 заголовков
        }
        
    except Exception as e:
        logger.error(f"Ошибка анализа {normalized_path}: {e}")
        return None

def print_result_summary(result):
    """Выводит краткую сводку результата"""
    
    if not result:
        return
        
    print(f"   📈 Колонок: {result['normalized_columns']}")
    print(f"   📊 Строк данных: {result['data_rows']}")
    
    if result['routes']:
        print(f"   🚚 Маршруты: {', '.join(result['routes'])}")
    
    quality_scores = []
    if result['has_names']:
        quality_scores.append("Названия ✅")
    if result['has_prices']: 
        quality_scores.append("Цены ✅")
    if result['has_quantities']:
        quality_scores.append("Тиражи ✅")
        
    if quality_scores:
        print(f"   🎯 Качество: {' | '.join(quality_scores)}")

if __name__ == "__main__":
    test_multiple_tables()
