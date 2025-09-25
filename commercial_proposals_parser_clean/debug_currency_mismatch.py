#!/usr/bin/env python3
"""
Отладка несоответствия курса валют в товарах
"""

import openpyxl
from pathlib import Path
from database.manager_v4 import DatabaseManager  
from database.models_v4 import Product, SheetMetadata

def debug_currency_for_products(product_ids):
    """Отладка валют для конкретных товаров"""
    
    session = DatabaseManager.get_session()
    
    try:
        for product_id in product_ids:
            product = session.query(Product).get(product_id)
            if not product:
                print(f"❌ Товар ID {product_id} не найден")
                continue
                
            print(f"\n🔍 ТОВАР ID {product_id}: {product.name}")
            print("=" * 50)
            
            # Получаем метаданные таблицы
            sheet = session.query(SheetMetadata).get(product.sheet_id)
            if not sheet or not sheet.local_file_path:
                print(f"❌ Файл не найден для таблицы {product.sheet_id}")
                continue
            
            file_path = sheet.local_file_path
            print(f"📁 Файл: {file_path}")
            print(f"📋 Строка товара: {product.start_row}")
            
            # Открываем Excel файл
            if not Path(file_path).exists():
                print(f"❌ Файл не существует: {file_path}")
                continue
                
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                print(f"\n📊 ДАННЫЕ ИЗ EXCEL ФАЙЛА:")
                
                # Показываем заголовки (первые 3 строки)
                print("🔤 ЗАГОЛОВКИ:")
                for row_num in range(1, 4):
                    row_data = []
                    for col in range(1, 16):  # Первые 15 колонок
                        cell_value = ws.cell(row=row_num, column=col).value
                        if cell_value:
                            col_letter = openpyxl.utils.get_column_letter(col)
                            row_data.append(f"{col_letter}: {str(cell_value)[:20]}")
                    if row_data:
                        print(f"   Строка {row_num}: {' | '.join(row_data[:5])}")  # Первые 5 колонок
                
                # Показываем данные товара
                print(f"\n📦 ДАННЫЕ ТОВАРА (строка {product.start_row}):")
                row_data = []
                for col in range(1, 16):  # Первые 15 колонок
                    cell_value = ws.cell(row=product.start_row, column=col).value
                    col_letter = openpyxl.utils.get_column_letter(col)
                    row_data.append(f"{col_letter}: {cell_value}")
                
                # Выводим по 3 колонки в строке для читаемости
                for i in range(0, len(row_data), 3):
                    print(f"   {' | '.join(row_data[i:i+3])}")
                
                # Ищем колонки с ценами
                print(f"\n💰 АНАЛИЗ ЦЕНОВЫХ КОЛОНОК:")
                for col in range(1, 16):
                    # Проверяем заголовки
                    headers = []
                    for row in range(1, 4):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            headers.append(str(cell_value).lower())
                    
                    combined_header = ' '.join(headers)
                    col_letter = openpyxl.utils.get_column_letter(col)
                    
                    # Определяем является ли колонка ценовой
                    is_usd = any(word in combined_header for word in ['$', 'usd', 'долл']) and 'цена' in combined_header
                    is_rub = any(word in combined_header for word in ['руб', '₽', 'rub']) and 'цена' in combined_header
                    
                    if is_usd or is_rub:
                        value = ws.cell(row=product.start_row, column=col).value
                        currency = "USD" if is_usd else "RUB"
                        print(f"   {col_letter} ({currency}): {combined_header} = {value}")
                
                wb.close()
                
            except Exception as e:
                print(f"❌ Ошибка чтения Excel: {e}")
    
    finally:
        session.close()

if __name__ == "__main__":
    # Отлаживаем товары с неправильными курсами
    debug_currency_for_products([104, 114])

