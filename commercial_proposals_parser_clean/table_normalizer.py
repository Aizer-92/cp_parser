#!/usr/bin/env python3
"""
Нормализация Excel таблиц перед парсингом
Приводит все таблицы к единому стандартному формату
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple, Optional
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TableNormalizer:
    """Нормализатор таблиц к стандартному формату"""
    
    def __init__(self):
        # Словарь переводов заголовков
        self.header_translations = {
            # Основные поля товара
            'name': 'Наименование товара',
            'product name': 'Наименование товара', 
            'product': 'Наименование товара',
            'товар': 'Наименование товара',
            'наименование': 'Наименование товара',
            'артикул': 'Артикул',
            'article': 'Артикул',
            'sku': 'Артикул',
            
            # Описание и характеристики
            'description': 'Описание',
            'описание': 'Описание', 
            'характеристики': 'Описание',
            'material': 'Материал',
            'материал': 'Материал',
            'size': 'Размер',
            'размер': 'Размер',
            'color': 'Цвет',
            'цвет': 'Цвет',
            'packaging': 'Упаковка',
            'упаковка': 'Упаковка',
            
            # Тираж/количество
            'quantity': 'Тираж, шт',
            'qty': 'Тираж, шт',
            'тираж': 'Тираж, шт',
            'количество': 'Тираж, шт',
            'кол-во': 'Тираж, шт',
            'шт': 'Тираж, шт',
            
            # Цены USD
            'price usd': 'Цена за шт., $',
            'price $': 'Цена за шт., $',
            'usd': 'Цена за шт., $',
            'dollar': 'Цена за шт., $',
            'цена $': 'Цена за шт., $',
            'цена usd': 'Цена за шт., $',
            'цена долл': 'Цена за шт., $',
            
            # Цены RUB
            'price rub': 'Цена за шт., ₽',
            'price руб': 'Цена за шт., ₽', 
            'rub': 'Цена за шт., ₽',
            'руб': 'Цена за шт., ₽',
            'рубль': 'Цена за шт., ₽',
            'цена руб': 'Цена за шт., ₽',
            'цена ₽': 'Цена за шт., ₽',
            
            # Сроки доставки
            'delivery time': 'Срок доставки, дн.',
            'срок': 'Срок доставки, дн.',
            'дни': 'Срок доставки, дн.',
            'время': 'Срок доставки, дн.',
            'delivery': 'Срок доставки, дн.',
            
            # Маршруты доставки
            'стандарт': 'СТАНДАРТ',
            'standard': 'СТАНДАРТ', 
            'жд': 'ЖД',
            'railway': 'ЖД',
            'train': 'ЖД',
            'авиа': 'АВИА',
            'air': 'АВИА',
            'plane': 'АВИА',
            'aviation': 'АВИА',
            'образец': 'ОБРАЗЕЦ',
            'sample': 'ОБРАЗЕЦ'
        }
        
        # Стандартные колонки для каждого маршрута
        self.route_columns = {
            'СТАНДАРТ': ['Тираж (СТАНДАРТ), шт', 'Цена (СТАНДАРТ), $', 'Цена (СТАНДАРТ), ₽', 'Срок (СТАНДАРТ), дн.'],
            'ЖД': ['Тираж (ЖД), шт', 'Цена (ЖД), $', 'Цена (ЖД), ₽', 'Срок (ЖД), дн.'],
            'АВИА': ['Тираж (АВИА), шт', 'Цена (АВИА), $', 'Цена (АВИА), ₽', 'Срок (АВИА), дн.'],
            'ОБРАЗЕЦ': ['Тираж (ОБРАЗЕЦ), шт', 'Цена (ОБРАЗЕЦ), $', 'Цена (ОБРАЗЕЦ), ₽', 'Срок (ОБРАЗЕЦ), дн.']
        }
        
        # Базовые колонки товара
        self.base_columns = [
            'Наименование товара',
            'Артикул', 
            'Описание',
            'Материал',
            'Размер', 
            'Цвет',
            'Упаковка'
        ]

    def normalize_table(self, file_path: str, output_path: str = None) -> str:
        """Нормализует одну таблицу"""
        
        if not output_path:
            output_path = file_path.replace('.xlsx', '_normalized.xlsx')
        
        logger.info(f"🔧 Нормализация таблицы: {file_path}")
        
        # Загружаем таблицу
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 1. Анализируем текущую структуру
        current_structure = self._analyze_current_structure(ws)
        logger.info(f"Найдено маршрутов: {list(current_structure['routes'].keys())}")
        logger.info(f"Найдено базовых колонок: {list(current_structure['base_columns'].keys())}")
        
        # 2. Создаем новую нормализованную таблицу
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "Нормализованная таблица"
        
        # 3. Создаем заголовки
        normalized_headers = self._create_normalized_headers(current_structure)
        self._write_headers(new_ws, normalized_headers)
        
        # 4. Копируем данные с преобразованием
        self._copy_normalized_data(ws, new_ws, current_structure, normalized_headers)
        
        # 5. Применяем форматирование
        self._apply_formatting(new_ws, normalized_headers)
        
        # Сохраняем результат
        new_wb.save(output_path)
        wb.close()
        new_wb.close()
        
        logger.info(f"✅ Нормализованная таблица сохранена: {output_path}")
        return output_path

    def _analyze_current_structure(self, ws) -> Dict:
        """Анализирует текущую структуру таблицы"""
        
        structure = {
            'base_columns': {},  # {тип: колонка}
            'routes': {},        # {маршрут: {тип: колонка}}
            'header_rows': [],   # строки с заголовками
            'data_start_row': None
        }
        
        # Ищем заголовки в первых 5 строках
        for row in range(1, 6):
            has_headers = False
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    header_text = str(cell.value).lower().strip()
                    if len(header_text) > 2:  # Минимальная длина заголовка
                        has_headers = True
            
            if has_headers:
                structure['header_rows'].append(row)
        
        # Определяем начало данных (обычно это строка 4 для таких таблиц)
        structure['data_start_row'] = 4 if structure['header_rows'] else 3
        
        # Анализируем каждую колонку
        for col in range(1, min(20, ws.max_column + 1)):
            
            # Собираем все заголовки колонки
            column_headers = []
            for row in structure['header_rows']:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    column_headers.append(str(cell_value).lower().strip())
            
            combined_header = ' '.join(column_headers)
            
            if not combined_header:
                continue
            
            # Получаем контекст соседних колонок для лучшей классификации
            context_headers = []
            for ctx_col in range(max(1, col-2), min(ws.max_column + 1, col+3)):
                if ctx_col != col:  # Исключаем текущую колонку
                    ctx_combined = []
                    for row in structure['header_rows']:
                        cell_value = ws.cell(row=row, column=ctx_col).value
                        if cell_value:
                            ctx_combined.append(str(cell_value).lower().strip())
                    if ctx_combined:
                        context_headers.append(' '.join(ctx_combined))
            
            # Определяем тип колонки и маршрут с учетом контекста
            column_type, route = self._classify_column(combined_header, context_headers)
            
            if column_type and route:
                # Маршрут-специфичная колонка
                if route not in structure['routes']:
                    structure['routes'][route] = {}
                structure['routes'][route][column_type] = col
                
            elif column_type:
                # Базовая колонка товара
                structure['base_columns'][column_type] = col
            
            logger.info(f"Колонка {get_column_letter(col)}: '{combined_header}' → {column_type} ({route or 'базовая'})")
        
        return structure

    def _classify_column(self, header_text: str, context_headers: List[str] = None) -> Tuple[Optional[str], Optional[str]]:
        """Классифицирует колонку по заголовку с учетом контекста соседних колонок"""
        
        if context_headers is None:
            context_headers = []
        
        # УЛУЧШЕННОЕ определение маршрута с учетом контекста
        route = None
        route_mappings = {
            'АВИА': ['авиа', 'авто', 'air', 'aviation', 'avia', 'самолет', 'дубай', 'dubai'],  # Добавляем дубай
            'ЖД': ['жд', 'жел', 'поезд', 'railway', 'train', 'железн', 'контейнер'],
            'ОБРАЗЕЦ': ['образец', 'sample', 'пробн', 'проба']  # НЕ добавляем общие слова!
        }
        
        for route_name, keywords in route_mappings.items():
            if any(kw in header_text for kw in keywords):
                route = route_name
                break
                
        # Если маршрут не найден в текущей колонке, ищем в контексте
        if not route:
            for ctx_header in context_headers:
                for route_name, keywords in route_mappings.items():
                    if any(kw in ctx_header for kw in keywords):
                        # Проверяем что это соседняя колонка того же типа
                        if any(price_word in header_text for price_word in ['цена', 'price', '$', '₽', 'руб']):
                            route = route_name
                            break
                if route:
                    break
        
        # УЛУЧШЕННОЕ определение типа колонки
        column_type = None
        
        # Цены: сначала точное совпадение, потом общие слова
        if any(word in header_text for word in ['$', 'usd', 'долл']):
            if any(word in header_text for word in ['цена', 'price']):
                column_type = 'price_usd'
        elif any(word in header_text for word in ['₽', 'руб', 'rub']):
            if any(word in header_text for word in ['цена', 'price']):
                column_type = 'price_rub'
        elif 'цена' in header_text or 'price' in header_text:
            # Определяем валюту из контекста или по умолчанию
            is_rub_context = any('₽' in ctx or 'руб' in ctx for ctx in context_headers)
            column_type = 'price_rub' if is_rub_context else 'price_usd'
        
        # Тираж/количество (более строгая проверка)
        elif any(word in header_text for word in ['тираж', 'количество', 'кол-во', 'quantity', 'qty']):
            # Убеждаемся что это не цена с "шт"
            if not any(price_word in header_text for price_word in ['цена', 'price', '$', '₽', 'руб']):
                # КРИТИЧНО: Если это просто "Тираж, шт" - это БАЗОВОЕ поле, НЕ образец!
                if not route:  # Если нет явного указания на маршрут - это базовый тираж
                    column_type = 'quantity'
                    route = None  # Сбрасываем маршрут
                else:
                    column_type = 'quantity'  # С маршрутом
        
        # Сроки доставки
        elif any(word in header_text for word in ['срок', 'дн', 'дней', 'время', 'delivery', 'time']):
            column_type = 'delivery_time'
        
        # ПРИОРИТЕТ: Базовые поля товара (точные названия имеют высший приоритет)
        if any(word in header_text for word in ['наименование', 'товар', 'name', 'product']):
            column_type = 'name'
            route = None  # Сбрасываем маршрут - это базовое поле!
        elif any(word in header_text for word in ['артикул', 'article', 'sku']):
            column_type = 'article'  
            route = None
        elif any(word in header_text for word in ['описание', 'description', 'характеристики']):
            column_type = 'description'
            route = None
        elif any(word in header_text for word in ['материал', 'material']):
            column_type = 'material'
            route = None
        elif any(word in header_text for word in ['размер', 'size']):
            column_type = 'size'
            route = None
        elif any(word in header_text for word in ['цвет', 'color']):
            column_type = 'color'
            route = None
        elif any(word in header_text for word in ['упаковка', 'packaging']):
            column_type = 'packaging'
            route = None
        
        # Если нет точного совпадения базового поля и есть маршрут, проверяем другие типы
        elif route:
            pass  # Эти проверки перенесены выше
        
        return column_type, route

    def _create_normalized_headers(self, structure: Dict) -> List[str]:
        """Создает список нормализованных заголовков"""
        
        headers = []
        
        # Добавляем базовые колонки
        for base_col in self.base_columns:
            headers.append(base_col)
        
        # Добавляем колонки для каждого найденного маршрута
        for route in sorted(structure['routes'].keys()):
            if route in self.route_columns:
                headers.extend(self.route_columns[route])
        
        return headers
    
    def _save_row_ranges_metadata(self, normalized_path: str, structure: Dict):
        """Сохраняет метаданные с диапазонами строк для привязки изображений"""
        
        import json
        
        metadata = {
            'normalized_file': os.path.basename(normalized_path),
            'original_file': structure.get('original_file', ''),
            'data_ranges': {
                'original_start_row': structure['data_start_row'],
                'original_end_row': structure.get('original_end_row', 0),
                'normalized_start_row': structure['normalized_data_range']['start_row'],
                'normalized_end_row': structure['normalized_data_range']['end_row'],
                'total_products': structure['normalized_data_range']['total_products']
            },
            'row_mappings': structure['row_mappings'],  # оригинальная_строка → нормализованная_строка
            'found_routes': list(structure['routes'].keys()),
            'base_columns': list(structure['base_columns'].keys()),
            'created_at': str(datetime.now())
        }
        
        # Сохраняем в файл рядом с нормализованной таблицей
        metadata_path = normalized_path.replace('.xlsx', '_metadata.json')
        
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
            
        logger.info(f"📋 Метаданные сохранены: {metadata_path}")
        
        return metadata_path

    def _write_headers(self, ws, headers: List[str]):
        """Записывает заголовки в таблицу"""
        
        # Заголовок таблицы  
        ws.cell(row=1, column=1, value="НОРМАЛИЗОВАННАЯ ТАБЛИЦА КОММЕРЧЕСКОГО ПРЕДЛОЖЕНИЯ")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        
        # Основные заголовки
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)

    def _copy_normalized_data(self, source_ws, target_ws, structure: Dict, headers: List[str]):
        """Копирует данные с нормализацией"""
        
        data_start = structure['data_start_row']
        target_row = 3  # Начинаем с 3-й строки (после заголовков)
        
        # Создаем маппинг заголовков к колонкам
        header_to_col = {header: col for col, header in enumerate(headers, 1)}
        
        # Сохраняем информацию о диапазонах строк для привязки изображений
        structure['row_mappings'] = {}
        
        # Обрабатываем каждую строку данных
        logger.info(f"Копируем данные с строки {data_start} до {source_ws.max_row}")
        
        for row in range(data_start, source_ws.max_row + 1):
            
            # Проверяем есть ли данные в строке (проверяем key колонки)
            key_columns_to_check = [col for col in structure['base_columns'].values()] + \
                                 [col for route_data in structure['routes'].values() 
                                  for col in route_data.values()]
            
            has_data = False
            for col in key_columns_to_check[:5]:  # Проверяем первые 5 важных колонок
                cell_value = source_ws.cell(row=row, column=col).value
                if cell_value and str(cell_value).strip():
                    has_data = True
                    break
            
            if not has_data:
                continue
            
            logger.debug(f"Обрабатываем строку {row} → нормализованная строка {target_row}")
            
            # Сохраняем маппинг строк для изображений (оригинальная строка → нормализованная строка)
            structure['row_mappings'][row] = target_row
            
            # Копируем базовые поля
            for field_type, source_col in structure['base_columns'].items():
                if field_type == 'name':
                    target_col = header_to_col.get('Наименование товара')
                elif field_type == 'article':
                    target_col = header_to_col.get('Артикул')
                elif field_type == 'description':
                    target_col = header_to_col.get('Описание')
                elif field_type == 'material':
                    target_col = header_to_col.get('Материал')
                elif field_type == 'size':
                    target_col = header_to_col.get('Размер')
                elif field_type == 'color':
                    target_col = header_to_col.get('Цвет')
                elif field_type == 'packaging':
                    target_col = header_to_col.get('Упаковка')
                else:
                    continue
                
                if target_col:
                    value = source_ws.cell(row=row, column=source_col).value
                    if value is not None:  # Копируем только непустые значения
                        target_ws.cell(row=target_row, column=target_col, value=value)
            
            # Копируем маршрут-специфичные данные
            for route, route_fields in structure['routes'].items():
                for field_type, source_col in route_fields.items():
                    
                    if field_type == 'quantity':
                        target_col = header_to_col.get(f'Тираж ({route}), шт')
                    elif field_type == 'price_usd':
                        target_col = header_to_col.get(f'Цена ({route}), $')
                    elif field_type == 'price_rub':
                        target_col = header_to_col.get(f'Цена ({route}), ₽')
                    elif field_type == 'delivery_time':
                        target_col = header_to_col.get(f'Срок ({route}), дн.')
                    else:
                        continue
                    
                    if target_col:
                        value = source_ws.cell(row=row, column=source_col).value
                        if value is not None:  # Копируем только непустые значения
                            target_ws.cell(row=target_row, column=target_col, value=value)
                            logger.debug(f"Скопировано {field_type} для {route}: {value} из колонки {source_col} в {target_col}")
            
            # ДОПОЛНИТЕЛЬНО: Копируем базовые количества и цены если они есть  
            if 'quantity' in structure['base_columns']:
                base_quantity = source_ws.cell(row=row, column=structure['base_columns']['quantity']).value
                # Копируем базовое количество во все маршруты где нет своего
                for route in structure['routes'].keys():
                    if 'quantity' not in structure['routes'][route]:
                        target_col = header_to_col.get(f'Тираж ({route}), шт')
                        if target_col and base_quantity:
                            target_ws.cell(row=target_row, column=target_col, value=base_quantity)
            
            target_row += 1
            
        # Сохраняем общую информацию о диапазонах
        structure['normalized_data_range'] = {
            'start_row': 3,  # Начальная строка данных в нормализованной таблице
            'end_row': target_row - 1,  # Конечная строка данных
            'total_products': target_row - 3  # Количество товаров
        }

    def _apply_formatting(self, ws, headers: List[str]):
        """Применяет форматирование к таблице"""
        
        # Заголовок таблицы
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовки колонок
        for col in range(1, len(headers) + 1):
            header_cell = ws.cell(row=2, column=col)
            header_cell.font = Font(bold=True, size=11)
            header_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15


def test_normalization():
    """Тестируем нормализацию на одной таблице"""
    
    import os
    
    normalizer = TableNormalizer()
    
    # Находим тестовую таблицу
    excel_files_dir = "storage/excel_files"
    test_files = []
    
    if os.path.exists(excel_files_dir):
        for file in os.listdir(excel_files_dir):
            if file.endswith('.xlsx') and not file.startswith('~'):
                test_files.append(os.path.join(excel_files_dir, file))
    
    if not test_files:
        print("❌ Не найдено Excel файлов для тестирования")
        return
    
    # Берем файл с реальными данными
    test_file = "storage/excel_files/original_sheet.xlsx"
    print(f"🧪 ТЕСТИРУЕМ НОРМАЛИЗАЦИЮ НА ФАЙЛЕ: {test_file}")
    print("=" * 80)
    
    # Показываем оригинальную структуру
    print("\n📊 АНАЛИЗ ОРИГИНАЛЬНОЙ СТРУКТУРЫ:")
    wb = openpyxl.load_workbook(test_file, data_only=True)
    ws = wb.active
    
    print(f"Размер таблицы: {ws.max_row} строк × {ws.max_column} колонок")
    
    # Показываем заголовки
    print("\n🔍 ТЕКУЩИЕ ЗАГОЛОВКИ:")
    for row in range(1, min(4, ws.max_row + 1)):
        headers = []
        for col in range(1, min(15, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                headers.append(f"{get_column_letter(col)}={cell_value}")
        if headers:
            print(f"  Строка {row}: {' | '.join(headers)}")
    
    wb.close()
    
    # Выполняем нормализацию
    print(f"\n🔧 ВЫПОЛНЯЕМ НОРМАЛИЗАЦИЮ...")
    normalized_file = normalizer.normalize_table(test_file)
    
    # Показываем результат
    print(f"\n✅ РЕЗУЛЬТАТ НОРМАЛИЗАЦИИ:")
    print(f"Нормализованный файл: {normalized_file}")
    
    # Показываем нормализованную структуру
    wb_norm = openpyxl.load_workbook(normalized_file, data_only=True)
    ws_norm = wb_norm.active
    
    print(f"Размер нормализованной таблицы: {ws_norm.max_row} строк × {ws_norm.max_column} колонок")
    
    print("\n🎯 НОРМАЛИЗОВАННЫЕ ЗАГОЛОВКИ:")
    for col in range(1, ws_norm.max_column + 1):
        header = ws_norm.cell(row=2, column=col).value
        if header:
            print(f"  {get_column_letter(col)}: {header}")
    
    # Показываем первые 3 строки данных
    print(f"\n📋 ОБРАЗЕЦ ДАННЫХ (первые 3 строки):")
    for row in range(3, min(6, ws_norm.max_row + 1)):
        data = []
        for col in range(1, min(8, ws_norm.max_column + 1)):  # Показываем первые 7 колонок
            cell_value = ws_norm.cell(row=row, column=col).value
            data.append(str(cell_value) if cell_value else "")
        print(f"  Строка {row-2}: {' | '.join(data)}")
    
    wb_norm.close()
    
    return normalized_file

if __name__ == "__main__":
    test_normalization()
