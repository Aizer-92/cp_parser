#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ИСПРАВЛЕННЫЙ парсер для нормализованных таблиц - читает заголовки динамически
"""

import os
import sys
import json
import logging
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from database.manager_v4 import CommercialProposalsDB
from database.models_v4 import Product, PriceOffer, ProductImage, SheetMetadata
from config import DATABASE_URL_V4

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FixedNormalizedParser:
    """ИСПРАВЛЕННЫЙ парсер для нормализованных таблиц с динамическим чтением заголовков"""
    
    def __init__(self):
        self.db = CommercialProposalsDB(DATABASE_URL_V4)
    
    def parse_normalized_table(self, file_path: str, metadata_path: str) -> Tuple[int, int]:
        """Парсит одну нормализованную таблицу"""
        
        logger.info(f"📋 Парсим нормализованную таблицу: {Path(file_path).name}")
        
        # Загружаем метаданные
        try:
            with open(metadata_path, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
        except:
            metadata = {}
        
        # Открываем таблицу
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        session = self.db.get_session()
        
        try:
            # Определяем или создаем SheetMetadata
            sheet_metadata = self._get_or_create_sheet_metadata(file_path, metadata, session)
            
            # ДИНАМИЧЕСКИ читаем заголовки из строки 2
            column_mapping = self._read_headers(ws)
            
            if not column_mapping:
                logger.error(f"❌ Не удалось прочитать заголовки из {file_path}")
                return 0, 0
                
            logger.info(f"📊 Найдено колонок: {len(column_mapping)}")
            
            products_count = 0
            price_offers_count = 0
            
            # Читаем данные начиная с 3-й строки
            for row in range(3, ws.max_row + 1):
                
                # Извлекаем базовые данные товара
                product_data = self._extract_product_data(ws, row, column_mapping)
                
                if not product_data['name']:
                    continue
                
                # Создаем товар
                product = Product(
                    name=product_data['name'],
                    description=product_data['description'],
                    characteristics=product_data['characteristics'],
                    custom_design=product_data['custom_design'],
                    sheet_id=sheet_metadata.id
                )
                
                session.add(product)
                session.flush()  # Получаем ID товара
                
                products_count += 1
                
                # Извлекаем варианты цен для всех маршрутов
                price_offers = self._extract_price_offers(ws, row, product.id, column_mapping)
                
                for offer in price_offers:
                    session.add(offer)
                    price_offers_count += 1
            
            # Сохраняем изменения
            session.commit()
            wb.close()
            
            logger.info(f"✅ Обработано: {products_count} товаров, {price_offers_count} цен")
            return products_count, price_offers_count
            
        except Exception as e:
            session.rollback()
            wb.close()
            logger.error(f"❌ Ошибка парсинга {file_path}: {e}")
            return 0, 0
            
        finally:
            session.close()
    
    def _read_headers(self, ws) -> Dict[str, int]:
        """ДИНАМИЧЕСКИ читает заголовки из строки 2 и возвращает маппинг название->колонка"""
        
        column_mapping = {}
        
        # Читаем заголовки из строки 2
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=2, column=col).value
            
            if header_value:
                header = str(header_value).strip()
                
                # Классифицируем заголовки
                field_name = self._classify_header(header)
                if field_name:
                    column_mapping[field_name] = col
        
        logger.info(f"🔍 Найденные колонки: {column_mapping}")
        return column_mapping
    
    def _classify_header(self, header: str) -> Optional[str]:
        """Классифицирует заголовок и возвращает стандартное название поля"""
        
        header_lower = header.lower()
        
        # Базовые поля товара
        if 'наименование' in header_lower or header_lower == 'name':
            return 'name'
        elif 'артикул' in header_lower or header_lower == 'article':
            return 'article'
        elif 'описание' in header_lower or header_lower == 'description':
            return 'description'
        elif 'материал' in header_lower or header_lower == 'material':
            return 'material'
        elif 'размер' in header_lower or header_lower == 'size':
            return 'size'
        elif 'цвет' in header_lower or header_lower == 'color':
            return 'color'
        elif 'упаковка' in header_lower or header_lower == 'packaging':
            return 'packaging'
        
        # Поля для маршрутов (АВИА, ЖД, ОБРАЗЕЦ)
        elif 'авиа' in header_lower and 'тираж' in header_lower:
            return 'quantity_avia'
        elif 'авиа' in header_lower and ('цена' in header_lower and '$' in header):
            return 'price_avia_usd'
        elif 'авиа' in header_lower and ('цена' in header_lower and '₽' in header):
            return 'price_avia_rub'
        elif 'авиа' in header_lower and 'срок' in header_lower:
            return 'delivery_avia'
            
        elif 'жд' in header_lower and 'тираж' in header_lower:
            return 'quantity_jd'
        elif 'жд' in header_lower and ('цена' in header_lower and '$' in header):
            return 'price_jd_usd'
        elif 'жд' in header_lower and ('цена' in header_lower and '₽' in header):
            return 'price_jd_rub'
        elif 'жд' in header_lower and 'срок' in header_lower:
            return 'delivery_jd'
            
        elif 'образец' in header_lower and 'тираж' in header_lower:
            return 'quantity_sample'
        elif 'образец' in header_lower and ('цена' in header_lower and '$' in header):
            return 'price_sample_usd'
        elif 'образец' in header_lower and ('цена' in header_lower and '₽' in header):
            return 'price_sample_rub'
        elif 'образец' in header_lower and 'срок' in header_lower:
            return 'delivery_sample'
        
        return None
    
    def _extract_product_data(self, ws, row: int, column_mapping: Dict[str, int]) -> Dict:
        """Извлекает базовые данные товара из строки"""
        
        def get_cell_value(field_name: str) -> Optional[str]:
            if field_name in column_mapping:
                col = column_mapping[field_name]
                value = ws.cell(row=row, column=col).value
                return str(value).strip() if value else None
            return None
        
        # Соединяем все дополнительные характеристики
        characteristics_parts = []
        for field in ['material', 'size', 'color', 'packaging']:
            value = get_cell_value(field)
            if value:
                characteristics_parts.append(f"{field.title()}: {value}")
        
        characteristics = "; ".join(characteristics_parts) if characteristics_parts else None
        
        return {
            'name': get_cell_value('name'),
            'description': get_cell_value('description'),
            'characteristics': characteristics,
            'custom_design': None,  # Пока не используем
        }
    
    def _extract_price_offers(self, ws, row: int, product_id: int, column_mapping: Dict[str, int]) -> List[PriceOffer]:
        """Извлекает все варианты цен для товара"""
        
        offers = []
        
        # Проверяем каждый маршрут
        routes = [
            ('АВИА', 'quantity_avia', 'price_avia_usd', 'price_avia_rub', 'delivery_avia'),
            ('ЖД', 'quantity_jd', 'price_jd_usd', 'price_jd_rub', 'delivery_jd'), 
            ('ОБРАЗЕЦ', 'quantity_sample', 'price_sample_usd', 'price_sample_rub', 'delivery_sample')
        ]
        
        for route_name, qty_field, usd_field, rub_field, delivery_field in routes:
            
            # Извлекаем данные для этого маршрута
            quantity = self._get_cell_number(ws, row, column_mapping.get(qty_field))
            price_usd = self._get_cell_number(ws, row, column_mapping.get(usd_field))
            price_rub = self._get_cell_number(ws, row, column_mapping.get(rub_field))
            delivery_time = self._get_cell_text(ws, row, column_mapping.get(delivery_field))
            
            # Создаем предложение если есть хоть какие-то данные
            if quantity or price_usd or price_rub:
                
                offer = PriceOffer(
                    product_id=product_id,
                    route_name=route_name,
                    quantity=int(quantity) if quantity else None,
                    price_usd=float(price_usd) if price_usd else None,
                    price_rub=float(price_rub) if price_rub else None,
                    delivery_time=delivery_time,
                    is_sample=(route_name == 'ОБРАЗЕЦ'),
                    is_available=True
                )
                
                offers.append(offer)
        
        return offers
    
    def _get_cell_number(self, ws, row: int, col: Optional[int]) -> Optional[float]:
        """Извлекает число из ячейки"""
        
        if not col:
            return None
            
        value = ws.cell(row=row, column=col).value
        
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        # Очищаем строку
        clean_str = str(value).replace(' ', '').replace(',', '.').strip()
        
        # Убираем символы валют
        for symbol in ['$', '₽', 'руб', 'шт', 'pcs']:
            clean_str = clean_str.replace(symbol, '')
        
        try:
            return float(clean_str) if clean_str else None
        except:
            return None
    
    def _get_cell_text(self, ws, row: int, col: Optional[int]) -> Optional[str]:
        """Извлекает текст из ячейки"""
        
        if not col:
            return None
            
        value = ws.cell(row=row, column=col).value
        return str(value).strip() if value else None
    
    def _get_or_create_sheet_metadata(self, file_path: str, metadata: Dict, session) -> SheetMetadata:
        """Получает или создает SheetMetadata"""
        
        # Генерируем уникальный sheet_id из пути
        sheet_id_str = Path(file_path).stem.replace('_normalized', '')
        
        # Ищем существующий
        sheet_meta = session.query(SheetMetadata).filter_by(sheet_id=sheet_id_str).first()
        
        if not sheet_meta:
            # Создаем новый
            sheet_meta = SheetMetadata(
                sheet_url=f"file://{file_path}",
                sheet_title=Path(file_path).name,
                sheet_id=sheet_id_str,
                status='normalized_parsed',
                products_count=0,
                local_file_path=file_path
            )
            session.add(sheet_meta)
            session.flush()
        
        return sheet_meta

def clear_all_products():
    """Полностью очищает БД от товаров и цен (изображения остаются)"""
    
    print("🗑️ ПОЛНАЯ ОЧИСТКА БД")
    print("=" * 60)
    
    db = CommercialProposalsDB(DATABASE_URL_V4)
    session = db.get_session()
    
    try:
        # Считаем что удаляем
        products_count = session.query(Product).count()
        prices_count = session.query(PriceOffer).count()
        images_count = session.query(ProductImage).count()
        
        print(f"📊 К удалению:")
        print(f"   • Товаров: {products_count}")
        print(f"   • Вариантов цен: {prices_count}")
        print(f"🖼️ ИЗОБРАЖЕНИЯ ОСТАЮТСЯ: {images_count} (не трогаем)")
        
        # Удаляем только цены и товары - изображения НЕ ТРОГАЕМ!
        session.query(PriceOffer).delete()
        session.query(Product).delete()
        
        session.commit()
        
        print("✅ БД полностью очищена!")
        print("🖼️ Изображения сохранены!")
        
    except Exception as e:
        session.rollback()
        print(f"❌ Ошибка очистки: {e}")
        raise
        
    finally:
        session.close()

def parse_all_fixed():
    """Парсит все нормализованные таблицы ИСПРАВЛЕННЫМ парсером"""
    
    print("🎯 ИСПРАВЛЕННЫЙ ПАРСИНГ НОРМАЛИЗОВАННЫХ ТАБЛИЦ")
    print("=" * 80)
    
    # Находим все нормализованные файлы
    excel_dir = Path("storage/excel_files")
    normalized_files = list(excel_dir.glob("*_normalized.xlsx"))
    
    print(f"📊 Найдено нормализованных таблиц: {len(normalized_files)}")
    
    parser = FixedNormalizedParser()
    
    total_products = 0
    total_prices = 0
    success_count = 0
    error_count = 0
    
    for i, file_path in enumerate(normalized_files, 1):
        
        # Ищем файл метаданных
        metadata_path = str(file_path).replace('_normalized.xlsx', '_normalized_metadata.json')
        
        if not Path(metadata_path).exists():
            logger.warning(f"⚠️ Нет метаданных для {file_path.name}")
            metadata_path = None
        
        print(f"\n📋 [{i}/{len(normalized_files)}] {file_path.name}")
        
        try:
            products, prices = parser.parse_normalized_table(str(file_path), metadata_path or "")
            
            if products > 0:
                total_products += products
                total_prices += prices
                success_count += 1
                print(f"   ✅ {products} товаров, {prices} цен")
            else:
                error_count += 1
                print(f"   ⚠️ Пустая таблица")
                
        except Exception as e:
            error_count += 1
            logger.error(f"   ❌ Ошибка: {e}")
        
        # Прогресс каждые 25 файлов
        if i % 25 == 0:
            print(f"   📊 Прогресс: {i}/{len(normalized_files)} | Товаров: {total_products} | Цен: {total_prices}")
    
    # Итоговая статистика
    print(f"\n🎯 ИТОГИ ИСПРАВЛЕННОГО ПАРСИНГА:")
    print("=" * 80)
    print(f"✅ Успешно: {success_count}")
    print(f"❌ Ошибок: {error_count}")
    print(f"📊 Товаров создано: {total_products}")
    print(f"💰 Вариантов цен: {total_prices}")
    print(f"📋 Среднее цен на товар: {total_prices/total_products:.1f}" if total_products > 0 else "")

if __name__ == "__main__":
    
    print("🎯 ИСПРАВЛЕННОЕ ОБНОВЛЕНИЕ БД")
    print("=" * 80)
    
    # Подтверждение
    response = input("⚠️ Полностью пересоздать БД из нормализованных таблиц? (y/N): ")
    
    if response.lower() in ['y', 'yes', 'да']:
        
        # 1. Полностью очищаем БД
        clear_all_products()
        
        # 2. Парсим ИСПРАВЛЕННЫМ парсером
        parse_all_fixed()
        
        print(f"\n🎉 ОБНОВЛЕНИЕ ЗАВЕРШЕНО!")
        print("🖼️ Изображения остались с оригинальными привязками!")
        
    else:
        print("❌ Отменено пользователем")
