#!/usr/bin/env python3
"""
Исправление проблем с дублированием и неправильными тиражами
"""

from smart_adaptive_parser import SmartAdaptiveParser
from database.manager_v4 import DatabaseManager
from database.models_v4 import Product, PriceOffer, SheetMetadata
import openpyxl
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FixedSmartParser(SmartAdaptiveParser):
    """Исправленный умный парсер без дублирования"""
    
    def _extract_products_smart(self, ws, structure, columns):
        """ИСПРАВЛЕННОЕ извлечение товаров без дублирования"""
        
        products = []
        name_col = None
        quantity_col = None
        
        # Находим ПРАВИЛЬНУЮ колонку name
        for col_key, col_info in columns.items():
            if col_key == 'name' and hasattr(col_info, 'index'):
                name_col = col_info.index
                break
        
        # Находим ПРАВИЛЬНУЮ колонку quantity (НЕ из цен!)
        for col_key, col_info in columns.items():
            if col_key == 'quantity' and hasattr(col_info, 'index'):
                quantity_col = col_info.index
                break
        
        if not name_col:
            logger.error("Не найдена колонка с названиями")
            return products
        
        logger.info(f"Используем колонку name: {name_col} (колонка {openpyxl.utils.get_column_letter(name_col)})")
        if quantity_col:
            logger.info(f"Используем колонку quantity: {quantity_col} (колонка {openpyxl.utils.get_column_letter(quantity_col)})")
        
        # ОДНОКРАТНЫЙ проход по строкам товаров
        processed_rows = set()  # Избегаем дублирования
        
        for row in range(structure.data_start_row, structure.data_end_row + 1):
            if row in processed_rows:
                continue
                
            name_value = ws.cell(row=row, column=name_col).value
            
            if not name_value or not self._is_valid_product_name(str(name_value)):
                continue
            
            product_name = str(name_value).strip()
            
            # Извлекаем базовую информацию товара
            product_data = {
                'name': product_name,
                'row': row,
                'prices': []
            }
            
            # Правильное извлечение тиража
            base_quantity = None
            if quantity_col:
                qty_value = ws.cell(row=row, column=quantity_col).value
                if qty_value:
                    base_quantity = self._validate_quantity(qty_value)
                    if base_quantity:
                        logger.info(f"Товар '{product_name}' (строка {row}): тираж {base_quantity}")
                    else:
                        logger.warning(f"Товар '{product_name}' (строка {row}): недействительный тираж {qty_value}")
            
            if not base_quantity:
                base_quantity = 1  # По умолчанию
            
            product_data['base_quantity'] = base_quantity
            
            # Извлекаем ВСЕ ценовые предложения для этого товара
            self._extract_all_price_variants(ws, row, columns, product_data)
            
            # Добавляем товар только если есть ценовые предложения
            if product_data['prices']:
                products.append(product_data)
                processed_rows.add(row)
                logger.info(f"Добавлен товар: {product_name} с {len(product_data['prices'])} ценовыми предложениями")
        
        return products
    
    def _extract_all_price_variants(self, ws, row, columns, product_data):
        """Извлекает ВСЕ ценовые варианты для одного товара"""
        
        base_quantity = product_data.get('base_quantity', 1)
        
        # Находим все ценовые колонки и группируем по маршрутам
        price_variants = {}
        
        for col_key, col_info in columns.items():
            if not col_key.startswith('price_'):
                continue
                
            if not hasattr(col_info, 'index'):
                continue
            
            col_type = getattr(col_info, 'type', '')
            route_type = self._extract_route_from_key(col_key)
            
            # Извлекаем значение цены
            cell_value = ws.cell(row=row, column=col_info.index).value
            if not cell_value:
                continue
                
            price = self._validate_price(cell_value)
            if not price:
                continue
            
            # Группируем по маршруту
            if route_type not in price_variants:
                price_variants[route_type] = {
                    'route_name': self._get_route_display_name(route_type),
                    'quantity': base_quantity
                }
            
            # Добавляем цену
            if col_type == 'price_usd':
                price_variants[route_type]['price_usd'] = price
            elif col_type == 'price_rub':
                price_variants[route_type]['price_rub'] = price
        
        # Создаем ценовые предложения
        for route_type, variant in price_variants.items():
            if variant.get('price_usd') or variant.get('price_rub'):
                product_data['prices'].append({
                    'route_name': variant['route_name'],
                    'quantity': variant['quantity'],
                    'price_usd': variant.get('price_usd'),
                    'price_rub': variant.get('price_rub'),
                    'delivery_time': None
                })
    
    def _extract_route_from_key(self, col_key):
        """Извлекает тип маршрута из ключа колонки"""
        if '_railway' in col_key:
            return 'railway'
        elif '_air' in col_key:
            return 'air'
        elif '_sample' in col_key:
            return 'sample'
        else:
            return 'standard'
    
    def _get_route_display_name(self, route_type):
        """Возвращает читабельное название маршрута"""
        route_names = {
            'railway': 'ЖД доставка',
            'air': 'АВИА доставка', 
            'sample': 'Образец',
            'standard': 'Стандарт'
        }
        return route_names.get(route_type, route_type)

def test_fixed_parser():
    """Тестируем исправленный парсер"""
    
    print("🔧 ТЕСТИРОВАНИЕ ИСПРАВЛЕННОГО ПАРСЕРА")
    print("=" * 60)
    
    # Очищаем предыдущие результаты
    session = DatabaseManager.get_session()
    
    try:
        # Удаляем товары из проблемной таблицы
        sheet = session.query(SheetMetadata).filter(
            SheetMetadata.local_file_path.contains("sheet_1nav9w2d_public.xlsx")
        ).first()
        
        if sheet:
            # Удаляем связанные ценовые предложения и товары
            products = session.query(Product).filter(Product.sheet_id == sheet.id).all()
            for product in products:
                session.query(PriceOffer).filter(PriceOffer.product_id == product.id).delete()
            session.query(Product).filter(Product.sheet_id == sheet.id).delete()
            session.commit()
            
            print(f"✅ Очищены предыдущие результаты для {sheet.id}")
            
            # Тестируем исправленный парсер
            fixed_parser = FixedSmartParser()
            success = fixed_parser.parse_excel_smart(sheet.id)
            
            print(f"Результат: {'✅ Успешно' if success else '❌ Ошибка'}")
            
        else:
            print("❌ Тестовая таблица не найдена")
            
    finally:
        session.close()

if __name__ == "__main__":
    test_fixed_parser()
