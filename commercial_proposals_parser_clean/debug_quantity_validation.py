#!/usr/bin/env python3
"""
Отладка валидации тиража - где теряются реальные значения
"""

import openpyxl
from pathlib import Path
import sys
sys.path.append(str(Path(__file__).parent))

from scripts.complete_parsing_pipeline_v5 import EnhancedParser

def test_quantity_validation():
    """Тестируем валидацию тиража на реальных значениях"""
    
    parser = EnhancedParser()
    
    # Тестовые значения из реальной таблицы  
    test_values = [
        2130.0,   # Реальное значение из E17
        2150.0,   # Реальное значение из E19  
        "2130",   # Как строка
        "2150.0", # Как строка с .0
        3.67,     # Цена (НЕ должна проходить как тираж)
        3,        # То что сейчас в БД
        "3",      # Как строка
        None,     # Пустое значение
    ]
    
    print("🧪 ТЕСТИРОВАНИЕ ВАЛИДАЦИИ ТИРАЖА:")
    print("=" * 50)
    
    for value in test_values:
        try:
            result = parser.validate_quantity_enhanced(value)
            print(f"Значение: {value} ({type(value).__name__}) → Результат: {result}")
        except Exception as e:
            print(f"Значение: {value} ({type(value).__name__}) → ОШИБКА: {e}")

def debug_real_excel_parsing():
    """Отлаживаем что реально происходит при парсинге Excel"""
    
    file_path = "/Users/bakirovresad/Downloads/Reshad 1/projects/commercial_proposals_parser_clean/storage/excel_files/sheet_1nav9w2d_public.xlsx"
    
    if not Path(file_path).exists():
        print("❌ Файл не найден!")
        return
    
    print(f"\n📁 ОТЛАДКА РЕАЛЬНОГО ПАРСИНГА:")
    print("=" * 50)
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Проверяем конкретную строку 17 (Ёлочная игрушка)
    row = 17
    print(f"🔍 СТРОКА {row}:")
    
    # Проверяем колонки E, F, G (5, 6, 7)
    for col in range(5, 8):
        cell_value = ws.cell(row=row, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"   {col_letter}{row}: {cell_value} (тип: {type(cell_value).__name__})")
        
        # Тестируем валидацию
        parser = EnhancedParser()
        
        if col == 5:  # Колонка E - тираж
            qty_result = parser.validate_quantity_enhanced(cell_value)
            print(f"      → validate_quantity_enhanced: {qty_result}")
        
        elif col == 6:  # Колонка F - цена USD
            price_result = parser.validate_price_enhanced(cell_value)
            print(f"      → validate_price_enhanced: {price_result}")
    
    wb.close()

def debug_columns_detection():
    """Проверяем правильно ли определяются колонки"""
    
    parser = EnhancedParser()
    
    # Имитируем поиск колонок как в реальном парсере
    print(f"\n🎯 ОТЛАДКА ОПРЕДЕЛЕНИЯ КОЛОНОК:")
    print("=" * 50)
    
    # Из анализа знаем что заголовки:
    # E: "Тираж, шт" 
    # F: "Доставка ЖД" -> "Цена за шт., $"
    # G: "Цена за шт., руб"
    
    headers = {
        5: "Тираж, шт",
        6: "Цена за шт., $", 
        7: "Цена за шт., руб"
    }
    
    for col_num, header in headers.items():
        col_letter = openpyxl.utils.get_column_letter(col_num)
        print(f"Колонка {col_letter} ({col_num}): '{header}'")
        
        # Проверяем как парсер определяет тип колонки
        header_lower = header.lower()
        
        if any(word in header_lower for word in ['тираж', 'quantity', 'шт', 'pcs']):
            print(f"   → Определена как QUANTITY ✅")
        elif any(word in header_lower for word in ['price', 'цена', '$', 'usd']):
            print(f"   → Определена как PRICE_USD ✅")
        elif any(word in header_lower for word in ['руб', 'rub', '₽']):
            print(f"   → Определена как PRICE_RUB ✅")
        else:
            print(f"   → НЕ ОПРЕДЕЛЕНА ❌")

if __name__ == "__main__":
    test_quantity_validation()
    debug_real_excel_parsing()
    debug_columns_detection()


