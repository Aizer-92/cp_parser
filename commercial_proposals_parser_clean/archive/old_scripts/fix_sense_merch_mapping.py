#!/usr/bin/env python3
"""
Скрипт для исправления привязки изображений для таблицы "Мерч для Sense"
"""

import sys
import os
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage

def fix_sense_merch_mapping():
    """Исправляет привязку изображений для таблицы Мерч для Sense"""
    session = DatabaseManager.get_session()
    
    print("=== ИСПРАВЛЕНИЕ ПРИВЯЗКИ ДЛЯ МЕРЧ ДЛЯ SENSE ===\n")
    
    # Находим таблицу "Мерч для Sense"
    sense_sheet = session.query(SheetMetadata).filter(
        SheetMetadata.sheet_title.like('%Sense%')
    ).first()
    
    if not sense_sheet:
        print("❌ Таблица 'Мерч для Sense' не найдена")
        return
    
    print(f"📊 Обрабатываем таблицу: {sense_sheet.sheet_title}")
    
    # Анализируем структуру Excel файла
    try:
        wb = openpyxl.load_workbook(sense_sheet.local_file_path)
        ws = wb.active
        
        print(f"📏 Размер: {ws.max_row} строк, {ws.max_column} столбцов")
        
        # В этой таблице товары находятся в столбце 3 (C), а не в столбце 2
        name_column = 3
        
        # Находим все товары в этой таблице
        products_in_sheet = []
        for row in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=name_column).value
            
            if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower() and 'менеджер' not in str(cell_value).lower():
                product_name = str(cell_value).strip()
                products_in_sheet.append((row, product_name))
                print(f"   📦 Найден товар в строке {row}: {product_name}")
        
        wb.close()
        
        # Находим изображения, которые не привязаны к товарам
        unassigned_images = session.query(ProductImage).join(Product).filter(
            Product.sheet_id == sense_sheet.id,
            ProductImage.image_type == 'main'
        ).all()
        
        print(f"\n📥 Найдено изображений для перепривязки: {len(unassigned_images)}")
        
        # Перепривязываем изображения
        fixed_count = 0
        
        for image in unassigned_images:
            # Извлекаем номер строки из имени файла
            filename = os.path.basename(image.local_path)
            if 'A' in filename:
                try:
                    # Ищем паттерн A{number} в имени файла
                    import re
                    match = re.search(r'A(\d+)', filename)
                    if match:
                        image_row = int(match.group(1))
                        
                        # Находим товар для этой строки
                        correct_product = None
                        for product_row, product_name in products_in_sheet:
                            if product_row == image_row:
                                correct_product = product_name
                                break
                        
                        if correct_product:
                            # Находим товар в базе данных
                            product = session.query(Product).filter(
                                Product.name == correct_product,
                                Product.sheet_id == sense_sheet.id
                            ).first()
                            
                            if product:
                                # Обновляем привязку изображения
                                old_product_id = image.product_id
                                image.product_id = product.id
                                session.add(image)
                                fixed_count += 1
                                
                                print(f"   ✅ {filename} → {correct_product} (строка {image_row})")
                            else:
                                print(f"   ⚠️  Товар '{correct_product}' не найден в базе данных")
                        else:
                            print(f"   ⚠️  Не найден товар для строки {image_row}: {filename}")
                            
                except Exception as e:
                    print(f"   ❌ Ошибка при обработке {filename}: {e}")
        
        session.commit()
        print(f"\n✅ Исправлено изображений: {fixed_count}")
        
    except Exception as e:
        print(f"❌ Ошибка при обработке таблицы: {e}")
    
    session.close()

if __name__ == "__main__":
    try:
        fix_sense_merch_mapping()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

