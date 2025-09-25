#!/usr/bin/env python3
"""
Скрипт для анализа структуры Excel файлов и определения правильной позиции изображений
"""

import sys
import os
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata

def analyze_excel_structure():
    """Анализирует структуру Excel файлов для понимания позиций изображений"""
    session = DatabaseManager.get_session()
    
    print("=== АНАЛИЗ СТРУКТУРЫ EXCEL ФАЙЛОВ ===\n")
    
    # Получаем все таблицы
    sheets = session.query(SheetMetadata).all()
    
    for sheet in sheets:
        print(f"📊 Таблица: {sheet.sheet_title}")
        print(f"   Файл: {sheet.local_file_path}")
        
        if not os.path.exists(sheet.local_file_path):
            print(f"   ❌ Файл не найден: {sheet.local_file_path}")
            continue
            
        try:
            # Открываем Excel файл
            wb = openpyxl.load_workbook(sheet.local_file_path)
            ws = wb.active
            
            print(f"   Размер: {ws.max_row} строк, {ws.max_column} столбцов")
            
            # Анализируем первые 20 строк для понимания структуры
            print("   Первые 20 строк:")
            for row in range(1, min(21, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(6, ws.max_column + 1)):  # Первые 5 столбцов
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        row_data.append(f"{col}:{str(cell_value)[:20]}")
                    else:
                        row_data.append(f"{col}:")
                
                if any(row_data):
                    print(f"     Строка {row:2d}: {' | '.join(row_data)}")
            
            # Ищем столбец с названиями товаров
            name_column = None
            for col in range(1, ws.max_column + 1):
                for row in range(1, min(10, ws.max_row + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and 'наименование' in str(cell_value).lower():
                        name_column = col
                        print(f"   📝 Столбец с названиями: {col} (найден в строке {row})")
                        break
                if name_column:
                    break
            
            # Анализируем товары в столбце названий
            if name_column:
                print(f"   Товары в столбце {name_column}:")
                for row in range(1, min(50, ws.max_row + 1)):
                    cell_value = ws.cell(row=row, column=name_column).value
                    if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower():
                        print(f"     Строка {row:2d}: {str(cell_value)[:50]}")
            
            wb.close()
            print()
            
        except Exception as e:
            print(f"   ❌ Ошибка при анализе: {e}")
            print()
    
    session.close()

if __name__ == "__main__":
    try:
        analyze_excel_structure()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()
