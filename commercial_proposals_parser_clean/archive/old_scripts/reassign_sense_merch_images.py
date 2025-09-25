#!/usr/bin/env python3
"""
Скрипт для перепривязки изображений из таблицы "Мерч для Sense"
"""

import sys
import os
import re
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage

def reassign_sense_merch_images():
    """Перепривязывает изображения из таблицы Мерч для Sense"""
    session = DatabaseManager.get_session()
    
    print("=== ПЕРЕПРИВЯЗКА ИЗОБРАЖЕНИЙ МЕРЧ ДЛЯ SENSE ===\n")
    
    # Находим таблицу "Мерч для Sense"
    sense_sheet = session.query(SheetMetadata).filter(
        SheetMetadata.sheet_title.like('%Sense%')
    ).first()
    
    if not sense_sheet:
        print("❌ Таблица 'Мерч для Sense' не найдена")
        return
    
    print(f"📊 Обрабатываем таблицу: {sense_sheet.sheet_title}")
    
    # Анализируем структуру Excel файла
    try:
        wb = openpyxl.load_workbook(sense_sheet.local_file_path)
        ws = wb.active
        
        # В этой таблице товары находятся в столбце 3 (C)
        name_column = 3
        
        # Создаем карту товаров по строкам
        product_map = {}
        for row in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=name_column).value
            
            if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower() and 'менеджер' not in str(cell_value).lower():
                product_name = str(cell_value).strip()
                product_map[row] = product_name
                print(f"   📦 Строка {row}: {product_name}")
        
        wb.close()
        
        # Находим все изображения из этой таблицы
        sense_images = session.query(ProductImage).join(Product).filter(
            Product.sheet_id == sense_sheet.id
        ).all()
        
        print(f"\n📥 Найдено изображений для перепривязки: {len(sense_images)}")
        
        # Перепривязываем изображения
        fixed_count = 0
        
        for image in sense_images:
            filename = os.path.basename(image.local_path)
            
            # Извлекаем номер строки из имени файла
            match = re.search(r'A(\d+)', filename)
            if match:
                image_row = int(match.group(1))
                
                # Находим товар для этой строки
                if image_row in product_map:
                    product_name = product_map[image_row]
                    
                    # Находим товар в базе данных
                    product = session.query(Product).filter(
                        Product.name == product_name,
                        Product.sheet_id == sense_sheet.id
                    ).first()
                    
                    if product:
                        # Обновляем привязку изображения
                        old_product_id = image.product_id
                        image.product_id = product.id
                        session.add(image)
                        fixed_count += 1
                        
                        print(f"   ✅ {filename} → {product_name} (строка {image_row})")
                    else:
                        print(f"   ⚠️  Товар '{product_name}' не найден в базе данных")
                else:
                    print(f"   ⚠️  Не найден товар для строки {image_row}: {filename}")
            else:
                print(f"   ⚠️  Не удалось извлечь номер строки из: {filename}")
        
        session.commit()
        print(f"\n✅ Перепривязано изображений: {fixed_count}")
        
    except Exception as e:
        print(f"❌ Ошибка при обработке таблицы: {e}")
        import traceback
        traceback.print_exc()
    
    session.close()

if __name__ == "__main__":
    try:
        reassign_sense_merch_images()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

