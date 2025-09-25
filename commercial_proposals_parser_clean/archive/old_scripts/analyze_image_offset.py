#!/usr/bin/env python3
"""
Скрипт для анализа сдвига изображений относительно названий товаров
"""

import sys
import os
import re
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent.absolute()))

import openpyxl
from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, ProductImage

def analyze_image_offset():
    """Анализирует сдвиг изображений относительно названий товаров"""
    session = DatabaseManager.get_session()
    
    print("=== АНАЛИЗ СДВИГА ИЗОБРАЖЕНИЙ ===\n")
    
    # Анализируем original_sheet
    original_sheet = session.query(SheetMetadata).filter(
        SheetMetadata.sheet_title == 'original_sheet'
    ).first()
    
    if not original_sheet:
        print("❌ Таблица 'original_sheet' не найдена")
        return
    
    print(f"📊 Анализируем таблицу: {original_sheet.sheet_title}")
    
    try:
        wb = openpyxl.load_workbook(original_sheet.local_file_path)
        ws = wb.active
        
        # Находим все товары в Excel
        products_in_excel = []
        for row in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=2).value  # столбец B
            
            if cell_value and str(cell_value).strip() and 'наименование' not in str(cell_value).lower() and 'менеджер' not in str(cell_value).lower():
                product_name = str(cell_value).strip()
                products_in_excel.append((row, product_name))
                print(f"   📦 Строка {row}: {product_name}")
        
        wb.close()
        
        # Находим все изображения из этой таблицы
        images_dir = Path("storage/images")
        original_images = list(images_dir.glob("original_sheet_*"))
        
        print(f"\n📥 Найдено изображений: {len(original_images)}")
        
        # Анализируем соответствие
        for image_path in original_images:
            filename = image_path.name
            
            # Извлекаем номер строки из имени файла
            match = re.search(r'A(\d+)', filename)
            if match:
                image_row = int(match.group(1))
                
                # Находим ближайший товар
                closest_product = None
                min_distance = float('inf')
                
                for excel_row, product_name in products_in_excel:
                    distance = abs(excel_row - image_row)
                    if distance < min_distance:
                        min_distance = distance
                        closest_product = (excel_row, product_name)
                
                if closest_product:
                    excel_row, product_name = closest_product
                    print(f"   🖼️  {filename} (строка {image_row}) → ближайший товар: {product_name} (строка {excel_row}, расстояние: {min_distance})")
                    
                    # Проверяем, есть ли этот товар в базе данных
                    product = session.query(Product).filter(
                        Product.name == product_name,
                        Product.sheet_id == original_sheet.id
                    ).first()
                    
                    if product:
                        print(f"      ✅ Товар найден в БД: строки {product.start_row}-{product.end_row}")
                        
                        # Проверяем, привязано ли изображение к этому товару
                        existing_image = session.query(ProductImage).filter(
                            ProductImage.local_path == str(image_path)
                        ).first()
                        
                        if existing_image:
                            print(f"      ✅ Изображение уже привязано к товару ID {existing_image.product_id}")
                        else:
                            print(f"      ⚠️  Изображение не привязано к товару")
                    else:
                        print(f"      ❌ Товар не найден в БД")
                else:
                    print(f"   ⚠️  {filename} (строка {image_row}) → товар не найден")
            else:
                print(f"   ⚠️  Не удалось извлечь номер строки из: {filename}")
        
    except Exception as e:
        print(f"❌ Ошибка при анализе: {e}")
        import traceback
        traceback.print_exc()
    
    session.close()

if __name__ == "__main__":
    try:
        analyze_image_offset()
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

