#!/usr/bin/env python3
"""
Отладочный скрипт для проверки парсинга конкретной таблицы
"""

import requests
import openpyxl
from io import BytesIO

def debug_parse_table():
    # URL проблемной таблицы
    url = 'https://docs.google.com/spreadsheets/d/18mCRfsNMhKGX8nCOxd4Taa4NzJ4LBbC4WE6KMDzSw-c/edit'
    sheet_id = url.split('/d/')[1].split('/')[0]
    excel_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
    
    print("🔍 ОТЛАДКА ПАРСИНГА ТАБЛИЦЫ")
    print("=" * 50)
    
    # Скачиваем файл
    response = requests.get(excel_url, timeout=30)
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        print(f"\n📋 Лист: {sheet_name}")
        
        # ШАГ 1: Поиск заголовков
        print("\n1️⃣ ПОИСК ЗАГОЛОВКОВ:")
        header_row = None
        header_keywords = ['название', 'товар', 'product', 'name', 'наименование', 'артикул', 'код']
        
        for row_idx in range(1, 10):
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            print(f"   Строка {row_idx}: {[str(cell)[:15] if cell else '' for cell in row[:5]]}")
            
            if any(cell and any(keyword in str(cell).lower() for keyword in header_keywords) for cell in row):
                header_row = row_idx
                print(f"   ✅ Найдена строка заголовков: {row_idx}")
                break
        
        if not header_row:
            print("   ❌ Заголовки не найдены!")
            continue
        
        # ШАГ 2: Анализ заголовков
        print(f"\n2️⃣ АНАЛИЗ ЗАГОЛОВКОВ (строка {header_row}):")
        headers = list(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        
        name_col = None
        for col_idx, header in enumerate(headers):
            print(f"   Столбец {col_idx}: '{header}'")
            if header and any(keyword in str(header).lower() for keyword in header_keywords):
                name_col = col_idx
                print(f"   ✅ Найден столбец с названием: {col_idx} ('{header}')")
        
        if name_col is None:
            print("   ❌ Столбец с названием товара не найден!")
            continue
        
        # ШАГ 3: Парсинг товаров
        print(f"\n3️⃣ ПАРСИНГ ТОВАРОВ (столбец {name_col}):")
        products_found = 0
        
        for row_idx in range(header_row + 1, min(header_row + 20, worksheet.max_row + 1)):
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            
            if len(row) > name_col and row[name_col]:
                product_name = str(row[name_col]).strip()
                if len(product_name) > 2:  # Минимальная длина названия
                    products_found += 1
                    print(f"   Строка {row_idx}: '{product_name}'")
                    
                    if products_found >= 5:  # Показываем первые 5
                        print("   ... (показаны первые 5)")
                        break
        
        print(f"\n📊 ИТОГО найдено товаров: {products_found}")

if __name__ == "__main__":
    debug_parse_table()
