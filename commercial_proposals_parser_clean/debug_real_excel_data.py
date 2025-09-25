#!/usr/bin/env python3
"""
Отладка реальных данных в Excel файле vs БД
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from database.manager_v4 import DatabaseManager
from database.models_v4 import Product, PriceOffer, SheetMetadata

def debug_excel_file(file_path, specific_rows=None):
    """Детально анализируем Excel файл"""
    
    print(f"📁 АНАЛИЗ ФАЙЛА: {file_path}")
    print("=" * 80)
    
    if not Path(file_path).exists():
        print("❌ Файл не найден!")
        return
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"📊 Размер таблицы: {ws.max_row} строк, {ws.max_column} колонок")
        print()
        
        # Показываем заголовки (первые 3 строки)
        print("📋 ЗАГОЛОВКИ ТАБЛИЦЫ:")
        for row in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                col_letter = get_column_letter(col)
                if cell.value is not None:
                    row_data.append(f"{col_letter}: {str(cell.value)[:50]}")
            
            if row_data:
                print(f"   Строка {row}: {' | '.join(row_data)}")
        
        print("\n" + "="*80)
        
        # Если указаны конкретные строки - анализируем их детально
        if specific_rows:
            print("🔍 ДЕТАЛЬНЫЙ АНАЛИЗ УКАЗАННЫХ СТРОК:")
            for row_num in specific_rows:
                print(f"\n📍 СТРОКА {row_num}:")
                print("-" * 40)
                
                for col in range(1, min(20, ws.max_column + 1)):
                    cell = ws.cell(row=row_num, column=col)
                    col_letter = get_column_letter(col)
                    
                    if cell.value is not None:
                        # Проверяем числовое значение
                        cell_value = cell.value
                        cell_type = type(cell_value).__name__
                        
                        print(f"   {col_letter}{row_num}: {cell_value} (тип: {cell_type})")
                        
                        # Если это число - проверим может ли быть тиражом
                        if isinstance(cell_value, (int, float)):
                            if 1 <= cell_value <= 100000:
                                print(f"      🎯 ПОТЕНЦИАЛЬНЫЙ ТИРАЖ: {cell_value}")
                            elif cell_value > 0.1 and cell_value < 1000:
                                print(f"      💰 ПОТЕНЦИАЛЬНАЯ ЦЕНА: {cell_value}")
        
        # Ищем колонки с числовыми данными (потенциальные тиражи/цены)
        print(f"\n🔍 ПОИСК ЧИСЛОВЫХ КОЛОНОК (потенциальные тиражи и цены):")
        print("=" * 60)
        
        for col in range(1, min(15, ws.max_column + 1)):
            col_letter = get_column_letter(col)
            numbers = []
            
            # Смотрим первые 30 строк
            for row in range(1, min(31, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)) and cell_value > 0:
                    numbers.append((row, cell_value))
            
            if numbers:
                print(f"\n📊 Колонка {col_letter}:")
                for row, value in numbers[:10]:  # Первые 10 значений
                    if isinstance(value, float) and 1 <= value <= 100000:
                        print(f"      Строка {row}: {value} ← ПОХОЖЕ НА ТИРАЖ")
                    elif isinstance(value, int) and 1 <= value <= 100000:
                        print(f"      Строка {row}: {value} ← ПОХОЖЕ НА ТИРАЖ")  
                    elif 0.1 <= value <= 1000:
                        print(f"      Строка {row}: {value} ← ПОХОЖЕ НА ЦЕНУ")
                    else:
                        print(f"      Строка {row}: {value}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Ошибка анализа файла: {e}")

def compare_with_db():
    """Сравниваем что извлек парсер с реальными данными"""
    
    session = DatabaseManager.get_session()
    
    try:
        # Найдем товары "Ёлочная игрушка"
        xmas_products = session.query(Product).filter(
            Product.name.contains('лочная игрушка')
        ).all()
        
        print("\n🔍 СРАВНЕНИЕ С БАЗОЙ ДАННЫХ:")
        print("=" * 60)
        
        for product in xmas_products:
            sheet = session.query(SheetMetadata).get(product.sheet_id)
            prices = session.query(PriceOffer).filter(PriceOffer.product_id == product.id).all()
            
            print(f"\n📦 {product.name}")
            print(f"   📏 Строки в Excel: {product.start_row}-{product.end_row}")
            print(f"   📊 Файл: {sheet.local_file_path if sheet else 'Unknown'}")
            
            for price in prices:
                print(f"   💰 Парсер извлек:")
                print(f"      Тираж: {price.quantity}")
                print(f"      USD: ${price.price_usd}")
                print(f"      RUB: {price.price_rub} ₽")
            
            # Анализируем соответствующий Excel файл
            if sheet and sheet.local_file_path:
                print(f"   🔍 Реальные данные из Excel:")
                debug_excel_file(sheet.local_file_path, [product.start_row])
    
    finally:
        session.close()

if __name__ == "__main__":
    # Анализируем конкретный файл
    file_path = "/Users/bakirovresad/Downloads/Reshad 1/projects/commercial_proposals_parser_clean/storage/excel_files/sheet_1nav9w2d_public.xlsx"
    
    print("🎯 ДЕТАЛЬНЫЙ АНАЛИЗ ФАЙЛА С ЁЛОЧНЫМИ ИГРУШКАМИ")
    print("=" * 80)
    
    # Анализируем строки с товарами (из предыдущего анализа)
    problem_rows = [17, 19, 21, 23, 25]
    debug_excel_file(file_path, problem_rows)
    
    # Сравниваем с БД
    compare_with_db()


