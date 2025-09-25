#!/usr/bin/env python3
"""
Проверка точности данных - сверка товаров в БД с исходными Excel файлами
"""

from database.manager_v4 import DatabaseManager
from database.models_v4 import SheetMetadata, Product, PriceOffer
from pathlib import Path
import openpyxl
import pandas as pd
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_specific_products():
    """Проверяем конкретные товары из интерфейса"""
    
    session = DatabaseManager.get_session()
    
    try:
        # Найдем товары с названием "Ёлочная игрушка" из скриншота
        xmas_toys = session.query(Product).filter(
            Product.name.contains('лочная игрушка')
        ).all()
        
        print("🎄 НАЙДЕННЫЕ ТОВАРЫ 'ЁЛОЧНАЯ ИГРУШКА':")
        print("=" * 60)
        
        for product in xmas_toys:
            sheet = session.query(SheetMetadata).get(product.sheet_id)
            prices = session.query(PriceOffer).filter(PriceOffer.product_id == product.id).all()
            
            print(f"\n📦 Товар: {product.name}")
            print(f"   📊 Таблица: {sheet.sheet_title if sheet else 'Unknown'}")
            print(f"   📁 Файл: {sheet.local_file_path if sheet else 'Unknown'}")
            print(f"   📏 Строки: {product.start_row}-{product.end_row}")
            print(f"   💰 Ценовых предложений: {len(prices)}")
            
            for i, price in enumerate(prices, 1):
                print(f"      {i}. Тираж: {price.quantity} шт")
                print(f"         USD: ${price.price_usd}")
                print(f"         RUB: {price.price_rub} ₽" if price.price_rub else "         RUB: None")
                print(f"         Маршрут: {price.route_name}")
                print(f"         Заметки: {price.notes or 'None'}")
            
            # Теперь проверим исходный Excel файл
            if sheet and sheet.local_file_path and Path(sheet.local_file_path).exists():
                print(f"\n🔍 ПРОВЕРКА ИСХОДНОГО ФАЙЛА:")
                verify_excel_data(sheet.local_file_path, product.start_row, product.end_row)
            
            print("-" * 60)
        
        # Также проверим любые товары с ценами около $2.3-$3.7 как на скриншоте
        print(f"\n💰 ТОВАРЫ С ЦЕНАМИ $2-$4 (как на скриншоте):")
        print("=" * 60)
        
        similar_price_products = session.query(Product).join(PriceOffer).filter(
            PriceOffer.price_usd >= 2.0,
            PriceOffer.price_usd <= 4.0
        ).limit(5).all()
        
        for product in similar_price_products:
            sheet = session.query(SheetMetadata).get(product.sheet_id)
            prices = session.query(PriceOffer).filter(PriceOffer.product_id == product.id).all()
            
            print(f"\n📦 {product.name}")
            print(f"   📊 {sheet.sheet_title if sheet else 'Unknown'}")
            for price in prices:
                if 2.0 <= price.price_usd <= 4.0:
                    print(f"   💰 ${price.price_usd} | {price.quantity} шт | RUB: {price.price_rub or 'None'}")
            
            # Проверим исходник
            if sheet and sheet.local_file_path and Path(sheet.local_file_path).exists():
                verify_excel_data(sheet.local_file_path, product.start_row, product.end_row, brief=True)
        
        return xmas_toys + similar_price_products
        
    finally:
        session.close()

def verify_excel_data(file_path, start_row, end_row, brief=False):
    """Проверяем исходные данные в Excel файле"""
    try:
        file_path = Path(file_path)
        
        if not file_path.exists():
            print(f"      ❌ Файл не найден: {file_path}")
            return
            
        # Определяем формат файла
        if file_path.suffix.lower() == '.csv':
            # CSV файл
            try:
                df = pd.read_csv(file_path, encoding='utf-8')
                print(f"      📄 CSV файл, строки {start_row}-{end_row}:")
                
                for row_idx in range(max(0, start_row-1), min(len(df), end_row)):
                    row_data = df.iloc[row_idx]
                    if not brief:
                        print(f"         Строка {row_idx+1}: {dict(row_data.dropna())}")
                    else:
                        # Ищем ценовые столбцы
                        price_cols = [col for col in df.columns if any(word in str(col).lower() 
                                    for word in ['price', 'цена', 'usd', 'rub', 'руб'])]
                        qty_cols = [col for col in df.columns if any(word in str(col).lower() 
                                  for word in ['quantity', 'тираж', 'шт', 'pcs'])]
                        
                        prices = {col: row_data[col] for col in price_cols if pd.notna(row_data[col])}
                        qtys = {col: row_data[col] for col in qty_cols if pd.notna(row_data[col])}
                        
                        if prices or qtys:
                            print(f"         Строка {row_idx+1}: Цены={prices}, Тиражи={qtys}")
                        
            except Exception as e:
                print(f"      ❌ Ошибка чтения CSV: {e}")
                
        else:
            # Excel файл
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                print(f"      📄 Excel файл, строки {start_row}-{end_row}:")
                
                for row_idx in range(start_row, end_row + 1):
                    row_data = []
                    for col_idx in range(1, min(15, ws.max_column + 1)):  # Первые 15 колонок
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}: {cell_value}")
                    
                    if row_data and not brief:
                        print(f"         Строка {row_idx}: {', '.join(row_data[:5])}{'...' if len(row_data) > 5 else ''}")
                    elif row_data and brief:
                        # Ищем числовые значения (цены/тиражи)
                        numbers = []
                        for item in row_data[:10]:  # Первые 10 колонок
                            try:
                                if ':' in item:
                                    val = item.split(':')[1].strip()
                                    if val.replace('.', '').replace(',', '').isdigit():
                                        numbers.append(item)
                            except:
                                pass
                        
                        if numbers:
                            print(f"         Строка {row_idx}: {', '.join(numbers)}")
                
                wb.close()
                
            except Exception as e:
                print(f"      ❌ Ошибка чтения Excel: {e}")
                
    except Exception as e:
        print(f"      ❌ Общая ошибка: {e}")

def check_rub_prices_display():
    """Проверяем почему рубли не отображаются в интерфейсе"""
    
    session = DatabaseManager.get_session()
    
    try:
        # Найдем товары с RUB ценами
        rub_prices = session.query(PriceOffer).filter(
            PriceOffer.price_rub.isnot(None)
        ).limit(10).all()
        
        print(f"\n🔍 ПРОВЕРКА ОТОБРАЖЕНИЯ РУБЛЕЙ:")
        print("=" * 50)
        print(f"Найдено ценовых предложений с RUB: {len(rub_prices)}")
        
        for price in rub_prices[:5]:
            product = session.query(Product).get(price.product_id)
            print(f"\n📦 {product.name}")
            print(f"   USD: ${price.price_usd}")
            print(f"   RUB: {price.price_rub} ₽")
            print(f"   Тираж: {price.quantity}")
            print(f"   Заметки: {price.notes}")
        
        # Проверим веб-интерфейс
        web_app_path = Path("web_app_v4/main.py")
        if web_app_path.exists():
            print(f"\n🌐 ПРОВЕРКА ВЕБА:")
            with open(web_app_path, 'r', encoding='utf-8') as f:
                content = f.read()
                if 'price_rub' in content:
                    print("   ✅ price_rub упоминается в коде веба")
                else:
                    print("   ❌ price_rub НЕ найден в коде веба!")
                    
                if 'RUB' in content or 'руб' in content:
                    print("   ✅ Рублевые обозначения найдены")
                else:
                    print("   ❌ Рублевые обозначения НЕ найдены!")
        
        return len(rub_prices)
        
    finally:
        session.close()

if __name__ == "__main__":
    print("🔍 ПРОВЕРКА ТОЧНОСТИ ДАННЫХ")
    print("=" * 80)
    
    products = verify_specific_products()
    rub_count = check_rub_prices_display()
    
    print(f"\n📊 ИТОГ:")
    print(f"   Проверено товаров: {len(products) if products else 0}")
    print(f"   Товаров с RUB ценами: {rub_count}")
    print(f"   \n🎯 РЕКОМЕНДАЦИИ:")
    print(f"   1. Проверить правильность извлечения цен и тиражей")
    print(f"   2. Исправить отображение рублей в веб-интерфейсе") 
    print(f"   3. Сверить формулы конвертации валют")


