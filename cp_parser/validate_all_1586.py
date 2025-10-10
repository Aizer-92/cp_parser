#!/usr/bin/env python3
"""
Валидация всех скачанных Excel файлов
Проверяет какие файлы подходят под существующую структуру парсера
"""

import sys
from pathlib import Path
from datetime import datetime
import json
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from src.structure_parser import CommercialProposalParser

def validate_single_file(file_path: Path, parser: CommercialProposalParser, index: int, total: int):
    """Валидирует один Excel файл"""
    
    file_name = file_path.name
    
    try:
        # Загружаем Excel
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if len(workbook.sheetnames) == 0:
            return {
                'file': file_name,
                'status': 'error',
                'error': 'No sheets in workbook'
            }
        
        worksheet = workbook[workbook.sheetnames[0]]
        sheet_name = workbook.sheetnames[0]
        
        # Валидация структуры
        validation = parser.validate_table_structure(worksheet, sheet_name)
        
        if validation['is_valid']:
            return {
                'file': file_name,
                'status': 'valid',
                'main_columns': validation.get('matched_main_columns_count', 0),
                'routes': validation.get('matched_routes_count', 0),
                'structure': validation
            }
        else:
            return {
                'file': file_name,
                'status': 'invalid',
                'errors': validation.get('validation_errors', []),
                'main_columns': validation.get('matched_main_columns_count', 0),
                'routes': validation.get('matched_routes_count', 0)
            }
        
    except Exception as e:
        return {
            'file': file_name,
            'status': 'error',
            'error': str(e)
        }

def main():
    """Главная функция"""
    
    print("="*80)
    print("✅ ВАЛИДАЦИЯ ВСЕХ СКАЧАННЫХ ФАЙЛОВ")
    print("="*80)
    print()
    
    # 1. Находим все Excel файлы
    excel_dir = Path(__file__).parent / 'storage' / 'excel_files'
    excel_files = sorted(excel_dir.glob("*.xlsx"))
    
    if len(excel_files) == 0:
        print("❌ Не найдено Excel файлов")
        return
    
    print(f"📁 Папка: {excel_dir}")
    print(f"📊 Найдено файлов: {len(excel_files)}")
    print()
    
    # 2. Создаем парсер
    parser = CommercialProposalParser()
    
    # 3. Валидируем все файлы
    results = {
        'valid': [],
        'invalid': [],
        'error': []
    }
    
    print("⏱️  Начинаем валидацию...\n")
    
    start_time = datetime.now()
    
    for i, file_path in enumerate(excel_files, 1):
        if i % 50 == 0 or i == 1:
            print(f"[{i}/{len(excel_files)}] Проверяю: {file_path.name}")
        
        result = validate_single_file(file_path, parser, i, len(excel_files))
        
        status = result['status']
        results[status].append(result)
        
        # Прогресс каждые 100 файлов
        if i % 100 == 0:
            valid_count = len(results['valid'])
            invalid_count = len(results['invalid'])
            error_count = len(results['error'])
            
            print(f"\n📊 Прогресс: {i}/{len(excel_files)}")
            print(f"   ✅ Валидных: {valid_count} ({valid_count/i*100:.1f}%)")
            print(f"   ⚠️  Невалидных: {invalid_count} ({invalid_count/i*100:.1f}%)")
            print(f"   ❌ Ошибок: {error_count} ({error_count/i*100:.1f}%)")
            print()
    
    # 4. Итоговая статистика
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    print("\n" + "="*80)
    print("📊 ИТОГОВАЯ СТАТИСТИКА ВАЛИДАЦИИ")
    print("="*80)
    
    total = len(excel_files)
    valid_count = len(results['valid'])
    invalid_count = len(results['invalid'])
    error_count = len(results['error'])
    
    print(f"\n✅ Валидных файлов: {valid_count} ({valid_count/total*100:.1f}%)")
    print(f"⚠️  Невалидных файлов: {invalid_count} ({invalid_count/total*100:.1f}%)")
    print(f"❌ Ошибок чтения: {error_count} ({error_count/total*100:.1f}%)")
    
    print(f"\n⏱️  Время валидации: {duration:.0f} сек ({duration/60:.1f} мин)")
    print(f"⚡ Скорость: {total/duration:.1f} файлов/сек")
    
    # 5. Анализ причин невалидности
    if invalid_count > 0:
        print(f"\n⚠️  ПРИЧИНЫ НЕВАЛИДНОСТИ (ТОП-10):")
        
        # Собираем все ошибки
        error_reasons = {}
        for result in results['invalid']:
            errors = result.get('errors', [])
            if isinstance(errors, list):
                for err in errors:
                    error_str = str(err)[:100]  # Первые 100 символов
                    error_reasons[error_str] = error_reasons.get(error_str, 0) + 1
            else:
                error_str = str(errors)[:100]
                error_reasons[error_str] = error_reasons.get(error_str, 0) + 1
        
        # Сортируем по частоте
        sorted_errors = sorted(error_reasons.items(), key=lambda x: x[1], reverse=True)
        
        for i, (error, count) in enumerate(sorted_errors[:10], 1):
            print(f"   {i}. ({count} файлов) {error}")
    
    # 6. Примеры валидных файлов
    if valid_count > 0:
        print(f"\n✅ ПРИМЕРЫ ВАЛИДНЫХ ФАЙЛОВ (первые 5):")
        for result in results['valid'][:5]:
            print(f"   - {result['file']}")
            print(f"     Столбцы: {result['main_columns']}, Маршруты: {result['routes']}")
    
    # 7. Сохраняем детальный отчет
    report = {
        'timestamp': datetime.now().isoformat(),
        'total_files': total,
        'statistics': {
            'valid': valid_count,
            'invalid': invalid_count,
            'errors': error_count,
            'valid_percentage': round(valid_count/total*100, 2),
            'duration_seconds': round(duration, 2)
        },
        'results': {
            'valid': results['valid'][:100],  # Первые 100
            'invalid': results['invalid'][:100],  # Первые 100
            'error': results['error'][:100]  # Первые 100
        },
        'error_reasons': dict(sorted(error_reasons.items(), key=lambda x: x[1], reverse=True)[:20]) if invalid_count > 0 else {}
    }
    
    with open('validation_report.json', 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Детальный отчет сохранен: validation_report.json")
    
    # 8. Создаем список файлов для парсинга
    if valid_count > 0:
        valid_files = [r['file'] for r in results['valid']]
        with open('valid_files_list.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(valid_files))
        
        print(f"✅ Список валидных файлов: valid_files_list.txt")
    
    print("\n" + "="*80)
    print("💡 РЕКОМЕНДАЦИИ")
    print("="*80)
    
    if valid_count > 0:
        print(f"\n1️⃣  Готово {valid_count} файлов для парсинга")
        print(f"2️⃣  Запустите: python3 parse_valid_files.py")
    
    if invalid_count > 0:
        print(f"\n3️⃣  {invalid_count} файлов требуют новых шаблонов парсера")
        print(f"4️⃣  Проверьте причины в validation_report.json")
    
    print("\n" + "="*80)

if __name__ == "__main__":
    main()




