#!/usr/bin/env python3
"""
Парсер для Шаблона 4 (идеальные проекты, confidence 9/9)
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import re
import hashlib

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template4Parser:
    """
    Парсер для Шаблона 4
    
    Особенности:
    - Тираж в колонке E (один для всех маршрутов)
    - Multi-row товары (доп. тиражи без названия)
    - Колонка D = Дизайн/Кастом
    - Игнорируем "Итого, руб" (H, L)
    """
    
    DATA_START_ROW = 4
    
    # Маппинг колонок
    COLUMNS = {
        'photo': 1,           # A
        'name': 2,            # B
        'description': 3,     # C
        'custom_field': 4,    # D (Дизайн/Кастом)
        'quantity': 5,        # E (ОДИН для всех маршрутов)
        
        # ЖД маршрут (F-I)
        'jd_price_usd': 6,    # F
        'jd_price_rub': 7,    # G
        # H - Итого руб (SKIP)
        'jd_delivery': 9,     # I
        
        # АВИА маршрут (J-M)
        'avia_price_usd': 10, # J
        'avia_price_rub': 11, # K
        # L - Итого руб (SKIP)
        'avia_delivery': 13,  # M
        
        # Образец (N-P)
        'sample_price': 14,   # N
        'sample_photo_days': 15, # O
        'sample_delivery': 16,   # P
        
        'extra_photo': 17     # Q
    }
    
    def __init__(self):
        self.db = PostgreSQLManager()
        self.storage_dir = Path('storage/images')
        self.storage_dir.mkdir(parents=True, exist_ok=True)
    
    def parse_project(self, project_id):
        """Парсит один проект"""
        try:
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx'))
            
            if not matching_files:
                return {'success': False, 'error': 'File not found'}
            
            file_path = matching_files[0]
            
            # Парсим
            wb_data = load_workbook(str(file_path), data_only=True)
            wb_images = load_workbook(str(file_path), data_only=False)
            
            ws_data = wb_data.active
            ws_images = wb_images.active
            
            # Получаем table_id из БД
            with self.db.get_session() as session:
                result = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).fetchone()
                
                if not result:
                    return {'success': False, 'error': 'Project not in DB'}
                
                table_id = result[0]
            
            # Парсим товары
            products = self._parse_products(ws_data, ws_images, project_id, table_id)
            
            wb_data.close()
            wb_images.close()
            
            return {
                'success': True,
                'products': len(products),
                'total_offers': sum(len(p['offers']) for p in products),
                'total_images': sum(len(p['images']) for p in products)
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def _parse_products(self, ws_data, ws_images, project_id, table_id):
        """Парсит все товары из листа"""
        products = []
        current_product = None
        
        max_row = ws_data.max_row
        
        for row in range(self.DATA_START_ROW, max_row + 1):
            name = self._get_cell_value(ws_data, row, self.COLUMNS['name'])
            quantity = self._get_cell_value(ws_data, row, self.COLUMNS['quantity'])
            
            # Пропускаем пустые строки
            if not name and not quantity:
                continue
            
            # Новый товар
            if name and len(str(name).strip()) >= 3:
                # Сохраняем предыдущий товар
                if current_product:
                    self._save_product(current_product, project_id, table_id)
                    products.append(current_product)
                
                # Создаем новый товар
                current_product = self._create_product(ws_data, ws_images, row, table_id)
            
            # Парсим предложения для текущего товара
            if current_product and quantity:
                offers = self._parse_offers(ws_data, row)
                current_product['offers'].extend(offers)
        
        # Сохраняем последний товар
        if current_product:
            self._save_product(current_product, project_id, table_id)
            products.append(current_product)
        
        return products
    
    def _create_product(self, ws_data, ws_images, row, table_id):
        """Создает структуру товара"""
        name = self._get_cell_value(ws_data, row, self.COLUMNS['name'])
        description = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
        custom_field = self._get_cell_value(ws_data, row, self.COLUMNS['custom_field'])
        
        # Парсим данные об образце (колонки N, O, P)
        sample_price_raw = self._get_cell_value(ws_data, row, self.COLUMNS['sample_price'])
        sample_delivery_raw = self._get_cell_value(ws_data, row, self.COLUMNS['sample_delivery'])
        
        # Парсим цену и срок образца
        sample_price = self._parse_price(sample_price_raw) if sample_price_raw else None
        sample_delivery = self._parse_delivery_time(sample_delivery_raw) if sample_delivery_raw else None
        
        # Извлекаем изображения
        images = self._extract_images(ws_images, row, table_id)
        
        return {
            'name': str(name).strip(),
            'description': str(description).strip() if description else None,
            'custom_field': str(custom_field).strip() if custom_field else None,
            'sample_price': str(sample_price) if sample_price else None,
            'sample_delivery_time': sample_delivery,
            'row_number': row,
            'offers': [],
            'images': images
        }
    
    def _parse_offers(self, ws_data, row):
        """Парсит предложения из строки"""
        offers = []
        quantity = self._get_cell_value(ws_data, row, self.COLUMNS['quantity'])
        
        if not quantity:
            return offers
        
        # Парсим количество
        qty = self._parse_quantity(quantity)
        if not qty or qty <= 0 or qty > 1_000_000:
            return offers
        
        # ЖД маршрут
        jd_usd = self._get_cell_value(ws_data, row, self.COLUMNS['jd_price_usd'])
        jd_rub = self._get_cell_value(ws_data, row, self.COLUMNS['jd_price_rub'])
        
        if jd_usd or jd_rub:
            price_usd = self._parse_price(jd_usd)
            price_rub = self._parse_price(jd_rub)
            delivery = self._parse_delivery_time(
                self._get_cell_value(ws_data, row, self.COLUMNS['jd_delivery'])
            )
            
            if price_usd and price_usd < 10_000 and price_rub and price_rub < 1_000_000:
                offers.append({
                    'quantity': qty,
                    'price_usd': f"{price_usd:.2f}",
                    'price_rub': f"{price_rub:.2f}",
                    'route': 'ЖД',
                    'delivery_time_days': delivery if delivery and delivery <= 365 else 60
                })
        
        # АВИА маршрут
        avia_usd = self._get_cell_value(ws_data, row, self.COLUMNS['avia_price_usd'])
        avia_rub = self._get_cell_value(ws_data, row, self.COLUMNS['avia_price_rub'])
        
        if avia_usd or avia_rub:
            price_usd = self._parse_price(avia_usd)
            price_rub = self._parse_price(avia_rub)
            delivery = self._parse_delivery_time(
                self._get_cell_value(ws_data, row, self.COLUMNS['avia_delivery'])
            )
            
            if price_usd and price_usd < 10_000 and price_rub and price_rub < 1_000_000:
                offers.append({
                    'quantity': qty,
                    'price_usd': f"{price_usd:.2f}",
                    'price_rub': f"{price_rub:.2f}",
                    'route': 'АВИА',
                    'delivery_time_days': delivery if delivery and delivery <= 365 else 40
                })
        
        # Образец (колонка N) - добавляем КАК OFFER для совместимости с фронтендом
        # И также оставляем в custom_field
        sample_price = self._get_cell_value(ws_data, row, self.COLUMNS['sample_price'])
        if sample_price:
            price_rub = self._parse_price(sample_price)
            delivery = self._parse_delivery_time(
                self._get_cell_value(ws_data, row, self.COLUMNS['sample_delivery'])
            )
            
            if price_rub and price_rub < 1_000_000:
                offers.append({
                    'quantity': 1,
                    'price_usd': None,
                    'price_rub': f"{price_rub:.2f}",
                    'route': 'Образец',
                    'delivery_time_days': delivery if delivery and delivery <= 365 else 14
                })
        
        return offers
    
    def _extract_images(self, ws_images, row, table_id):
        """Извлекает изображения из строки"""
        images = []
        
        # Основное фото (колонка A)
        img_main = self._extract_image_from_cell(ws_images, row, self.COLUMNS['photo'], table_id)
        if img_main:
            images.append(img_main)
        
        # Дополнительные фото (колонка Q) - может быть НЕСКОЛЬКО!
        extra_images = self._extract_all_images_from_cell(ws_images, row, self.COLUMNS['extra_photo'], table_id)
        images.extend(extra_images)
        
        return images
    
    def _extract_all_images_from_cell(self, ws, row, col, table_id):
        """Извлекает ВСЕ изображения из конкретной ячейки (для колонки Q)"""
        images = []
        try:
            col_letter = chr(64 + col)
            cell_pos = f"{col_letter}{row}"
            img_index = 0
            
            for image in ws._images:
                if not hasattr(image, 'anchor') or not hasattr(image.anchor, '_from'):
                    continue
                
                img_row = image.anchor._from.row + 1
                img_col = image.anchor._from.col + 1
                
                if img_row == row and img_col == col:
                    img_data = image._data()
                    img_hash = hashlib.md5(img_data).hexdigest()[:8]
                    
                    # Для первого изображения используем существующее имя (не дублируем)
                    if img_index == 0:
                        filename = f"{table_id}_{cell_pos}_{img_hash}.png"
                    else:
                        # Для остальных добавляем суффикс
                        filename = f"{table_id}_{cell_pos}_{img_index}_{img_hash}.png"
                    
                    filepath = self.storage_dir / filename
                    
                    # Сохраняем изображение
                    with open(filepath, 'wb') as f:
                        f.write(img_data)
                    
                    images.append({
                        'filename': filename,
                        'cell_position': cell_pos,
                        'row_number': row,
                        'column_number': col,
                        'is_main': False  # Дополнительные фото не главные
                    })
                    
                    img_index += 1
        except Exception as e:
            pass
        
        return images
    
    def _extract_image_from_cell(self, ws, row, col, table_id):
        """Извлекает изображение из конкретной ячейки"""
        try:
            for image in ws._images:
                if not hasattr(image, 'anchor') or not hasattr(image.anchor, '_from'):
                    continue
                
                img_row = image.anchor._from.row + 1
                img_col = image.anchor._from.col + 1
                
                if img_row == row and img_col == col:
                    # Генерируем имя файла
                    col_letter = chr(64 + col)
                    cell_pos = f"{col_letter}{row}"
                    
                    img_data = image._data()
                    img_hash = hashlib.md5(img_data).hexdigest()[:8]
                    
                    filename = f"{table_id}_{cell_pos}_{img_hash}.png"
                    filepath = self.storage_dir / filename
                    
                    # Сохраняем изображение
                    with open(filepath, 'wb') as f:
                        f.write(img_data)
                    
                    return {
                        'filename': filename,
                        'cell_position': cell_pos,
                        'row_number': row,
                        'column_number': col,
                        'is_main': col == self.COLUMNS['photo']
                    }
        except Exception as e:
            pass
        
        return None
    
    def _save_product(self, product, project_id, table_id):
        """Сохраняет товар в БД"""
        try:
            with self.db.get_session() as session:
                # Вставка товара
                result = session.execute(text("""
                    INSERT INTO products (
                        project_id, table_id, name, description, custom_field, 
                        sample_price, sample_delivery_time,
                        row_number, created_at, updated_at
                    ) VALUES (
                        :project_id, :table_id, :name, :description, :custom_field,
                        :sample_price, :sample_delivery_time,
                        :row_number, :created_at, :updated_at
                    ) RETURNING id
                """), {
                    'project_id': project_id,
                    'table_id': table_id,
                    'name': product['name'],
                    'description': product['description'],
                    'custom_field': product['custom_field'],
                    'sample_price': product.get('sample_price'),
                    'sample_delivery_time': product.get('sample_delivery_time'),
                    'row_number': product['row_number'],
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat()
                })
                
                product_id = result.fetchone()[0]
                session.commit()
                
                # Вставка предложений
                for offer in product['offers']:
                    session.execute(text("""
                        INSERT INTO price_offers (
                            product_id, table_id, quantity, price_usd, price_rub,
                            route, delivery_time_days, created_at, updated_at
                        ) VALUES (
                            :product_id, :table_id, :quantity, :price_usd, :price_rub,
                            :route, :delivery_time_days, :created_at, :updated_at
                        )
                    """), {
                        'product_id': product_id,
                        'table_id': table_id,
                        'quantity': offer['quantity'],
                        'price_usd': offer['price_usd'],
                        'price_rub': offer['price_rub'],
                        'route': offer['route'],
                        'delivery_time_days': offer['delivery_time_days'],
                        'created_at': datetime.now().isoformat(),
                        'updated_at': datetime.now().isoformat()
                    })
                
                # Вставка изображений
                for img in product['images']:
                    session.execute(text("""
                        INSERT INTO product_images (
                            product_id, table_id, image_filename, cell_position,
                            row_number, column_number, created_at, updated_at
                        ) VALUES (
                            :product_id, :table_id, :image_filename, :cell_position,
                            :row_number, :column_number, :created_at, :updated_at
                        )
                    """), {
                        'product_id': product_id,
                        'table_id': table_id,
                        'image_filename': img['filename'],
                        'cell_position': img['cell_position'],
                        'row_number': img['row_number'],
                        'column_number': img['column_number'],
                        'created_at': datetime.now().isoformat(),
                        'updated_at': datetime.now().isoformat()
                    })
                
                session.commit()
                
        except Exception as e:
            print(f"      ⚠️ Ошибка сохранения: {str(e)[:100]}")
    
    def _get_cell_value(self, ws, row, col):
        """Получает значение ячейки"""
        try:
            val = ws.cell(row, col).value
            return val if val is not None else None
        except:
            return None
    
    def _parse_quantity(self, value):
        """Парсит количество"""
        if not value:
            return None
        
        val_str = str(value).strip().replace(' ', '')
        
        # Если есть дефис - берем первое число
        if '-' in val_str:
            val_str = val_str.split('-')[0].strip()
        
        # Правильное преобразование через float (для "10000.0" → 10000)
        try:
            return int(float(val_str))
        except (ValueError, TypeError):
            return None
    
    def _parse_price(self, value):
        """Парсит цену"""
        if not value:
            return None
        
        try:
            val_str = str(value).strip()
            val_str = val_str.replace(',', '.').replace(' ', '')
            val_str = ''.join(c for c in val_str if c.isdigit() or c == '.')
            
            if val_str:
                return float(val_str)
        except:
            pass
        
        return None
    
    def _parse_delivery_time(self, value):
        """Парсит срок доставки"""
        if not value:
            return None
        
        try:
            val_str = str(value).strip()
            
            # Если диапазон - берем максимум (хуже для клиента)
            if '-' in val_str:
                val_str = val_str.split('-')[-1].strip()
            
            # Извлекаем цифры
            digits = ''.join(filter(str.isdigit, val_str))
            
            if digits:
                days = int(digits)
                return days if days <= 365 else None
        except:
            pass
        
        return None
    
    def parse_all_perfect(self):
        """Парсит все идеальные проекты"""
        # Читаем список ID
        ids_file = Path('template_4_perfect_ids.txt')
        if not ids_file.exists():
            print("❌ Файл template_4_perfect_ids.txt не найден!")
            return
        
        with open(ids_file, 'r') as f:
            project_ids = [int(line.strip()) for line in f if line.strip()]
        
        print("=" * 80)
        print("🔍 ПАРСИНГ ШАБЛОНА 4 (ИДЕАЛЬНЫЕ ПРОЕКТЫ)")
        print("=" * 80)
        print(f"\nВсего проектов: {len(project_ids)}")
        print("")
        
        results = {
            'success': 0,
            'failed': 0,
            'total_products': 0,
            'total_offers': 0,
            'total_images': 0
        }
        
        for i, project_id in enumerate(project_ids):
            if i % 10 == 0:
                print(f"   Обработано: {i}/{len(project_ids)}...")
            
            result = self.parse_project(project_id)
            
            if result['success']:
                results['success'] += 1
                results['total_products'] += result['products']
                results['total_offers'] += result['total_offers']
                results['total_images'] += result['total_images']
                
                # Обновляем статус в БД
                with self.db.get_session() as session:
                    session.execute(text("""
                        UPDATE projects 
                        SET parsing_status = 'completed',
                            parsed_at = :parsed_at
                        WHERE id = :pid
                    """), {
                        'pid': project_id,
                        'parsed_at': datetime.now().isoformat()
                    })
                    session.commit()
            else:
                results['failed'] += 1
                print(f"      ❌ Project {project_id}: {result.get('error', 'Unknown error')}")
        
        print(f"   Обработано: {len(project_ids)}/{len(project_ids)} ✅")
        print("")
        
        # Итоги
        print("=" * 80)
        print("📊 РЕЗУЛЬТАТЫ ПАРСИНГА")
        print("=" * 80)
        print("")
        print(f"✅ Успешно:      {results['success']:4d} ({results['success']/len(project_ids)*100:5.1f}%)")
        print(f"❌ Ошибки:       {results['failed']:4d} ({results['failed']/len(project_ids)*100:5.1f}%)")
        print("")
        print(f"📦 Товары:       {results['total_products']:,}")
        print(f"💰 Предложения:  {results['total_offers']:,}")
        print(f"🖼️ Изображения:   {results['total_images']:,}")
        print("")
        print("=" * 80)


if __name__ == '__main__':
    parser = Template4Parser()
    parser.parse_all_perfect()

