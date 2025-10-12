#!/usr/bin/env python3
"""
Упрощенная автоматическая проверка тиражей
Использует публичные ссылки для скачивания
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path.cwd()))

import json
import requests
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text
import os
from openpyxl import load_workbook
import csv
from datetime import datetime
import time

# Railway DB
os.environ['USE_RAILWAY_DB'] = 'true'

def download_excel_simple(sheet_id, output_path):
    """Скачивает Google Sheet как Excel через публичную ссылку"""
    try:
        # Используем публичную ссылку для экспорта
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid=0"
        
        response = requests.get(export_url, timeout=30)
        
        if response.status_code == 200:
            with open(output_path, 'wb') as f:
                f.write(response.content)
            return True
        else:
            print(f"  ❌ Код ответа: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"  ❌ Ошибка: {e}")
        return False

def extract_quantity_from_excel(excel_path, row_number):
    """Извлекает тираж из колонки E"""
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        
        # Колонка E = 5
        cell_value = ws.cell(row=row_number, column=5).value
        
        if cell_value is None:
            return None
            
        # Пытаемся преобразовать в число
        try:
            return int(float(str(cell_value).replace(',', '').replace(' ', '')))
        except:
            return str(cell_value)
            
    except Exception as e:
        return None

def verify_project(project_data, db, temp_dir):
    """Проверяет один проект"""
    project_id = project_data['id']
    sheet_id = project_data['sheet_id']
    
    print(f"\n{'='*80}")
    print(f"🔍 ID {project_id}: {project_data['name'][:60]}")
    print(f"{'='*80}")
    
    # Скачиваем
    excel_path = temp_dir / f"project_{project_id}.xlsx"
    print(f"📥 Скачиваю...", end=' ', flush=True)
    
    if not download_excel_simple(sheet_id, excel_path):
        return None
    
    size_kb = excel_path.stat().st_size / 1024
    print(f"✅ {size_kb:.0f} KB")
    
    # Получаем товары из БД
    with db.get_session() as session:
        products = session.execute(text("""
            SELECT 
                p.id, p.name, p.row_number,
                po.quantity, po.route
            FROM products p
            INNER JOIN price_offers po ON p.id = po.product_id
            WHERE p.project_id = :pid
            AND po.quantity >= 300
            AND po.quantity % 10 = 0
            ORDER BY p.row_number
        """), {'pid': project_id}).fetchall()
        
        if not products:
            print(f"⚠️  Нет подозрительных товаров")
            return None
        
        print(f"📊 Проверяю {len(products)} записей...", end=' ', flush=True)
        
        results = []
        checked_rows = {}
        
        for prod in products:
            row_num = prod.row_number
            
            if row_num not in checked_rows:
                checked_rows[row_num] = extract_quantity_from_excel(excel_path, row_num)
            
            excel_qty = checked_rows[row_num]
            db_qty = int(prod.quantity)
            
            is_x10 = False
            status = "✓"
            
            if excel_qty is None:
                status = "⚠️"
            elif isinstance(excel_qty, str):
                status = "⚠️"
            elif excel_qty * 10 == db_qty:
                is_x10 = True
                status = "❌"
            elif excel_qty == db_qty:
                status = "✅"
            else:
                status = "⚠️"
            
            results.append({
                'product_id': prod.id,
                'product_name': prod.name,
                'row_number': row_num,
                'excel_quantity': excel_qty,
                'db_quantity': db_qty,
                'route': prod.route,
                'is_x10': is_x10,
                'status': status
            })
        
        # Статистика
        x10_count = sum(1 for r in results if r['is_x10'])
        ok_count = sum(1 for r in results if r['status'] == "✅")
        
        print(f"Done!")
        print(f"  ❌ Ошибка x10: {x10_count}/{len(results)} ({x10_count/len(results)*100:.0f}%)")
        print(f"  ✅ Правильно: {ok_count}/{len(results)} ({ok_count/len(results)*100:.0f}%)")
        
        # Примеры
        if x10_count > 0:
            print(f"  📋 Примеры ошибок:")
            for r in results[:3]:
                if r['is_x10']:
                    print(f"    Строка {r['row_number']}: {r['excel_quantity']} → {r['db_quantity']}")
        
        return {
            'project_id': project_id,
            'project_name': project_data['name'],
            'url': project_data['url'],
            'total_records': len(results),
            'x10_errors': x10_count,
            'ok_count': ok_count,
            'results': results
        }

def main():
    print("="*80)
    print("🔍 АВТОМАТИЧЕСКАЯ ПРОВЕРКА ТИРАЖЕЙ (Упрощенная версия)")
    print("="*80)
    
    # Загружаем список
    with open('projects_to_verify.json', 'r', encoding='utf-8') as f:
        projects = json.load(f)
    
    print(f"\n📋 Проектов для проверки: {len(projects)}")
    print(f"⏱️  Проверяю первые 10 проектов...")
    
    temp_dir = Path('temp_verification')
    temp_dir.mkdir(exist_ok=True)
    
    db = PostgreSQLManager()
    
    all_results = []
    
    for i, project in enumerate(projects[:10], 1):
        print(f"\n{'#'*80}")
        print(f"# {i}/10")
        
        result = verify_project(project, db, temp_dir)
        if result:
            all_results.append(result)
        
        time.sleep(1)  # Пауза между запросами
    
    # Итог
    print("\n\n" + "="*80)
    print("📊 ИТОГОВЫЙ ОТЧЕТ")
    print("="*80)
    
    if not all_results:
        print("\n⚠️  Не удалось проверить ни одного проекта")
        return
    
    total_x10 = sum(r['x10_errors'] for r in all_results)
    total_records = sum(r['total_records'] for r in all_results)
    
    print(f"\n✅ Проверено проектов: {len(all_results)}")
    print(f"📊 Всего записей: {total_records:,}")
    print(f"❌ Ошибок x10: {total_x10:,} ({total_x10/total_records*100:.1f}%)")
    
    print(f"\n📋 ДЕТАЛИ:")
    for r in all_results:
        percent = r['x10_errors'] / r['total_records'] * 100 if r['total_records'] > 0 else 0
        status_icon = "❌" if r['x10_errors'] > 0 else "✅"
        print(f"\n{status_icon} ID {r['project_id']}: {r['project_name'][:55]}")
        print(f"    Ошибок: {r['x10_errors']}/{r['total_records']} ({percent:.0f}%)")
    
    # Сохраняем CSV
    csv_path = Path(f"АВТОПРОВЕРКА_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'ID Проекта', 'Название', 'ID Товара', 'Товар',
            'Строка', 'Excel', 'БД', 'Маршрут', 'Статус', 'x10?'
        ])
        
        for proj in all_results:
            for r in proj['results']:
                writer.writerow([
                    proj['project_id'], proj['project_name'],
                    r['product_id'], r['product_name'],
                    r['row_number'], r['excel_quantity'], r['db_quantity'],
                    r['route'], r['status'], 'ДА' if r['is_x10'] else 'НЕТ'
                ])
    
    print(f"\n✅ Детальный отчет: {csv_path}")
    print("="*80)

if __name__ == '__main__':
    main()




