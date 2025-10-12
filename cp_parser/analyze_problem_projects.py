#!/usr/bin/env python3
"""
Анализ проблемных проектов Template 6
Скачивает и анализирует структуру файлов
"""

import sys
from pathlib import Path
from sqlalchemy import text
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from database.postgresql_manager import PostgreSQLManager
from download_google_sheet import download_sheet_as_excel

# Проблемные проекты для анализа
PROBLEM_PROJECTS = [860, 846, 794, 292]

def analyze_project(project_id):
    """Анализирует структуру Excel файла проекта"""
    
    db = PostgreSQLManager()
    
    # Получаем table_id проекта
    with db.get_session() as session:
        table_id = session.execute(text("""
            SELECT table_id FROM projects WHERE id = :id
        """), {'id': project_id}).scalar()
    
    if not table_id:
        print(f"❌ Проект {project_id}: table_id не найден")
        return
    
    # Скачиваем файл
    print(f"\n{'='*80}")
    print(f"📥 ПРОЕКТ {project_id} (table_id: {table_id})")
    print(f"{'='*80}")
    
    excel_path = Path(f"excel_files/project_{project_id}_{table_id}.xlsx")
    
    if not excel_path.exists():
        print(f"📥 Скачиваю файл...")
        try:
            download_sheet_as_excel(table_id, str(excel_path))
        except Exception as e:
            print(f"❌ Ошибка скачивания: {e}")
            return
    else:
        print(f"✅ Файл уже есть")
    
    # Анализируем структуру
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        print(f"\n📋 СТРУКТУРА ФАЙЛА:")
        print(f"  Строк: {ws.max_row}")
        print(f"  Столбцов: {ws.max_column}")
        
        # Строка 2 (заголовки)
        print(f"\n📌 СТРОКА 2 (заголовки):")
        headers = []
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(2, col).value
            if val:
                headers.append((col, str(val)[:40]))
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Строка 3 (подзаголовки)
        print(f"\n📌 СТРОКА 3 (подзаголовки):")
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(3, col).value
            if val:
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Строка 4 (первые данные)
        print(f"\n📌 СТРОКА 4 (первые данные):")
        for col in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(4, col).value
            if val:
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Проверяем изображения
        print(f"\n🖼️  ИЗОБРАЖЕНИЯ:")
        if hasattr(ws, '_images'):
            print(f"  Всего: {len(ws._images)}")
            if ws._images:
                for i, img in enumerate(ws._images[:3], 1):
                    if hasattr(img.anchor, '_from'):
                        row = img.anchor._from.row + 1
                        col = img.anchor._from.col + 1
                        print(f"  {i}. Строка {row}, Колонка {col}")
        else:
            print(f"  ❌ Нет изображений!")
        
        # Товары в БД
        with db.get_session() as session:
            product_count = session.execute(text("""
                SELECT COUNT(*) FROM products WHERE project_id = :id
            """), {'id': project_id}).scalar()
            
            print(f"\n📊 В БАЗЕ ДАННЫХ:")
            print(f"  Товары: {product_count}")
        
    except Exception as e:
        print(f"❌ Ошибка анализа: {e}")

def main():
    print("=" * 80)
    print("🔍 АНАЛИЗ ПРОБЛЕМНЫХ ПРОЕКТОВ TEMPLATE 6")
    print("=" * 80)
    
    for project_id in PROBLEM_PROJECTS:
        try:
            analyze_project(project_id)
        except Exception as e:
            print(f"\n❌ Ошибка проекта {project_id}: {e}")
    
    print("\n" + "=" * 80)
    print("✅ АНАЛИЗ ЗАВЕРШЕН")
    print("=" * 80)

if __name__ == '__main__':
    main()



"""
Анализ проблемных проектов Template 6
Скачивает и анализирует структуру файлов
"""

import sys
from pathlib import Path
from sqlalchemy import text
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from database.postgresql_manager import PostgreSQLManager
from download_google_sheet import download_sheet_as_excel

# Проблемные проекты для анализа
PROBLEM_PROJECTS = [860, 846, 794, 292]

def analyze_project(project_id):
    """Анализирует структуру Excel файла проекта"""
    
    db = PostgreSQLManager()
    
    # Получаем table_id проекта
    with db.get_session() as session:
        table_id = session.execute(text("""
            SELECT table_id FROM projects WHERE id = :id
        """), {'id': project_id}).scalar()
    
    if not table_id:
        print(f"❌ Проект {project_id}: table_id не найден")
        return
    
    # Скачиваем файл
    print(f"\n{'='*80}")
    print(f"📥 ПРОЕКТ {project_id} (table_id: {table_id})")
    print(f"{'='*80}")
    
    excel_path = Path(f"excel_files/project_{project_id}_{table_id}.xlsx")
    
    if not excel_path.exists():
        print(f"📥 Скачиваю файл...")
        try:
            download_sheet_as_excel(table_id, str(excel_path))
        except Exception as e:
            print(f"❌ Ошибка скачивания: {e}")
            return
    else:
        print(f"✅ Файл уже есть")
    
    # Анализируем структуру
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        print(f"\n📋 СТРУКТУРА ФАЙЛА:")
        print(f"  Строк: {ws.max_row}")
        print(f"  Столбцов: {ws.max_column}")
        
        # Строка 2 (заголовки)
        print(f"\n📌 СТРОКА 2 (заголовки):")
        headers = []
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(2, col).value
            if val:
                headers.append((col, str(val)[:40]))
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Строка 3 (подзаголовки)
        print(f"\n📌 СТРОКА 3 (подзаголовки):")
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(3, col).value
            if val:
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Строка 4 (первые данные)
        print(f"\n📌 СТРОКА 4 (первые данные):")
        for col in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(4, col).value
            if val:
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Проверяем изображения
        print(f"\n🖼️  ИЗОБРАЖЕНИЯ:")
        if hasattr(ws, '_images'):
            print(f"  Всего: {len(ws._images)}")
            if ws._images:
                for i, img in enumerate(ws._images[:3], 1):
                    if hasattr(img.anchor, '_from'):
                        row = img.anchor._from.row + 1
                        col = img.anchor._from.col + 1
                        print(f"  {i}. Строка {row}, Колонка {col}")
        else:
            print(f"  ❌ Нет изображений!")
        
        # Товары в БД
        with db.get_session() as session:
            product_count = session.execute(text("""
                SELECT COUNT(*) FROM products WHERE project_id = :id
            """), {'id': project_id}).scalar()
            
            print(f"\n📊 В БАЗЕ ДАННЫХ:")
            print(f"  Товары: {product_count}")
        
    except Exception as e:
        print(f"❌ Ошибка анализа: {e}")

def main():
    print("=" * 80)
    print("🔍 АНАЛИЗ ПРОБЛЕМНЫХ ПРОЕКТОВ TEMPLATE 6")
    print("=" * 80)
    
    for project_id in PROBLEM_PROJECTS:
        try:
            analyze_project(project_id)
        except Exception as e:
            print(f"\n❌ Ошибка проекта {project_id}: {e}")
    
    print("\n" + "=" * 80)
    print("✅ АНАЛИЗ ЗАВЕРШЕН")
    print("=" * 80)

if __name__ == '__main__':
    main()










