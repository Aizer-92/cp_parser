#!/usr/bin/env python3
"""
Детальный анализ причин невалидности pending проектов
"""
import sys
from pathlib import Path
from collections import defaultdict
import json
from openpyxl import load_workbook

sys.path.insert(0, str(Path.cwd()))
from src.structure_parser import CommercialProposalParser
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text

class PendingProjectsAnalyzer:
    def __init__(self):
        self.parser = CommercialProposalParser()
        self.db = PostgreSQLManager()
        self.results = []
        
    def get_pending_projects(self):
        """Получает список pending проектов из БД"""
        with self.db.get_session() as session:
            projects = session.execute(text("""
                SELECT id, file_name, file_path 
                FROM projects 
                WHERE parsing_status = 'pending'
                ORDER BY id
            """)).fetchall()
        
        return projects
    
    def analyze_file_structure(self, file_path):
        """Детальный анализ структуры файла"""
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            
            # Читаем первые 10 строк
            headers = {}
            for row in range(1, 11):
                row_data = {}
                for col in range(1, 30):  # Первые 30 колонок
                    cell_value = ws.cell(row, col).value
                    if cell_value:
                        col_letter = chr(64 + col)  # A, B, C...
                        row_data[col_letter] = str(cell_value).strip()
                
                if row_data:
                    headers[f"row_{row}"] = row_data
            
            wb.close()
            return headers
            
        except Exception as e:
            return {"error": str(e)}
    
    def detect_missing_columns(self, headers):
        """Определяет какие обязательные колонки отсутствуют"""
        missing = []
        
        # Ищем колонку "Фото" или "Photo"
        has_photo = False
        for row_key, row_data in headers.items():
            for col, value in row_data.items():
                if any(word in value.lower() for word in ['фото', 'photo', 'изображ']):
                    has_photo = True
                    break
        
        if not has_photo:
            missing.append("Photo column (A)")
        
        # Ищем колонку "Наименование"
        has_name = False
        for row_key, row_data in headers.items():
            for col, value in row_data.items():
                if any(word in value.lower() for word in ['наименование', 'название', 'товар', 'name', 'product']):
                    has_name = True
                    break
        
        if not has_name:
            missing.append("Name column (B)")
        
        # Ищем маршруты
        has_route = False
        for row_key, row_data in headers.items():
            for col, value in row_data.items():
                if any(word in value.lower() for word in ['жд', 'авиа', 'маршрут', 'доставка']):
                    has_route = True
                    break
        
        if not has_route:
            missing.append("Route columns")
        
        # Ищем тираж
        has_quantity = False
        for row_key, row_data in headers.items():
            for col, value in row_data.items():
                if any(word in value.lower() for word in ['тираж', 'количество', 'кол-во', 'qty']):
                    has_quantity = True
                    break
        
        if not has_quantity:
            missing.append("Quantity column")
        
        # Ищем цены
        has_price = False
        for row_key, row_data in headers.items():
            for col, value in row_data.items():
                if any(word in value.lower() for word in ['цена', 'price', 'usd', 'руб', 'rub', 'стоимость']):
                    has_price = True
                    break
        
        if not has_price:
            missing.append("Price columns")
        
        return missing
    
    def categorize_failure(self, headers, missing_columns):
        """Категоризирует причину невалидности"""
        # Проверяем специфические паттерны
        
        # Категория 1: Пустой файл или только заголовки
        total_cells = sum(len(row_data) for row_data in headers.values())
        if total_cells < 5:
            return "empty_or_minimal", "Пустой файл или минимум данных"
        
        # Категория 2: Отсутствует фото колонка
        if "Photo column (A)" in missing_columns:
            return "no_photo_column", "Отсутствует колонка 'Фото'"
        
        # Категория 3: Отсутствует название
        if "Name column (B)" in missing_columns:
            return "no_name_column", "Отсутствует колонка 'Наименование'"
        
        # Категория 4: Нет маршрутов доставки
        if "Route columns" in missing_columns:
            return "no_routes", "Отсутствуют маршруты доставки"
        
        # Категория 5: Нет тиражей
        if "Quantity column" in missing_columns:
            return "no_quantity", "Отсутствует колонка 'Тираж'"
        
        # Категория 6: Нет цен
        if "Price columns" in missing_columns:
            return "no_prices", "Отсутствуют цены"
        
        # Категория 7: Нестандартная структура
        return "non_standard", "Нестандартная структура (все колонки есть, но не проходит валидацию)"
    
    def analyze_all_pending(self):
        """Анализирует все pending проекты"""
        projects = self.get_pending_projects()
        
        print(f"🔍 Анализ {len(projects)} pending проектов...")
        print("")
        
        categories = defaultdict(list)
        
        for i, (project_id, file_name, file_path) in enumerate(projects):
            if i % 50 == 0:
                print(f"   Обработано: {i}/{len(projects)}...")
            
            # Ищем файл по project_id (формат: project_{id}_*.xlsx)
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx'))
            
            if not matching_files:
                categories['file_not_found'].append({
                    'project_id': project_id,
                    'file_name': file_name,
                    'reason': 'Файл не найден локально'
                })
                continue
            
            # Используем первый найденный файл
            excel_path = matching_files[0]
            
            # Анализируем структуру
            headers = self.analyze_file_structure(str(excel_path))
            
            if 'error' in headers:
                categories['parse_error'].append({
                    'project_id': project_id,
                    'file_name': file_name,
                    'reason': f"Ошибка чтения: {headers['error']}"
                })
                continue
            
            # Определяем отсутствующие колонки
            missing = self.detect_missing_columns(headers)
            
            # Категоризируем
            category, reason = self.categorize_failure(headers, missing)
            
            categories[category].append({
                'project_id': project_id,
                'file_name': file_name,
                'reason': reason,
                'missing_columns': missing,
                'sample_headers': {
                    k: v for k, v in list(headers.items())[:3]  # Первые 3 строки
                }
            })
        
        print(f"   Обработано: {len(projects)}/{len(projects)} ✅")
        print("")
        
        return categories
    
    def generate_report(self, categories):
        """Генерирует детальный отчет"""
        
        # Markdown отчет
        report_md = []
        report_md.append("# 🔍 Детальный анализ pending проектов")
        report_md.append("")
        report_md.append(f"**Дата**: {Path.cwd().name}")
        report_md.append(f"**Всего pending проектов**: {sum(len(items) for items in categories.values())}")
        report_md.append("")
        report_md.append("---")
        report_md.append("")
        
        # Сортируем категории по количеству проектов
        sorted_categories = sorted(
            categories.items(),
            key=lambda x: len(x[1]),
            reverse=True
        )
        
        # Оглавление
        report_md.append("## 📊 Сводка по категориям")
        report_md.append("")
        report_md.append("| Категория | Количество | % |")
        report_md.append("|-----------|------------|---|")
        
        total = sum(len(items) for items in categories.values())
        for category, items in sorted_categories:
            percentage = len(items) / total * 100 if total > 0 else 0
            category_name = self._get_category_name(category)
            report_md.append(f"| {category_name} | {len(items)} | {percentage:.1f}% |")
        
        report_md.append("")
        report_md.append("---")
        report_md.append("")
        
        # Детали по каждой категории
        for category, items in sorted_categories:
            category_name = self._get_category_name(category)
            report_md.append(f"## {category_name} ({len(items)} проектов)")
            report_md.append("")
            
            # Первые 10 примеров
            for i, item in enumerate(items[:10]):
                report_md.append(f"### Пример {i+1}: Project {item['project_id']}")
                report_md.append("")
                report_md.append(f"**Файл**: `{item['file_name']}`")
                report_md.append(f"**Причина**: {item['reason']}")
                
                if item.get('missing_columns'):
                    report_md.append(f"**Отсутствует**: {', '.join(item['missing_columns'])}")
                
                if item.get('sample_headers'):
                    report_md.append("")
                    report_md.append("**Заголовки (первые 3 строки)**:")
                    report_md.append("```")
                    for row_key, row_data in item['sample_headers'].items():
                        report_md.append(f"{row_key}:")
                        for col, val in row_data.items():
                            report_md.append(f"  {col}: {val[:50]}...")
                    report_md.append("```")
                
                report_md.append("")
            
            if len(items) > 10:
                report_md.append(f"**...и еще {len(items) - 10} проектов**")
                report_md.append("")
                report_md.append("**Полный список ID**:")
                ids = [str(item['project_id']) for item in items]
                report_md.append(f"```{', '.join(ids)}```")
                report_md.append("")
            
            report_md.append("---")
            report_md.append("")
        
        # Сохраняем отчет
        report_path = Path('PENDING_ANALYSIS_REPORT.md')
        report_path.write_text('\n'.join(report_md), encoding='utf-8')
        
        # JSON для программной обработки
        json_data = {
            category: [
                {
                    'project_id': item['project_id'],
                    'file_name': item['file_name'],
                    'reason': item['reason'],
                    'missing_columns': item.get('missing_columns', [])
                }
                for item in items
            ]
            for category, items in categories.items()
        }
        
        json_path = Path('pending_analysis.json')
        json_path.write_text(json.dumps(json_data, indent=2, ensure_ascii=False), encoding='utf-8')
        
        print("=" * 80)
        print("✅ АНАЛИЗ ЗАВЕРШЕН")
        print("=" * 80)
        print("")
        print(f"📄 Markdown отчет: {report_path}")
        print(f"📊 JSON данные: {json_path}")
        print("")
        print("📊 СВОДКА ПО КАТЕГОРИЯМ:")
        print("")
        
        for category, items in sorted_categories:
            category_name = self._get_category_name(category)
            percentage = len(items) / total * 100 if total > 0 else 0
            print(f"   {category_name:50s}: {len(items):4d} ({percentage:5.1f}%)")
        
        print("")
        print("=" * 80)
    
    def _get_category_name(self, category):
        """Получает читаемое название категории"""
        names = {
            'empty_or_minimal': '🗑️ Пустые или минимальные данные',
            'no_photo_column': '📷 Отсутствует колонка "Фото"',
            'no_name_column': '📝 Отсутствует колонка "Наименование"',
            'no_routes': '🚚 Отсутствуют маршруты доставки',
            'no_quantity': '📦 Отсутствует колонка "Тираж"',
            'no_prices': '💰 Отсутствуют цены',
            'non_standard': '🔧 Нестандартная структура',
            'file_not_found': '❌ Файл не найден',
            'parse_error': '⚠️ Ошибка чтения файла'
        }
        return names.get(category, category)

if __name__ == '__main__':
    analyzer = PendingProjectsAnalyzer()
    
    print("=" * 80)
    print("🔍 АНАЛИЗ PENDING ПРОЕКТОВ")
    print("=" * 80)
    print("")
    
    categories = analyzer.analyze_all_pending()
    analyzer.generate_report(categories)

