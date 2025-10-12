#!/usr/bin/env python3
"""
Парсер для Шаблона 5 (230 проектов)

Отличия от Шаблона 4:
- Заголовки в строке 2 (вместо 4)
- Единая цена (без разделения ЖД/Авиа)
- Нет отдельной колонки "Кастом"
- Образец в колонках H, I (вместо N, P)
- Доп. фото в колонке J (вместо Q)
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import re
import hashlib
from io import BytesIO

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template5Parser:
    """
    Парсер для Шаблона 5
    
    Особенности:
    - Заголовки в строке 2
    - Единая цена (колонки E, F, G)
    - Multi-row товары (доп. опции без названия)
    - Образец в H, I
    """
    
    DATA_START_ROW = 3  # Строка 2 = заголовки, строка 3 = данные
    
    # Маппинг колонок
    COLUMNS = {
        'photo': 1,           # A - главное фото
        'name': 2,            # B - наименование
        'description': 3,     # C - описание (включает дизайн)
        'quantity': 4,        # D - тираж (может быть несколько)
        
        # Единая цена (без разделения на маршруты)
        'price_usd': 5,       # E - цена USD
        'price_rub': 6,       # F - цена RUB
        'delivery': 7,        # G - срок производства + доставки
        
        # Образец
        'sample_price': 8,    # H - цена образца
        'sample_delivery': 9, # I - срок образца
        
        'extra_photo': 10     # J - дополнительные фото
    }
    
    def __init__(self):
        self.db = PostgreSQLManager()
        self.storage_dir = Path('storage/images')
        self.storage_dir.mkdir(parents=True, exist_ok=True)
    
    def parse_project(self, project_id):
        """Парсит один проект"""
        try:
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx'))
            
            if not matching_files:
                return {'success': False, 'error': 'File not found'}
            
            file_path = matching_files[0]
            
            # Парсим
            wb_data = load_workbook(str(file_path), data_only=True)
            wb_images = load_workbook(str(file_path), data_only=False)
            
            # Берем первый лист (не "Цена")
            sheet_name = None
            for name in wb_data.sheetnames:
                if 'цена' not in name.lower():
                    sheet_name = name
                    break
            
            if not sheet_name:
                return {'success': False, 'error': 'No working sheet found'}
            
            ws_data = wb_data[sheet_name]
            ws_images = wb_images[sheet_name]
            
            # Получаем table_id из БД
            with self.db.get_session() as session:
                result = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).fetchone()
                
                if not result:
                    return {'success': False, 'error': 'Project not in DB'}
                
                table_id = result[0]
            
            # Парсим товары
            products = self._parse_products(ws_data, ws_images, project_id, table_id)
            
            wb_data.close()
            wb_images.close()
            
            return {
                'success': True,
                'products': len(products),
                'total_offers': sum(len(p['offers']) for p in products),
                'total_images': sum(len(p['images']) for p in products)
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def _parse_products(self, ws_data, ws_images, project_id, table_id):
        """Парсит все товары из листа"""
        products = []
        current_product = None
        
        max_row = ws_data.max_row if ws_data.max_row else 1000
        
        for row in range(self.DATA_START_ROW, min(max_row + 1, 1000)):
            name = self._get_cell_value(ws_data, row, self.COLUMNS['name'])
            
            # Проверяем на multi-row (дополнительная строка для товара)
            has_prices = (
                self._get_cell_value(ws_data, row, self.COLUMNS['price_usd']) or
                self._get_cell_value(ws_data, row, self.COLUMNS['price_rub'])
            )
            has_description = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
            
            # Пропускаем полностью пустые строки
            if not name and not has_prices and not has_description:
                continue
            
            # Новый товар (есть наименование)
            if name and len(str(name).strip()) >= 3:
                # Сохраняем предыдущий товар
                if current_product:
                    self._save_product(current_product, project_id, table_id)
                    products.append(current_product)
                
                # Создаем новый товар
                current_product = self._create_product(ws_data, ws_images, row, table_id)
            
            # Дополнительная строка для текущего товара
            elif current_product and (has_prices or has_description):
                # Добавляем дополнительное описание
                add_desc = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
                if add_desc:
                    if current_product['custom_field']:
                        current_product['custom_field'] += "\n" + str(add_desc).strip()
                    else:
                        current_product['custom_field'] = str(add_desc).strip()
                
                # Добавляем дополнительные офферы
                if has_prices:
                    offers = self._parse_offers(ws_data, row, current_product.get('last_quantity'))
                    current_product['offers'].extend(offers)
                
                # Обновляем данные об образце если есть
                sample_update = self._parse_sample(ws_data, row)
                if sample_update:
                    if current_product.get('sample_price'):
                        current_product['sample_price'] += " / " + sample_update['price']
                    else:
                        current_product['sample_price'] = sample_update['price']
                    
                    if sample_update['delivery_time']:
                        current_product['sample_delivery_time'] = max(
                            current_product.get('sample_delivery_time', 0),
                            sample_update['delivery_time']
                        )
        
        # Сохраняем последний товар
        if current_product:
            self._save_product(current_product, project_id, table_id)
            products.append(current_product)
        
        return products
    
    def _create_product(self, ws_data, ws_images, row, table_id):
        """Создает структуру товара"""
        name = self._get_cell_value(ws_data, row, self.COLUMNS['name'])
        description = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
        
        # Парсим данные об образце
        sample_data = self._parse_sample(ws_data, row)
        
        # Извлекаем изображения
        images = self._extract_images(ws_images, row, table_id)
        
        # Парсим офферы
        offers = self._parse_offers(ws_data, row)
        
        # Запоминаем последний тираж для multi-row
        last_quantity = None
        if offers:
            last_quantity = offers[-1]['quantity']
        
        return {
            'name': str(name).strip(),
            'custom_field': str(description).strip() if description else None,
            'sample_price': sample_data.get('price') if sample_data else None,
            'sample_delivery_time': sample_data.get('delivery_time') if sample_data else None,
            'row_number': row,
            'offers': offers,
            'images': images,
            'last_quantity': last_quantity
        }
    
    def _parse_sample(self, ws_data, row):
        """Парсит данные об образце"""
        sample_price_raw = self._get_cell_value(ws_data, row, self.COLUMNS['sample_price'])
        sample_delivery_raw = self._get_cell_value(ws_data, row, self.COLUMNS['sample_delivery'])
        
        if not sample_price_raw or str(sample_price_raw).strip() == '-':
            return None
        
        # Парсим цену (может быть несколько значений через перенос строки)
        price_text = str(sample_price_raw).strip()
        
        # Парсим срок доставки
        delivery_time = self._parse_delivery_time(sample_delivery_raw) if sample_delivery_raw else None
        
        return {
            'price': price_text,
            'delivery_time': delivery_time
        }
    
    def _parse_offers(self, ws_data, row, fallback_quantity=None):
        """Парсит предложения из строки"""
        offers = []
        
        # Получаем данные из ячеек
        quantity_raw = self._get_cell_value(ws_data, row, self.COLUMNS['quantity'])
        price_usd_raw = self._get_cell_value(ws_data, row, self.COLUMNS['price_usd'])
        price_rub_raw = self._get_cell_value(ws_data, row, self.COLUMNS['price_rub'])
        delivery_raw = self._get_cell_value(ws_data, row, self.COLUMNS['delivery'])
        
        # Если нет цен - выходим
        if not price_usd_raw and not price_rub_raw:
            return offers
        
        # Парсим множественные значения из ячеек
        quantities = self._parse_multiline_values(quantity_raw) if quantity_raw else []
        prices_usd = self._parse_multiline_values(price_usd_raw) if price_usd_raw else []
        prices_rub = self._parse_multiline_values(price_rub_raw) if price_rub_raw else []
        deliveries = self._parse_multiline_values(delivery_raw) if delivery_raw else []
        
        # Если нет тиража, используем fallback (для multi-row)
        if not quantities and fallback_quantity:
            quantities = [fallback_quantity]
        
        # Группируем: сначала группы по пустым строкам, потом комбинируем
        # Пример: D="2000\n4000", E="0.49\n0.75\n\n0.46\n0.71"
        # Результат: (2000,0.49), (4000,0.75), (2000,0.46), (4000,0.71)
        
        # Простой подход: берем все комбинации
        max_len = max(len(quantities), len(prices_usd), len(prices_rub))
        
        for i in range(max_len):
            qty = quantities[i % len(quantities)] if quantities else fallback_quantity
            usd = prices_usd[i] if i < len(prices_usd) else None
            rub = prices_rub[i] if i < len(prices_rub) else None
            delivery_time = deliveries[i] if i < len(deliveries) else (deliveries[0] if deliveries else None)
            
            # Парсим количество
            if qty:
                try:
                    qty_int = self._parse_quantity(qty)
                    if not qty_int or qty_int <= 0 or qty_int > 1_000_000:
                        continue
                except:
                    continue
            else:
                continue
            
            # Парсим цены
            try:
                price_usd = float(str(usd).replace(',', '.').replace('$', '').strip()) if usd else None
                price_rub = float(str(rub).replace(',', '.').replace('₽', '').replace(' ', '').strip()) if rub else None
                
                # Проверка адекватности
                if price_usd and (price_usd <= 0 or price_usd > 10_000):
                    price_usd = None
                if price_rub and (price_rub <= 0 or price_rub > 1_000_000):
                    price_rub = None
                
                if not price_usd and not price_rub:
                    continue
                
            except:
                continue
            
            # Парсим срок
            delivery_days = self._parse_delivery_time(delivery_time) if delivery_time else None
            if delivery_days and delivery_days > 365:
                delivery_days = 60  # default
            
            offers.append({
                'quantity': qty_int,
                'price_usd': f"{price_usd:.2f}" if price_usd else None,
                'price_rub': f"{price_rub:.2f}" if price_rub else None,
                'route': None,  # Нет разделения на маршруты в Шаблоне 5
                'delivery_time_days': delivery_days or 45
            })
        
        return offers
    
    def _parse_multiline_values(self, cell_value):
        """Парсит множественные значения из ячейки, разделенные переносами"""
        if not cell_value:
            return []
        
        text = str(cell_value).strip()
        
        # Разделяем по переносам строк
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Фильтруем пустые и "None"
        values = [v for v in lines if v and v != 'None' and v != '-']
        
        return values
    
    def _extract_images(self, ws_images, row, table_id):
        """Извлекает изображения из строки"""
        images = []
        
        # Главное фото из колонки A
        main_image = self._extract_image_from_cell(ws_images, row, self.COLUMNS['photo'])
        if main_image:
            images.append({
                'data': main_image['data'],
                'filename': f"{table_id}_A{row}_{main_image['hash']}.png",
                'cell_position': f"A{row}",
                'is_main': True
            })
        
        # Дополнительные фото из колонки J (может быть несколько)
        extra_images = self._extract_all_images_from_cell(ws_images, row, self.COLUMNS['extra_photo'])
        for idx, img in enumerate(extra_images, 1):
            images.append({
                'data': img['data'],
                'filename': f"{table_id}_J{row}_{idx}_{img['hash']}.png",
                'cell_position': f"J{row}",
                'is_main': False
            })
        
        return images
    
    def _extract_image_from_cell(self, ws, row, col):
        """Извлекает ОДНО изображение из ячейки"""
        try:
            if not hasattr(ws, '_images'):
                return None
            
            for image in ws._images:
                if image.anchor and hasattr(image.anchor, 'from'):
                    if image.anchor._from.row == row - 1 and image.anchor._from.col == col - 1:
                        img_data = image._data()
                        if img_data:
                            img_hash = hashlib.md5(img_data).hexdigest()[:16]
                            return {'data': img_data, 'hash': img_hash}
        except:
            pass
        
        return None
    
    def _extract_all_images_from_cell(self, ws, row, col):
        """Извлекает ВСЕ изображения из ячейки"""
        images = []
        
        try:
            if not hasattr(ws, '_images'):
                return images
            
            for image in ws._images:
                if image.anchor and hasattr(image.anchor, '_from'):
                    if image.anchor._from.row == row - 1 and image.anchor._from.col == col - 1:
                        img_data = image._data()
                        if img_data:
                            img_hash = hashlib.md5(img_data).hexdigest()[:16]
                            images.append({'data': img_data, 'hash': img_hash})
        except:
            pass
        
        return images
    
    def _save_product(self, product, project_id, table_id):
        """Сохраняет товар в БД"""
        with self.db.get_session() as session:
            # Вставляем товар
            result = session.execute(text("""
                INSERT INTO products (
                    project_id, table_id, name, custom_field,
                    sample_price, sample_delivery_time, row_number, created_at, updated_at
                )
                VALUES (:project_id, :table_id, :name, :custom_field,
                        :sample_price, :sample_delivery_time, :row_number, NOW(), NOW())
                RETURNING id
            """), {
                'project_id': project_id,
                'table_id': table_id,
                'name': product['name'],
                'custom_field': product.get('custom_field'),
                'sample_price': product.get('sample_price'),
                'sample_delivery_time': product.get('sample_delivery_time'),
                'row_number': product['row_number']
            })
            
            product_id = result.fetchone()[0]
            
            # Сохраняем изображения
            for img in product['images']:
                # Сохраняем файл
                img_path = self.storage_dir / img['filename']
                with open(img_path, 'wb') as f:
                    f.write(img['data'])
                
                # Записываем в БД
                session.execute(text("""
                    INSERT INTO product_images (
                        product_id, table_id, image_filename, local_path,
                        cell_position, is_main_image, row_number, created_at, updated_at
                    )
                    VALUES (:product_id, :table_id, :filename, :local_path,
                            :cell_position, :is_main, :row_number, NOW(), NOW())
                """), {
                    'product_id': product_id,
                    'table_id': table_id,
                    'filename': img['filename'],
                    'local_path': str(img_path),
                    'cell_position': img['cell_position'],
                    'is_main': img['is_main'],
                    'row_number': product['row_number']
                })
            
            # Сохраняем офферы
            for offer in product['offers']:
                session.execute(text("""
                    INSERT INTO price_offers (
                        product_id, quantity, price_usd, price_rub,
                        route, delivery_time_days, created_at, updated_at
                    )
                    VALUES (:product_id, :quantity, :price_usd, :price_rub,
                            :route, :delivery_days, NOW(), NOW())
                """), {
                    'product_id': product_id,
                    'quantity': offer['quantity'],
                    'price_usd': offer['price_usd'],
                    'price_rub': offer['price_rub'],
                    'route': offer['route'],
                    'delivery_days': offer['delivery_time_days']
                })
            
            session.commit()
    
    def _get_cell_value(self, ws, row, col):
        """Получает значение ячейки"""
        try:
            cell = ws.cell(row, col)
            return cell.value if cell.value is not None else None
        except:
            return None
    
    def _parse_quantity(self, value):
        """Парсит количество"""
        if not value:
            return None
        
        val_str = str(value).strip().replace(' ', '').replace(',', '')
        
        # Убираем все кроме цифр и точки
        val_str = re.sub(r'[^\d.]', '', val_str)
        
        if not val_str:
            return None
        
        try:
            # Преобразуем через float чтобы обработать "10000.0"
            return int(float(val_str))
        except:
            return None
    
    def _parse_price(self, value):
        """Парсит цену"""
        if not value:
            return None
        
        val_str = str(value).strip()
        
        # Убираем символы валют и пробелы
        val_str = val_str.replace('$', '').replace('₽', '').replace(' ', '').replace(',', '.')
        
        try:
            return float(val_str)
        except:
            return None
    
    def _parse_delivery_time(self, value):
        """Парсит срок доставки, возвращает максимальное значение"""
        if not value:
            return None
        
        val_str = str(value).lower().strip()
        
        # Ищем все числа
        numbers = re.findall(r'\d+', val_str)
        
        if not numbers:
            return None
        
        # Возвращаем максимальное (например, "20-22 дня" → 22)
        return max(int(n) for n in numbers)


def main():
    """Парсит все проекты Шаблона 5"""
    print("=" * 100)
    print("🚀 ПАРСИНГ ШАБЛОНА 5")
    print("=" * 100)
    
    # Загружаем ID
    ids_file = Path('template_5_candidate_ids.txt')
    with open(ids_file, 'r') as f:
        project_ids = [int(line.strip()) for line in f if line.strip()]
    
    print(f"\nВсего проектов: {len(project_ids)}")
    
    parser = Template5Parser()
    
    success_count = 0
    error_count = 0
    total_products = 0
    total_offers = 0
    total_images = 0
    
    for i, project_id in enumerate(project_ids, 1):
        print(f"\n[{i}/{len(project_ids)}] Парсинг проекта {project_id}...", end=' ')
        
        result = parser.parse_project(project_id)
        
        if result['success']:
            print(f"✅ {result['products']} товаров, {result['total_offers']} офферов, {result['total_images']} изображений")
            success_count += 1
            total_products += result['products']
            total_offers += result['total_offers']
            total_images += result['total_images']
        else:
            print(f"❌ {result['error']}")
            error_count += 1
        
        # Прогресс каждые 10 проектов
        if i % 10 == 0:
            print(f"\n📊 Прогресс: {success_count} успешно, {error_count} ошибок")
    
    print("\n" + "=" * 100)
    print("📊 ИТОГИ:")
    print("=" * 100)
    print(f"✅ Успешно:      {success_count}")
    print(f"❌ Ошибок:       {error_count}")
    print(f"📦 Товаров:      {total_products}")
    print(f"💰 Офферов:      {total_offers}")
    print(f"🖼️  Изображений:  {total_images}")
    print("=" * 100)


if __name__ == '__main__':
    main()


Парсер для Шаблона 5 (230 проектов)

Отличия от Шаблона 4:
- Заголовки в строке 2 (вместо 4)
- Единая цена (без разделения ЖД/Авиа)
- Нет отдельной колонки "Кастом"
- Образец в колонках H, I (вместо N, P)
- Доп. фото в колонке J (вместо Q)
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import re
import hashlib
from io import BytesIO

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template5Parser:
    """
    Парсер для Шаблона 5
    
    Особенности:
    - Заголовки в строке 2
    - Единая цена (колонки E, F, G)
    - Multi-row товары (доп. опции без названия)
    - Образец в H, I
    """
    
    DATA_START_ROW = 3  # Строка 2 = заголовки, строка 3 = данные
    
    # Маппинг колонок
    COLUMNS = {
        'photo': 1,           # A - главное фото
        'name': 2,            # B - наименование
        'description': 3,     # C - описание (включает дизайн)
        'quantity': 4,        # D - тираж (может быть несколько)
        
        # Единая цена (без разделения на маршруты)
        'price_usd': 5,       # E - цена USD
        'price_rub': 6,       # F - цена RUB
        'delivery': 7,        # G - срок производства + доставки
        
        # Образец
        'sample_price': 8,    # H - цена образца
        'sample_delivery': 9, # I - срок образца
        
        'extra_photo': 10     # J - дополнительные фото
    }
    
    def __init__(self):
        self.db = PostgreSQLManager()
        self.storage_dir = Path('storage/images')
        self.storage_dir.mkdir(parents=True, exist_ok=True)
    
    def parse_project(self, project_id):
        """Парсит один проект"""
        try:
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx'))
            
            if not matching_files:
                return {'success': False, 'error': 'File not found'}
            
            file_path = matching_files[0]
            
            # Парсим
            wb_data = load_workbook(str(file_path), data_only=True)
            wb_images = load_workbook(str(file_path), data_only=False)
            
            # Берем первый лист (не "Цена")
            sheet_name = None
            for name in wb_data.sheetnames:
                if 'цена' not in name.lower():
                    sheet_name = name
                    break
            
            if not sheet_name:
                return {'success': False, 'error': 'No working sheet found'}
            
            ws_data = wb_data[sheet_name]
            ws_images = wb_images[sheet_name]
            
            # Получаем table_id из БД
            with self.db.get_session() as session:
                result = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).fetchone()
                
                if not result:
                    return {'success': False, 'error': 'Project not in DB'}
                
                table_id = result[0]
            
            # Парсим товары
            products = self._parse_products(ws_data, ws_images, project_id, table_id)
            
            wb_data.close()
            wb_images.close()
            
            return {
                'success': True,
                'products': len(products),
                'total_offers': sum(len(p['offers']) for p in products),
                'total_images': sum(len(p['images']) for p in products)
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def _parse_products(self, ws_data, ws_images, project_id, table_id):
        """Парсит все товары из листа"""
        products = []
        current_product = None
        
        max_row = ws_data.max_row if ws_data.max_row else 1000
        
        for row in range(self.DATA_START_ROW, min(max_row + 1, 1000)):
            name = self._get_cell_value(ws_data, row, self.COLUMNS['name'])
            
            # Проверяем на multi-row (дополнительная строка для товара)
            has_prices = (
                self._get_cell_value(ws_data, row, self.COLUMNS['price_usd']) or
                self._get_cell_value(ws_data, row, self.COLUMNS['price_rub'])
            )
            has_description = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
            
            # Пропускаем полностью пустые строки
            if not name and not has_prices and not has_description:
                continue
            
            # Новый товар (есть наименование)
            if name and len(str(name).strip()) >= 3:
                # Сохраняем предыдущий товар
                if current_product:
                    self._save_product(current_product, project_id, table_id)
                    products.append(current_product)
                
                # Создаем новый товар
                current_product = self._create_product(ws_data, ws_images, row, table_id)
            
            # Дополнительная строка для текущего товара
            elif current_product and (has_prices or has_description):
                # Добавляем дополнительное описание
                add_desc = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
                if add_desc:
                    if current_product['custom_field']:
                        current_product['custom_field'] += "\n" + str(add_desc).strip()
                    else:
                        current_product['custom_field'] = str(add_desc).strip()
                
                # Добавляем дополнительные офферы
                if has_prices:
                    offers = self._parse_offers(ws_data, row, current_product.get('last_quantity'))
                    current_product['offers'].extend(offers)
                
                # Обновляем данные об образце если есть
                sample_update = self._parse_sample(ws_data, row)
                if sample_update:
                    if current_product.get('sample_price'):
                        current_product['sample_price'] += " / " + sample_update['price']
                    else:
                        current_product['sample_price'] = sample_update['price']
                    
                    if sample_update['delivery_time']:
                        current_product['sample_delivery_time'] = max(
                            current_product.get('sample_delivery_time', 0),
                            sample_update['delivery_time']
                        )
        
        # Сохраняем последний товар
        if current_product:
            self._save_product(current_product, project_id, table_id)
            products.append(current_product)
        
        return products
    
    def _create_product(self, ws_data, ws_images, row, table_id):
        """Создает структуру товара"""
        name = self._get_cell_value(ws_data, row, self.COLUMNS['name'])
        description = self._get_cell_value(ws_data, row, self.COLUMNS['description'])
        
        # Парсим данные об образце
        sample_data = self._parse_sample(ws_data, row)
        
        # Извлекаем изображения
        images = self._extract_images(ws_images, row, table_id)
        
        # Парсим офферы
        offers = self._parse_offers(ws_data, row)
        
        # Запоминаем последний тираж для multi-row
        last_quantity = None
        if offers:
            last_quantity = offers[-1]['quantity']
        
        return {
            'name': str(name).strip(),
            'custom_field': str(description).strip() if description else None,
            'sample_price': sample_data.get('price') if sample_data else None,
            'sample_delivery_time': sample_data.get('delivery_time') if sample_data else None,
            'row_number': row,
            'offers': offers,
            'images': images,
            'last_quantity': last_quantity
        }
    
    def _parse_sample(self, ws_data, row):
        """Парсит данные об образце"""
        sample_price_raw = self._get_cell_value(ws_data, row, self.COLUMNS['sample_price'])
        sample_delivery_raw = self._get_cell_value(ws_data, row, self.COLUMNS['sample_delivery'])
        
        if not sample_price_raw or str(sample_price_raw).strip() == '-':
            return None
        
        # Парсим цену (может быть несколько значений через перенос строки)
        price_text = str(sample_price_raw).strip()
        
        # Парсим срок доставки
        delivery_time = self._parse_delivery_time(sample_delivery_raw) if sample_delivery_raw else None
        
        return {
            'price': price_text,
            'delivery_time': delivery_time
        }
    
    def _parse_offers(self, ws_data, row, fallback_quantity=None):
        """Парсит предложения из строки"""
        offers = []
        
        # Получаем данные из ячеек
        quantity_raw = self._get_cell_value(ws_data, row, self.COLUMNS['quantity'])
        price_usd_raw = self._get_cell_value(ws_data, row, self.COLUMNS['price_usd'])
        price_rub_raw = self._get_cell_value(ws_data, row, self.COLUMNS['price_rub'])
        delivery_raw = self._get_cell_value(ws_data, row, self.COLUMNS['delivery'])
        
        # Если нет цен - выходим
        if not price_usd_raw and not price_rub_raw:
            return offers
        
        # Парсим множественные значения из ячеек
        quantities = self._parse_multiline_values(quantity_raw) if quantity_raw else []
        prices_usd = self._parse_multiline_values(price_usd_raw) if price_usd_raw else []
        prices_rub = self._parse_multiline_values(price_rub_raw) if price_rub_raw else []
        deliveries = self._parse_multiline_values(delivery_raw) if delivery_raw else []
        
        # Если нет тиража, используем fallback (для multi-row)
        if not quantities and fallback_quantity:
            quantities = [fallback_quantity]
        
        # Группируем: сначала группы по пустым строкам, потом комбинируем
        # Пример: D="2000\n4000", E="0.49\n0.75\n\n0.46\n0.71"
        # Результат: (2000,0.49), (4000,0.75), (2000,0.46), (4000,0.71)
        
        # Простой подход: берем все комбинации
        max_len = max(len(quantities), len(prices_usd), len(prices_rub))
        
        for i in range(max_len):
            qty = quantities[i % len(quantities)] if quantities else fallback_quantity
            usd = prices_usd[i] if i < len(prices_usd) else None
            rub = prices_rub[i] if i < len(prices_rub) else None
            delivery_time = deliveries[i] if i < len(deliveries) else (deliveries[0] if deliveries else None)
            
            # Парсим количество
            if qty:
                try:
                    qty_int = self._parse_quantity(qty)
                    if not qty_int or qty_int <= 0 or qty_int > 1_000_000:
                        continue
                except:
                    continue
            else:
                continue
            
            # Парсим цены
            try:
                price_usd = float(str(usd).replace(',', '.').replace('$', '').strip()) if usd else None
                price_rub = float(str(rub).replace(',', '.').replace('₽', '').replace(' ', '').strip()) if rub else None
                
                # Проверка адекватности
                if price_usd and (price_usd <= 0 or price_usd > 10_000):
                    price_usd = None
                if price_rub and (price_rub <= 0 or price_rub > 1_000_000):
                    price_rub = None
                
                if not price_usd and not price_rub:
                    continue
                
            except:
                continue
            
            # Парсим срок
            delivery_days = self._parse_delivery_time(delivery_time) if delivery_time else None
            if delivery_days and delivery_days > 365:
                delivery_days = 60  # default
            
            offers.append({
                'quantity': qty_int,
                'price_usd': f"{price_usd:.2f}" if price_usd else None,
                'price_rub': f"{price_rub:.2f}" if price_rub else None,
                'route': None,  # Нет разделения на маршруты в Шаблоне 5
                'delivery_time_days': delivery_days or 45
            })
        
        return offers
    
    def _parse_multiline_values(self, cell_value):
        """Парсит множественные значения из ячейки, разделенные переносами"""
        if not cell_value:
            return []
        
        text = str(cell_value).strip()
        
        # Разделяем по переносам строк
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Фильтруем пустые и "None"
        values = [v for v in lines if v and v != 'None' and v != '-']
        
        return values
    
    def _extract_images(self, ws_images, row, table_id):
        """Извлекает изображения из строки"""
        images = []
        
        # Главное фото из колонки A
        main_image = self._extract_image_from_cell(ws_images, row, self.COLUMNS['photo'])
        if main_image:
            images.append({
                'data': main_image['data'],
                'filename': f"{table_id}_A{row}_{main_image['hash']}.png",
                'cell_position': f"A{row}",
                'is_main': True
            })
        
        # Дополнительные фото из колонки J (может быть несколько)
        extra_images = self._extract_all_images_from_cell(ws_images, row, self.COLUMNS['extra_photo'])
        for idx, img in enumerate(extra_images, 1):
            images.append({
                'data': img['data'],
                'filename': f"{table_id}_J{row}_{idx}_{img['hash']}.png",
                'cell_position': f"J{row}",
                'is_main': False
            })
        
        return images
    
    def _extract_image_from_cell(self, ws, row, col):
        """Извлекает ОДНО изображение из ячейки"""
        try:
            if not hasattr(ws, '_images'):
                return None
            
            for image in ws._images:
                if image.anchor and hasattr(image.anchor, 'from'):
                    if image.anchor._from.row == row - 1 and image.anchor._from.col == col - 1:
                        img_data = image._data()
                        if img_data:
                            img_hash = hashlib.md5(img_data).hexdigest()[:16]
                            return {'data': img_data, 'hash': img_hash}
        except:
            pass
        
        return None
    
    def _extract_all_images_from_cell(self, ws, row, col):
        """Извлекает ВСЕ изображения из ячейки"""
        images = []
        
        try:
            if not hasattr(ws, '_images'):
                return images
            
            for image in ws._images:
                if image.anchor and hasattr(image.anchor, '_from'):
                    if image.anchor._from.row == row - 1 and image.anchor._from.col == col - 1:
                        img_data = image._data()
                        if img_data:
                            img_hash = hashlib.md5(img_data).hexdigest()[:16]
                            images.append({'data': img_data, 'hash': img_hash})
        except:
            pass
        
        return images
    
    def _save_product(self, product, project_id, table_id):
        """Сохраняет товар в БД"""
        with self.db.get_session() as session:
            # Вставляем товар
            result = session.execute(text("""
                INSERT INTO products (
                    project_id, table_id, name, custom_field,
                    sample_price, sample_delivery_time, row_number, created_at, updated_at
                )
                VALUES (:project_id, :table_id, :name, :custom_field,
                        :sample_price, :sample_delivery_time, :row_number, NOW(), NOW())
                RETURNING id
            """), {
                'project_id': project_id,
                'table_id': table_id,
                'name': product['name'],
                'custom_field': product.get('custom_field'),
                'sample_price': product.get('sample_price'),
                'sample_delivery_time': product.get('sample_delivery_time'),
                'row_number': product['row_number']
            })
            
            product_id = result.fetchone()[0]
            
            # Сохраняем изображения
            for img in product['images']:
                # Сохраняем файл
                img_path = self.storage_dir / img['filename']
                with open(img_path, 'wb') as f:
                    f.write(img['data'])
                
                # Записываем в БД
                session.execute(text("""
                    INSERT INTO product_images (
                        product_id, table_id, image_filename, local_path,
                        cell_position, is_main_image, row_number, created_at, updated_at
                    )
                    VALUES (:product_id, :table_id, :filename, :local_path,
                            :cell_position, :is_main, :row_number, NOW(), NOW())
                """), {
                    'product_id': product_id,
                    'table_id': table_id,
                    'filename': img['filename'],
                    'local_path': str(img_path),
                    'cell_position': img['cell_position'],
                    'is_main': img['is_main'],
                    'row_number': product['row_number']
                })
            
            # Сохраняем офферы
            for offer in product['offers']:
                session.execute(text("""
                    INSERT INTO price_offers (
                        product_id, quantity, price_usd, price_rub,
                        route, delivery_time_days, created_at, updated_at
                    )
                    VALUES (:product_id, :quantity, :price_usd, :price_rub,
                            :route, :delivery_days, NOW(), NOW())
                """), {
                    'product_id': product_id,
                    'quantity': offer['quantity'],
                    'price_usd': offer['price_usd'],
                    'price_rub': offer['price_rub'],
                    'route': offer['route'],
                    'delivery_days': offer['delivery_time_days']
                })
            
            session.commit()
    
    def _get_cell_value(self, ws, row, col):
        """Получает значение ячейки"""
        try:
            cell = ws.cell(row, col)
            return cell.value if cell.value is not None else None
        except:
            return None
    
    def _parse_quantity(self, value):
        """Парсит количество"""
        if not value:
            return None
        
        val_str = str(value).strip().replace(' ', '').replace(',', '')
        
        # Убираем все кроме цифр и точки
        val_str = re.sub(r'[^\d.]', '', val_str)
        
        if not val_str:
            return None
        
        try:
            # Преобразуем через float чтобы обработать "10000.0"
            return int(float(val_str))
        except:
            return None
    
    def _parse_price(self, value):
        """Парсит цену"""
        if not value:
            return None
        
        val_str = str(value).strip()
        
        # Убираем символы валют и пробелы
        val_str = val_str.replace('$', '').replace('₽', '').replace(' ', '').replace(',', '.')
        
        try:
            return float(val_str)
        except:
            return None
    
    def _parse_delivery_time(self, value):
        """Парсит срок доставки, возвращает максимальное значение"""
        if not value:
            return None
        
        val_str = str(value).lower().strip()
        
        # Ищем все числа
        numbers = re.findall(r'\d+', val_str)
        
        if not numbers:
            return None
        
        # Возвращаем максимальное (например, "20-22 дня" → 22)
        return max(int(n) for n in numbers)


def main():
    """Парсит все проекты Шаблона 5"""
    print("=" * 100)
    print("🚀 ПАРСИНГ ШАБЛОНА 5")
    print("=" * 100)
    
    # Загружаем ID
    ids_file = Path('template_5_candidate_ids.txt')
    with open(ids_file, 'r') as f:
        project_ids = [int(line.strip()) for line in f if line.strip()]
    
    print(f"\nВсего проектов: {len(project_ids)}")
    
    parser = Template5Parser()
    
    success_count = 0
    error_count = 0
    total_products = 0
    total_offers = 0
    total_images = 0
    
    for i, project_id in enumerate(project_ids, 1):
        print(f"\n[{i}/{len(project_ids)}] Парсинг проекта {project_id}...", end=' ')
        
        result = parser.parse_project(project_id)
        
        if result['success']:
            print(f"✅ {result['products']} товаров, {result['total_offers']} офферов, {result['total_images']} изображений")
            success_count += 1
            total_products += result['products']
            total_offers += result['total_offers']
            total_images += result['total_images']
        else:
            print(f"❌ {result['error']}")
            error_count += 1
        
        # Прогресс каждые 10 проектов
        if i % 10 == 0:
            print(f"\n📊 Прогресс: {success_count} успешно, {error_count} ошибок")
    
    print("\n" + "=" * 100)
    print("📊 ИТОГИ:")
    print("=" * 100)
    print(f"✅ Успешно:      {success_count}")
    print(f"❌ Ошибок:       {error_count}")
    print(f"📦 Товаров:      {total_products}")
    print(f"💰 Офферов:      {total_offers}")
    print(f"🖼️  Изображений:  {total_images}")
    print("=" * 100)


if __name__ == '__main__':
    main()

