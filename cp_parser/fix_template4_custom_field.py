#!/usr/bin/env python3
"""
Исправление custom_field в Шаблоне 4:
1. Убрать "Образец" из custom_field
2. Парсить цену и срок образца из Excel
3. Записать в sample_price и sample_delivery_time
"""

import sys
from pathlib import Path
from sqlalchemy import text
import re
from openpyxl import load_workbook

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager

db = PostgreSQLManager()

print("=" * 100)
print("🔧 ИСПРАВЛЕНИЕ CUSTOM_FIELD И ОБРАЗЦОВ В ШАБЛОНЕ 4")
print("=" * 100)

with open('template_4_perfect_ids.txt', 'r') as f:
    project_ids = [int(line.strip()) for line in f if line.strip()]

with db.get_session() as session:
    # 1. Убираем "Образец" из custom_field
    print("\n📝 ШАГ 1: Убираем 'Образец' из custom_field...")
    print("-" * 100)
    
    products_with_obrazec = session.execute(text("""
        SELECT id, custom_field
        FROM products
        WHERE project_id = ANY(:ids)
        AND custom_field LIKE '%Образец%'
    """), {'ids': project_ids}).fetchall()
    
    print(f"Найдено товаров с 'Образец' в custom_field: {len(products_with_obrazec)}")
    
    updated = 0
    for prod_id, custom in products_with_obrazec:
        # Убираем ", Образец" или ",Образец" из конца строки
        custom_clean = re.sub(r',\s*Образец$', '', custom, flags=re.IGNORECASE)
        custom_clean = re.sub(r'\s*Образец$', '', custom_clean, flags=re.IGNORECASE)
        custom_clean = custom_clean.strip()
        
        if custom_clean != custom:
            session.execute(text("""
                UPDATE products
                SET custom_field = :custom_clean
                WHERE id = :prod_id
            """), {'custom_clean': custom_clean, 'prod_id': prod_id})
            updated += 1
    
    session.commit()
    print(f"✅ Обновлено custom_field: {updated}")
    
    # 2. Парсим образцы из Excel и записываем в sample_price/sample_delivery_time
    print("\n📝 ШАГ 2: Парсим образцы из Excel...")
    print("-" * 100)
    
    excel_dir = Path('storage/excel_files')
    
    # Получаем проекты
    projects = session.execute(text("""
        SELECT id, table_id
        FROM projects
        WHERE id = ANY(:ids)
        ORDER BY id
    """), {'ids': project_ids}).fetchall()
    
    total_files = len(projects)
    processed = 0
    sample_updates = 0
    
    for proj_id, table_id in projects:
        # Ищем Excel файл
        excel_files = list(excel_dir.glob(f'project_{proj_id}_*.xlsx'))
        if not excel_files:
            continue
        
        excel_file = excel_files[0]
        
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Получаем товары этого проекта
            products = session.execute(text("""
                SELECT id, row_number
                FROM products
                WHERE project_id = :proj_id
            """), {'proj_id': proj_id}).fetchall()
            
            for prod_id, row_num in products:
                # Колонка N - цена образца (руб)
                sample_price_cell = ws.cell(row_num, 14)  # N = 14
                # Колонка P - срок образца (к.д.)
                sample_delivery_cell = ws.cell(row_num, 16)  # P = 16
                
                sample_price = None
                sample_delivery = None
                
                if sample_price_cell.value:
                    try:
                        # Парсим цену
                        val = str(sample_price_cell.value).replace(' ', '').replace(',', '.')
                        sample_price = float(val)
                    except:
                        pass
                
                if sample_delivery_cell.value:
                    try:
                        # Парсим срок
                        val = str(sample_delivery_cell.value).replace(' ', '')
                        sample_delivery = int(float(val))
                    except:
                        pass
                
                if sample_price or sample_delivery:
                    session.execute(text("""
                        UPDATE products
                        SET sample_price = :price,
                            sample_delivery_time = :delivery
                        WHERE id = :prod_id
                    """), {
                        'price': str(sample_price) if sample_price else None,
                        'delivery': sample_delivery,
                        'prod_id': prod_id
                    })
                    sample_updates += 1
            
            wb.close()
            processed += 1
            
            if processed % 50 == 0:
                print(f"   ⏳ Обработано файлов: {processed}/{total_files}")
        
        except Exception as e:
            print(f"   ⚠️  Ошибка в проекте {proj_id}: {str(e)[:50]}")
            continue
    
    session.commit()
    
    print(f"\n✅ ИТОГО:")
    print(f"   Обработано Excel файлов: {processed}/{total_files}")
    print(f"   Обновлено образцов: {sample_updates}")
    
    # 3. Проверяем результат
    print("\n📊 ШАГ 3: Проверка результата...")
    print("-" * 100)
    
    stats = session.execute(text("""
        SELECT 
            COUNT(*) as total,
            COUNT(CASE WHEN custom_field LIKE '%Образец%' THEN 1 END) as with_obrazec_in_custom,
            COUNT(CASE WHEN sample_price IS NOT NULL THEN 1 END) as with_sample_price,
            COUNT(CASE WHEN sample_delivery_time IS NOT NULL THEN 1 END) as with_sample_delivery
        FROM products
        WHERE project_id = ANY(:ids)
    """), {'ids': project_ids}).fetchone()
    
    total, with_obrazec, with_price, with_delivery = stats
    
    print(f"Всего товаров Шаблона 4:     {total:,}")
    print(f"С 'Образец' в custom_field:  {with_obrazec:,} (должно быть 0!)")
    print(f"С sample_price:              {with_price:,}")
    print(f"С sample_delivery_time:      {with_delivery:,}")
    
    print("\n" + "=" * 100)
    
    if with_obrazec == 0:
        print("✅ УСПЕШНО! Все 'Образцы' убраны из custom_field")
    else:
        print(f"⚠️  Осталось {with_obrazec} товаров с 'Образец' в custom_field")
    
    print("=" * 100)


