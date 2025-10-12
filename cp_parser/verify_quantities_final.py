#!/usr/bin/env python3
"""
Финальная версия автоматической проверки тиражей
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path.cwd()))

import json
import gspread
from google.oauth2.service_account import Credentials
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text
import os
from openpyxl import load_workbook
import csv
from datetime import datetime
import time
import io

# Railway DB
os.environ['USE_RAILWAY_DB'] = 'true'

def download_excel_gspread(sheet_id, output_path):
    """Скачивает через gspread API"""
    try:
        scope = [
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.readonly'
        ]
        creds = Credentials.from_service_account_file('service_account.json', scopes=scope)
        client = gspread.authorize(creds)
        
        # Открываем таблицу
        spreadsheet = client.open_by_key(sheet_id)
        
        # Экспортируем как Excel
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        
        import requests
        response = requests.get(export_url, headers={'Authorization': f'Bearer {creds.token}'})
        
        if response.status_code == 200:
            with open(output_path, 'wb') as f:
                f.write(response.content)
            return True
        else:
            return False
            
    except Exception as e:
        print(f"  ❌ {str(e)[:50]}")
        return False

def extract_quantity_from_excel(excel_path, row_number):
    """Извлекает тираж из колонки E"""
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        
        cell_value = ws.cell(row=row_number, column=5).value
        
        if cell_value is None:
            return None
            
        try:
            return int(float(str(cell_value).replace(',', '').replace(' ', '')))
        except:
            return str(cell_value)
            
    except Exception as e:
        return None

def verify_project(project_data, db, temp_dir):
    """Проверяет проект"""
    project_id = project_data['id']
    sheet_id = project_data['sheet_id']
    
    print(f"\n{'='*70}")
    print(f"🔍 ID {project_id}: {project_data['name'][:55]}")
    
    excel_path = temp_dir / f"proj_{project_id}.xlsx"
    print(f"📥 Скачиваю...", end=' ', flush=True)
    
    if not download_excel_gspread(sheet_id, excel_path):
        return None
    
    print(f"✅ {excel_path.stat().st_size/1024:.0f} KB")
    
    with db.get_session() as session:
        products = session.execute(text("""
            SELECT 
                p.id, p.name, p.row_number,
                po.quantity, po.route
            FROM products p
            INNER JOIN price_offers po ON p.id = po.product_id
            WHERE p.project_id = :pid
            AND po.quantity >= 300
            AND po.quantity % 10 = 0
            ORDER BY p.row_number
        """), {'pid': project_id}).fetchall()
        
        if not products:
            print(f"⚠️  Нет подозрительных товаров")
            return None
        
        print(f"📊 Проверяю {len(products)} записей...")
        
        results = []
        checked_rows = {}
        
        for prod in products:
            row_num = prod.row_number
            
            if row_num not in checked_rows:
                checked_rows[row_num] = extract_quantity_from_excel(excel_path, row_num)
            
            excel_qty = checked_rows[row_num]
            db_qty = int(prod.quantity)
            
            is_x10 = False
            if excel_qty and isinstance(excel_qty, int) and excel_qty * 10 == db_qty:
                is_x10 = True
            
            results.append({
                'product_id': prod.id,
                'product_name': prod.name,
                'row': row_num,
                'excel': excel_qty,
                'db': db_qty,
                'route': prod.route,
                'is_x10': is_x10
            })
        
        x10_count = sum(1 for r in results if r['is_x10'])
        ok_count = sum(1 for r in results if r['excel'] == r['db'])
        
        print(f"  ❌ x10: {x10_count}/{len(results)} ({x10_count/len(results)*100:.0f}%)")
        print(f"  ✅ OK: {ok_count}/{len(results)} ({ok_count/len(results)*100:.0f}%)")
        
        if x10_count > 0:
            print(f"  Примеры:")
            shown = 0
            for r in results:
                if r['is_x10'] and shown < 2:
                    print(f"    Р{r['row']}: {r['excel']} → {r['db']}")
                    shown += 1
        
        return {
            'id': project_id,
            'name': project_data['name'],
            'url': project_data['url'],
            'total': len(results),
            'x10': x10_count,
            'ok': ok_count,
            'results': results
        }

def main():
    print("="*70)
    print("🔍 АВТОПРОВЕРКА ТИРАЖЕЙ Template 4")
    print("="*70)
    
    with open('projects_to_verify.json', 'r', encoding='utf-8') as f:
        projects = json.load(f)
    
    print(f"\n📋 Проверяю все {len(projects)} проектов\n")
    
    temp_dir = Path('temp_verification')
    temp_dir.mkdir(exist_ok=True)
    
    db = PostgreSQLManager()
    all_results = []
    
    for i, proj in enumerate(projects, 1):
        print(f"\n{'#'*70}")
        print(f"# {i}/{len(projects)}")
        
        result = verify_project(proj, db, temp_dir)
        if result:
            all_results.append(result)
        
        time.sleep(0.5)
    
    # Итог
    print("\n\n" + "="*70)
    print("📊 ИТОГ")
    print("="*70)
    
    if not all_results:
        print("\n⚠️  Не удалось проверить")
        return
    
    total_x10 = sum(r['x10'] for r in all_results)
    total_rec = sum(r['total'] for r in all_results)
    
    print(f"\n✅ Проверено: {len(all_results)} проектов")
    print(f"📊 Записей: {total_rec:,}")
    print(f"❌ Ошибок x10: {total_x10:,} ({total_x10/total_rec*100:.1f}%)")
    
    print(f"\n📋 ПО ПРОЕКТАМ:")
    for r in all_results:
        pct = r['x10']/r['total']*100 if r['total'] > 0 else 0
        ico = "❌" if r['x10'] > 0 else "✅"
        print(f"{ico} ID {r['id']}: {r['x10']}/{r['total']} ({pct:.0f}%) - {r['name'][:45]}")
    
    # CSV
    csv_path = Path(f"АВТОПРОВЕРКА_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")
    
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'ID_Проекта', 'Название', 'ID_Товара', 'Товар',
            'Строка', 'Тираж_Excel', 'Тираж_БД', 'Маршрут', 'Ошибка_x10'
        ])
        
        for proj in all_results:
            for r in proj['results']:
                writer.writerow([
                    proj['id'], proj['name'],
                    r['product_id'], r['product_name'],
                    r['row'], r['excel'], r['db'],
                    r['route'], 'ДА' if r['is_x10'] else 'НЕТ'
                ])
    
    print(f"\n💾 Отчет: {csv_path}")
    
    # Вывод: нужно ли исправлять
    if total_x10 > total_rec * 0.5:  # Если >50% ошибок
        print(f"\n🚨 КРИТИЧНО: {total_x10/total_rec*100:.0f}% ошибок x10!")
        print(f"   Рекомендуется массовое исправление (÷10)")
    
    print("="*70)

if __name__ == '__main__':
    main()

