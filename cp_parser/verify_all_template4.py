#!/usr/bin/env python3
"""
Проверка тиражей для ВСЕХ проектов Template 4 (943 проекта)
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path.cwd()))

import json
import gspread
from google.oauth2.service_account import Credentials
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text
import os
from openpyxl import load_workbook
import csv
from datetime import datetime
import time

# Railway DB
os.environ['USE_RAILWAY_DB'] = 'true'

def download_excel_gspread(sheet_id, output_path):
    """Скачивает через gspread API"""
    try:
        scope = [
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.readonly'
        ]
        creds = Credentials.from_service_account_file('service_account.json', scopes=scope)
        
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        
        import requests
        response = requests.get(export_url, headers={'Authorization': f'Bearer {creds.token}'})
        
        if response.status_code == 200:
            with open(output_path, 'wb') as f:
                f.write(response.content)
            return True
        else:
            return False
            
    except Exception as e:
        return False

def extract_quantity_from_excel(excel_path, row_number):
    """Извлекает тираж из колонки E"""
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        
        cell_value = ws.cell(row=row_number, column=5).value
        
        if cell_value is None:
            return None
            
        try:
            return int(float(str(cell_value).replace(',', '').replace(' ', '')))
        except:
            return str(cell_value)
            
    except Exception as e:
        return None

def verify_project(project_data, db, temp_dir):
    """Проверяет проект"""
    project_id = project_data['id']
    sheet_id = project_data['sheet_id']
    
    excel_path = temp_dir / f"proj_{project_id}.xlsx"
    
    # Скачиваем только если еще не скачан
    if not excel_path.exists():
        if not download_excel_gspread(sheet_id, excel_path):
            return None
    
    with db.get_session() as session:
        products = session.execute(text("""
            SELECT 
                p.id, p.name, p.row_number,
                po.quantity, po.route
            FROM products p
            INNER JOIN price_offers po ON p.id = po.product_id
            WHERE p.project_id = :pid
            AND po.quantity >= 300
            AND po.quantity % 10 = 0
            ORDER BY p.row_number
        """), {'pid': project_id}).fetchall()
        
        if not products:
            return None
        
        results = []
        checked_rows = {}
        
        for prod in products:
            row_num = prod.row_number
            
            if row_num not in checked_rows:
                checked_rows[row_num] = extract_quantity_from_excel(excel_path, row_num)
            
            excel_qty = checked_rows[row_num]
            db_qty = int(prod.quantity)
            
            is_x10 = False
            if excel_qty and isinstance(excel_qty, int) and excel_qty * 10 == db_qty:
                is_x10 = True
            
            results.append({
                'product_id': prod.id,
                'product_name': prod.name,
                'row': row_num,
                'excel': excel_qty,
                'db': db_qty,
                'route': prod.route,
                'is_x10': is_x10
            })
        
        x10_count = sum(1 for r in results if r['is_x10'])
        
        return {
            'id': project_id,
            'name': project_data['name'],
            'url': project_data['url'],
            'total': len(results),
            'x10': x10_count,
            'results': results
        }

def main():
    print("="*70)
    print("🔍 ПОЛНАЯ ПРОВЕРКА ВСЕХ ПРОЕКТОВ TEMPLATE 4")
    print("="*70)
    
    db = PostgreSQLManager()
    
    # Получаем ВСЕ проекты Template 4 из БД
    print("\n📊 Получаю список проектов Template 4...")
    
    with db.get_session() as session:
        projects_data = session.execute(text("""
            SELECT DISTINCT
                pr.id,
                pr.project_name,
                pr.google_sheets_url
            FROM projects pr
            INNER JOIN products p ON pr.id = p.project_id
            WHERE p.sample_price IS NOT NULL
            AND pr.google_sheets_url IS NOT NULL
            ORDER BY pr.id
        """)).fetchall()
    
    print(f"✅ Найдено проектов Template 4: {len(projects_data)}")
    
    # Преобразуем в список
    projects = []
    for proj in projects_data:
        if '/d/' in proj.google_sheets_url:
            sheet_id = proj.google_sheets_url.split('/d/')[1].split('/')[0]
            projects.append({
                'id': proj.id,
                'name': proj.project_name,
                'sheet_id': sheet_id,
                'url': proj.google_sheets_url
            })
    
    print(f"📋 Проектов с валидными URL: {len(projects)}")
    print(f"📁 Excel в: temp_verification/\n")
    
    temp_dir = Path('temp_verification')
    temp_dir.mkdir(exist_ok=True)
    
    all_results = []
    failed_count = 0
    
    start_time = time.time()
    
    for i, proj in enumerate(projects, 1):
        # Краткий вывод
        if i % 10 == 1 or i % 10 == 0 or i == len(projects):
            print(f"\r🔍 {i}/{len(projects)} | ID {proj['id']}: {proj['name'][:40]:<40}", end='', flush=True)
        
        result = verify_project(proj, db, temp_dir)
        if result:
            all_results.append(result)
        else:
            failed_count += 1
        
        # Прогресс каждые 100 проектов
        if i % 100 == 0:
            elapsed = time.time() - start_time
            avg_time = elapsed / i
            remaining = (len(projects) - i) * avg_time
            projects_with_errors = len([r for r in all_results if r['x10'] > 0])
            print(f"\n⏱️  Прогресс: {i}/{len(projects)} ({i/len(projects)*100:.1f}%) | "
                  f"С ошибками: {projects_with_errors} | "
                  f"Осталось: ~{remaining/60:.0f} мин")
        
        time.sleep(0.2)  # Небольшая пауза
    
    # Итоговый отчет
    print("\n\n" + "="*70)
    print("📊 ИТОГОВЫЙ ОТЧЕТ")
    print("="*70)
    
    if not all_results:
        print("\n⚠️  Не удалось проверить проекты")
        return
    
    total_x10 = sum(r['x10'] for r in all_results)
    total_rec = sum(r['total'] for r in all_results)
    projects_with_errors = len([r for r in all_results if r['x10'] > 0])
    
    print(f"\n✅ Проверено: {len(all_results)} проектов")
    print(f"❌ Не удалось: {failed_count} проектов")
    print(f"📊 Всего записей: {total_rec:,}")
    print(f"❌ Ошибок x10: {total_x10:,} ({total_x10/total_rec*100:.1f}%)")
    print(f"⚠️  Проектов с ошибками: {projects_with_errors} ({projects_with_errors/len(all_results)*100:.1f}%)")
    
    # Топ проектов с ошибками
    error_projects = sorted([r for r in all_results if r['x10'] > 0], 
                          key=lambda x: x['x10'], reverse=True)
    
    if error_projects:
        print(f"\n📋 ТОП-30 ПРОЕКТОВ С НАИБОЛЬШИМ КОЛ-ВОМ ОШИБОК x10:")
        for r in error_projects[:30]:
            pct = r['x10']/r['total']*100 if r['total'] > 0 else 0
            print(f"  ❌ ID {r['id']}: {r['x10']}/{r['total']} ({pct:.0f}%) - {r['name'][:45]}")
    
    # Сохраняем CSV
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    csv_report = Path(f"ПОЛНАЯ_ПРОВЕРКА_TEMPLATE4_{timestamp}.csv")
    
    print(f"\n💾 Сохраняю детальный отчет...")
    
    with open(csv_report, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'ID_Проекта', 'Название', 'ID_Товара', 'Товар',
            'Строка', 'Тираж_Google', 'Тираж_БД', 'Маршрут', 'Ошибка_x10', 'URL'
        ])
        
        for proj in all_results:
            for r in proj['results']:
                writer.writerow([
                    proj['id'], proj['name'],
                    r['product_id'], r['product_name'],
                    r['row'], r['excel'], r['db'],
                    r['route'], 'ДА' if r['is_x10'] else 'НЕТ',
                    proj['url']
                ])
    
    print(f"✅ Отчет: {csv_report}")
    
    # Сохраняем список проектов с ошибками для исправления
    if error_projects:
        fix_list = Path(f"ПРОЕКТЫ_ДЛЯ_ИСПРАВЛЕНИЯ_{timestamp}.csv")
        
        with open(fix_list, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                'ID_Проекта', 'Название', 'Ошибок', 'Всего', 'Процент', 'URL'
            ])
            
            for r in error_projects:
                pct = r['x10']/r['total']*100 if r['total'] > 0 else 0
                writer.writerow([
                    r['id'], r['name'], r['x10'], r['total'], f"{pct:.1f}%", r['url']
                ])
        
        print(f"✅ Список для исправления: {fix_list}")
    
    # Рекомендации
    if total_x10 > 0:
        error_pct = total_x10 / total_rec * 100
        print(f"\n💡 РЕКОМЕНДАЦИЯ:")
        if error_pct > 10:
            print(f"   🚨 КРИТИЧНО: {error_pct:.1f}% ошибок!")
            print(f"   Массовое исправление для {projects_with_errors} проектов")
        elif error_pct > 5:
            print(f"   ⚠️  Средний уровень: {error_pct:.1f}%")
            print(f"   Исправить проекты с ошибками >10%")
        else:
            print(f"   ✅ Низкий уровень: {error_pct:.1f}%")
            print(f"   Выборочная проверка")
    
    elapsed = time.time() - start_time
    print(f"\n⏱️  Время: {elapsed/60:.1f} минут")
    print("="*70)

if __name__ == '__main__':
    main()


