#!/usr/bin/env python3
"""
Валидация проектов по Шаблону 4 (Нестандартная структура)
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict
import json

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template4Validator:
    """
    Валидатор для Шаблона 4
    
    Характеристики:
    - Строка 1: Информация о менеджере
    - Строка 2: Заголовки колонок (Фото, Наименование, Характеристики, Кастом, Тираж, Доставка ЖД, Доставка АВИА, Образец)
    - Строка 3: Подзаголовки (Цена за шт., $, Цена за шт., руб, Итого руб, Срок тиража)
    - Строка 4+: Данные товаров
    
    Структура колонок:
    A: Фото
    B: Наименование
    C: Характеристики
    D: Кастом
    E: Тираж, шт (ОДИН для всех маршрутов)
    F-I: ЖД (Цена $, Цена руб, Итого руб, Срок)
    J-M: АВИА (Цена $, Цена руб, Итого руб, Срок)
    N-P: Образец (Цена руб, Срок фото/видео, Срок с доставкой)
    Q: Доп. фото
    """
    
    def __init__(self):
        self.db = PostgreSQLManager()
    
    def validate_file(self, file_path):
        """Валидация одного файла по Шаблону 4"""
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            
            # Проверяем структуру заголовков
            validation_result = {
                'is_valid': False,
                'template': None,
                'reasons': [],
                'structure': {},
                'confidence': 0
            }
            
            # Читаем первые 3 строки
            row1 = self._read_row(ws, 1, 20)
            row2 = self._read_row(ws, 2, 20)
            row3 = self._read_row(ws, 3, 20)
            
            # Проверяем признаки Шаблона 4
            checks = {
                'has_photo': self._check_photo_column(row2),
                'has_name': self._check_name_column(row2),
                'has_quantity': self._check_quantity_column(row2),
                'has_jd_route': self._check_jd_route(row2),
                'has_avia_route': self._check_avia_route(row2),
                'has_sample': self._check_sample_column(row2),
                'has_price_structure': self._check_price_structure(row3),
                'quantity_in_E': self._check_quantity_in_column_E(ws, row2),
                'has_custom_column': self._check_custom_column(row2),
            }
            
            validation_result['structure'] = checks
            
            # Подсчет confidence
            confidence = sum(checks.values())
            validation_result['confidence'] = confidence
            
            # Минимальные требования для Шаблона 4
            required_checks = [
                'has_photo',
                'has_name',
                'has_quantity',
                'quantity_in_E',  # Ключевое отличие Шаблона 4
                'has_price_structure'
            ]
            
            # Должен быть хотя бы один маршрут (ЖД или АВИА)
            has_route = checks['has_jd_route'] or checks['has_avia_route']
            
            if all(checks.get(req, False) for req in required_checks) and has_route:
                validation_result['is_valid'] = True
                validation_result['template'] = 'template_4'
            else:
                # Определяем что не хватает
                missing = []
                if not checks['has_photo']:
                    missing.append('Колонка "Фото"')
                if not checks['has_name']:
                    missing.append('Колонка "Наименование"')
                if not checks['has_quantity']:
                    missing.append('Колонка "Тираж"')
                if not checks['quantity_in_E']:
                    missing.append('Тираж не в колонке E')
                if not has_route:
                    missing.append('Маршруты доставки')
                if not checks['has_price_structure']:
                    missing.append('Структура цен')
                
                validation_result['reasons'] = missing
            
            wb.close()
            return validation_result
            
        except Exception as e:
            return {
                'is_valid': False,
                'template': None,
                'reasons': [f'Ошибка чтения: {str(e)}'],
                'structure': {},
                'confidence': 0
            }
    
    def _read_row(self, ws, row_num, max_col):
        """Читает строку из листа"""
        row_data = {}
        for col in range(1, max_col + 1):
            val = ws.cell(row_num, col).value
            if val:
                col_letter = chr(64 + col)
                row_data[col_letter] = str(val).strip().lower()
        return row_data
    
    def _check_photo_column(self, row2):
        """Проверка колонки Фото (должна быть в A)"""
        return 'A' in row2 and 'фото' in row2['A']
    
    def _check_name_column(self, row2):
        """Проверка колонки Наименование (должна быть в B)"""
        return 'B' in row2 and ('наименование' in row2['B'] or 'название' in row2['B'])
    
    def _check_quantity_column(self, row2):
        """Проверка колонки Тираж"""
        for col, val in row2.items():
            if 'тираж' in val or 'количество' in val:
                return True
        return False
    
    def _check_quantity_in_column_E(self, ws, row2):
        """Проверка что тираж именно в колонке E (ключевой признак Шаблона 4)"""
        if 'E' in row2:
            val = row2['E']
            return 'тираж' in val or 'кол' in val
        return False
    
    def _check_jd_route(self, row2):
        """Проверка маршрута ЖД"""
        for col, val in row2.items():
            if 'жд' in val or 'доставка жд' in val:
                return True
        return False
    
    def _check_avia_route(self, row2):
        """Проверка маршрута АВИА"""
        for col, val in row2.items():
            if 'авиа' in val or 'доставка авиа' in val:
                return True
        return False
    
    def _check_sample_column(self, row2):
        """Проверка колонки Образец"""
        for col, val in row2.items():
            if 'образец' in val:
                return True
        return False
    
    def _check_price_structure(self, row3):
        """Проверка структуры цен в строке 3"""
        has_usd = False
        has_rub = False
        
        for col, val in row3.items():
            if '$' in val or 'usd' in val:
                has_usd = True
            if 'руб' in val or 'rub' in val:
                has_rub = True
        
        return has_usd and has_rub
    
    def _check_custom_column(self, row2):
        """Проверка колонки Кастом"""
        for col, val in row2.items():
            if 'кастом' in val or 'custom' in val:
                return True
        return False
    
    def validate_all_template_4(self):
        """Валидация всех проектов из template_4_ids.txt"""
        
        # Читаем список ID
        ids_file = Path('template_4_ids.txt')
        if not ids_file.exists():
            print("❌ Файл template_4_ids.txt не найден!")
            return
        
        with open(ids_file, 'r') as f:
            project_ids = [int(line.strip()) for line in f if line.strip()]
        
        print("=" * 80)
        print("🔍 ВАЛИДАЦИЯ ПРОЕКТОВ ПО ШАБЛОНУ 4")
        print("=" * 80)
        print(f"\nВсего проектов для проверки: {len(project_ids)}")
        print("")
        
        results = {
            'valid': [],
            'invalid': [],
            'not_found': [],
            'errors': []
        }
        
        confidence_distribution = defaultdict(int)
        
        for i, project_id in enumerate(project_ids):
            if i % 50 == 0:
                print(f"   Обработано: {i}/{len(project_ids)}...")
            
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx'))
            
            if not matching_files:
                results['not_found'].append(project_id)
                continue
            
            # Валидация
            file_path = matching_files[0]
            validation = self.validate_file(str(file_path))
            
            confidence_distribution[validation['confidence']] += 1
            
            if validation['is_valid']:
                results['valid'].append({
                    'project_id': project_id,
                    'file_name': file_path.name,
                    'confidence': validation['confidence'],
                    'structure': validation['structure']
                })
            else:
                results['invalid'].append({
                    'project_id': project_id,
                    'file_name': file_path.name,
                    'reasons': validation['reasons'],
                    'confidence': validation['confidence'],
                    'structure': validation['structure']
                })
        
        print(f"   Обработано: {len(project_ids)}/{len(project_ids)} ✅")
        print("")
        
        # Отчет
        print("=" * 80)
        print("📊 РЕЗУЛЬТАТЫ ВАЛИДАЦИИ")
        print("=" * 80)
        print("")
        print(f"✅ Валидные:       {len(results['valid']):4d} ({len(results['valid'])/len(project_ids)*100:5.1f}%)")
        print(f"❌ Невалидные:     {len(results['invalid']):4d} ({len(results['invalid'])/len(project_ids)*100:5.1f}%)")
        print(f"📁 Не найдены:     {len(results['not_found']):4d} ({len(results['not_found'])/len(project_ids)*100:5.1f}%)")
        print("")
        
        # Распределение по confidence
        print("📈 РАСПРЕДЕЛЕНИЕ ПО CONFIDENCE:")
        print("")
        for conf in sorted(confidence_distribution.keys(), reverse=True):
            count = confidence_distribution[conf]
            bar = "█" * int(count / 10)
            print(f"   {conf:2d}/9: {count:4d} проектов {bar}")
        print("")
        
        # Топ причины невалидности
        if results['invalid']:
            print("🔍 ТОП ПРИЧИНЫ НЕВАЛИДНОСТИ:")
            print("")
            reason_counts = defaultdict(int)
            for item in results['invalid']:
                for reason in item['reasons']:
                    reason_counts[reason] += 1
            
            for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1])[:5]:
                print(f"   {count:4d} проектов: {reason}")
            print("")
        
        # Примеры валидных
        if results['valid']:
            print("✅ ПРИМЕРЫ ВАЛИДНЫХ ПРОЕКТОВ (первые 10):")
            print("")
            for item in results['valid'][:10]:
                print(f"   Project {item['project_id']:5d} (confidence: {item['confidence']}/9)")
            print("")
        
        # Сохраняем результаты
        output_file = Path('template_4_validation_results.json')
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        
        print(f"💾 Результаты сохранены: {output_file}")
        print("")
        
        # Создаем список валидных ID
        if results['valid']:
            valid_ids_file = Path('template_4_valid_ids.txt')
            with open(valid_ids_file, 'w') as f:
                for item in results['valid']:
                    f.write(f"{item['project_id']}\n")
            
            print(f"📋 Список валидных ID: {valid_ids_file}")
            print("")
        
        print("=" * 80)
        print("✅ ВАЛИДАЦИЯ ЗАВЕРШЕНА")
        print("=" * 80)
        
        return results


if __name__ == '__main__':
    validator = Template4Validator()
    validator.validate_all_template_4()



