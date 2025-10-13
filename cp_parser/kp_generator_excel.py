"""
Генератор Excel коммерческого предложения
Группирует товары и показывает все варианты маршрутов для каждого товара
"""

import os
from pathlib import Path
import sys
from datetime import datetime
from collections import defaultdict

# Добавляем путь к модулям проекта
sys.path.append(str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import requests
from io import BytesIO
from PIL import Image as PILImage

from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class KPExcelGenerator:
    """Генератор Excel файлов для коммерческого предложения"""
    
    def __init__(self):
        self.db_manager = PostgreSQLManager()
        
        # Стили для Excel
        self.title_font = Font(name='Arial', size=16, bold=True)
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.product_font = Font(name='Arial', size=12, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.product_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def get_kp_items(self, session_id):
        """Получает товары из КП и группирует по product_id"""
        
        db_session = self.db_manager.get_session_direct()
        
        try:
            result = db_session.execute(text("""
                SELECT 
                    ki.id as kp_item_id,
                    p.id as product_id,
                    p.name as product_name,
                    p.description,
                    p.sample_price,
                    p.sample_delivery_time,
                    po.id as price_offer_id,
                    po.quantity,
                    po.route,
                    po.price_usd,
                    po.price_rub,
                    po.delivery_time_days
                FROM kp_items ki
                JOIN products p ON p.id = ki.product_id
                JOIN price_offers po ON po.id = ki.price_offer_id
                WHERE ki.session_id = :session_id
                ORDER BY p.name, po.quantity
            """), {'session_id': session_id})
            
            # Группируем по product_id
            products_grouped = defaultdict(lambda: {
                'info': None,
                'offers': [],
                'images': []
            })
            
            for row in result:
                product_id = row[1]
                
                # Информация о товаре (заполняем один раз)
                if products_grouped[product_id]['info'] is None:
                    products_grouped[product_id]['info'] = {
                        'name': row[2],
                        'description': row[3],
                        'sample_price': float(row[4]) if row[4] else None,
                        'sample_delivery_time': row[5]
                    }
                
                # Добавляем ценовое предложение
                products_grouped[product_id]['offers'].append({
                    'quantity': row[7],
                    'route': row[8],
                    'price_usd': float(row[9]) if row[9] else None,
                    'price_rub': float(row[10]) if row[10] else None,
                    'delivery_days': row[11]
                })
            
            # Получаем все изображения для каждого товара
            for product_id in products_grouped.keys():
                img_result = db_session.execute(text("""
                    SELECT image_url, image_filename
                    FROM product_images
                    WHERE product_id = :product_id
                    AND (image_url IS NOT NULL OR image_filename IS NOT NULL)
                    ORDER BY CASE WHEN column_number = 1 THEN 0 ELSE 1 END, id
                    LIMIT 5
                """), {'product_id': product_id})
                
                for img_row in img_result:
                    image_url = img_row[0]
                    if not image_url and img_row[1]:
                        image_url = f"https://s3.ru1.storage.beget.cloud/73d16f7545b3-promogoods/images/{img_row[1]}"
                    
                    if image_url:
                        products_grouped[product_id]['images'].append(image_url)
            
            return products_grouped
        finally:
            db_session.close()
    
    def download_image(self, url, max_size=(150, 150)):
        """Скачивает и подготавливает изображение для Excel"""
        
        if not url:
            return None
        
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                # Открываем изображение
                img = PILImage.open(BytesIO(response.content))
                
                # Конвертируем в RGB если нужно
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = PILImage.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Resize с сохранением пропорций
                img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
                
                # Сохраняем в BytesIO
                img_byte_arr = BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                
                return img_byte_arr
            else:
                print(f"⚠️  Не удалось загрузить изображение: {url} (статус {response.status_code})")
                return None
        except Exception as e:
            print(f"⚠️  Ошибка загрузки изображения {url}: {e}")
            return None
    
    def generate(self, session_id, output_path=None):
        """Генерирует Excel файл коммерческого предложения"""
        
        print("📊 [KP EXCEL] Начинаю генерацию Excel...")
        
        # Получаем товары
        products = self.get_kp_items(session_id)
        
        if not products:
            raise ValueError("КП пусто - нет товаров для генерации")
        
        print(f"   Найдено товаров: {len(products)}")
        print(f"   Общее количество предложений: {sum(len(p['offers']) for p in products.values())}")
        
        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Коммерческое предложение"
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 20  # Основное изображение
        ws.column_dimensions['B'].width = 35  # Товар/Описание
        ws.column_dimensions['C'].width = 12  # Тираж
        ws.column_dimensions['D'].width = 10  # USD
        ws.column_dimensions['E'].width = 12  # RUB
        ws.column_dimensions['F'].width = 15  # Доставка
        ws.column_dimensions['G'].width = 10  # Срок
        # Дополнительные изображения
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        
        current_row = 1
        
        # Заголовок документа
        ws.merge_cells(f'A{current_row}:K{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = 'КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ'
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment
        current_row += 1
        
        # Дата
        ws.merge_cells(f'A{current_row}:K{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f'от {datetime.now().strftime("%d.%m.%Y")}'
        date_cell.font = Font(name='Arial', size=10, italic=True)
        date_cell.alignment = self.center_alignment
        current_row += 2
        
        # Генерируем товары
        for product_id, product_data in products.items():
            product_info = product_data['info']
            offers = product_data['offers']
            images = product_data['images']
            
            print(f"   Обрабатываю: {product_info['name']} ({len(offers)} вариантов, {len(images)} изображений)")
            
            start_row = current_row
            
            # Название товара (объединяем ячейки B-G для текста)
            ws.merge_cells(f'B{current_row}:G{current_row}')
            name_cell = ws[f'B{current_row}']
            name_cell.value = product_info['name']
            name_cell.font = self.product_font
            name_cell.alignment = self.left_alignment
            name_cell.fill = self.product_fill
            
            # Применяем fill ко всем ячейкам в строке (включая колонки для доп. изображений)
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{col}{current_row}'].fill = self.product_fill
                ws[f'{col}{current_row}'].border = self.thin_border
            
            current_row += 1
            
            # Описание (если есть)
            if product_info['description']:
                ws.merge_cells(f'B{current_row}:G{current_row}')
                desc_cell = ws[f'B{current_row}']
                desc_cell.value = product_info['description'][:200]  # Ограничиваем длину
                desc_cell.font = Font(name='Arial', size=9, italic=True)
                desc_cell.alignment = self.left_alignment
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    ws[f'{col}{current_row}'].border = self.thin_border
                
                current_row += 1
            
            # Информация об образце (если есть)
            if product_info['sample_price'] or product_info['sample_delivery_time']:
                ws.merge_cells(f'B{current_row}:G{current_row}')
                sample_cell = ws[f'B{current_row}']
                
                sample_parts = []
                if product_info['sample_price']:
                    sample_parts.append(f"Образец: ${product_info['sample_price']:.2f}")
                if product_info['sample_delivery_time']:
                    sample_parts.append(f"Срок: {product_info['sample_delivery_time']} дн.")
                
                sample_cell.value = ' | '.join(sample_parts)
                sample_cell.font = Font(name='Arial', size=9, bold=True, color='0066CC')
                sample_cell.alignment = self.left_alignment
                sample_cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    ws[f'{col}{current_row}'].border = self.thin_border
                    if col != 'B':
                        ws[f'{col}{current_row}'].fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                
                current_row += 1
            
            # Заголовки таблицы с ценами
            headers = ['', '', 'Тираж', 'USD', 'RUB', 'Доставка', 'Срок']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Заголовок для дополнительных изображений
            if len(images) > 1:
                ws.merge_cells(f'H{current_row}:K{current_row}')
                img_header_cell = ws[f'H{current_row}']
                img_header_cell.value = 'Доп. фото'
                img_header_cell.font = self.header_font
                img_header_cell.fill = self.header_fill
                img_header_cell.alignment = self.center_alignment
                for col in ['H', 'I', 'J', 'K']:
                    ws[f'{col}{current_row}'].border = self.thin_border
            
            current_row += 1
            offers_start_row = current_row
            
            # Ценовые предложения
            for offer in offers:
                ws.cell(row=current_row, column=3, value=f"{offer['quantity']:,.0f} шт").alignment = self.center_alignment
                ws.cell(row=current_row, column=4, value=f"${offer['price_usd']:.2f}" if offer['price_usd'] else '-').alignment = self.center_alignment
                ws.cell(row=current_row, column=5, value=f"₽{offer['price_rub']:.2f}" if offer['price_rub'] else '-').alignment = self.center_alignment
                ws.cell(row=current_row, column=6, value=offer['route'] or '-').alignment = self.center_alignment
                ws.cell(row=current_row, column=7, value=f"{offer['delivery_days']} дн." if offer['delivery_days'] else '-').alignment = self.center_alignment
                
                # Границы для всех колонок (включая доп. изображения)
                for col in range(1, 12):  # A-K (1-11)
                    ws.cell(row=current_row, column=col).border = self.thin_border
                    ws.cell(row=current_row, column=col).font = self.normal_font
                
                current_row += 1
            
            # Вставляем изображения
            # Основное изображение (колонка A)
            if images and len(images) > 0:
                img_data = self.download_image(images[0])
                if img_data:
                    try:
                        img = XLImage(img_data)
                        img.width = 130
                        img.height = 130
                        img.anchor = f'A{start_row}'
                        ws.add_image(img)
                    except Exception as e:
                        print(f"   ⚠️  Не удалось вставить основное изображение: {e}")
            
            # Дополнительные изображения (колонки H, I, J, K)
            additional_images = images[1:5] if len(images) > 1 else []  # Берем до 4 доп. изображений
            image_columns = ['H', 'I', 'J', 'K']
            
            for idx, img_url in enumerate(additional_images):
                if idx >= len(image_columns):
                    break
                
                img_data = self.download_image(img_url, max_size=(100, 100))
                if img_data:
                    try:
                        img = XLImage(img_data)
                        img.width = 90
                        img.height = 90
                        img.anchor = f'{image_columns[idx]}{start_row}'
                        ws.add_image(img)
                        print(f"      Доп. изображение {idx+1} добавлено в колонку {image_columns[idx]}")
                    except Exception as e:
                        print(f"   ⚠️  Не удалось вставить доп. изображение {idx+1}: {e}")
            
            # Устанавливаем высоту строк для изображений
            total_rows = current_row - start_row
            if total_rows > 0:
                row_height = 140 / total_rows  # Распределяем высоту
                for r in range(start_row, current_row):
                    if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < row_height:
                        ws.row_dimensions[r].height = row_height
            
            # Пустая строка между товарами
            current_row += 1
        
        # Сохранение
        if output_path is None:
            output_dir = Path(__file__).parent / 'output'
            output_dir.mkdir(exist_ok=True)
            output_path = output_dir / f'KP_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(output_path)
        print(f"✅ [KP EXCEL] Файл сохранен: {output_path}")
        
        return str(output_path)


def main():
    """Тестовая генерация"""
    
    # Для теста используем session_id из первого найденного КП
    from database.postgresql_manager import PostgreSQLManager
    from sqlalchemy import text
    
    db_manager = PostgreSQLManager()
    db_session = db_manager.get_session_direct()
    
    try:
        result = db_session.execute(text("""
            SELECT DISTINCT session_id 
            FROM kp_items 
            LIMIT 1
        """))
        
        row = result.fetchone()
        if not row:
            print("❌ Нет товаров в КП для теста")
            return
        
        session_id = row[0]
        print(f"🧪 Тестовая генерация для session_id: {session_id}")
        
        generator = KPExcelGenerator()
        output_path = generator.generate(session_id)
        
        print(f"✅ Файл создан: {output_path}")
    finally:
        db_session.close()


if __name__ == '__main__':
    main()

