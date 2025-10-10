#!/usr/bin/env python3
"""
Парсинг 272 валидных файлов с использованием правильного парсера
Работает напрямую с PostgreSQL через postgresql_manager
"""

import sys
from pathlib import Path
from datetime import datetime
import json
from sqlalchemy import text
import openpyxl
from openpyxl import load_workbook
from PIL import Image
import io
import hashlib
import re

sys.path.insert(0, str(Path(__file__).parent))

from database.postgresql_manager import PostgreSQLManager
from src.structure_parser import CommercialProposalParser

class SimpleParser:
    def __init__(self, db: PostgreSQLManager):
        self.db = db
        self.structure_parser = CommercialProposalParser()
        self.images_dir = Path(__file__).parent / "storage" / "images"
        self.images_dir.mkdir(parents=True, exist_ok=True)
    
    def save_image_from_excel(self, image_obj, table_id: str, row: int, col: int) -> dict:
        """Сохраняет изображение из Excel и возвращает данные о файле"""
        try:
            # Получаем данные изображения
            img_data = image_obj._data()
            
            # Создаем PIL Image
            pil_img = Image.open(io.BytesIO(img_data))
            
            # Генерируем cell_position (например A1, B5, N7)
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_position = f"{col_letter}{row}"
            
            # Генерируем хеш и имя файла
            img_hash = hashlib.md5(img_data).hexdigest()[:8]
            filename = f"{table_id}_{cell_position}_{img_hash}.png"
            filepath = self.images_dir / filename
            
            # Сохраняем
            pil_img.save(filepath, "PNG")
            
            return {
                'filename': filename,
                'cell_position': cell_position,
                'row_number': row,
                'column_number': col
            }
        except Exception as e:
            print(f"      ⚠️  Ошибка сохранения изображения: {e}")
            return None
    
    def parse_file(self, file_path: Path, project_id: int) -> dict:
        """
        Полный парсинг файла
        """
        
        table_id = file_path.stem.split('_')[2] if len(file_path.stem.split('_')) > 2 else 'unknown'
        
        # Загружаем файл для изображений (БЕЗ data_only чтобы получить blob данные)
        workbook_images = load_workbook(file_path, data_only=False)
        worksheet_images = workbook_images[workbook_images.sheetnames[0]]
        
        # Загружаем файл для данных (С data_only чтобы получить вычисленные значения)
        workbook_data = load_workbook(file_path, data_only=True)
        worksheet_data = workbook_data[workbook_data.sheetnames[0]]
        sheet_name = workbook_data.sheetnames[0]
        
        # Валидация структуры
        validation = self.structure_parser.validate_table_structure(worksheet_data, sheet_name)
        
        if not validation['is_valid']:
            return {'success': False, 'error': 'Validation failed', 'products': 0, 'offers': 0, 'images': 0}
        
        # Извлекаем изображения СНАЧАЛА
        images_by_row = {}  # {row_idx: [{'filename': ..., 'cell_position': ..., 'row_number': ..., 'column_number': ...}]}
        
        if hasattr(worksheet_images, '_images'):
            for img in worksheet_images._images:
                try:
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        row = img.anchor._from.row + 1
                        col = img.anchor._from.col + 1
                        
                        img_data = self.save_image_from_excel(img, table_id, row, col)
                        
                        if img_data:
                            if row not in images_by_row:
                                images_by_row[row] = []
                            images_by_row[row].append(img_data)
                except Exception as e:
                    print(f"      ⚠️  Ошибка извлечения изображения: {e}")
        
        # Определяем начало данных (строка 4 согласно структуре)
        data_start_row = 4
        
        # Определяем типы маршрутов из заголовков (строки 2-3)
        # Сначала проверяем строку 2
        route_jd_name = str(worksheet_data.cell(2, 6).value or '').strip() if worksheet_data.cell(2, 6).value else ''
        route_avia_name = str(worksheet_data.cell(2, 9).value or '').strip() if worksheet_data.cell(2, 9).value else ''
        
        # Если в строке 2 пусто, проверяем строку 3
        if not route_jd_name:
            route_jd_name = str(worksheet_data.cell(3, 6).value or '').strip() if worksheet_data.cell(3, 6).value else 'ЖД'
        if not route_avia_name:
            route_avia_name = str(worksheet_data.cell(3, 9).value or '').strip() if worksheet_data.cell(3, 9).value else 'АВИА'
        
        # Проверяем, есть ли "Образец" в названии маршрута
        is_sample_route = 'образец' in route_avia_name.lower()
        
        # Парсим товары (с учетом многострочных товаров)
        products_data = []
        skip_rows = set()  # Строки которые уже обработали как дополнительные предложения
        
        for row_idx in range(data_start_row, min(worksheet_data.max_row + 1, 500)):
            if row_idx in skip_rows:
                continue
                
            # Название (столбец B = 2)
            name_cell = worksheet_data.cell(row_idx, 2)
            name = str(name_cell.value).strip() if name_cell.value else None
            
            if not name or name in ['None', ''] or len(name) < 3:
                continue
            
            # Пропускаем заголовки
            if any(kw in name.lower() for kw in ['фото', 'наименование', 'название', 'характеристики', 'тираж', 'цена']):
                continue
            
            # Характеристики (столбец C = 3)
            desc_cell = worksheet_data.cell(row_idx, 3)
            description = str(desc_cell.value) if desc_cell.value else None
            
            # Кастом (столбец D = 4)
            custom_cell = worksheet_data.cell(row_idx, 4)
            custom_field = str(custom_cell.value) if custom_cell.value else None
            
            # Парсим ценовые предложения для этого товара
            offers = []
            
            # Функция для парсинга предложений из строки
            def parse_offers_from_row(r_idx):
                row_offers = []
                
                # Тираж (столбец E = 5)
                qty_cell = worksheet_data.cell(r_idx, 5)
                if not qty_cell.value:
                    return row_offers
                    
                try:
                    qty_str = str(qty_cell.value).strip()
                    if '-' in qty_str:
                        qty_str = qty_str.split('-')[0].strip()
                    
                    # Фильтр аномальных тиражей
                    qty = int(''.join(filter(str.isdigit, qty_str))) if qty_str else None
                    if not qty or qty <= 0 or qty > 1_000_000:
                        return row_offers
                    
                    # ЖД маршрут (столбцы F, G, H = 6, 7, 8)
                    price_usd_jd = None
                    price_rub_jd = None
                    delivery_time_jd = 60
                    
                    cell_f = worksheet_data.cell(r_idx, 6)
                    if cell_f.value:
                        try:
                            price_usd_jd = float(str(cell_f.value).replace(',', '.').strip())
                            # Фильтр аномальных цен
                            if price_usd_jd > 10_000:
                                price_usd_jd = None
                        except:
                            pass
                    
                    cell_g = worksheet_data.cell(r_idx, 7)
                    if cell_g.value:
                        try:
                            price_rub_jd = float(str(cell_g.value).replace(',', '.').replace(' ', '').strip())
                            # Фильтр аномальных цен
                            if price_rub_jd > 1_000_000:
                                price_rub_jd = None
                        except:
                            pass
                    
                    cell_h = worksheet_data.cell(r_idx, 8)
                    if cell_h.value:
                        try:
                            # ИСПРАВЛЕНО: Парсинг срока доставки с учетом тире "17-19" -> 19
                            delivery_str = str(cell_h.value).strip()
                            if '-' in delivery_str:
                                # Берем число ПОСЛЕ тире
                                delivery_str = delivery_str.split('-')[-1].strip()
                            delivery_time_jd = int(''.join(filter(str.isdigit, delivery_str)))
                            # Фильтр аномальных сроков
                            if delivery_time_jd > 365:
                                delivery_time_jd = 60
                        except:
                            pass
                    
                    if price_usd_jd or price_rub_jd:
                        row_offers.append({
                            'quantity': qty,
                            'price_usd': price_usd_jd,
                            'price_rub': price_rub_jd,
                            'route': 'ЖД',
                            'delivery_time_days': delivery_time_jd
                        })
                    
                    # АВИА/Образец маршрут (столбцы I, J, K = 9, 10, 11)
                    # ИСПРАВЛЕНО: Пропускаем если это "Образец" и нет цен
                    if not is_sample_route:
                        price_usd_avia = None
                        price_rub_avia = None
                        delivery_time_avia = 30
                        
                        cell_i = worksheet_data.cell(r_idx, 9)
                        if cell_i.value:
                            try:
                                price_usd_avia = float(str(cell_i.value).replace(',', '.').strip())
                                # Фильтр аномальных цен
                                if price_usd_avia > 10_000:
                                    price_usd_avia = None
                            except:
                                pass
                        
                        cell_j = worksheet_data.cell(r_idx, 10)
                        if cell_j.value:
                            try:
                                price_rub_avia = float(str(cell_j.value).replace(',', '.').replace(' ', '').strip())
                                # Фильтр аномальных цен
                                if price_rub_avia > 1_000_000:
                                    price_rub_avia = None
                            except:
                                pass
                        
                        cell_k = worksheet_data.cell(r_idx, 11)
                        if cell_k.value:
                            try:
                                # ИСПРАВЛЕНО: Парсинг срока доставки с учетом тире
                                delivery_str = str(cell_k.value).strip()
                                if '-' in delivery_str:
                                    # Берем число ПОСЛЕ тире
                                    delivery_str = delivery_str.split('-')[-1].strip()
                                delivery_time_avia = int(''.join(filter(str.isdigit, delivery_str)))
                                # Фильтр аномальных сроков
                                if delivery_time_avia > 365:
                                    delivery_time_avia = 30
                            except:
                                pass
                        
                        if price_usd_avia or price_rub_avia:
                            row_offers.append({
                                'quantity': qty,
                                'price_usd': price_usd_avia,
                                'price_rub': price_rub_avia,
                                'route': 'АВИА',
                                'delivery_time_days': delivery_time_avia
                            })
                    
                except Exception as e:
                    print(f"         ⚠️  Ошибка парсинга строки {r_idx}: {e}")
                
                return row_offers
            
            # Парсим предложения из текущей строки
            offers.extend(parse_offers_from_row(row_idx))
            
            # Проверяем следующие строки на дополнительные предложения
            # (пустое название, но есть тираж и цены)
            check_row = row_idx + 1
            while check_row < min(worksheet_data.max_row + 1, 500):
                next_name = worksheet_data.cell(check_row, 2).value
                next_qty = worksheet_data.cell(check_row, 5).value
                
                # Если есть название - это новый товар, прекращаем
                if next_name and str(next_name).strip() and len(str(next_name).strip()) >= 3:
                    break
                
                # Если есть тираж - это дополнительное предложение
                if next_qty:
                    additional_offers = parse_offers_from_row(check_row)
                    offers.extend(additional_offers)
                    skip_rows.add(check_row)
                    check_row += 1
                else:
                    break
            
            # Ищем изображения для этого товара (±2 строки от начальной)
            product_images = []
            for check_row in range(max(1, row_idx - 2), row_idx + 3):
                if check_row in images_by_row:
                    product_images.extend(images_by_row[check_row])
            
            products_data.append({
                'name': name,
                'description': description,
                'custom_field': custom_field,
                'row': row_idx,
                'offers': offers,
                'images': product_images
            })
        
        if len(products_data) == 0:
            return {'success': False, 'error': 'No products found', 'products': 0, 'offers': 0, 'images': 0}
        
        # Сохраняем в БД
        with self.db.get_session() as session:
            products_saved = 0
            offers_saved = 0
            images_saved = 0
            
            for product_data in products_data:
                # Проверяем дубликат
                existing = session.execute(text("""
                    SELECT id FROM products 
                    WHERE project_id = :project_id AND name = :name
                """), {'project_id': project_id, 'name': product_data['name']}).fetchone()
                
                if existing:
                    continue
                
                # Сохраняем товар
                result = session.execute(text("""
                    INSERT INTO products (
                        project_id, name, description, custom_field,
                        created_at, updated_at
                    ) VALUES (
                        :project_id, :name, :description, :custom_field,
                        :created_at, :updated_at
                    ) RETURNING id
                """), {
                    'project_id': project_id,
                    'name': product_data['name'],
                    'description': product_data.get('description'),
                    'custom_field': product_data.get('custom_field'),
                    'created_at': datetime.now(),
                    'updated_at': datetime.now()
                })
                
                product_id_db = result.fetchone()[0]
                products_saved += 1
                
                # Сохраняем предложения
                for offer in product_data.get('offers', []):
                    session.execute(text("""
                        INSERT INTO price_offers (
                            product_id, quantity,
                            price_usd, price_rub,
                            route, delivery_time_days,
                            created_at, updated_at
                        ) VALUES (
                            :product_id, :quantity,
                            :price_usd, :price_rub,
                            :route, :delivery_time_days,
                            :created_at, :updated_at
                        )
                    """), {
                        'product_id': product_id_db,
                        'quantity': offer.get('quantity'),
                        'price_usd': offer.get('price_usd'),
                        'price_rub': offer.get('price_rub'),
                        'route': offer.get('route'),
                        'delivery_time_days': offer.get('delivery_time_days'),
                        'created_at': datetime.now(),
                        'updated_at': datetime.now()
                    })
                    offers_saved += 1
                
                # Сохраняем изображения
                for img_data in product_data.get('images', []):
                    try:
                        session.execute(text("""
                            INSERT INTO product_images (
                                product_id, table_id, image_filename,
                                cell_position, row_number, column_number,
                                created_at, updated_at
                            ) VALUES (
                                :product_id, :table_id, :image_filename,
                                :cell_position, :row_number, :column_number,
                                :created_at, :updated_at
                            )
                        """), {
                            'product_id': product_id_db,
                            'table_id': table_id,
                            'image_filename': img_data.get('filename'),
                            'cell_position': img_data.get('cell_position'),
                            'row_number': img_data.get('row_number'),
                            'column_number': img_data.get('column_number'),
                            'created_at': datetime.now(),
                            'updated_at': datetime.now()
                        })
                        images_saved += 1
                    except Exception as e:
                        print(f"      ⚠️  Ошибка сохранения изображения в БД: {e}")
            
            # Обновляем статус проекта
            session.execute(text("""
                UPDATE projects 
                SET parsing_status = 'completed',
                    updated_at = :updated_at
                WHERE id = :project_id
            """), {'updated_at': datetime.now(), 'project_id': project_id})
            
            session.commit()
            
            return {
                'success': True,
                'products': products_saved,
                'offers': offers_saved,
                'images': images_saved
            }

def main():
    # Загружаем список валидных pending файлов
    valid_list = Path(__file__).parent / "valid_pending_files.txt"
    if not valid_list.exists():
        print("❌ Файл valid_pending_files.txt не найден!")
        print("💡 Запустите: python3 validate_pending_only.py")
        return
    
    with open(valid_list, 'r', encoding='utf-8') as f:
        valid_files = [line.strip() for line in f if line.strip()]
    
    total_files = len(valid_files)
    print(f"🚀 ПАРСИНГ {total_files} ВАЛИДНЫХ PENDING ФАЙЛОВ")
    print("=" * 80)
    
    db = PostgreSQLManager()
    parser = SimpleParser(db)
    excel_dir = Path(__file__).parent / "storage" / "excel_files"
    
    success_count = 0
    error_count = 0
    
    products_total = 0
    offers_total = 0
    images_total = 0
    
    for i, filename in enumerate(valid_files, 1):
        file_path = excel_dir / filename
        
        # Извлекаем project_id
        match = re.search(r'project_(\d+)_', filename)
        if not match:
            print(f"[{i}/{total_files}] ❌ Неверный формат имени: {filename}")
            error_count += 1
            continue
        
        project_id = int(match.group(1))
        
        print(f"[{i}/{total_files}] {filename}")
        
        try:
            result = parser.parse_file(file_path, project_id)
            
            if result['success']:
                success_count += 1
                products_total += result['products']
                offers_total += result['offers']
                images_total += result['images']
                
                img_icon = '🖼️' if result['images'] > 0 else '❌'
                print(f"   ✅ Товары: {result['products']} | Предложения: {result['offers']} | Изображения: {result['images']} {img_icon}")
            else:
                error_count += 1
                print(f"   ⚠️  {result.get('error', 'Unknown error')}")
        
        except Exception as e:
            error_count += 1
            print(f"   ❌ {str(e)[:100]}")
        
        # Прогресс каждые 20 файлов
        if i % 20 == 0:
            print(f"\n📊 Прогресс: {i}/{total_files} | ✅ {success_count} | ❌ {error_count}\n")
    
    # Финальный отчет
    print("\n" + "=" * 80)
    print("✅ ПАРСИНГ ЗАВЕРШЕН")
    print("=" * 80)
    print(f"Успешно: {success_count}/{total_files} ({success_count/total_files*100:.1f}%)")
    print(f"❌ Ошибки: {error_count}")
    print()
    print(f"📦 Извлечено:")
    print(f"   Товары: {products_total}")
    print(f"   Предложения: {offers_total}")
    print(f"   Изображения: {images_total} 🖼️")
    print()
    
    # Сохраняем отчет
    report = {
        'timestamp': datetime.now().isoformat(),
        'total_files': total_files,
        'success': success_count,
        'errors': error_count,
        'products_total': products_total,
        'offers_total': offers_total,
        'images_total': images_total
    }
    
    with open('parsing_final_report.json', 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    
    print("✅ Отчет: parsing_final_report.json")
    print("=" * 80)

if __name__ == "__main__":
    main()

