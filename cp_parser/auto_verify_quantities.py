#!/usr/bin/env python3
"""
Автоматическая проверка тиражей в проектах Template 4
Скачивает Excel файлы и сравнивает тиражи с БД
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path.cwd()))

import json
import requests
import gspread
from google.oauth2.service_account import Credentials
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text
import os
from openpyxl import load_workbook
import csv
from datetime import datetime

# Railway DB
os.environ['USE_RAILWAY_DB'] = 'true'

def download_excel(sheet_id, output_path):
    """Скачивает Google Sheet как Excel без изображений"""
    try:
        # Авторизация
        scope = ['https://www.googleapis.com/auth/spreadsheets.readonly',
                 'https://www.googleapis.com/auth/drive.readonly']
        creds = Credentials.from_service_account_file('service_account.json', scopes=scope)
        
        # Получаем URL для экспорта (без изображений - быстро)
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        
        # Скачиваем с токеном
        response = requests.get(export_url, headers={'Authorization': f'Bearer {creds.token}'})
        
        if response.status_code == 200:
            with open(output_path, 'wb') as f:
                f.write(response.content)
            return True
        else:
            print(f"  ❌ Ошибка скачивания: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"  ❌ Ошибка: {e}")
        return False

def extract_quantity_from_excel(excel_path, row_number):
    """Извлекает тираж из колонки E указанной строки"""
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        
        # Колонка E = 5
        cell_value = ws.cell(row=row_number, column=5).value
        
        if cell_value is None:
            return None
            
        # Пытаемся преобразовать в число
        try:
            return int(float(str(cell_value).replace(',', '').replace(' ', '')))
        except:
            return str(cell_value)
            
    except Exception as e:
        print(f"    ⚠️  Ошибка чтения Excel: {e}")
        return None

def verify_project(project_data, db, temp_dir):
    """Проверяет один проект"""
    project_id = project_data['id']
    sheet_id = project_data['sheet_id']
    
    print(f"\n{'='*80}")
    print(f"🔍 ПРОЕКТ ID {project_id}: {project_data['name'][:60]}")
    print(f"{'='*80}")
    
    # Скачиваем Excel
    excel_path = temp_dir / f"project_{project_id}.xlsx"
    print(f"📥 Скачиваю Excel...")
    
    if not download_excel(sheet_id, excel_path):
        return None
    
    print(f"✅ Скачан: {excel_path.stat().st_size / 1024:.1f} KB")
    
    # Получаем товары и тиражи из БД
    with db.get_session() as session:
        products = session.execute(text("""
            SELECT 
                p.id as product_id,
                p.name as product_name,
                p.row_number,
                po.quantity as db_quantity,
                po.route
            FROM products p
            INNER JOIN price_offers po ON p.id = po.product_id
            WHERE p.project_id = :project_id
            AND po.quantity >= 300
            AND po.quantity % 10 = 0
            ORDER BY p.row_number, po.route
        """), {'project_id': project_id}).fetchall()
        
        if not products:
            print(f"⚠️  Нет подозрительных товаров")
            return None
        
        print(f"\n📊 Проверяю {len(products)} записей...")
        
        results = []
        checked_rows = {}  # Кэш: row_number -> excel_quantity
        
        for prod in products:
            row_num = prod.row_number
            
            # Проверяем строку только один раз
            if row_num in checked_rows:
                excel_qty = checked_rows[row_num]
            else:
                excel_qty = extract_quantity_from_excel(excel_path, row_num)
                checked_rows[row_num] = excel_qty
            
            db_qty = int(prod.db_quantity)
            
            # Анализируем
            is_x10 = False
            status = "✓"
            
            if excel_qty is None:
                status = "⚠️  Нет данных"
            elif isinstance(excel_qty, str):
                status = "⚠️  Не число"
            elif excel_qty * 10 == db_qty:
                is_x10 = True
                status = "❌ x10"
            elif excel_qty == db_qty:
                status = "✅ OK"
            else:
                status = f"⚠️  Не совпадает ({excel_qty})"
            
            results.append({
                'product_id': prod.product_id,
                'product_name': prod.product_name,
                'row_number': row_num,
                'excel_quantity': excel_qty,
                'db_quantity': db_qty,
                'route': prod.route,
                'is_x10': is_x10,
                'status': status
            })
        
        # Статистика
        x10_count = sum(1 for r in results if r['is_x10'])
        ok_count = sum(1 for r in results if r['status'] == "✅ OK")
        
        print(f"\n📊 РЕЗУЛЬТАТЫ:")
        print(f"  ❌ Ошибка x10: {x10_count} ({x10_count/len(results)*100:.1f}%)")
        print(f"  ✅ Правильно: {ok_count} ({ok_count/len(results)*100:.1f}%)")
        
        # Показываем примеры ошибок
        if x10_count > 0:
            print(f"\n❌ ПРИМЕРЫ ОШИБОК x10:")
            for r in results[:5]:
                if r['is_x10']:
                    print(f"  Строка {r['row_number']}: Excel={r['excel_quantity']} → БД={r['db_quantity']}")
        
        return {
            'project_id': project_id,
            'project_name': project_data['name'],
            'url': project_data['url'],
            'total_records': len(results),
            'x10_errors': x10_count,
            'ok_count': ok_count,
            'results': results
        }

def main():
    print("="*80)
    print("🔍 АВТОМАТИЧЕСКАЯ ПРОВЕРКА ТИРАЖЕЙ")
    print("="*80)
    
    # Загружаем список проектов
    with open('projects_to_verify.json', 'r', encoding='utf-8') as f:
        projects = json.load(f)
    
    print(f"\n📋 Проектов для проверки: {len(projects)}")
    
    # Создаем временную папку
    temp_dir = Path('temp_verification')
    temp_dir.mkdir(exist_ok=True)
    
    # БД
    db = PostgreSQLManager()
    
    # Проверяем каждый проект
    all_results = []
    
    # Проверяем все 50 проектов
    for i, project in enumerate(projects, 1):
        print(f"\n\n{'#'*80}")
        print(f"# ПРОЕКТ {i}/{len(projects)}")
        print(f"{'#'*80}")
        
        result = verify_project(project, db, temp_dir)
        if result:
            all_results.append(result)
    
    # Итоговый отчет
    print("\n\n" + "="*80)
    print("📊 ИТОГОВЫЙ ОТЧЕТ")
    print("="*80)
    
    total_x10_errors = sum(r['x10_errors'] for r in all_results)
    total_records = sum(r['total_records'] for r in all_results)
    
    print(f"\n✅ Проверено проектов: {len(all_results)}")
    print(f"📊 Всего записей: {total_records:,}")
    
    if total_records > 0:
        print(f"❌ Ошибок x10: {total_x10_errors:,} ({total_x10_errors/total_records*100:.1f}%)")
    else:
        print(f"⚠️  Не удалось проверить ни одного проекта")
    
    print(f"\n📋 ПРОЕКТЫ С ОШИБКАМИ:")
    for r in all_results:
        if r['x10_errors'] > 0:
            print(f"\n  ID {r['project_id']}: {r['project_name'][:60]}")
            print(f"    Ошибок: {r['x10_errors']}/{r['total_records']} ({r['x10_errors']/r['total_records']*100:.1f}%)")
    
    # Сохраняем детальный отчет
    csv_path = Path(f"АВТОПРОВЕРКА_РЕЗУЛЬТАТЫ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'ID Проекта', 'Название проекта', 'ID Товара', 'Название товара',
            'Строка', 'Тираж Excel', 'Тираж БД', 'Маршрут', 'Статус', 'Ошибка x10'
        ])
        
        for proj_result in all_results:
            for r in proj_result['results']:
                writer.writerow([
                    proj_result['project_id'],
                    proj_result['project_name'],
                    r['product_id'],
                    r['product_name'],
                    r['row_number'],
                    r['excel_quantity'],
                    r['db_quantity'],
                    r['route'],
                    r['status'],
                    'ДА' if r['is_x10'] else 'НЕТ'
                ])
    
    print(f"\n✅ Детальный отчет: {csv_path}")
    print(f"\n💾 Excel файлы сохранены в: {temp_dir}")
    
    print("\n" + "="*80)

if __name__ == '__main__':
    main()

