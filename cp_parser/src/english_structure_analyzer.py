#!/usr/bin/env python3
"""
Анализатор английской структуры таблиц.
Поддерживает файлы с листом 'Calculation' и английскими заголовками.
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Добавляем путь к модулям проекта
sys.path.append(str(Path(__file__).parent.parent))

from database import db_manager
from database.models import Project
from loguru import logger

class EnglishStructureAnalyzer:
    """Анализатор английской структуры таблиц"""
    
    def __init__(self):
        self.logger = logger
        
        # Английские листы для поиска
        self.expected_sheets = ['Calculation']
        
        # Основные заголовки во 2 строке (столбцы A-E)
        self.expected_main_columns = [
            'photo',
            'name', 
            'description',
            'custom',
            'quantity'  # может быть "quantity, pcs" или просто "quantity"
        ]
        
        # Возможные маршруты доставки во 2 строке
        self.delivery_routes = [
            'price per item, including air delivery to dubai',  # АВИА
            'air delivery',
            'price per item, including sea delivery to dubai',  # МОРЕ
            'sea delivery',
            'price per item, including',  # общий паттерн
        ]
        
        # Заголовки цен в 3 строке
        self.price_headers = [
            'price per item, $',
            'price per item, aed', 
            'circulation period',
            'period'
        ]
        
        # Заголовки образцов во 2 строке
        self.sample_headers = [
            'sample price (does not include delivery cost), $',
            'sample price',
            'stock sample',
            'customed sample',
            'sample period'
        ]
        
        # Дополнительные фото
        self.additional_photos = [
            'additional photos',
            'add photos'
        ]
    
    def find_matching_sheet(self, workbook):
        """Находит подходящий лист в файле"""
        available_sheets = [sheet.lower() for sheet in workbook.sheetnames]
        
        for expected_sheet in self.expected_sheets:
            if expected_sheet.lower() in available_sheets:
                # Возвращаем оригинальное имя листа
                for original_name in workbook.sheetnames:
                    if original_name.lower() == expected_sheet.lower():
                        return original_name
        
        return None
    
    def validate_english_structure(self, worksheet):
        """Валидация английской структуры таблицы"""
        result = {
            'is_valid': False,
            'confidence_score': 0.0,
            'matched_main_columns': {},
            'matched_delivery_routes': {},
            'matched_price_headers': {},
            'matched_sample_headers': {},
            'validation_errors': [],
            'structure_details': {}
        }
        
        try:
            matches = 0
            total_checks = 0
            
            # 1. Проверка основных столбцов (A-E) во 2 строке
            main_column_matches = 0
            for col_idx, expected_header in enumerate(self.expected_main_columns, 1):
                cell = worksheet.cell(row=2, column=col_idx)
                actual_value = str(cell.value).strip().lower() if cell.value else ""
                
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                total_checks += 1
                
                # Более гибкая проверка для quantity
                if expected_header == 'quantity':
                    is_match = 'quantity' in actual_value or 'pcs' in actual_value
                else:
                    is_match = expected_header in actual_value or actual_value in expected_header
                
                if is_match:
                    matches += 1
                    main_column_matches += 1
                    result['matched_main_columns'][col_letter] = {
                        'expected': expected_header,
                        'actual': actual_value,
                        'matched': True
                    }
                else:
                    result['matched_main_columns'][col_letter] = {
                        'expected': expected_header,
                        'actual': actual_value,
                        'matched': False
                    }
            
            # 2. Поиск маршрутов доставки (столбцы F и далее) во 2 строке
            delivery_matches = 0
            for col_idx in range(6, 16):  # Проверяем столбцы F-O
                cell = worksheet.cell(row=2, column=col_idx)
                actual_value = str(cell.value).strip().lower() if cell.value else ""
                
                if actual_value:  # Только если есть значение
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    total_checks += 1
                    
                    # Проверяем на соответствие маршрутам доставки
                    route_match = False
                    for route in self.delivery_routes:
                        if route in actual_value or any(word in actual_value for word in ['delivery', 'air', 'sea']):
                            route_match = True
                            break
                    
                    if route_match:
                        matches += 1
                        delivery_matches += 1
                        result['matched_delivery_routes'][col_letter] = {
                            'actual': actual_value,
                            'matched': True
                        }
            
            # 3. Проверка заголовков цен в 3 строке
            price_matches = 0
            for col_idx in range(6, 16):  # Проверяем столбцы F-O
                cell = worksheet.cell(row=3, column=col_idx)
                actual_value = str(cell.value).strip().lower() if cell.value else ""
                
                if actual_value:  # Только если есть значение
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    total_checks += 1
                    
                    # Проверяем на соответствие ценовым заголовкам
                    price_match = False
                    for price_header in self.price_headers:
                        if price_header in actual_value or any(word in actual_value for word in ['price', '$', 'aed', 'period', 'circulation']):
                            price_match = True
                            break
                    
                    if price_match:
                        matches += 1
                        price_matches += 1
                        result['matched_price_headers'][col_letter] = {
                            'actual': actual_value,
                            'matched': True
                        }
            
            # 4. Поиск заголовков образцов во 2 строке
            sample_matches = 0
            for col_idx in range(10, 20):  # Проверяем правые столбцы
                cell = worksheet.cell(row=2, column=col_idx)
                actual_value = str(cell.value).strip().lower() if cell.value else ""
                
                if actual_value:  # Только если есть значение
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    # Проверяем на соответствие образцам
                    sample_match = False
                    for sample_header in self.sample_headers:
                        if sample_header in actual_value or 'sample' in actual_value:
                            sample_match = True
                            break
                    
                    if sample_match:
                        sample_matches += 1
                        result['matched_sample_headers'][col_letter] = {
                            'actual': actual_value,
                            'matched': True
                        }
            
            # Вычисляем рейтинг соответствия
            confidence_score = (matches / total_checks) * 100 if total_checks > 0 else 0
            result['confidence_score'] = confidence_score
            
            # Дополнительные бонусы за английскую структуру
            if main_column_matches >= 4:  # Если основные столбцы совпадают
                confidence_score += 10
            if delivery_matches >= 1:  # Если есть маршруты доставки
                confidence_score += 15
            if sample_matches >= 1:  # Если есть образцы
                confidence_score += 5
            
            result['confidence_score'] = min(confidence_score, 100)  # Ограничиваем 100%
            
            # Считаем структуру валидной если соответствие >= 70%
            result['is_valid'] = result['confidence_score'] >= 70
            
            # Сохраняем детали структуры
            result['structure_details'] = {
                'main_columns_matches': main_column_matches,
                'delivery_routes_matches': delivery_matches,
                'price_headers_matches': price_matches,
                'sample_headers_matches': sample_matches,
                'total_checks': total_checks,
                'total_matches': matches
            }
            
            self.logger.info(f"Английская структура: совпадений {matches}/{total_checks} ({result['confidence_score']:.1f}%)")
            
        except Exception as e:
            result['validation_errors'].append(f"Ошибка валидации английской структуры: {str(e)}")
        
        return result
    
    def analyze_english_file(self, file_path: Path):
        """Анализ одного файла на соответствие английской структуре"""
        result = {
            'is_parseable': False,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'error': None,
            'sheet_name': None,
            'structure_matches': 0,
            'total_expected': 15,
            'confidence_score': 0.0
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet_name = self.find_matching_sheet(workbook)
            
            if sheet_name:
                result['sheet_name'] = sheet_name
                worksheet = workbook[sheet_name]
                validation_result = self.validate_english_structure(worksheet)
                
                result['is_parseable'] = validation_result['is_valid']
                result['confidence_score'] = validation_result['confidence_score']
                result['structure_matches'] = validation_result['structure_details'].get('total_matches', 0)
                result['validation_details'] = validation_result
                
                if not validation_result['is_valid']:
                    result['error'] = f"Низкий рейтинг соответствия английской структуре: {validation_result['confidence_score']:.1f}%"
            else:
                result['error'] = f"Не найден лист 'Calculation' в файле. Доступные листы: {workbook.sheetnames}"
                
        except InvalidFileException:
            result['error'] = "Файл поврежден или не является Excel файлом"
        except Exception as e:
            result['error'] = f"Ошибка анализа файла: {str(e)}"
        
        return result
    
    def analyze_all_english_files(self, max_files: int = 50):
        """Анализ всех файлов на соответствие английской структуре"""
        
        # Получаем проекты из базы данных
        with db_manager.get_session() as session:
            projects = session.query(Project).limit(max_files).all()
        
        results = {
            'total_files': len(projects),
            'parseable_files': [],
            'non_parseable_files': [],
            'error_files': [],
            'statistics': {
                'parseable_count': 0,
                'non_parseable_count': 0,
                'error_count': 0,
                'calculation_sheet_count': 0
            }
        }
        
        self.logger.info(f"🚀 Анализирую {len(projects)} файлов на английскую структуру...")
        
        for project in projects:
            file_path = Path(project.file_path)
            if not file_path.exists():
                continue
            
            analysis_result = self.analyze_english_file(file_path)
            
            # Считаем файлы с листом Calculation
            if analysis_result.get('sheet_name') == 'Calculation':
                results['statistics']['calculation_sheet_count'] += 1
            
            if analysis_result['is_parseable']:
                results['parseable_files'].append(analysis_result)
                results['statistics']['parseable_count'] += 1
            else:
                error = analysis_result.get('error', 'Неизвестная ошибка')
                if 'не найден лист' in error.lower():
                    results['non_parseable_files'].append(analysis_result)
                    results['statistics']['non_parseable_count'] += 1
                elif 'поврежден' in error.lower():
                    results['error_files'].append(analysis_result)
                    results['statistics']['error_count'] += 1
                else:
                    results['non_parseable_files'].append(analysis_result)
                    results['statistics']['non_parseable_count'] += 1
        
        return results

if __name__ == "__main__":
    analyzer = EnglishStructureAnalyzer()
    results = analyzer.analyze_all_english_files()
    
    print(f"\n📊 АНАЛИЗ АНГЛИЙСКОЙ СТРУКТУРЫ ТАБЛИЦ")
    print(f"=" * 60)
    print(f"📁 Всего файлов проанализировано: {results['total_files']}")
    print(f"✅ Подходят для парсинга: {results['statistics']['parseable_count']}")
    print(f"❌ Не подходят: {results['statistics']['non_parseable_count']}")
    print(f"💥 Ошибки: {results['statistics']['error_count']}")
    print(f"📋 Файлов с листом 'Calculation': {results['statistics']['calculation_sheet_count']}")
    
    success_rate = (results['statistics']['parseable_count'] / results['total_files']) * 100 if results['total_files'] > 0 else 0
    print(f"🎯 Процент успешного парсинга английской структуры: {success_rate:.1f}%")
    
    if results['parseable_files']:
        print(f"\n✅ ФАЙЛЫ, ПОДХОДЯЩИЕ ДЛЯ АНГЛИЙСКОГО ПАРСИНГА:")
        for i, file_result in enumerate(results['parseable_files'][:10], 1):
            confidence = file_result.get('confidence_score', 0)
            print(f"  {i:2d}. {file_result['file_name']} - {confidence:.1f}% качества")
    
    print(f"\n💡 РЕКОМЕНДАЦИЯ:")
    if results['statistics']['parseable_count'] > 0:
        print(f"   Найдено {results['statistics']['parseable_count']} файлов с английской структурой!")
        print(f"   Рекомендуется добавить поддержку английской структуры в основной парсер.")
    else:
        print(f"   Файлов с подходящей английской структурой не найдено.")
        print(f"   Возможно, нужно скорректировать критерии поиска.")


