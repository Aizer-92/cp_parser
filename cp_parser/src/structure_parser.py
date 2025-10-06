#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Парсер коммерческих предложений с проверкой структуры столбцов
Проверяет соответствие названий столбцов ожидаемым значениям
"""

import sys
import os
from pathlib import Path
import pandas as pd
import logging
from typing import Dict, List, Tuple, Optional, Any
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import re

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

from database.manager import db_manager, project_service, product_service, price_offer_service

class CommercialProposalParser:
    """Парсер коммерческих предложений с проверкой структуры"""
    
    def __init__(self):
        self.setup_logging()
        
        # Ожидаемые названия основных столбцов (строка 2) - русские и английские
        self.expected_main_columns = {
            'A': ['фото', 'photo'],  # Добавляем поддержку столбца A
            'B': ['наименование', 'название', 'товар', 'product', 'name'],
            'C': ['характеристики', 'описание', 'спецификация', 'description'],
            'D': ['кастом', 'custom'],
            'E': ['тираж', 'тираж, шт', 'количество', 'qty', 'quantity', 'quantity, pcs']
        }
        
        # Ожидаемые маршруты (строка 2) - русские и английские
        self.expected_routes = {
            'F': ['доставка жд', 'доставка мск', 'мск', 'москва', 'доставка', 'sea delivery', 'price per item, including sea delivery',
                  # Английские альтернативы для авиа в столбце F
                  'price per item, including air delivery', 'price per item, including air delivery to dubai', 'air delivery'],
            'I': ['доставка авиа', 'авиа', 'авиа доставка', 'air delivery', 'price per item, including air delivery',
                  # Английские альтернативы для образцов в столбце I  
                  'sample price', 'sample price (does not include delivery cost)', 'sample'],
            'L': ['образец', 'sample', 'sample price']
        }
        
        # Ожидаемые названия ценовых столбцов (строка 3) - русские и английские
        self.expected_price_columns = {
            'F': ['цена за шт', 'цена за шт.', 'цена за шт., $', 'price', 'price per item, $',
                  # Английские термины для цен в долларах
                  'price per pcs ($)', 'price per item, $'],
            'G': ['цена за шт', 'цена за шт., руб', 'цена за шт.', 'руб', 'price per item, aed',
                  # Английские термины для цен в других валютах и сроков
                  'price per item, aed', 'circulation period', 'period (days)'],
            'H': ['срок тиража', 'срок тиража, к.д.', 'срок', 'delivery', 'period', 'circulation period',
                  # Расширенные английские термины для сроков
                  'period (days)', 'circulation period', 'production', 'calendar days'],
            'I': ['цена за шт', 'цена за шт., $', 'price', 'price per item, $',
                  # Английские термины для образцов в столбце I
                  'sample price', 'sample price (does not include delivery cost)', 'stock sample'],
            'J': ['цена за шт', 'цена за шт., руб', 'руб', 'price per item, aed',
                  # Английские термины для дополнительных столбцов
                  'add photos', 'additional photos', 'customed sample'],
            'K': ['срок тиража', 'срок тиража, к.д.', 'срок', 'period', 'circulation period',
                  # Английские термины для сроков образцов
                  'sample period', 'customed sample'],
            'L': ['цена за шт', 'образец', 'цена образца', 'цена за шт., руб', 'sample price', 'price',
                  # Расширенные термины для образцов
                  'stock sample', 'sample period'],
            'N': ['срок с доставкой', 'срок с доставкой, к.д.', 'срок образца', 'period', 'sample period',
                  # Английские термины для сроков образцов
                  'sample period', 'stock 17-27 days']
        }
        
        # Ожидаемые листы (русские и английские)
        self.expected_sheets = [
            # Русские листы
            'Copy of Просчет', 'Просчет', 'Copy of Прочет', 'Прочет',
            # Английские листы
            'Calculation'
        ]
    
    def setup_logging(self):
        """Настройка логирования"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
    
    def normalize_text(self, text: str) -> str:
        """Нормализация текста для сравнения"""
        if not text or not isinstance(text, str):
            return ""
        # Убираем лишние пробелы, переводим в нижний регистр
        return re.sub(r'\s+', ' ', text.strip().lower())
    
    def check_column_match(self, actual_value: str, expected_values: List[str], allow_empty: bool = False) -> bool:
        """Проверка соответствия значения столбца ожидаемым"""
        if not actual_value:
            return allow_empty
        
        normalized_actual = self.normalize_text(actual_value)
        
        for expected in expected_values:
            expected_normalized = self.normalize_text(expected)
            if expected_normalized in normalized_actual or normalized_actual in expected_normalized:
                return True
        
        return False
    
    def find_matching_sheet(self, workbook) -> Optional[str]:
        """Поиск подходящего листа"""
        for sheet_name in workbook.sheetnames:
            sheet_name_normalized = self.normalize_text(sheet_name)
            for expected_sheet in self.expected_sheets:
                expected_normalized = self.normalize_text(expected_sheet)
                if expected_normalized in sheet_name_normalized:
                    return sheet_name
        return None
    
    def validate_table_structure(self, worksheet, sheet_name=None) -> Dict[str, Any]:
        """Проверка структуры таблицы с двухуровневыми заголовками"""
        result = {
            'is_valid': False,
            'matched_main_columns': {},
            'matched_routes': {},
            'matched_price_columns': {},
            'validation_errors': [],
            'confidence_score': 0
        }
        
        try:
            # Читаем строки 2 и 3
            row2_values = {}
            row3_values = {}
            
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                col_num = ord(col_letter) - ord('A') + 1
                
                cell2 = worksheet.cell(2, col_num)
                cell3 = worksheet.cell(3, col_num)
                
                row2_values[col_letter] = cell2.value if cell2.value else ""
                row3_values[col_letter] = cell3.value if cell3.value else ""
            
            matches = 0
            total_checks = 0
            
            # 1. Проверяем основные столбцы (B, C, D, E) в строке 2
            for col_letter, expected_values in self.expected_main_columns.items():
                total_checks += 1
                actual_value = row2_values.get(col_letter, "")
                
                if self.check_column_match(actual_value, expected_values):
                    matches += 1
                    result['matched_main_columns'][col_letter] = {
                        'expected': expected_values,
                        'actual': actual_value,
                        'matched': True
                    }
                else:
                    result['matched_main_columns'][col_letter] = {
                        'expected': expected_values,
                        'actual': actual_value,
                        'matched': False
                    }
                    result['validation_errors'].append(f"Основной столбец {col_letter}: ожидалось {expected_values}, получено '{actual_value}'")
            
            # 2. Проверяем маршруты (F, I, L) в строке 2
            for col_letter, expected_routes in self.expected_routes.items():
                total_checks += 1
                actual_route = row2_values.get(col_letter, "")
                
                if self.check_column_match(actual_route, expected_routes):
                    matches += 1
                    result['matched_routes'][col_letter] = {
                        'expected': expected_routes,
                        'actual': actual_route,
                        'matched': True
                    }
                else:
                    result['matched_routes'][col_letter] = {
                        'expected': expected_routes,
                        'actual': actual_route,
                        'matched': False
                    }
                    result['validation_errors'].append(f"Маршрут {col_letter}: ожидалось {expected_routes}, получено '{actual_route}'")
            
            # 3. Проверяем ценовые столбцы в строке 3
            for col_letter, expected_values in self.expected_price_columns.items():
                total_checks += 1
                actual_value = row3_values.get(col_letter, "")
                
                # Специальная обработка для столбца N - разрешаем пустое значение
                allow_empty = (col_letter == 'N')
                
                if self.check_column_match(actual_value, expected_values, allow_empty):
                    matches += 1
                    result['matched_price_columns'][col_letter] = {
                        'expected': expected_values,
                        'actual': actual_value,
                        'matched': True
                    }
                else:
                    result['matched_price_columns'][col_letter] = {
                        'expected': expected_values,
                        'actual': actual_value,
                        'matched': False
                    }
                    if not allow_empty or actual_value:  # Не показываем ошибку для пустого N
                        result['validation_errors'].append(f"Ценовой столбец {col_letter}: ожидалось {expected_values}, получено '{actual_value}'")
            
            # Вычисляем рейтинг соответствия
            confidence_score = (matches / total_checks) * 100
            result['confidence_score'] = confidence_score
            
            # Новый алгоритм валидации на основе минимально необходимых столбцов
            result['is_valid'] = self._validate_minimum_required_columns(result)
            result['validation_details'] = self._get_validation_details(result)
            
            self.logger.info(f"Структура таблицы: совпадений {matches}/{total_checks} ({confidence_score:.1f}%)")
            
        except Exception as e:
            result['validation_errors'].append(f"Ошибка валидации: {str(e)}")
            self.logger.error(f"Ошибка валидации структуры: {e}")
        
        return result
    
    def _validate_minimum_required_columns(self, validation_result) -> bool:
        """
        Проверяет минимально необходимые столбцы для валидного парсинга:
        1. Фотография (столбец A)
        2. Название (столбец B)
        3. Хотя бы один полный маршрут (тип + тираж + цена + сроки)
        """
        
        # 1. Проверяем обязательные основные столбцы
        required_main_columns = ['A', 'B']  # Фото и Название
        main_columns_found = []
        
        for col_letter in required_main_columns:
            if col_letter in validation_result['matched_main_columns']:
                if validation_result['matched_main_columns'][col_letter]['matched']:
                    main_columns_found.append(col_letter)
        
        # Если нет фото И названия - не валидно
        if len(main_columns_found) < 2:
            self.logger.warning(f"❌ Не найдены обязательные столбцы. Найдено: {main_columns_found}, требуется: {required_main_columns}")
            return False
        
        # 2. Проверяем наличие хотя бы одного полного маршрута
        complete_routes = self._find_complete_routes(validation_result)
        
        if not complete_routes:
            self.logger.warning("❌ Не найдено ни одного полного маршрута (тип + тираж + цена + сроки)")
            return False
        
        self.logger.info(f"✅ Найдены обязательные столбцы: {main_columns_found}")
        self.logger.info(f"✅ Найдено полных маршрутов: {len(complete_routes)}")
        
        return True
    
    def _find_complete_routes(self, validation_result) -> list:
        """
        Находит полные маршруты, где определены:
        - Тип маршрута (в matched_routes)
        - Тираж (столбец E)
        - Цена в одной из валют
        - Сроки производства
        
        ВАЖНО: Все данные маршрута должны быть в соседних столбцах!
        """
        complete_routes = []
        
        # Проверяем есть ли тираж (столбец E)
        quantity_found = False
        if 'E' in validation_result['matched_main_columns']:
            if validation_result['matched_main_columns']['E']['matched']:
                quantity_found = True
        
        if not quantity_found:
            self.logger.warning("❌ Не найден столбец 'Тираж' (E)")
            return complete_routes
        
        # Определяем ожидаемые группы соседних столбцов для каждого маршрута
        route_groups = {
            'sea_route': {  # Морской маршрут
                'route_column': 'F',        # Тип маршрута в строке 2
                'price_columns': ['F', 'G'], # Цены в строке 3 (USD, RUB)
                'delivery_column': 'H'       # Сроки в строке 3
            },
            'air_route': {  # Авиа маршрут
                'route_column': 'I',        # Тип маршрута в строке 2
                'price_columns': ['I', 'J'], # Цены в строке 3 (USD, RUB)
                'delivery_column': 'K'       # Сроки в строке 3
            },
            # АЛЬТЕРНАТИВНАЯ КОНФИГУРАЦИЯ ДЛЯ АНГЛИЙСКИХ ФАЙЛОВ
            'english_air_route': {  # Английский авиа маршрут (F-G-H)
                'route_column': 'F',        # Тип маршрута в строке 2 (air delivery)
                'price_columns': ['F', 'G'], # Цены в строке 3 (USD, AED)
                'delivery_column': 'H'       # Сроки в строке 3 (circulation period)
            },
            'english_sample_route': {  # Английский образец (I)
                'route_column': 'I',        # Тип маршрута в строке 2 (sample price)
                'price_columns': ['I'],      # Цена в строке 3
                'delivery_column': None      # Сроки могут отсутствовать для образцов
            },
            'sample_route': {  # Образец
                'route_column': 'L',        # Тип маршрута в строке 2
                'price_columns': ['L'],      # Цена в строке 3
                'delivery_column': 'N'       # Сроки в строке 3 (пропускаем M)
            }
        }
        
        for route_name, route_config in route_groups.items():
            route_info = {
                'route_name': route_name,
                'column': route_config['route_column'],
                'route_type_found': False,
                'price_found': False,
                'delivery_time_found': False,
                'adjacent_columns_valid': True
            }
            
            # 1. Проверяем тип маршрута (строка 2)
            route_col = route_config['route_column']
            if route_col in validation_result['matched_routes']:
                if validation_result['matched_routes'][route_col]['matched']:
                    route_info['route_type_found'] = True
                    self.logger.debug(f"✅ Найден тип маршрута в столбце {route_col}")
            
            # 2. Проверяем цены в соседних столбцах (строка 3)
            price_found_in_group = False
            for price_col in route_config['price_columns']:
                if price_col in validation_result['matched_price_columns']:
                    if validation_result['matched_price_columns'][price_col]['matched']:
                        # Проверяем что это действительно цена
                        actual_value = validation_result['matched_price_columns'][price_col]['actual'].lower()
                        if any(word in actual_value for word in ['price', '$', 'руб', 'aed', 'цена']):
                            price_found_in_group = True
                            self.logger.debug(f"✅ Найдена цена в столбце {price_col}: {actual_value}")
                            break
            
            route_info['price_found'] = price_found_in_group
            
            # 3. Проверяем сроки в ожидаемом столбце (строка 3)
            delivery_col = route_config['delivery_column']
            if delivery_col is None:
                # Для некоторых маршрутов (например, английских образцов) сроки необязательны
                route_info['delivery_time_found'] = True
                self.logger.debug(f"✅ Сроки не требуются для маршрута {route_name}")
            elif delivery_col in validation_result['matched_price_columns']:
                if validation_result['matched_price_columns'][delivery_col]['matched']:
                    # Проверяем что это действительно сроки
                    actual_value = validation_result['matched_price_columns'][delivery_col]['actual'].lower()
                    if any(word in actual_value for word in ['срок', 'period', 'delivery', 'день', 'к.д', 'calendar']):
                        route_info['delivery_time_found'] = True
                        self.logger.debug(f"✅ Найдены сроки в столбце {delivery_col}: {actual_value}")
            
            # 4. Дополнительная проверка: убеждаемся что нет "чужих" данных в группе столбцов
            route_info['adjacent_columns_valid'] = self._validate_adjacent_columns(
                validation_result, route_config, route_name
            )
            
            # Считаем маршрут полным если найдены ВСЕ компоненты И столбцы соседние
            if (route_info['route_type_found'] and 
                route_info['price_found'] and 
                route_info['delivery_time_found'] and
                route_info['adjacent_columns_valid']):
                complete_routes.append(route_info)
                # Формируем описание диапазона столбцов для лога
                if route_config['delivery_column'] is not None:
                    column_range = f"{route_config['route_column']}-{route_config['delivery_column']}"
                else:
                    column_range = f"{route_config['route_column']}"
                self.logger.info(f"✅ Полный маршрут '{route_name}' найден в группе столбцов {column_range}")
            else:
                missing = []
                if not route_info['route_type_found']:
                    missing.append("тип маршрута")
                if not route_info['price_found']:
                    missing.append("цена")
                if not route_info['delivery_time_found']:
                    missing.append("сроки")
                if not route_info['adjacent_columns_valid']:
                    missing.append("соседние столбцы")
                
                self.logger.debug(f"❌ Маршрут '{route_name}' неполный. Не хватает: {', '.join(missing)}")
        
        return complete_routes
    
    def _validate_adjacent_columns(self, validation_result, route_config, route_name) -> bool:
        """
        Проверяет, что в группе столбцов маршрута нет посторонних данных
        """
        route_col = route_config['route_column']
        price_cols = route_config['price_columns']
        delivery_col = route_config['delivery_column']
        
        # Получаем все столбцы группы (исключаем None)
        all_route_columns = [route_col] + price_cols
        if delivery_col is not None:
            all_route_columns.append(delivery_col)
        
        # КРИТИЧЕСКАЯ ПРОВЕРКА: убеждаемся что в столбце сроков нет названия другого маршрута
        if delivery_col is not None and delivery_col in validation_result.get('matched_routes', {}):
            if validation_result['matched_routes'][delivery_col].get('matched', False):
                delivery_route_value = validation_result['matched_routes'][delivery_col].get('actual', '').lower()
                self.logger.warning(f"⚠️ В столбце сроков {delivery_col} найден маршрут: '{delivery_route_value}' - данные сдвинуты!")
                return False
        
        # Проверяем что между столбцами нет больших разрывов
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
        column_indices = [column_letters.index(col) for col in all_route_columns if col in column_letters]
        
        if not column_indices:
            return False
        
        column_indices.sort()
        
        # Проверяем что разрыв между первым и последним столбцом не больше ожидаемого
        expected_max_gap = {
            'sea_route': 2,              # F, G, H (разрыв 2)
            'air_route': 2,              # I, J, K (разрыв 2) 
            'english_air_route': 2,      # F, G, H (разрыв 2) - английский авиа
            'english_sample_route': 0,   # I (только один столбец)
            'sample_route': 2            # L, N (разрыв 1, но пропускаем M)
        }
        
        actual_gap = column_indices[-1] - column_indices[0]
        max_allowed_gap = expected_max_gap.get(route_name, 3)
        
        if actual_gap > max_allowed_gap:
            self.logger.warning(f"⚠️ Большой разрыв в столбцах маршрута '{route_name}': {actual_gap} > {max_allowed_gap}")
            return False
        
        return True
    
    def _get_validation_details(self, validation_result) -> dict:
        """Получает детальную информацию о качестве парсинга"""
        details = {
            'total_columns_found': 0,
            'main_columns_found': 0,
            'routes_found': 0,
            'price_columns_found': 0,
            'complete_routes_count': 0,
            'quality_score': 0.0
        }
        
        # Считаем найденные основные столбцы
        for col_data in validation_result.get('matched_main_columns', {}).values():
            if col_data.get('matched', False):
                details['main_columns_found'] += 1
                details['total_columns_found'] += 1
        
        # Считаем найденные маршруты
        for col_data in validation_result.get('matched_routes', {}).values():
            if col_data.get('matched', False):
                details['routes_found'] += 1
                details['total_columns_found'] += 1
        
        # Считаем найденные ценовые столбцы
        for col_data in validation_result.get('matched_price_columns', {}).values():
            if col_data.get('matched', False):
                details['price_columns_found'] += 1
                details['total_columns_found'] += 1
        
        # Считаем полные маршруты
        complete_routes = self._find_complete_routes(validation_result)
        details['complete_routes_count'] = len(complete_routes)
        
        # Вычисляем качество парсинга (0-100)
        max_possible_columns = 16  # Примерное максимальное количество столбцов
        details['quality_score'] = min((details['total_columns_found'] / max_possible_columns) * 100, 100)
        
        return details
    
    def parse_delivery_time(self, time_str: str) -> Optional[int]:
        """Парсинг времени доставки из строки типа '20-25' -> 25"""
        if not time_str or not isinstance(time_str, str):
            return None
        
        # Ищем числа в строке
        numbers = re.findall(r'\d+', str(time_str))
        if numbers:
            # Берем последнее число (как указано в требованиях)
            return int(numbers[-1])
        
        return None
    
    def clean_quantity(self, quantity_str: str) -> Optional[int]:
        """Очистка тиража от пробелов и парсинг"""
        if not quantity_str:
            return None
        
        # Убираем все пробелы и нечисловые символы
        cleaned = re.sub(r'[^\d]', '', str(quantity_str))
        
        try:
            return int(cleaned) if cleaned else None
        except (ValueError, TypeError):
            return None
    
    def analyze_file(self, file_path: Path) -> Dict[str, Any]:
        """Анализ файла на соответствие структуре"""
        result = {
            'file_path': str(file_path),
            'file_name': file_path.name,
            'analysis_status': 'unknown',
            'is_parseable': False,
            'sheet_found': None,
            'validation_result': None,
            'error': None
        }
        
        try:
            self.logger.info(f"🔍 Анализирую файл: {file_path.name}")
            
            # Открываем файл
            wb = load_workbook(file_path, read_only=True)
            
            # Ищем подходящий лист
            sheet_name = self.find_valid_sheet(wb)
            if not sheet_name:
                result['analysis_status'] = 'no_valid_sheet'
                result['error'] = f"Не найдены листы: {self.expected_sheets}"
                wb.close()
                return result
            
            result['sheet_found'] = sheet_name
            ws = wb[sheet_name]
            
            # Проверяем структуру таблицы
            validation_result = self.validate_table_structure(ws, sheet_name)
            result['validation_result'] = validation_result
            
            if validation_result['is_valid']:
                result['is_parseable'] = True
                result['analysis_status'] = 'parseable'
                self.logger.info(f"✅ Файл {file_path.name} подходит для парсинга (рейтинг: {validation_result['confidence_score']:.1f}%)")
            else:
                result['analysis_status'] = 'structure_mismatch'
                result['error'] = f"Структура не соответствует ожидаемой (рейтинг: {validation_result['confidence_score']:.1f}%)"
                self.logger.warning(f"❌ Файл {file_path.name} не подходит для парсинга")
            
            wb.close()
            
        except Exception as e:
            result['analysis_status'] = 'error'
            result['error'] = str(e)
            self.logger.error(f"❌ Ошибка анализа {file_path.name}: {e}")
        
        return result
    
    def analyze_test_files(self, max_files: int = 50) -> Dict[str, Any]:
        """Анализ тестовых файлов на соответствие структуре"""
        
        # Получаем проекты из базы данных
        with db_manager.get_session() as session:
            from database.models import Project
            projects = session.query(Project).limit(max_files).all()
        
        results = {
            'total_files': len(projects),
            'parseable_files': [],
            'non_parseable_files': [],
            'error_files': [],
            'statistics': {
                'parseable_count': 0,
                'structure_mismatch_count': 0,
                'no_sheet_count': 0,
                'error_count': 0
            }
        }
        
        self.logger.info(f"🚀 Начинаю анализ {len(projects)} файлов...")
        
        for project in projects:
            if not project.file_path or not Path(project.file_path).exists():
                self.logger.warning(f"⚠️ Файл не найден: {project.file_path}")
                continue
            
            file_path = Path(project.file_path)
            analysis_result = self.analyze_file(file_path)
            
            # Дополняем результат информацией о проекте
            analysis_result['project_id'] = project.id
            analysis_result['table_id'] = project.table_id
            analysis_result['file_name'] = file_path.name
            
            # Классифицируем результаты
            if analysis_result['is_parseable']:
                results['parseable_files'].append(analysis_result)
                results['statistics']['parseable_count'] += 1
            else:
                error = analysis_result.get('error', 'Неизвестная ошибка')
                if 'не найдены листы' in error.lower():
                    results['non_parseable_files'].append(analysis_result)
                    results['statistics']['no_sheet_count'] += 1
                elif 'file is not a zip file' in error.lower():
                    results['error_files'].append(analysis_result)
                    results['statistics']['error_count'] += 1
                else:
                    results['non_parseable_files'].append(analysis_result)
                    results['statistics']['structure_mismatch_count'] += 1
        
        return results
    
    def analyze_file(self, file_path: Path) -> Dict[str, Any]:
        """Анализ одного файла на соответствие структуре"""
        result = {
            'is_parseable': False,
            'file_path': str(file_path),
            'error': None
        }
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet_name = self.find_matching_sheet(workbook)
            
            if sheet_name:
                worksheet = workbook[sheet_name]
                validation_result = self.validate_table_structure(worksheet, sheet_name)
                result['is_parseable'] = validation_result['is_valid']
                result['confidence_score'] = validation_result['confidence_score']
                result['sheet_name'] = sheet_name
                
                if not validation_result['is_valid']:
                    result['error'] = f"Низкий рейтинг соответствия: {validation_result['confidence_score']:.1f}%"
            else:
                result['error'] = f"Не найдены листы: {self.expected_sheets}"
                
            workbook.close()
            
        except InvalidFileException:
            result['error'] = "File is not a zip file"
        except Exception as e:
            result['error'] = str(e)
        
        return result

    def print_analysis_report(self, results: Dict[str, Any]):
        """Вывод отчета об анализе"""
        print("\n" + "="*80)
        print("📊 ОТЧЕТ АНАЛИЗА СТРУКТУРЫ КОММЕРЧЕСКИХ ПРЕДЛОЖЕНИЙ")
        print("="*80)
        
        stats = results['statistics']
        total = results['total_files']
        
        print(f"\n📁 ОБЩАЯ СТАТИСТИКА:")
        print(f"Всего файлов проанализировано: {total}")
        print(f"✅ Подходят для парсинга: {stats['parseable_count']} ({stats['parseable_count']/total*100:.1f}%)")
        print(f"❌ Не подходящая структура: {stats['structure_mismatch_count']} ({stats['structure_mismatch_count']/total*100:.1f}%)")
        print(f"📄 Нет нужных листов: {stats['no_sheet_count']} ({stats['no_sheet_count']/total*100:.1f}%)")
        print(f"🚫 Ошибки анализа: {stats['error_count']} ({stats['error_count']/total*100:.1f}%)")
        
        if results['parseable_files']:
            print(f"\n✅ ФАЙЛЫ ГОТОВЫЕ К ПАРСИНГУ ({len(results['parseable_files'])} шт.):")
            print(f"{'№':<3} {'Файл':<50} {'Рейтинг':<8} {'Лист':<20}")
            print("-" * 85)
            for i, file_result in enumerate(results['parseable_files'][:15], 1):
                confidence = file_result['validation_result']['confidence_score']
                file_name = file_result['file_name'][:47] + "..." if len(file_result['file_name']) > 50 else file_result['file_name']
                sheet_name = file_result['sheet_found'][:17] + "..." if len(file_result['sheet_found']) > 20 else file_result['sheet_found']
                print(f"{i:2d}. {file_name:<50} {confidence:5.1f}% {sheet_name:<20}")
            
            if len(results['parseable_files']) > 15:
                print(f"      ... и еще {len(results['parseable_files']) - 15} файлов")
        
        # Показываем файлы с высоким рейтингом, но не прошедшие валидацию
        high_rating_files = [f for f in results['non_parseable_files'] 
                           if f.get('validation_result') and f['validation_result']['confidence_score'] >= 70]
        
        if high_rating_files:
            print(f"\n⚠️ ФАЙЛЫ С ВЫСОКИМ РЕЙТИНГОМ, НО НЕ ПРОШЕДШИЕ ВАЛИДАЦИЮ ({len(high_rating_files)} шт.):")
            print(f"{'№':<3} {'Файл':<50} {'Рейтинг':<8} {'Причина':<25}")
            print("-" * 90)
            for i, file_result in enumerate(high_rating_files[:10], 1):
                confidence = file_result['validation_result']['confidence_score']
                file_name = file_result['file_name'][:47] + "..." if len(file_result['file_name']) > 50 else file_result['file_name']
                error = file_result.get('error', 'Неизвестная ошибка')[:22] + "..." if len(file_result.get('error', '')) > 25 else file_result.get('error', 'Неизвестная ошибка')
                print(f"{i:2d}. {file_name:<50} {confidence:5.1f}% {error:<25}")
        
        if results['non_parseable_files'] and not high_rating_files:
            print(f"\n❌ ФАЙЛЫ С НЕПОДХОДЯЩЕЙ СТРУКТУРОЙ (первые 5):")
            for i, file_result in enumerate(results['non_parseable_files'][:5], 1):
                error = file_result.get('error', 'Неизвестная ошибка')
                confidence = ""
                if file_result.get('validation_result'):
                    confidence = f" (рейтинг: {file_result['validation_result']['confidence_score']:.1f}%)"
                print(f"  {i}. {file_result['file_name']}: {error}{confidence}")


if __name__ == "__main__":
    parser = CommercialProposalParser()
    
    # Анализируем ВСЕ файлы для подсчета
    results = parser.analyze_test_files(max_files=200)  # Увеличиваем лимит
    parser.print_analysis_report(results)
