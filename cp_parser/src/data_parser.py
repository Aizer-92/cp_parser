#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Полноценный парсер коммерческих предложений с извлечением изображений
Парсит товары, ценовые предложения и извлекает изображения с позициями
"""

import sys
import os
from pathlib import Path
import pandas as pd
import logging
from typing import Dict, List, Tuple, Optional, Any, Union
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import re
from datetime import datetime
import uuid
from PIL import Image
import io

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

from database.manager import db_manager, ProjectService, ProductService, PriceOfferService, ImageService
from src.structure_parser import CommercialProposalParser

class CommercialProposalDataParser:
    """Парсер данных коммерческих предложений с извлечением изображений"""
    
    def __init__(self):
        self.setup_logging()
        self.structure_parser = CommercialProposalParser()
        
        # Сервисы для работы с БД
        self.project_service = ProjectService(db_manager)
        self.product_service = ProductService(db_manager)
        self.price_offer_service = PriceOfferService(db_manager)
        self.image_service = ImageService(db_manager)
        
        # Создаем папку для изображений
        self.images_dir = project_root / "storage" / "images"
        self.images_dir.mkdir(parents=True, exist_ok=True)
        
        # Маппинг маршрутов
        self.route_mapping = {
            'F': 'ЖД',  # Доставка ЖД
            'I': 'АВИА',  # Доставка АВИА  
            'L': 'ОБРАЗЕЦ'  # Образец
        }
    
    def setup_logging(self):
        """Настройка логирования"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
    
    def extract_images_from_workbook(self, workbook_path: Path, table_id: str) -> List[Dict[str, Any]]:
        """Извлечение изображений из Excel файла с позициями"""
        images_data = []
        
        try:
            # Открываем workbook для извлечения изображений
            wb = load_workbook(workbook_path, data_only=False)
            
            # Ищем нужный лист
            sheet_name = self.structure_parser.find_matching_sheet(wb)
            if not sheet_name:
                self.logger.warning(f"Не найден подходящий лист в {workbook_path.name}")
                return images_data
            
            ws = wb[sheet_name]
            
            # Извлекаем изображения
            if hasattr(ws, '_images') and ws._images:
                for img in ws._images:
                    try:
                        # Получаем позицию изображения
                        anchor = img.anchor
                        
                        # Определяем ячейку
                        if hasattr(anchor, '_from'):
                            col_idx = anchor._from.col
                            row_idx = anchor._from.row
                        else:
                            # Fallback для других типов anchor
                            col_idx = 0
                            row_idx = 0
                        
                        # Конвертируем в буквенно-цифровое обозначение
                        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                        cell_position = f"{col_letter}{row_idx + 1}"
                        
                        # Проверяем, не существует ли уже изображение с такой позицией
                        existing_image = self.image_service.get_image_by_table_id_and_position(table_id, cell_position)
                        
                        if existing_image:
                            self.logger.info(f"Изображение уже существует: {cell_position} -> {existing_image.image_filename}")
                            continue
                        
                        # Генерируем уникальное имя файла
                        image_filename = f"{table_id}_{cell_position}_{uuid.uuid4().hex[:8]}.png"
                        image_path = self.images_dir / image_filename
                        
                        # Извлекаем и сохраняем изображение
                        if hasattr(img, '_data'):
                            # Получаем данные изображения
                            img_data = img._data()
                            
                            # Сохраняем как PIL Image
                            pil_image = Image.open(io.BytesIO(img_data))
                            pil_image.save(image_path, 'PNG')
                            
                            # Получаем размеры
                            width, height = pil_image.size
                            file_size_kb = len(img_data) / 1024
                            
                            # Определяем тип изображения (главное или дополнительное)
                            is_main = col_letter == 'A'  # Столбец A - главные изображения
                            
                            images_data.append({
                                'table_id': table_id,
                                'local_path': str(image_path),
                                'image_filename': image_filename,
                                'sheet_name': sheet_name,
                                'cell_position': cell_position,
                                'row_number': row_idx + 1,
                                'column_number': col_idx + 1,
                                'width_px': width,
                                'height_px': height,
                                'file_size_kb': round(file_size_kb, 2),
                                'format': 'PNG',
                                'is_main_image': is_main,
                                'extraction_method': 'openpyxl_blob',
                                'processing_status': 'extracted'
                            })
                            
                            self.logger.info(f"Извлечено изображение: {cell_position} -> {image_filename}")
                        
                    except Exception as e:
                        self.logger.error(f"Ошибка извлечения изображения: {e}")
                        continue
            
            wb.close()
            
        except Exception as e:
            self.logger.error(f"Ошибка извлечения изображений из {workbook_path}: {e}")
        
        return images_data
    
    def parse_delivery_time(self, time_str: str) -> Optional[int]:
        """Парсинг времени доставки из строки типа '20-25' -> 25"""
        if not time_str or not isinstance(time_str, str):
            return None
        
        # Ищем числа в строке
        numbers = re.findall(r'\d+', str(time_str))
        if numbers:
            # Берем последнее число (как указано в требованиях)
            return int(numbers[-1])
        
        return None
    
    def clean_quantity(self, quantity_str: str) -> Optional[int]:
        """Очистка тиража от пробелов и парсинг"""
        if not quantity_str:
            return None
        
        # Убираем все пробелы и нечисловые символы кроме цифр
        cleaned = re.sub(r'[^\d]', '', str(quantity_str))
        
        try:
            return int(cleaned) if cleaned else None
        except (ValueError, TypeError):
            return None
    
    def clean_price(self, price_str: str) -> Optional[float]:
        """Очистка и парсинг цены"""
        if not price_str:
            return None
        
        # Убираем пробелы, запятые заменяем на точки
        cleaned = str(price_str).replace(' ', '').replace(',', '.')
        
        # Извлекаем числа с точкой
        match = re.search(r'(\d+\.?\d*)', cleaned)
        if match:
            try:
                return float(match.group(1))
            except (ValueError, TypeError):
                return None
        
        return None
    
    def get_products_by_table_id(self, table_id: str) -> List[Dict[str, Any]]:
        """Получение товаров по table_id"""
        products = []
        try:
            with db_manager.get_session() as session:
                from database.models import Product
                db_products = session.query(Product).filter(Product.table_id == table_id).all()
                
                for product in db_products:
                    products.append({
                        'id': product.id,
                        'table_id': product.table_id,
                        'row_number': product.row_number,
                        'row_number_end': product.row_number_end,
                        'name': product.name
                    })
        except Exception as e:
            self.logger.error(f"Ошибка получения товаров по table_id {table_id}: {e}")
        
        return products
    
    def find_product_for_image(self, img_data: Dict[str, Any], products: List[Dict[str, Any]]) -> Optional[int]:
        """Поиск товара для изображения по table_id + row_number"""
        img_row = img_data.get('row_number')
        img_table_id = img_data.get('table_id')
        
        if not img_row or not img_table_id:
            return None
        
        # Ищем товар, в диапазон строк которого попадает изображение
        for product in products:
            if product['table_id'] == img_table_id:
                row_start = product['row_number']
                row_end = product['row_number_end'] or product['row_number']
                
                # Проверяем, попадает ли строка изображения в диапазон товара
                if row_start <= img_row <= row_end:
                    self.logger.info(f"Привязываю изображение {img_data['cell_position']} к товару {product['name']} (строки {row_start}-{row_end})")
                    return product['id']
        
        self.logger.info(f"Изображение {img_data['cell_position']} (строка {img_row}) не привязано к товару")
        return None
    
    
    
    def find_product_name_column(self, worksheet, start_row=4, max_rows=6):
        """
        УЛУЧШЕННОЕ динамическое определение столбца с названиями товаров
        Приоритет: краткие названия > длинные описания
        """
        # Проверяем столбцы A, B, C, D для поиска данных товаров
        candidate_columns = ['A', 'B', 'C', 'D']
        column_scores = {}
        
        for col_letter in candidate_columns:
            col_num = ord(col_letter) - ord('A') + 1
            score = 0
            non_empty_count = 0
            total_length = 0
            description_indicators = 0
            
            for row_num in range(start_row, start_row + max_rows):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = str(cell.value or '').strip()
                
                if value and value not in ['None', 'nan']:
                    non_empty_count += 1
                    value_length = len(value)
                    total_length += value_length
                    
                    # НОВАЯ ЛОГИКА: Определяем тип содержимого
                    
                    # 1. Проверяем на описания (Material:, Size:, Package: и т.д.)
                    description_keywords = ['material:', 'размер:', 'size:', 'package:', 'упаковка:', 
                                          'capacity:', 'объем:', 'color:', 'цвет:', 'design:']
                    
                    if any(keyword in value.lower() for keyword in description_keywords):
                        description_indicators += 1
                        # Описания получают МЕНЬШЕ баллов
                        score += 1
                    
                    # 2. Проверяем на короткие названия товаров (приоритет)
                    elif 5 <= value_length <= 30:
                        # Короткие названия - это то, что нам нужно!
                        score += 5
                    
                    # 3. Средние тексты
                    elif 3 <= value_length <= 50:
                        score += 3
                    
                    # 4. Очень длинные тексты (скорее всего описания)
                    elif value_length > 100:
                        score += 1  # Минимальный балл
                    
                    # 5. Очень короткие тексты
                    else:
                        score += 2
            
            # ШТРАФ за столбцы с большим количеством описаний
            if non_empty_count > 0:
                description_ratio = description_indicators / non_empty_count
                if description_ratio > 0.5:  # Больше 50% описаний
                    score = max(1, score // 2)  # Уменьшаем балл вдвое
                
                # БОНУС за оптимальную длину (названия товаров обычно 10-25 символов)
                if non_empty_count > 0:
                    avg_length = total_length / non_empty_count
                    if 10 <= avg_length <= 25:
                        score += 3  # Бонус за оптимальную длину названий
            
            column_scores[col_letter] = score
            
            # Логирование для отладки
            self.logger.info(f"🔍 Столбец {col_letter}: {non_empty_count} ячеек, "
                           f"ср.длина: {total_length/max(1,non_empty_count):.1f}, "
                           f"описаний: {description_indicators}, счет: {score}")
        
        # Выбираем столбец с максимальным счетом
        if not column_scores or max(column_scores.values()) == 0:
            return 'B'  # Возвращаемся к стандартному столбцу B
        
        best_column = max(column_scores.items(), key=lambda x: x[1])
        
        self.logger.info(f"🎯 Выбран столбец {best_column[0]} (счет: {best_column[1]})")
        
        return best_column[0]
    
    def parse_products_and_offers(self, workbook_path: Path, project_id: int, table_id: str) -> Dict[str, Any]:
        """Парсинг товаров и ценовых предложений с динамическим определением столбцов"""
        result = {
            'products_created': 0,
            'price_offers_created': 0,
            'errors': []
        }
        
        try:
            # Открываем файл
            wb = load_workbook(workbook_path, data_only=True)
            
            # Ищем подходящий лист (БЕЗ ИЗМЕНЕНИЙ)
            sheet_name = self.structure_parser.find_matching_sheet(wb)
            if not sheet_name:
                result['errors'].append(f"Не найден подходящий лист")
                return result
            
            ws = wb[sheet_name]
            
            # Читаем маршруты из строки 2 (БЕЗ ИЗМЕНЕНИЙ)
            routes = {}
            for col_letter, route_name in self.route_mapping.items():
                col_num = ord(col_letter) - ord('A') + 1
                cell_value = ws.cell(2, col_num).value
                if cell_value:
                    routes[col_letter] = str(cell_value)
                else:
                    routes[col_letter] = route_name
            
            # НОВОЕ: Динамически определяем столбец с названиями товаров
            product_name_column = self.find_product_name_column(ws)
            product_name_col_num = ord(product_name_column) - ord('A') + 1
            
            self.logger.info(f"📦 Используем столбец {product_name_column} для названий товаров")
            
            # Парсим строки начиная с 4-й (УЛУЧШЕННАЯ ЛОГИКА)
            current_product = None
            row_num = 4
            
            while row_num <= ws.max_row:
                try:
                    # Читаем основные поля (ИСПОЛЬЗУЕМ ДИНАМИЧЕСКИЙ СТОЛБЕЦ)
                    name_cell = ws.cell(row_num, product_name_col_num)  # Динамический столбец
                    characteristics_cell = ws.cell(row_num, 3)  # C - Характеристики (фиксированный)
                    custom_cell = ws.cell(row_num, 4)  # D - Кастом (фиксированный)
                    quantity_cell = ws.cell(row_num, 5)  # E - Тираж (фиксированный)
                    
                    name = str(name_cell.value).strip() if name_cell.value else ""
                    characteristics = str(characteristics_cell.value).strip() if characteristics_cell.value else ""
                    custom_field = str(custom_cell.value).strip() if custom_cell.value else ""
                    quantity_str = str(quantity_cell.value).strip() if quantity_cell.value else ""
                    
                    quantity = self.clean_quantity(quantity_str)
                    
                    # УЛУЧШЕННАЯ проверка логики: есть ли наименование
                    if name and len(name) > 2 and name.lower() not in ['', 'none', 'null']:
                        # Проверяем, не существует ли уже товар с таким table_id и row_number
                        existing_product = self.product_service.get_product_by_table_id_and_row(table_id, row_num)
                        
                        if existing_product:
                            self.logger.info(f"Товар уже существует: {existing_product.name} (ID: {existing_product.id}, строка: {existing_product.row_number})")
                            current_product = existing_product
                        else:
                            # Создаем новый товар
                            product_data = {
                                'project_id': project_id,
                                'table_id': table_id,
                                'name': name,
                                'description': characteristics if characteristics else None,
                                'custom_field': custom_field if custom_field else None,
                                'sheet_name': sheet_name,
                                'row_number': row_num,
                                'row_number_end': row_num  # Пока одна строка
                            }
                            
                            # Парсим образец (столбцы L и N) - БЕЗ ИЗМЕНЕНИЙ
                            sample_price_cell = ws.cell(row_num, 12)  # L - Цена образца
                            sample_time_cell = ws.cell(row_num, 14)   # N - Срок образца
                            
                            if sample_price_cell.value:
                                product_data['sample_price'] = self.clean_price(str(sample_price_cell.value))
                            
                            if sample_time_cell.value:
                                product_data['sample_delivery_time'] = self.parse_delivery_time(str(sample_time_cell.value))
                            
                            # Создаем товар в БД
                            current_product = self.product_service.create_product(**product_data)
                            result['products_created'] += 1
                            
                            self.logger.info(f"Создан товар: {name} (ID: {current_product.id}, строка: {row_num})")
                    
                    # Создаем ценовые предложения, если есть тираж (БЕЗ ИЗМЕНЕНИЙ)
                    if quantity and current_product:
                        # Парсим ценовые предложения для каждого маршрута
                        for col_letter, route_name in routes.items():
                            if col_letter in ['F', 'I']:  # ЖД и АВИА маршруты
                                price_usd_cell = ws.cell(row_num, ord(col_letter) - ord('A') + 1)
                                price_rub_cell = ws.cell(row_num, ord(col_letter) - ord('A') + 2)  # Следующий столбец
                                time_cell = ws.cell(row_num, ord(col_letter) - ord('A') + 3)      # Через столбец
                                
                                price_usd = self.clean_price(str(price_usd_cell.value)) if price_usd_cell.value else None
                                price_rub = self.clean_price(str(price_rub_cell.value)) if price_rub_cell.value else None
                                delivery_time = self.parse_delivery_time(str(time_cell.value)) if time_cell.value else None
                                
                                if price_usd or price_rub:  # Создаем предложение только если есть цена
                                    # Проверяем, не существует ли уже предложение
                                    existing_offer = self.price_offer_service.get_offer_by_product_and_route(
                                        current_product.id, route_name, quantity
                                    )
                                    
                                    if not existing_offer:
                                        offer_data = {
                                            'product_id': current_product.id,
                                            'table_id': table_id,
                                            'route': route_name,
                                            'quantity': quantity,
                                            'price_usd': price_usd,
                                            'price_rub': price_rub,
                                            'delivery_time_days': delivery_time,
                                            'row_position': str(row_num),
                                        }
                                        
                                        self.price_offer_service.create_offer(**offer_data)
                                        result['price_offers_created'] += 1
                                        
                                        self.logger.info(f"Создано предложение: {route_name}, {quantity} шт, ${price_usd}")
                
                except Exception as e:
                    result['errors'].append(f"Ошибка в строке {row_num}: {str(e)}")
                    self.logger.error(f"Ошибка парсинга строки {row_num}: {e}")
                
                row_num += 1
            
            wb.close()
            
        except Exception as e:
            result['errors'].append(f"Общая ошибка парсинга: {str(e)}")
            self.logger.error(f"Ошибка парсинга файла {workbook_path}: {e}")
        
        return result
    
    def parse_single_file(self, project_id: int) -> Dict[str, Any]:
        """Парсинг одного файла по ID проекта"""
        
        # Получаем проект из БД
        project = self.project_service.get_project(project_id)
        if not project:
            return {'error': f'Проект с ID {project_id} не найден'}
        
        if not project.file_path or not Path(project.file_path).exists():
            return {'error': f'Файл не найден: {project.file_path}'}
        
        file_path = Path(project.file_path)
        table_id = project.table_id
        
        self.logger.info(f"🚀 Начинаю парсинг файла: {file_path.name} (table_id: {table_id})")
        
        result = {
            'project_id': project_id,
            'table_id': table_id,
            'file_name': file_path.name,
            'parsing_started_at': datetime.now(),
            'products_created': 0,
            'price_offers_created': 0,
            'images_extracted': 0,
            'errors': []
        }
        
        try:
            # 1. Проверяем структуру файла
            analysis = self.structure_parser.analyze_file(file_path)
            if not analysis['is_parseable']:
                result['errors'].append(f"Файл не подходит для парсинга: {analysis.get('error', 'Неизвестная ошибка')}")
                return result
            
            # 2. Парсим товары и ценовые предложения СНАЧАЛА
            self.logger.info(f"📦 Парсю товары и ценовые предложения...")
            parsing_result = self.parse_products_and_offers(file_path, project_id, table_id)
            
            result['products_created'] = parsing_result['products_created']
            result['price_offers_created'] = parsing_result['price_offers_created']
            result['errors'].extend(parsing_result['errors'])
            
            # 3. Извлекаем изображения и привязываем к товарам
            self.logger.info(f"📷 Извлекаю изображения и привязываю к товарам...")
            images_data = self.extract_images_from_workbook(file_path, table_id)
            
            # Привязываем изображения к товарам по позициям
            created_products = self.get_products_by_table_id(table_id)
            for img_data in images_data:
                # Ищем товар по позиции изображения
                product_id = self.find_product_for_image(img_data, created_products)
                if product_id:
                    img_data['product_id'] = product_id
                    self.image_service.create_image(**img_data)
                    result['images_extracted'] += 1
                else:
                    # Сохраняем изображение без привязки к товару (например, логотипы)
                    self.image_service.create_image(**img_data)
                    result['images_extracted'] += 1
            
            # 4. Обновляем статус проекта
            project.parsing_status = 'completed' if not result['errors'] else 'completed_with_errors'
            project.parsed_at = datetime.now()
            
            with db_manager.get_session() as session:
                session.merge(project)
                session.commit()
            
            result['parsing_completed_at'] = datetime.now()
            result['success'] = True
            
            self.logger.info(f"✅ Парсинг завершен: товаров {result['products_created']}, предложений {result['price_offers_created']}, изображений {result['images_extracted']}")
            
        except Exception as e:
            result['errors'].append(f"Критическая ошибка: {str(e)}")
            result['success'] = False
            self.logger.error(f"❌ Критическая ошибка парсинга: {e}")
        
        return result
    
    def parse_multiple_files(self, max_files: int = 5) -> Dict[str, Any]:
        """Парсинг нескольких файлов для тестирования"""
        
        # Получаем проекты с подходящими файлами
        results = self.structure_parser.analyze_test_files(max_files=200)
        parseable_files = results['parseable_files'][:max_files]
        
        if not parseable_files:
            return {'error': 'Нет файлов подходящих для парсинга'}
        
        batch_result = {
            'total_files': len(parseable_files),
            'successful_parses': 0,
            'failed_parses': 0,
            'total_products': 0,
            'total_offers': 0,
            'total_images': 0,
            'file_results': []
        }
        
        self.logger.info(f"🚀 Начинаю пакетный парсинг {len(parseable_files)} файлов...")
        
        for file_info in parseable_files:
            file_name = file_info['file_name']
            
            # Получаем project_id по имени файла
            project = self.project_service.get_project_by_file_name(file_name)
            if not project:
                self.logger.error(f"❌ Проект для файла {file_name} не найден в БД")
                batch_result['failed_parses'] += 1
                batch_result['file_results'].append({
                    'file_name': file_name,
                    'error': 'Проект не найден в БД',
                    'success': False
                })
                continue
            
            project_id = project.id
            
            self.logger.info(f"\n📁 Обрабатываю файл {batch_result['successful_parses'] + batch_result['failed_parses'] + 1}/{len(parseable_files)}: {file_name}")
            
            parse_result = self.parse_single_file(project_id)
            
            if parse_result.get('success'):
                batch_result['successful_parses'] += 1
                batch_result['total_products'] += parse_result['products_created']
                batch_result['total_offers'] += parse_result['price_offers_created']
                batch_result['total_images'] += parse_result['images_extracted']
            else:
                batch_result['failed_parses'] += 1
            
            batch_result['file_results'].append(parse_result)
        
        self.logger.info(f"\n🎉 ПАКЕТНЫЙ ПАРСИНГ ЗАВЕРШЕН!")
        self.logger.info(f"✅ Успешно: {batch_result['successful_parses']}")
        self.logger.info(f"❌ Ошибки: {batch_result['failed_parses']}")
        self.logger.info(f"📦 Всего товаров: {batch_result['total_products']}")
        self.logger.info(f"💰 Всего предложений: {batch_result['total_offers']}")
        self.logger.info(f"📷 Всего изображений: {batch_result['total_images']}")
        
        return batch_result


if __name__ == "__main__":
    parser = CommercialProposalDataParser()
    
    # Запускаем полный парсинг всех подходящих файлов с нуля
    print("🚀 Запускаю полный парсинг всех подходящих файлов с чистой базой...")
    results = parser.parse_multiple_files(max_files=30)
    
    print(f"\n📊 ФИНАЛЬНЫЕ РЕЗУЛЬТАТЫ ПОЛНОГО ПАРСИНГА:")
    print(f"=" * 60)
    print(f"✅ Успешно обработано: {results.get('successful_parses', 0)} файлов")
    print(f"❌ Ошибок: {results.get('failed_parses', 0)} файлов") 
    print(f"📦 Всего товаров создано: {results.get('total_products', 0)}")
    print(f"💰 Всего предложений создано: {results.get('total_offers', 0)}")
    print(f"📷 Всего изображений извлечено: {results.get('total_images', 0)}")
    
    if results.get('file_results'):
        print(f"\n📋 ДЕТАЛИ ПО ФАЙЛАМ:")
        for i, file_result in enumerate(results['file_results'], 1):
            status = "✅" if file_result.get('success') else "❌"
            file_name = file_result.get('file_name', 'Неизвестный файл')
            products = file_result.get('products_created', 0)
            offers = file_result.get('price_offers_created', 0)
            images = file_result.get('images_extracted', 0)
            
            print(f"{i:2d}. {status} {file_name}")
            if file_result.get('success'):
                print(f"     📦 {products} товаров, 💰 {offers} предложений, 📷 {images} изображений")
            else:
                error = file_result.get('error', 'Неизвестная ошибка')
                print(f"     ❌ {error}")
                
    print(f"\n🎯 Полный парсинг завершен! Откройте http://localhost:5000 для просмотра результатов")
