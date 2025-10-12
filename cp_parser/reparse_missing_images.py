#!/usr/bin/env python3
"""
Допарсинг изображений для проектов без изображений
Извлекает только изображения из уже скачанных Excel файлов
"""

import sys
from pathlib import Path
from sqlalchemy import text
import openpyxl
from datetime import datetime

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager

def extract_images_from_excel(excel_path, table_id):
    """Извлекает изображения из Excel файла"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        images = []
        for img in ws._images:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                row = anchor._from.row + 1  # openpyxl начинает с 0
                col = anchor._from.col + 1
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_position = f"{col_letter}{row}"
                
                # Генерируем имя файла
                img_filename = f"{table_id}_{cell_position}_{abs(hash(img.ref))}.png"
                
                # Получаем данные изображения
                img_data = img.ref
                if hasattr(img_data, 'read'):
                    # Если это файловый объект
                    img_data.seek(0)
                    img_bytes = img_data.read()
                elif isinstance(img_data, bytes):
                    img_bytes = img_data
                else:
                    # Пропускаем если непонятный формат
                    continue
                
                images.append({
                    'filename': img_filename,
                    'data': img_bytes,
                    'cell_position': cell_position,
                    'row': row,
                    'col': col
                })
        
        wb.close()
        return images
        
    except Exception as e:
        print(f"    ❌ Ошибка извлечения: {e}")
        return []

def save_images_to_db(project_id, table_id, images, storage_dir):
    """Сохраняет изображения в БД"""
    db = PostgreSQLManager()
    
    with db.get_session() as session:
        # Получаем товары проекта с их row_number
        products = session.execute(text("""
            SELECT id, row_number
            FROM products
            WHERE project_id = :project_id
            ORDER BY row_number
        """), {'project_id': project_id}).fetchall()
        
        if not products:
            return 0
        
        saved_count = 0
        
        for img in images:
            # Находим товар для этого изображения по row_number
            product_id = None
            for prod_id, row_num in products:
                if row_num == img['row']:
                    product_id = prod_id
                    break
            
            if not product_id:
                # Если точное совпадение не найдено, берем ближайший товар
                for prod_id, row_num in products:
                    if row_num <= img['row']:
                        product_id = prod_id
                    else:
                        break
            
            if not product_id:
                continue
            
            # Сохраняем файл
            img_path = storage_dir / img['filename']
            with open(img_path, 'wb') as f:
                f.write(img['data'])
            
            # Проверяем не существует ли уже
            exists = session.execute(text("""
                SELECT COUNT(*) FROM product_images
                WHERE product_id = :product_id
                AND cell_position = :cell_position
            """), {
                'product_id': product_id,
                'cell_position': img['cell_position']
            }).scalar()
            
            if exists > 0:
                continue
            
            # Определяем is_main_image (первая колонка A)
            is_main = img['col'] == 1
            
            # Записываем в БД
            session.execute(text("""
                INSERT INTO product_images (
                    product_id, table_id, image_filename, local_path,
                    cell_position, is_main_image, row_number,
                    created_at, updated_at
                )
                VALUES (
                    :product_id, :table_id, :filename, :local_path,
                    :cell_position, :is_main, :row_number,
                    NOW(), NOW()
                )
            """), {
                'product_id': product_id,
                'table_id': table_id,
                'filename': img['filename'],
                'local_path': str(img_path),
                'cell_position': img['cell_position'],
                'is_main': is_main,
                'row_number': img['row']
            })
            
            saved_count += 1
        
        session.commit()
        return saved_count

def main():
    print("=" * 80)
    print("🖼️  ДОПАРСИНГ ИЗОБРАЖЕНИЙ ДЛЯ ПРОЕКТОВ БЕЗ ИЗОБРАЖЕНИЙ")
    print("=" * 80)
    
    # Читаем списки
    template4_missing = []
    with open('missing_images_template4.txt', 'r') as f:
        for line in f:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                template4_missing.append((int(parts[0]), parts[1]))
    
    template5_missing = []
    with open('missing_images_template5.txt', 'r') as f:
        for line in f:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                template5_missing.append((int(parts[0]), parts[1]))
    
    all_missing = template4_missing + template5_missing
    
    print(f"\n📊 К обработке:")
    print(f"  Шаблон 4: {len(template4_missing)} проектов")
    print(f"  Шаблон 5: {len(template5_missing)} проектов")
    print(f"  ВСЕГО: {len(all_missing)} проектов")
    
    # Директории
    excel_dir = Path('storage/excel_files')
    images_dir = Path('storage/images')
    images_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n🚀 Начинаю допарсинг изображений...")
    
    processed = 0
    skipped = 0
    errors = 0
    total_images = 0
    
    for idx, (proj_id, table_id) in enumerate(all_missing, 1):
        excel_path = excel_dir / f"{table_id}.xlsx"
        
        if not excel_path.exists():
            print(f"  [{idx}/{len(all_missing)}] ⏩ Проект {proj_id}: нет Excel файла")
            skipped += 1
            continue
        
        print(f"  [{idx}/{len(all_missing)}] 🖼️  Проект {proj_id}...", end='')
        
        # Извлекаем изображения
        images = extract_images_from_excel(excel_path, table_id)
        
        if not images:
            print(f" ⚠️  Нет изображений в файле")
            skipped += 1
            continue
        
        # Сохраняем в БД
        saved = save_images_to_db(proj_id, table_id, images, images_dir)
        
        print(f" ✅ {saved} изображений")
        processed += 1
        total_images += saved
    
    print("\n" + "=" * 80)
    print("✅ ДОПАРСИНГ ЗАВЕРШЕН")
    print("=" * 80)
    print(f"\n📊 Статистика:")
    print(f"  • Обработано:    {processed}")
    print(f"  • Пропущено:     {skipped}")
    print(f"  • Ошибок:        {errors}")
    print(f"  • Всего проектов: {len(all_missing)}")
    print(f"  • Всего изображений: {total_images}")
    print("=" * 80)

if __name__ == '__main__':
    main()


Допарсинг изображений для проектов без изображений
Извлекает только изображения из уже скачанных Excel файлов
"""

import sys
from pathlib import Path
from sqlalchemy import text
import openpyxl
from datetime import datetime

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager

def extract_images_from_excel(excel_path, table_id):
    """Извлекает изображения из Excel файла"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        images = []
        for img in ws._images:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                row = anchor._from.row + 1  # openpyxl начинает с 0
                col = anchor._from.col + 1
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_position = f"{col_letter}{row}"
                
                # Генерируем имя файла
                img_filename = f"{table_id}_{cell_position}_{abs(hash(img.ref))}.png"
                
                # Получаем данные изображения
                img_data = img.ref
                if hasattr(img_data, 'read'):
                    # Если это файловый объект
                    img_data.seek(0)
                    img_bytes = img_data.read()
                elif isinstance(img_data, bytes):
                    img_bytes = img_data
                else:
                    # Пропускаем если непонятный формат
                    continue
                
                images.append({
                    'filename': img_filename,
                    'data': img_bytes,
                    'cell_position': cell_position,
                    'row': row,
                    'col': col
                })
        
        wb.close()
        return images
        
    except Exception as e:
        print(f"    ❌ Ошибка извлечения: {e}")
        return []

def save_images_to_db(project_id, table_id, images, storage_dir):
    """Сохраняет изображения в БД"""
    db = PostgreSQLManager()
    
    with db.get_session() as session:
        # Получаем товары проекта с их row_number
        products = session.execute(text("""
            SELECT id, row_number
            FROM products
            WHERE project_id = :project_id
            ORDER BY row_number
        """), {'project_id': project_id}).fetchall()
        
        if not products:
            return 0
        
        saved_count = 0
        
        for img in images:
            # Находим товар для этого изображения по row_number
            product_id = None
            for prod_id, row_num in products:
                if row_num == img['row']:
                    product_id = prod_id
                    break
            
            if not product_id:
                # Если точное совпадение не найдено, берем ближайший товар
                for prod_id, row_num in products:
                    if row_num <= img['row']:
                        product_id = prod_id
                    else:
                        break
            
            if not product_id:
                continue
            
            # Сохраняем файл
            img_path = storage_dir / img['filename']
            with open(img_path, 'wb') as f:
                f.write(img['data'])
            
            # Проверяем не существует ли уже
            exists = session.execute(text("""
                SELECT COUNT(*) FROM product_images
                WHERE product_id = :product_id
                AND cell_position = :cell_position
            """), {
                'product_id': product_id,
                'cell_position': img['cell_position']
            }).scalar()
            
            if exists > 0:
                continue
            
            # Определяем is_main_image (первая колонка A)
            is_main = img['col'] == 1
            
            # Записываем в БД
            session.execute(text("""
                INSERT INTO product_images (
                    product_id, table_id, image_filename, local_path,
                    cell_position, is_main_image, row_number,
                    created_at, updated_at
                )
                VALUES (
                    :product_id, :table_id, :filename, :local_path,
                    :cell_position, :is_main, :row_number,
                    NOW(), NOW()
                )
            """), {
                'product_id': product_id,
                'table_id': table_id,
                'filename': img['filename'],
                'local_path': str(img_path),
                'cell_position': img['cell_position'],
                'is_main': is_main,
                'row_number': img['row']
            })
            
            saved_count += 1
        
        session.commit()
        return saved_count

def main():
    print("=" * 80)
    print("🖼️  ДОПАРСИНГ ИЗОБРАЖЕНИЙ ДЛЯ ПРОЕКТОВ БЕЗ ИЗОБРАЖЕНИЙ")
    print("=" * 80)
    
    # Читаем списки
    template4_missing = []
    with open('missing_images_template4.txt', 'r') as f:
        for line in f:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                template4_missing.append((int(parts[0]), parts[1]))
    
    template5_missing = []
    with open('missing_images_template5.txt', 'r') as f:
        for line in f:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                template5_missing.append((int(parts[0]), parts[1]))
    
    all_missing = template4_missing + template5_missing
    
    print(f"\n📊 К обработке:")
    print(f"  Шаблон 4: {len(template4_missing)} проектов")
    print(f"  Шаблон 5: {len(template5_missing)} проектов")
    print(f"  ВСЕГО: {len(all_missing)} проектов")
    
    # Директории
    excel_dir = Path('storage/excel_files')
    images_dir = Path('storage/images')
    images_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n🚀 Начинаю допарсинг изображений...")
    
    processed = 0
    skipped = 0
    errors = 0
    total_images = 0
    
    for idx, (proj_id, table_id) in enumerate(all_missing, 1):
        excel_path = excel_dir / f"{table_id}.xlsx"
        
        if not excel_path.exists():
            print(f"  [{idx}/{len(all_missing)}] ⏩ Проект {proj_id}: нет Excel файла")
            skipped += 1
            continue
        
        print(f"  [{idx}/{len(all_missing)}] 🖼️  Проект {proj_id}...", end='')
        
        # Извлекаем изображения
        images = extract_images_from_excel(excel_path, table_id)
        
        if not images:
            print(f" ⚠️  Нет изображений в файле")
            skipped += 1
            continue
        
        # Сохраняем в БД
        saved = save_images_to_db(proj_id, table_id, images, images_dir)
        
        print(f" ✅ {saved} изображений")
        processed += 1
        total_images += saved
    
    print("\n" + "=" * 80)
    print("✅ ДОПАРСИНГ ЗАВЕРШЕН")
    print("=" * 80)
    print(f"\n📊 Статистика:")
    print(f"  • Обработано:    {processed}")
    print(f"  • Пропущено:     {skipped}")
    print(f"  • Ошибок:        {errors}")
    print(f"  • Всего проектов: {len(all_missing)}")
    print(f"  • Всего изображений: {total_images}")
    print("=" * 80)

if __name__ == '__main__':
    main()

