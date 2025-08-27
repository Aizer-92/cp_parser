#!/usr/bin/env python3
"""
Smart Plan/Fact Analyzer for Main mega ТОП
Умный анализатор с правильным разбором структуры сообщений
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

class SmartPlanFactAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/smart_plan_fact.log", rotation="1 day", retention="7 days")
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_plan_fact_messages(self, weeks_back=12):
        """Получение сообщений с планами и фактами"""
        try:
            print(f"📥 Загрузка сообщений с планами/фактами за последние {weeks_back} недель...")
            
            today = datetime.now()
            start_date = today - timedelta(weeks=weeks_back)
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=5000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            # Фильтруем сообщения с планами и фактами
            plan_fact_messages = []
            
            for msg in messages:
                if msg.text and len(msg.text.strip()) > 10:
                    text = msg.text.lower()
                    
                    # Проверяем наличие планов и фактов
                    has_plan = re.search(r'план[:\s]*(\d+(?:\.\d+)?)', text)
                    has_fact = re.search(r'факт[:\s]*(\d+(?:\.\d+)?)', text)
                    has_pf = re.search(r'п/ф[:\s]*(\d+(?:\.\d+)?)', text)
                    has_numbers = re.search(r'(\d+(?:\.\d+)?)[/\-]\s*(\d+(?:\.\d+)?)', text)
                    has_brackets = re.search(r'\((\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)\)', text)
                    
                    if has_plan or has_fact or has_pf or has_numbers or has_brackets:
                        sender_name = await self.get_sender_name(msg)
                        plan_fact_messages.append({
                            'id': msg.id,
                            'date': msg.date,
                            'text': msg.text.strip(),
                            'sender': sender_name
                        })
            
            print(f"📊 Найдено {len(plan_fact_messages)} сообщений с планами/фактами")
            return plan_fact_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return []
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def parse_message_structure(self, text, message_date, sender_name, message_id):
        """Умный разбор структуры сообщения"""
        try:
            lines = text.split('\n')
            tasks = []
            
            # Ищем разделы "План:" и "Факт:"
            plan_section = []
            fact_section = []
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Определяем раздел
                if re.match(r'^план[:\s]*', line.lower()):
                    current_section = 'plan'
                    continue
                elif re.match(r'^факт[:\s]*', line.lower()):
                    current_section = 'fact'
                    continue
                
                # Добавляем в соответствующий раздел
                if current_section == 'plan':
                    plan_section.append(line)
                elif current_section == 'fact':
                    fact_section.append(line)
            
            # Разбираем задачи из разделов
            plan_tasks = self.extract_tasks_from_section(plan_section, 'plan')
            fact_tasks = self.extract_tasks_from_section(fact_section, 'fact')
            
            # Сопоставляем планы и факты
            tasks = self.match_plan_fact_tasks(plan_tasks, fact_tasks, message_date, sender_name, message_id)
            
            return tasks
            
        except Exception as e:
            logger.error(f"❌ Ошибка разбора структуры: {e}")
            return []
    
    def extract_tasks_from_section(self, section_lines, section_type):
        """Извлечение задач из раздела"""
        tasks = []
        
        for line in section_lines:
            # Ищем нумерованные задачи
            numbered_match = re.match(r'^(\d+)[\.\)]\s*(.+)', line)
            if numbered_match:
                task_num = numbered_match.group(1)
                task_desc = numbered_match.group(2)
                
                # Ищем числа в задаче
                numbers = self.extract_numbers_from_text(task_desc)
                
                tasks.append({
                    'number': task_num,
                    'description': task_desc,
                    'numbers': numbers,
                    'section': section_type
                })
                continue
            
            # Ищем задачи с маркерами
            marker_match = re.match(r'^[-•]\s*(.+)', line)
            if marker_match:
                task_desc = marker_match.group(1)
                numbers = self.extract_numbers_from_text(task_desc)
                
                tasks.append({
                    'number': len(tasks) + 1,
                    'description': task_desc,
                    'numbers': numbers,
                    'section': section_type
                })
                continue
            
            # Ищем строки с числами
            numbers = self.extract_numbers_from_text(line)
            if numbers:
                tasks.append({
                    'number': len(tasks) + 1,
                    'description': line,
                    'numbers': numbers,
                    'section': section_type
                })
        
        return tasks
    
    def extract_numbers_from_text(self, text):
        """Извлечение чисел из текста"""
        numbers = []
        
        # Паттерны для поиска чисел
        patterns = [
            r'(\d+(?:\.\d+)?)[/\-]\s*(\d+(?:\.\d+)?)',  # X/Y или X-Y
            r'\((\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)\)',  # (X Y)
            r'план[:\s]*(\d+(?:\.\d+)?)',  # план: X
            r'факт[:\s]*(\d+(?:\.\d+)?)',  # факт: X
            r'п/ф[:\s]*(\d+(?:\.\d+)?)',   # п/ф: X
        ]
        
        for pattern in patterns:
            matches = re.finditer(pattern, text.lower())
            for match in matches:
                try:
                    if len(match.groups()) == 2:
                        num1 = float(match.group(1))
                        num2 = float(match.group(2))
                        numbers.append((num1, num2))
                    else:
                        num = float(match.group(1))
                        numbers.append((num, None))
                except ValueError:
                    continue
        
        return numbers
    
    def match_plan_fact_tasks(self, plan_tasks, fact_tasks, message_date, sender_name, message_id):
        """Сопоставление планов и фактов"""
        matched_tasks = []
        
        # Создаем словари для быстрого поиска
        plan_dict = {task['number']: task for task in plan_tasks}
        fact_dict = {task['number']: task for task in fact_tasks}
        
        # Сопоставляем по номерам
        all_numbers = set(plan_dict.keys()) | set(fact_dict.keys())
        
        for num in all_numbers:
            plan_task = plan_dict.get(num)
            fact_task = fact_dict.get(num)
            
            # Извлекаем планы и факты
            plan_value = None
            fact_value = None
            
            if plan_task and plan_task['numbers']:
                for nums in plan_task['numbers']:
                    if len(nums) == 2 and nums[1] is not None:
                        plan_value = nums[0]
                        break
                    elif len(nums) == 2 and nums[1] is None:
                        plan_value = nums[0]
                        break
            
            if fact_task and fact_task['numbers']:
                for nums in fact_task['numbers']:
                    if len(nums) == 2 and nums[1] is not None:
                        fact_value = nums[1]
                        break
                    elif len(nums) == 2 and nums[1] is None:
                        fact_value = nums[0]
                        break
            
            # Если нашли и план, и факт
            if plan_value is not None and fact_value is not None:
                completion_percent = (fact_value / plan_value * 100) if plan_value > 0 else 0
                
                matched_tasks.append({
                    'message_id': message_id,
                    'date': message_date,
                    'sender': sender_name,
                    'task_number': num,
                    'task_description': plan_task['description'] if plan_task else fact_task['description'],
                    'plan': plan_value,
                    'fact': fact_value,
                    'difference': fact_value - plan_value,
                    'completion_percent': completion_percent,
                    'status': self.get_status(completion_percent),
                    'is_duplicate': False
                })
        
        return matched_tasks
    
    def get_status(self, completion_percent):
        """Определение статуса выполнения"""
        if completion_percent >= 100:
            return "Перевыполнение"
        elif completion_percent >= 90:
            return "Хорошо"
        elif completion_percent >= 70:
            return "Удовлетворительно"
        else:
            return "Недовыполнение"
    
    def find_duplicates(self, all_tasks):
        """Поиск дублирующихся задач"""
        try:
            # Группируем по отправителю, описанию и значениям
            grouped_tasks = {}
            
            for task in all_tasks:
                key = (task['sender'], task['task_description'][:50], task['plan'], task['fact'])
                if key not in grouped_tasks:
                    grouped_tasks[key] = []
                grouped_tasks[key].append(task)
            
            # Помечаем дубликаты
            for key, tasks in grouped_tasks.items():
                if len(tasks) > 1:
                    for task in tasks:
                        task['is_duplicate'] = True
                        task['duplicate_count'] = len(tasks)
            
            return all_tasks
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска дубликатов: {e}")
            return all_tasks
    
    def create_smart_report(self, all_tasks, filename=None):
        """Создание умного отчета"""
        try:
            if not filename:
                filename = f"output/smart_plan_fact_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # Лист 1: Все задачи
                if all_tasks:
                    df_tasks = pd.DataFrame(all_tasks)
                    df_tasks['date'] = pd.to_datetime(df_tasks['date']).dt.tz_localize(None)
                    df_tasks = df_tasks.sort_values(['sender', 'date'], ascending=[True, False])
                    
                    # Переименовываем колонки
                    df_tasks = df_tasks.rename(columns={
                        'message_id': 'ID сообщения',
                        'date': 'Дата',
                        'sender': 'Руководитель',
                        'task_number': 'Номер задачи',
                        'task_description': 'Описание задачи',
                        'plan': 'План',
                        'fact': 'Факт',
                        'difference': 'Разница',
                        'completion_percent': '% выполнения',
                        'status': 'Статус',
                        'is_duplicate': 'Дубликат',
                        'duplicate_count': 'Количество дубликатов'
                    })
                    
                    df_tasks.to_excel(writer, sheet_name='Все задачи', index=False)
                    
                    # Форматирование
                    worksheet = writer.sheets['Все задачи']
                    
                    # Подсветка дубликатов и недовыполнений
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    
                    for row in worksheet.iter_rows(min_row=2):
                        if row[df_tasks.columns.get_loc('Дубликат')].value:
                            for cell in row:
                                cell.fill = red_fill
                        elif row[df_tasks.columns.get_loc('% выполнения')].value and row[df_tasks.columns.get_loc('% выполнения')].value < 70:
                            for cell in row:
                                cell.fill = yellow_fill
                
                # Лист 2: Анализ по руководителям
                if all_tasks:
                    sender_analysis = {}
                    
                    for task in all_tasks:
                        sender = task['sender']
                        if sender not in sender_analysis:
                            sender_analysis[sender] = {
                                'total_plan': 0,
                                'total_fact': 0,
                                'tasks_count': 0,
                                'over_performance': 0,
                                'under_performance': 0,
                                'duplicates': 0,
                                'dates': set(),
                                'tasks': []
                            }
                        
                        sender_analysis[sender]['total_plan'] += task['plan']
                        sender_analysis[sender]['total_fact'] += task['fact']
                        sender_analysis[sender]['tasks_count'] += 1
                        sender_analysis[sender]['dates'].add(task['date'].strftime('%Y-%m-%d'))
                        
                        if task['completion_percent'] >= 100:
                            sender_analysis[sender]['over_performance'] += 1
                        else:
                            sender_analysis[sender]['under_performance'] += 1
                        
                        if task['is_duplicate']:
                            sender_analysis[sender]['duplicates'] += 1
                        
                        sender_analysis[sender]['tasks'].append({
                            'date': task['date'].strftime('%Y-%m-%d'),
                            'task': task['task_description'],
                            'plan': task['plan'],
                            'fact': task['fact'],
                            'completion': task['completion_percent'],
                            'status': task['status']
                        })
                    
                    # Создаем данные для отчета
                    sender_data = []
                    for sender, data in sender_analysis.items():
                        avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                        sender_data.append({
                            'Руководитель': sender,
                            'Общий план': round(data['total_plan'], 2),
                            'Общий факт': round(data['total_fact'], 2),
                            'Средний % выполнения': round(avg_completion, 2),
                            'Количество задач': data['tasks_count'],
                            'Перевыполнений': data['over_performance'],
                            'Недовыполнений': data['under_performance'],
                            'Дубликатов': data['duplicates'],
                            'Количество дат': len(data['dates']),
                            'Даты активности': ', '.join(sorted(data['dates']))
                        })
                    
                    df_senders = pd.DataFrame(sender_data)
                    df_senders = df_senders.sort_values('Средний % выполнения', ascending=False)
                    df_senders.to_excel(writer, sheet_name='Анализ по руководителям', index=False)
                
                # Лист 3: Дубликаты
                duplicates_tasks = [task for task in all_tasks if task['is_duplicate']]
                if duplicates_tasks:
                    df_duplicates = pd.DataFrame(duplicates_tasks)
                    df_duplicates['date'] = pd.to_datetime(df_duplicates['date']).dt.tz_localize(None)
                    df_duplicates = df_duplicates.sort_values(['sender', 'date'], ascending=[True, False])
                    
                    df_duplicates = df_duplicates.rename(columns={
                        'date': 'Дата',
                        'sender': 'Руководитель',
                        'task_description': 'Описание задачи',
                        'plan': 'План',
                        'fact': 'Факт',
                        'completion_percent': '% выполнения',
                        'duplicate_count': 'Количество дубликатов'
                    })
                    
                    df_duplicates.to_excel(writer, sheet_name='Дубликаты', index=False)
            
            print(f"✅ Умный отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    async def run_smart_analysis(self):
        """Запуск умного анализа"""
        try:
            print("🚀 ЗАПУСК УМНОГО АНАЛИЗА ПЛАНОВ И ФАКТОВ")
            print("=" * 60)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение сообщений
            messages = await self.get_plan_fact_messages(weeks_back=12)
            
            if not messages:
                print("❌ Не найдено сообщений с планами/фактами")
                return
            
            # Разбор структуры сообщений
            all_tasks = []
            
            for msg in messages:
                tasks = self.parse_message_structure(msg['text'], msg['date'], msg['sender'], msg['id'])
                all_tasks.extend(tasks)
            
            print(f"📊 Извлечено {len(all_tasks)} задач из {len(messages)} сообщений")
            
            if not all_tasks:
                print("❌ Не удалось извлечь задачи")
                return
            
            # Поиск дубликатов
            all_tasks = self.find_duplicates(all_tasks)
            duplicates_count = len([task for task in all_tasks if task['is_duplicate']])
            print(f"🔍 Найдено {duplicates_count} дублирующихся задач")
            
            # Создание отчета
            report_file = self.create_smart_report(all_tasks)
            
            if report_file:
                print(f"\n🎉 Умный анализ завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Всего задач: {len(all_tasks)}")
                print(f"Уникальных руководителей: {len(set(task['sender'] for task in all_tasks))}")
                print(f"Дубликатов: {duplicates_count}")
                
                # Показываем статистику по руководителям
                sender_stats = {}
                for task in all_tasks:
                    sender = task['sender']
                    if sender not in sender_stats:
                        sender_stats[sender] = {'count': 0, 'total_completion': 0}
                    sender_stats[sender]['count'] += 1
                    sender_stats[sender]['total_completion'] += task['completion_percent']
                
                print(f"\n👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ:")
                for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['count'], reverse=True):
                    avg_completion = stats['total_completion'] / stats['count']
                    print(f"  {sender}: {stats['count']} задач, средний % выполнения: {avg_completion:.1f}%")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = SmartPlanFactAnalyzer()
    await analyzer.run_smart_analysis()

if __name__ == "__main__":
    asyncio.run(main())
