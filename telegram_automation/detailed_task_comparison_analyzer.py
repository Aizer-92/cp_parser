#!/usr/bin/env python3
"""
Detailed Task Comparison Analyzer for Main mega ТОП
Анализатор с детальным сопоставлением задач плана и факта
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class DetailedTaskComparisonAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/detailed_comparison.log", rotation="1 day", retention="7 days")
        
        # Цветовая схема
        self.colors = {
            'green': '28A745',      # Выполнено
            'yellow': 'FFC107',     # Частично выполнено
            'red': 'DC3545',        # Не выполнено
            'blue': '007BFF',       # Заголовки
            'orange': 'FD7E14',     # Новые задачи
            'light_gray': 'F8F9FA', # Фон
            'dark_gray': '6C757D'   # Текст
        }
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_weekly_messages(self, weeks_back=8):
        """Получение сообщений по неделям"""
        try:
            print(f"📥 Загрузка сообщений за последние {weeks_back} недель...")
            
            today = datetime.now()
            start_date = today - timedelta(weeks=weeks_back)
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=5000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            # Группируем сообщения по неделям
            weekly_messages = {}
            
            for msg in messages:
                if msg.text and len(msg.text.strip()) > 50:
                    # Определяем неделю
                    msg_date = msg.date.replace(tzinfo=None)
                    week_start = msg_date - timedelta(days=msg_date.weekday())
                    week_key = week_start.strftime('%Y-%m-%d')
                    
                    if week_key not in weekly_messages:
                        weekly_messages[week_key] = []
                    
                    sender_name = await self.get_sender_name(msg)
                    weekly_messages[week_key].append({
                        'id': msg.id,
                        'date': msg_date,
                        'text': msg.text.strip(),
                        'sender': sender_name,
                        'week_start': week_start,
                        'week_key': week_key
                    })
            
            print(f"📊 Сгруппировано по {len(weekly_messages)} неделям")
            return weekly_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return {}
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def extract_plan_fact_sections(self, text, sender, message_date):
        """Извлечение разделов план и факт из текста"""
        try:
            # Разбиваем на строки
            lines = text.split('\n')
            
            plan_section = []
            fact_section = []
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Определяем раздел
                if re.match(r'^план[:\s]*', line.lower()):
                    current_section = 'plan'
                    continue
                elif re.match(r'^факт[:\s]*', line.lower()):
                    current_section = 'fact'
                    continue
                
                # Добавляем в соответствующий раздел
                if current_section == 'plan':
                    plan_section.append(line)
                elif current_section == 'fact':
                    fact_section.append(line)
            
            return {
                'sender': sender,
                'date': message_date,
                'plan_tasks': plan_section,
                'fact_tasks': fact_section,
                'has_plan': len(plan_section) > 0,
                'has_fact': len(fact_section) > 0
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка извлечения разделов: {e}")
            return None
    
    def analyze_weekly_comparison(self, weekly_messages):
        """Анализ сравнения недель с детальным сопоставлением"""
        try:
            print("🔍 Анализ сравнения недель с детальным сопоставлением...")
            
            detailed_comparisons = []
            week_keys = sorted(weekly_messages.keys())
            
            # Проходим по неделям (кроме последней)
            for i in range(len(week_keys) - 1):
                current_week = week_keys[i]
                next_week = week_keys[i + 1]
                
                current_week_messages = weekly_messages[current_week]
                next_week_messages = weekly_messages[next_week]
                
                # Анализируем планы текущей недели
                current_week_plans = {}
                for msg in current_week_messages:
                    sections = self.extract_plan_fact_sections(msg['text'], msg['sender'], msg['date'])
                    if sections and sections['has_plan']:
                        if sections['sender'] not in current_week_plans:
                            current_week_plans[sections['sender']] = []
                        current_week_plans[sections['sender']].extend(sections['plan_tasks'])
                
                # Анализируем факты следующей недели
                next_week_facts = {}
                for msg in next_week_messages:
                    sections = self.extract_plan_fact_sections(msg['text'], msg['sender'], msg['date'])
                    if sections and sections['has_fact']:
                        if sections['sender'] not in next_week_facts:
                            next_week_facts[sections['sender']] = []
                        next_week_facts[sections['sender']].extend(sections['fact_tasks'])
                
                # Сравниваем для каждого руководителя
                for sender in set(current_week_plans.keys()) | set(next_week_facts.keys()):
                    plan_tasks = current_week_plans.get(sender, [])
                    fact_tasks = next_week_facts.get(sender, [])
                    
                    if plan_tasks or fact_tasks:
                        detailed_comparison = self.create_detailed_task_comparison(
                            plan_tasks, fact_tasks, sender, current_week, next_week
                        )
                        if detailed_comparison:
                            detailed_comparisons.append(detailed_comparison)
            
            print(f"📊 Создано {len(detailed_comparisons)} детальных сравнений")
            return detailed_comparisons
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа сравнения: {e}")
            return []
    
    def create_detailed_task_comparison(self, plan_tasks, fact_tasks, sender, plan_week, fact_week):
        """Создание детального сопоставления задач"""
        try:
            # Сопоставляем задачи
            matched_tasks = []
            unmatched_plans = []
            unmatched_facts = []
            
            # Создаем копии для работы
            remaining_facts = fact_tasks.copy()
            
            # Ищем соответствия для каждой задачи из плана
            for plan_task in plan_tasks:
                best_match = None
                best_similarity = 0
                best_fact_index = -1
                
                # Ищем лучшее соответствие среди фактов
                for i, fact_task in enumerate(remaining_facts):
                    similarity = self.calculate_task_similarity(plan_task, fact_task)
                    
                    if similarity > best_similarity and similarity >= 0.3:  # Минимальный порог
                        best_similarity = similarity
                        best_match = fact_task
                        best_fact_index = i
                
                if best_match:
                    # Найдено соответствие
                    matched_tasks.append({
                        'plan_task': plan_task,
                        'fact_task': best_match,
                        'similarity': best_similarity,
                        'status': self.get_match_status(best_similarity)
                    })
                    # Удаляем использованный факт
                    remaining_facts.pop(best_fact_index)
                else:
                    # Не найдено соответствие
                    unmatched_plans.append(plan_task)
            
            # Оставшиеся факты - новые задачи
            unmatched_facts = remaining_facts
            
            # Вычисляем статистику
            total_planned = len(plan_tasks)
            total_facts = len(fact_tasks)
            total_matched = len(matched_tasks)
            total_unmatched_plans = len(unmatched_plans)
            total_new_tasks = len(unmatched_facts)
            
            # Процент выполнения (только по сопоставленным задачам)
            completion_rate = (total_matched / total_planned * 100) if total_planned > 0 else 0
            
            return {
                'sender': sender,
                'plan_week': plan_week,
                'fact_week': fact_week,
                'plan_tasks': plan_tasks,
                'fact_tasks': fact_tasks,
                'matched_tasks': matched_tasks,
                'unmatched_plans': unmatched_plans,
                'unmatched_facts': unmatched_facts,
                'total_planned': total_planned,
                'total_facts': total_facts,
                'total_matched': total_matched,
                'total_unmatched_plans': total_unmatched_plans,
                'total_new_tasks': total_new_tasks,
                'completion_rate': completion_rate,
                'overall_status': self.get_overall_status(completion_rate, total_matched, total_planned)
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания детального сопоставления: {e}")
            return None
    
    def calculate_task_similarity(self, task1, task2):
        """Вычисление сходства между задачами"""
        try:
            # Приводим к нижнему регистру
            t1 = task1.lower()
            t2 = task2.lower()
            
            # Убираем цифры и специальные символы
            t1_clean = re.sub(r'[^\w\s]', '', t1)
            t2_clean = re.sub(r'[^\w\s]', '', t2)
            
            # Разбиваем на слова
            words1 = set(t1_clean.split())
            words2 = set(t2_clean.split())
            
            if not words1 or not words2:
                return 0.0
            
            # Вычисляем Jaccard similarity
            intersection = len(words1.intersection(words2))
            union = len(words1.union(words2))
            
            return intersection / union if union > 0 else 0.0
            
        except Exception as e:
            return 0.0
    
    def get_match_status(self, similarity):
        """Определение статуса сопоставления"""
        if similarity >= 0.7:
            return "✅ Выполнено"
        elif similarity >= 0.5:
            return "⚠️ Частично выполнено"
        elif similarity >= 0.3:
            return "🔄 Изменено"
        else:
            return "❌ Не выполнено"
    
    def get_overall_status(self, completion_rate, matched, planned):
        """Определение общего статуса"""
        if completion_rate >= 80 and matched > 0:
            return "✅ Отлично"
        elif completion_rate >= 60:
            return "⚠️ Хорошо"
        elif completion_rate >= 40:
            return "⚠️ Средне"
        elif planned > 0:
            return "❌ Плохо"
        else:
            return "📝 Только новые задачи"
    
    def create_detailed_report(self, comparisons, filename=None):
        """Создание детального отчета"""
        try:
            if not filename:
                filename = f"output/detailed_task_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            # Удаляем дефолтный лист
            workbook.remove(workbook.active)
            
            # Создаем листы
            self.create_summary_sheet(workbook, comparisons)
            self.create_detailed_comparison_sheet(workbook, comparisons)
            self.create_task_matching_sheet(workbook, comparisons)
            self.create_unmatched_tasks_sheet(workbook, comparisons)
            
            workbook.save(filename)
            print(f"✅ Детальный отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    def create_summary_sheet(self, workbook, comparisons):
        """Создание сводного листа"""
        try:
            ws = workbook.create_sheet("Сводка")
            
            # Заголовок
            ws['A1'] = "ДЕТАЛЬНОЕ СОПОСТАВЛЕНИЕ ПЛАНОВ И ФАКТОВ"
            ws['A2'] = "Main mega ТОП - Анализ сопоставления задач"
            
            # Общая статистика
            total_comparisons = len(comparisons)
            total_planned = sum(comp['total_planned'] for comp in comparisons)
            total_facts = sum(comp['total_facts'] for comp in comparisons)
            total_matched = sum(comp['total_matched'] for comp in comparisons)
            total_unmatched_plans = sum(comp['total_unmatched_plans'] for comp in comparisons)
            total_new_tasks = sum(comp['total_new_tasks'] for comp in comparisons)
            
            avg_completion = (total_matched / total_planned * 100) if total_planned > 0 else 0
            
            ws['A4'] = "📊 ОБЩАЯ СТАТИСТИКА СОПОСТАВЛЕНИЯ:"
            ws['A5'] = f"Всего сравнений: {total_comparisons}"
            ws['A6'] = f"Всего запланировано: {total_planned}"
            ws['A7'] = f"Всего выполнено (факт): {total_facts}"
            ws['A8'] = f"Сопоставлено задач: {total_matched}"
            ws['A9'] = f"Не выполнено (план): {total_unmatched_plans}"
            ws['A10'] = f"Новых задач (факт): {total_new_tasks}"
            ws['A11'] = f"Процент сопоставления: {avg_completion:.1f}%"
            
            # Статистика по руководителям
            sender_stats = {}
            for comp in comparisons:
                sender = comp['sender']
                if sender not in sender_stats:
                    sender_stats[sender] = {
                        'planned': 0, 'facts': 0, 'matched': 0,
                        'unmatched_plans': 0, 'new_tasks': 0, 'comparisons': 0
                    }
                
                sender_stats[sender]['planned'] += comp['total_planned']
                sender_stats[sender]['facts'] += comp['total_facts']
                sender_stats[sender]['matched'] += comp['total_matched']
                sender_stats[sender]['unmatched_plans'] += comp['total_unmatched_plans']
                sender_stats[sender]['new_tasks'] += comp['total_new_tasks']
                sender_stats[sender]['comparisons'] += 1
            
            ws['A13'] = "👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ:"
            
            headers = ['Руководитель', 'Сравнений', 'Запланировано', 'Факт', 'Сопоставлено', 'Не выполнено', 'Новых', '% сопоставления']
            for col, header in enumerate(headers, 1):
                ws.cell(row=14, column=col, value=header)
            
            row = 15
            for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['matched'], reverse=True):
                completion_rate = (stats['matched'] / stats['planned'] * 100) if stats['planned'] > 0 else 0
                
                ws.cell(row=row, column=1, value=sender)
                ws.cell(row=row, column=2, value=stats['comparisons'])
                ws.cell(row=row, column=3, value=stats['planned'])
                ws.cell(row=row, column=4, value=stats['facts'])
                ws.cell(row=row, column=5, value=stats['matched'])
                ws.cell(row=row, column=6, value=stats['unmatched_plans'])
                ws.cell(row=row, column=7, value=stats['new_tasks'])
                ws.cell(row=row, column=8, value=f"{completion_rate:.1f}%")
                
                row += 1
            
            # Форматирование
            self.format_summary_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания сводного листа: {e}")
    
    def create_detailed_comparison_sheet(self, workbook, comparisons):
        """Создание листа детального сравнения"""
        try:
            ws = workbook.create_sheet("Детальное сравнение")
            
            # Заголовки
            headers = [
                'Руководитель', 'Неделя плана', 'Неделя факта', 
                'Запланировано', 'Факт', 'Сопоставлено', 'Не выполнено', 'Новых',
                '% сопоставления', 'Статус'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for comp in comparisons:
                ws.cell(row=row, column=1, value=comp['sender'])
                ws.cell(row=row, column=2, value=comp['plan_week'])
                ws.cell(row=row, column=3, value=comp['fact_week'])
                ws.cell(row=row, column=4, value=comp['total_planned'])
                ws.cell(row=row, column=5, value=comp['total_facts'])
                ws.cell(row=row, column=6, value=comp['total_matched'])
                ws.cell(row=row, column=7, value=comp['total_unmatched_plans'])
                ws.cell(row=row, column=8, value=comp['total_new_tasks'])
                ws.cell(row=row, column=9, value=f"{comp['completion_rate']:.1f}%")
                ws.cell(row=row, column=10, value=comp['overall_status'])
                
                row += 1
            
            # Форматирование
            self.format_detailed_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа детального сравнения: {e}")
    
    def create_task_matching_sheet(self, workbook, comparisons):
        """Создание листа сопоставления задач"""
        try:
            ws = workbook.create_sheet("Сопоставление задач")
            
            # Заголовки
            headers = [
                'Руководитель', 'Неделя плана', 'Неделя факта',
                'Задача из плана', 'Задача из факта', 'Сходство', 'Статус'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for comp in comparisons:
                for match in comp['matched_tasks']:
                    ws.cell(row=row, column=1, value=comp['sender'])
                    ws.cell(row=row, column=2, value=comp['plan_week'])
                    ws.cell(row=row, column=3, value=comp['fact_week'])
                    ws.cell(row=row, column=4, value=match['plan_task'])
                    ws.cell(row=row, column=5, value=match['fact_task'])
                    ws.cell(row=row, column=6, value=f"{match['similarity']:.2f}")
                    ws.cell(row=row, column=7, value=match['status'])
                    
                    # Форматирование ячеек с текстом
                    for col in [4, 5]:
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        ws.row_dimensions[row].height = 60
                    
                    row += 1
            
            # Форматирование
            self.format_task_matching_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа сопоставления задач: {e}")
    
    def create_unmatched_tasks_sheet(self, workbook, comparisons):
        """Создание листа несопоставленных задач"""
        try:
            ws = workbook.create_sheet("Несопоставленные задачи")
            
            # Заголовки
            headers = [
                'Руководитель', 'Неделя плана', 'Неделя факта',
                'Тип', 'Задача', 'Примечание'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for comp in comparisons:
                # Невыполненные планы
                for plan_task in comp['unmatched_plans']:
                    ws.cell(row=row, column=1, value=comp['sender'])
                    ws.cell(row=row, column=2, value=comp['plan_week'])
                    ws.cell(row=row, column=3, value=comp['fact_week'])
                    ws.cell(row=row, column=4, value="❌ Не выполнено")
                    ws.cell(row=row, column=5, value=plan_task)
                    ws.cell(row=row, column=6, value="Задача из плана не найдена в фактах")
                    
                    # Форматирование
                    cell = ws.cell(row=row, column=4)
                    cell.fill = PatternFill(start_color=self.colors['red'], end_color=self.colors['red'], fill_type='solid')
                    cell.font = Font(color='FFFFFF')
                    
                    row += 1
                
                # Новые задачи
                for fact_task in comp['unmatched_facts']:
                    ws.cell(row=row, column=1, value=comp['sender'])
                    ws.cell(row=row, column=2, value=comp['plan_week'])
                    ws.cell(row=row, column=3, value=comp['fact_week'])
                    ws.cell(row=row, column=4, value="🆕 Новая задача")
                    ws.cell(row=row, column=5, value=fact_task)
                    ws.cell(row=row, column=6, value="Задача из фактов не была в планах")
                    
                    # Форматирование
                    cell = ws.cell(row=row, column=4)
                    cell.fill = PatternFill(start_color=self.colors['orange'], end_color=self.colors['orange'], fill_type='solid')
                    cell.font = Font(color='FFFFFF')
                    
                    row += 1
            
            # Форматирование
            self.format_unmatched_tasks_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа несопоставленных задач: {e}")
    
    def format_summary_sheet(self, ws):
        """Форматирование сводного листа"""
        # Заголовок
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['blue'])
        ws['A2'].font = Font(size=14, bold=True)
        
        # Объединение ячеек
        ws.merge_cells('A1:H1')
        ws.merge_cells('A2:H2')
        
        # Ширина колонок
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def format_detailed_sheet(self, ws):
        """Форматирование детального листа"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Ширина колонок
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def format_task_matching_sheet(self, ws):
        """Форматирование листа сопоставления задач"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
    
    def format_unmatched_tasks_sheet(self, ws):
        """Форматирование листа несопоставленных задач"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 30
    
    async def run_detailed_comparison(self):
        """Запуск детального анализа сопоставления"""
        try:
            print("🚀 ЗАПУСК ДЕТАЛЬНОГО АНАЛИЗА СОПОСТАВЛЕНИЯ")
            print("=" * 60)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение сообщений по неделям
            weekly_messages = await self.get_weekly_messages(weeks_back=8)
            
            if not weekly_messages:
                print("❌ Не найдено сообщений по неделям")
                return
            
            # Анализ сравнения
            comparisons = self.analyze_weekly_comparison(weekly_messages)
            
            if not comparisons:
                print("❌ Не найдено сравнений планов и фактов")
                return
            
            # Создание отчета
            report_file = self.create_detailed_report(comparisons)
            
            if report_file:
                print(f"\n🎉 Детальный анализ сопоставления завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА СОПОСТАВЛЕНИЯ:")
                print(f"Проанализировано недель: {len(weekly_messages)}")
                print(f"Создано сравнений: {len(comparisons)}")
                
                # Статистика сопоставления
                total_planned = sum(comp['total_planned'] for comp in comparisons)
                total_facts = sum(comp['total_facts'] for comp in comparisons)
                total_matched = sum(comp['total_matched'] for comp in comparisons)
                total_unmatched_plans = sum(comp['total_unmatched_plans'] for comp in comparisons)
                total_new_tasks = sum(comp['total_new_tasks'] for comp in comparisons)
                
                print(f"Всего запланировано: {total_planned}")
                print(f"Всего выполнено (факт): {total_facts}")
                print(f"Сопоставлено задач: {total_matched}")
                print(f"Не выполнено (план): {total_unmatched_plans}")
                print(f"Новых задач (факт): {total_new_tasks}")
                
                avg_completion = (total_matched / total_planned * 100) if total_planned > 0 else 0
                print(f"Процент сопоставления: {avg_completion:.1f}%")
                
                # Показываем топ руководителей по сопоставлению
                sender_stats = {}
                for comp in comparisons:
                    sender = comp['sender']
                    if sender not in sender_stats:
                        sender_stats[sender] = {'planned': 0, 'matched': 0}
                    sender_stats[sender]['planned'] += comp['total_planned']
                    sender_stats[sender]['matched'] += comp['total_matched']
                
                print(f"\n👥 ТОП РУКОВОДИТЕЛЕЙ ПО СОПОСТАВЛЕНИЮ:")
                for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['matched'], reverse=True)[:5]:
                    completion_rate = (stats['matched'] / stats['planned'] * 100) if stats['planned'] > 0 else 0
                    print(f"  {sender}: {stats['matched']}/{stats['planned']} ({completion_rate:.1f}%)")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = DetailedTaskComparisonAnalyzer()
    await analyzer.run_detailed_comparison()

if __name__ == "__main__":
    asyncio.run(main())
