#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Единый список задач с сопоставлением планов и фактов
Main mega ТОП - Июль-Август 2025
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from telethon.tl.types import PeerChannel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class UnifiedTaskComparisonAnalyzer:
    def __init__(self):
        # Загрузка конфигурации
        with open('config/personal_config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        self.api_id = config['api_id']
        self.api_hash = config['api_hash']
        self.phone = config['phone']
        self.session_name = config['session_name']
        self.chat_name = 'Main mega ТОП'
        
        # Инициализация клиента
        self.client = TelegramClient(self.session_name, self.api_id, self.api_hash)
        
    async def connect(self):
        """Подключение к Telegram"""
        try:
            await self.client.start(phone=self.phone)
            me = await self.client.get_me()
            logger.info(f"👤 Подключен как: {me.first_name}")
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_chat(self):
        """Поиск чата по имени"""
        try:
            logger.info(f"🔍 Поиск чата '{self.chat_name}'...")
            async for dialog in self.client.iter_dialogs():
                if self.chat_name in dialog.name:
                    logger.info(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return dialog
            logger.error(f"❌ Чат '{self.chat_name}' не найден")
            return None
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return None
    
    async def get_july_august_messages(self):
        """Загрузка сообщений за июль и август 2025"""
        try:
            logger.info("📥 Загрузка сообщений за июль и август 2025...")
            
            # Определяем период (убираем timezone)
            july_start = datetime(2025, 7, 1)
            august_end = datetime(2025, 8, 31, 23, 59, 59)
            
            # Получаем сообщения
            messages = await self.client.get_messages(
                self.chat,
                offset_date=july_start,
                limit=20000
            )
            
            # Фильтруем по дате (убираем timezone из дат сообщений)
            filtered_messages = []
            for msg in messages:
                if msg.date:
                    # Убираем timezone для сравнения
                    msg_date = msg.date.replace(tzinfo=None)
                    if july_start <= msg_date <= august_end:
                        filtered_messages.append(msg)
            
            logger.info(f"📊 Отфильтровано {len(filtered_messages)} сообщений за июль-август")
            
            # Если сообщений мало, пробуем загрузить больше
            if len(filtered_messages) < 50:
                logger.info("🔄 Попытка загрузки большего количества сообщений...")
                messages = await self.client.get_messages(
                    self.chat,
                    limit=50000
                )
                
                filtered_messages = []
                for msg in messages:
                    if msg.date:
                        # Убираем timezone для сравнения
                        msg_date = msg.date.replace(tzinfo=None)
                        if july_start <= msg_date <= august_end:
                            filtered_messages.append(msg)
                
                logger.info(f"📊 После расширенной загрузки: {len(filtered_messages)} сообщений за июль-август")
            
            return filtered_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки сообщений: {e}")
            return []
    
    def extract_plan_fact_sections(self, text, sender, message_date):
        """Извлечение разделов план и факт из текста"""
        try:
            # Разбиваем на строки
            lines = text.split('\n')
            
            plan_section = []
            fact_section = []
            current_section = None
            
            # Метод 1: Поиск по ключевым словам "План:" и "Факт:"
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Определяем раздел по ключевым словам
                if re.match(r'^план[:\s]*', line.lower()):
                    current_section = 'plan'
                    continue
                elif re.match(r'^факт[:\s]*', line.lower()):
                    current_section = 'fact'
                    continue
                
                # Добавляем в соответствующий раздел
                if current_section == 'plan':
                    plan_section.append(line)
                elif current_section == 'fact':
                    fact_section.append(line)
            
            # Метод 2: Если не найдены разделы с ключевыми словами, ищем числовые списки
            if not plan_section and not fact_section:
                numbered_lines = []
                for line in lines:
                    line = line.strip()
                    if re.match(r'^\d+[\.\)]', line):
                        numbered_lines.append(line)
                
                if len(numbered_lines) > 0:
                    # Разделяем на две части (предполагаем, что первая часть - план, вторая - факт)
                    mid_point = len(numbered_lines) // 2
                    plan_section = numbered_lines[:mid_point]
                    fact_section = numbered_lines[mid_point:]
            
            # Метод 3: Ищем строки с дефисами или маркерами
            if not plan_section and not fact_section:
                bullet_lines = []
                for line in lines:
                    line = line.strip()
                    if re.match(r'^[-•*]\s', line):
                        bullet_lines.append(line)
                
                if len(bullet_lines) > 0:
                    mid_point = len(bullet_lines) // 2
                    plan_section = bullet_lines[:mid_point]
                    fact_section = bullet_lines[mid_point:]
            
            # Метод 4: Простое разделение по количеству строк (для больших сообщений)
            if not plan_section and not fact_section:
                non_empty_lines = [line.strip() for line in lines if line.strip() and len(line.strip()) > 3]
                if len(non_empty_lines) > 3:  # Минимум 4 строки для разделения
                    mid_point = len(non_empty_lines) // 2
                    plan_section = non_empty_lines[:mid_point]
                    fact_section = non_empty_lines[mid_point:]
            
            # Метод 5: Поиск по контексту (если в тексте есть слова "планирую", "сделал", "выполнил")
            if not plan_section and not fact_section:
                plan_keywords = ['планирую', 'планируем', 'будем', 'собираемся', 'намерены']
                fact_keywords = ['сделал', 'сделали', 'выполнил', 'выполнили', 'завершил', 'завершили']
                
                plan_lines = []
                fact_lines = []
                
                for line in lines:
                    line = line.strip()
                    if not line or len(line) < 5:
                        continue
                    
                    line_lower = line.lower()
                    if any(keyword in line_lower for keyword in plan_keywords):
                        plan_lines.append(line)
                    elif any(keyword in line_lower for keyword in fact_keywords):
                        fact_lines.append(line)
                
                if plan_lines and fact_lines:
                    plan_section = plan_lines
                    fact_section = fact_lines
            
            return {
                'sender': sender,
                'date': message_date,
                'plan_tasks': plan_section,
                'fact_tasks': fact_section,
                'has_plan': len(plan_section) > 0,
                'has_fact': len(fact_section) > 0
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка извлечения разделов: {e}")
            return None
    
    def get_week_range(self, date):
        """Получение диапазона недели для даты"""
        # Находим понедельник недели
        days_since_monday = date.weekday()
        monday = date - timedelta(days=days_since_monday)
        sunday = monday + timedelta(days=6)
        
        return monday, sunday
    
    def format_week_range(self, monday, sunday):
        """Форматирование диапазона недели"""
        return f"{monday.strftime('%d.%m')}-{sunday.strftime('%d.%m')}"
    
    def calculate_task_similarity(self, task1, task2):
        """Вычисление схожести между задачами"""
        try:
            # Очищаем задачи от лишних символов
            clean_task1 = re.sub(r'[^\w\s]', '', task1.lower())
            clean_task2 = re.sub(r'[^\w\s]', '', task2.lower())
            
            # Разбиваем на слова
            words1 = set(clean_task1.split())
            words2 = set(clean_task2.split())
            
            # Убираем стоп-слова
            stop_words = {'и', 'в', 'на', 'с', 'по', 'для', 'от', 'до', 'за', 'из', 'к', 'о', 'у', 'я', 'мы', 'он', 'она', 'оно', 'они', 'это', 'то', 'так', 'как', 'что', 'где', 'когда', 'почему', 'какой', 'какая', 'какое', 'какие', 'чей', 'чья', 'чьё', 'чьи', 'много', 'мало', 'больше', 'меньше', 'все', 'вся', 'всё', 'всех', 'всех', 'каждый', 'каждая', 'каждое', 'каждые', 'любой', 'любая', 'любое', 'любые', 'никто', 'ничто', 'нигде', 'никогда', 'нипочему', 'никакой', 'никакая', 'никакое', 'никакие', 'ничей', 'ничья', 'ничьё', 'ничьи'}
            words1 = words1 - stop_words
            words2 = words2 - stop_words
            
            if not words1 or not words2:
                return 0.0
            
            # Вычисляем Jaccard similarity
            intersection = len(words1.intersection(words2))
            union = len(words1.union(words2))
            
            return intersection / union if union > 0 else 0.0
            
        except Exception as e:
            logger.error(f"❌ Ошибка вычисления схожести: {e}")
            return 0.0
    
    def find_best_match(self, plan_task, fact_tasks, threshold=0.3):
        """Поиск лучшего совпадения для задачи плана среди фактов"""
        best_match = None
        best_similarity = 0.0
        
        for fact_task in fact_tasks:
            similarity = self.calculate_task_similarity(plan_task, fact_task)
            if similarity > best_similarity and similarity >= threshold:
                best_similarity = similarity
                best_match = fact_task
        
        return best_match, best_similarity
    
    def analyze_unified_tasks(self, messages):
        """Анализ всех задач с созданием единого списка"""
        try:
            logger.info("🔍 Анализ сообщений и извлечение задач...")
            
            # Извлекаем планы и факты из сообщений
            plan_fact_data = []
            for msg in messages:
                if msg.text and msg.sender:
                    sender_name = msg.sender.first_name if msg.sender.first_name else "Неизвестный"
                    if hasattr(msg.sender, 'last_name') and msg.sender.last_name:
                        sender_name += f" {msg.sender.last_name}"
                    
                    # Проверяем, что сообщение достаточно большое (больше 50 символов)
                    if len(msg.text.strip()) > 50:
                        extracted = self.extract_plan_fact_sections(msg.text, sender_name, msg.date)
                        if extracted and (extracted['has_plan'] or extracted['has_fact']):
                            plan_fact_data.append(extracted)
                            logger.info(f"📝 Найдено сообщение от {sender_name}: {len(extracted['plan_tasks'])} планов, {len(extracted['fact_tasks'])} фактов")
            
            logger.info(f"📊 Извлечено {len(plan_fact_data)} больших сообщений с планами/фактами")
            
            # Группируем по неделям
            weekly_data = {}
            for data in plan_fact_data:
                plan_monday, plan_sunday = self.get_week_range(data['date'])
                plan_week = self.format_week_range(plan_monday, plan_sunday)
                
                if plan_week not in weekly_data:
                    weekly_data[plan_week] = {
                        'plan_week': plan_week,
                        'fact_week': self.format_week_range(plan_monday + timedelta(days=7), plan_sunday + timedelta(days=7)),
                        'plans': [],
                        'facts': [],
                        'managers': set()
                    }
                
                weekly_data[plan_week]['managers'].add(data['sender'])
                
                # Добавляем планы
                for task in data['plan_tasks']:
                    weekly_data[plan_week]['plans'].append({
                        'task': task,
                        'manager': data['sender'],
                        'date': data['date']
                    })
                
                # Добавляем факты (для следующей недели)
                fact_week = self.format_week_range(plan_monday + timedelta(days=7), plan_sunday + timedelta(days=7))
                if fact_week not in weekly_data:
                    weekly_data[fact_week] = {
                        'plan_week': self.format_week_range(plan_monday - timedelta(days=7), plan_sunday - timedelta(days=7)),
                        'fact_week': fact_week,
                        'plans': [],
                        'facts': [],
                        'managers': set()
                    }
                
                weekly_data[fact_week]['managers'].add(data['sender'])
                for task in data['fact_tasks']:
                    weekly_data[fact_week]['facts'].append({
                        'task': task,
                        'manager': data['sender'],
                        'date': data['date']
                    })
            
            logger.info(f"📊 Сгруппировано по {len(weekly_data)} неделям")
            
            # Создаем единый список всех задач
            unified_tasks = []
            
            # Сортируем недели в хронологическом порядке
            sorted_weeks = sorted(weekly_data.keys())
            
            for week in sorted_weeks:
                week_data = weekly_data[week]
                
                # Получаем планы предыдущей недели и факты текущей недели
                plan_tasks = week_data['plans']
                fact_tasks = week_data['facts']
                
                logger.info(f"📅 Неделя {week}: {len(plan_tasks)} планов, {len(fact_tasks)} фактов")
                
                # Сопоставляем задачи
                matched_plans = set()
                matched_facts = set()
                
                for plan_item in plan_tasks:
                    best_match, similarity = self.find_best_match(plan_item['task'], [f['task'] for f in fact_tasks])
                    
                    if best_match:
                        # Находим соответствующий факт
                        fact_item = next(f for f in fact_tasks if f['task'] == best_match)
                        
                        unified_tasks.append({
                            'Неделя плана': week_data['plan_week'],
                            'Неделя факта': week_data['fact_week'],
                            'Руководитель': plan_item['manager'],
                            'Задача плана': plan_item['task'],
                            'Задача факта': fact_item['task'],
                            'Схожесть': f"{similarity:.2f}",
                            'Статус': self.get_match_status(similarity),
                            'Тип': 'Сопоставлено'
                        })
                        
                        matched_plans.add(plan_item['task'])
                        matched_facts.add(best_match)
                
                # Несопоставленные планы
                for plan_item in plan_tasks:
                    if plan_item['task'] not in matched_plans:
                        unified_tasks.append({
                            'Неделя плана': week_data['plan_week'],
                            'Неделя факта': week_data['fact_week'],
                            'Руководитель': plan_item['manager'],
                            'Задача плана': plan_item['task'],
                            'Задача факта': '',
                            'Схожесть': '',
                            'Статус': '❌ Не выполнено',
                            'Тип': 'Не выполнено'
                        })
                
                # Новые задачи (факты без планов)
                for fact_item in fact_tasks:
                    if fact_item['task'] not in matched_facts:
                        unified_tasks.append({
                            'Неделя плана': week_data['plan_week'],
                            'Неделя факта': week_data['fact_week'],
                            'Руководитель': fact_item['manager'],
                            'Задача плана': '',
                            'Задача факта': fact_item['task'],
                            'Схожесть': '',
                            'Статус': '🆕 Новая задача',
                            'Тип': 'Новая задача'
                        })
            
            logger.info(f"📊 Создан единый список из {len(unified_tasks)} задач")
            return unified_tasks
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа задач: {e}")
            return []
    
    def get_match_status(self, similarity):
        """Получение статуса сопоставления"""
        if similarity >= 0.7:
            return "✅ Выполнено"
        elif similarity >= 0.5:
            return "🔄 Частично выполнено"
        elif similarity >= 0.3:
            return "🔄 Изменено"
        else:
            return "❌ Не выполнено"
    
    def create_unified_report(self, unified_tasks):
        """Создание единого отчета"""
        try:
            logger.info("📊 Создание единого отчета...")
            
            # Создаем DataFrame
            df = pd.DataFrame(unified_tasks)
            
            # Сортируем по неделе плана, руководителю и типу
            df = df.sort_values(['Неделя плана', 'Руководитель', 'Тип'])
            
            # Создаем Excel файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"output/unified_task_comparison_{timestamp}.xlsx"
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Основной лист с единым списком
                df.to_excel(writer, sheet_name='Единый список задач', index=False)
                
                # Получаем workbook для форматирования
                workbook = writer.book
                worksheet = writer.sheets['Единый список задач']
                
                # Форматирование
                self.format_unified_worksheet(worksheet, df)
                
                # Создаем сводный лист
                self.create_summary_sheet(workbook, df)
                
                # Создаем лист по руководителям
                self.create_managers_sheet(workbook, df)
            
            logger.info(f"✅ Единый отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    def format_unified_worksheet(self, worksheet, df):
        """Форматирование основного листа"""
        try:
            # Настройка ширины колонок
            column_widths = {
                'A': 15,  # Неделя плана
                'B': 15,  # Неделя факта
                'C': 20,  # Руководитель
                'D': 50,  # Задача плана
                'E': 50,  # Задача факта
                'F': 10,  # Схожесть
                'G': 20,  # Статус
                'H': 15   # Тип
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Заголовки
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Форматирование строк по типу
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                task_type = row[7].value  # Тип
                
                if task_type == 'Сопоставлено':
                    # Зеленый фон для сопоставленных
                    for cell in row:
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif task_type == 'Не выполнено':
                    # Красный фон для невыполненных
                    for cell in row:
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                elif task_type == 'Новая задача':
                    # Синий фон для новых задач
                    for cell in row:
                        cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                
                # Выравнивание текста
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Границы
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1):
                for cell in row:
                    cell.border = thin_border
            
        except Exception as e:
            logger.error(f"❌ Ошибка форматирования: {e}")
    
    def create_summary_sheet(self, workbook, df):
        """Создание сводного листа"""
        try:
            worksheet = workbook.create_sheet("Сводка")
            
            # Статистика
            total_tasks = len(df)
            matched_tasks = len(df[df['Тип'] == 'Сопоставлено'])
            unmatched_plans = len(df[df['Тип'] == 'Не выполнено'])
            new_tasks = len(df[df['Тип'] == 'Новая задача'])
            
            # Заголовок
            worksheet['A1'] = "📊 СВОДКА ЕДИНОГО СПИСКА ЗАДАЧ"
            worksheet['A1'].font = Font(bold=True, size=16)
            
            # Статистика
            worksheet['A3'] = "Общая статистика:"
            worksheet['A3'].font = Font(bold=True)
            
            stats = [
                ["Всего задач", total_tasks],
                ["Сопоставлено", matched_tasks],
                ["Не выполнено", unmatched_plans],
                ["Новых задач", new_tasks],
                ["% сопоставления", f"{(matched_tasks/total_tasks*100):.1f}%" if total_tasks > 0 else "0%"]
            ]
            
            for i, (label, value) in enumerate(stats, 4):
                worksheet[f'A{i}'] = label
                worksheet[f'B{i}'] = value
                worksheet[f'A{i}'].font = Font(bold=True)
            
            # Статистика по неделям
            worksheet['A10'] = "Статистика по неделям:"
            worksheet['A10'].font = Font(bold=True)
            
            week_stats = df.groupby('Неделя плана').agg({
                'Тип': 'count',
                'Схожесть': lambda x: len([s for s in x if s != ''])
            }).reset_index()
            
            week_stats.columns = ['Неделя', 'Всего задач', 'Сопоставлено']
            week_stats['% сопоставления'] = (week_stats['Сопоставлено'] / week_stats['Всего задач'] * 100).round(1)
            
            # Заголовки
            headers = ['Неделя', 'Всего задач', 'Сопоставлено', '% сопоставления']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=11, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Данные
            for row_idx, (_, row) in enumerate(week_stats.iterrows(), 12):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 4:  # % сопоставления
                        cell.value = f"{value}%"
            
            # Настройка ширины колонок
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 15
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания сводки: {e}")
    
    def create_managers_sheet(self, workbook, df):
        """Создание листа по руководителям"""
        try:
            worksheet = workbook.create_sheet("По руководителям")
            
            # Статистика по руководителям
            manager_stats = df.groupby('Руководитель').agg({
                'Тип': 'count',
                'Схожесть': lambda x: len([s for s in x if s != ''])
            }).reset_index()
            
            manager_stats.columns = ['Руководитель', 'Всего задач', 'Сопоставлено']
            manager_stats['% сопоставления'] = (manager_stats['Сопоставлено'] / manager_stats['Всего задач'] * 100).round(1)
            manager_stats = manager_stats.sort_values('% сопоставления', ascending=False)
            
            # Заголовок
            worksheet['A1'] = "👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ"
            worksheet['A1'].font = Font(bold=True, size=16)
            
            # Заголовки
            headers = ['Руководитель', 'Всего задач', 'Сопоставлено', '% сопоставления', 'Статус']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Данные
            for row_idx, (_, row) in enumerate(manager_stats.iterrows(), 4):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 4:  # % сопоставления
                        cell.value = f"{value}%"
                
                # Статус
                status_cell = worksheet.cell(row=row_idx, column=5)
                if row['% сопоставления'] >= 60:
                    status_cell.value = "✅ Отлично"
                    status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif row['% сопоставления'] >= 30:
                    status_cell.value = "⚠️ Средне"
                    status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                elif row['% сопоставления'] >= 10:
                    status_cell.value = "❌ Плохо"
                    status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                else:
                    status_cell.value = "🚨 Критически"
                    status_cell.fill = PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid")
            
            # Настройка ширины колонок
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа руководителей: {e}")
    
    async def run_analysis(self):
        """Запуск полного анализа"""
        try:
            logger.info("🚀 ЗАПУСК СОЗДАНИЯ ЕДИНОГО СПИСКА ЗАДАЧ")
            logger.info("=" * 70)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск чата
            self.chat = await self.find_chat()
            if not self.chat:
                return
            
            # Загрузка сообщений
            messages = await self.get_july_august_messages()
            if not messages:
                logger.error("❌ Не найдено сообщений для анализа")
                return
            
            # Анализ задач
            unified_tasks = self.analyze_unified_tasks(messages)
            if not unified_tasks:
                logger.error("❌ Не удалось проанализировать задачи")
                return
            
            # Создание отчета
            filename = self.create_unified_report(unified_tasks)
            if filename:
                logger.info(f"🎉 Единый список задач создан!")
                logger.info(f"📁 Файл сохранен: {filename}")
                
                # Краткая сводка
                total_tasks = len(unified_tasks)
                matched_tasks = len([t for t in unified_tasks if t['Тип'] == 'Сопоставлено'])
                unmatched_plans = len([t for t in unified_tasks if t['Тип'] == 'Не выполнено'])
                new_tasks = len([t for t in unified_tasks if t['Тип'] == 'Новая задача'])
                
                logger.info(f"\n📊 КРАТКАЯ СВОДКА ЕДИНОГО СПИСКА:")
                logger.info(f"Всего задач: {total_tasks}")
                logger.info(f"Сопоставлено: {matched_tasks}")
                logger.info(f"Не выполнено: {unmatched_plans}")
                logger.info(f"Новых задач: {new_tasks}")
                logger.info(f"Средний % сопоставления: {(matched_tasks/total_tasks*100):.1f}%" if total_tasks > 0 else "0%")
            
        except Exception as e:
            logger.error(f"❌ Ошибка выполнения анализа: {e}")
        finally:
            await self.client.disconnect()

async def main():
    """Главная функция"""
    analyzer = UnifiedTaskComparisonAnalyzer()
    await analyzer.run_analysis()

if __name__ == "__main__":
    asyncio.run(main())
