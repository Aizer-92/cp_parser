#!/usr/bin/env python3
"""
Full Chat Analyzer for Main mega ТОП
Полный анализатор чата за все время
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

class FullChatAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/full_chat_analyzer.log", rotation="1 day", retention="7 days")
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_all_messages(self):
        """Получение ВСЕХ сообщений из чата"""
        try:
            print("📥 Загрузка ВСЕХ сообщений из чата...")
            
            all_messages = []
            offset_id = 0
            limit = 100
            
            while True:
                messages = await self.client.get_messages(
                    self.target_chat_id,
                    limit=limit,
                    offset_id=offset_id
                )
                
                if not messages:
                    break
                
                # Фильтруем сообщения с текстом
                text_messages = []
                for msg in messages:
                    if msg.text and len(msg.text.strip()) > 5:
                        sender_name = await self.get_sender_name(msg)
                        text_messages.append({
                            'id': msg.id,
                            'date': msg.date,
                            'text': msg.text.strip(),
                            'sender': sender_name,
                            'length': len(msg.text.strip())
                        })
                
                all_messages.extend(text_messages)
                offset_id = messages[-1].id
                
                print(f"📊 Загружено {len(all_messages)} сообщений...")
                
                # Ограничиваем для тестирования (можно убрать)
                if len(all_messages) >= 10000:
                    print("⚠️ Достигнут лимит 10000 сообщений для анализа")
                    break
            
            print(f"✅ Всего загружено {len(all_messages)} сообщений")
            return all_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return []
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def analyze_message_for_plan_fact(self, text, message_date, sender_name, message_id):
        """Расширенный анализ сообщения на наличие планов и фактов"""
        try:
            # Расширенные паттерны для поиска
            patterns = [
                # Классические паттерны
                (r'план[:\s]*(\d+(?:\.\d+)?)', 'план_явный'),
                (r'факт[:\s]*(\d+(?:\.\d+)?)', 'факт_явный'),
                (r'п/ф[:\s]*(\d+(?:\.\d+)?)', 'пф_сокращение'),
                
                # Числа через слеш или дефис (план/факт)
                (r'(\d+(?:\.\d+)?)[/\-]\s*(\d+(?:\.\d+)?)', 'числа_слеш'),
                
                # Числа в скобках (план факт)
                (r'\((\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)\)', 'числа_скобки'),
                
                # Даты с числами (формат: дата + числа)
                (r'(\d{1,2}[\.\/]\d{1,2}[\.\/]\d{2,4})[^\d]*(\d+(?:\.\d+)?)', 'дата_числа'),
                
                # Просто числа в контексте задач
                (r'(\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)', 'два_числа'),
                
                # Проценты выполнения
                (r'(\d+(?:\.\d+)?)%[^\d]*(\d+(?:\.\d+)?)', 'проценты'),
            ]
            
            found_patterns = []
            extracted_data = []
            
            for pattern, pattern_name in patterns:
                matches = re.finditer(pattern, text.lower())
                for match in matches:
                    try:
                        if len(match.groups()) == 2:
                            num1 = float(match.group(1))
                            num2 = float(match.group(2))
                            
                            # Проверяем разумность значений
                            if 0 <= num1 <= 1000000 and 0 <= num2 <= 1000000:
                                found_patterns.append(pattern_name)
                                
                                # Определяем план и факт
                                plan_value, fact_value = self.determine_plan_fact(num1, num2, pattern_name, text)
                                
                                if plan_value is not None and fact_value is not None:
                                    completion_percent = (fact_value / plan_value * 100) if plan_value > 0 else 0
                                    
                                    extracted_data.append({
                                        'message_id': message_id,
                                        'date': message_date,
                                        'sender': sender_name,
                                        'plan': plan_value,
                                        'fact': fact_value,
                                        'difference': fact_value - plan_value,
                                        'completion_percent': completion_percent,
                                        'pattern_used': pattern_name,
                                        'match_text': match.group(0),
                                        'original_text': text[:300],
                                        'status': self.get_status(completion_percent),
                                        'is_duplicate': False
                                    })
                        elif len(match.groups()) == 1:
                            # Одиночные числа (может быть план или факт)
                            num = float(match.group(1))
                            if 0 <= num <= 1000000:
                                found_patterns.append(pattern_name)
                                
                                # Ищем парное число в том же сообщении
                                pair_data = self.find_pair_number(text, num, pattern_name, message_id, message_date, sender_name)
                                if pair_data:
                                    extracted_data.append(pair_data)
                                    
                    except (ValueError, ZeroDivisionError):
                        continue
            
            return extracted_data, found_patterns
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа сообщения: {e}")
            return [], []
    
    def determine_plan_fact(self, num1, num2, pattern_name, text):
        """Определение что есть план, а что факт"""
        try:
            # Для явных паттернов
            if pattern_name in ['план_явный', 'факт_явный']:
                if 'план' in text.lower():
                    return num1, num2
                elif 'факт' in text.lower():
                    return num2, num1
            
            # Для чисел через слеш (обычно план/факт)
            elif pattern_name == 'числа_слеш':
                return num1, num2
            
            # Для чисел в скобках
            elif pattern_name == 'числа_скобки':
                return num1, num2
            
            # Для дат с числами (обычно факт/план)
            elif pattern_name == 'дата_числа':
                return num2, num1
            
            # Для процентов
            elif pattern_name == 'проценты':
                return num1, num2
            
            # По умолчанию первое число - план, второе - факт
            else:
                return num1, num2
                
        except Exception as e:
            return None, None
    
    def find_pair_number(self, text, single_num, pattern_name, message_id, message_date, sender_name):
        """Поиск парного числа для одиночного"""
        try:
            # Ищем другие числа в том же сообщении
            numbers = re.findall(r'(\d+(?:\.\d+)?)', text)
            numbers = [float(n) for n in numbers if 0 <= float(n) <= 1000000]
            
            if len(numbers) >= 2:
                # Берем следующее число после найденного
                for i, num in enumerate(numbers):
                    if abs(num - single_num) < 0.01:  # Нашли наше число
                        if i + 1 < len(numbers):
                            pair_num = numbers[i + 1]
                            completion_percent = (pair_num / single_num * 100) if single_num > 0 else 0
                            
                            return {
                                'message_id': message_id,
                                'date': message_date,
                                'sender': sender_name,
                                'plan': single_num,
                                'fact': pair_num,
                                'difference': pair_num - single_num,
                                'completion_percent': completion_percent,
                                'pattern_used': f"{pattern_name}_парное",
                                'match_text': f"{single_num}/{pair_num}",
                                'original_text': text[:300],
                                'status': self.get_status(completion_percent),
                                'is_duplicate': False
                            }
            
            return None
            
        except Exception as e:
            return None
    
    def get_status(self, completion_percent):
        """Определение статуса выполнения"""
        if completion_percent >= 100:
            return "Перевыполнение"
        elif completion_percent >= 90:
            return "Хорошо"
        elif completion_percent >= 70:
            return "Удовлетворительно"
        else:
            return "Недовыполнение"
    
    def find_duplicates(self, all_data):
        """Поиск дублирующихся записей"""
        try:
            grouped_data = {}
            
            for item in all_data:
                key = (item['sender'], item['plan'], item['fact'])
                if key not in grouped_data:
                    grouped_data[key] = []
                grouped_data[key].append(item)
            
            # Помечаем дубликаты
            for key, items in grouped_data.items():
                if len(items) > 1:
                    for item in items:
                        item['is_duplicate'] = True
                        item['duplicate_count'] = len(items)
            
            return all_data
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска дубликатов: {e}")
            return all_data
    
    def analyze_by_periods(self, all_data):
        """Анализ по временным периодам"""
        try:
            periods = {
                '2025': [],
                '2024': [],
                '2023': [],
                '2022': [],
                '2021': []
            }
            
            for item in all_data:
                year = item['date'].year
                year_str = str(year)
                if year_str in periods:
                    periods[year_str].append(item)
            
            return periods
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа по периодам: {e}")
            return {}
    
    def create_full_report(self, all_data, periods_analysis, filename=None):
        """Создание полного отчета"""
        try:
            if not filename:
                filename = f"output/full_chat_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # Лист 1: Все данные
                if all_data:
                    df_all = pd.DataFrame(all_data)
                    df_all['date'] = pd.to_datetime(df_all['date']).dt.tz_localize(None)
                    df_all = df_all.sort_values(['sender', 'date'], ascending=[True, False])
                    
                    # Переименовываем колонки
                    df_all = df_all.rename(columns={
                        'message_id': 'ID сообщения',
                        'date': 'Дата',
                        'sender': 'Руководитель',
                        'plan': 'План',
                        'fact': 'Факт',
                        'difference': 'Разница',
                        'completion_percent': '% выполнения',
                        'pattern_used': 'Формат',
                        'match_text': 'Найденный текст',
                        'original_text': 'Исходный текст',
                        'status': 'Статус',
                        'is_duplicate': 'Дубликат',
                        'duplicate_count': 'Количество дубликатов'
                    })
                    
                    df_all.to_excel(writer, sheet_name='Все данные', index=False)
                    
                    # Форматирование
                    worksheet = writer.sheets['Все данные']
                    
                    # Подсветка дубликатов и недовыполнений
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    
                    for row in worksheet.iter_rows(min_row=2):
                        if row[df_all.columns.get_loc('Дубликат')].value:
                            for cell in row:
                                cell.fill = red_fill
                        elif row[df_all.columns.get_loc('% выполнения')].value and row[df_all.columns.get_loc('% выполнения')].value < 70:
                            for cell in row:
                                cell.fill = yellow_fill
                
                # Лист 2: Анализ по годам
                if periods_analysis:
                    year_data = []
                    for year, data in periods_analysis.items():
                        if data:
                            total_plan = sum(item['plan'] for item in data)
                            total_fact = sum(item['fact'] for item in data)
                            avg_completion = sum(item['completion_percent'] for item in data) / len(data)
                            
                            year_data.append({
                                'Год': year,
                                'Количество записей': len(data),
                                'Общий план': round(total_plan, 2),
                                'Общий факт': round(total_fact, 2),
                                'Средний % выполнения': round(avg_completion, 2),
                                'Уникальных руководителей': len(set(item['sender'] for item in data)),
                                'Дубликатов': len([item for item in data if item['is_duplicate']])
                            })
                    
                    df_years = pd.DataFrame(year_data)
                    df_years = df_years.sort_values('Год', ascending=False)
                    df_years.to_excel(writer, sheet_name='Анализ по годам', index=False)
                
                # Лист 3: Анализ по руководителям
                if all_data:
                    sender_analysis = {}
                    
                    for item in all_data:
                        sender = item['sender']
                        if sender not in sender_analysis:
                            sender_analysis[sender] = {
                                'total_plan': 0,
                                'total_fact': 0,
                                'tasks_count': 0,
                                'over_performance': 0,
                                'under_performance': 0,
                                'duplicates': 0,
                                'dates': set(),
                                'patterns': set()
                            }
                        
                        sender_analysis[sender]['total_plan'] += item['plan']
                        sender_analysis[sender]['total_fact'] += item['fact']
                        sender_analysis[sender]['tasks_count'] += 1
                        sender_analysis[sender]['dates'].add(item['date'].strftime('%Y-%m-%d'))
                        sender_analysis[sender]['patterns'].add(item['pattern_used'])
                        
                        if item['completion_percent'] >= 100:
                            sender_analysis[sender]['over_performance'] += 1
                        else:
                            sender_analysis[sender]['under_performance'] += 1
                        
                        if item['is_duplicate']:
                            sender_analysis[sender]['duplicates'] += 1
                    
                    # Создаем данные для отчета
                    sender_data = []
                    for sender, data in sender_analysis.items():
                        avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                        sender_data.append({
                            'Руководитель': sender,
                            'Общий план': round(data['total_plan'], 2),
                            'Общий факт': round(data['total_fact'], 2),
                            'Средний % выполнения': round(avg_completion, 2),
                            'Количество задач': data['tasks_count'],
                            'Перевыполнений': data['over_performance'],
                            'Недовыполнений': data['under_performance'],
                            'Дубликатов': data['duplicates'],
                            'Количество дат': len(data['dates']),
                            'Используемые форматы': ', '.join(sorted(data['patterns']))
                        })
                    
                    df_senders = pd.DataFrame(sender_data)
                    df_senders = df_senders.sort_values('Средний % выполнения', ascending=False)
                    df_senders.to_excel(writer, sheet_name='Анализ по руководителям', index=False)
                
                # Лист 4: Статистика по форматам
                if all_data:
                    pattern_stats = {}
                    for item in all_data:
                        pattern = item['pattern_used']
                        if pattern not in pattern_stats:
                            pattern_stats[pattern] = {
                                'count': 0,
                                'total_plan': 0,
                                'total_fact': 0,
                                'senders': set()
                            }
                        
                        pattern_stats[pattern]['count'] += 1
                        pattern_stats[pattern]['total_plan'] += item['plan']
                        pattern_stats[pattern]['total_fact'] += item['fact']
                        pattern_stats[pattern]['senders'].add(item['sender'])
                    
                    pattern_data = []
                    for pattern, stats in pattern_stats.items():
                        avg_completion = (stats['total_fact'] / stats['total_plan'] * 100) if stats['total_plan'] > 0 else 0
                        pattern_data.append({
                            'Формат': pattern,
                            'Количество использований': stats['count'],
                            'Общий план': round(stats['total_plan'], 2),
                            'Общий факт': round(stats['total_fact'], 2),
                            'Средний % выполнения': round(avg_completion, 2),
                            'Уникальных отправителей': len(stats['senders'])
                        })
                    
                    df_patterns = pd.DataFrame(pattern_data)
                    df_patterns = df_patterns.sort_values('Количество использований', ascending=False)
                    df_patterns.to_excel(writer, sheet_name='Статистика по форматам', index=False)
            
            print(f"✅ Полный отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    async def run_full_analysis(self):
        """Запуск полного анализа"""
        try:
            print("🚀 ЗАПУСК ПОЛНОГО АНАЛИЗА ЧАТА ЗА ВСЕ ВРЕМЯ")
            print("=" * 70)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение всех сообщений
            messages = await self.get_all_messages()
            
            if not messages:
                print("❌ Не найдено сообщений")
                return
            
            # Анализ сообщений
            all_extracted_data = []
            analyzed_messages = 0
            
            print("🔍 Анализ сообщений на наличие планов и фактов...")
            
            for msg in messages:
                extracted_data, patterns = self.analyze_message_for_plan_fact(
                    msg['text'], msg['date'], msg['sender'], msg['id']
                )
                
                if extracted_data:
                    all_extracted_data.extend(extracted_data)
                    analyzed_messages += 1
                
                # Прогресс
                if len(messages) > 1000 and analyzed_messages % 100 == 0:
                    print(f"📊 Проанализировано {analyzed_messages} сообщений, найдено {len(all_extracted_data)} записей...")
            
            print(f"📊 Проанализировано {analyzed_messages} сообщений из {len(messages)}")
            print(f"📊 Найдено {len(all_extracted_data)} записей планов/фактов")
            
            if not all_extracted_data:
                print("❌ Не найдено данных планов/фактов")
                return
            
            # Поиск дубликатов
            all_extracted_data = self.find_duplicates(all_extracted_data)
            duplicates_count = len([item for item in all_extracted_data if item['is_duplicate']])
            print(f"🔍 Найдено {duplicates_count} дублирующихся записей")
            
            # Анализ по периодам
            periods_analysis = self.analyze_by_periods(all_extracted_data)
            
            # Создание отчета
            report_file = self.create_full_report(all_extracted_data, periods_analysis)
            
            if report_file:
                print(f"\n🎉 Полный анализ завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Всего сообщений: {len(messages)}")
                print(f"Проанализировано: {analyzed_messages}")
                print(f"Найдено записей планов/фактов: {len(all_extracted_data)}")
                print(f"Дубликатов: {duplicates_count}")
                print(f"Уникальных руководителей: {len(set(item['sender'] for item in all_extracted_data))}")
                
                # Показываем статистику по годам
                if periods_analysis:
                    print(f"\n📅 СТАТИСТИКА ПО ГОДАМ:")
                    for year, data in sorted(periods_analysis.items(), reverse=True):
                        if data:
                            print(f"  {year}: {len(data)} записей")
                
                # Показываем топ руководителей
                sender_stats = {}
                for item in all_extracted_data:
                    sender = item['sender']
                    if sender not in sender_stats:
                        sender_stats[sender] = {'count': 0, 'total_completion': 0}
                    sender_stats[sender]['count'] += 1
                    sender_stats[sender]['total_completion'] += item['completion_percent']
                
                print(f"\n👥 ТОП РУКОВОДИТЕЛЕЙ:")
                for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['count'], reverse=True)[:5]:
                    avg_completion = stats['total_completion'] / stats['count']
                    print(f"  {sender}: {stats['count']} записей, средний % выполнения: {avg_completion:.1f}%")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = FullChatAnalyzer()
    await analyzer.run_full_analysis()

if __name__ == "__main__":
    asyncio.run(main())
