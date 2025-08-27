#!/usr/bin/env python3
"""
Precise Plan/Fact Analyzer for Main mega ТОП
Точный анализатор для поиска структурированных планов и фактов
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

class PrecisePlanFactAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/precise_plan_fact.log", rotation="1 day", retention="7 days")
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_structured_messages(self, weeks_back=52):
        """Получение структурированных сообщений"""
        try:
            print(f"📥 Загрузка сообщений за последние {weeks_back} недель...")
            
            today = datetime.now()
            start_date = today - timedelta(weeks=weeks_back)
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=10000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            # Фильтруем сообщения с явными разделами план/факт
            structured_messages = []
            
            for msg in messages:
                if msg.text and len(msg.text.strip()) > 100:  # Минимум 100 символов
                    text = msg.text.lower()
                    
                    # Ищем явные разделы план и факт
                    has_plan_section = re.search(r'план[:\s]*', text)
                    has_fact_section = re.search(r'факт[:\s]*', text)
                    
                    # Должны быть оба раздела
                    if has_plan_section and has_fact_section:
                        sender_name = await self.get_sender_name(msg)
                        structured_messages.append({
                            'id': msg.id,
                            'date': msg.date,
                            'text': msg.text.strip(),
                            'sender': sender_name,
                            'length': len(msg.text.strip()),
                            'weekday': msg.date.weekday(),
                            'weekday_name': ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'][msg.date.weekday()]
                        })
            
            print(f"📊 Найдено {len(structured_messages)} структурированных сообщений")
            return structured_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return []
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def parse_structured_message(self, text, message_date, sender_name, message_id):
        """Парсинг структурированного сообщения"""
        try:
            # Разбиваем на строки
            lines = text.split('\n')
            
            # Ищем разделы
            plan_section = []
            fact_section = []
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Определяем раздел
                if re.match(r'^план[:\s]*', line.lower()):
                    current_section = 'plan'
                    continue
                elif re.match(r'^факт[:\s]*', line.lower()):
                    current_section = 'fact'
                    continue
                
                # Добавляем в соответствующий раздел
                if current_section == 'plan':
                    plan_section.append(line)
                elif current_section == 'fact':
                    fact_section.append(line)
            
            # Извлекаем задачи из разделов
            plan_tasks = self.extract_tasks_from_section(plan_section, 'plan')
            fact_tasks = self.extract_tasks_from_section(fact_section, 'fact')
            
            # Сопоставляем планы и факты
            matched_tasks = self.match_plan_fact_tasks(plan_tasks, fact_tasks, message_date, sender_name, message_id)
            
            return matched_tasks
            
        except Exception as e:
            logger.error(f"❌ Ошибка парсинга сообщения: {e}")
            return []
    
    def extract_tasks_from_section(self, lines, section_type):
        """Извлечение задач из раздела"""
        tasks = []
        
        for line in lines:
            # Ищем нумерованные задачи
            numbered_match = re.match(r'^(\d+)[\.\)]\s*(.+)', line)
            if numbered_match:
                task_num = numbered_match.group(1)
                task_desc = numbered_match.group(2)
                
                # Ищем числа в задаче
                numbers = self.extract_numbers_from_task(task_desc)
                
                if numbers:
                    tasks.append({
                        'number': task_num,
                        'description': task_desc,
                        'numbers': numbers,
                        'section': section_type
                    })
                continue
            
            # Ищем задачи с маркерами
            marker_match = re.match(r'^[-•]\s*(.+)', line)
            if marker_match:
                task_desc = marker_match.group(1)
                numbers = self.extract_numbers_from_task(task_desc)
                
                if numbers:
                    tasks.append({
                        'number': len(tasks) + 1,
                        'description': task_desc,
                        'numbers': numbers,
                        'section': section_type
                    })
                continue
        
        return tasks
    
    def extract_numbers_from_task(self, task_text):
        """Извлечение чисел из задачи"""
        numbers = []
        
        # Паттерны для поиска чисел
        patterns = [
            r'(\d+(?:\.\d+)?)[/\-]\s*(\d+(?:\.\d+)?)',  # X/Y или X-Y
            r'\((\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)\)',  # (X Y)
            r'план[:\s]*(\d+(?:\.\d+)?)',  # план: X
            r'факт[:\s]*(\d+(?:\.\d+)?)',  # факт: X
            r'п/ф[:\s]*(\d+(?:\.\d+)?)',   # п/ф: X
            r'(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)',  # X / Y
        ]
        
        for pattern in patterns:
            matches = re.finditer(pattern, task_text.lower())
            for match in matches:
                try:
                    if len(match.groups()) == 2:
                        num1 = float(match.group(1))
                        num2 = float(match.group(2))
                        
                        # Фильтруем разумные значения
                        if 0 <= num1 <= 1000 and 0 <= num2 <= 1000:
                            numbers.append((num1, num2))
                    else:
                        num = float(match.group(1))
                        if 0 <= num <= 1000:
                            numbers.append((num, None))
                except ValueError:
                    continue
        
        return numbers
    
    def match_plan_fact_tasks(self, plan_tasks, fact_tasks, message_date, sender_name, message_id):
        """Сопоставление планов и фактов"""
        matched_tasks = []
        
        # Создаем словари для быстрого поиска
        plan_dict = {task['number']: task for task in plan_tasks}
        fact_dict = {task['number']: task for task in fact_tasks}
        
        # Сопоставляем по номерам
        all_numbers = set(plan_dict.keys()) | set(fact_dict.keys())
        
        for num in all_numbers:
            plan_task = plan_dict.get(num)
            fact_task = fact_dict.get(num)
            
            # Извлекаем планы и факты
            plan_value = None
            fact_value = None
            
            if plan_task and plan_task['numbers']:
                for nums in plan_task['numbers']:
                    if len(nums) == 2 and nums[1] is not None:
                        plan_value = nums[0]
                        break
                    elif len(nums) == 2 and nums[1] is None:
                        plan_value = nums[0]
                        break
            
            if fact_task and fact_task['numbers']:
                for nums in fact_task['numbers']:
                    if len(nums) == 2 and nums[1] is not None:
                        fact_value = nums[1]
                        break
                    elif len(nums) == 2 and nums[1] is None:
                        fact_value = nums[0]
                        break
            
            # Если нашли и план, и факт
            if plan_value is not None and fact_value is not None:
                completion_percent = (fact_value / plan_value * 100) if plan_value > 0 else 0
                
                # Фильтруем нереалистичные значения
                if 0 <= completion_percent <= 300:  # Максимум 300%
                    matched_tasks.append({
                        'message_id': message_id,
                        'date': message_date,
                        'sender': sender_name,
                        'task_number': num,
                        'task_description': plan_task['description'] if plan_task else fact_task['description'],
                        'plan': plan_value,
                        'fact': fact_value,
                        'difference': fact_value - plan_value,
                        'completion_percent': completion_percent,
                        'status': self.get_status(completion_percent),
                        'is_duplicate': False
                    })
        
        return matched_tasks
    
    def get_status(self, completion_percent):
        """Определение статуса выполнения"""
        if completion_percent >= 100:
            return "Перевыполнение"
        elif completion_percent >= 90:
            return "Хорошо"
        elif completion_percent >= 70:
            return "Удовлетворительно"
        else:
            return "Недовыполнение"
    
    def find_duplicates(self, all_tasks):
        """Поиск дублирующихся задач"""
        try:
            grouped_data = {}
            
            for task in all_tasks:
                key = (task['sender'], task['task_description'][:50], task['plan'], task['fact'])
                if key not in grouped_data:
                    grouped_data[key] = []
                grouped_data[key].append(task)
            
            # Помечаем дубликаты
            for key, tasks in grouped_data.items():
                if len(tasks) > 1:
                    for task in tasks:
                        task['is_duplicate'] = True
                        task['duplicate_count'] = len(tasks)
            
            return all_tasks
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска дубликатов: {e}")
            return all_tasks
    
    def create_precise_report(self, all_tasks, messages, filename=None):
        """Создание точного отчета"""
        try:
            if not filename:
                filename = f"output/precise_plan_fact_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # Лист 1: Все задачи
                if all_tasks:
                    df_tasks = pd.DataFrame(all_tasks)
                    df_tasks['date'] = pd.to_datetime(df_tasks['date']).dt.tz_localize(None)
                    df_tasks = df_tasks.sort_values(['sender', 'date'], ascending=[True, False])
                    
                    # Переименовываем колонки
                    df_tasks = df_tasks.rename(columns={
                        'message_id': 'ID сообщения',
                        'date': 'Дата',
                        'sender': 'Руководитель',
                        'task_number': 'Номер задачи',
                        'task_description': 'Описание задачи',
                        'plan': 'План',
                        'fact': 'Факт',
                        'difference': 'Разница',
                        'completion_percent': '% выполнения',
                        'status': 'Статус',
                        'is_duplicate': 'Дубликат',
                        'duplicate_count': 'Количество дубликатов'
                    })
                    
                    df_tasks.to_excel(writer, sheet_name='Все задачи', index=False)
                    
                    # Форматирование
                    worksheet = writer.sheets['Все задачи']
                    
                    # Подсветка дубликатов и недовыполнений
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    
                    for row in worksheet.iter_rows(min_row=2):
                        if row[df_tasks.columns.get_loc('Дубликат')].value:
                            for cell in row:
                                cell.fill = red_fill
                        elif row[df_tasks.columns.get_loc('% выполнения')].value and row[df_tasks.columns.get_loc('% выполнения')].value < 70:
                            for cell in row:
                                cell.fill = yellow_fill
                
                # Лист 2: Исходные сообщения
                if messages:
                    messages_data = []
                    for msg in messages:
                        messages_data.append({
                            'ID': msg['id'],
                            'Дата': msg['date'].strftime('%Y-%m-%d %H:%M'),
                            'День недели': msg['weekday_name'],
                            'Отправитель': msg['sender'],
                            'Текст': msg['text'],
                            'Длина': msg['length']
                        })
                    
                    df_messages = pd.DataFrame(messages_data)
                    df_messages = df_messages.sort_values(['Отправитель', 'Дата'], ascending=[True, False])
                    df_messages.to_excel(writer, sheet_name='Исходные сообщения', index=False)
                
                # Лист 3: Анализ по руководителям
                if all_tasks:
                    sender_analysis = {}
                    
                    for task in all_tasks:
                        sender = task['sender']
                        if sender not in sender_analysis:
                            sender_analysis[sender] = {
                                'total_plan': 0,
                                'total_fact': 0,
                                'tasks_count': 0,
                                'over_performance': 0,
                                'under_performance': 0,
                                'duplicates': 0,
                                'dates': set(),
                                'messages': set()
                            }
                        
                        sender_analysis[sender]['total_plan'] += task['plan']
                        sender_analysis[sender]['total_fact'] += task['fact']
                        sender_analysis[sender]['tasks_count'] += 1
                        sender_analysis[sender]['dates'].add(task['date'].strftime('%Y-%m-%d'))
                        sender_analysis[sender]['messages'].add(task['message_id'])
                        
                        if task['completion_percent'] >= 100:
                            sender_analysis[sender]['over_performance'] += 1
                        else:
                            sender_analysis[sender]['under_performance'] += 1
                        
                        if task['is_duplicate']:
                            sender_analysis[sender]['duplicates'] += 1
                    
                    # Создаем данные для отчета
                    sender_data = []
                    for sender, data in sender_analysis.items():
                        avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                        sender_data.append({
                            'Руководитель': sender,
                            'Общий план': round(data['total_plan'], 2),
                            'Общий факт': round(data['total_fact'], 2),
                            'Средний % выполнения': round(avg_completion, 2),
                            'Количество задач': data['tasks_count'],
                            'Перевыполнений': data['over_performance'],
                            'Недовыполнений': data['under_performance'],
                            'Дубликатов': data['duplicates'],
                            'Количество дат': len(data['dates']),
                            'Количество сообщений': len(data['messages'])
                        })
                    
                    df_senders = pd.DataFrame(sender_data)
                    df_senders = df_senders.sort_values('Средний % выполнения', ascending=False)
                    df_senders.to_excel(writer, sheet_name='Анализ по руководителям', index=False)
            
            print(f"✅ Точный отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    async def run_precise_analysis(self):
        """Запуск точного анализа"""
        try:
            print("🚀 ЗАПУСК ТОЧНОГО АНАЛИЗА ПЛАНОВ И ФАКТОВ")
            print("=" * 55)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение структурированных сообщений
            messages = await self.get_structured_messages(weeks_back=52)
            
            if not messages:
                print("❌ Не найдено структурированных сообщений")
                return
            
            # Парсинг сообщений
            all_tasks = []
            
            print("🔍 Парсинг структурированных сообщений...")
            
            for msg in messages:
                tasks = self.parse_structured_message(msg['text'], msg['date'], msg['sender'], msg['id'])
                if tasks:
                    all_tasks.extend(tasks)
                    print(f"📝 Найдено {len(tasks)} задач в сообщении от {msg['sender']} ({msg['date'].strftime('%Y-%m-%d')})")
            
            print(f"📊 Всего найдено {len(all_tasks)} задач")
            
            if not all_tasks:
                print("❌ Не найдено задач с планами и фактами")
                return
            
            # Поиск дубликатов
            all_tasks = self.find_duplicates(all_tasks)
            duplicates_count = len([task for task in all_tasks if task['is_duplicate']])
            print(f"🔍 Найдено {duplicates_count} дублирующихся задач")
            
            # Создание отчета
            report_file = self.create_precise_report(all_tasks, messages)
            
            if report_file:
                print(f"\n🎉 Точный анализ завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Найдено структурированных сообщений: {len(messages)}")
                print(f"Найдено задач: {len(all_tasks)}")
                print(f"Дубликатов: {duplicates_count}")
                print(f"Уникальных руководителей: {len(set(task['sender'] for task in all_tasks))}")
                
                # Показываем статистику по руководителям
                sender_stats = {}
                for task in all_tasks:
                    sender = task['sender']
                    if sender not in sender_stats:
                        sender_stats[sender] = {'count': 0, 'total_completion': 0}
                    sender_stats[sender]['count'] += 1
                    sender_stats[sender]['total_completion'] += task['completion_percent']
                
                print(f"\n👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ:")
                for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['count'], reverse=True):
                    avg_completion = stats['total_completion'] / stats['count']
                    print(f"  {sender}: {stats['count']} задач, средний % выполнения: {avg_completion:.1f}%")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = PrecisePlanFactAnalyzer()
    await analyzer.run_precise_analysis()

if __name__ == "__main__":
    asyncio.run(main())
