#!/usr/bin/env python3
"""
Manager Individual Reports Analyzer for Main mega ТОП
Анализатор с отдельными отчетами для каждого руководителя за последние 2 месяца
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class ManagerIndividualReportsAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/manager_reports.log", rotation="1 day", retention="7 days")
        
        # Цветовая схема
        self.colors = {
            'green': '28A745',      # Выполнено
            'yellow': 'FFC107',     # Частично выполнено
            'red': 'DC3545',        # Не выполнено
            'blue': '007BFF',       # Заголовки
            'orange': 'FD7E14',     # Новые задачи
            'purple': '6F42C1',     # Акценты
            'light_gray': 'F8F9FA', # Фон
            'dark_gray': '6C757D'   # Текст
        }
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_recent_messages(self, months_back=2):
        """Получение сообщений за последние 2 месяца"""
        try:
            print(f"📥 Загрузка сообщений за последние {months_back} месяца...")
            
            today = datetime.now()
            start_date = today - timedelta(days=months_back * 30)
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=10000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            # Группируем сообщения по неделям
            weekly_messages = {}
            
            for msg in messages:
                if msg.text and len(msg.text.strip()) > 50:
                    # Определяем неделю
                    msg_date = msg.date.replace(tzinfo=None)
                    week_start = msg_date - timedelta(days=msg_date.weekday())
                    week_key = week_start.strftime('%Y-%m-%d')
                    
                    if week_key not in weekly_messages:
                        weekly_messages[week_key] = []
                    
                    sender_name = await self.get_sender_name(msg)
                    weekly_messages[week_key].append({
                        'id': msg.id,
                        'date': msg_date,
                        'text': msg.text.strip(),
                        'sender': sender_name,
                        'week_start': week_start,
                        'week_key': week_key
                    })
            
            print(f"📊 Сгруппировано по {len(weekly_messages)} неделям")
            return weekly_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return {}
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def extract_plan_fact_sections(self, text, sender, message_date):
        """Извлечение разделов план и факт из текста"""
        try:
            # Разбиваем на строки
            lines = text.split('\n')
            
            plan_section = []
            fact_section = []
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Определяем раздел
                if re.match(r'^план[:\s]*', line.lower()):
                    current_section = 'plan'
                    continue
                elif re.match(r'^факт[:\s]*', line.lower()):
                    current_section = 'fact'
                    continue
                
                # Добавляем в соответствующий раздел
                if current_section == 'plan':
                    plan_section.append(line)
                elif current_section == 'fact':
                    fact_section.append(line)
            
            return {
                'sender': sender,
                'date': message_date,
                'plan_tasks': plan_section,
                'fact_tasks': fact_section,
                'has_plan': len(plan_section) > 0,
                'has_fact': len(fact_section) > 0
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка извлечения разделов: {e}")
            return None
    
    def analyze_manager_data(self, weekly_messages):
        """Анализ данных по руководителям"""
        try:
            print("🔍 Анализ данных по руководителям...")
            
            manager_data = {}
            week_keys = sorted(weekly_messages.keys())
            
            # Проходим по неделям (кроме последней)
            for i in range(len(week_keys) - 1):
                current_week = week_keys[i]
                next_week = week_keys[i + 1]
                
                current_week_messages = weekly_messages[current_week]
                next_week_messages = weekly_messages[next_week]
                
                # Анализируем планы текущей недели
                current_week_plans = {}
                for msg in current_week_messages:
                    sections = self.extract_plan_fact_sections(msg['text'], msg['sender'], msg['date'])
                    if sections and sections['has_plan']:
                        if sections['sender'] not in current_week_plans:
                            current_week_plans[sections['sender']] = []
                        current_week_plans[sections['sender']].extend(sections['plan_tasks'])
                
                # Анализируем факты следующей недели
                next_week_facts = {}
                for msg in next_week_messages:
                    sections = self.extract_plan_fact_sections(msg['text'], msg['sender'], msg['date'])
                    if sections and sections['has_fact']:
                        if sections['sender'] not in next_week_facts:
                            next_week_facts[sections['sender']] = []
                        next_week_facts[sections['sender']].extend(sections['fact_tasks'])
                
                # Сравниваем для каждого руководителя
                for sender in set(current_week_plans.keys()) | set(next_week_facts.keys()):
                    plan_tasks = current_week_plans.get(sender, [])
                    fact_tasks = next_week_facts.get(sender, [])
                    
                    if plan_tasks or fact_tasks:
                        # Инициализируем данные руководителя
                        if sender not in manager_data:
                            manager_data[sender] = {
                                'weeks': [],
                                'total_planned': 0,
                                'total_facts': 0,
                                'total_matched': 0,
                                'total_unmatched_plans': 0,
                                'total_new_tasks': 0,
                                'all_plan_tasks': [],
                                'all_fact_tasks': [],
                                'all_matches': []
                            }
                        
                        # Создаем детальное сопоставление
                        comparison = self.create_detailed_task_comparison(
                            plan_tasks, fact_tasks, sender, current_week, next_week
                        )
                        
                        if comparison:
                            manager_data[sender]['weeks'].append(comparison)
                            manager_data[sender]['total_planned'] += comparison['total_planned']
                            manager_data[sender]['total_facts'] += comparison['total_facts']
                            manager_data[sender]['total_matched'] += comparison['total_matched']
                            manager_data[sender]['total_unmatched_plans'] += comparison['total_unmatched_plans']
                            manager_data[sender]['total_new_tasks'] += comparison['total_new_tasks']
                            manager_data[sender]['all_plan_tasks'].extend(plan_tasks)
                            manager_data[sender]['all_fact_tasks'].extend(fact_tasks)
                            manager_data[sender]['all_matches'].extend(comparison['matched_tasks'])
            
            print(f"📊 Проанализировано {len(manager_data)} руководителей")
            return manager_data
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа данных: {e}")
            return {}
    
    def create_detailed_task_comparison(self, plan_tasks, fact_tasks, sender, plan_week, fact_week):
        """Создание детального сопоставления задач"""
        try:
            # Сопоставляем задачи
            matched_tasks = []
            unmatched_plans = []
            unmatched_facts = []
            
            # Создаем копии для работы
            remaining_facts = fact_tasks.copy()
            
            # Ищем соответствия для каждой задачи из плана
            for plan_task in plan_tasks:
                best_match = None
                best_similarity = 0
                best_fact_index = -1
                
                # Ищем лучшее соответствие среди фактов
                for i, fact_task in enumerate(remaining_facts):
                    similarity = self.calculate_task_similarity(plan_task, fact_task)
                    
                    if similarity > best_similarity and similarity >= 0.3:  # Минимальный порог
                        best_similarity = similarity
                        best_match = fact_task
                        best_fact_index = i
                
                if best_match:
                    # Найдено соответствие
                    matched_tasks.append({
                        'plan_task': plan_task,
                        'fact_task': best_match,
                        'similarity': best_similarity,
                        'status': self.get_match_status(best_similarity)
                    })
                    # Удаляем использованный факт
                    remaining_facts.pop(best_fact_index)
                else:
                    # Не найдено соответствие
                    unmatched_plans.append(plan_task)
            
            # Оставшиеся факты - новые задачи
            unmatched_facts = remaining_facts
            
            # Вычисляем статистику
            total_planned = len(plan_tasks)
            total_facts = len(fact_tasks)
            total_matched = len(matched_tasks)
            total_unmatched_plans = len(unmatched_plans)
            total_new_tasks = len(unmatched_facts)
            
            # Процент выполнения (только по сопоставленным задачам)
            completion_rate = (total_matched / total_planned * 100) if total_planned > 0 else 0
            
            return {
                'sender': sender,
                'plan_week': plan_week,
                'fact_week': fact_week,
                'plan_tasks': plan_tasks,
                'fact_tasks': fact_tasks,
                'matched_tasks': matched_tasks,
                'unmatched_plans': unmatched_plans,
                'unmatched_facts': unmatched_facts,
                'total_planned': total_planned,
                'total_facts': total_facts,
                'total_matched': total_matched,
                'total_unmatched_plans': total_unmatched_plans,
                'total_new_tasks': total_new_tasks,
                'completion_rate': completion_rate,
                'overall_status': self.get_overall_status(completion_rate, total_matched, total_planned)
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания детального сопоставления: {e}")
            return None
    
    def calculate_task_similarity(self, task1, task2):
        """Вычисление сходства между задачами"""
        try:
            # Приводим к нижнему регистру
            t1 = task1.lower()
            t2 = task2.lower()
            
            # Убираем цифры и специальные символы
            t1_clean = re.sub(r'[^\w\s]', '', t1)
            t2_clean = re.sub(r'[^\w\s]', '', t2)
            
            # Разбиваем на слова
            words1 = set(t1_clean.split())
            words2 = set(t2_clean.split())
            
            if not words1 or not words2:
                return 0.0
            
            # Вычисляем Jaccard similarity
            intersection = len(words1.intersection(words2))
            union = len(words1.union(words2))
            
            return intersection / union if union > 0 else 0.0
            
        except Exception as e:
            return 0.0
    
    def get_match_status(self, similarity):
        """Определение статуса сопоставления"""
        if similarity >= 0.7:
            return "✅ Выполнено"
        elif similarity >= 0.5:
            return "⚠️ Частично выполнено"
        elif similarity >= 0.3:
            return "🔄 Изменено"
        else:
            return "❌ Не выполнено"
    
    def get_overall_status(self, completion_rate, matched, planned):
        """Определение общего статуса"""
        if completion_rate >= 80 and matched > 0:
            return "✅ Отлично"
        elif completion_rate >= 60:
            return "⚠️ Хорошо"
        elif completion_rate >= 40:
            return "⚠️ Средне"
        elif planned > 0:
            return "❌ Плохо"
        else:
            return "📝 Только новые задачи"
    
    def create_manager_reports(self, manager_data, filename=None):
        """Создание отчетов по руководителям"""
        try:
            if not filename:
                filename = f"output/manager_individual_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            # Удаляем дефолтный лист
            workbook.remove(workbook.active)
            
            # Создаем общую сводку
            self.create_summary_sheet(workbook, manager_data)
            
            # Создаем отдельные листы для каждого руководителя
            for sender, data in manager_data.items():
                if data['weeks']:  # Только если есть данные
                    self.create_manager_sheet(workbook, sender, data)
            
            workbook.save(filename)
            print(f"✅ Отчеты по руководителям созданы: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчетов: {e}")
            return None
    
    def create_summary_sheet(self, workbook, manager_data):
        """Создание сводного листа"""
        try:
            ws = workbook.create_sheet("Общая сводка")
            
            # Заголовок
            ws['A1'] = "ОТЧЕТЫ ПО РУКОВОДИТЕЛЯМ - ПОСЛЕДНИЕ 2 МЕСЯЦА"
            ws['A2'] = "Main mega ТОП - Индивидуальные отчеты"
            
            # Общая статистика
            total_managers = len(manager_data)
            total_planned = sum(data['total_planned'] for data in manager_data.values())
            total_facts = sum(data['total_facts'] for data in manager_data.values())
            total_matched = sum(data['total_matched'] for data in manager_data.values())
            total_unmatched_plans = sum(data['total_unmatched_plans'] for data in manager_data.values())
            total_new_tasks = sum(data['total_new_tasks'] for data in manager_data.values())
            
            avg_completion = (total_matched / total_planned * 100) if total_planned > 0 else 0
            
            ws['A4'] = "📊 ОБЩАЯ СТАТИСТИКА:"
            ws['A5'] = f"Руководителей с данными: {total_managers}"
            ws['A6'] = f"Всего запланировано: {total_planned}"
            ws['A7'] = f"Всего выполнено (факт): {total_facts}"
            ws['A8'] = f"Сопоставлено задач: {total_matched}"
            ws['A9'] = f"Не выполнено (план): {total_unmatched_plans}"
            ws['A10'] = f"Новых задач (факт): {total_new_tasks}"
            ws['A11'] = f"Средний % сопоставления: {avg_completion:.1f}%"
            
            # Статистика по руководителям
            ws['A13'] = "👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ:"
            
            headers = ['Руководитель', 'Недель', 'Запланировано', 'Факт', 'Сопоставлено', 'Не выполнено', 'Новых', '% сопоставления', 'Статус']
            for col, header in enumerate(headers, 1):
                ws.cell(row=14, column=col, value=header)
            
            row = 15
            for sender, data in sorted(manager_data.items(), key=lambda x: x[1]['total_matched'], reverse=True):
                weeks_count = len(data['weeks'])
                completion_rate = (data['total_matched'] / data['total_planned'] * 100) if data['total_planned'] > 0 else 0
                status = self.get_overall_status(completion_rate, data['total_matched'], data['total_planned'])
                
                ws.cell(row=row, column=1, value=sender)
                ws.cell(row=row, column=2, value=weeks_count)
                ws.cell(row=row, column=3, value=data['total_planned'])
                ws.cell(row=row, column=4, value=data['total_facts'])
                ws.cell(row=row, column=5, value=data['total_matched'])
                ws.cell(row=row, column=6, value=data['total_unmatched_plans'])
                ws.cell(row=row, column=7, value=data['total_new_tasks'])
                ws.cell(row=row, column=8, value=f"{completion_rate:.1f}%")
                ws.cell(row=row, column=9, value=status)
                
                row += 1
            
            # Форматирование
            self.format_summary_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания сводного листа: {e}")
    
    def create_manager_sheet(self, workbook, sender, data):
        """Создание листа для конкретного руководителя"""
        try:
            # Очищаем имя для использования в названии листа
            sheet_name = self.clean_sheet_name(sender)
            ws = workbook.create_sheet(sheet_name)
            
            # Заголовок руководителя
            ws['A1'] = f"ОТЧЕТ ПО РУКОВОДИТЕЛЮ: {sender}"
            ws['A2'] = f"Период: Последние 2 месяца"
            
            # Общая статистика руководителя
            completion_rate = (data['total_matched'] / data['total_planned'] * 100) if data['total_planned'] > 0 else 0
            status = self.get_overall_status(completion_rate, data['total_matched'], data['total_planned'])
            
            ws['A4'] = "📊 СТАТИСТИКА РУКОВОДИТЕЛЯ:"
            ws['A5'] = f"Недель с данными: {len(data['weeks'])}"
            ws['A6'] = f"Всего запланировано: {data['total_planned']}"
            ws['A7'] = f"Всего выполнено (факт): {data['total_facts']}"
            ws['A8'] = f"Сопоставлено задач: {data['total_matched']}"
            ws['A9'] = f"Не выполнено (план): {data['total_unmatched_plans']}"
            ws['A10'] = f"Новых задач (факт): {data['total_new_tasks']}"
            ws['A11'] = f"Процент сопоставления: {completion_rate:.1f}%"
            ws['A12'] = f"Общий статус: {status}"
            
            # Детализация по неделям
            ws['A14'] = "📅 ДЕТАЛИЗАЦИЯ ПО НЕДЕЛЯМ:"
            
            headers = ['Неделя плана', 'Неделя факта', 'Запланировано', 'Факт', 'Сопоставлено', 'Не выполнено', 'Новых', '% сопоставления', 'Статус']
            for col, header in enumerate(headers, 1):
                ws.cell(row=15, column=col, value=header)
            
            row = 16
            for week_data in data['weeks']:
                ws.cell(row=row, column=1, value=week_data['plan_week'])
                ws.cell(row=row, column=2, value=week_data['fact_week'])
                ws.cell(row=row, column=3, value=week_data['total_planned'])
                ws.cell(row=row, column=4, value=week_data['total_facts'])
                ws.cell(row=row, column=5, value=week_data['total_matched'])
                ws.cell(row=row, column=6, value=week_data['total_unmatched_plans'])
                ws.cell(row=row, column=7, value=week_data['total_new_tasks'])
                ws.cell(row=row, column=8, value=f"{week_data['completion_rate']:.1f}%")
                ws.cell(row=row, column=9, value=week_data['overall_status'])
                
                row += 1
            
            # Сопоставленные задачи
            if data['all_matches']:
                ws['A' + str(row + 2)] = "✅ СОПОСТАВЛЕННЫЕ ЗАДАЧИ:"
                
                match_headers = ['Неделя плана', 'Неделя факта', 'Задача из плана', 'Задача из факта', 'Сходство', 'Статус']
                for col, header in enumerate(match_headers, 1):
                    ws.cell(row=row + 3, column=col, value=header)
                
                match_row = row + 4
                for match in data['all_matches'][:20]:  # Показываем первые 20
                    # Находим неделю для этого матча
                    week_info = self.find_week_for_match(match, data['weeks'])
                    
                    ws.cell(row=match_row, column=1, value=week_info['plan_week'])
                    ws.cell(row=match_row, column=2, value=week_info['fact_week'])
                    ws.cell(row=match_row, column=3, value=match['plan_task'])
                    ws.cell(row=match_row, column=4, value=match['fact_task'])
                    ws.cell(row=match_row, column=5, value=f"{match['similarity']:.2f}")
                    ws.cell(row=match_row, column=6, value=match['status'])
                    
                    # Форматирование ячеек с текстом
                    for col in [3, 4]:
                        cell = ws.cell(row=match_row, column=col)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        ws.row_dimensions[match_row].height = 60
                    
                    match_row += 1
                
                if len(data['all_matches']) > 20:
                    ws['A' + str(match_row)] = f"... и еще {len(data['all_matches']) - 20} сопоставленных задач"
            
            # Форматирование
            self.format_manager_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа руководителя {sender}: {e}")
    
    def find_week_for_match(self, match, weeks):
        """Найти неделю для конкретного матча"""
        for week in weeks:
            for week_match in week['matched_tasks']:
                if (week_match['plan_task'] == match['plan_task'] and 
                    week_match['fact_task'] == match['fact_task']):
                    return {
                        'plan_week': week['plan_week'],
                        'fact_week': week['fact_week']
                    }
        return {'plan_week': 'N/A', 'fact_week': 'N/A'}
    
    def clean_sheet_name(self, name):
        """Очистка имени для использования в названии листа"""
        # Убираем недопустимые символы
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '')
        
        # Ограничиваем длину
        if len(name) > 31:
            name = name[:28] + "..."
        
        return name
    
    def format_summary_sheet(self, ws):
        """Форматирование сводного листа"""
        # Заголовок
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['blue'])
        ws['A2'].font = Font(size=14, bold=True)
        
        # Объединение ячеек
        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:I2')
        
        # Ширина колонок
        for col in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def format_manager_sheet(self, ws):
        """Форматирование листа руководителя"""
        # Заголовок
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['purple'])
        ws['A2'].font = Font(size=14, bold=True)
        
        # Объединение ячеек
        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:I2')
        
        # Ширина колонок
        for col in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    async def run_manager_reports(self):
        """Запуск создания отчетов по руководителям"""
        try:
            print("🚀 ЗАПУСК СОЗДАНИЯ ОТЧЕТОВ ПО РУКОВОДИТЕЛЯМ")
            print("=" * 65)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение сообщений за последние 2 месяца
            weekly_messages = await self.get_recent_messages(months_back=2)
            
            if not weekly_messages:
                print("❌ Не найдено сообщений за последние 2 месяца")
                return
            
            # Анализ данных по руководителям
            manager_data = self.analyze_manager_data(weekly_messages)
            
            if not manager_data:
                print("❌ Не найдено данных по руководителям")
                return
            
            # Создание отчетов
            report_file = self.create_manager_reports(manager_data)
            
            if report_file:
                print(f"\n🎉 Отчеты по руководителям созданы!")
                print(f"📁 Файл сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Проанализировано недель: {len(weekly_messages)}")
                print(f"Руководителей с данными: {len(manager_data)}")
                
                # Статистика по руководителям
                total_planned = sum(data['total_planned'] for data in manager_data.values())
                total_facts = sum(data['total_facts'] for data in manager_data.values())
                total_matched = sum(data['total_matched'] for data in manager_data.values())
                
                print(f"Всего запланировано: {total_planned}")
                print(f"Всего выполнено (факт): {total_facts}")
                print(f"Сопоставлено задач: {total_matched}")
                
                avg_completion = (total_matched / total_planned * 100) if total_planned > 0 else 0
                print(f"Средний % сопоставления: {avg_completion:.1f}%")
                
                # Показываем руководителей
                print(f"\n👥 РУКОВОДИТЕЛИ С ДАННЫМИ:")
                for sender, data in sorted(manager_data.items(), key=lambda x: x[1]['total_matched'], reverse=True):
                    completion_rate = (data['total_matched'] / data['total_planned'] * 100) if data['total_planned'] > 0 else 0
                    weeks_count = len(data['weeks'])
                    print(f"  {sender}: {data['total_matched']}/{data['total_planned']} ({completion_rate:.1f}%) - {weeks_count} недель")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = ManagerIndividualReportsAnalyzer()
    await analyzer.run_manager_reports()

if __name__ == "__main__":
    asyncio.run(main())
