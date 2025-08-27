#!/usr/bin/env python3
"""
Detailed Plan/Fact Analyzer for Main mega ТОП
Детальный анализатор с разбором отдельных пунктов
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class DetailedPlanFactAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/detailed_plan_fact.log", rotation="1 day", retention="7 days")
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    def extract_detailed_plan_fact_data(self, message_text, message_date, sender_name, message_id):
        """Детальное извлечение данных с разбором отдельных пунктов"""
        try:
            text = message_text.strip()
            extracted_data = []
            
            # Разбиваем текст на строки
            lines = text.split('\n')
            
            # Паттерны для поиска планов и фактов
            patterns = [
                # Паттерн 1: План: X, Факт: Y
                (r'план[:\s]*(\d+(?:\.\d+)?)[^\d]*факт[:\s]*(\d+(?:\.\d+)?)', 'План: X, Факт: Y'),
                # Паттерн 2: П/Ф: X/Y
                (r'п/ф[:\s]*(\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)', 'П/Ф: X/Y'),
                # Паттерн 3: План X Факт Y
                (r'план\s+(\d+(?:\.\d+)?)[^\d]*факт\s+(\d+(?:\.\d+)?)', 'План X Факт Y'),
                # Паттерн 4: Просто числа через слеш или дефис
                (r'(\d+(?:\.\d+)?)[/\-]\s*(\d+(?:\.\d+)?)', 'X/Y или X-Y'),
                # Паттерн 5: П: X Ф: Y
                (r'п[:\s]*(\d+(?:\.\d+)?)[^\d]*ф[:\s]*(\d+(?:\.\d+)?)', 'П: X Ф: Y'),
                # Паттерн 6: План - X, Факт - Y
                (r'план\s*[-\s]*(\d+(?:\.\d+)?)[^\d]*факт\s*[-\s]*(\d+(?:\.\d+)?)', 'План - X, Факт - Y'),
                # Паттерн 7: Числа в скобках
                (r'\((\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)\)', '(X Y)'),
                # Паттерн 8: План/Факт: X/Y
                (r'план/факт[:\s]*(\d+(?:\.\d+)?)[^\d]*(\d+(?:\.\d+)?)', 'План/Факт: X/Y'),
            ]
            
            # Ищем задачи и описания
            task_patterns = [
                r'(\d+)[\.\)]\s*(.+)',  # 1. Описание задачи
                r'[-•]\s*(.+)',         # - Описание задачи
                r'(\w+)[:\s]*(.+)',     # Задача: описание
            ]
            
            # Извлекаем все найденные планы/факты
            plan_fact_items = []
            for pattern, pattern_name in patterns:
                matches = re.finditer(pattern, text.lower())
                for match in matches:
                    try:
                        plan_value = float(match.group(1))
                        fact_value = float(match.group(2))
                        
                        if 0 <= plan_value <= 1000000 and 0 <= fact_value <= 1000000:
                            plan_fact_items.append({
                                'plan': plan_value,
                                'fact': fact_value,
                                'pattern': pattern_name,
                                'match_text': match.group(0)
                            })
                    except (ValueError, ZeroDivisionError):
                        continue
            
            # Ищем описания задач
            task_descriptions = []
            for pattern in task_patterns:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                for match in matches:
                    task_descriptions.append(match.group(0).strip())
            
            # Если нашли планы/факты, создаем записи
            for i, item in enumerate(plan_fact_items):
                # Ищем описание для этого пункта
                task_description = ""
                if i < len(task_descriptions):
                    task_description = task_descriptions[i]
                elif task_descriptions:
                    task_description = task_descriptions[0]  # Берем первое описание
                
                completion_percent = (item['fact'] / item['plan'] * 100) if item['plan'] > 0 else 0
                
                extracted_data.append({
                    'message_id': message_id,
                    'date': message_date,
                    'sender': sender_name,
                    'task_description': task_description,
                    'plan': item['plan'],
                    'fact': item['fact'],
                    'difference': item['fact'] - item['plan'],
                    'completion_percent': completion_percent,
                    'pattern_used': item['pattern'],
                    'match_text': item['match_text'],
                    'original_text': text[:500],
                    'status': self.get_status(completion_percent),
                    'is_duplicate': False  # Будет заполнено позже
                })
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"❌ Ошибка извлечения данных: {e}")
            return []
    
    def get_status(self, completion_percent):
        """Определение статуса выполнения"""
        if completion_percent >= 100:
            return "Перевыполнение"
        elif completion_percent >= 90:
            return "Хорошо"
        elif completion_percent >= 70:
            return "Удовлетворительно"
        else:
            return "Недовыполнение"
    
    def find_duplicates(self, all_data):
        """Поиск дублирующихся записей"""
        try:
            # Группируем по отправителю и значениям плана/факта
            grouped_data = {}
            
            for item in all_data:
                key = (item['sender'], item['plan'], item['fact'])
                if key not in grouped_data:
                    grouped_data[key] = []
                grouped_data[key].append(item)
            
            # Помечаем дубликаты
            for key, items in grouped_data.items():
                if len(items) > 1:
                    for item in items:
                        item['is_duplicate'] = True
                        item['duplicate_count'] = len(items)
            
            return all_data
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска дубликатов: {e}")
            return all_data
    
    async def get_weekly_messages(self, weeks_back=6):
        """Получение сообщений за последние недели"""
        try:
            print(f"📅 Получение сообщений за последние {weeks_back} недели...")
            
            today = datetime.now()
            days_since_tuesday = (today.weekday() - 1) % 7
            start_date = today - timedelta(days=days_since_tuesday + (weeks_back * 7))
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=3000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            tuesday_messages = []
            for msg in messages:
                if msg.text and msg.date.weekday() == 1:  # 1 = вторник
                    sender_name = await self.get_sender_name(msg)
                    tuesday_messages.append({
                        'id': msg.id,
                        'date': msg.date,
                        'text': msg.text,
                        'sender': sender_name
                    })
            
            print(f"📊 Найдено {len(tuesday_messages)} сообщений по вторникам")
            return tuesday_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return []
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def create_detailed_excel_report(self, all_data, filename=None):
        """Создание детального Excel отчета"""
        try:
            if not filename:
                filename = f"output/detailed_plan_fact_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # Лист 1: Все данные с дубликатами
                if all_data:
                    df_all = pd.DataFrame(all_data)
                    df_all['date'] = pd.to_datetime(df_all['date']).dt.tz_localize(None)
                    df_all = df_all.sort_values(['sender', 'date'], ascending=[True, False])
                    
                    # Переименовываем колонки
                    df_all = df_all.rename(columns={
                        'message_id': 'ID сообщения',
                        'date': 'Дата',
                        'sender': 'Руководитель',
                        'task_description': 'Описание задачи',
                        'plan': 'План',
                        'fact': 'Факт',
                        'difference': 'Разница',
                        'completion_percent': '% выполнения',
                        'pattern_used': 'Формат',
                        'match_text': 'Найденный текст',
                        'original_text': 'Исходный текст',
                        'status': 'Статус',
                        'is_duplicate': 'Дубликат',
                        'duplicate_count': 'Количество дубликатов'
                    })
                    
                    df_all.to_excel(writer, sheet_name='Все данные', index=False)
                    
                    # Форматирование с подсветкой дубликатов
                    worksheet = writer.sheets['Все данные']
                    
                    # Подсветка дубликатов
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    
                    for row in worksheet.iter_rows(min_row=2):
                        if row[df_all.columns.get_loc('Дубликат')].value:
                            for cell in row:
                                cell.fill = red_fill
                        elif row[df_all.columns.get_loc('% выполнения')].value and row[df_all.columns.get_loc('% выполнения')].value < 70:
                            for cell in row:
                                cell.fill = yellow_fill
                    
                    # Настройка ширины колонок
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Лист 2: Анализ по руководителям
                if all_data:
                    sender_analysis = {}
                    
                    for item in all_data:
                        sender = item['sender']
                        if sender not in sender_analysis:
                            sender_analysis[sender] = {
                                'total_plan': 0,
                                'total_fact': 0,
                                'tasks_count': 0,
                                'over_performance': 0,
                                'under_performance': 0,
                                'duplicates': 0,
                                'dates': set(),
                                'tasks': []
                            }
                        
                        sender_analysis[sender]['total_plan'] += item['plan']
                        sender_analysis[sender]['total_fact'] += item['fact']
                        sender_analysis[sender]['tasks_count'] += 1
                        sender_analysis[sender]['dates'].add(item['date'].strftime('%Y-%m-%d'))
                        
                        if item['completion_percent'] >= 100:
                            sender_analysis[sender]['over_performance'] += 1
                        else:
                            sender_analysis[sender]['under_performance'] += 1
                        
                        if item['is_duplicate']:
                            sender_analysis[sender]['duplicates'] += 1
                        
                        sender_analysis[sender]['tasks'].append({
                            'date': item['date'].strftime('%Y-%m-%d'),
                            'task': item['task_description'],
                            'plan': item['plan'],
                            'fact': item['fact'],
                            'completion': item['completion_percent'],
                            'status': item['status']
                        })
                    
                    # Создаем данные для отчета
                    sender_data = []
                    for sender, data in sender_analysis.items():
                        avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                        sender_data.append({
                            'Руководитель': sender,
                            'Общий план': round(data['total_plan'], 2),
                            'Общий факт': round(data['total_fact'], 2),
                            'Средний % выполнения': round(avg_completion, 2),
                            'Количество задач': data['tasks_count'],
                            'Перевыполнений': data['over_performance'],
                            'Недовыполнений': data['under_performance'],
                            'Дубликатов': data['duplicates'],
                            'Количество дат': len(data['dates']),
                            'Даты активности': ', '.join(sorted(data['dates']))
                        })
                    
                    df_senders = pd.DataFrame(sender_data)
                    df_senders = df_senders.sort_values('Средний % выполнения', ascending=False)
                    df_senders.to_excel(writer, sheet_name='Анализ по руководителям', index=False)
                
                # Лист 3: Детализация по руководителям
                if all_data:
                    # Создаем отдельные листы для каждого руководителя
                    sender_groups = {}
                    for item in all_data:
                        sender = item['sender']
                        if sender not in sender_groups:
                            sender_groups[sender] = []
                        sender_groups[sender].append(item)
                    
                    # Создаем лист с детализацией
                    detailed_data = []
                    for sender, items in sender_groups.items():
                        for item in items:
                            detailed_data.append({
                                'Руководитель': sender,
                                'Дата': item['date'].strftime('%Y-%m-%d'),
                                'Описание задачи': item['task_description'],
                                'План': item['plan'],
                                'Факт': item['fact'],
                                '% выполнения': round(item['completion_percent'], 2),
                                'Статус': item['status'],
                                'Дубликат': 'Да' if item['is_duplicate'] else 'Нет',
                                'Формат данных': item['pattern_used']
                            })
                    
                    df_detailed = pd.DataFrame(detailed_data)
                    df_detailed = df_detailed.sort_values(['Руководитель', 'Дата'], ascending=[True, False])
                    df_detailed.to_excel(writer, sheet_name='Детализация по руководителям', index=False)
                
                # Лист 4: Дубликаты
                duplicates_data = [item for item in all_data if item['is_duplicate']]
                if duplicates_data:
                    df_duplicates = pd.DataFrame(duplicates_data)
                    df_duplicates['date'] = pd.to_datetime(df_duplicates['date']).dt.tz_localize(None)
                    df_duplicates = df_duplicates.sort_values(['sender', 'date'], ascending=[True, False])
                    
                    df_duplicates = df_duplicates.rename(columns={
                        'date': 'Дата',
                        'sender': 'Руководитель',
                        'task_description': 'Описание задачи',
                        'plan': 'План',
                        'fact': 'Факт',
                        'completion_percent': '% выполнения',
                        'duplicate_count': 'Количество дубликатов'
                    })
                    
                    df_duplicates.to_excel(writer, sheet_name='Дубликаты', index=False)
                
                # Лист 5: Статистика
                if all_data:
                    total_plan = sum(item['plan'] for item in all_data)
                    total_fact = sum(item['fact'] for item in all_data)
                    duplicates_count = len([item for item in all_data if item['is_duplicate']])
                    
                    stats_data = {
                        'Показатель': [
                            'Всего задач',
                            'Общий план',
                            'Общий факт',
                            'Средний % выполнения',
                            'Количество дубликатов',
                            'Уникальных руководителей',
                            'Период анализа (недель)',
                            'Самый активный руководитель',
                            'Руководитель с лучшим % выполнения',
                            'Руководитель с худшим % выполнения'
                        ],
                        'Значение': [
                            len(all_data),
                            round(total_plan, 2),
                            round(total_fact, 2),
                            round(sum(item['completion_percent'] for item in all_data) / len(all_data), 2),
                            duplicates_count,
                            len(set(item['sender'] for item in all_data)),
                            len(set(item['date'].strftime('%Y-%W') for item in all_data)),
                            max(set(item['sender'] for item in all_data), key=lambda x: len([item for item in all_data if item['sender'] == x])),
                            max(set(item['sender'] for item in all_data), key=lambda x: sum([item['completion_percent'] for item in all_data if item['sender'] == x]) / len([item for item in all_data if item['sender'] == x])),
                            min(set(item['sender'] for item in all_data), key=lambda x: sum([item['completion_percent'] for item in all_data if item['sender'] == x]) / len([item for item in all_data if item['sender'] == x]))
                        ]
                    }
                    
                    df_stats = pd.DataFrame(stats_data)
                    df_stats.to_excel(writer, sheet_name='Общая статистика', index=False)
            
            print(f"✅ Детальный отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    async def run_detailed_analysis(self):
        """Запуск детального анализа"""
        try:
            print("🚀 ЗАПУСК ДЕТАЛЬНОГО АНАЛИЗА ПЛАНОВ И ФАКТОВ")
            print("=" * 60)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение сообщений
            messages = await self.get_weekly_messages(weeks_back=8)  # 8 недель для анализа
            
            if not messages:
                print("❌ Не найдено сообщений для анализа")
                return
            
            # Извлечение данных
            all_extracted_data = []
            
            for msg in messages:
                extracted = self.extract_detailed_plan_fact_data(msg['text'], msg['date'], msg['sender'], msg['id'])
                all_extracted_data.extend(extracted)
            
            print(f"📊 Извлечено {len(all_extracted_data)} отдельных пунктов планов/фактов")
            
            if not all_extracted_data:
                print("❌ Не удалось извлечь данные планов/фактов")
                return
            
            # Поиск дубликатов
            all_extracted_data = self.find_duplicates(all_extracted_data)
            duplicates_count = len([item for item in all_extracted_data if item['is_duplicate']])
            print(f"🔍 Найдено {duplicates_count} дублирующихся записей")
            
            # Создание отчета
            report_file = self.create_detailed_excel_report(all_extracted_data)
            
            if report_file:
                print(f"\n🎉 Детальный анализ завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Всего пунктов: {len(all_extracted_data)}")
                print(f"Уникальных руководителей: {len(set(item['sender'] for item in all_extracted_data))}")
                print(f"Дубликатов: {duplicates_count}")
                
                # Показываем статистику по руководителям
                sender_stats = {}
                for item in all_extracted_data:
                    sender = item['sender']
                    if sender not in sender_stats:
                        sender_stats[sender] = {'count': 0, 'total_completion': 0}
                    sender_stats[sender]['count'] += 1
                    sender_stats[sender]['total_completion'] += item['completion_percent']
                
                print(f"\n👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ:")
                for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['count'], reverse=True):
                    avg_completion = stats['total_completion'] / stats['count']
                    print(f"  {sender}: {stats['count']} задач, средний % выполнения: {avg_completion:.1f}%")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = DetailedPlanFactAnalyzer()
    await analyzer.run_detailed_analysis()

if __name__ == "__main__":
    asyncio.run(main())
