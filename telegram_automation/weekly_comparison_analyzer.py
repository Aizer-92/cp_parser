#!/usr/bin/env python3
"""
Weekly Comparison Analyzer for Main mega ТОП
Анализатор для сравнения планов прошлой недели с фактами текущей недели
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class WeeklyComparisonAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/weekly_comparison.log", rotation="1 day", retention="7 days")
        
        # Цветовая схема
        self.colors = {
            'green': '28A745',      # Выполнено
            'yellow': 'FFC107',     # Частично выполнено
            'red': 'DC3545',        # Не выполнено
            'blue': '007BFF',       # Заголовки
            'light_gray': 'F8F9FA', # Фон
            'dark_gray': '6C757D'   # Текст
        }
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_weekly_messages(self, weeks_back=12):
        """Получение сообщений по неделям"""
        try:
            print(f"📥 Загрузка сообщений за последние {weeks_back} недель...")
            
            today = datetime.now()
            start_date = today - timedelta(weeks=weeks_back)
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=10000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            # Группируем сообщения по неделям
            weekly_messages = {}
            
            for msg in messages:
                if msg.text and len(msg.text.strip()) > 50:
                    # Определяем неделю
                    msg_date = msg.date.replace(tzinfo=None)
                    week_start = msg_date - timedelta(days=msg_date.weekday())
                    week_key = week_start.strftime('%Y-%m-%d')
                    
                    if week_key not in weekly_messages:
                        weekly_messages[week_key] = []
                    
                    sender_name = await self.get_sender_name(msg)
                    weekly_messages[week_key].append({
                        'id': msg.id,
                        'date': msg_date,
                        'text': msg.text.strip(),
                        'sender': sender_name,
                        'week_start': week_start,
                        'week_key': week_key
                    })
            
            print(f"📊 Сгруппировано по {len(weekly_messages)} неделям")
            return weekly_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return {}
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def extract_plan_fact_sections(self, text, sender, message_date):
        """Извлечение разделов план и факт из текста"""
        try:
            # Разбиваем на строки
            lines = text.split('\n')
            
            plan_section = []
            fact_section = []
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Определяем раздел
                if re.match(r'^план[:\s]*', line.lower()):
                    current_section = 'plan'
                    continue
                elif re.match(r'^факт[:\s]*', line.lower()):
                    current_section = 'fact'
                    continue
                
                # Добавляем в соответствующий раздел
                if current_section == 'plan':
                    plan_section.append(line)
                elif current_section == 'fact':
                    fact_section.append(line)
            
            return {
                'sender': sender,
                'date': message_date,
                'plan_tasks': plan_section,
                'fact_tasks': fact_section,
                'has_plan': len(plan_section) > 0,
                'has_fact': len(fact_section) > 0
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка извлечения разделов: {e}")
            return None
    
    def analyze_weekly_comparison(self, weekly_messages):
        """Анализ сравнения недель"""
        try:
            print("🔍 Анализ сравнения недель...")
            
            comparisons = []
            week_keys = sorted(weekly_messages.keys())
            
            # Проходим по неделям (кроме последней)
            for i in range(len(week_keys) - 1):
                current_week = week_keys[i]
                next_week = week_keys[i + 1]
                
                current_week_messages = weekly_messages[current_week]
                next_week_messages = weekly_messages[next_week]
                
                # Анализируем планы текущей недели
                current_week_plans = {}
                for msg in current_week_messages:
                    sections = self.extract_plan_fact_sections(msg['text'], msg['sender'], msg['date'])
                    if sections and sections['has_plan']:
                        if sections['sender'] not in current_week_plans:
                            current_week_plans[sections['sender']] = []
                        current_week_plans[sections['sender']].extend(sections['plan_tasks'])
                
                # Анализируем факты следующей недели
                next_week_facts = {}
                for msg in next_week_messages:
                    sections = self.extract_plan_fact_sections(msg['text'], msg['sender'], msg['date'])
                    if sections and sections['has_fact']:
                        if sections['sender'] not in next_week_facts:
                            next_week_facts[sections['sender']] = []
                        next_week_facts[sections['sender']].extend(sections['fact_tasks'])
                
                # Сравниваем для каждого руководителя
                for sender in set(current_week_plans.keys()) | set(next_week_facts.keys()):
                    plan_tasks = current_week_plans.get(sender, [])
                    fact_tasks = next_week_facts.get(sender, [])
                    
                    if plan_tasks or fact_tasks:
                        comparison = self.compare_plan_fact_tasks(plan_tasks, fact_tasks, sender, current_week, next_week)
                        if comparison:
                            comparisons.append(comparison)
            
            print(f"📊 Создано {len(comparisons)} сравнений")
            return comparisons
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа сравнения: {e}")
            return []
    
    def compare_plan_fact_tasks(self, plan_tasks, fact_tasks, sender, plan_week, fact_week):
        """Сравнение планов и фактов"""
        try:
            # Анализируем выполнение задач
            completed_tasks = []
            partially_completed = []
            not_completed = []
            new_tasks = []
            
            # Проверяем каждую задачу из плана
            for plan_task in plan_tasks:
                task_completed = False
                task_partial = False
                
                # Ищем соответствие в фактах
                for fact_task in fact_tasks:
                    similarity = self.calculate_task_similarity(plan_task, fact_task)
                    
                    if similarity >= 0.7:  # Высокое сходство
                        completed_tasks.append({
                            'plan': plan_task,
                            'fact': fact_task,
                            'similarity': similarity
                        })
                        task_completed = True
                        break
                    elif similarity >= 0.4:  # Частичное сходство
                        task_partial = True
                
                if not task_completed:
                    if task_partial:
                        partially_completed.append(plan_task)
                    else:
                        not_completed.append(plan_task)
            
            # Находим новые задачи (есть в фактах, но не было в планах)
            for fact_task in fact_tasks:
                is_new = True
                for plan_task in plan_tasks:
                    if self.calculate_task_similarity(plan_task, fact_task) >= 0.4:
                        is_new = False
                        break
                
                if is_new:
                    new_tasks.append(fact_task)
            
            return {
                'sender': sender,
                'plan_week': plan_week,
                'fact_week': fact_week,
                'plan_tasks': plan_tasks,
                'fact_tasks': fact_tasks,
                'completed_tasks': completed_tasks,
                'partially_completed': partially_completed,
                'not_completed': not_completed,
                'new_tasks': new_tasks,
                'completion_rate': len(completed_tasks) / len(plan_tasks) * 100 if plan_tasks else 0,
                'total_planned': len(plan_tasks),
                'total_completed': len(completed_tasks),
                'total_partial': len(partially_completed),
                'total_not_completed': len(not_completed),
                'total_new': len(new_tasks)
            }
            
        except Exception as e:
            logger.error(f"❌ Ошибка сравнения задач: {e}")
            return None
    
    def calculate_task_similarity(self, task1, task2):
        """Вычисление сходства между задачами"""
        try:
            # Приводим к нижнему регистру
            t1 = task1.lower()
            t2 = task2.lower()
            
            # Убираем цифры и специальные символы
            t1_clean = re.sub(r'[^\w\s]', '', t1)
            t2_clean = re.sub(r'[^\w\s]', '', t2)
            
            # Разбиваем на слова
            words1 = set(t1_clean.split())
            words2 = set(t2_clean.split())
            
            if not words1 or not words2:
                return 0.0
            
            # Вычисляем Jaccard similarity
            intersection = len(words1.intersection(words2))
            union = len(words1.union(words2))
            
            return intersection / union if union > 0 else 0.0
            
        except Exception as e:
            return 0.0
    
    def create_comparison_report(self, comparisons, filename=None):
        """Создание отчета сравнения"""
        try:
            if not filename:
                filename = f"output/weekly_comparison_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            # Удаляем дефолтный лист
            workbook.remove(workbook.active)
            
            # Создаем листы
            self.create_summary_sheet(workbook, comparisons)
            self.create_detailed_sheet(workbook, comparisons)
            self.create_sender_sheet(workbook, comparisons)
            
            workbook.save(filename)
            print(f"✅ Отчет сравнения создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return None
    
    def create_summary_sheet(self, workbook, comparisons):
        """Создание сводного листа"""
        try:
            ws = workbook.create_sheet("Сводка")
            
            # Заголовок
            ws['A1'] = "СРАВНЕНИЕ ПЛАНОВ И ФАКТОВ ПО НЕДЕЛЯМ"
            ws['A2'] = "Main mega ТОП"
            
            # Общая статистика
            total_comparisons = len(comparisons)
            total_planned = sum(comp['total_planned'] for comp in comparisons)
            total_completed = sum(comp['total_completed'] for comp in comparisons)
            avg_completion = (total_completed / total_planned * 100) if total_planned > 0 else 0
            
            ws['A4'] = "📊 ОБЩАЯ СТАТИСТИКА:"
            ws['A5'] = f"Всего сравнений: {total_comparisons}"
            ws['A6'] = f"Всего запланировано: {total_planned}"
            ws['A7'] = f"Всего выполнено: {total_completed}"
            ws['A8'] = f"Средний % выполнения: {avg_completion:.1f}%"
            
            # Топ руководителей
            sender_stats = {}
            for comp in comparisons:
                sender = comp['sender']
                if sender not in sender_stats:
                    sender_stats[sender] = {
                        'total_planned': 0,
                        'total_completed': 0,
                        'comparisons': 0
                    }
                
                sender_stats[sender]['total_planned'] += comp['total_planned']
                sender_stats[sender]['total_completed'] += comp['total_completed']
                sender_stats[sender]['comparisons'] += 1
            
            ws['A10'] = "👥 СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ:"
            
            headers = ['Руководитель', 'Сравнений', 'Запланировано', 'Выполнено', '% выполнения']
            for col, header in enumerate(headers, 1):
                ws.cell(row=11, column=col, value=header)
            
            row = 12
            for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['total_completed'], reverse=True):
                completion_rate = (stats['total_completed'] / stats['total_planned'] * 100) if stats['total_planned'] > 0 else 0
                
                ws.cell(row=row, column=1, value=sender)
                ws.cell(row=row, column=2, value=stats['comparisons'])
                ws.cell(row=row, column=3, value=stats['total_planned'])
                ws.cell(row=row, column=4, value=stats['total_completed'])
                ws.cell(row=row, column=5, value=f"{completion_rate:.1f}%")
                
                row += 1
            
            # Форматирование
            self.format_summary_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания сводного листа: {e}")
    
    def create_detailed_sheet(self, workbook, comparisons):
        """Создание детального листа"""
        try:
            ws = workbook.create_sheet("Детализация")
            
            # Заголовки
            headers = [
                'Руководитель', 'Неделя плана', 'Неделя факта', 
                'Запланировано', 'Выполнено', 'Частично', 'Не выполнено', 'Новых задач',
                '% выполнения', 'Статус'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for comp in comparisons:
                status = self.get_completion_status(comp['completion_rate'])
                
                ws.cell(row=row, column=1, value=comp['sender'])
                ws.cell(row=row, column=2, value=comp['plan_week'])
                ws.cell(row=row, column=3, value=comp['fact_week'])
                ws.cell(row=row, column=4, value=comp['total_planned'])
                ws.cell(row=row, column=5, value=comp['total_completed'])
                ws.cell(row=row, column=6, value=comp['total_partial'])
                ws.cell(row=row, column=7, value=comp['total_not_completed'])
                ws.cell(row=row, column=8, value=comp['total_new'])
                ws.cell(row=row, column=9, value=f"{comp['completion_rate']:.1f}%")
                ws.cell(row=row, column=10, value=status)
                
                row += 1
            
            # Форматирование
            self.format_detailed_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания детального листа: {e}")
    
    def create_sender_sheet(self, workbook, comparisons):
        """Создание листа по руководителям"""
        try:
            ws = workbook.create_sheet("По руководителям")
            
            # Группируем по руководителям
            sender_data = {}
            for comp in comparisons:
                sender = comp['sender']
                if sender not in sender_data:
                    sender_data[sender] = []
                sender_data[sender].append(comp)
            
            row = 1
            for sender, sender_comps in sender_data.items():
                # Заголовок руководителя
                ws.cell(row=row, column=1, value=f"РУКОВОДИТЕЛЬ: {sender}")
                ws.cell(row=row, column=1).font = Font(bold=True, size=14)
                row += 1
                
                # Заголовки таблицы
                headers = ['Неделя плана', 'Неделя факта', 'План', 'Факт', 'Статус']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                    ws.cell(row=row, column=col).font = Font(bold=True)
                row += 1
                
                # Данные по неделям
                for comp in sender_comps:
                    # План
                    plan_text = '\n'.join(comp['plan_tasks'][:3])  # Первые 3 задачи
                    if len(comp['plan_tasks']) > 3:
                        plan_text += f"\n... и еще {len(comp['plan_tasks']) - 3} задач"
                    
                    # Факт
                    fact_text = '\n'.join(comp['fact_tasks'][:3])  # Первые 3 задачи
                    if len(comp['fact_tasks']) > 3:
                        fact_text += f"\n... и еще {len(comp['fact_tasks']) - 3} задач"
                    
                    # Статус
                    status = self.get_completion_status(comp['completion_rate'])
                    
                    ws.cell(row=row, column=1, value=comp['plan_week'])
                    ws.cell(row=row, column=2, value=comp['fact_week'])
                    ws.cell(row=row, column=3, value=plan_text)
                    ws.cell(row=row, column=4, value=fact_text)
                    ws.cell(row=row, column=5, value=status)
                    
                    # Форматирование ячеек с текстом
                    for col in [3, 4]:
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        ws.row_dimensions[row].height = 80
                    
                    row += 1
                
                row += 2  # Пропуск между руководителями
            
            # Форматирование
            self.format_sender_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа по руководителям: {e}")
    
    def get_completion_status(self, completion_rate):
        """Определение статуса выполнения"""
        if completion_rate >= 80:
            return "✅ Отлично"
        elif completion_rate >= 60:
            return "⚠️ Хорошо"
        elif completion_rate >= 40:
            return "⚠️ Средне"
        else:
            return "❌ Плохо"
    
    def format_summary_sheet(self, ws):
        """Форматирование сводного листа"""
        # Заголовок
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['blue'])
        ws['A2'].font = Font(size=14, bold=True)
        
        # Объединение ячеек
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def format_detailed_sheet(self, ws):
        """Форматирование детального листа"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Ширина колонок
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def format_sender_sheet(self, ws):
        """Форматирование листа по руководителям"""
        # Ширина колонок
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
    
    async def run_weekly_comparison(self):
        """Запуск анализа сравнения недель"""
        try:
            print("🚀 ЗАПУСК АНАЛИЗА СРАВНЕНИЯ НЕДЕЛЬ")
            print("=" * 55)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение сообщений по неделям
            weekly_messages = await self.get_weekly_messages(weeks_back=12)
            
            if not weekly_messages:
                print("❌ Не найдено сообщений по неделям")
                return
            
            # Анализ сравнения
            comparisons = self.analyze_weekly_comparison(weekly_messages)
            
            if not comparisons:
                print("❌ Не найдено сравнений планов и фактов")
                return
            
            # Создание отчета
            report_file = self.create_comparison_report(comparisons)
            
            if report_file:
                print(f"\n🎉 Анализ сравнения завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Проанализировано недель: {len(weekly_messages)}")
                print(f"Создано сравнений: {len(comparisons)}")
                
                # Статистика по руководителям
                senders = set(comp['sender'] for comp in comparisons)
                print(f"Уникальных руководителей: {len(senders)}")
                
                # Средний % выполнения
                total_planned = sum(comp['total_planned'] for comp in comparisons)
                total_completed = sum(comp['total_completed'] for comp in comparisons)
                avg_completion = (total_completed / total_planned * 100) if total_planned > 0 else 0
                print(f"Средний % выполнения: {avg_completion:.1f}%")
                
                # Показываем топ руководителей
                sender_stats = {}
                for comp in comparisons:
                    sender = comp['sender']
                    if sender not in sender_stats:
                        sender_stats[sender] = {'planned': 0, 'completed': 0}
                    sender_stats[sender]['planned'] += comp['total_planned']
                    sender_stats[sender]['completed'] += comp['total_completed']
                
                print(f"\n👥 ТОП РУКОВОДИТЕЛЕЙ:")
                for sender, stats in sorted(sender_stats.items(), key=lambda x: x[1]['completed'], reverse=True)[:5]:
                    completion_rate = (stats['completed'] / stats['planned'] * 100) if stats['planned'] > 0 else 0
                    print(f"  {sender}: {stats['completed']}/{stats['planned']} ({completion_rate:.1f}%)")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = WeeklyComparisonAnalyzer()
    await analyzer.run_weekly_comparison()

if __name__ == "__main__":
    asyncio.run(main())
