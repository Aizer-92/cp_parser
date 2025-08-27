#!/usr/bin/env python3
"""
Dashboard Plan/Fact Analyzer for Main mega ТОП
Анализатор для создания дашборда по планам и фактам
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from telethon import TelegramClient
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class DashboardPlanFactAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "Main mega ТОП"
        self.target_chat_id = None
        
        # Настройка логирования
        os.makedirs("output/logs", exist_ok=True)
        logger.add("output/logs/dashboard_plan_fact.log", rotation="1 day", retention="7 days")
        
        # Цветовая схема
        self.colors = {
            'green': '28A745',      # Выполнение 90%+
            'yellow': 'FFC107',     # Выполнение 70-89%
            'red': 'DC3545',        # Выполнение <70%
            'blue': '007BFF',       # Заголовки
            'light_gray': 'F8F9FA', # Фон
            'dark_gray': '6C757D'   # Текст
        }
        
    def load_config(self):
        with open("config/personal_config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"👤 Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_target_chat(self):
        """Поиск целевого чата"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            return False
            
        except Exception as e:
            logger.error(f"❌ Ошибка поиска чата: {e}")
            return False
    
    async def get_filtered_messages(self, weeks_back=52):
        """Получение отфильтрованных сообщений"""
        try:
            print(f"📥 Загрузка сообщений за последние {weeks_back} недель...")
            
            today = datetime.now()
            start_date = today - timedelta(weeks=weeks_back)
            
            messages = await self.client.get_messages(
                self.target_chat_id,
                offset_date=start_date,
                limit=10000
            )
            
            print(f"✅ Получено {len(messages)} сообщений")
            
            # Фильтруем сообщения с планами и фактами
            filtered_messages = []
            
            for msg in messages:
                if msg.text and len(msg.text.strip()) > 50:
                    text = msg.text.lower()
                    
                    # Ищем сообщения с планами и фактами
                    has_plan_fact = (
                        ('план' in text and 'факт' in text) or
                        ('п/ф' in text) or
                        (re.search(r'\d+/\d+', text)) or
                        (re.search(r'\(\d+.*\d+\)', text))
                    )
                    
                    if has_plan_fact:
                        sender_name = await self.get_sender_name(msg)
                        filtered_messages.append({
                            'id': msg.id,
                            'date': msg.date,
                            'text': msg.text.strip(),
                            'sender': sender_name,
                            'length': len(msg.text.strip()),
                            'weekday': msg.date.weekday(),
                            'weekday_name': ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'][msg.date.weekday()]
                        })
            
            print(f"📊 Отфильтровано {len(filtered_messages)} сообщений с планами/фактами")
            return filtered_messages
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения сообщений: {e}")
            return []
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            sender = await message.get_sender()
            if sender:
                return f"{sender.first_name} {sender.last_name or ''}".strip()
            return "Unknown"
        except:
            return "Unknown"
    
    def extract_plan_fact_data(self, messages):
        """Извлечение данных планов и фактов"""
        try:
            print("🔍 Извлечение данных планов и фактов...")
            
            all_tasks = []
            
            for msg in messages:
                # Извлекаем числа из сообщения
                numbers = re.findall(r'(\d+(?:\.\d+)?)', msg['text'])
                numbers = [float(n) for n in numbers if 0 <= float(n) <= 1000]
                
                if len(numbers) >= 2:
                    # Группируем числа попарно (план/факт)
                    for i in range(0, len(numbers) - 1, 2):
                        plan = numbers[i]
                        fact = numbers[i + 1]
                        
                        # Фильтруем разумные значения
                        completion_percent = (fact / plan * 100) if plan > 0 else 0
                        
                        if 0 <= completion_percent <= 300:  # Максимум 300%
                            all_tasks.append({
                                'message_id': msg['id'],
                                'date': msg['date'],
                                'sender': msg['sender'],
                                'plan': plan,
                                'fact': fact,
                                'difference': fact - plan,
                                'completion_percent': completion_percent,
                                'status': self.get_status(completion_percent),
                                'weekday': msg['weekday_name'],
                                'is_duplicate': False
                            })
            
            print(f"📊 Извлечено {len(all_tasks)} задач")
            return all_tasks
            
        except Exception as e:
            logger.error(f"❌ Ошибка извлечения данных: {e}")
            return []
    
    def get_status(self, completion_percent):
        """Определение статуса выполнения"""
        if completion_percent >= 100:
            return "Перевыполнение"
        elif completion_percent >= 90:
            return "Хорошо"
        elif completion_percent >= 70:
            return "Удовлетворительно"
        else:
            return "Недовыполнение"
    
    def analyze_by_senders(self, tasks):
        """Анализ по руководителям"""
        try:
            sender_analysis = {}
            
            for task in tasks:
                sender = task['sender']
                if sender not in sender_analysis:
                    sender_analysis[sender] = {
                        'total_plan': 0,
                        'total_fact': 0,
                        'tasks_count': 0,
                        'over_performance': 0,
                        'under_performance': 0,
                        'dates': set(),
                        'messages': set(),
                        'last_activity': None
                    }
                
                sender_analysis[sender]['total_plan'] += task['plan']
                sender_analysis[sender]['total_fact'] += task['fact']
                sender_analysis[sender]['tasks_count'] += 1
                sender_analysis[sender]['dates'].add(task['date'].strftime('%Y-%m-%d'))
                sender_analysis[sender]['messages'].add(task['message_id'])
                
                if task['completion_percent'] >= 100:
                    sender_analysis[sender]['over_performance'] += 1
                else:
                    sender_analysis[sender]['under_performance'] += 1
                
                # Обновляем последнюю активность
                if (sender_analysis[sender]['last_activity'] is None or 
                    task['date'] > sender_analysis[sender]['last_activity']):
                    sender_analysis[sender]['last_activity'] = task['date']
            
            return sender_analysis
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа по руководителям: {e}")
            return {}
    
    def analyze_by_weeks(self, tasks):
        """Анализ по неделям"""
        try:
            weekly_analysis = {}
            
            for task in tasks:
                # Определяем неделю
                date = task['date'].replace(tzinfo=None)
                week_start = date - timedelta(days=date.weekday())
                week_key = week_start.strftime('%Y-%m-%d')
                
                if week_key not in weekly_analysis:
                    weekly_analysis[week_key] = {
                        'total_plan': 0,
                        'total_fact': 0,
                        'tasks_count': 0,
                        'senders': set(),
                        'sender_tasks': {}
                    }
                
                weekly_analysis[week_key]['total_plan'] += task['plan']
                weekly_analysis[week_key]['total_fact'] += task['fact']
                weekly_analysis[week_key]['tasks_count'] += 1
                weekly_analysis[week_key]['senders'].add(task['sender'])
                
                # Подсчитываем задачи по руководителям
                if task['sender'] not in weekly_analysis[week_key]['sender_tasks']:
                    weekly_analysis[week_key]['sender_tasks'][task['sender']] = 0
                weekly_analysis[week_key]['sender_tasks'][task['sender']] += 1
            
            return weekly_analysis
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа по неделям: {e}")
            return {}
    
    def create_dashboard_report(self, tasks, sender_analysis, weekly_analysis, filename=None):
        """Создание дашборда отчета"""
        try:
            if not filename:
                filename = f"output/dashboard_plan_fact_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            os.makedirs("output", exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            # Удаляем дефолтный лист
            workbook.remove(workbook.active)
            
            # Создаем листы
            self.create_summary_sheet(workbook, tasks, sender_analysis)
            self.create_senders_sheet(workbook, sender_analysis)
            self.create_tasks_sheet(workbook, tasks)
            self.create_weekly_sheet(workbook, weekly_analysis)
            self.create_problems_sheet(workbook, sender_analysis)
            
            workbook.save(filename)
            print(f"✅ Дашборд создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания дашборда: {e}")
            return None
    
    def create_summary_sheet(self, workbook, tasks, sender_analysis):
        """Создание сводного листа"""
        try:
            ws = workbook.create_sheet("Сводка")
            
            # Заголовок
            ws['A1'] = "СВОДКА ПО ПЛАНАМ И ФАКТАМ"
            ws['A2'] = "Main mega ТОП"
            
            # Общая статистика
            total_tasks = len(tasks)
            total_plan = sum(task['plan'] for task in tasks)
            total_fact = sum(task['fact'] for task in tasks)
            avg_completion = (total_fact / total_plan * 100) if total_plan > 0 else 0
            completed_tasks = len([t for t in tasks if t['completion_percent'] >= 70])
            
            ws['A4'] = "📅 Период:"
            ws['B4'] = f"{min(task['date'] for task in tasks).strftime('%d.%m.%Y')} - {max(task['date'] for task in tasks).strftime('%d.%m.%Y')}"
            
            ws['A5'] = "👥 Всего руководителей:"
            ws['B5'] = len(sender_analysis)
            
            ws['A6'] = "📊 Всего задач:"
            ws['B6'] = total_tasks
            
            ws['A7'] = "✅ Выполнено:"
            ws['B7'] = f"{completed_tasks} ({completed_tasks/total_tasks*100:.1f}%)"
            
            ws['A8'] = "⚠️ Недовыполнено:"
            ws['B8'] = f"{total_tasks - completed_tasks} ({(total_tasks - completed_tasks)/total_tasks*100:.1f}%)"
            
            # Топ-5 руководителей
            ws['A10'] = "ТОП-5 РУКОВОДИТЕЛЕЙ ПО АКТИВНОСТИ"
            
            # Сортируем руководителей по количеству задач
            sorted_senders = sorted(sender_analysis.items(), key=lambda x: x[1]['tasks_count'], reverse=True)[:5]
            
            headers = ['Руководитель', 'Задачи', 'План', 'Факт', '% Выполнения', 'Статус']
            for col, header in enumerate(headers, 1):
                ws.cell(row=11, column=col, value=header)
            
            for row, (sender, data) in enumerate(sorted_senders, 12):
                avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                status = self.get_status(avg_completion)
                
                ws.cell(row=row, column=1, value=sender)
                ws.cell(row=row, column=2, value=data['tasks_count'])
                ws.cell(row=row, column=3, value=round(data['total_plan'], 2))
                ws.cell(row=row, column=4, value=round(data['total_fact'], 2))
                ws.cell(row=row, column=5, value=f"{avg_completion:.1f}%")
                ws.cell(row=row, column=6, value=status)
            
            # Форматирование
            self.format_summary_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания сводного листа: {e}")
    
    def create_senders_sheet(self, workbook, sender_analysis):
        """Создание листа руководителей"""
        try:
            ws = workbook.create_sheet("Руководители")
            
            # Заголовки
            headers = [
                'Руководитель', 'Задачи', 'Общий план', 'Общий факт', 
                '% выполнения', 'Статус', 'Перевыполнено', 'Недовыполнено',
                'Последняя активность', 'Тренд'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for sender, data in sender_analysis.items():
                avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                status = self.get_status(avg_completion)
                
                ws.cell(row=row, column=1, value=sender)
                ws.cell(row=row, column=2, value=data['tasks_count'])
                ws.cell(row=row, column=3, value=round(data['total_plan'], 2))
                ws.cell(row=row, column=4, value=round(data['total_fact'], 2))
                ws.cell(row=row, column=5, value=f"{avg_completion:.1f}%")
                ws.cell(row=row, column=6, value=status)
                ws.cell(row=row, column=7, value=data['over_performance'])
                ws.cell(row=row, column=8, value=data['under_performance'])
                ws.cell(row=row, column=9, value=data['last_activity'].strftime('%Y-%m-%d') if data['last_activity'] else '')
                ws.cell(row=row, column=10, value='↗️' if avg_completion >= 90 else '→' if avg_completion >= 70 else '↘️')
                
                row += 1
            
            # Форматирование
            self.format_senders_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа руководителей: {e}")
    
    def create_tasks_sheet(self, workbook, tasks):
        """Создание листа задач"""
        try:
            ws = workbook.create_sheet("Детализация")
            
            # Заголовки
            headers = [
                'Дата', 'Руководитель', 'План', 'Факт', 'Разница', 
                '% выполнения', 'Статус', 'День недели'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task['date'].strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=task['sender'])
                ws.cell(row=row, column=3, value=task['plan'])
                ws.cell(row=row, column=4, value=task['fact'])
                ws.cell(row=row, column=5, value=task['difference'])
                ws.cell(row=row, column=6, value=f"{task['completion_percent']:.1f}%")
                ws.cell(row=row, column=7, value=task['status'])
                ws.cell(row=row, column=8, value=task['weekday'])
            
            # Форматирование
            self.format_tasks_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа задач: {e}")
    
    def create_weekly_sheet(self, workbook, weekly_analysis):
        """Создание листа по неделям"""
        try:
            ws = workbook.create_sheet("По неделям")
            
            # Заголовки
            headers = [
                'Неделя', 'Всего задач', 'Общий план', 'Общий факт', 
                '% выполнения', 'Руководителей', 'Топ-3 руководителя'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for week, data in sorted(weekly_analysis.items(), reverse=True):
                avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                
                # Топ-3 руководителя недели
                top_senders = sorted(data['sender_tasks'].items(), key=lambda x: x[1], reverse=True)[:3]
                top_senders_str = ', '.join([f"{sender}({tasks})" for sender, tasks in top_senders])
                
                ws.cell(row=row, column=1, value=week)
                ws.cell(row=row, column=2, value=data['tasks_count'])
                ws.cell(row=row, column=3, value=round(data['total_plan'], 2))
                ws.cell(row=row, column=4, value=round(data['total_fact'], 2))
                ws.cell(row=row, column=5, value=f"{avg_completion:.1f}%")
                ws.cell(row=row, column=6, value=len(data['senders']))
                ws.cell(row=row, column=7, value=top_senders_str)
                
                row += 1
            
            # Форматирование
            self.format_weekly_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа по неделям: {e}")
    
    def create_problems_sheet(self, workbook, sender_analysis):
        """Создание листа проблем"""
        try:
            ws = workbook.create_sheet("Проблемы")
            
            # Критические проблемы
            ws['A1'] = "🚨 КРИТИЧЕСКИЕ ПРОБЛЕМЫ"
            
            row = 3
            for sender, data in sender_analysis.items():
                avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                
                if avg_completion < 70:
                    ws.cell(row=row, column=1, value=f"• {sender} - стабильно недовыполняет планы ({avg_completion:.1f}% в среднем)")
                    row += 1
            
            # Рекомендации
            ws['A10'] = "💡 РЕКОМЕНДАЦИИ"
            ws['A12'] = "• Провести встречи с проблемными руководителями"
            ws['A13'] = "• Улучшить планирование задач (снизить план на 10-15%)"
            ws['A14'] = "• Внедрить еженедельные чек-апы по выполнению задач"
            ws['A15'] = "• Добавить систему уведомлений о приближающихся дедлайнах"
            
            # Форматирование
            self.format_problems_sheet(ws)
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания листа проблем: {e}")
    
    def format_summary_sheet(self, ws):
        """Форматирование сводного листа"""
        # Заголовок
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['blue'])
        ws['A2'].font = Font(size=14, bold=True)
        
        # Объединение ячеек
        ws.merge_cells('A1:B1')
        ws.merge_cells('A2:B2')
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def format_senders_sheet(self, ws):
        """Форматирование листа руководителей"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Ширина колонок
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def format_tasks_sheet(self, ws):
        """Форматирование листа задач"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Условное форматирование для % выполнения
        for row in range(2, ws.max_row + 1):
            completion_cell = ws.cell(row=row, column=6)
            try:
                completion_value = float(completion_cell.value.replace('%', ''))
                if completion_value >= 90:
                    completion_cell.fill = PatternFill(start_color=self.colors['green'], end_color=self.colors['green'], fill_type='solid')
                elif completion_value >= 70:
                    completion_cell.fill = PatternFill(start_color=self.colors['yellow'], end_color=self.colors['yellow'], fill_type='solid')
                else:
                    completion_cell.fill = PatternFill(start_color=self.colors['red'], end_color=self.colors['red'], fill_type='solid')
            except:
                pass
    
    def format_weekly_sheet(self, ws):
        """Форматирование листа по неделям"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['blue'], end_color=self.colors['blue'], fill_type='solid')
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['G'].width = 30
    
    def format_problems_sheet(self, ws):
        """Форматирование листа проблем"""
        # Заголовки
        ws['A1'].font = Font(size=14, bold=True, color=self.colors['red'])
        ws['A10'].font = Font(size=14, bold=True, color=self.colors['blue'])
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 80
    
    async def run_dashboard_analysis(self):
        """Запуск анализа дашборда"""
        try:
            print("🚀 ЗАПУСК АНАЛИЗА ДАШБОРДА ПЛАНОВ И ФАКТОВ")
            print("=" * 60)
            
            # Подключение
            if not await self.connect():
                return
            
            # Поиск целевого чата
            if not await self.find_target_chat():
                return
            
            # Получение отфильтрованных сообщений
            messages = await self.get_filtered_messages(weeks_back=52)
            
            if not messages:
                print("❌ Не найдено подходящих сообщений")
                return
            
            # Извлечение данных
            tasks = self.extract_plan_fact_data(messages)
            
            if not tasks:
                print("❌ Не найдено задач с планами и фактами")
                return
            
            # Анализ данных
            sender_analysis = self.analyze_by_senders(tasks)
            weekly_analysis = self.analyze_by_weeks(tasks)
            
            # Создание дашборда
            report_file = self.create_dashboard_report(tasks, sender_analysis, weekly_analysis)
            
            if report_file:
                print(f"\n🎉 Анализ дашборда завершен!")
                print(f"📁 Отчет сохранен: {report_file}")
                
                # Выводим краткую сводку
                print(f"\n📊 КРАТКАЯ СВОДКА:")
                print(f"Проанализировано сообщений: {len(messages)}")
                print(f"Найдено задач: {len(tasks)}")
                print(f"Уникальных руководителей: {len(sender_analysis)}")
                
                # Показываем топ руководителей
                sorted_senders = sorted(sender_analysis.items(), key=lambda x: x[1]['tasks_count'], reverse=True)[:5]
                print(f"\n👥 ТОП-5 РУКОВОДИТЕЛЕЙ:")
                for i, (sender, data) in enumerate(sorted_senders, 1):
                    avg_completion = (data['total_fact'] / data['total_plan'] * 100) if data['total_plan'] > 0 else 0
                    print(f"  {i}. {sender}: {data['tasks_count']} задач, {avg_completion:.1f}% выполнения")
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа: {e}")
            print(f"❌ Ошибка: {e}")

async def main():
    analyzer = DashboardPlanFactAnalyzer()
    await analyzer.run_dashboard_analysis()

if __name__ == "__main__":
    asyncio.run(main())
