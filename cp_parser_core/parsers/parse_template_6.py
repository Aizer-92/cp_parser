#!/usr/bin/env python3
"""
Универсальный парсер для Шаблона 6 (заголовки в строке 2)

Поддерживает:
- 27 столбцов (222 файла)
- 30 столбцов (179 файлов)
- 33 столбца (34 файла)
- ИТОГО: 435 файлов (65% покрытие!)

Особенности:
- Строка 1: информация о менеджере (пропускается)
- Строка 2: заголовки товаров
- Строка 3+: данные
- Автоматическое определение столбцов по названиям
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import re
import hashlib

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template6Parser:
    """
    Универсальный парсер для Шаблона 6
    Автоматически определяет структуру по заголовкам в строке 2
    """
    
    HEADER_ROW = 2  # Заголовки всегда в строке 2!
    
    def __init__(self):
        self.db = PostgreSQLManager()
        self.storage_dir = Path('storage/images')
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        
        # Словарь для маппинга столбцов (заполняется автоматически)
        self.columns = {}
        self.data_start_row = None  # Определяется динамически!
    
    def _detect_data_start_row(self, ws):
        """Динамически определяет с какой строки начинаются данные"""
        # Проверяем строки 3 и 4
        for row in [3, 4]:
            # Получаем значения из первых столбцов
            name_col = 2  # Обычно наименование во 2-м столбце
            name_val = ws.cell(row, name_col).value
            
            # Если строка содержит подзаголовки типа "Цена за шт." - пропускаем
            if name_val:
                name_str = str(name_val).lower()
                # Ключевые слова подзаголовков
                subheader_keywords = ['цена за шт', 'price', 'итого', 'срок тиража', 'delivery']
                
                if any(kw in name_str for kw in subheader_keywords):
                    continue  # Это подзаголовок, пропускаем
                else:
                    # Это реальные данные
                    return row
        
        # По умолчанию строка 3
        return 3
    
    def _detect_columns(self, ws):
        """Автоматически определяет индексы столбцов по названиям в строке 2 и 3
        
        Использует групповое сканирование:
        1. Находит группу маршрута в строке 2 (например "Доставка ЖД")
        2. Сканирует следующие N колонок для подзаголовков в строке 3
        3. Привязывает их к маршруту
        """
        self.columns = {}
        
        # Проверяем есть ли строка 3 с подзаголовками
        has_subheaders = False
        for col in range(1, min(10, ws.max_column + 1)):
            val3 = ws.cell(3, col).value
            if val3 and ('цена' in str(val3).lower() or 'price' in str(val3).lower()):
                has_subheaders = True
                break
        
        # Основные поля (не требуют группового сканирования)
        for col in range(1, ws.max_column + 1):
            header2 = ws.cell(self.HEADER_ROW, col).value
            header2_str = str(header2).lower().strip() if header2 else ''
            
            header3 = ws.cell(3, col).value if has_subheaders else None
            header3_str = str(header3).lower().strip() if header3 else ''
            
            combined_header = f"{header2_str} {header3_str}".strip()
            header_lower = combined_header if combined_header else header2_str
            
            # Основные поля
            if 'фото' in header_lower or 'photo' in header_lower:
                if 'доп' not in header_lower and 'extra' not in header_lower:
                    self.columns['photo'] = col
                else:
                    self.columns['extra_photo'] = col
            
            elif 'наименование' in header_lower or ('name' in header_lower and 'product' not in header_lower):
                self.columns['name'] = col
            
            elif 'описание мм' in header_lower or 'описание ММ' in header_lower:
                self.columns['description_mm'] = col
            
            elif 'описание' in header_lower or 'description' in header_lower:
                if 'description' not in self.columns:  # Первое описание
                    self.columns['description'] = col
            
            elif 'тз' in header_lower and len(header_lower) < 10:
                self.columns['tz'] = col
            
            elif 'кастом' in header_lower or 'custom' in header_lower or 'дизайн' in header_lower:
                self.columns['custom'] = col
            
            elif 'образец' in header_lower and 'срок' not in header_lower:
                self.columns['sample_price'] = col
            
            elif 'образец' in header_lower and 'срок' in header_lower:
                self.columns['sample_delivery'] = col
            
            # Тираж (ищем первый столбец с "тираж" или "quantity")
            elif ('тираж' in header_lower or 'quantity' in header_lower) and 'тираж_col' not in self.columns:
                self.columns['тираж_col'] = col
        
        # ГРУППОВОЕ СКАНИРОВАНИЕ для маршрутов доставки
        # Ищем группы типа "Доставка ЖД", "Доставка Авиа" и сканируем следующие колонки
        if has_subheaders:
            self._detect_route_groups(ws)
        
        # Если нет подзаголовков, используем старый метод
        else:
            self._detect_simple_prices(ws)
        
        return len(self.columns) > 0
    
    def _detect_route_groups(self, ws):
        """Детектит группы маршрутов с подзаголовками"""
        for col in range(1, ws.max_column + 1):
            header2 = ws.cell(2, col).value
            if not header2:
                continue
            
            header2_lower = str(header2).lower().strip()
            
            # Ищем маршруты
            route_name = None
            prefix = None
            
            if 'долгое' in header2_lower and 'жд' in header2_lower:
                route_name = 'Долгое ЖД'
                prefix = 'long_jd'
            elif 'жд' in header2_lower and 'долгое' not in header2_lower:
                route_name = 'ЖД'
                prefix = 'jd'
            elif 'авиа' in header2_lower or 'avia' in header2_lower:
                route_name = 'Авиа'
                prefix = 'avia'
            
            # Если нашли маршрут, сканируем следующие 5 колонок
            if prefix:
                for offset in range(6):  # Проверяем текущую и следующие 5 колонок
                    check_col = col + offset
                    if check_col > ws.max_column:
                        break
                    
                    header3 = ws.cell(3, check_col).value
                    if not header3:
                        continue
                    
                    header3_lower = str(header3).lower().strip()
                    
                    # Ищем подзаголовки
                    if ('цена' in header3_lower or 'price' in header3_lower) and ('$' in header3_lower or 'usd' in header3_lower):
                        if f'{prefix}_usd' not in self.columns:
                            self.columns[f'{prefix}_usd'] = check_col
                    
                    elif ('цена' in header3_lower or 'price' in header3_lower) and ('руб' in header3_lower or 'rub' in header3_lower):
                        if f'{prefix}_rub' not in self.columns:
                            self.columns[f'{prefix}_rub'] = check_col
                    
                    elif 'срок' in header3_lower or 'delivery' in header3_lower or 'к.д' in header3_lower:
                        if f'{prefix}_delivery' not in self.columns:
                            self.columns[f'{prefix}_delivery'] = check_col
    
    def _detect_simple_prices(self, ws):
        """Детектит простые цены без подзаголовков (для файлов без строки 3)"""
        for col in range(1, ws.max_column + 1):
            header2 = ws.cell(2, col).value
            if not header2:
                continue
            
            header_lower = str(header2).lower().strip()
            
            # Цены и сроки по маршрутам (старый метод для простых файлов)
            # Долгое ЖД
            if ('долгое' in header_lower or 'доставка жд долгое' in header_lower) and 'жд' in header_lower:
                if '$' in header_lower or 'usd' in header_lower or 'цена за шт., $' in header_lower:
                    self.columns['long_jd_usd'] = col
                elif 'руб' in header_lower or 'rub' in header_lower or 'цена за шт., руб' in header_lower:
                    self.columns['long_jd_rub'] = col
                elif 'срок' in header_lower or 'delivery' in header_lower:
                    self.columns['long_jd_delivery'] = col
            
            # ЖД (обычное, не долгое)
            elif 'жд' in header_lower and 'долгое' not in header_lower:
                if '$' in header_lower or 'usd' in header_lower or 'цена за шт., $' in header_lower:
                    self.columns['jd_usd'] = col
                elif 'руб' in header_lower or 'rub' in header_lower or 'цена за шт., руб' in header_lower:
                    self.columns['jd_rub'] = col
                elif 'срок' in header_lower or 'delivery' in header_lower or 'к.д' in header_lower:
                    self.columns['jd_delivery'] = col
            
            # АВИА
            elif 'авиа' in header_lower or 'avia' in header_lower:
                if '$' in header_lower or 'usd' in header_lower or 'цена за шт., $' in header_lower:
                    self.columns['avia_usd'] = col
                elif 'руб' in header_lower or 'rub' in header_lower or 'цена за шт., руб' in header_lower:
                    self.columns['avia_rub'] = col
                elif 'срок' in header_lower or 'delivery' in header_lower or 'к.д' in header_lower:
                    self.columns['avia_delivery'] = col
            
            # Универсальные цена/срок (для 27 столбцов без указания маршрута)
            elif 'цена' in header_lower and 'шт' in header_lower:
                if '$' in header_lower and 'price_usd' not in self.columns:
                    self.columns['price_usd'] = col
                elif 'руб' in header_lower and 'price_rub' not in self.columns:
                    self.columns['price_rub'] = col
            
            elif 'срок тиража' in header_lower and 'delivery' not in self.columns:
                self.columns['delivery'] = col
        
        return len(self.columns) > 0
    
    def _get_cell_value(self, ws, row, col_name, default=None):
        """Безопасно получает значение ячейки по имени столбца"""
        if col_name not in self.columns:
            return default
        
        col = self.columns[col_name]
        value = ws.cell(row, col).value
        
        return value if value is not None else default
    
    def _parse_quantity(self, value):
        """Парсит тираж"""
        if not value:
            return None
        
        # Преобразуем число в int (напр. 10000.0 -> 10000)
        if isinstance(value, (int, float)):
            return int(float(value))
        
        # Убираем пробелы и запятые
        value_str = str(value).replace(' ', '').replace(',', '')
        
        # Извлекаем число
        match = re.search(r'(\d+)', value_str)
        return int(match.group(1)) if match else None
    
    def _parse_price(self, value):
        """Парсит цену"""
        if not value:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        # Убираем валюту, пробелы
        value_str = str(value).replace('₽', '').replace('$', '').replace(' ', '').replace(',', '.')
        
        try:
            return float(value_str)
        except:
            return None
    
    def _parse_delivery_time(self, value):
        """Парсит срок доставки"""
        if not value:
            return None
        
        value_str = str(value).lower()
        
        # Ищем число (берем первое или среднее из диапазона)
        match = re.search(r'(\d+)[-–—]?(\d+)?', value_str)
        if match:
            num1 = int(match.group(1))
            num2 = int(match.group(2)) if match.group(2) else num1
            return (num1 + num2) // 2  # Среднее значение
        
        return None
    
    def _extract_images(self, ws_images, row, table_id):
        """Извлекает изображения из строки товара (старый метод, оставлен для совместимости)"""
        return self._extract_images_from_range(ws_images, row, row, table_id)
    
    def _extract_images_from_range(self, ws_images, start_row, end_row, table_id):
        """Извлекает изображения из диапазона строк товара"""
        images = []
        
        # Проходим по всем строкам диапазона
        for row in range(start_row, end_row + 1):
            # Главное фото (колонка 'photo')
            photo_col = self.columns.get('photo', 1)
            row_images = self._extract_all_images_from_cell(
                ws_images, row, photo_col, table_id, is_main=(row == start_row)  # Первая строка = main
            )
            images.extend(row_images)
            
            # Дополнительные фото
            if 'extra_photo' in self.columns:
                extra_col = self.columns['extra_photo']
                extra_images = self._extract_all_images_from_cell(
                    ws_images, row, extra_col, table_id, is_main=False
                )
                images.extend(extra_images)
        
        return images
    
    def _extract_all_images_from_cell(self, ws, row, col, table_id, is_main=False):
        """Извлекает ВСЕ изображения из ячейки"""
        images = []
        cell_images = []
        
        # Находим все изображения в ячейке
        for img in ws._images:
            if hasattr(img.anchor, '_from'):
                img_row = img.anchor._from.row + 1
                img_col = img.anchor._from.col + 1
                
                if img_row == row and img_col == col:
                    cell_images.append(img)
        
        # Обрабатываем найденные изображения
        for idx, img in enumerate(cell_images):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            cell_pos = f"{col_letter}{row}"
            
            # Генерируем уникальное имя файла
            suffix = f"_{idx + 1}_" if idx > 0 else "_"
            img_filename = f"{table_id}_{cell_pos}{suffix}{abs(hash(img.ref))}.png"
            
            # Получаем данные изображения
            img_data = img.ref
            if hasattr(img_data, 'read'):
                img_data.seek(0)
                img_bytes = img_data.read()
            elif isinstance(img_data, bytes):
                img_bytes = img_data
            else:
                continue
            
            images.append({
                'filename': img_filename,
                'data': img_bytes,
                'cell_position': cell_pos,
                'is_main': is_main and idx == 0,  # Только первое = главное
                'row': row
            })
        
        return images
    
    def _parse_products(self, ws_data, ws_images, project_id, table_id):
        """Парсит все товары из файла"""
        
        # Автоматически определяем структуру
        if not self._detect_columns(ws_data):
            raise Exception("Не удалось определить структуру файла!")
        
        # Динамически определяем строку начала данных
        self.data_start_row = self._detect_data_start_row(ws_data)
        
        print(f"\n📋 Обнаруженные столбцы:")
        for name, col in sorted(self.columns.items(), key=lambda x: x[1]):
            print(f"  • {name}: колонка {col}")
        print(f"📍 Данные начинаются со строки: {self.data_start_row}")
        
        products = []
        current_product = None
        product_start_row = None  # Отслеживаем начало товара
        
        for row in range(self.data_start_row, ws_data.max_row + 1):
            # Читаем основные поля
            name = self._get_cell_value(ws_data, row, 'name')
            
            # Если есть название - новый товар
            if name:
                # Сохраняем предыдущий товар И извлекаем его изображения из диапазона строк
                if current_product and product_start_row:
                    # Извлекаем изображения из диапазона [product_start_row, row-1]
                    current_product['images'] = self._extract_images_from_range(
                        ws_images, product_start_row, row - 1, table_id
                    )
                    products.append(current_product)
                
                product_start_row = row  # Запоминаем начало нового товара
                
                # Создаем новый товар
                description = self._get_cell_value(ws_data, row, 'description', '')
                
                # Для 33-столбцовой структуры может быть ТЗ и Описание ММ
                tz = self._get_cell_value(ws_data, row, 'tz', '')
                desc_mm = self._get_cell_value(ws_data, row, 'description_mm', '')
                
                # Объединяем описания
                full_desc = '\n'.join(filter(None, [tz, description, desc_mm]))
                
                custom = self._get_cell_value(ws_data, row, 'custom', '')
                
                # Образец
                sample_price = self._parse_price(self._get_cell_value(ws_data, row, 'sample_price'))
                sample_delivery = self._parse_delivery_time(self._get_cell_value(ws_data, row, 'sample_delivery'))
                
                current_product = {
                    'name': str(name).strip(),
                    'description': full_desc.strip() if full_desc else '',
                    'custom_field': str(custom).strip() if custom else '',
                    'sample_price': sample_price,
                    'sample_delivery_time': sample_delivery,
                    'row_number': row,
                    'images': [],  # Заполним потом из диапазона
                    'offers': []
                }
            
            # Парсим офферы (есть тираж = есть оффер)
            quantity = self._parse_quantity(self._get_cell_value(ws_data, row, 'тираж_col'))
            
            if quantity:
                offers = self._parse_offers_from_row(ws_data, row, quantity)
                if current_product:
                    current_product['offers'].extend(offers)
        
        # Сохраняем последний товар (извлекаем изображения до конца файла)
        if current_product and product_start_row:
            current_product['images'] = self._extract_images_from_range(
                ws_images, product_start_row, ws_data.max_row, table_id
            )
            products.append(current_product)
        
        return products
    
    def _parse_offers_from_row(self, ws, row, quantity):
        """Парсит все офферы из строки"""
        offers = []
        
        # Проверяем все возможные маршруты
        routes_to_check = [
            ('long_jd_usd', 'long_jd_rub', 'long_jd_delivery', 'Долгое ЖД'),
            ('jd_usd', 'jd_rub', 'jd_delivery', 'ЖД'),
            ('avia_usd', 'avia_rub', 'avia_delivery', 'Авиа'),
            ('price_usd', 'price_rub', 'delivery', None)  # Универсальный (без маршрута)
        ]
        
        for usd_col, rub_col, del_col, route_name in routes_to_check:
            price_usd = self._parse_price(self._get_cell_value(ws, row, usd_col))
            price_rub = self._parse_price(self._get_cell_value(ws, row, rub_col))
            delivery = self._parse_delivery_time(self._get_cell_value(ws, row, del_col))
            
            # Если есть хотя бы цена - создаем оффер
            if price_usd or price_rub:
                offers.append({
                    'quantity': quantity,
                    'price_usd': price_usd,
                    'price_rub': price_rub,
                    'delivery_time_days': delivery,
                    'route': route_name
                })
        
        return offers
    
    def _save_product(self, product, project_id, table_id):
        """Сохраняет товар в БД"""
        try:
            with self.db.get_session() as session:
                # Вставка товара
                result = session.execute(text("""
                    INSERT INTO products (
                        project_id, table_id, name, description, custom_field,
                        sample_price, sample_delivery_time,
                        row_number, created_at, updated_at
                    ) VALUES (
                        :project_id, :table_id, :name, :description, :custom_field,
                        :sample_price, :sample_delivery_time,
                        :row_number, NOW(), NOW()
                    ) RETURNING id
                """), {
                    'project_id': project_id,
                    'table_id': table_id,
                    'name': product['name'],
                    'description': product['description'],
                    'custom_field': product['custom_field'],
                    'sample_price': product.get('sample_price'),
                    'sample_delivery_time': product.get('sample_delivery_time'),
                    'row_number': product['row_number']
                })
                
                product_id = result.fetchone()[0]
                session.commit()
                
                # Сохраняем изображения
                for img in product['images']:
                    # Сохраняем файл
                    img_path = self.storage_dir / img['filename']
                    with open(img_path, 'wb') as f:
                        f.write(img['data'])
                    
                    # Записываем в БД
                    session.execute(text("""
                        INSERT INTO product_images (
                            product_id, table_id, image_filename, local_path, image_url,
                            cell_position, is_main_image, row_number, created_at, updated_at
                        )
                        VALUES (:product_id, :table_id, :filename, :local_path, :image_url,
                                :cell_position, :is_main, :row_number, NOW(), NOW())
                    """), {
                        'product_id': product_id,
                        'table_id': table_id,
                        'filename': img['filename'],
                        'local_path': str(img_path),
                        'image_url': str(img_path),  # Пока локальный путь, потом обновим на FTP
                        'cell_position': img['cell_position'],
                        'is_main': img['is_main'],
                        'row_number': img['row']
                    })
                    session.commit()
                
                # Сохраняем офферы
                for offer in product['offers']:
                    session.execute(text("""
                        INSERT INTO price_offers (
                            product_id, quantity, price_usd, price_rub,
                            route, delivery_time_days, created_at, updated_at
                        )
                        VALUES (:product_id, :quantity, :price_usd, :price_rub,
                                :route, :delivery_days, NOW(), NOW())
                    """), {
                        'product_id': product_id,
                        'quantity': offer['quantity'],
                        'price_usd': offer['price_usd'],
                        'price_rub': offer['price_rub'],
                        'route': offer['route'],
                        'delivery_days': offer['delivery_time_days']
                    })
                    session.commit()
                
                return True
                
        except Exception as e:
            print(f"    ❌ Ошибка сохранения товара '{product['name']}': {e}")
            return False
    
    def parse_project(self, project_id):
        """Парсит один проект"""
        try:
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx')) + \
                           list(excel_dir.glob(f'project_{project_id}_*.xls'))
            
            if not matching_files:
                return {'success': False, 'error': 'File not found'}
            
            file_path = matching_files[0]
            
            # Парсим
            wb_data = load_workbook(str(file_path), data_only=True)
            wb_images = load_workbook(str(file_path), data_only=False)
            
            ws_data = wb_data.active
            ws_images = wb_images.active
            
            # Получаем table_id из БД
            with self.db.get_session() as session:
                result = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).fetchone()
                
                if not result:
                    return {'success': False, 'error': 'Project not in DB'}
                
                table_id = result[0]
            
            # Парсим товары
            products = self._parse_products(ws_data, ws_images, project_id, table_id)
            
            wb_data.close()
            wb_images.close()
            
            # Сохраняем в БД
            saved_products = 0
            saved_images = 0
            saved_offers = 0
            
            for product in products:
                if self._save_product(product, project_id, table_id):
                    saved_products += 1
                    saved_images += len(product['images'])
                    saved_offers += len(product['offers'])
            
            # Обновляем статус проекта
            with self.db.get_session() as session:
                session.execute(text("""
                    UPDATE projects
                    SET parsing_status = 'complete',
                        total_products_found = :products,
                        total_images_found = :images,
                        updated_at = NOW()
                    WHERE id = :pid
                """), {
                    'products': saved_products,
                    'images': saved_images,
                    'pid': project_id
                })
                session.commit()
            
            return {
                'success': True,
                'products': saved_products,
                'images': saved_images,
                'offers': saved_offers
            }
            
        except Exception as e:
            # Отмечаем ошибку в БД
            with self.db.get_session() as session:
                session.execute(text("""
                    UPDATE projects
                    SET parsing_status = 'error',
                        updated_at = NOW()
                    WHERE id = :pid
                """), {'pid': project_id})
                session.commit()
            
            return {'success': False, 'error': str(e)}
    
    def reparse_images_only(self, project_id, excel_path):
        """Допарсивает ТОЛЬКО изображения для существующих товаров"""
        try:
            # Открываем файл
            wb_data = load_workbook(excel_path, data_only=True)
            wb_images = load_workbook(excel_path, data_only=False)
            ws_data = wb_data.active
            ws_images = wb_images.active
            
            # Получаем table_id
            with self.db.get_session() as session:
                table_id = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).scalar()
            
            # Парсим товары (чтобы получить диапазоны строк и изображения)
            if not self._detect_columns(ws_data):
                return {'success': False, 'images': 0, 'error': 'Не удалось определить структуру'}
            
            self.data_start_row = self._detect_data_start_row(ws_data)
            
            # Получаем существующие товары из БД
            with self.db.get_session() as session:
                existing_products = session.execute(text("""
                    SELECT id, row_number, name
                    FROM products
                    WHERE project_id = :pid
                    ORDER BY row_number
                """), {'pid': project_id}).fetchall()
            
            if not existing_products:
                return {'success': False, 'images': 0, 'error': 'Нет товаров в БД'}
            
            # Создаем карту: row_number -> product_id
            product_map = {}
            for prod_id, row_num, name in existing_products:
                product_map[row_num] = {'id': prod_id, 'name': name}
            
            # Определяем диапазоны строк для каждого товара
            sorted_rows = sorted(product_map.keys())
            saved_images = 0
            
            for i, start_row in enumerate(sorted_rows):
                # Конец диапазона = начало следующего товара - 1
                end_row = sorted_rows[i + 1] - 1 if i < len(sorted_rows) - 1 else ws_data.max_row
                
                # Извлекаем изображения из диапазона
                images = self._extract_images_from_range(ws_images, start_row, end_row, table_id)
                
                # Сохраняем изображения
                product_id = product_map[start_row]['id']
                
                with self.db.get_session() as session:
                    for img in images:
                        # Проверяем существует ли уже такое изображение для этого товара
                        exists = session.execute(text("""
                            SELECT COUNT(*) 
                            FROM product_images 
                            WHERE product_id = :product_id 
                            AND image_filename = :filename
                        """), {
                            'product_id': product_id,
                            'filename': img['filename']
                        }).scalar()
                        
                        if exists > 0:
                            # Изображение уже есть, пропускаем
                            continue
                        
                        # Сохраняем файл
                        img_path = self.storage_dir / img['filename']
                        with open(img_path, 'wb') as f:
                            f.write(img['data'])
                        
                        # Записываем в БД
                        session.execute(text("""
                            INSERT INTO product_images (
                                product_id, table_id, image_filename, image_url,
                                cell_position, is_main_image, row_number, created_at, updated_at
                            )
                            VALUES (:product_id, :table_id, :filename, :url,
                                    :cell_position, :is_main, :row_number, NOW(), NOW())
                        """), {
                            'product_id': product_id,
                            'table_id': table_id,
                            'filename': img['filename'],
                            'url': str(img_path),
                            'cell_position': img['cell_position'],
                            'is_main': img['is_main'],
                            'row_number': img['row']
                        })
                        session.commit()
                        saved_images += 1
            
            wb_data.close()
            wb_images.close()
            
            return {'success': True, 'images': saved_images}
        
        except Exception as e:
            return {'success': False, 'images': 0, 'error': str(e)}


def main():
    """Тестовый запуск"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python parse_template_6.py <project_id>")
        print("Example: python parse_template_6.py 2101")
        return
    
    project_id = int(sys.argv[1])
    
    print("=" * 80)
    print("🚀 ПАРСИНГ ШАБЛОНА 6 (заголовки в строке 2)")
    print("=" * 80)
    print(f"\n📋 Проект ID: {project_id}")
    
    parser = Template6Parser()
    result = parser.parse_project(project_id)
    
    print(f"\n{'=' * 80}")
    
    if result['success']:
        print("✅ УСПЕШНО!")
        print(f"\n📊 Результаты:")
        print(f"  • Товары: {result['products']}")
        print(f"  • Изображения: {result['images']}")
        print(f"  • Предложения: {result['offers']}")
    else:
        print("❌ ОШИБКА!")
        print(f"  {result['error']}")
    
    print("=" * 80)


if __name__ == '__main__':
    main()


Универсальный парсер для Шаблона 6 (заголовки в строке 2)

Поддерживает:
- 27 столбцов (222 файла)
- 30 столбцов (179 файлов)
- 33 столбца (34 файла)
- ИТОГО: 435 файлов (65% покрытие!)

Особенности:
- Строка 1: информация о менеджере (пропускается)
- Строка 2: заголовки товаров
- Строка 3+: данные
- Автоматическое определение столбцов по названиям
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import re
import hashlib

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template6Parser:
    """
    Универсальный парсер для Шаблона 6
    Автоматически определяет структуру по заголовкам в строке 2
    """
    
    HEADER_ROW = 2  # Заголовки всегда в строке 2!
    
    def __init__(self):
        self.db = PostgreSQLManager()
        self.storage_dir = Path('storage/images')
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        
        # Словарь для маппинга столбцов (заполняется автоматически)
        self.columns = {}
        self.data_start_row = None  # Определяется динамически!
    
    def _detect_data_start_row(self, ws):
        """Динамически определяет с какой строки начинаются данные"""
        # Проверяем строки 3 и 4
        for row in [3, 4]:
            # Получаем значения из первых столбцов
            name_col = 2  # Обычно наименование во 2-м столбце
            name_val = ws.cell(row, name_col).value
            
            # Если строка содержит подзаголовки типа "Цена за шт." - пропускаем
            if name_val:
                name_str = str(name_val).lower()
                # Ключевые слова подзаголовков
                subheader_keywords = ['цена за шт', 'price', 'итого', 'срок тиража', 'delivery']
                
                if any(kw in name_str for kw in subheader_keywords):
                    continue  # Это подзаголовок, пропускаем
                else:
                    # Это реальные данные
                    return row
        
        # По умолчанию строка 3
        return 3
    
    def _detect_columns(self, ws):
        """Автоматически определяет индексы столбцов по названиям в строке 2 и 3
        
        Использует групповое сканирование:
        1. Находит группу маршрута в строке 2 (например "Доставка ЖД")
        2. Сканирует следующие N колонок для подзаголовков в строке 3
        3. Привязывает их к маршруту
        """
        self.columns = {}
        
        # Проверяем есть ли строка 3 с подзаголовками
        has_subheaders = False
        for col in range(1, min(10, ws.max_column + 1)):
            val3 = ws.cell(3, col).value
            if val3 and ('цена' in str(val3).lower() or 'price' in str(val3).lower()):
                has_subheaders = True
                break
        
        # Основные поля (не требуют группового сканирования)
        for col in range(1, ws.max_column + 1):
            header2 = ws.cell(self.HEADER_ROW, col).value
            header2_str = str(header2).lower().strip() if header2 else ''
            
            header3 = ws.cell(3, col).value if has_subheaders else None
            header3_str = str(header3).lower().strip() if header3 else ''
            
            combined_header = f"{header2_str} {header3_str}".strip()
            header_lower = combined_header if combined_header else header2_str
            
            # Основные поля
            if 'фото' in header_lower or 'photo' in header_lower:
                if 'доп' not in header_lower and 'extra' not in header_lower:
                    self.columns['photo'] = col
                else:
                    self.columns['extra_photo'] = col
            
            elif 'наименование' in header_lower or ('name' in header_lower and 'product' not in header_lower):
                self.columns['name'] = col
            
            elif 'описание мм' in header_lower or 'описание ММ' in header_lower:
                self.columns['description_mm'] = col
            
            elif 'описание' in header_lower or 'description' in header_lower:
                if 'description' not in self.columns:  # Первое описание
                    self.columns['description'] = col
            
            elif 'тз' in header_lower and len(header_lower) < 10:
                self.columns['tz'] = col
            
            elif 'кастом' in header_lower or 'custom' in header_lower or 'дизайн' in header_lower:
                self.columns['custom'] = col
            
            elif 'образец' in header_lower and 'срок' not in header_lower:
                self.columns['sample_price'] = col
            
            elif 'образец' in header_lower and 'срок' in header_lower:
                self.columns['sample_delivery'] = col
            
            # Тираж (ищем первый столбец с "тираж" или "quantity")
            elif ('тираж' in header_lower or 'quantity' in header_lower) and 'тираж_col' not in self.columns:
                self.columns['тираж_col'] = col
        
        # ГРУППОВОЕ СКАНИРОВАНИЕ для маршрутов доставки
        # Ищем группы типа "Доставка ЖД", "Доставка Авиа" и сканируем следующие колонки
        if has_subheaders:
            self._detect_route_groups(ws)
        
        # Если нет подзаголовков, используем старый метод
        else:
            self._detect_simple_prices(ws)
        
        return len(self.columns) > 0
    
    def _detect_route_groups(self, ws):
        """Детектит группы маршрутов с подзаголовками"""
        for col in range(1, ws.max_column + 1):
            header2 = ws.cell(2, col).value
            if not header2:
                continue
            
            header2_lower = str(header2).lower().strip()
            
            # Ищем маршруты
            route_name = None
            prefix = None
            
            if 'долгое' in header2_lower and 'жд' in header2_lower:
                route_name = 'Долгое ЖД'
                prefix = 'long_jd'
            elif 'жд' in header2_lower and 'долгое' not in header2_lower:
                route_name = 'ЖД'
                prefix = 'jd'
            elif 'авиа' in header2_lower or 'avia' in header2_lower:
                route_name = 'Авиа'
                prefix = 'avia'
            
            # Если нашли маршрут, сканируем следующие 5 колонок
            if prefix:
                for offset in range(6):  # Проверяем текущую и следующие 5 колонок
                    check_col = col + offset
                    if check_col > ws.max_column:
                        break
                    
                    header3 = ws.cell(3, check_col).value
                    if not header3:
                        continue
                    
                    header3_lower = str(header3).lower().strip()
                    
                    # Ищем подзаголовки
                    if ('цена' in header3_lower or 'price' in header3_lower) and ('$' in header3_lower or 'usd' in header3_lower):
                        if f'{prefix}_usd' not in self.columns:
                            self.columns[f'{prefix}_usd'] = check_col
                    
                    elif ('цена' in header3_lower or 'price' in header3_lower) and ('руб' in header3_lower or 'rub' in header3_lower):
                        if f'{prefix}_rub' not in self.columns:
                            self.columns[f'{prefix}_rub'] = check_col
                    
                    elif 'срок' in header3_lower or 'delivery' in header3_lower or 'к.д' in header3_lower:
                        if f'{prefix}_delivery' not in self.columns:
                            self.columns[f'{prefix}_delivery'] = check_col
    
    def _detect_simple_prices(self, ws):
        """Детектит простые цены без подзаголовков (для файлов без строки 3)"""
        for col in range(1, ws.max_column + 1):
            header2 = ws.cell(2, col).value
            if not header2:
                continue
            
            header_lower = str(header2).lower().strip()
            
            # Цены и сроки по маршрутам (старый метод для простых файлов)
            # Долгое ЖД
            if ('долгое' in header_lower or 'доставка жд долгое' in header_lower) and 'жд' in header_lower:
                if '$' in header_lower or 'usd' in header_lower or 'цена за шт., $' in header_lower:
                    self.columns['long_jd_usd'] = col
                elif 'руб' in header_lower or 'rub' in header_lower or 'цена за шт., руб' in header_lower:
                    self.columns['long_jd_rub'] = col
                elif 'срок' in header_lower or 'delivery' in header_lower:
                    self.columns['long_jd_delivery'] = col
            
            # ЖД (обычное, не долгое)
            elif 'жд' in header_lower and 'долгое' not in header_lower:
                if '$' in header_lower or 'usd' in header_lower or 'цена за шт., $' in header_lower:
                    self.columns['jd_usd'] = col
                elif 'руб' in header_lower or 'rub' in header_lower or 'цена за шт., руб' in header_lower:
                    self.columns['jd_rub'] = col
                elif 'срок' in header_lower or 'delivery' in header_lower or 'к.д' in header_lower:
                    self.columns['jd_delivery'] = col
            
            # АВИА
            elif 'авиа' in header_lower or 'avia' in header_lower:
                if '$' in header_lower or 'usd' in header_lower or 'цена за шт., $' in header_lower:
                    self.columns['avia_usd'] = col
                elif 'руб' in header_lower or 'rub' in header_lower or 'цена за шт., руб' in header_lower:
                    self.columns['avia_rub'] = col
                elif 'срок' in header_lower or 'delivery' in header_lower or 'к.д' in header_lower:
                    self.columns['avia_delivery'] = col
            
            # Универсальные цена/срок (для 27 столбцов без указания маршрута)
            elif 'цена' in header_lower and 'шт' in header_lower:
                if '$' in header_lower and 'price_usd' not in self.columns:
                    self.columns['price_usd'] = col
                elif 'руб' in header_lower and 'price_rub' not in self.columns:
                    self.columns['price_rub'] = col
            
            elif 'срок тиража' in header_lower and 'delivery' not in self.columns:
                self.columns['delivery'] = col
        
        return len(self.columns) > 0
    
    def _get_cell_value(self, ws, row, col_name, default=None):
        """Безопасно получает значение ячейки по имени столбца"""
        if col_name not in self.columns:
            return default
        
        col = self.columns[col_name]
        value = ws.cell(row, col).value
        
        return value if value is not None else default
    
    def _parse_quantity(self, value):
        """Парсит тираж"""
        if not value:
            return None
        
        # Преобразуем число в int (напр. 10000.0 -> 10000)
        if isinstance(value, (int, float)):
            return int(float(value))
        
        # Убираем пробелы и запятые
        value_str = str(value).replace(' ', '').replace(',', '')
        
        # Извлекаем число
        match = re.search(r'(\d+)', value_str)
        return int(match.group(1)) if match else None
    
    def _parse_price(self, value):
        """Парсит цену"""
        if not value:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        # Убираем валюту, пробелы
        value_str = str(value).replace('₽', '').replace('$', '').replace(' ', '').replace(',', '.')
        
        try:
            return float(value_str)
        except:
            return None
    
    def _parse_delivery_time(self, value):
        """Парсит срок доставки"""
        if not value:
            return None
        
        value_str = str(value).lower()
        
        # Ищем число (берем первое или среднее из диапазона)
        match = re.search(r'(\d+)[-–—]?(\d+)?', value_str)
        if match:
            num1 = int(match.group(1))
            num2 = int(match.group(2)) if match.group(2) else num1
            return (num1 + num2) // 2  # Среднее значение
        
        return None
    
    def _extract_images(self, ws_images, row, table_id):
        """Извлекает изображения из строки товара (старый метод, оставлен для совместимости)"""
        return self._extract_images_from_range(ws_images, row, row, table_id)
    
    def _extract_images_from_range(self, ws_images, start_row, end_row, table_id):
        """Извлекает изображения из диапазона строк товара"""
        images = []
        
        # Проходим по всем строкам диапазона
        for row in range(start_row, end_row + 1):
            # Главное фото (колонка 'photo')
            photo_col = self.columns.get('photo', 1)
            row_images = self._extract_all_images_from_cell(
                ws_images, row, photo_col, table_id, is_main=(row == start_row)  # Первая строка = main
            )
            images.extend(row_images)
            
            # Дополнительные фото
            if 'extra_photo' in self.columns:
                extra_col = self.columns['extra_photo']
                extra_images = self._extract_all_images_from_cell(
                    ws_images, row, extra_col, table_id, is_main=False
                )
                images.extend(extra_images)
        
        return images
    
    def _extract_all_images_from_cell(self, ws, row, col, table_id, is_main=False):
        """Извлекает ВСЕ изображения из ячейки"""
        images = []
        cell_images = []
        
        # Находим все изображения в ячейке
        for img in ws._images:
            if hasattr(img.anchor, '_from'):
                img_row = img.anchor._from.row + 1
                img_col = img.anchor._from.col + 1
                
                if img_row == row and img_col == col:
                    cell_images.append(img)
        
        # Обрабатываем найденные изображения
        for idx, img in enumerate(cell_images):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            cell_pos = f"{col_letter}{row}"
            
            # Генерируем уникальное имя файла
            suffix = f"_{idx + 1}_" if idx > 0 else "_"
            img_filename = f"{table_id}_{cell_pos}{suffix}{abs(hash(img.ref))}.png"
            
            # Получаем данные изображения
            img_data = img.ref
            if hasattr(img_data, 'read'):
                img_data.seek(0)
                img_bytes = img_data.read()
            elif isinstance(img_data, bytes):
                img_bytes = img_data
            else:
                continue
            
            images.append({
                'filename': img_filename,
                'data': img_bytes,
                'cell_position': cell_pos,
                'is_main': is_main and idx == 0,  # Только первое = главное
                'row': row
            })
        
        return images
    
    def _parse_products(self, ws_data, ws_images, project_id, table_id):
        """Парсит все товары из файла"""
        
        # Автоматически определяем структуру
        if not self._detect_columns(ws_data):
            raise Exception("Не удалось определить структуру файла!")
        
        # Динамически определяем строку начала данных
        self.data_start_row = self._detect_data_start_row(ws_data)
        
        print(f"\n📋 Обнаруженные столбцы:")
        for name, col in sorted(self.columns.items(), key=lambda x: x[1]):
            print(f"  • {name}: колонка {col}")
        print(f"📍 Данные начинаются со строки: {self.data_start_row}")
        
        products = []
        current_product = None
        product_start_row = None  # Отслеживаем начало товара
        
        for row in range(self.data_start_row, ws_data.max_row + 1):
            # Читаем основные поля
            name = self._get_cell_value(ws_data, row, 'name')
            
            # Если есть название - новый товар
            if name:
                # Сохраняем предыдущий товар И извлекаем его изображения из диапазона строк
                if current_product and product_start_row:
                    # Извлекаем изображения из диапазона [product_start_row, row-1]
                    current_product['images'] = self._extract_images_from_range(
                        ws_images, product_start_row, row - 1, table_id
                    )
                    products.append(current_product)
                
                product_start_row = row  # Запоминаем начало нового товара
                
                # Создаем новый товар
                description = self._get_cell_value(ws_data, row, 'description', '')
                
                # Для 33-столбцовой структуры может быть ТЗ и Описание ММ
                tz = self._get_cell_value(ws_data, row, 'tz', '')
                desc_mm = self._get_cell_value(ws_data, row, 'description_mm', '')
                
                # Объединяем описания
                full_desc = '\n'.join(filter(None, [tz, description, desc_mm]))
                
                custom = self._get_cell_value(ws_data, row, 'custom', '')
                
                # Образец
                sample_price = self._parse_price(self._get_cell_value(ws_data, row, 'sample_price'))
                sample_delivery = self._parse_delivery_time(self._get_cell_value(ws_data, row, 'sample_delivery'))
                
                current_product = {
                    'name': str(name).strip(),
                    'description': full_desc.strip() if full_desc else '',
                    'custom_field': str(custom).strip() if custom else '',
                    'sample_price': sample_price,
                    'sample_delivery_time': sample_delivery,
                    'row_number': row,
                    'images': [],  # Заполним потом из диапазона
                    'offers': []
                }
            
            # Парсим офферы (есть тираж = есть оффер)
            quantity = self._parse_quantity(self._get_cell_value(ws_data, row, 'тираж_col'))
            
            if quantity:
                offers = self._parse_offers_from_row(ws_data, row, quantity)
                if current_product:
                    current_product['offers'].extend(offers)
        
        # Сохраняем последний товар (извлекаем изображения до конца файла)
        if current_product and product_start_row:
            current_product['images'] = self._extract_images_from_range(
                ws_images, product_start_row, ws_data.max_row, table_id
            )
            products.append(current_product)
        
        return products
    
    def _parse_offers_from_row(self, ws, row, quantity):
        """Парсит все офферы из строки"""
        offers = []
        
        # Проверяем все возможные маршруты
        routes_to_check = [
            ('long_jd_usd', 'long_jd_rub', 'long_jd_delivery', 'Долгое ЖД'),
            ('jd_usd', 'jd_rub', 'jd_delivery', 'ЖД'),
            ('avia_usd', 'avia_rub', 'avia_delivery', 'Авиа'),
            ('price_usd', 'price_rub', 'delivery', None)  # Универсальный (без маршрута)
        ]
        
        for usd_col, rub_col, del_col, route_name in routes_to_check:
            price_usd = self._parse_price(self._get_cell_value(ws, row, usd_col))
            price_rub = self._parse_price(self._get_cell_value(ws, row, rub_col))
            delivery = self._parse_delivery_time(self._get_cell_value(ws, row, del_col))
            
            # Если есть хотя бы цена - создаем оффер
            if price_usd or price_rub:
                offers.append({
                    'quantity': quantity,
                    'price_usd': price_usd,
                    'price_rub': price_rub,
                    'delivery_time_days': delivery,
                    'route': route_name
                })
        
        return offers
    
    def _save_product(self, product, project_id, table_id):
        """Сохраняет товар в БД"""
        try:
            with self.db.get_session() as session:
                # Вставка товара
                result = session.execute(text("""
                    INSERT INTO products (
                        project_id, table_id, name, description, custom_field,
                        sample_price, sample_delivery_time,
                        row_number, created_at, updated_at
                    ) VALUES (
                        :project_id, :table_id, :name, :description, :custom_field,
                        :sample_price, :sample_delivery_time,
                        :row_number, NOW(), NOW()
                    ) RETURNING id
                """), {
                    'project_id': project_id,
                    'table_id': table_id,
                    'name': product['name'],
                    'description': product['description'],
                    'custom_field': product['custom_field'],
                    'sample_price': product.get('sample_price'),
                    'sample_delivery_time': product.get('sample_delivery_time'),
                    'row_number': product['row_number']
                })
                
                product_id = result.fetchone()[0]
                session.commit()
                
                # Сохраняем изображения
                for img in product['images']:
                    # Сохраняем файл
                    img_path = self.storage_dir / img['filename']
                    with open(img_path, 'wb') as f:
                        f.write(img['data'])
                    
                    # Записываем в БД
                    session.execute(text("""
                        INSERT INTO product_images (
                            product_id, table_id, image_filename, local_path, image_url,
                            cell_position, is_main_image, row_number, created_at, updated_at
                        )
                        VALUES (:product_id, :table_id, :filename, :local_path, :image_url,
                                :cell_position, :is_main, :row_number, NOW(), NOW())
                    """), {
                        'product_id': product_id,
                        'table_id': table_id,
                        'filename': img['filename'],
                        'local_path': str(img_path),
                        'image_url': str(img_path),  # Пока локальный путь, потом обновим на FTP
                        'cell_position': img['cell_position'],
                        'is_main': img['is_main'],
                        'row_number': img['row']
                    })
                    session.commit()
                
                # Сохраняем офферы
                for offer in product['offers']:
                    session.execute(text("""
                        INSERT INTO price_offers (
                            product_id, quantity, price_usd, price_rub,
                            route, delivery_time_days, created_at, updated_at
                        )
                        VALUES (:product_id, :quantity, :price_usd, :price_rub,
                                :route, :delivery_days, NOW(), NOW())
                    """), {
                        'product_id': product_id,
                        'quantity': offer['quantity'],
                        'price_usd': offer['price_usd'],
                        'price_rub': offer['price_rub'],
                        'route': offer['route'],
                        'delivery_days': offer['delivery_time_days']
                    })
                    session.commit()
                
                return True
                
        except Exception as e:
            print(f"    ❌ Ошибка сохранения товара '{product['name']}': {e}")
            return False
    
    def parse_project(self, project_id):
        """Парсит один проект"""
        try:
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx')) + \
                           list(excel_dir.glob(f'project_{project_id}_*.xls'))
            
            if not matching_files:
                return {'success': False, 'error': 'File not found'}
            
            file_path = matching_files[0]
            
            # Парсим
            wb_data = load_workbook(str(file_path), data_only=True)
            wb_images = load_workbook(str(file_path), data_only=False)
            
            ws_data = wb_data.active
            ws_images = wb_images.active
            
            # Получаем table_id из БД
            with self.db.get_session() as session:
                result = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).fetchone()
                
                if not result:
                    return {'success': False, 'error': 'Project not in DB'}
                
                table_id = result[0]
            
            # Парсим товары
            products = self._parse_products(ws_data, ws_images, project_id, table_id)
            
            wb_data.close()
            wb_images.close()
            
            # Сохраняем в БД
            saved_products = 0
            saved_images = 0
            saved_offers = 0
            
            for product in products:
                if self._save_product(product, project_id, table_id):
                    saved_products += 1
                    saved_images += len(product['images'])
                    saved_offers += len(product['offers'])
            
            # Обновляем статус проекта
            with self.db.get_session() as session:
                session.execute(text("""
                    UPDATE projects
                    SET parsing_status = 'complete',
                        total_products_found = :products,
                        total_images_found = :images,
                        updated_at = NOW()
                    WHERE id = :pid
                """), {
                    'products': saved_products,
                    'images': saved_images,
                    'pid': project_id
                })
                session.commit()
            
            return {
                'success': True,
                'products': saved_products,
                'images': saved_images,
                'offers': saved_offers
            }
            
        except Exception as e:
            # Отмечаем ошибку в БД
            with self.db.get_session() as session:
                session.execute(text("""
                    UPDATE projects
                    SET parsing_status = 'error',
                        updated_at = NOW()
                    WHERE id = :pid
                """), {'pid': project_id})
                session.commit()
            
            return {'success': False, 'error': str(e)}
    
    def reparse_images_only(self, project_id, excel_path):
        """Допарсивает ТОЛЬКО изображения для существующих товаров"""
        try:
            # Открываем файл
            wb_data = load_workbook(excel_path, data_only=True)
            wb_images = load_workbook(excel_path, data_only=False)
            ws_data = wb_data.active
            ws_images = wb_images.active
            
            # Получаем table_id
            with self.db.get_session() as session:
                table_id = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).scalar()
            
            # Парсим товары (чтобы получить диапазоны строк и изображения)
            if not self._detect_columns(ws_data):
                return {'success': False, 'images': 0, 'error': 'Не удалось определить структуру'}
            
            self.data_start_row = self._detect_data_start_row(ws_data)
            
            # Получаем существующие товары из БД
            with self.db.get_session() as session:
                existing_products = session.execute(text("""
                    SELECT id, row_number, name
                    FROM products
                    WHERE project_id = :pid
                    ORDER BY row_number
                """), {'pid': project_id}).fetchall()
            
            if not existing_products:
                return {'success': False, 'images': 0, 'error': 'Нет товаров в БД'}
            
            # Создаем карту: row_number -> product_id
            product_map = {}
            for prod_id, row_num, name in existing_products:
                product_map[row_num] = {'id': prod_id, 'name': name}
            
            # Определяем диапазоны строк для каждого товара
            sorted_rows = sorted(product_map.keys())
            saved_images = 0
            
            for i, start_row in enumerate(sorted_rows):
                # Конец диапазона = начало следующего товара - 1
                end_row = sorted_rows[i + 1] - 1 if i < len(sorted_rows) - 1 else ws_data.max_row
                
                # Извлекаем изображения из диапазона
                images = self._extract_images_from_range(ws_images, start_row, end_row, table_id)
                
                # Сохраняем изображения
                product_id = product_map[start_row]['id']
                
                with self.db.get_session() as session:
                    for img in images:
                        # Проверяем существует ли уже такое изображение для этого товара
                        exists = session.execute(text("""
                            SELECT COUNT(*) 
                            FROM product_images 
                            WHERE product_id = :product_id 
                            AND image_filename = :filename
                        """), {
                            'product_id': product_id,
                            'filename': img['filename']
                        }).scalar()
                        
                        if exists > 0:
                            # Изображение уже есть, пропускаем
                            continue
                        
                        # Сохраняем файл
                        img_path = self.storage_dir / img['filename']
                        with open(img_path, 'wb') as f:
                            f.write(img['data'])
                        
                        # Записываем в БД
                        session.execute(text("""
                            INSERT INTO product_images (
                                product_id, table_id, image_filename, image_url,
                                cell_position, is_main_image, row_number, created_at, updated_at
                            )
                            VALUES (:product_id, :table_id, :filename, :url,
                                    :cell_position, :is_main, :row_number, NOW(), NOW())
                        """), {
                            'product_id': product_id,
                            'table_id': table_id,
                            'filename': img['filename'],
                            'url': str(img_path),
                            'cell_position': img['cell_position'],
                            'is_main': img['is_main'],
                            'row_number': img['row']
                        })
                        session.commit()
                        saved_images += 1
            
            wb_data.close()
            wb_images.close()
            
            return {'success': True, 'images': saved_images}
        
        except Exception as e:
            return {'success': False, 'images': 0, 'error': str(e)}


def main():
    """Тестовый запуск"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python parse_template_6.py <project_id>")
        print("Example: python parse_template_6.py 2101")
        return
    
    project_id = int(sys.argv[1])
    
    print("=" * 80)
    print("🚀 ПАРСИНГ ШАБЛОНА 6 (заголовки в строке 2)")
    print("=" * 80)
    print(f"\n📋 Проект ID: {project_id}")
    
    parser = Template6Parser()
    result = parser.parse_project(project_id)
    
    print(f"\n{'=' * 80}")
    
    if result['success']:
        print("✅ УСПЕШНО!")
        print(f"\n📊 Результаты:")
        print(f"  • Товары: {result['products']}")
        print(f"  • Изображения: {result['images']}")
        print(f"  • Предложения: {result['offers']}")
    else:
        print("❌ ОШИБКА!")
        print(f"  {result['error']}")
    
    print("=" * 80)


if __name__ == '__main__':
    main()

