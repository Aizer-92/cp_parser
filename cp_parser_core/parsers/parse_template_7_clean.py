#!/usr/bin/env python3
"""
Парсер для Шаблона 7 (Multiple Routes - 406 проектов)

АДАПТИРОВАН ИЗ TEMPLATE 4 С КЛЮЧЕВЫМИ ИЗМЕНЕНИЯМИ:
- Динамическое определение колонок (не фиксированные)
- Динамический поиск строки заголовков (1-3, не фиксированная 4)
- Поддержка 2+ маршрутов (не только ЖД/АВИА)
- Сохранение изображений ЛОКАЛЬНО (не FTP)

КРИТИЧЕСКИЕ ИСПРАВЛЕНИЯ:
- parse_quantity: int(float(value)) - правильная обработка 100.0 → 100
- Интервалы тиражей: берём ПЕРВОЕ число
- Интервалы сроков: берём ПОСЛЕДНЕЕ число
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import re
import hashlib

sys.path.insert(0, str(Path.cwd()))
from database.postgresql_manager import PostgreSQLManager
from sqlalchemy import text


class Template7Parser:
    """
    Парсер для Шаблона 7 (Multiple Routes)
    
    Особенности:
    - 406 проектов с 2+ маршрутами
    - ДИНАМИЧЕСКОЕ определение структуры
    - Multi-row товары (доп. тиражи без названия)
    - Поддержка русского и английского
    """
    
    # Паттерны для поиска колонок (динамические!)
    COLUMN_PATTERNS = {
        'photo': ['фото', 'photo', 'изображение'],
        'name': ['наименование', 'name', 'название', 'товар', 'product'],
        'description': ['характеристики', 'description', 'описание', 'spec'],
        'custom': ['кастом', 'custom', 'персонализация'],
        'quantity': ['тираж', 'quantity', 'qty', 'количество', 'circulation'],
        'sample': ['образец', 'sample']
    }
    
    # Паттерны для маршрутов
    ROUTE_PATTERNS = {
        'sea': ['sea', 'море', 'морской', 'долгое жд', 'контейнер', 'container'],
        'air': ['air', 'авиа', 'авио', 'долгое авиа'],
        'railway': ['жд', 'railway', 'ж.д', 'железная дорога'],
        'contract': ['контракт', 'contract']
    }
    
    def __init__(self):
        self.db = PostgreSQLManager()
        self.storage_dir = Path('storage/images')
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        
        # Динамические колонки (определяются для каждого файла!)
        self.columns = {}
        self.routes = {}
        self.header_row = None
        self.data_start_row = None
    
    def _find_header_and_detect_structure(self, ws):
        """
        ДИНАМИЧЕСКИ находит строку заголовков и определяет структуру
        
        Возвращает: True если структура найдена
        """
        # Ищем строку с заголовками (1-3)
        for row_idx in range(1, 4):
            row_values = [str(cell.value or '').strip().lower() for cell in ws[row_idx]]
            
            # Проверяем наличие ключевых колонок
            has_name = any(pattern in val for val in row_values for pattern in self.COLUMN_PATTERNS['name'])
            has_quantity = any(pattern in val for val in row_values for pattern in self.COLUMN_PATTERNS['quantity'])
            
            if has_name and has_quantity:
                self.header_row = row_idx
                self.data_start_row = row_idx + 1
                print(f"   📍 Найдена строка заголовков: {row_idx}")
                break
        
        if not self.header_row:
            return False
        
        # Получаем заголовки
        header_values = [str(cell.value or '').strip() for cell in ws[self.header_row]]
        
        # Определяем колонки
        for col_idx, header in enumerate(header_values, start=1):
            header_lower = header.lower()
            
            # Основные колонки
            for col_type, patterns in self.COLUMN_PATTERNS.items():
                if any(pattern in header_lower for pattern in patterns):
                    if col_type not in self.columns:  # берём первое совпадение
                        self.columns[col_type] = col_idx
                        print(f"   ✓ {col_type}: колонка {col_idx} ({header})")
            
            # Маршруты (находим группы колонок для каждого маршрута)
            for route_key, patterns in self.ROUTE_PATTERNS.items():
                if any(pattern in header_lower for pattern in patterns):
                    if route_key not in self.routes:
                        self.routes[route_key] = {'columns': [], 'start_col': col_idx}
                        print(f"   🚢 {route_key}: колонка {col_idx} ({header})")
                        
                        # ВАЖНО: Проверяем соседние колонки справа (до следующего маршрута)
                        # Обычно структура: [USD, RUB, сро к, дополнительные]
                        next_route_col = 999
                        for other_route, other_patterns in self.ROUTE_PATTERNS.items():
                            if other_route != route_key:
                                for check_col in range(col_idx + 1, min(col_idx + 10, len(header_values))):
                                    check_header = header_values[check_col].lower() if check_col < len(header_values) else ""
                                    if any(p in check_header for p in other_patterns):
                                        next_route_col = min(next_route_col, check_col + 1)
                                        break
                        
                        # Добавляем текущую колонку и до 4 соседних (обычно USD, RUB, срок, +1)
                        max_cols = min(col_idx + 5, next_route_col, len(header_values) + 1)
                        for related_col in range(col_idx, max_cols):
                            if related_col <= len(header_values):
                                self.routes[route_key]['columns'].append(related_col)
                        
                        print(f"      └─ Колонки маршрута: {self.routes[route_key]['columns']}")
        
        # Проверяем минимальные требования
        has_required = 'name' in self.columns and 'quantity' in self.columns
        has_routes = len(self.routes) >= 2
        
        print(f"   📊 Найдено: {len(self.columns)} колонок, {len(self.routes)} маршрутов")
        
        return has_required and has_routes
    
    def parse_project(self, project_id):
        """Парсит один проект"""
        try:
            print(f"\n{'='*80}")
            print(f"🔍 Парсинг проекта #{project_id}")
            print(f"{'='*80}")
            
            # Ищем файл
            excel_dir = Path('storage/excel_files')
            matching_files = list(excel_dir.glob(f'project_{project_id}_*.xlsx'))
            
            if not matching_files:
                return {'success': False, 'error': 'File not found'}
            
            file_path = matching_files[0]
            print(f"📁 Файл: {file_path.name}")
            
            # Парсим
            wb_data = load_workbook(str(file_path), data_only=True)
            wb_images = load_workbook(str(file_path), data_only=False)
            
            ws_data = wb_data.active
            ws_images = wb_images.active
            
            # КРИТИЧЕСКИ ВАЖНО: Определяем структуру ДИНАМИЧЕСКИ
            print("\n🔎 Определение структуры...")
            if not self._find_header_and_detect_structure(ws_data):
                wb_data.close()
                wb_images.close()
                return {'success': False, 'error': 'Не удалось определить структуру Template 7'}
            
            # Получаем table_id из БД
            with self.db.get_session() as session:
                result = session.execute(
                    text("SELECT table_id FROM projects WHERE id = :pid"),
                    {'pid': project_id}
                ).fetchone()
                
                if not result:
                    return {'success': False, 'error': 'Project not in DB'}
                
                table_id = result[0]
            
            # Парсим товары
            print(f"\n📦 Парсинг товаров (начало со строки {self.data_start_row})...")
            products = self._parse_products(ws_data, ws_images, project_id, table_id)
            
            wb_data.close()
            wb_images.close()
            
            return {
                'success': True,
                'products': len(products),
                'total_offers': sum(len(p['offers']) for p in products),
                'total_images': sum(len(p['images']) for p in products)
            }
            
        except Exception as e:
            import traceback
            return {'success': False, 'error': str(e), 'traceback': traceback.format_exc()}
    
    def _parse_products(self, ws_data, ws_images, project_id, table_id):
        """Парсит все товары из листа"""
        products = []
        current_product = None
        
        max_row = ws_data.max_row
        
        for row in range(self.data_start_row, max_row + 1):
            name = self._get_cell_value(ws_data, row, self.columns.get('name'))
            quantity = self._get_cell_value(ws_data, row, self.columns.get('quantity'))
            
            # Пропускаем пустые строки
            if not name and not quantity:
                continue
            
            # Новый товар
            if name and len(str(name).strip()) >= 3:
                # Сохраняем предыдущий товар
                if current_product:
                    self._save_product(current_product, project_id, table_id)
                    products.append(current_product)
                
                # Создаем новый товар
                current_product = self._create_product(ws_data, ws_images, row, table_id)
            
            # Парсим предложения для текущего товара
            if current_product and quantity:
                offers = self._parse_offers(ws_data, row)
                current_product['offers'].extend(offers)
        
        # Сохраняем последний товар
        if current_product:
            self._save_product(current_product, project_id, table_id)
            products.append(current_product)
        
        return products
    
    def _create_product(self, ws_data, ws_images, row, table_id):
        """Создает структуру товара"""
        name = self._get_cell_value(ws_data, row, self.columns.get('name'))
        description = self._get_cell_value(ws_data, row, self.columns.get('description'))
        custom_field = self._get_cell_value(ws_data, row, self.columns.get('custom'))
        
        # Парсим данные об образце (если есть)
        sample_price_raw = self._get_cell_value(ws_data, row, self.columns.get('sample'))
        sample_price = self._parse_price(sample_price_raw) if sample_price_raw else None
        
        # Извлекаем изображения
        images = self._extract_images(ws_images, row, table_id)
        
        return {
            'name': str(name).strip(),
            'description': str(description).strip() if description else None,
            'custom_field': str(custom_field).strip() if custom_field else None,
            'sample_price': str(sample_price) if sample_price else None,
            'sample_delivery_time': None,
            'row_number': row,
            'offers': [],
            'images': images
        }
    
    def _parse_offers(self, ws_data, row):
        """
        Парсит предложения из строки для ВСЕХ маршрутов
        
        Template 7 особенность: несколько маршрутов с разными ценами
        """
        offers = []
        quantity = self._get_cell_value(ws_data, row, self.columns.get('quantity'))
        
        if not quantity:
            return offers
        
        # Парсим количество (КРИТИЧНО: используем проверенный метод Template 4)
        qty = self._parse_quantity(quantity)
        if not qty or qty <= 0 or qty > 1_000_000:
            return offers
        
        # Парсим КАЖДЫЙ маршрут
        for route_key, route_data in self.routes.items():
            route_cols = route_data['columns']
            
            # Предполагаем что первая колонка маршрута = USD, вторая = RUB, третья = срок
            # Но это гибко - адаптируется под реальную структуру
            price_usd = None
            price_rub = None
            delivery = None
            
            if len(route_cols) >= 1:
                price_usd = self._parse_price(self._get_cell_value(ws_data, row, route_cols[0]))
            if len(route_cols) >= 2:
                price_rub = self._parse_price(self._get_cell_value(ws_data, row, route_cols[1]))
            if len(route_cols) >= 3:
                delivery = self._parse_delivery_time(self._get_cell_value(ws_data, row, route_cols[2]))
            
            # Создаём оффер если есть хотя бы одна цена
            if price_usd or price_rub:
                # Валидация цен (как в Template 4)
                if price_usd and price_usd >= 10_000:
                    price_usd = None
                if price_rub and price_rub >= 1_000_000:
                    price_rub = None
                
                if price_usd or price_rub:  # еще раз проверяем после валидации
                    offers.append({
                        'quantity': qty,
                        'price_usd': f"{price_usd:.2f}" if price_usd else None,
                        'price_rub': f"{price_rub:.2f}" if price_rub else None,
                        'route': route_key.upper(),  # 'SEA', 'AIR', 'RAILWAY', 'CONTRACT'
                        'delivery_time_days': delivery if delivery and delivery <= 365 else 60
                    })
        
        return offers
    
    def _extract_images(self, ws_images, row, table_id):
        """
        Извлекает изображения из Excel (blob) и сохраняет ЛОКАЛЬНО
        
        В БД записываем ПУТЬ к будущему облачному файлу (для последующей загрузки на S3)
        
        ПРАВИЛЬНЫЙ СПОСОБ (как в Template 4): image.anchor._from.row/col
        """
        images = []
        
        try:
            # Извлекаем все изображения из листа
            if not hasattr(ws_images, '_images'):
                return images
            
            img_index = 0
            for image in ws_images._images:
                try:
                    # ПРАВИЛЬНЫЙ СПОСОБ: используем anchor._from
                    if not hasattr(image, 'anchor') or not hasattr(image.anchor, '_from'):
                        continue
                    
                    img_row = image.anchor._from.row + 1  # +1 потому что openpyxl индексы с 0
                    img_col = image.anchor._from.col + 1
                    
                    # Проверяем принадлежность к текущей строке
                    if img_row != row:
                        continue
                    
                    # Генерируем позицию и имя файла
                    col_letter = chr(64 + img_col)  # 1=A, 2=B, ...
                    cell_pos = f"{col_letter}{row}"
                    
                    # Получаем данные изображения из blob
                    img_data = image._data()
                    img_hash = hashlib.md5(img_data).hexdigest()[:8]
                    
                    # Имя файла (с индексом если несколько изображений в строке)
                    if img_index == 0:
                        filename = f"{table_id}_{cell_pos}_{img_hash}.png"
                    else:
                        filename = f"{table_id}_{cell_pos}_{img_index}_{img_hash}.png"
                    
                    # Сохраняем ЛОКАЛЬНО в storage/images/
                    file_path = self.storage_dir / filename
                    with open(file_path, 'wb') as f:
                        f.write(img_data)
                    
                    # КРИТИЧЕСКИ ВАЖНО: В БД сохраняем путь к БУДУЩЕМУ облачному файлу
                    cloud_path = f"73d16f7545b3-promogoods/images/{filename}"
                    
                    is_main = len(images) == 0  # первое изображение = главное
                    images.append({
                        'filename': filename,
                        'cloud_path': cloud_path,
                        'position': cell_pos,
                        'is_main': is_main
                    })
                    
                    img_index += 1
                    
                except Exception as img_err:
                    print(f"      ⚠️  Ошибка обработки одного изображения: {img_err}")
                    import traceback
                    traceback.print_exc()
                    continue
                    
        except Exception as e:
            print(f"      ⚠️  Ошибка извлечения изображений: {e}")
            import traceback
            traceback.print_exc()
        
        return images
    
    def _save_product(self, product, project_id, table_id):
        """
        Сохраняет товар в БД (логика Template 4)
        """
        try:
            with self.db.get_session() as session:
                # Вставляем товар (с created_at!)
                product_result = session.execute(
                    text("""
                        INSERT INTO products (
                            project_id, name, description, custom_field,
                            sample_price, sample_delivery_time, row_number,
                            created_at, updated_at
                        )
                        VALUES (
                            :project_id, :name, :description, :custom_field,
                            :sample_price, :sample_delivery_time, :row_number,
                            NOW(), NOW()
                        )
                        RETURNING id
                    """),
                    {
                        'project_id': project_id,
                        'name': product['name'],
                        'description': product['description'],
                        'custom_field': product['custom_field'],
                        'sample_price': product['sample_price'],
                        'sample_delivery_time': product['sample_delivery_time'],
                        'row_number': product['row_number']
                    }
                )
                product_id = product_result.fetchone()[0]
                
                # Вставляем офферы
                for offer in product['offers']:
                    session.execute(
                        text("""
                            INSERT INTO price_offers (
                                product_id, quantity, price_usd, price_rub,
                                route, delivery_time_days,
                                created_at, updated_at
                            )
                            VALUES (
                                :product_id, :quantity, :price_usd, :price_rub,
                                :route, :delivery_time_days,
                                NOW(), NOW()
                            )
                        """),
                        {
                            'product_id': product_id,
                            'quantity': offer['quantity'],
                            'price_usd': offer['price_usd'],
                            'price_rub': offer['price_rub'],
                            'route': offer['route'],
                            'delivery_time_days': offer['delivery_time_days']
                        }
                    )
                
                # Вставляем изображения
                for img in product['images']:
                    # КРИТИЧЕСКИ ВАЖНО: 
                    # image_filename = только имя файла
                    # image_url = полный URL к облачному файлу (для будущей загрузки на S3)
                    full_s3_url = f"https://s3.ru1.storage.beget.cloud/{img['cloud_path']}"
                    
                    session.execute(
                        text("""
                            INSERT INTO product_images (
                                product_id, table_id, cell_position, image_filename,
                                image_url, is_main_image,
                                created_at, updated_at
                            )
                            VALUES (
                                :product_id, :table_id, :cell_position, :image_filename,
                                :image_url, :is_main_image,
                                NOW(), NOW()
                            )
                        """),
                        {
                            'product_id': product_id,
                            'table_id': table_id,
                            'cell_position': img['position'],
                            'image_filename': img['filename'],  # только имя файла
                            'image_url': full_s3_url,  # полный URL на S3 (для будущей загрузки)
                            'is_main_image': img['is_main']
                        }
                    )
                
                session.commit()
                print(f"   ✅ Товар #{product_id}: {product['name']} ({len(product['offers'])} офферов, {len(product['images'])} изображений)")
                
        except Exception as e:
            print(f"   ❌ Ошибка сохранения товара: {e}")
    
    # ========================================================================
    # ПРОВЕРЕННЫЕ МЕТОДЫ ПАРСИНГА ИЗ TEMPLATE 4
    # ========================================================================
    
    def _get_cell_value(self, ws, row, col):
        """Безопасное получение значения ячейки"""
        if not col:
            return None
        try:
            return ws.cell(row=row, column=col).value
        except:
            return None
    
    def _parse_quantity(self, value):
        """
        Парсит количество (ПРОВЕРЕННАЯ ЛОГИКА TEMPLATE 4)
        
        🔴 КРИТИЧЕСКАЯ ОБРАБОТКА x10:
        "100.0" → 100 (НЕ 1000!)
        """
        if not value:
            return None
        
        val_str = str(value).strip().replace(' ', '')
        
        # Если есть дефис - берем первое число (ИНТЕРВАЛ)
        if '-' in val_str:
            val_str = val_str.split('-')[0].strip()
        
        # Правильное преобразование через float (для "10000.0" → 10000)
        try:
            return int(float(val_str))
        except (ValueError, TypeError):
            return None
    
    def _parse_price(self, value):
        """Парсит цену (ПРОВЕРЕННАЯ ЛОГИКА TEMPLATE 4)"""
        if not value:
            return None
        
        try:
            val_str = str(value).strip()
            val_str = val_str.replace(',', '.').replace(' ', '')
            val_str = ''.join(c for c in val_str if c.isdigit() or c == '.')
            
            if val_str:
                return float(val_str)
        except:
            pass
        
        return None
    
    def _parse_delivery_time(self, value):
        """
        Парсит срок доставки (ПРОВЕРЕННАЯ ЛОГИКА TEMPLATE 4)
        
        ИНТЕРВАЛ: берём ПОСЛЕДНЕЕ число (максимум)
        """
        if not value:
            return None
        
        try:
            val_str = str(value).strip()
            
            # Если диапазон - берем максимум (хуже для клиента)
            if '-' in val_str:
                val_str = val_str.split('-')[-1].strip()
            
            # Извлекаем цифры
            digits = ''.join(filter(str.isdigit, val_str))
            
            if digits:
                days = int(digits)
                return days if days <= 365 else None
        except:
            pass
        
        return None


# ========================================================================
# ТЕСТОВЫЙ ЗАПУСК
# ========================================================================

if __name__ == '__main__':
    import json
    
    # Загружаем список Template 7 проектов
    with open('TEMPLATE7_FILTERED_RESULTS.json', 'r', encoding='utf-8') as f:
        template7_data = json.load(f)
    
    # Это список project_id (строки)
    project_ids = template7_data.get('template7_projects', [])
    
    print(f"\n{'='*80}")
    print(f"🧪 ТЕСТОВЫЙ ЗАПУСК ПАРСЕРА TEMPLATE 7")
    print(f"{'='*80}")
    print(f"Всего проектов Template 7: {len(project_ids)}")
    print(f"Тестируем первые 10 проектов...")
    
    parser = Template7Parser()
    results = []
    
    for i, project_id in enumerate(project_ids[:10], 1):
        project_id = int(project_id)
        result = parser.parse_project(project_id)
        results.append({
            'project_id': project_id,
            'result': result
        })
        
        # Очищаем динамические данные для следующего проекта
        parser.columns = {}
        parser.routes = {}
        parser.header_row = None
        parser.data_start_row = None
    
    # Итоговый отчет
    print(f"\n{'='*80}")
    print(f"📊 ИТОГОВЫЙ ОТЧЕТ")
    print(f"{'='*80}")
    
    successful = [r for r in results if r['result']['success']]
    failed = [r for r in results if not r['result']['success']]
    
    print(f"✅ Успешно: {len(successful)}/10")
    print(f"❌ Ошибки: {len(failed)}/10")
    
    if successful:
        print(f"\n📦 Статистика успешных:")
        total_products = sum(r['result']['products'] for r in successful)
        total_offers = sum(r['result']['total_offers'] for r in successful)
        total_images = sum(r['result']['total_images'] for r in successful)
        print(f"   Товаров: {total_products}")
        print(f"   Офферов: {total_offers}")
        print(f"   Изображений: {total_images}")
    
    if failed:
        print(f"\n❌ Ошибки:")
        for r in failed:
            print(f"   Проект #{r['project_id']}: {r['result']['error']}")
    
    # Сохраняем результаты
    with open('TEMPLATE7_TEST_RESULTS.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Результаты сохранены в TEMPLATE7_TEST_RESULTS.json")
    print(f"{'='*80}\n")

