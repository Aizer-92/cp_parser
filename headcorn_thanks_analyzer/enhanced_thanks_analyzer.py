#!/usr/bin/env python3
"""
Enhanced HeadCorn Thanks Analyzer
Улучшенный анализатор благодарностей в чате HeadCorn

Улучшения:
- Более точное распознавание упоминаний
- Фильтрация общих слов (всем, большое)
- Лучшая обработка имен пользователей
- Детальная статистика по месяцам
"""

import asyncio
import json
import re
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from telethon import TelegramClient
from telethon.tl.types import User, Chat, Channel
from loguru import logger
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from dateutil.relativedelta import relativedelta

class EnhancedHeadCornThanksAnalyzer:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.target_chat_name = "HeadCorn"
        self.target_chat_id = None
        
        # Создаем папки для вывода
        os.makedirs("output", exist_ok=True)
        os.makedirs("output/logs", exist_ok=True)
        
        # Настройка логирования
        logger.add("output/logs/enhanced_thanks_analyzer.log", rotation="1 day", retention="7 days")
        
        # Структуры для анализа
        self.thanks_data = defaultdict(lambda: defaultdict(int))  # sender -> receiver -> count
        self.monthly_limits = defaultdict(lambda: defaultdict(int))  # sender -> month -> count
        self.user_names = {}  # user_id -> name
        self.mentioned_users = set()  # Все упомянутые пользователи
        self.monthly_stats = defaultdict(lambda: defaultdict(int))  # month -> sender -> count
        
        # Слова для фильтрации (не считаем как имена пользователей)
        self.filter_words = {
            'всем', 'все', 'большое', 'огромное', 'спасибо', 'благодарю', 
            'спс', 'thanks', 'thank', 'you', 'сенкс', 'сенкью', 'сенк',
            'спасибочки', 'спасибочки', 'очень', 'много', 'всех', 'всех',
            'всем', 'всем', 'всем', 'всем', 'всем', 'всем', 'всем'
        }
        
    def load_config(self):
        """Загрузка конфигурации"""
        try:
            with open("config.json", 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Ошибка загрузки конфигурации: {e}")
            raise
    
    async def connect(self):
        """Подключение к Telegram"""
        try:
            print("🔗 Подключение к Telegram...")
            
            self.client = TelegramClient(
                self.config['session_name'],
                self.config['api_id'],
                self.config['api_hash']
            )
            
            await self.client.start()
            
            if await self.client.is_user_authorized():
                me = await self.client.get_me()
                print(f"✅ Подключен как: {me.first_name}")
                return True
            else:
                print("❌ Не удалось авторизоваться")
                return False
                
        except Exception as e:
            logger.error(f"Ошибка подключения: {e}")
            print(f"❌ Ошибка подключения: {e}")
            return False
    
    async def find_headcorn_chat(self):
        """Поиск чата HeadCorn"""
        try:
            print(f"🔍 Поиск чата '{self.target_chat_name}'...")
            
            dialogs = await self.client.get_dialogs()
            
            for dialog in dialogs:
                if self.target_chat_name.lower() in dialog.name.lower():
                    self.target_chat_id = dialog.id
                    print(f"✅ Найден чат: {dialog.name} (ID: {dialog.id})")
                    return True
            
            print(f"❌ Чат '{self.target_chat_name}' не найден")
            print("📋 Доступные чаты:")
            for dialog in dialogs[:10]:
                print(f"  - {dialog.name}")
            return False
            
        except Exception as e:
            logger.error(f"Ошибка поиска чата: {e}")
            return False
    
    async def get_sender_name(self, message):
        """Получение имени отправителя"""
        try:
            if message.sender_id:
                if message.sender_id in self.user_names:
                    return self.user_names[message.sender_id]
                
                sender = await self.client.get_entity(message.sender_id)
                if isinstance(sender, User):
                    name = f"{sender.first_name or ''} {sender.last_name or ''}".strip()
                    if not name:
                        name = sender.username or f"User{message.sender_id}"
                else:
                    name = getattr(sender, 'title', f"Chat{message.sender_id}")
                
                self.user_names[message.sender_id] = name
                return name
            return "Unknown"
        except Exception as e:
            logger.warning(f"Ошибка получения имени отправителя: {e}")
            return f"User{message.sender_id}"
    
    def extract_mentions(self, text):
        """Улучшенное извлечение упоминаний пользователей из текста"""
        mentions = []
        
        # Паттерны для поиска упоминаний
        patterns = [
            r'@(\w+)',  # @username
            r'(\w+),?\s+спасибо',  # имя, спасибо
            r'спасибо\s+(\w+)',  # спасибо имя
            r'(\w+)\s+благодарю',  # имя благодарю
            r'благодарю\s+(\w+)',  # благодарю имя
            r'(\w+)\s+спс',  # имя спс
            r'спс\s+(\w+)',  # спс имя
            r'(\w+)\s+сенкс',  # имя сенкс
            r'сенкс\s+(\w+)',  # сенкс имя
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                # Фильтруем общие слова
                if match.lower() not in self.filter_words and len(match) > 2:
                    mentions.append(match)
        
        return list(set(mentions))  # Убираем дубликаты
    
    def is_thanks_message(self, text):
        """Проверка, является ли сообщение благодарностью"""
        thanks_keywords = [
            'спасибо', 'благодарю', 'спс', 'thanks', 'thank you',
            'сенкс', 'сенкью', 'сенк', 'спасибочки', 'спасибочки'
        ]
        
        text_lower = text.lower()
        return any(keyword in text_lower for keyword in thanks_keywords)
    
    async def analyze_messages(self):
        """Анализ всех сообщений в чате"""
        try:
            print("📥 Загрузка сообщений из чата HeadCorn...")
            
            all_messages = []
            offset_id = 0
            limit = 100
            processed_count = 0
            
            while True:
                messages = await self.client.get_messages(
                    self.target_chat_id,
                    limit=limit,
                    offset_id=offset_id
                )
                
                if not messages:
                    break
                
                # Фильтруем сообщения с текстом
                text_messages = []
                for msg in messages:
                    if msg.text and len(msg.text.strip()) > 2:
                        sender_name = await self.get_sender_name(msg)
                        text_messages.append({
                            'id': msg.id,
                            'sender': sender_name,
                            'sender_id': msg.sender_id,
                            'text': msg.text,
                            'date': msg.date,
                            'month': msg.date.strftime('%Y-%m')
                        })
                
                all_messages.extend(text_messages)
                processed_count += len(messages)
                
                if processed_count % 1000 == 0:
                    print(f"📊 Обработано сообщений: {processed_count}")
                
                offset_id = messages[-1].id if messages else 0
                
                # Ограничиваем количество сообщений для анализа
                if len(all_messages) >= 15000:  # Увеличили лимит
                    print("⚠️ Достигнут лимит сообщений (15000)")
                    break
            
            print(f"✅ Загружено {len(all_messages)} сообщений")
            
            # Анализируем благодарности
            await self.process_thanks(all_messages)
            
        except Exception as e:
            logger.error(f"Ошибка анализа сообщений: {e}")
            print(f"❌ Ошибка анализа сообщений: {e}")
    
    async def process_thanks(self, messages):
        """Обработка благодарностей с учетом ограничений"""
        print("🔍 Анализ благодарностей...")
        
        thanks_count = 0
        limited_count = 0
        filtered_count = 0
        
        for msg in messages:
            if not self.is_thanks_message(msg['text']):
                continue
            
            sender = msg['sender']
            month = msg['month']
            mentions = self.extract_mentions(msg['text'])
            
            # Проверяем ограничение по месяцам
            if self.monthly_limits[sender][month] >= 3:
                limited_count += 1
                continue
            
            # Обрабатываем упоминания
            valid_mentions = 0
            for mention in mentions:
                # Ищем пользователя по упоминанию
                mentioned_user = await self.find_user_by_mention(mention)
                if mentioned_user and mentioned_user.lower() not in self.filter_words:
                    self.thanks_data[sender][mentioned_user] += 1
                    self.monthly_limits[sender][month] += 1
                    self.monthly_stats[month][sender] += 1
                    self.mentioned_users.add(mentioned_user)
                    thanks_count += 1
                    valid_mentions += 1
                    
                    # Проверяем лимит после добавления
                    if self.monthly_limits[sender][month] > 3:
                        self.thanks_data[sender][mentioned_user] -= 1
                        self.monthly_limits[sender][month] -= 1
                        self.monthly_stats[month][sender] -= 1
                        thanks_count -= 1
                        limited_count += 1
                        valid_mentions -= 1
                        break
                else:
                    filtered_count += 1
            
            if valid_mentions == 0 and len(mentions) > 0:
                filtered_count += 1
        
        print(f"✅ Найдено благодарностей: {thanks_count}")
        print(f"⚠️ Ограничено по лимиту: {limited_count}")
        print(f"🚫 Отфильтровано: {filtered_count}")
    
    async def find_user_by_mention(self, mention):
        """Поиск пользователя по упоминанию"""
        try:
            # Фильтруем общие слова
            if mention.lower() in self.filter_words:
                return None
            
            # Сначала ищем в уже известных пользователях
            for user_id, name in self.user_names.items():
                if mention.lower() in name.lower() or name.lower() in mention.lower():
                    return name
            
            # Пытаемся найти пользователя через API
            try:
                user = await self.client.get_entity(f"@{mention}")
                if isinstance(user, User):
                    name = f"{user.first_name or ''} {user.last_name or ''}".strip()
                    if not name:
                        name = user.username or mention
                    return name
            except:
                pass
            
            # Если не нашли, возвращаем упоминание как есть (если это не общее слово)
            if mention.lower() not in self.filter_words:
                return mention
            
            return None
            
        except Exception as e:
            logger.warning(f"Ошибка поиска пользователя {mention}: {e}")
            return None
    
    def create_excel_report(self):
        """Создание Excel отчета"""
        try:
            print("📊 Создание Excel отчета...")
            
            # Создаем DataFrame
            data = []
            for sender, receivers in self.thanks_data.items():
                for receiver, count in receivers.items():
                    data.append({
                        'Отправитель': sender,
                        'Получатель': receiver,
                        'Количество спасибо': count
                    })
            
            df = pd.DataFrame(data)
            
            if df.empty:
                print("⚠️ Нет данных для отчета")
                return
            
            # Сортируем по количеству благодарностей
            df = df.sort_values('Количество спасибо', ascending=False)
            
            # Создаем Excel файл
            filename = f"output/enhanced_headcorn_thanks_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Основная таблица
                df.to_excel(writer, sheet_name='Благодарности', index=False)
                
                # Получаем рабочую книгу и лист
                workbook = writer.book
                worksheet = writer.sheets['Благодарности']
                
                # Стилизация
                self.style_worksheet(worksheet, df)
                
                # Создаем сводную таблицу
                self.create_summary_sheet(workbook, df)
                
                # Создаем лист статистики
                self.create_stats_sheet(workbook)
                
                # Создаем лист месячной статистики
                self.create_monthly_stats_sheet(workbook)
            
            print(f"✅ Отчет сохранен: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Ошибка создания отчета: {e}")
            print(f"❌ Ошибка создания отчета: {e}")
    
    def style_worksheet(self, worksheet, df):
        """Стилизация рабочего листа"""
        # Заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Границы
        thin_border = Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, workbook, df):
        """Создание сводного листа"""
        worksheet = workbook.create_sheet("Сводка")
        
        # Подсчет общих благодарностей
        total_thanks = df['Количество спасибо'].sum()
        unique_senders = df['Отправитель'].nunique()
        unique_receivers = df['Получатель'].nunique()
        
        # Топ отправителей
        top_senders = df.groupby('Отправитель')['Количество спасибо'].sum().sort_values(ascending=False).head(10)
        
        # Топ получателей
        top_receivers = df.groupby('Получатель')['Количество спасибо'].sum().sort_values(ascending=False).head(10)
        
        # Записываем данные
        worksheet['A1'] = "СВОДКА ПО БЛАГОДАРНОСТЯМ В ЧАТЕ HEADCORN (УЛУЧШЕННАЯ ВЕРСИЯ)"
        worksheet['A1'].font = Font(bold=True, size=14)
        
        worksheet['A3'] = "Общая статистика:"
        worksheet['A3'].font = Font(bold=True)
        
        worksheet['A4'] = f"Всего благодарностей: {total_thanks}"
        worksheet['A5'] = f"Уникальных отправителей: {unique_senders}"
        worksheet['A6'] = f"Уникальных получателей: {unique_receivers}"
        
        # Топ отправителей
        worksheet['A8'] = "Топ-10 отправителей благодарностей:"
        worksheet['A8'].font = Font(bold=True)
        
        row = 9
        for sender, count in top_senders.items():
            worksheet[f'A{row}'] = sender
            worksheet[f'B{row}'] = count
            row += 1
        
        # Топ получателей
        worksheet['A20'] = "Топ-10 получателей благодарностей:"
        worksheet['A20'].font = Font(bold=True)
        
        row = 21
        for receiver, count in top_receivers.items():
            worksheet[f'A{row}'] = receiver
            worksheet[f'B{row}'] = count
            row += 1
    
    def create_stats_sheet(self, workbook):
        """Создание листа статистики"""
        worksheet = workbook.create_sheet("Статистика")
        
        # Месячные лимиты
        worksheet['A1'] = "МЕСЯЧНЫЕ ЛИМИТЫ БЛАГОДАРНОСТЕЙ"
        worksheet['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for sender, months in self.monthly_limits.items():
            for month, count in months.items():
                worksheet[f'A{row}'] = sender
                worksheet[f'B{row}'] = month
                worksheet[f'C{row}'] = count
                row += 1
        
        # Заголовки
        worksheet['A3'] = "Отправитель"
        worksheet['B3'] = "Месяц"
        worksheet['C3'] = "Количество"
        
        for cell in worksheet[3]:
            cell.font = Font(bold=True)
    
    def create_monthly_stats_sheet(self, workbook):
        """Создание листа месячной статистики"""
        worksheet = workbook.create_sheet("Месячная статистика")
        
        worksheet['A1'] = "МЕСЯЧНАЯ СТАТИСТИКА БЛАГОДАРНОСТЕЙ"
        worksheet['A1'].font = Font(bold=True, size=14)
        
        # Заголовки
        worksheet['A3'] = "Месяц"
        worksheet['B3'] = "Отправитель"
        worksheet['C3'] = "Количество благодарностей"
        
        for cell in worksheet[3]:
            cell.font = Font(bold=True)
        
        row = 4
        for month in sorted(self.monthly_stats.keys()):
            for sender, count in self.monthly_stats[month].items():
                worksheet[f'A{row}'] = month
                worksheet[f'B{row}'] = sender
                worksheet[f'C{row}'] = count
                row += 1
    
    async def run_analysis(self):
        """Запуск полного анализа"""
        try:
            print("🚀 Запуск улучшенного анализа благодарностей в чате HeadCorn")
            print("=" * 60)
            
            # Подключение
            if not await self.connect():
                return False
            
            # Поиск чата
            if not await self.find_headcorn_chat():
                return False
            
            # Анализ сообщений
            await self.analyze_messages()
            
            # Создание отчета
            report_file = self.create_excel_report()
            
            if report_file:
                print("\n" + "=" * 60)
                print("✅ Анализ завершен успешно!")
                print(f"📊 Отчет: {report_file}")
                print("\n📈 Основные результаты:")
                
                # Показываем топ-5 отправителей
                sender_totals = defaultdict(int)
                for sender, receivers in self.thanks_data.items():
                    sender_totals[sender] = sum(receivers.values())
                
                top_senders = sorted(sender_totals.items(), key=lambda x: x[1], reverse=True)[:5]
                print("\n🏆 Топ-5 отправителей благодарностей:")
                for i, (sender, count) in enumerate(top_senders, 1):
                    print(f"  {i}. {sender}: {count}")
                
                # Показываем топ-5 получателей
                receiver_totals = defaultdict(int)
                for receivers in self.thanks_data.values():
                    for receiver, count in receivers.items():
                        receiver_totals[receiver] += count
                
                top_receivers = sorted(receiver_totals.items(), key=lambda x: x[1], reverse=True)[:5]
                print("\n🎯 Топ-5 получателей благодарностей:")
                for i, (receiver, count) in enumerate(top_receivers, 1):
                    print(f"  {i}. {receiver}: {count}")
                
                # Показываем месячную статистику
                print("\n📅 Активность по месяцам:")
                for month in sorted(self.monthly_stats.keys())[-6:]:  # Последние 6 месяцев
                    total_month = sum(self.monthly_stats[month].values())
                    print(f"  {month}: {total_month} благодарностей")
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка выполнения анализа: {e}")
            print(f"❌ Ошибка выполнения анализа: {e}")
            return False
        
        finally:
            if self.client:
                await self.client.disconnect()

async def main():
    """Главная функция"""
    analyzer = EnhancedHeadCornThanksAnalyzer()
    await analyzer.run_analysis()

if __name__ == "__main__":
    asyncio.run(main())
