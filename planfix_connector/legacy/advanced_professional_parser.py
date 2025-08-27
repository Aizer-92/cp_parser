import requests
import sqlite3
import json
import pandas as pd
from datetime import datetime
import re
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class AdvancedProfessionalParser:
    def __init__(self):
        self.config = self.load_config()
        self.db_path = 'output/planfix_tasks_correct.db'
        self.output_dir = 'output'
        
    def load_config(self) -> Dict[str, Any]:
        """Загружает конфигурацию из JSON файла"""
        try:
            with open('planfix_config.json', 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ Ошибка загрузки конфигурации: {e}")
            return {}
    
    def get_all_tasks_with_custom_fields(self) -> List[Dict[str, Any]]:
        """Получает все задачи с кастомными полями из Planfix API"""
        headers = {
            'Authorization': f'Bearer {self.config["rest_api"]["auth_token"]}',
            'Content-Type': 'application/json'
        }
        
        # Расширенный список полей для получения
        fields = [
            'id', 'name', 'description', 'status', 'startDateTime', 'endDateTime', 
            'deadline', 'assigner', 'assignees', 'priority', 'importance', 
            'workTime', 'planTime', 'factTime', 'createDate', 'modifyDate', 
            'closeDate', 'startDate', 'finishDate'
        ]
        
        # Добавляем кастомные поля
        custom_fields = [
            'custom_1', 'custom_2', 'custom_3', 'custom_4', 'custom_5',
            'custom_6', 'custom_7', 'custom_8', 'custom_9', 'custom_10',
            'custom_11', 'custom_12', 'custom_13', 'custom_14', 'custom_15'
        ]
        fields.extend(custom_fields)
        
        request_data = {
            'offset': 0,
            'pageSize': 100,
            'filters': [
                {
                    "type": 10,  # Статус
                    "operator": "equal",
                    "value": [127, 128, 129]  # Статусы: Поиск, Согласование, Согласовано
                }
            ],
            'fields': ','.join(fields)
        }
        
        try:
            response = requests.post(
                f"{self.config['rest_api']['base_url']}/task/list",
                headers=headers,
                json=request_data,
                timeout=self.config['rest_api']['timeout']
            )
            
            if response.status_code == 200:
                data = response.json()
                tasks = data.get('tasks', [])
                print(f"✅ Получено {len(tasks)} задач из API")
                return tasks
            else:
                print(f"❌ Ошибка API: {response.status_code} - {response.text}")
                return []
                
        except Exception as e:
            print(f"❌ Ошибка при получении задач: {e}")
            return []
    
    def extract_custom_field_value(self, task: Dict[str, Any], field_name: str) -> Optional[Any]:
        """Извлекает значение кастомного поля из задачи"""
        # Пытаемся найти поле в различных местах
        if 'customFields' in task:
            for field in task['customFields']:
                if field.get('name') == field_name:
                    return field.get('value')
        
        # Проверяем прямые поля
        for key in task.keys():
            if field_name.lower() in key.lower():
                return task.get(key)
        
        return None
    
    def extract_task_data(self, task: Dict[str, Any]) -> Dict[str, Any]:
        """Извлекает все необходимые данные из задачи"""
        task_data = {
            'id': task.get('id'),
            'name': task.get('name', ''),
            'description': task.get('description', ''),
            'status': task.get('status', {}).get('name', ''),
            'assigner': task.get('assigner', {}).get('name', '') if task.get('assigner') else '',
            'assignees': self.extract_assignees(task.get('assignees', [])),
            'priority': task.get('priority', ''),
            'importance': task.get('importance', ''),
            'create_date': None,
            'start_date': None,
            'end_date': None,
            'deadline': None,
            'client_grade': None,
            'order_percentage': None,
            'deal_sum': None,
            'calculation_priority': None,
            'calculation_sum': None
        }
        
        # Извлекаем даты
        if task.get('endDateTime') and isinstance(task['endDateTime'], dict):
            task_data['end_date'] = task['endDateTime'].get('datetime')
        if task.get('startDateTime') and isinstance(task['startDateTime'], dict):
            task_data['start_date'] = task['startDateTime'].get('datetime')
        if task.get('createDate'):
            task_data['create_date'] = task['createDate']
        if task.get('deadline'):
            task_data['deadline'] = task['deadline']
        
        # Извлекаем кастомные поля
        task_data['client_grade'] = self.extract_custom_field_value(task, 'Грейд клиента')
        task_data['order_percentage'] = self.extract_custom_field_value(task, '% заказа')
        task_data['deal_sum'] = self.extract_custom_field_value(task, 'Сумма сделки')
        task_data['calculation_priority'] = self.extract_custom_field_value(task, 'Расчет - Приоритет просчета')
        task_data['calculation_sum'] = self.extract_custom_field_value(task, 'Сумма просчета')
        
        return task_data
    
    def extract_assignees(self, assignees) -> str:
        """Извлекает список исполнителей"""
        if not assignees:
            return ''
        
        if isinstance(assignees, list):
            names = []
            for assignee in assignees:
                if isinstance(assignee, dict):
                    names.append(assignee.get('name', ''))
                else:
                    names.append(str(assignee))
            return ', '.join(filter(None, names))
        elif isinstance(assignees, dict):
            return assignees.get('name', '')
        else:
            return str(assignees)
    
    def calculate_advanced_priority(self, task_data: Dict[str, Any]) -> str:
        """Рассчитывает приоритет по новой формуле"""
        try:
            # Получаем значения полей
            client_grade = self.parse_numeric(task_data.get('client_grade'), 3.0)  # Среднее значение по умолчанию
            order_percentage = self.parse_numeric(task_data.get('order_percentage'), 50.0)  # 50% по умолчанию
            deal_sum = self.parse_numeric(task_data.get('deal_sum'), 0)
            calculation_sum = self.parse_numeric(task_data.get('calculation_sum'), 0)
            
            # Используем максимальную сумму из двух полей
            max_sum = max(deal_sum, calculation_sum)
            
            # Формула приоритета (адаптированная из Excel)
            grade_score = (client_grade / 5.0) * 0.4
            
            # Расчет по сумме
            if max_sum <= 250000:
                sum_score = 0.2
            elif max_sum <= 1000000:
                sum_score = 0.4
            elif max_sum <= 5000000:
                sum_score = 0.6
            elif max_sum <= 10000000:
                sum_score = 0.8
            else:
                sum_score = 1.0
            
            sum_score = sum_score * 0.4
            
            # Расчет по проценту заказа
            percentage_score = (order_percentage / 100.0) * 0.2
            
            # Итоговый балл
            total_score = grade_score + sum_score + percentage_score
            
            # Определение приоритета
            if total_score >= 0.8:
                return 'A'
            elif total_score >= 0.6:
                return 'B'
            elif total_score >= 0.4:
                return 'C'
            else:
                return 'D'
                
        except Exception as e:
            print(f"❌ Ошибка расчета приоритета для задачи {task_data.get('id')}: {e}")
            return 'D'
    
    def parse_numeric(self, value: Any, default: float = 0.0) -> float:
        """Парсит числовое значение из строки"""
        if value is None:
            return default
        
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            # Убираем все нечисловые символы, кроме точки и минуса
            cleaned = re.sub(r'[^\d.-]', '', value)
            try:
                return float(cleaned) if cleaned else default
            except ValueError:
                return default
        
        return default
    
    def update_database_with_custom_fields(self, tasks_with_data: List[Dict[str, Any]]):
        """Обновляет базу данных с кастомными полями"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Создаем таблицу для кастомных полей, если её нет
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS custom_field_values (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id INTEGER,
                    field_name TEXT,
                    field_value TEXT,
                    field_type TEXT,
                    UNIQUE(task_id, field_name)
                )
            ''')
            
            # Обновляем основную таблицу tasks
            for task_data in tasks_with_data:
                cursor.execute('''
                    INSERT OR REPLACE INTO tasks 
                    (id, name, description, status_id, status_name, assignees, assigner, start_date_time, export_timestamp)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    task_data['id'],
                    task_data['name'],
                    task_data['description'],
                    self.get_status_id(task_data['status']),
                    task_data['status'],
                    task_data['assignees'],
                    task_data['assigner'],
                    task_data['end_date'] or task_data['start_date'],
                    datetime.now().isoformat()
                ))
                
                # Добавляем кастомные поля
                custom_fields = [
                    ('Грейд клиента', task_data['client_grade'], 'numeric'),
                    ('% заказа', task_data['order_percentage'], 'percentage'),
                    ('Сумма сделки', task_data['deal_sum'], 'currency'),
                    ('Расчет - Приоритет просчета', task_data['calculation_priority'], 'text'),
                    ('Сумма просчета', task_data['calculation_sum'], 'currency')
                ]
                
                for field_name, field_value, field_type in custom_fields:
                    if field_value is not None:
                        cursor.execute('''
                            INSERT OR REPLACE INTO custom_field_values 
                            (task_id, field_name, field_value, field_type)
                            VALUES (?, ?, ?, ?)
                        ''', (task_data['id'], field_name, str(field_value), field_type))
            
            conn.commit()
            conn.close()
            print(f"✅ База данных обновлена для {len(tasks_with_data)} задач")
            
        except Exception as e:
            print(f"❌ Ошибка обновления базы данных: {e}")
    
    def get_status_id(self, status_name: str) -> int:
        """Возвращает ID статуса по имени"""
        status_mapping = {
            'Поиск и расчет товара': 128,
            'КП Согласование': 129,
            'КП Согласовано': 127
        }
        return status_mapping.get(status_name, 128)
    
    def create_advanced_analysis_report(self, tasks_with_data: List[Dict[str, Any]]):
        """Создает расширенный анализ с кастомными полями"""
        try:
            # Создаем DataFrame
            df = pd.DataFrame(tasks_with_data)
            
            # Добавляем рассчитанный приоритет
            df['calculated_priority'] = df.apply(self.calculate_advanced_priority, axis=1)
            
            # Анализ по приоритетам
            priority_analysis = df['calculated_priority'].value_counts().sort_index()
            
            # Анализ по статусам
            status_analysis = df['status'].value_counts()
            
            # Анализ по грейдам клиентов
            client_grades = df[df['client_grade'].notna()]['client_grade'].value_counts().sort_index()
            
            # Анализ по суммам сделок
            deal_sums = df[df['deal_sum'].notna()]['deal_sum'].apply(self.parse_numeric)
            sum_analysis = {
                'Общая сумма': deal_sums.sum(),
                'Средняя сумма': deal_sums.mean(),
                'Максимальная сумма': deal_sums.max(),
                'Минимальная сумма': deal_sums.min()
            }
            
            # Анализ по исполнителям
            assignee_analysis = df['assignees'].value_counts().head(10)
            
            # Анализ по постановщикам
            assigner_analysis = df['assigner'].value_counts().head(10)
            
            # Создаем HTML отчет
            self.generate_advanced_html_report(
                df, status_analysis, priority_analysis, client_grades, 
                sum_analysis, assignee_analysis, assigner_analysis
            )
            
            # Создаем Excel отчет
            self.generate_advanced_excel_report(
                df, status_analysis, priority_analysis, client_grades, 
                sum_analysis, assignee_analysis, assigner_analysis
            )
            
            print("✅ Расширенные отчеты созданы")
            
        except Exception as e:
            print(f"❌ Ошибка создания анализа: {e}")
    
    def generate_advanced_html_report(self, df: pd.DataFrame, status_analysis: pd.Series, 
                                    priority_analysis: pd.Series, client_grades: pd.Series,
                                    sum_analysis: Dict[str, float], assignee_analysis: pd.Series,
                                    assigner_analysis: pd.Series):
        """Генерирует современный HTML отчет"""
        
        html_content = f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Продвинутый Анализ Задач Planfix</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}
        
        .header {{
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 20px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            text-align: center;
        }}
        
        .header h1 {{
            color: #2c3e50;
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 300;
        }}
        
        .header p {{
            color: #7f8c8d;
            font-size: 1.2em;
        }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 25px;
            margin-bottom: 30px;
        }}
        
        .stat-card {{
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 20px;
            padding: 25px;
            box-shadow: 0 15px 35px rgba(0,0,0,0.1);
            transition: transform 0.3s ease, box-shadow 0.3s ease;
        }}
        
        .stat-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 25px 50px rgba(0,0,0,0.15);
        }}
        
        .stat-card h3 {{
            color: #2c3e50;
            margin-bottom: 15px;
            font-size: 1.3em;
            font-weight: 500;
        }}
        
        .stat-value {{
            font-size: 2.5em;
            font-weight: 700;
            color: #3498db;
            margin-bottom: 10px;
        }}
        
        .priority-a {{ color: #e74c3c; }}
        .priority-b {{ color: #f39c12; }}
        .priority-c {{ color: #f1c40f; }}
        .priority-d {{ color: #95a5a6; }}
        
        .chart-container {{
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 20px;
            padding: 25px;
            margin-bottom: 30px;
            box-shadow: 0 15px 35px rgba(0,0,0,0.1);
        }}
        
        .chart-container h3 {{
            color: #2c3e50;
            margin-bottom: 20px;
            font-size: 1.4em;
            font-weight: 500;
        }}
        
        .data-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        
        .data-table th, .data-table td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ecf0f1;
        }}
        
        .data-table th {{
            background: #f8f9fa;
            color: #2c3e50;
            font-weight: 600;
        }}
        
        .data-table tr:hover {{
            background: #f8f9fa;
        }}
        
        .footer {{
            text-align: center;
            color: rgba(255, 255, 255, 0.8);
            margin-top: 40px;
            font-size: 0.9em;
        }}
        
        @media (max-width: 768px) {{
            .stats-grid {{
                grid-template-columns: 1fr;
            }}
            
            .header h1 {{
                font-size: 2em;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Продвинутый Анализ Задач Planfix</h1>
            <p>Комплексный анализ с кастомными полями и новой формулой приоритетов</p>
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <h3>📊 Общее количество задач</h3>
                <div class="stat-value">{len(df)}</div>
                <p>Активных задач в системе</p>
            </div>
            
            <div class="stat-card">
                <h3>💰 Общая сумма сделок</h3>
                <div class="stat-value">{sum_analysis.get('Общая сумма', 0):,.0f} ₽</div>
                <p>Сумма всех просчетов</p>
            </div>
            
            <div class="stat-card">
                <h3>📈 Средняя сумма</h3>
                <div class="stat-value">{sum_analysis.get('Средняя сумма', 0):,.0f} ₽</div>
                <p>Средний размер сделки</p>
            </div>
            
            <div class="stat-card">
                <h3>🎯 Максимальная сумма</h3>
                <div class="stat-value">{sum_analysis.get('Максимальная сумма', 0):,.0f} ₽</div>
                <p>Самая крупная сделка</p>
            </div>
        </div>
        
        <div class="chart-container">
            <h3>🎯 Распределение по приоритетам (новая формула)</h3>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Приоритет</th>
                        <th>Количество</th>
                        <th>Процент</th>
                    </tr>
                </thead>
                <tbody>
                    {self.generate_priority_table_rows(priority_analysis)}
                </tbody>
            </table>
        </div>
        
        <div class="chart-container">
            <h3>📋 Распределение по статусам</h3>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Статус</th>
                        <th>Количество</th>
                        <th>Процент</th>
                    </tr>
                </thead>
                <tbody>
                    {self.generate_status_table_rows(status_analysis)}
                </tbody>
            </table>
        </div>
        
        <div class="chart-container">
            <h3>👥 Топ исполнителей</h3>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Исполнитель</th>
                        <th>Количество задач</th>
                    </tr>
                </thead>
                <tbody>
                    {self.generate_assignee_table_rows(assignee_analysis)}
                </tbody>
            </table>
        </div>
        
        <div class="chart-container">
            <h3>👤 Топ постановщиков</h3>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Постановщик</th>
                        <th>Количество задач</th>
                    </tr>
                </thead>
                <tbody>
                    {self.generate_assigner_table_rows(assigner_analysis)}
                </tbody>
            </table>
        </div>
        
        <div class="chart-container">
            <h3>📊 Детальные данные по задачам</h3>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Название</th>
                        <th>Статус</th>
                        <th>Приоритет</th>
                        <th>Грейд клиента</th>
                        <th>% заказа</th>
                        <th>Сумма сделки</th>
                        <th>Исполнитель</th>
                    </tr>
                </thead>
                <tbody>
                    {self.generate_tasks_table_rows(df.head(20))}
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
        """
        
        with open(f'{self.output_dir}/advanced_analysis_report.html', 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print("✅ HTML отчет создан: advanced_analysis_report.html")
    
    def generate_priority_table_rows(self, priority_analysis: pd.Series) -> str:
        """Генерирует строки таблицы приоритетов"""
        rows = ""
        total = priority_analysis.sum()
        
        for priority, count in priority_analysis.items():
            percentage = (count / total * 100) if total > 0 else 0
            priority_class = f"priority-{priority.lower()}"
            rows += f"""
                <tr>
                    <td><span class="{priority_class}">Приоритет {priority}</span></td>
                    <td>{count}</td>
                    <td>{percentage:.1f}%</td>
                </tr>
            """
        return rows
    
    def generate_status_table_rows(self, status_analysis: pd.Series) -> str:
        """Генерирует строки таблицы статусов"""
        rows = ""
        total = status_analysis.sum()
        
        for status, count in status_analysis.items():
            percentage = (count / total * 100) if total > 0 else 0
            rows += f"""
                <tr>
                    <td>{status}</td>
                    <td>{count}</td>
                    <td>{percentage:.1f}%</td>
                </tr>
            """
        return rows
    
    def generate_assignee_table_rows(self, assignee_analysis: pd.Series) -> str:
        """Генерирует строки таблицы исполнителей"""
        rows = ""
        for assignee, count in assignee_analysis.items():
            if assignee:  # Пропускаем пустые значения
                rows += f"""
                    <tr>
                        <td>{assignee}</td>
                        <td>{count}</td>
                    </tr>
                """
        return rows
    
    def generate_assigner_table_rows(self, assigner_analysis: pd.Series) -> str:
        """Генерирует строки таблицы постановщиков"""
        rows = ""
        for assigner, count in assigner_analysis.items():
            if assigner:  # Пропускаем пустые значения
                rows += f"""
                    <tr>
                        <td>{assigner}</td>
                        <td>{count}</td>
                    </tr>
                """
        return rows
    
    def generate_tasks_table_rows(self, df: pd.DataFrame) -> str:
        """Генерирует строки таблицы задач"""
        rows = ""
        for _, task in df.iterrows():
            rows += f"""
                <tr>
                    <td>{task['id']}</td>
                    <td>{task['name'][:50]}...</td>
                    <td>{task['status']}</td>
                    <td><span class="priority-{task['calculated_priority'].lower()}">{task['calculated_priority']}</span></td>
                    <td>{task['client_grade'] or 'Н/Д'}</td>
                    <td>{task['order_percentage'] or 'Н/Д'}</td>
                    <td>{task['deal_sum'] or 'Н/Д'}</td>
                    <td>{task['assignees'] or 'Н/Д'}</td>
                </tr>
            """
        return rows
    
    def generate_advanced_excel_report(self, df: pd.DataFrame, status_analysis: pd.Series, 
                                     priority_analysis: pd.Series, client_grades: pd.Series,
                                     sum_analysis: Dict[str, float], assignee_analysis: pd.Series,
                                     assigner_analysis: pd.Series):
        """Генерирует расширенный Excel отчет"""
        try:
            # Создаем новый Excel файл
            wb = openpyxl.Workbook()
            
            # Удаляем лист по умолчанию
            wb.remove(wb.active)
            
            # Создаем лист с общими данными
            ws_summary = wb.create_sheet("Общий анализ")
            self.add_summary_sheet(ws_summary, df, status_analysis, priority_analysis, sum_analysis)
            
            # Создаем лист с детальными данными
            ws_details = wb.create_sheet("Детальные данные")
            self.add_details_sheet(ws_details, df)
            
            # Создаем лист с анализом по исполнителям
            ws_assignees = wb.create_sheet("Анализ исполнителей")
            self.add_assignee_sheet(ws_assignees, assignee_analysis)
            
            # Создаем лист с анализом по постановщикам
            ws_assigners = wb.create_sheet("Анализ постановщиков")
            self.add_assigner_sheet(ws_assigners, assigner_analysis)
            
            # Сохраняем файл
            wb.save(f'{self.output_dir}/advanced_analysis_report.xlsx')
            print("✅ Excel отчет создан: advanced_analysis_report.xlsx")
            
        except Exception as e:
            print(f"❌ Ошибка создания Excel отчета: {e}")
    
    def add_summary_sheet(self, ws, df: pd.DataFrame, status_analysis: pd.Series, 
                         priority_analysis: pd.Series, sum_analysis: Dict[str, float]):
        """Добавляет лист с общим анализом"""
        # Заголовки
        headers = ['Показатель', 'Значение', 'Описание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Данные
        data = [
            ['Общее количество задач', len(df), 'Всего активных задач в системе'],
            ['Общая сумма сделок', f"{sum_analysis.get('Общая сумма', 0):,.0f} ₽", 'Сумма всех просчетов'],
            ['Средняя сумма сделки', f"{sum_analysis.get('Средняя сумма', 0):,.0f} ₽", 'Средний размер сделки'],
            ['Максимальная сумма', f"{sum_analysis.get('Максимальная сумма', 0):,.0f} ₽", 'Самая крупная сделка'],
            ['Минимальная сумма', f"{sum_analysis.get('Минимальная сумма', 0):,.0f} ₽", 'Самая мелкая сделка']
        ]
        
        for row, (indicator, value, description) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=indicator)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=description)
        
        # Анализ по статусам
        ws.cell(row=8, column=1, value='Анализ по статусам').font = Font(bold=True, size=14)
        ws.cell(row=9, column=1, value='Статус')
        ws.cell(row=9, column=2, value='Количество')
        ws.cell(row=9, column=3, value='Процент')
        
        for row, (status, count) in enumerate(status_analysis.items(), 10):
            percentage = (count / len(df) * 100) if len(df) > 0 else 0
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        # Анализ по приоритетам
        start_row = 10 + len(status_analysis)
        ws.cell(row=start_row, column=1, value='Анализ по приоритетам').font = Font(bold=True, size=14)
        ws.cell(row=start_row+1, column=1, value='Приоритет')
        ws.cell(row=start_row+1, column=2, value='Количество')
        ws.cell(row=start_row+1, column=3, value='Процент')
        
        for row, (priority, count) in enumerate(priority_analysis.items(), start_row+2):
            percentage = (count / len(df) * 100) if len(df) > 0 else 0
            ws.cell(row=row, column=1, value=f"Приоритет {priority}")
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_details_sheet(self, ws, df: pd.DataFrame):
        """Добавляет лист с детальными данными"""
        # Заголовки
        headers = ['ID', 'Название', 'Статус', 'Приоритет', 'Грейд клиента', '% заказа', 'Сумма сделки', 'Исполнитель', 'Постановщик']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Данные
        for row, (_, task) in enumerate(df.iterrows(), 2):
            ws.cell(row=row, column=1, value=task['id'])
            ws.cell(row=row, column=2, value=task['name'][:100])  # Ограничиваем длину названия
            ws.cell(row=row, column=3, value=task['status'])
            ws.cell(row=row, column=4, value=task['calculated_priority'])
            ws.cell(row=row, column=5, value=task['client_grade'])
            ws.cell(row=row, column=6, value=task['order_percentage'])
            ws.cell(row=row, column=7, value=task['deal_sum'])
            ws.cell(row=row, column=8, value=task['assignees'])
            ws.cell(row=row, column=9, value=task['assigner'])
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_assignee_sheet(self, ws, assignee_analysis: pd.Series):
        """Добавляет лист с анализом исполнителей"""
        ws.cell(row=1, column=1, value='Топ исполнителей').font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value='Исполнитель').font = Font(bold=True)
        ws.cell(row=2, column=2, value='Количество задач').font = Font(bold=True)
        
        for row, (assignee, count) in enumerate(assignee_analysis.items(), 3):
            if assignee:  # Пропускаем пустые значения
                ws.cell(row=row, column=1, value=assignee)
                ws.cell(row=row, column=2, value=count)
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_assigner_sheet(self, ws, assigner_analysis: pd.Series):
        """Добавляет лист с анализом постановщиков"""
        ws.cell(row=1, column=1, value='Топ постановщиков').font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value='Постановщик').font = Font(bold=True)
        ws.cell(row=2, column=2, value='Количество задач').font = Font(bold=True)
        
        for row, (assigner, count) in enumerate(assigner_analysis.items(), 3):
            if assigner:  # Пропускаем пустые значения
                ws.cell(row=row, column=1, value=assigner)
                ws.cell(row=row, column=2, value=count)
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def run_full_analysis(self):
        """Запускает полный анализ"""
        print("🚀 Запуск продвинутого профессионального анализа...")
        
        # Получаем все задачи с кастомными полями
        print("📡 Получение задач из API...")
        tasks = self.get_all_tasks_with_custom_fields()
        
        if not tasks:
            print("❌ Не удалось получить задачи из API")
            return
        
        # Извлекаем данные из задач
        print("🔍 Извлечение данных из задач...")
        tasks_with_data = []
        for task in tasks:
            task_data = self.extract_task_data(task)
            tasks_with_data.append(task_data)
        
        # Обновляем базу данных
        print("💾 Обновление базы данных...")
        self.update_database_with_custom_fields(tasks_with_data)
        
        # Создаем расширенный анализ
        print("📊 Создание расширенного анализа...")
        self.create_advanced_analysis_report(tasks_with_data)
        
        print("✅ Продвинутый анализ завершен!")
        print(f"📁 Отчеты сохранены в папке: {self.output_dir}")

if __name__ == "__main__":
    parser = AdvancedProfessionalParser()
    parser.run_full_analysis()
